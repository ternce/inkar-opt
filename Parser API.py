import requests
import pandas as pd
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# =========================
# НАСТРОЙКИ
# =========================

TOKEN = os.getenv("PROVISOR_TOKEN", "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJqdGkiOiJiM2E3N2U3NS02ODEwLTRiY2YtOTIyMS1jNGEwZGI5ZWRjNDUiLCJodHRwOi8vc2NoZW1hcy54bWxzb2FwLm9yZy93cy8yMDA1LzA1L2lkZW50aXR5L2NsYWltcy9uYW1laWRlbnRpZmllciI6IjVlZmIyOTZkLTk5NTItNGI5ZS1iYjcwLTgzZmFkMzE2MTM4MyIsImh0dHA6Ly9zY2hlbWFzLnhtbHNvYXAub3JnL3dzLzIwMDUvMDUvaWRlbnRpdHkvY2xhaW1zL25hbWUiOiJBa3NhaTQvODMiLCJodHRwOi8vc2NoZW1hcy54bWxzb2FwLm9yZy93cy8yMDA1LzA1L2lkZW50aXR5L2NsYWltcy9oYXNoIjoiQVFBQUFBRUFBQ2NRQUFBQUVFTTJ3azFLU1U5TEhlaW91eUZuZmIxRk1WZi85NzVZOXRzajRVNjRNZGQvTnFuQkRjeGRpZWY3MjlTNjRPRmNYdz09IiwiQ29tcGFueVR5cGUiOiLQkNC_0YLQtdC60LAiLCJodHRwOi8vc2NoZW1hcy5taWNyb3NvZnQuY29tL3dzLzIwMDgvMDYvaWRlbnRpdHkvY2xhaW1zL3JvbGUiOiJQaGFybWFjeSIsImV4cCI6MTc3Njk2Nzg5MiwiaXNzIjoiaHR0cHM6Ly9QaGFybWNlbnRlci5reiIsImF1ZCI6IlBoYXJtYWNldXRpY2FsIG1hcmtldCJ9.E4UbZGiFEsdbw7641DBrJtNXzNlCUiAznSDSAp4ZPBI")

FILIALS = {
    128: "Инкар Алматы",
    148: "Медсервис Караганда",
    149: "Медсервис Астана",
    151: "Медсервис Талдыкорган",
    152: "Медсервис Усть-Каменогорск",
    153: "Медсервис (Уральск)",
    154: "Медсервис (Костанай)",
    155: "Медсервис (Павлодар)",
    158: "Медсервис (Кызылорда)",
    159: "Медсервис (Актау)",
    162: "Медсервис (Актобе)",
}

FILIAL_IDS = list(FILIALS.keys())

BASE_URL = "https://api.provisor.kz"

HEADERS = {
    "Authorization": f"Bearer {TOKEN}",
    "Accept": "application/json",
    "User-Agent": "Mozilla/5.0"
}

EXPORT_DIR = "exports"
os.makedirs(EXPORT_DIR, exist_ok=True)

# =========================
# API
# =========================

def fetch_price_data(session, filial_id):
    url = f"{BASE_URL}/Price/GetByFilialId?filialId={filial_id}"
    print(f"[LOAD] filial={filial_id}")
    r = session.get(url, headers=HEADERS, timeout=60)
    r.raise_for_status()
    return r.json()


# =========================
# DATA
# =========================

def flatten_item(item):
    return {
        "goods_id": item.get("goodsId"),
        "filial_id": item.get("filialId"),
        "distributor_goods_id": str(item.get("distributorGoodsId")).strip() if item.get("distributorGoodsId") else None,
        "goods_full_name": item.get("goods", {}).get("fullName"),
        "distributor_goods_name": item.get("distributorGoodsName"),
        "price": item.get("goodsPrice"),
        "stored": item.get("stored"),
    }


def build_dataframe(data):
    df = pd.DataFrame([flatten_item(x) for x in data])

    df["price"] = pd.to_numeric(df["price"], errors="coerce")
    df["stored"] = pd.to_numeric(df["stored"], errors="coerce")

    return df


# =========================
# PIVOT
# =========================

def build_pivot(df, filial_ids):
    df = df[df["goods_id"].notna()].copy()

    # составной SKU
    df["SKU"] = (
        df["goods_id"].astype(str)
        + "_"
        + df["distributor_goods_id"].fillna("no_sku")
    )

    df["Название"] = df["goods_full_name"].fillna(df["distributor_goods_name"])

    base = df[[
        "SKU",
        "goods_id",
        "distributor_goods_id",
        "Название"
    ]].drop_duplicates()

    base = base.rename(columns={
        "goods_id": "goodsId",
        "distributor_goods_id": "SKU дистрибьютора"
    })

    result = base.copy()

    for fid in filial_ids:
        label = FILIALS.get(fid, str(fid))

        subset = df[df["filial_id"] == fid][["SKU", "price", "stored"]]

        subset = subset.groupby("SKU", as_index=False).agg({
            "price": "min",
            "stored": "sum"
        })

        subset = subset.rename(columns={
            "price": f"Цена {label}",
            "stored": f"Остаток {label}",
        })

        result = result.merge(subset, on="SKU", how="left")

    return result.sort_values(["goodsId", "SKU"]).reset_index(drop=True)


# =========================
# EXCEL FORMAT
# =========================

def format_excel(file_path):
    wb = load_workbook(file_path)
    ws = wb.active

    # 🔥 авто-ширина + фикс для названия
    for col in ws.columns:
        col_letter = col[0].column_letter

        if col_letter == "D":  # Название
            ws.column_dimensions[col_letter].width = 70
            continue

        max_length = 0
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = max_length + 2

    # 🔥 перенос текста
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)

    # 🔥 закрепить шапку
    ws.freeze_panes = "A2"

    wb.save(file_path)


# =========================
# MAIN
# =========================

def main():
    session = requests.Session()
    frames = []

    for fid in FILIAL_IDS:
        try:
            data = fetch_price_data(session, fid)

            if not isinstance(data, list):
                print(f"[ERROR] filial {fid}")
                continue

            print(f"[OK] {fid}: {len(data)}")

            df = build_dataframe(data)
            frames.append(df)

        except Exception as e:
            print(f"[ERROR] {fid}:", e)

    if not frames:
        raise Exception("Нет данных")

    df_all = pd.concat(frames, ignore_index=True)

    pivot = build_pivot(df_all, FILIAL_IDS)

    now = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    file_path = os.path.join(EXPORT_DIR, f"prices_{now}.xlsx")

    pivot.to_excel(file_path, index=False)

    # 🔥 форматирование
    format_excel(file_path)

    print(f"\n[OK] Файл готов: {os.path.abspath(file_path)}")


if __name__ == "__main__":
    main()