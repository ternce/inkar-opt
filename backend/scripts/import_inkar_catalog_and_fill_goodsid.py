from __future__ import annotations

import csv
import io
import json
import os
import re
import urllib.parse
import urllib.request
from collections import Counter, defaultdict
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import create_engine, text


ROOT_DIR = Path(__file__).resolve().parents[2]
BACKEND_DIR = ROOT_DIR / "backend"
DIAGNOSTICS_DIR = BACKEND_DIR / "diagnostics"
DEFAULT_CATALOG = Path(r"C:\Users\Dilsh\Downloads\Инкар_Ассортимент_2.xlsx")
PHARMCENTER_URL = "https://ph.center/api/Report/PricesAnalysis"
REFERENCE_FILIALS = (128, 133)
UNKNOWN_GOODS_NAME = "неизвестный товар"


def load_env_file(path: Path) -> None:
    if not path.exists():
        return
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if line and not line.startswith("#") and "=" in line:
            key, value = line.split("=", 1)
            os.environ.setdefault(key.strip(), value.strip().strip('"').strip("'"))


def database_url() -> str:
    raw = os.getenv("DATABASE_URL", "postgresql://apteka:apteka@127.0.0.1:5432/apteka")
    if "@postgres:" in raw:
        raw = raw.replace("@postgres:", "@127.0.0.1:")
    if raw.startswith("postgres://"):
        raw = "postgresql://" + raw[len("postgres://") :]
    if raw.startswith("postgresql://") and "+psycopg" not in raw:
        raw = "postgresql+psycopg://" + raw[len("postgresql://") :]
    return raw


def safe_db_url(url: str) -> str:
    return re.sub(r"://([^:/@]+):([^@]+)@", r"://\1:***@", url.replace("+psycopg", ""))


def txt(value: Any) -> str:
    return "" if value is None else str(value).strip()


def normalize_sku(value: Any, pad_to: int = 18) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if value != value:
            return ""
        s = str(int(value)) if value.is_integer() else str(value)
    else:
        s = str(value).strip()
    s = s.replace("\ufeff", "").replace("\u200b", "").replace("\xa0", " ")
    s = re.sub(r"\.0+$", "", s)
    if "_" in s:
        s = s.split("_")[-1].strip()
    digits = re.sub(r"\D+", "", s)
    if not digits:
        return ""
    if len(digits) < pad_to:
        digits = digits.zfill(pad_to)
    elif len(digits) > pad_to:
        digits = digits[-pad_to:]
    return digits


def normalize_name(value: Any) -> str:
    s = txt(value).lower().replace("ё", "е")
    s = re.sub(r"\s+", " ", s)
    s = re.sub(r"\s*/\s*", "/", s)
    s = re.sub(r"/+", "/", s)
    return s.strip(" /")


def norm_header(value: Any) -> str:
    return re.sub(r"\s+", " ", txt(value).lower())


def as_decimal(value: Any, default: Decimal | None = None) -> Decimal | None:
    if value in (None, ""):
        return default
    try:
        if isinstance(value, (int, float, Decimal)):
            return Decimal(str(value))
        return Decimal(txt(value).replace(" ", "").replace(",", "."))
    except Exception:
        return default


def first_nonempty(*values: Any) -> str:
    for value in values:
        s = txt(value)
        if s:
            return s
    return ""


def parse_raw_json(value: Any) -> dict[str, Any]:
    if not value:
        return {}
    try:
        parsed = json.loads(str(value))
        return parsed if isinstance(parsed, dict) else {}
    except json.JSONDecodeError:
        return {}


def nested_get(payload: dict[str, Any], path: tuple[str, ...]) -> Any:
    cur: Any = payload
    for key in path:
        if not isinstance(cur, dict):
            return None
        cur = cur.get(key)
    return cur


def write_csv(path: Path, rows: list[dict[str, Any]], fields: list[str]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def load_catalog(path: Path) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    wb = load_workbook(io.BytesIO(path.read_bytes()), data_only=True)
    ws = wb.worksheets[0]
    header_row = None
    required_header_pairs = (
        ("материал", "артикул"),
        ("sku", "name"),
        ("код", "наименование"),
        ("код", "название"),
    )
    for row in range(1, min(ws.max_row, 30) + 1):
        row_headers = {
            norm_header(ws.cell(row=row, column=col).value)
            for col in range(1, ws.max_column + 1)
            if ws.cell(row=row, column=col).value not in (None, "")
        }
        if any(left in row_headers and right in row_headers for left, right in required_header_pairs):
            header_row = row
            break
    if header_row is None:
        raise RuntimeError("Catalog has no header row")

    headers: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        key = norm_header(ws.cell(row=header_row, column=col).value)
        if key:
            headers[key] = col

    def cell(row: int, *names: str) -> Any:
        for name in names:
            col = headers.get(norm_header(name))
            if col:
                return ws.cell(row=row, column=col).value
        return None

    parsed: list[dict[str, Any]] = []
    raw_skus: list[str] = []
    for row in range(header_row + 1, ws.max_row + 1):
        sku = normalize_sku(cell(row, "Материал", "SKU", "Код"))
        if not sku:
            continue
        raw_skus.append(sku)
        cost = as_decimal(cell(row, "Себестоимость", "Учетная себестоимость", "Учётная себестоимость"), Decimal("0")) or Decimal("0")
        stock = as_decimal(cell(row, "Остаток", "Остатки"))
        parsed.append(
            {
                "code": sku,
                "name": txt(cell(row, "Артикул", "Наименование", "Название")) or sku,
                "manufacturer": txt(cell(row, "Производитель")),
                "stock": stock,
                "cost": cost,
                "source_row": row,
            }
        )

    rows_by_sku: dict[str, dict[str, Any]] = {}
    for row in parsed:
        rows_by_sku[row["code"]] = row
    duplicates = sorted(code for code, count in Counter(raw_skus).items() if count > 1)
    return list(rows_by_sku.values()), {
        "source_file": str(path),
        "excel_rows": len(parsed),
        "unique_skus": len(rows_by_sku),
        "duplicate_skus_count": len(duplicates),
        "duplicate_skus": duplicates[:200],
    }


def fetch_pharmcenter() -> list[dict[str, Any]]:
    token = os.getenv("PHCENTER_TOKEN", "")
    query = urllib.parse.urlencode({"region": 1, "price_mode": 0, "distributors": 1})
    request = urllib.request.Request(
        f"{PHARMCENTER_URL}?{query}",
        headers={"Authorization": token} if token else {},
    )
    with urllib.request.urlopen(request, timeout=120) as response:
        data = json.loads(response.read().decode("utf-8"))
    goods = nested_get(data, ("root", "goods", "good")) if isinstance(data, dict) else None
    if isinstance(goods, list):
        return goods
    if isinstance(data, list):
        return data
    raise RuntimeError(f"Unexpected PharmCenter response shape: {type(data).__name__}")


def source_key(row: dict[str, Any]) -> tuple[str, str, str]:
    return (txt(row.get("price_list_id")), txt(row.get("account_id")), txt(row.get("filial_id")))


def compact_sources(rows: list[dict[str, Any]], limit: int = 20) -> list[dict[str, Any]]:
    seen: set[tuple[str, str, str]] = set()
    out: list[dict[str, Any]] = []
    for row in rows:
        key = source_key(row)
        if key in seen:
            continue
        seen.add(key)
        out.append(
            {
                "price_list_id": row["price_list_id"],
                "filialId": row["filial_id"],
                "filialName": row["branch_name"],
                "accountId": row["account_id"],
                "accountLogin": row["account_login"],
            }
        )
        if len(out) >= limit:
            break
    return out


def main() -> int:
    load_env_file(ROOT_DIR / ".env")
    load_env_file(BACKEND_DIR / ".env")
    DIAGNOSTICS_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    catalog_path = Path(os.getenv("INKAR_CATALOG_PATH", str(DEFAULT_CATALOG)))
    catalog_rows, catalog_meta = load_catalog(catalog_path)

    engine = create_engine(database_url(), connect_args={"connect_timeout": 10})
    safe_url = safe_db_url(str(engine.url))

    safe_candidates_csv = DIAGNOSTICS_DIR / "safe_candidates.csv"
    conflicts_csv = DIAGNOSTICS_DIR / "conflicts.csv"
    coverage_summary_json = DIAGNOSTICS_DIR / "coverage_summary.json"
    import_summary_json = DIAGNOSTICS_DIR / "import_summary.json"
    backup_csv = DIAGNOSTICS_DIR / f"inkar_goodsid_fill_backup_{timestamp}.csv"
    backup_json = DIAGNOSTICS_DIR / f"inkar_goodsid_fill_backup_{timestamp}.json"

    with engine.begin() as conn:
        before = conn.execute(
            text(
                """
                SELECT
                    (SELECT count(*) FROM products) AS total,
                    (SELECT count(*) FROM products WHERE provisor_goods_id IS NOT NULL) AS with_goods_id,
                    (SELECT count(*) FROM products WHERE provisor_goods_id IS NULL) AS without_goods_id
                """
            )
        ).mappings().one()

        codes = [row["code"] for row in catalog_rows]
        existing_rows = conn.execute(
            text(
                """
                SELECT p.id, p.code, p.name, p.cost, p.provisor_goods_id, e.manufacturer, e.stock
                FROM products p
                LEFT JOIN product_extras e ON e.product_id = p.id
                WHERE p.code = ANY(:codes)
                """
            ),
            {"codes": codes},
        ).mappings().all()
        existing_by_code = {row["code"]: row for row in existing_rows}

        changed_existing = 0
        created = 0
        processed_existing = 0
        for row in catalog_rows:
            existing = existing_by_code.get(row["code"])
            if existing is None:
                product_id = conn.execute(
                    text(
                        """
                        INSERT INTO products (code, name, cost, created_at)
                        VALUES (:code, :name, :cost, CURRENT_TIMESTAMP)
                        RETURNING id
                        """
                    ),
                    {"code": row["code"], "name": row["name"], "cost": row["cost"]},
                ).scalar_one()
                conn.execute(
                    text(
                        """
                        INSERT INTO product_extras (product_id, stock, manufacturer, updated_at)
                        VALUES (:product_id, :stock, :manufacturer, CURRENT_TIMESTAMP)
                        ON CONFLICT (product_id) DO UPDATE
                        SET stock = EXCLUDED.stock,
                            manufacturer = EXCLUDED.manufacturer,
                            updated_at = CURRENT_TIMESTAMP
                        """
                    ),
                    {"product_id": product_id, "stock": row["stock"], "manufacturer": row["manufacturer"]},
                )
                created += 1
                continue

            processed_existing += 1
            old_cost = Decimal(str(existing["cost"] or 0))
            changed = (
                txt(existing["name"]) != row["name"]
                or txt(existing["manufacturer"]) != row["manufacturer"]
                or old_cost != row["cost"]
                or (existing["stock"] is None and row["stock"] is not None)
                or (existing["stock"] is not None and row["stock"] != Decimal(str(existing["stock"])))
            )
            if changed:
                changed_existing += 1
            conn.execute(
                text("UPDATE products SET name = :name, cost = :cost WHERE id = :id"),
                {"id": existing["id"], "name": row["name"], "cost": row["cost"]},
            )
            conn.execute(
                text(
                    """
                    INSERT INTO product_extras (product_id, stock, manufacturer, updated_at)
                    VALUES (:product_id, :stock, :manufacturer, CURRENT_TIMESTAMP)
                    ON CONFLICT (product_id) DO UPDATE
                    SET stock = EXCLUDED.stock,
                        manufacturer = EXCLUDED.manufacturer,
                        updated_at = CURRENT_TIMESTAMP
                    """
                ),
                {"product_id": existing["id"], "stock": row["stock"], "manufacturer": row["manufacturer"]},
            )

        after_import = conn.execute(
            text(
                """
                SELECT
                    (SELECT count(*) FROM products) AS total,
                    (SELECT count(*) FROM products WHERE provisor_goods_id IS NOT NULL) AS with_goods_id,
                    (SELECT count(*) FROM products WHERE provisor_goods_id IS NULL) AS without_goods_id
                """
            )
        ).mappings().one()

        products = conn.execute(
            text("SELECT id, code, name, provisor_goods_id FROM products ORDER BY id")
        ).mappings().all()
        product_by_code = {row["code"]: row for row in products}
        unresolved_codes = {row["code"] for row in products if row["provisor_goods_id"] is None}

        reference_rows = conn.execute(
            text(
                """
                SELECT
                    i.price_list_id,
                    i.provisor_goods_id,
                    i.filial_id,
                    i.distributor_goods_id,
                    i.name,
                    i.raw_json,
                    l.account_id,
                    l.account_login,
                    l.branch_name
                FROM competitor_price_list_items i
                JOIN competitor_price_lists l ON l.id = i.price_list_id
                WHERE l.source_type = 'provisor'
                  AND i.filial_id = ANY(:filials)
                  AND i.provisor_goods_id IS NOT NULL
                  AND i.distributor_goods_id <> ''
                """
            ),
            {"filials": list(REFERENCE_FILIALS)},
        ).mappings().all()

        refs: dict[str, dict[int, set[int]]] = defaultdict(lambda: defaultdict(set))
        ref_examples: dict[tuple[str, int, int], dict[str, Any]] = {}
        for raw in reference_rows:
            code = normalize_sku(raw["distributor_goods_id"])
            goods_id = int(raw["provisor_goods_id"]) if raw["provisor_goods_id"] is not None else None
            filial_id = int(raw["filial_id"]) if raw["filial_id"] is not None else None
            if not code or not goods_id or filial_id not in REFERENCE_FILIALS:
                continue
            refs[code][filial_id].add(goods_id)
            ref_examples.setdefault((code, filial_id, goods_id), dict(raw))

        candidates: dict[int, dict[str, Any]] = {}
        conflicts: list[dict[str, Any]] = []
        filled_via_reference = Counter()
        for code in sorted(unresolved_codes):
            product = product_by_code[code]
            by_filial = refs.get(code)
            if not by_filial:
                continue
            all_goods = sorted({goods for goods_set in by_filial.values() for goods in goods_set})
            if len(all_goods) != 1:
                conflicts.append(
                    {
                        "stage": "reference",
                        "product_id": product["id"],
                        "product_code": code,
                        "product_name": product["name"],
                        "conflicting_goodsIds": "; ".join(str(x) for x in all_goods),
                        "source_filialIds": "; ".join(str(x) for x in sorted(by_filial)),
                        "reason": "multiple_reference_goodsIds",
                    }
                )
                continue
            goods_id = all_goods[0]
            filials = sorted(fid for fid, goods_set in by_filial.items() if goods_id in goods_set)
            source_filial = filials[0]
            example = ref_examples.get((code, source_filial, goods_id), {})
            method = f"inkar_{source_filial}"
            if 128 in filials:
                method = "inkar_128"
            elif 133 in filials:
                method = "inkar_133"
            filled_via_reference[method] += 1
            candidates[int(product["id"])] = {
                "stage": "reference",
                "method": method,
                "product_id": product["id"],
                "product_code": code,
                "product_name": product["name"],
                "old_goodsId": None,
                "new_goodsId": goods_id,
                "PharmCenter.name": "",
                "Provisor.fullName": txt(example.get("name")),
                "source account": txt(example.get("account_login") or example.get("account_id")),
                "source filial": "; ".join(str(x) for x in filials),
                "source_price_list_id": txt(example.get("price_list_id")),
            }

        unresolved_after_reference = [
            row for row in products if row["provisor_goods_id"] is None and int(row["id"]) not in candidates
        ]

        pharm_rows = fetch_pharmcenter()
        pharm_by_code: dict[str, dict[str, Any]] = {}
        for row in pharm_rows:
            code = normalize_sku(row.get("id"))
            if code and code not in pharm_by_code:
                pharm_by_code[code] = row

        provisor_rows = conn.execute(
            text(
                """
                SELECT
                    i.price_list_id,
                    i.provisor_goods_id,
                    i.filial_id,
                    i.name,
                    i.raw_json,
                    l.account_id,
                    l.account_login,
                    l.branch_name
                FROM competitor_price_list_items i
                JOIN competitor_price_lists l ON l.id = i.price_list_id
                WHERE l.source_type = 'provisor'
                  AND i.provisor_goods_id IS NOT NULL
                """
            )
        ).mappings().all()
        provisor_by_name: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for raw in provisor_rows:
            row = dict(raw)
            payload = parse_raw_json(row.get("raw_json"))
            goods = payload.get("goods") if isinstance(payload.get("goods"), dict) else {}
            full_name = first_nonempty(
                nested_get(payload, ("goods", "fullName")),
                payload.get("goodsFullName"),
                payload.get("fullName"),
                nested_get(payload, ("raw", "goods", "fullName")),
                row.get("name"),
            )
            goods_id = first_nonempty(
                row.get("provisor_goods_id"),
                nested_get(payload, ("goods", "id")),
                payload.get("goodsId"),
                goods.get("goodsId") if isinstance(goods, dict) else None,
            )
            norm = normalize_name(full_name)
            if not norm or norm == UNKNOWN_GOODS_NAME or not goods_id:
                continue
            row["goods_id"] = int(goods_id)
            row["provisor_full_name"] = full_name
            provisor_by_name[norm].append(row)

        pharmcenter_stats = Counter()
        for product in unresolved_after_reference:
            pharm = pharm_by_code.get(product["code"])
            if not pharm:
                pharmcenter_stats["pharmcenter_not_found"] += 1
                continue
            pharm_name = txt(pharm.get("name"))
            pharm_norm = normalize_name(pharm_name)
            if not pharm_norm:
                pharmcenter_stats["pharmcenter_name_empty"] += 1
                continue
            matches = provisor_by_name.get(pharm_norm, [])
            if not matches:
                pharmcenter_stats["provisor_name_not_found"] += 1
                continue
            goods_ids = sorted({int(row["goods_id"]) for row in matches})
            if len(goods_ids) != 1:
                pharmcenter_stats["conflict_multiple_goodsIds"] += 1
                conflicts.append(
                    {
                        "stage": "pharmcenter_bridge",
                        "product_id": product["id"],
                        "product_code": product["code"],
                        "product_name": product["name"],
                        "pharmcenter_name": pharm_name,
                        "conflicting_goodsIds": "; ".join(str(x) for x in goods_ids),
                        "provisor_fullNames": " | ".join(sorted({txt(m["provisor_full_name"]) for m in matches})),
                        "source_examples": json.dumps(compact_sources(matches), ensure_ascii=False),
                        "reason": "multiple_goodsIds_for_exact_name",
                    }
                )
                continue
            example = sorted(matches, key=lambda m: int(m["price_list_id"]))[0]
            candidates[int(product["id"])] = {
                "stage": "pharmcenter_bridge",
                "method": "pharmcenter_bridge",
                "product_id": product["id"],
                "product_code": product["code"],
                "product_name": product["name"],
                "old_goodsId": None,
                "new_goodsId": goods_ids[0],
                "PharmCenter.name": pharm_name,
                "Provisor.fullName": example["provisor_full_name"],
                "source account": txt(example.get("account_login") or example.get("account_id")),
                "source filial": txt(example.get("filial_id") or example.get("branch_name")),
                "source_price_list_id": txt(example.get("price_list_id")),
            }
            pharmcenter_stats["safe_unique_goodsId"] += 1

        safe_candidates = sorted(candidates.values(), key=lambda row: (row["stage"], int(row["product_id"])))
        backup_rows = safe_candidates
        backup_fields = [
            "product_id",
            "product_code",
            "product_name",
            "old_goodsId",
            "new_goodsId",
            "PharmCenter.name",
            "Provisor.fullName",
            "source account",
            "source filial",
            "source_price_list_id",
            "stage",
            "method",
        ]
        write_csv(backup_csv, backup_rows, backup_fields)
        backup_json.write_text(
            json.dumps(
                {
                    "created_at": datetime.now().isoformat(timespec="seconds"),
                    "database_url": safe_url,
                    "rows": backup_rows,
                },
                ensure_ascii=False,
                indent=2,
            ),
            encoding="utf-8",
        )

        updated_rows = 0
        for row in safe_candidates:
            result = conn.execute(
                text(
                    """
                    UPDATE products
                    SET provisor_goods_id = :goods_id
                    WHERE id = :product_id
                      AND provisor_goods_id IS NULL
                    """
                ),
                {"goods_id": row["new_goodsId"], "product_id": row["product_id"]},
            )
            updated_rows += result.rowcount
        if updated_rows != len(safe_candidates):
            raise RuntimeError(f"Expected to update {len(safe_candidates)} rows, updated {updated_rows}; rolling back")

        after_matching = conn.execute(
            text(
                """
                SELECT
                    (SELECT count(*) FROM products) AS total,
                    (SELECT count(*) FROM products WHERE provisor_goods_id IS NOT NULL) AS with_goods_id,
                    (SELECT count(*) FROM products WHERE provisor_goods_id IS NULL) AS without_goods_id
                """
            )
        ).mappings().one()

        new_goods_ids = sorted({int(row["new_goodsId"]) for row in safe_candidates})
        selected_presence_rows = []
        if new_goods_ids:
            selected_presence_rows = conn.execute(
                text(
                    """
                    WITH selected_lists AS (
                        SELECT DISTINCT l.id
                        FROM competitor_price_lists l
                        JOIN price_format_competitor_assignments a
                          ON a.competitor_price_list_id = l.id
                         AND a.is_active = TRUE
                        WHERE l.source_type = 'provisor'
                    )
                    SELECT DISTINCT i.provisor_goods_id
                    FROM competitor_price_list_items i
                    JOIN selected_lists sl ON sl.id = i.price_list_id
                    WHERE i.provisor_goods_id = ANY(:goods_ids)
                    """
                ),
                {"goods_ids": new_goods_ids},
            ).scalars().all()
        present_goods_ids = sorted({int(x) for x in selected_presence_rows if x is not None})
        absent_goods_ids = sorted(set(new_goods_ids) - set(present_goods_ids))

    by_method = Counter(row["method"] for row in safe_candidates)
    import_summary = {
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "database_url": safe_url,
        "catalog": catalog_meta,
        "products_before_import": int(before["total"]),
        "products_after_import": int(after_import["total"]),
        "new_products_added": created,
        "existing_products_processed": processed_existing,
        "existing_products_changed": changed_existing,
        "duplicate_skus_detected": catalog_meta["duplicate_skus_count"],
    }
    coverage_summary = {
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "database_url": safe_url,
        "products_before_import": dict(before),
        "products_after_import": dict(after_import),
        "products_after_matching": dict(after_matching),
        "coverage_before_import_pct": (int(before["with_goods_id"]) / int(before["total"]) * 100) if before["total"] else 0,
        "coverage_after_import_before_matching_pct": (
            int(after_import["with_goods_id"]) / int(after_import["total"]) * 100
        )
        if after_import["total"]
        else 0,
        "coverage_after_matching_pct": (
            int(after_matching["with_goods_id"]) / int(after_matching["total"]) * 100
        )
        if after_matching["total"]
        else 0,
        "filled_via": {
            "Inkar 128": by_method["inkar_128"],
            "Inkar 133": by_method["inkar_133"],
            "Inkar 108": 0,
            "PharmCenter bridge": by_method["pharmcenter_bridge"],
        },
        "reference_matching": {
            "filials_used": list(REFERENCE_FILIALS),
            "matched_via_128": by_method["inkar_128"],
            "matched_via_133": by_method["inkar_133"],
            "matched_via_108": 0,
            "note_108": "Not used; optional filial 108 was left untouched.",
        },
        "pharmcenter_bridge": dict(pharmcenter_stats),
        "conflicts": len(conflicts),
        "skipped_rows": {
            "safe_candidate_update_skips": 0,
            "pharmcenter_not_found": pharmcenter_stats["pharmcenter_not_found"],
            "pharmcenter_name_empty": pharmcenter_stats["pharmcenter_name_empty"],
            "provisor_name_not_found": pharmcenter_stats["provisor_name_not_found"],
        },
        "remaining_unresolved": int(after_matching["without_goods_id"]),
        "estimated_competitor_coverage_impact": {
            "newly_assigned_goodsIds": len(new_goods_ids),
            "present_in_selected_competitor_lists": len(present_goods_ids),
            "absent_from_selected_competitor_lists": len(absent_goods_ids),
            "expected_competitor_coverage_improvement_products": len(present_goods_ids),
            "note": "Estimate only; no price list was generated or rebuilt.",
        },
        "backup_csv": str(backup_csv),
        "backup_json": str(backup_json),
        "no_changes_made_to": [
            "pricing",
            "MDC logic",
            "shoulder analytics",
            "Provisor parser",
            "SourceGoodsMatch",
            "competitor matching logic",
        ],
    }

    write_csv(safe_candidates_csv, safe_candidates, backup_fields)
    write_csv(
        conflicts_csv,
        conflicts,
        [
            "stage",
            "product_id",
            "product_code",
            "product_name",
            "pharmcenter_name",
            "conflicting_goodsIds",
            "provisor_fullNames",
            "source_filialIds",
            "source_examples",
            "reason",
        ],
    )
    import_summary_json.write_text(json.dumps(import_summary, ensure_ascii=False, indent=2), encoding="utf-8")
    coverage_summary_json.write_text(json.dumps(coverage_summary, ensure_ascii=False, indent=2), encoding="utf-8")

    output = {
        "import_summary": import_summary,
        "coverage_summary": coverage_summary,
        "files": {
            "safe_candidates.csv": str(safe_candidates_csv),
            "conflicts.csv": str(conflicts_csv),
            "coverage_summary.json": str(coverage_summary_json),
            "import_summary.json": str(import_summary_json),
        },
    }
    print(json.dumps(output, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
