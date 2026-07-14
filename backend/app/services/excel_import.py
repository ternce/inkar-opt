from __future__ import annotations

import io
import re
from datetime import datetime, date
from decimal import Decimal

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from ..models import (
    Product,
    PriceFormat,
    MarkupRange,
    UniversalList,
    ListItem,
    CompetitorPrice,
)
from .competitor_assignments import propagate_emit_assignments_to_new_price_format
from .sku import normalize_sku


def _norm_header(value: object) -> str:
    if value is None:
        return ""
    s = str(value).strip().lower()
    s = re.sub(r"\s+", " ", s)
    return s


def _first_nonempty_row(ws, max_scan: int = 25) -> int | None:
    for r in range(1, min(ws.max_row, max_scan) + 1):
        if any(ws.cell(row=r, column=c).value not in (None, "") for c in range(1, ws.max_column + 1)):
            return r
    return None


def _headers_map(ws) -> tuple[int, dict[str, int]]:
    header_row = _first_nonempty_row(ws)
    if header_row is None:
        return 1, {}

    headers: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        key = _norm_header(ws.cell(row=header_row, column=c).value)
        if key:
            headers[key] = c
    return header_row, headers


def _get(ws, row: int, headers: dict[str, int], *keys: str):
    for k in keys:
        col = headers.get(_norm_header(k))
        if col:
            return ws.cell(row=row, column=col).value
    return None


def _as_decimal(value: object, default: Decimal | None = None) -> Decimal | None:
    if value is None or value == "":
        return default
    try:
        if isinstance(value, (int, float, Decimal)):
            return Decimal(str(value))
        # excel may provide strings with commas
        s = str(value).strip().replace(" ", "").replace(",", ".")
        return Decimal(s)
    except Exception:
        return default


def _as_date(value: object) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    s = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            pass
    return None


def _detect_sheet_type(headers: set[str]) -> str | None:
    h = {x for x in headers if x}

    # products
    if ("код" in h or "code" in h or "product_code" in h) and (
        "наименование" in h or "name" in h or "товар" in h
    ) and ("себестоимость" in h or "cost" in h or "закуп" in h):
        return "products"

    # price formats
    if ("цф" in h or "price_format" in h or "format" in h or "код цф" in h) and (
        "филиал" in h or "branch" in h or "наименование" in h or "name" in h
    ):
        return "price_formats"

    # markup ranges
    if ("стоимость от" in h or "cost_from" in h or "нижняя граница" in h) and (
        "наценка" in h or "markup" in h or "markup_percent" in h
    ):
        return "markup_ranges"

    # universal lists
    if ("тип" in h or "type" in h) and ("статус" in h or "status" in h) and (
        "название" in h or "name" in h
    ):
        return "universal_lists"

    # list items
    if ("product_code" in h or "код" in h or "код товара" in h) and ("value" in h or "значение" in h):
        return "list_items"

    # competitor prices
    if ("source_name" in h or "источник" in h) and ("source_price" in h or "цена" in h) and (
        "product_code" in h or "код" in h
    ):
        return "competitor_prices"

    # competitor sources config
    if ("source_name" in h or "источник" in h) and ("coefficient" in h or "коэффициент" in h):
        return "competitor_sources"

    return None


def import_excel(*, db: Session, content: bytes) -> dict[str, int]:
    wb = load_workbook(io.BytesIO(content), data_only=True)

    counts = {
        "price_formats": 0,
        "markup_ranges": 0,
        "products": 0,
        "universal_lists": 0,
        "list_items": 0,
        "competitor_sources": 0,
        "competitor_prices": 0,
    }

    # Process sheets independently; detect by headers
    for ws in wb.worksheets:
        header_row, headers = _headers_map(ws)
        if not headers:
            continue

        kind = _detect_sheet_type(set(headers.keys()))
        if not kind:
            continue

        if kind == "products":
            for r in range(header_row + 1, ws.max_row + 1):
                code = _get(ws, r, headers, "код", "code", "product_code")
                name = _get(ws, r, headers, "наименование", "name", "товар")
                cost = _get(ws, r, headers, "себестоимость", "cost", "закуп")

                if code in (None, ""):
                    continue

                # Normalize SKU the same way as Provisor/ph.center persistence.
                # This prevents mismatches like "1016941" (Excel) vs "000000000001016941" (API).
                code_s = normalize_sku(code) or str(code).strip()
                name_s = str(name).strip() if name not in (None, "") else code_s
                cost_d = _as_decimal(cost, Decimal("0")) or Decimal("0")

                existing = db.execute(select(Product).where(Product.code == code_s)).scalars().first()
                if existing is None:
                    db.add(Product(code=code_s, name=name_s, cost=float(cost_d)))
                else:
                    existing.name = name_s
                    existing.cost = float(cost_d)

                counts["products"] += 1

        elif kind == "price_formats":
            for r in range(header_row + 1, ws.max_row + 1):
                code = _get(ws, r, headers, "код цф", "цф", "code", "price_format", "format")
                if code in (None, ""):
                    continue

                name = _get(ws, r, headers, "наименование", "name")
                branch = _get(ws, r, headers, "филиал", "branch")
                progib = _get(ws, r, headers, "прогиб", "deflection", "progib")

                code_s = str(code).strip()
                pf = db.execute(select(PriceFormat).where(PriceFormat.code == code_s)).scalars().first()
                if pf is None:
                    pf = PriceFormat(code=code_s, name=str(name).strip() if name else code_s)
                    db.add(pf)
                    db.flush()
                    propagate_emit_assignments_to_new_price_format(db=db, price_format_id=int(pf.id))

                if name not in (None, ""):
                    pf.name = str(name).strip()
                if branch not in (None, ""):
                    pf.branch = str(branch).strip()

                progib_d = _as_decimal(progib)
                if progib_d is not None:
                    pf.progib = float(progib_d)

                counts["price_formats"] += 1

        elif kind == "markup_ranges":
            # requires price_format_code
            pf_code = None
            if "код цф" in headers:
                pf_code = "код цф"
            elif "price_format_code" in headers:
                pf_code = "price_format_code"
            elif "цф" in headers:
                pf_code = "цф"

            for r in range(header_row + 1, ws.max_row + 1):
                pf_val = _get(ws, r, headers, pf_code or "код цф", "price_format_code", "цф")
                if pf_val in (None, ""):
                    continue

                cost_from = _get(ws, r, headers, "стоимость от", "cost_from", "нижняя граница")
                cost_to = _get(ws, r, headers, "стоимость до", "cost_to", "верхняя граница")
                markup = _get(ws, r, headers, "наценка", "markup", "markup_percent", "наценка %")

                pf = db.execute(select(PriceFormat).where(PriceFormat.code == str(pf_val).strip())).scalars().first()
                if not pf:
                    continue

                cost_from_d = _as_decimal(cost_from)
                if cost_from_d is None:
                    continue

                cost_to_d = _as_decimal(cost_to)
                markup_d = _as_decimal(markup, Decimal("0")) or Decimal("0")

                # Accept both 10 and 0.10. If markup > 1 => treat as percent.
                if markup_d > 1:
                    markup_d = markup_d / Decimal("100")

                # Upsert by (pf, from, to)
                existing = db.execute(
                    select(MarkupRange)
                    .where(MarkupRange.price_format_id == pf.id)
                    .where(MarkupRange.cost_from == float(cost_from_d))
                    .where(MarkupRange.cost_to == (float(cost_to_d) if cost_to_d is not None else None))
                ).scalars().first()

                if existing is None:
                    db.add(
                        MarkupRange(
                            price_format_id=pf.id,
                            cost_from=float(cost_from_d),
                            cost_to=float(cost_to_d) if cost_to_d is not None else None,
                            markup_percent=float(markup_d),
                        )
                    )
                else:
                    existing.markup_percent = float(markup_d)

                counts["markup_ranges"] += 1

        elif kind == "universal_lists":
            for r in range(header_row + 1, ws.max_row + 1):
                code = _get(ws, r, headers, "код", "code", "list_code")
                name = _get(ws, r, headers, "название", "name")
                list_type = _get(ws, r, headers, "тип", "type")
                status = _get(ws, r, headers, "статус", "status")
                start_date = _get(ws, r, headers, "дата начала", "start_date", "start")
                end_date = _get(ws, r, headers, "дата окончания", "end_date", "end")
                pf_code = _get(ws, r, headers, "код цф", "price_format_code", "цф")

                if name in (None, "") or list_type in (None, ""):
                    continue

                code_s = str(code).strip() if code not in (None, "") else None
                name_s = str(name).strip()
                type_s = str(list_type).strip()

                status_s = str(status).strip() if status not in (None, "") else "Не активный"
                if status_s.lower().startswith("актив"):
                    status_s = "Активный"

                pf_id = None
                if pf_code not in (None, ""):
                    pf = db.execute(select(PriceFormat).where(PriceFormat.code == str(pf_code).strip())).scalars().first()
                    pf_id = pf.id if pf else None

                start_d = _as_date(start_date)
                end_d = _as_date(end_date)

                ul = None
                if code_s:
                    ul = db.execute(select(UniversalList).where(UniversalList.code == code_s)).scalars().first()
                if ul is None:
                    ul = db.execute(
                        select(UniversalList).where(UniversalList.name == name_s).where(UniversalList.type == type_s)
                    ).scalars().first()

                if ul is None:
                    ul = UniversalList(code=code_s, name=name_s, type=type_s)
                    db.add(ul)
                    db.flush()

                ul.status = status_s
                ul.start_date = start_d
                ul.end_date = end_d
                ul.price_format_id = pf_id

                counts["universal_lists"] += 1

        elif kind == "list_items":
            for r in range(header_row + 1, ws.max_row + 1):
                list_code = _get(ws, r, headers, "list_code", "код списка", "код")
                list_name = _get(ws, r, headers, "list_name", "список", "название")
                product_code = _get(ws, r, headers, "product_code", "код товара", "код")
                value = _get(ws, r, headers, "value", "значение")

                if product_code in (None, "") or value in (None, ""):
                    continue

                ul = None
                if list_code not in (None, ""):
                    ul = db.execute(select(UniversalList).where(UniversalList.code == str(list_code).strip())).scalars().first()
                if ul is None and list_name not in (None, ""):
                    ul = db.execute(select(UniversalList).where(UniversalList.name == str(list_name).strip())).scalars().first()
                if ul is None:
                    continue

                sku = normalize_sku(product_code) or str(product_code).strip()
                p = db.execute(select(Product).where(Product.code == sku)).scalars().first()
                if not p:
                    continue

                v = _as_decimal(value)
                if v is None:
                    continue

                existing = db.execute(
                    select(ListItem)
                    .where(ListItem.universal_list_id == ul.id)
                    .where(ListItem.product_id == p.id)
                ).scalars().first()

                if existing is None:
                    db.add(ListItem(universal_list_id=ul.id, product_id=p.id, value=float(v)))
                else:
                    existing.value = float(v)

                counts["list_items"] += 1

        elif kind == "competitor_sources":
            for r in range(header_row + 1, ws.max_row + 1):
                pf_code = _get(ws, r, headers, "код цф", "price_format_code", "цф")
                source_name = _get(ws, r, headers, "source_name", "источник")
                coefficient = _get(ws, r, headers, "coefficient", "коэффициент")
                supplier = _get(ws, r, headers, "поставщик", "supplier")

                if pf_code in (None, "") or source_name in (None, ""):
                    continue

                pf = db.execute(select(PriceFormat).where(PriceFormat.code == str(pf_code).strip())).scalars().first()
                if not pf:
                    continue

                src = str(source_name).strip()
                coeff = _as_decimal(coefficient, Decimal("1")) or Decimal("1")

                existing = db.execute(
                    select(CompetitorPrice)
                    .where(CompetitorPrice.price_format_id == pf.id)
                    .where(CompetitorPrice.product_id.is_(None))
                    .where(CompetitorPrice.source_name == src)
                ).scalars().first()

                if existing is None:
                    existing = CompetitorPrice(
                        price_format_id=pf.id,
                        product_id=None,
                        source_name=src,
                        coefficient=float(coeff),
                    )
                    db.add(existing)

                existing.coefficient = float(coeff)
                if supplier not in (None, ""):
                    existing.supplier = str(supplier).strip()

                counts["competitor_sources"] += 1

        elif kind == "competitor_prices":
            for r in range(header_row + 1, ws.max_row + 1):
                pf_code = _get(ws, r, headers, "код цф", "price_format_code", "цф")
                product_code = _get(ws, r, headers, "product_code", "код", "код товара")
                source_name = _get(ws, r, headers, "source_name", "источник")
                supplier = _get(ws, r, headers, "поставщик", "supplier")
                price_date = _get(ws, r, headers, "price_date", "дата", "date")
                source_price = _get(ws, r, headers, "source_price", "цена")

                if pf_code in (None, "") or product_code in (None, "") or source_name in (None, ""):
                    continue

                pf = db.execute(select(PriceFormat).where(PriceFormat.code == str(pf_code).strip())).scalars().first()
                if not pf:
                    continue

                sku = normalize_sku(product_code) or str(product_code).strip()
                p = db.execute(select(Product).where(Product.code == sku)).scalars().first()
                if not p:
                    continue

                src = str(source_name).strip()
                price_d = _as_decimal(source_price)

                existing = db.execute(
                    select(CompetitorPrice)
                    .where(CompetitorPrice.price_format_id == pf.id)
                    .where(CompetitorPrice.product_id == p.id)
                    .where(CompetitorPrice.source_name == src)
                ).scalars().first()

                if existing is None:
                    existing = CompetitorPrice(
                        price_format_id=pf.id,
                        product_id=p.id,
                        source_name=src,
                        coefficient=1.0,
                    )
                    db.add(existing)

                if supplier not in (None, ""):
                    existing.supplier = str(supplier).strip()

                existing.price_date = _as_date(price_date)
                existing.source_price = float(price_d) if price_d is not None else None

                counts["competitor_prices"] += 1

    db.commit()
    return counts
