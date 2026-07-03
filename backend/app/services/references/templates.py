from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


RATING_TEMPLATE_TYPES = {"rating_global", "rating_local"}


REFERENCE_TEMPLATE_COLUMNS: dict[str, list[str]] = {
    "cost": ["Материал", "Артикул", "Производитель", "Учетная себ"],
    "stock": ["Материал", "Артикул", "Производитель", "Остаток"],
    "products": ["Материал", "Артикул", "Производитель"],
    "holdings": ["Наименование", "Код"],
    "counterparties": ["Наименование", "Код", "Код холдинга"],
    "delivery_points": ["Наименование", "Код", "Код контрагента", "Адрес"],
}


REFERENCE_TEMPLATE_EXAMPLES: dict[str, list[object]] = {
    "cost": ["000000000001234567", "Пример товара", "Пример производителя", 1250.75],
    "stock": ["000000000001234567", "Пример товара", "Пример производителя", 24],
    "products": ["000000000001234567", "Пример товара", "Пример производителя"],
    "holdings": ["Пример холдинга", "HOLD-001"],
    "counterparties": ["Пример контрагента", "CNT-001", "HOLD-001"],
    "delivery_points": ["Пример точки доставки", "DP-001", "CNT-001", "Пример адреса 1"],
}


def reference_template_filename(data_type: str) -> str:
    return f"{data_type}_import_template.xlsx"


def build_reference_template(data_type: str) -> bytes:
    if data_type in RATING_TEMPLATE_TYPES:
        raise ValueError("rating templates are customer-provided")
    columns = REFERENCE_TEMPLATE_COLUMNS.get(data_type)
    if not columns:
        raise ValueError("reference template is not supported")

    wb = Workbook()
    ws = wb.active
    ws.title = "Шаблон импорта"
    ws.append(columns)
    ws.append(REFERENCE_TEMPLATE_EXAMPLES.get(data_type) or [""] * len(columns))
    ws.freeze_panes = "A2"

    for cell in ws[1]:
        cell.font = Font(bold=True)

    for col_idx, header in enumerate(columns, start=1):
        values = [header, ws.cell(row=2, column=col_idx).value]
        width = max(len(str(value or "")) for value in values) + 2
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(width, 12), 48)

    stream = io.BytesIO()
    wb.save(stream)
    return stream.getvalue()
