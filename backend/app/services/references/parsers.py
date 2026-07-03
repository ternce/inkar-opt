from __future__ import annotations

import io
import re
from decimal import Decimal
from typing import Any

from openpyxl import load_workbook


HEADER_ALIASES = {
    "sku": ("sku", "код", "код товара", "материал", "номенклатура код"),
    "name": ("название", "наименование", "товар", "краткий текст материала", "номенклатура", "name"),
    "manufacturer": ("производитель", "manufacturer", "бренд"),
    "stock": ("остаток", "остатки", "количество", "stock"),
    "cost": ("себестоимость", "учетная себестоимость", "учётная себестоимость", "cost", "cost price"),
    "rating_global": ("рейтинг общий", "общий рейтинг", "rating global", "global rating", "top"),
    "rating_local": ("рейтинг локальный", "локальный рейтинг", "rating local", "local rating"),
    "branch_id": ("branch_id", "филиал id", "id филиала", "код филиала", "регион id"),
    "branch_name": ("филиал", "регион", "branch", "branch_name"),
    "external_id": ("id", "код", "external_id", "внешний id"),
    "holding_id": ("holding_id", "код холдинга", "холдинг id"),
    "counterparty_id": ("counterparty_id", "код контрагента", "контрагент id"),
    "address": ("адрес", "address"),
}

HEADER_ALIASES["name"] = (*HEADER_ALIASES["name"], "product name")
HEADER_ALIASES["external_id"] = (*HEADER_ALIASES["external_id"], "external id")
HEADER_ALIASES["holding_id"] = (*HEADER_ALIASES["holding_id"], "holding id")
HEADER_ALIASES["counterparty_id"] = (*HEADER_ALIASES["counterparty_id"], "counterparty id")


def _norm(value: object) -> str:
    s = str(value or "").strip().lower().replace("ё", "е")
    return re.sub(r"\s+", " ", s)


def _first_nonempty_row(ws, max_scan: int = 30) -> int | None:
    for row in range(1, min(ws.max_row, max_scan) + 1):
        if any(ws.cell(row=row, column=col).value not in (None, "") for col in range(1, ws.max_column + 1)):
            return row
    return None


def _header_map(ws) -> tuple[int, dict[str, int]]:
    header_row = _first_nonempty_row(ws)
    if header_row is None:
        return 1, {}
    raw_headers = {_norm(ws.cell(row=header_row, column=col).value): col for col in range(1, ws.max_column + 1)}
    out: dict[str, int] = {}
    for canonical, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            key = _norm(alias)
            if key in raw_headers:
                out[canonical] = raw_headers[key]
                break
    return header_row, out


def _get(ws, row: int, headers: dict[str, int], key: str) -> Any:
    col = headers.get(key)
    if not col:
        return None
    return ws.cell(row=row, column=col).value


def as_decimal(value: object) -> Decimal | None:
    if value is None or value == "":
        return None
    try:
        if isinstance(value, (int, float, Decimal)):
            return Decimal(str(value))
        s = str(value).strip().replace(" ", "").replace(",", ".")
        s = re.sub(r"[^0-9.\-]", "", s)
        return Decimal(s) if s else None
    except Exception:
        return None


def as_int(value: object) -> int | None:
    dec = as_decimal(value)
    if dec is None:
        return None
    try:
        return int(dec)
    except Exception:
        return None


def parse_excel_rows(content: bytes) -> tuple[list[dict], set[str]]:
    wb = load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.worksheets[0] if wb.worksheets else None
    if ws is None:
        raise ValueError("Excel file has no worksheets")
    header_row, headers = _header_map(ws)
    if not headers:
        raise ValueError("Не найдена строка заголовков")

    rows: list[dict] = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        row = {key: _get(ws, row_idx, headers, key) for key in HEADER_ALIASES}
        if not any(value not in (None, "") for value in row.values()):
            continue
        row["_row"] = row_idx
        rows.append(row)
    return rows, set(headers)
