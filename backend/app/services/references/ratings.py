from __future__ import annotations

import io
import json
import re
from decimal import Decimal, InvalidOperation
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from ...models import Product, ProductRating, ReferenceImportJob, ReferenceUpdateStatus
from ...timezone import local_iso, now_kz_naive
from ..sku import normalize_external_sku, normalize_sku
from .types import BRANCH_BY_ID


RATING_DATA_TYPES = {"rating_global": "global", "rating_local": "local"}
RATING_SHEET_NAME = "отчет"
RATING_HEADER_ALIASES = {
    "rank": ("№ рейтинга",),
    "sku": ("материал",),
    "name": ("полное имя товара (vi-ortis)",),
}


def _normalize_text(value: object) -> str:
    text = str(value or "").strip().casefold().replace("ё", "е")
    return re.sub(r"\s+", " ", text)


def _header_indexes(worksheet) -> tuple[int, dict[str, int]]:
    for row_number in range(1, min(worksheet.max_row, 30) + 1):
        raw = {
            _normalize_text(worksheet.cell(row=row_number, column=column).value): column
            for column in range(1, worksheet.max_column + 1)
        }
        indexes: dict[str, int] = {}
        for canonical, aliases in RATING_HEADER_ALIASES.items():
            for alias in aliases:
                column = raw.get(_normalize_text(alias))
                if column is not None:
                    indexes[canonical] = column
                    break
        if "rank" in indexes and "sku" in indexes:
            return row_number, indexes
    raise ValueError('На листе "Отчет" нужны колонки: № рейтинга и Материал')


def _sku_values(value: object) -> tuple[str, str, list[str]]:
    if value is None or isinstance(value, bool):
        return "", "", []
    if isinstance(value, float):
        if value != value:
            return "", "", []
        value = int(value) if value.is_integer() else value
    raw = normalize_external_sku(value)
    normalized = normalize_sku(value)
    candidates = [candidate for candidate in (raw, normalized) if candidate]
    candidates = list(dict.fromkeys(candidates))
    return raw, normalized or raw, candidates


def _positive_integer(value: object) -> int | None:
    if value is None or value == "" or isinstance(value, bool):
        return None
    try:
        parsed = Decimal(str(value).strip().replace(",", "."))
        if not parsed.is_finite() or parsed <= 0 or parsed != parsed.to_integral_value():
            return None
    except (InvalidOperation, ValueError):
        return None
    return int(parsed)


def _parse_rating_rows(content: bytes, filename: str) -> tuple[list[dict[str, Any]], dict[str, int], list[dict[str, Any]]]:
    suffix = filename.lower()
    if not suffix.endswith((".xlsx", ".xlsm")):
        raise ValueError("Поддерживаются только файлы .xlsx и .xlsm")
    try:
        workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    except Exception as exc:
        raise ValueError(f"Некорректный Excel-файл: {exc}") from exc

    try:
        worksheet = next((sheet for sheet in workbook.worksheets if _normalize_text(sheet.title) == RATING_SHEET_NAME), None)
        if worksheet is None:
            raise ValueError('Не найден лист "Отчет"')
        header_row, indexes = _header_indexes(worksheet)
        summary = {"total_rows": 0, "matched": 0, "not_found": 0, "duplicates": 0, "invalid_rows": 0, "updated": 0}
        errors: list[dict[str, Any]] = []
        parsed_rows: list[dict[str, Any]] = []
        seen_skus: set[str] = set()

        for row_number, values in enumerate(worksheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
            rank_value = values[indexes["rank"] - 1] if indexes["rank"] <= len(values) else None
            sku_value = values[indexes["sku"] - 1] if indexes["sku"] <= len(values) else None
            name_column = indexes.get("name")
            name_value = values[name_column - 1] if name_column is not None and name_column <= len(values) else None
            if rank_value in (None, "") and sku_value in (None, "") and name_value in (None, ""):
                continue
            summary["total_rows"] += 1
            raw_sku, normalized_sku, candidates = _sku_values(sku_value)
            if not normalized_sku:
                summary["invalid_rows"] += 1
                errors.append({"row": row_number, "sku": raw_sku, "reason": "SKU/Материал обязателен", "type": "invalid"})
                continue
            if normalized_sku in seen_skus:
                summary["duplicates"] += 1
                errors.append({"row": row_number, "sku": raw_sku, "reason": "Дубликат SKU; использована первая строка", "type": "duplicate"})
                continue
            seen_skus.add(normalized_sku)
            rank = _positive_integer(rank_value)
            if rank is None:
                summary["invalid_rows"] += 1
                errors.append({"row": row_number, "sku": raw_sku, "reason": "№ рейтинга должен быть положительным целым числом", "type": "invalid"})
                continue
            parsed_rows.append({"row": row_number, "sku": raw_sku, "normalized_sku": normalized_sku, "candidates": candidates, "rank": rank})
        return parsed_rows, summary, errors
    finally:
        workbook.close()


def _job_payload(job: ReferenceImportJob, summary: dict[str, int], errors: list[dict[str, Any]]) -> dict[str, Any]:
    return {
        "id": job.id,
        "dataType": job.data_type,
        "branchIds": job.branch_ids_json,
        "filename": job.filename,
        "sourceType": job.source_type,
        "status": job.status,
        "rowsTotal": job.rows_total,
        "rowsSuccess": job.rows_success,
        "rowsFailed": job.rows_failed,
        "error": job.error,
        "log": job.log_json,
        "createdAt": local_iso(job.created_at) if job.created_at else "",
        "startedAt": local_iso(job.started_at) if job.started_at else "",
        "finishedAt": local_iso(job.finished_at) if job.finished_at else "",
        "userName": job.user_name,
        "summary": summary,
        "errors": errors,
    }


def import_top_rating_excel(
    *,
    db: Session,
    data_type: str,
    branch_ids: list[str],
    content: bytes,
    filename: str,
    user_name: str = "",
) -> dict[str, Any]:
    rating_type = RATING_DATA_TYPES.get(data_type)
    if rating_type is None:
        raise ValueError("Тип рейтинга должен быть rating_global или rating_local")
    branch_ids = list(dict.fromkeys(str(value).strip() for value in branch_ids if str(value).strip()))
    if not branch_ids:
        raise ValueError("branch_ids is required")

    parsed_rows, summary, errors = _parse_rating_rows(content, filename)
    all_codes = {code for row in parsed_rows for code in row["candidates"]}

    try:
        products = db.execute(select(Product).where(Product.code.in_(all_codes))).scalars().all() if all_codes else []
        products_by_code = {product.code: product for product in products}
        matched_rows: list[tuple[dict[str, Any], Product]] = []
        for source_row in parsed_rows:
            product = next((products_by_code.get(code) for code in source_row["candidates"] if products_by_code.get(code)), None)
            if product is None:
                summary["not_found"] += 1
                errors.append({"row": source_row["row"], "sku": source_row["sku"], "reason": "Товар с таким SKU не найден", "type": "not_found"})
                continue
            summary["matched"] += 1
            matched_rows.append((source_row, product))

        product_ids = [product.id for _source, product in matched_rows]
        existing = (
            db.execute(
                select(ProductRating)
                .where(ProductRating.product_id.in_(product_ids))
                .where(ProductRating.rating_type == rating_type)
                .where(ProductRating.branch_id.in_(branch_ids))
            )
            .scalars()
            .all()
            if product_ids
            else []
        )
        ratings_by_scope = {(row.branch_id, row.product_id): row for row in existing}
        updated_at = now_kz_naive()
        for source_row, product in matched_rows:
            for branch_id in branch_ids:
                rating = ratings_by_scope.get((branch_id, product.id))
                if rating is None:
                    rating = ProductRating(branch_id=branch_id, product_id=product.id, sku=product.code, rating_type=rating_type)
                    db.add(rating)
                    ratings_by_scope[(branch_id, product.id)] = rating
                rating.rating = source_row["rank"]
                rating.source_type = "excel"
                rating.updated_at = updated_at
            if rating_type == "global":
                product.top_rank = source_row["rank"]

        summary["updated"] = len(matched_rows)
        failed = summary["not_found"] + summary["duplicates"] + summary["invalid_rows"]
        status = "success" if failed == 0 else "partial"
        job = ReferenceImportJob(
            data_type=data_type,
            branch_ids_json=json.dumps(branch_ids, ensure_ascii=False),
            filename=filename,
            source_type="excel",
            status=status,
            rows_total=summary["total_rows"],
            rows_success=summary["updated"],
            rows_failed=failed,
            log_json=json.dumps(errors[:500], ensure_ascii=False),
            started_at=updated_at,
            finished_at=updated_at,
            user_name=user_name,
        )
        db.add(job)
        for branch_id in branch_ids:
            reference_status = (
                db.execute(
                    select(ReferenceUpdateStatus)
                    .where(ReferenceUpdateStatus.branch_id == branch_id)
                    .where(ReferenceUpdateStatus.data_type == data_type)
                )
                .scalars()
                .first()
            )
            if reference_status is None:
                reference_status = ReferenceUpdateStatus(branch_id=branch_id, data_type=data_type)
                db.add(reference_status)
            reference_status.branch_name = BRANCH_BY_ID.get(branch_id, {}).get("name", branch_id)
            reference_status.last_updated_at = updated_at
            reference_status.rows_count = summary["updated"]
            reference_status.status = status
            reference_status.error = ""
        db.commit()
        db.refresh(job)
        return _job_payload(job, summary, errors)
    except Exception:
        db.rollback()
        raise
