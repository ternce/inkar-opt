from __future__ import annotations

import csv
import io
import json
import logging
import time
from datetime import date, datetime
from decimal import Decimal
from typing import Any

from openpyxl import Workbook, load_workbook
from sqlalchemy import delete, func, select
from sqlalchemy.orm import Session

from ..models import (
    BranchCost,
    BranchStock,
    CompetitorPrice,
    CompetitorCodeMapping,
    CompetitorPriceList,
    CompetitorPriceListItem,
    PriceSourceAccount,
    PriceFormat,
    Product,
    ProductRating,
    SourceGoodsMatch,
)
from .competitor_persist import _ensure_price_format
from .competitor_matching import (
    PROVISOR_REFERENCE_FILIAL_ID,
    PROVISOR_REFERENCE_FILIAL_IDS,
    _add_reference_lookup,
    _materialize_match_fields,
    _reference_sku_lookup_keys,
    _sync_provisor_reference_mapping_from_items,
    rebuild_competitor_prices_for_selected,
    rematch_price_list_items_by_product,
)
from .competitor_assignments import (
    get_assigned_competitor_price_lists,
    get_assignment,
    list_global_competitor_price_lists_for_format,
    set_competitor_assignments,
)
from .competitor_percentiles import DEFAULT_BRANCH, recalculate_competitor_percentiles
from .manufacturers import resolve_manufacturer
from .price_sources import UnifiedPriceItem, UnifiedPriceList
from .sku import normalize_external_sku, normalize_sku, normalize_sku_variants

logger = logging.getLogger(__name__)


def _timing(operation: str, step: str, started_at: float) -> None:
    logger.info("[TIMING] operation=%s step=%s elapsed_ms=%s", operation, step, round((time.perf_counter() - started_at) * 1000, 2))


def _db_save_timing(*, price_list_id: int | str, stage: str, rows: int, started_at: float) -> None:
    logger.info(
        "[DB_SAVE_TIMING] price_list_id=%s stage=%s rows=%s elapsed_ms=%s",
        price_list_id,
        stage,
        rows,
        round((time.perf_counter() - started_at) * 1000, 2),
    )


def _prepare_rows_timing(*, price_list_id: int | str, stage: str, rows: int, started_at: float) -> None:
    logger.info(
        "[PREPARE_ROWS_TIMING] price_list_id=%s stage=%s rows=%s elapsed_ms=%s",
        price_list_id,
        stage,
        rows,
        round((time.perf_counter() - started_at) * 1000, 2),
    )


PRICE_LIST_EXPORT_COLUMNS = [
    ("provisor_id", "Идентификатор Провизора"),
    ("filial_id", "Идентификатор филиала прайса"),
    ("name", "Наименование"),
    ("reg_number", "Регистрационный номер"),
    ("manufacturer", "Производитель"),
    ("distributor_goods_name", "Наименование у дистрибьютора"),
    ("distributor_goods_id", "Идентификатор дистрибьютора"),
    ("distributor_price", "Цена дистрибьютора"),
    ("stock", "Остаток"),
    ("package_count", "Количество упаковок"),
]


def _as_decimal(value: object) -> Decimal | None:
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    s = s.replace(" ", "").replace(",", ".")
    try:
        return Decimal(s)
    except Exception:
        return None


def _as_int(value: object) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(float(str(value).strip()))
    except Exception:
        return None


def _source_name(row: CompetitorPriceList) -> str:
    return f"{row.source_type}:{row.source_key}"


def _source_key_for_unified(price_list: UnifiedPriceList) -> str:
    return f"{price_list.account_id}:{price_list.price_list_id}"


def _parse_region_status(value: str | None) -> str:
    text = str(value or "")
    if "status:" not in text:
        return ""
    return text.split("status:", 1)[1].split(";", 1)[0].strip()


def _parse_source_updated_at(value: object) -> datetime | None:
    text = str(value or "").strip()
    if not text:
        return None
    candidates = [text, text[:19], text[:16]]
    for fmt in ("%d.%m.%Y %H:%M:%S", "%d.%m.%Y %H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S"):
        try:
            return next(datetime.strptime(candidate, fmt) for candidate in candidates if candidate)
        except Exception:
            continue
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00")).replace(tzinfo=None)
    except Exception:
        return None


def _move_vidman_selection_to_latest_duplicate(db: Session, row: CompetitorPriceList) -> None:
    if row.source_type != "vidman" or not row.branch_name or not row.competitor_name:
        return

    duplicates = (
        db.execute(
            select(CompetitorPriceList)
            .where(CompetitorPriceList.price_format_id == row.price_format_id)
            .where(CompetitorPriceList.source_type == row.source_type)
            .where(CompetitorPriceList.branch_name == row.branch_name)
            .where(CompetitorPriceList.competitor_name == row.competitor_name)
            .where(CompetitorPriceList.id != row.id)
        )
        .scalars()
        .all()
    )
    selected_duplicates = [item for item in duplicates if item.is_selected]
    if not selected_duplicates:
        return

    row_date = _parse_source_updated_at(row.source_updated_at)
    selected_date = max(
        (_parse_source_updated_at(item.source_updated_at) for item in selected_duplicates),
        default=None,
    )
    if selected_date is not None and (row_date is None or row_date < selected_date):
        return

    selected_source = max(
        selected_duplicates,
        key=lambda item: _parse_source_updated_at(item.source_updated_at) or datetime.min,
    )
    row.is_selected = True
    row.coefficient = selected_source.coefficient
    for item in selected_duplicates:
        item.is_selected = False
    logger.info(
        "[VW_PRICE_LIST_SELECTION_MOVED] account_id=%s old_account_id=%s old_price_id=%s new_price_id=%s price_name=%s old_date=%s new_date=%s",
        row.account_id,
        selected_source.account_id,
        selected_source.external_price_list_id or selected_source.source_key,
        row.external_price_list_id or row.source_key,
        row.competitor_name,
        selected_source.source_updated_at or "",
        row.source_updated_at or "",
    )


def _visible_competitor_price_list_rows(
    rows: list[CompetitorPriceList],
    counts: dict[int, int],
) -> list[CompetitorPriceList]:
    return [row for row in rows if int(counts.get(row.id, 0)) > 0]


def _price_list_assignment_hidden_reason(row: CompetitorPriceList, counts: dict[int, int]) -> str:
    real_items_count = int(counts.get(row.id, 0))
    if real_items_count == 0:
        return "real_items_count_zero"
    return ""


def _product_sku_map(db: Session) -> dict[str, Product]:
    products = db.execute(select(Product)).scalars().all()
    out: dict[str, Product] = {}
    for product in products:
        for key in [
            normalize_external_sku(product.code),
            normalize_sku(product.code),
            *normalize_sku_variants(product.code),
        ]:
            if key:
                out.setdefault(key, product)
    return out


def _identity_text(value: object) -> str:
    return str(value or "").strip().casefold()


def _competitor_item_identity_keys(
    *,
    source: str,
    price_list_id: object,
    provisor_id: object = None,
    provisor_goods_id: object = None,
    distributor_goods_id: object = None,
    name: object = None,
    producer: object = None,
    shelf_life: object = None,
    batch: object = None,
) -> list[tuple[str, ...]]:
    source_key = _identity_text(source)
    goods_key = _identity_text(_as_int(provisor_goods_id))
    provisor_key = _identity_text(_as_int(provisor_id))
    distributor_key = _identity_text(distributor_goods_id)
    name_key = _identity_text(name)
    producer_key = _identity_text(producer)
    shelf_key = _identity_text(shelf_life)
    batch_key = _identity_text(batch)
    keys: list[tuple[str, ...]] = []
    if goods_key and distributor_key:
        keys.append(("goods_distributor", source_key, goods_key, distributor_key, shelf_key, batch_key))
    if goods_key and name_key:
        keys.append(("goods_text", source_key, goods_key, name_key, producer_key, shelf_key, batch_key))
    if provisor_key:
        keys.append(("provisor_id", source_key, provisor_key))
    if distributor_key:
        keys.append(("distributor", source_key, distributor_key, shelf_key, batch_key))
    if name_key and producer_key:
        keys.append(("text", source_key, name_key, producer_key, shelf_key, batch_key))
    return keys


def _identity_keys_for_existing_item(item: CompetitorPriceListItem, source: str) -> list[tuple[str, ...]]:
    return _competitor_item_identity_keys(
        source=source,
        price_list_id=item.filial_id or item.price_list_id,
        provisor_id=item.provisor_id,
        provisor_goods_id=item.provisor_goods_id,
        distributor_goods_id=item.distributor_goods_id,
        name=item.distributor_goods_name or item.name,
        producer=item.raw_manufacturer,
        shelf_life=item.expiry_date,
    )


def _identity_keys_for_unified_item(item: UnifiedPriceItem) -> list[tuple[str, ...]]:
    raw = item.raw if isinstance(item.raw, dict) else {}
    item_name = item.product_name or item.distributor_product_name
    return _competitor_item_identity_keys(
        source=item.source,
        price_list_id=item.price_list_id,
        provisor_id=raw.get("id"),
        provisor_goods_id=raw.get("goodsId"),
        distributor_goods_id=item.distributor_product_id,
        name=item.distributor_product_name or item.product_name,
        producer=resolve_manufacturer(item.manufacturer, item_name, default=""),
        shelf_life=item.expiry_date,
        batch=raw.get("batch") or raw.get("series") or raw.get("batchNumber"),
    )


def _clean_expiry_date(value: object) -> str:
    return str(value or "").strip()[:64]


def _empty_match_structure_fields() -> dict[str, object]:
    return {
        "match_key": "",
        "normalized_name": "",
        "normalized_manufacturer": "",
        "parsed_base_name": "",
        "parsed_form": "",
        "parsed_forms_json": "",
        "parsed_dosage": None,
        "parsed_dosage_volume": None,
        "parsed_quantity": None,
        "parsed_volume": None,
        "parsed_weight": None,
        "parsed_percent_strength": None,
        "parsed_concentration": None,
        "parsed_iu_dosage": None,
        "parsed_strength_signature": "",
        "parsed_dimensions_json": "",
        "parsed_critical_tokens_json": "",
    }


def _manufacturer_cache_key(raw: object, name: object) -> tuple[str, str]:
    raw_text = str(raw or "")
    return (raw_text, "" if raw_text.strip() else str(name or ""))


def _update_provisor_reference_mapping_from_records(
    *,
    db: Session,
    account_id: str,
    records: list[tuple[object, object]],
) -> dict[str, int]:
    products_by_id: dict[int, Product] = {}
    lookup_paths: list[tuple[str, dict[str, set[int]]]] = [
        ("ProductCode", {}),
        ("SourceGoodsMatch", {}),
        ("CompetitorCodeMapping", {}),
        ("BranchStock", {}),
        ("BranchCost", {}),
        ("ProductRating", {}),
    ]
    lookups = {name: lookup for name, lookup in lookup_paths}
    for product in db.execute(select(Product)).scalars().all():
        products_by_id[int(product.id)] = product
        _add_reference_lookup(lookups["ProductCode"], product.code, product.id)
    for distributor_goods_id, product_id in db.execute(
        select(SourceGoodsMatch.distributor_goods_id, SourceGoodsMatch.product_id)
        .where(SourceGoodsMatch.source_type == "provisor")
        .where(SourceGoodsMatch.distributor_goods_id != "")
    ).all():
        _add_reference_lookup(lookups["SourceGoodsMatch"], distributor_goods_id, product_id)
    for source_external_key, product_id in db.execute(
        select(CompetitorCodeMapping.source_external_key, CompetitorCodeMapping.our_product_id)
        .where(CompetitorCodeMapping.platform == "provisor")
        .where(CompetitorCodeMapping.source_external_key.is_not(None))
        .where(CompetitorCodeMapping.our_product_id.is_not(None))
    ).all():
        _add_reference_lookup(lookups["CompetitorCodeMapping"], source_external_key, product_id)
    for sku, product_id in db.execute(select(BranchStock.sku, BranchStock.product_id).where(BranchStock.sku != "")).all():
        _add_reference_lookup(lookups["BranchStock"], sku, product_id)
    for sku, product_id in db.execute(select(BranchCost.sku, BranchCost.product_id).where(BranchCost.sku != "")).all():
        _add_reference_lookup(lookups["BranchCost"], sku, product_id)
    for sku, product_id in db.execute(select(ProductRating.sku, ProductRating.product_id).where(ProductRating.sku != "")).all():
        _add_reference_lookup(lookups["ProductRating"], sku, product_id)

    mapped = 0
    updated = 0
    skipped_empty = 0
    skipped_product_not_found = 0
    ambiguous_match = 0
    conflict = 0
    updated_by_path = {
        "ProductCode": 0,
        "SourceGoodsMatch": 0,
        "CompetitorCodeMapping": 0,
        "BranchStock": 0,
        "BranchCost": 0,
        "ProductRating": 0,
    }
    for distributor_goods_id, goods_id_raw in records:
        goods_id = _as_int(goods_id_raw)
        if not goods_id:
            skipped_empty += 1
            continue
        row_keys = _reference_sku_lookup_keys(distributor_goods_id)
        product = None
        matched_path = ""
        row_ambiguous = False
        for path_name, lookup in lookup_paths:
            product_ids: set[int] = set()
            for key in row_keys:
                product_ids.update(lookup.get(key, set()))
            if len(product_ids) > 1:
                ambiguous_match += 1
                row_ambiguous = True
                break
            if len(product_ids) == 1:
                product = products_by_id.get(next(iter(product_ids)))
                matched_path = path_name
                break
        if row_ambiguous:
            continue
        if product is None:
            skipped_product_not_found += 1
            continue
        mapped += 1
        product_goods_id = _as_int(product.provisor_goods_id)
        if product_goods_id and product_goods_id != goods_id:
            conflict += 1
            continue
        if not product_goods_id:
            product.provisor_goods_id = goods_id
            updated += 1
            if matched_path in updated_by_path:
                updated_by_path[matched_path] += 1
    logger.info(
        "[PROVISOR_REFERENCE] account_id=%s reference_filial_id=%s available=true mapped_products=%s updated_products=%s skipped_empty_goods_id=%s skipped_product_not_found=%s ambiguous_match=%s conflict=%s updatedViaProductCode=%s updatedViaSourceGoodsMatch=%s updatedViaCompetitorCodeMapping=%s updatedViaBranchStock=%s updatedViaBranchCost=%s updatedViaProductRating=%s",
        account_id or "",
        PROVISOR_REFERENCE_FILIAL_ID,
        mapped,
        updated,
        skipped_empty,
        skipped_product_not_found,
        ambiguous_match,
        conflict,
        updated_by_path["ProductCode"],
        updated_by_path["SourceGoodsMatch"],
        updated_by_path["CompetitorCodeMapping"],
        updated_by_path["BranchStock"],
        updated_by_path["BranchCost"],
        updated_by_path["ProductRating"],
    )
    return {
        "mapped_products": mapped,
        "updated_products": updated,
        "skipped_empty_goods_id": skipped_empty,
        "skipped_product_not_found": skipped_product_not_found,
        "updatedViaProductCode": updated_by_path["ProductCode"],
        "updatedViaSourceGoodsMatch": updated_by_path["SourceGoodsMatch"],
        "updatedViaCompetitorCodeMapping": updated_by_path["CompetitorCodeMapping"],
        "updatedViaBranchStock": updated_by_path["BranchStock"],
        "updatedViaBranchCost": updated_by_path["BranchCost"],
        "updatedViaProductRating": updated_by_path["ProductRating"],
        "ambiguousMatch": ambiguous_match,
        "conflict": conflict,
    }


def _region_meta(region: str | None) -> dict[str, str]:
    out: dict[str, str] = {}
    for part in (region or "").split(";"):
        if ":" not in part:
            continue
        key, value = part.split(":", 1)
        out[key.strip()] = value.strip()
    return out


def _branch_name(row: CompetitorPriceList) -> str:
    meta = _region_meta(row.region)
    return (row.branch_name or meta.get("branch") or DEFAULT_BRANCH).strip() or DEFAULT_BRANCH


def _competitor_name(row: CompetitorPriceList) -> str:
    meta = _region_meta(row.region)
    return (row.competitor_name or meta.get("competitor") or row.supplier or row.display_name or row.source_type).strip()


def _replace_legacy_price_rows_for_list(*, db: Session, price_list: CompetitorPriceList) -> None:
    src = _source_name(price_list)
    db.execute(
        delete(CompetitorPrice)
        .where(CompetitorPrice.price_format_id == price_list.price_format_id)
        .where(CompetitorPrice.product_id.is_not(None))
        .where(CompetitorPrice.source_name == src)
    )
    items = (
        db.execute(
            select(CompetitorPriceListItem)
            .where(CompetitorPriceListItem.price_list_id == price_list.id)
            .where(CompetitorPriceListItem.product_id.is_not(None))
            .where(CompetitorPriceListItem.matched_sku != "")
            .where(CompetitorPriceListItem.distributor_price.is_not(None))
        )
        .scalars()
        .all()
    )
    for item in items:
        if item.product_id is None or item.distributor_price is None:
            continue
        db.add(
            CompetitorPrice(
                price_format_id=price_list.price_format_id,
                product_id=item.product_id,
                source_name=src,
                supplier=price_list.supplier or price_list.display_name,
                price_date=price_list.price_date,
                coefficient=1.0,
                source_price=float(item.distributor_price),
            )
        )


def _display_name_from_provisor_item(item: dict[str, Any], fallback_filial_id: int) -> str:
    filial = item.get("filial") if isinstance(item.get("filial"), dict) else {}
    name = str(filial.get("name") or "").strip()
    if name:
        return name
    return f"Provisor {fallback_filial_id}"


def upsert_provisor_price_list(
    *,
    db: Session,
    price_format_code: str,
    filial_id: int,
    items: list[dict[str, Any]],
    region: str = "",
    as_of: date | None = None,
    run_matching: bool = True,
) -> CompetitorPriceList:
    pf = _ensure_price_format(db, price_format_code)
    today = as_of or date.today()
    first = next((x for x in items if isinstance(x, dict)), {})
    display_name = _display_name_from_provisor_item(first, filial_id)

    row = (
        db.execute(
            select(CompetitorPriceList)
            .where(CompetitorPriceList.source_type == "provisor")
            .where(CompetitorPriceList.source_key == str(filial_id))
            .order_by(CompetitorPriceList.updated_at.desc(), CompetitorPriceList.id.desc())
        )
        .scalars()
        .first()
    )
    if row is None:
        row = CompetitorPriceList(
            price_format_id=pf.id,
            source_type="provisor",
            source_key=str(filial_id),
            coefficient=1.0,
        )
        db.add(row)
        db.flush()

    row.display_name = display_name
    row.supplier = display_name
    row.region = region
    row.branch_id = str(filial_id)
    row.branch_code = str(filial_id)
    row.branch_name = display_name or DEFAULT_BRANCH
    row.competitor_name = display_name
    row.account_id = ""
    row.account_login = ""
    row.external_price_list_id = str(filial_id)
    row.sync_batch_id = datetime.utcnow().strftime("%Y%m%d%H%M%S%f")
    row.price_date = today
    row.updated_at = datetime.utcnow()
    _move_vidman_selection_to_latest_duplicate(db, row)

    db.execute(delete(CompetitorPriceListItem).where(CompetitorPriceListItem.price_list_id == row.id))

    code_to_id: dict[str, int] = {}
    if run_matching:
        codes = list(
            dict.fromkeys(
                normalize_sku(x.get("distributorGoodsId")) or ""
                for x in items
                if isinstance(x, dict)
            )
        )
        codes = [x for x in codes if x]
        code_to_id = dict(db.execute(select(Product.code, Product.id).where(Product.code.in_(codes))).all()) if codes else {}

    for item in items:
        if not isinstance(item, dict):
            continue
        goods = item.get("goods") if isinstance(item.get("goods"), dict) else {}
        sku = normalize_sku(item.get("distributorGoodsId")) or ""
        product_id = (
            next((code_to_id[v] for v in normalize_sku_variants(item.get("distributorGoodsId")) if v in code_to_id), None)
            if run_matching
            else None
        )

        price = _as_decimal(item.get("goodsPriceWithUserDiscount"))
        if price is None or price <= 0:
            price = _as_decimal(item.get("goodsPrice"))
        expiry_date = str(item.get("shelfLife") or "").strip()
        raw_name = str(item.get("distributorGoodsName") or goods.get("fullName") or "").strip()
        raw_manufacturer = resolve_manufacturer(
            item.get("distributorProducer") or item.get("manufacturer") or item.get("producer") or goods.get("producer"),
            raw_name,
            default="",
        )
        stock = _as_decimal(item.get("stored"))
        box = _as_decimal(item.get("box"))
        pack = _as_decimal(item.get("pack"))
        if box is not None and box > 0:
            package_count = box
        elif pack is not None and pack > 0:
            package_count = pack
        else:
            package_count = None

        db.add(
            CompetitorPriceListItem(
                price_list_id=row.id,
                product_id=product_id,
                provisor_id=_as_int(item.get("id")),
                provisor_goods_id=_as_int(item.get("goodsId")),
                filial_id=_as_int(item.get("filialId")) or filial_id,
                name=raw_name,
                reg_number=str(goods.get("regNumber") or "").strip(),
                distributor_goods_name=raw_name,
                distributor_goods_id=str(item.get("distributorGoodsId") or "").strip(),
                distributor_price=float(price) if price is not None and price > 0 else None,
                stock=float(stock) if stock is not None else None,
                package_count=float(package_count) if package_count is not None else None,
                expiry_date=expiry_date,
                match_type="sku" if product_id is not None else "unmatched",
                match_score=100 if product_id is not None else None,
                matched_sku=sku if product_id is not None else "",
                raw_name=raw_name,
                raw_manufacturer=raw_manufacturer,
                raw_json=json.dumps(item, ensure_ascii=False),
            )
        )

    db.flush()
    if int(filial_id) in PROVISOR_REFERENCE_FILIAL_IDS:
        _sync_provisor_reference_mapping_from_items(db, account_id="")
    if run_matching:
        rematch_price_list_items_by_product(db=db, price_list=row)
        _replace_legacy_price_rows_for_list(db=db, price_list=row)
        sync_selected_competitor_configs(db=db, price_format_id=pf.id)
        rebuild_competitor_prices_for_selected(db=db, price_format_id=pf.id)
        recalculate_competitor_percentiles(db=db, price_format_id=pf.id)
    db.commit()
    return row


def upsert_unified_price_list(
    *,
    db: Session,
    price_format_code: str,
    price_list: UnifiedPriceList,
    items: list[UnifiedPriceItem],
    status: str = "ok",
    as_of: date | None = None,
    run_matching: bool = True,
) -> CompetitorPriceList:
    total_started_at = time.perf_counter()
    pf = _ensure_price_format(db, price_format_code)
    today = as_of or date.today()
    source_key = _source_key_for_unified(price_list)

    row = (
        db.execute(
            select(CompetitorPriceList)
            .where(CompetitorPriceList.source_type == price_list.source)
            .where(CompetitorPriceList.source_key == source_key)
            .order_by(CompetitorPriceList.updated_at.desc(), CompetitorPriceList.id.desc())
        )
        .scalars()
        .first()
    )
    if row is None:
        row = CompetitorPriceList(
            price_format_id=pf.id,
            source_type=price_list.source,
            source_key=source_key,
            coefficient=1.0,
        )
        db.add(row)
        db.flush()

    branch_name = getattr(price_list, "branch_name", "") or "Без филиала"
    competitor_name = getattr(price_list, "competitor_name", "") or price_list.distributor_name or price_list.price_list_name
    account_login = getattr(price_list, "account_login", "") or price_list.account_id
    row.display_name = f"{branch_name} — {competitor_name} — {account_login}"
    row.supplier = price_list.distributor_name or price_list.price_list_name
    row.branch_id = getattr(price_list, "branch_id", "") or branch_name
    row.branch_code = getattr(price_list, "branch_code", "") or branch_name
    row.branch_name = branch_name
    row.competitor_name = competitor_name
    row.account_id = price_list.account_id
    row.account_login = account_login
    row.external_price_list_id = price_list.price_list_id
    row.sync_batch_id = datetime.utcnow().strftime("%Y%m%d%H%M%S%f")
    row.source_updated_at = getattr(price_list, "source_updated_at", "") or ""
    row.last_checked_at = datetime.utcnow()
    if status in {"ok", "updated", "checked_unchanged"}:
        row.last_success_at = row.last_checked_at
    row.last_refresh_status = "updated" if status == "ok" else status
    row.last_refresh_message = ""
    visible_status = ""
    if getattr(price_list, "enabled", None) is True:
        visible_status = "; visible:on"
    elif getattr(price_list, "enabled", None) is False:
        visible_status = "; visible:off"
    row.region = (
        f"branch:{branch_name}; competitor:{competitor_name}; account:{price_list.account_id}; "
        f"accountLogin:{account_login}; status:{status}{visible_status}"
    )
    row.price_date = today
    row.updated_at = datetime.utcnow()

    stage_started_at = time.perf_counter()
    preserved_match_fields: dict[tuple[str, ...], tuple[int | None, str, float | None, str]] = {}
    existing_count = 0
    if not run_matching:
        existing_items = (
            db.execute(select(CompetitorPriceListItem).where(CompetitorPriceListItem.price_list_id == row.id))
            .scalars()
            .all()
        )
        existing_count = len(existing_items)
        for existing_item in existing_items:
            fields = (
                existing_item.product_id,
                str(existing_item.match_type or "unmatched"),
                float(existing_item.match_score) if existing_item.match_score is not None else None,
                str(existing_item.matched_sku or ""),
            )
            for key in _identity_keys_for_existing_item(existing_item, price_list.source):
                preserved_match_fields.setdefault(key, fields)
    _db_save_timing(price_list_id=row.id, stage="load_existing_match_fields", rows=existing_count, started_at=stage_started_at)

    stage_started_at = time.perf_counter()
    delete_result = db.execute(
        delete(CompetitorPriceListItem)
        .where(CompetitorPriceListItem.price_list_id == row.id)
        .execution_options(synchronize_session=False)
    )
    _db_save_timing(price_list_id=row.id, stage="delete_old_items", rows=int(delete_result.rowcount or 0), started_at=stage_started_at)

    code_to_id: dict[str, int] = {}
    if run_matching:
        sku_candidates = [
            normalize_sku(item.distributor_product_id) or ""
            for item in items
            if item.distributor_product_id
        ]
        sku_candidates = [x for x in dict.fromkeys(sku_candidates) if x]
        code_to_id = (
            dict(db.execute(select(Product.code, Product.id).where(Product.code.in_(sku_candidates))).all())
            if sku_candidates
            else {}
        )

    preserved_count = 0
    reset_count = 0
    row_mappings: list[dict[str, object]] = []
    prepare_total_started_at = time.perf_counter()

    stage_started_at = time.perf_counter()
    item_names = [item.product_name or item.distributor_product_name for item in items]
    manufacturer_inputs = {
        _manufacturer_cache_key(item.manufacturer, item_names[index])
        for index, item in enumerate(items)
    }
    _prepare_rows_timing(price_list_id=row.id, stage="collect_unique_manufacturers", rows=len(manufacturer_inputs), started_at=stage_started_at)

    stage_started_at = time.perf_counter()
    manufacturer_cache = {
        key: resolve_manufacturer(key[0], key[1], default="")
        for key in manufacturer_inputs
    }
    _prepare_rows_timing(price_list_id=row.id, stage="resolve_manufacturers_bulk_or_cache", rows=len(manufacturer_cache), started_at=stage_started_at)

    stage_started_at = time.perf_counter()
    expiry_dates = [_clean_expiry_date(item.expiry_date) for item in items]
    _prepare_rows_timing(price_list_id=row.id, stage="date_cleanup", rows=len(expiry_dates), started_at=stage_started_at)

    stage_started_at = time.perf_counter()
    raw_values = [item.raw if isinstance(item.raw, dict) else {} for item in items]
    provisor_ids = [_as_int(raw.get("id")) for raw in raw_values]
    provisor_goods_ids = [_as_int(raw.get("goodsId")) for raw in raw_values]
    filial_ids = [_as_int(item.price_list_id) if item.source == "provisor" else None for item in items]
    _prepare_rows_timing(price_list_id=row.id, stage="stable_key_generation", rows=len(items), started_at=stage_started_at)

    stage_started_at = time.perf_counter()
    raw_manufacturers = [
        manufacturer_cache.get(_manufacturer_cache_key(item.manufacturer, item_names[index]), "")
        for index, item in enumerate(items)
    ]
    _prepare_rows_timing(price_list_id=row.id, stage="manufacturer_assignment", rows=len(raw_manufacturers), started_at=stage_started_at)

    stage_started_at = time.perf_counter()
    preserved_fields_by_index: list[tuple[int | None, str, float | None, str] | None] = []
    if not run_matching:
        for index, item in enumerate(items):
            raw = raw_values[index]
            keys = _competitor_item_identity_keys(
                source=item.source,
                price_list_id=item.price_list_id,
                provisor_id=provisor_ids[index],
                provisor_goods_id=provisor_goods_ids[index],
                distributor_goods_id=item.distributor_product_id,
                name=item.distributor_product_name or item.product_name,
                producer=raw_manufacturers[index],
                shelf_life=expiry_dates[index],
                batch=raw.get("batch") or raw.get("series") or raw.get("batchNumber"),
            )
            preserved_fields_by_index.append(next((preserved_match_fields[key] for key in keys if key in preserved_match_fields), None))
    _prepare_rows_timing(price_list_id=row.id, stage="build_preserve_keys", rows=len(preserved_fields_by_index), started_at=stage_started_at)

    stage_started_at = time.perf_counter()
    empty_structure_fields = _empty_match_structure_fields()
    for index, item in enumerate(items):
        item_name = item_names[index]
        raw_manufacturer = raw_manufacturers[index]
        preserved_fields = None
        if not run_matching:
            preserved_fields = preserved_fields_by_index[index]
            if preserved_fields is not None:
                preserved_count += 1
            else:
                reset_count += 1
        if run_matching:
            sku = normalize_sku(item.distributor_product_id) or ""
            product_id = next((code_to_id[v] for v in normalize_sku_variants(item.distributor_product_id) if v in code_to_id), None)
            match_type = "sku" if product_id is not None else "unmatched"
            match_score = 100 if product_id is not None else None
            matched_sku = sku if product_id is not None else ""
        elif preserved_fields is not None:
            product_id, match_type, match_score, matched_sku = preserved_fields
        else:
            product_id = None
            match_type = "unmatched"
            match_score = None
            matched_sku = ""
        mapping = {
            **empty_structure_fields,
            "price_list_id": row.id,
            "product_id": product_id,
            "provisor_id": provisor_ids[index],
            "provisor_goods_id": provisor_goods_ids[index],
            "filial_id": filial_ids[index],
            "name": item_name,
            "reg_number": item.registration_number,
            "distributor_goods_name": item.distributor_product_name or item.product_name,
            "distributor_goods_id": item.distributor_product_id,
            "distributor_price": float(item.distributor_price) if item.distributor_price is not None and item.distributor_price > 0 else None,
            "stock": float(item.stock) if item.stock is not None else None,
            "package_count": float(item.pack_quantity) if item.pack_quantity is not None else None,
            "expiry_date": expiry_dates[index],
            "match_type": match_type,
            "match_score": match_score,
            "matched_sku": matched_sku,
            "raw_name": item_name,
            "raw_manufacturer": raw_manufacturer,
            "raw_json": json.dumps(
                {
                    "source": item.source,
                    "accountId": item.account_id,
                    "priceListId": item.price_list_id,
                    "priceListName": item.price_list_name,
                    "distributorName": item.distributor_name,
                    "manufacturer": raw_manufacturer,
                    "expiryDate": item.expiry_date,
                    "raw": item.raw,
                },
                ensure_ascii=False,
                default=str,
            ),
        }
        if run_matching and price_list.source == "vidman":
            row_item = CompetitorPriceListItem(**mapping)
            _materialize_match_fields(row_item)
            for key in empty_structure_fields:
                mapping[key] = getattr(row_item, key)
        row_mappings.append(mapping)
    _prepare_rows_timing(price_list_id=row.id, stage="build_row_mapping_loop", rows=len(row_mappings), started_at=stage_started_at)
    _prepare_rows_timing(price_list_id=row.id, stage="total", rows=len(row_mappings), started_at=prepare_total_started_at)
    _db_save_timing(price_list_id=row.id, stage="prepare_rows", rows=len(row_mappings), started_at=prepare_total_started_at)
    if not run_matching:
        logger.info("[MATCH_FIELDS_PRESERVED] price_list_id=%s preserved_count=%s", row.id, preserved_count)
        logger.info("[MATCH_FIELDS_RESET] price_list_id=%s reset_count=%s", row.id, reset_count)

    stage_started_at = time.perf_counter()
    if row_mappings:
        db.bulk_insert_mappings(CompetitorPriceListItem, row_mappings)
    _db_save_timing(price_list_id=row.id, stage="bulk_insert", rows=len(row_mappings), started_at=stage_started_at)

    stage_started_at = time.perf_counter()
    db.flush()
    _db_save_timing(price_list_id=row.id, stage="flush", rows=len(row_mappings), started_at=stage_started_at)
    if price_list.source == "provisor" and _as_int(price_list.price_list_id) in PROVISOR_REFERENCE_FILIAL_IDS:
        _sync_provisor_reference_mapping_from_items(db, account_id=price_list.account_id)
    if run_matching:
        rematch_price_list_items_by_product(db=db, price_list=row)
        _replace_legacy_price_rows_for_list(db=db, price_list=row)
        sync_selected_competitor_configs(db=db, price_format_id=pf.id)
        rebuild_competitor_prices_for_selected(db=db, price_format_id=pf.id)
        recalculate_competitor_percentiles(db=db, price_format_id=pf.id)
    stage_started_at = time.perf_counter()
    db.commit()
    _db_save_timing(price_list_id=row.id, stage="commit", rows=len(row_mappings), started_at=stage_started_at)
    _db_save_timing(price_list_id=row.id, stage="total", rows=len(row_mappings), started_at=total_started_at)
    return row


def mark_unified_price_list_checked(
    *,
    db: Session,
    price_format_code: str,
    price_list: UnifiedPriceList,
    status: str,
    message: str = "",
) -> CompetitorPriceList | None:
    """Record a successful/failed check without replacing saved item rows."""
    pf = _ensure_price_format(db, price_format_code)
    source_key = _source_key_for_unified(price_list)
    row = (
        db.execute(
            select(CompetitorPriceList)
            .where(CompetitorPriceList.source_type == price_list.source)
            .where(CompetitorPriceList.source_key == source_key)
            .order_by(CompetitorPriceList.updated_at.desc(), CompetitorPriceList.id.desc())
        )
        .scalars()
        .first()
    )
    if row is None:
        return None

    checked_at = datetime.utcnow()
    source_updated_at = getattr(price_list, "source_updated_at", "") or ""
    if source_updated_at:
        row.source_updated_at = source_updated_at
    row.last_checked_at = checked_at
    if status in {"checked_unchanged", "updated", "success_zero_items"}:
        row.last_success_at = checked_at
    row.last_refresh_status = status
    row.last_refresh_message = str(message or "")[:512]
    db.flush()
    return row


def list_competitor_price_lists(
    *,
    db: Session,
    price_format_code: str,
    account_id: str | None = None,
    region: str | None = None,
) -> list[dict]:
    pf = _ensure_price_format(db, price_format_code)
    rows = list_global_competitor_price_lists_for_format(
        db=db,
        price_format=pf,
        account_id=str(account_id or "").strip() or None,
        region=str(region or "").strip() or None,
    )
    counts = dict(
        db.execute(
            select(CompetitorPriceListItem.price_list_id, func.count(CompetitorPriceListItem.id))
            .where(CompetitorPriceListItem.price_list_id.in_([row.id for row in rows]))
            .group_by(CompetitorPriceListItem.price_list_id)
        ).all()
    ) if rows else {}
    assignments = {
        int(item.price_list.id): item.assignment
        for item in get_assigned_competitor_price_lists(db=db, price_format_id=int(pf.id), active_only=False)
    }
    all_rows = rows
    visible_ids = {int(row.id) for row in all_rows}
    for row in all_rows:
        visible = int(row.id) in visible_ids
        hidden_reason = "" if visible else _price_list_assignment_hidden_reason(row, counts)
        logger.info(
            "[PRICE_LIST_ASSIGNMENT_DEBUG] source_type=%s account_id=%s price_list_id=%s price_name=%s items_count=%s real_items_count_from_db=%s visible=%s hidden_reason=%s",
            row.source_type or "",
            row.account_id or "",
            row.external_price_list_id or row.source_key or row.id,
            row.display_name or row.supplier or _source_name(row),
            int(counts.get(row.id, 0)),
            int(counts.get(row.id, 0)),
            str(visible).lower(),
            hidden_reason,
        )
    account_ids: set[int] = set()
    for row in rows:
        if ":" in (row.source_key or ""):
            try:
                account_ids.add(int(row.source_key.split(":", 1)[0]))
            except Exception:
                pass
    account_login_by_id = (
        {
            int(acc.id): acc.login
            for acc in db.execute(select(PriceSourceAccount).where(PriceSourceAccount.id.in_(account_ids))).scalars().all()
        }
        if account_ids
        else {}
    )
    return [
        {
            "id": row.id,
            "sourceType": row.source_type,
            "sourceKey": row.source_key,
            "accountId": row.account_id or (row.source_key.split(":", 1)[0] if ":" in row.source_key else ""),
            "accountLogin": row.account_login
            or (
                account_login_by_id.get(int(row.source_key.split(":", 1)[0]), "")
                if ":" in row.source_key and row.source_key.split(":", 1)[0].isdigit()
                else ""
            ),
            "supplier": row.supplier or row.display_name,
            "name": row.display_name or row.supplier or _source_name(row),
            "region": row.region or "",
            "branchId": _region_meta(row.region).get("branchId") or _region_meta(row.region).get("branch") or "Без филиала",
            "branchCode": _region_meta(row.region).get("branchCode") or _region_meta(row.region).get("branch") or "Без филиала",
            "branchName": _region_meta(row.region).get("branch") or "Без филиала",
            "competitorName": _region_meta(row.region).get("competitor") or row.supplier or row.display_name,
            "branchId": row.branch_id or _branch_name(row),
            "branchCode": row.branch_code or _branch_name(row),
            "branchName": _branch_name(row),
            "competitorName": _competitor_name(row),
            "priceDate": row.price_date.isoformat() if row.price_date else "",
            "sourceUpdatedAt": row.source_updated_at or "",
            "updatedAt": row.updated_at.isoformat() if row.updated_at else "",
            "syncBatchId": row.sync_batch_id or "",
            "status": row.last_refresh_status or _parse_region_status(row.region) or "ok",
            "refreshStatus": row.last_refresh_status or _parse_region_status(row.region) or "ok",
            "refreshMessage": row.last_refresh_message or "",
            "lastCheckedAt": row.last_checked_at.isoformat() if row.last_checked_at else "",
            "lastSuccessAt": row.last_success_at.isoformat() if row.last_success_at else "",
            "coefficient": float((assignments.get(int(row.id)).coefficient if assignments.get(int(row.id)) is not None else row.coefficient) or 1.0),
            "isSelected": bool(assignments.get(int(row.id)) is not None and assignments[int(row.id)].is_active),
            "itemsCount": int(counts.get(row.id, 0)),
            "items_count": int(counts.get(row.id, 0)),
            "visibleForFormatBranch": bool(getattr(row, "_visible_for_format_branch", True)),
            "branchMatchReason": str(getattr(row, "_branch_match_reason", "")),
            "branchMismatchReason": str(getattr(row, "_branch_mismatch_reason", "")),
            "filialId": row.external_price_list_id or row.branch_id or row.source_key,
        }
        for row in rows
    ]


def get_competitor_price_list_items(*, db: Session, price_list_id: int) -> dict:
    row = db.get(CompetitorPriceList, price_list_id)
    if row is None:
        raise ValueError("price list not found")
    items = (
        db.execute(
            select(CompetitorPriceListItem)
            .where(CompetitorPriceListItem.price_list_id == row.id)
            .order_by(CompetitorPriceListItem.name.asc(), CompetitorPriceListItem.id.asc())
        )
        .scalars()
        .all()
    )
    return {
        "meta": {
            "id": row.id,
            "name": row.display_name or row.supplier or _source_name(row),
            "supplier": row.supplier or row.display_name,
            "sourceType": row.source_type,
            "sourceKey": row.source_key,
            "region": row.region or "",
            "priceDate": row.price_date.isoformat() if row.price_date else "",
            "sourceUpdatedAt": row.source_updated_at or "",
        },
        "items": [
            {
                "provisor_id": x.provisor_id,
                "provisor_goods_id": x.provisor_goods_id,
                "filial_id": x.filial_id,
                "name": x.name,
                "reg_number": x.reg_number,
                "distributor_goods_name": x.distributor_goods_name,
                "distributor_goods_id": x.distributor_goods_id,
                "distributor_price": float(x.distributor_price) if x.distributor_price is not None else None,
                "stock": float(x.stock) if x.stock is not None else None,
                "package_count": float(x.package_count) if x.package_count is not None else None,
                "expiry_date": x.expiry_date or "",
                "match_type": x.match_type or "unmatched",
                "match_score": float(x.match_score) if x.match_score is not None else None,
                "matched_sku": x.matched_sku or "",
                "raw_name": x.raw_name or "",
                "manufacturer": resolve_manufacturer(x.raw_manufacturer, x.name or x.distributor_goods_name or x.raw_name),
                "raw_manufacturer": x.raw_manufacturer or "",
            }
            for x in items
        ],
    }


def sync_selected_competitor_configs(*, db: Session, price_format_id: int) -> None:
    selected = get_assigned_competitor_price_lists(db=db, price_format_id=price_format_id)
    selected_sources = {_source_name(item.price_list) for item in selected}

    existing = (
        db.execute(
            select(CompetitorPrice)
            .where(CompetitorPrice.price_format_id == price_format_id)
            .where(CompetitorPrice.product_id.is_(None))
        )
        .scalars()
        .all()
    )
    for cfg in existing:
        if (cfg.source_name or "") not in selected_sources:
            db.delete(cfg)

    for item in selected:
        row = item.price_list
        src = _source_name(row)
        cfg = next((x for x in existing if x.source_name == src), None)
        if cfg is None:
            cfg = CompetitorPrice(
                price_format_id=price_format_id,
                product_id=None,
                source_name=src,
            )
            db.add(cfg)
        cfg.supplier = row.supplier or row.display_name
        cfg.coefficient = float(item.assignment.coefficient or 1.0)
        cfg.price_date = row.price_date


def set_selected_competitor_price_lists(
    *,
    db: Session,
    price_format_code: str,
    selected_ids: list[int],
    coefficients: dict[int, float] | None = None,
) -> list[dict]:
    operation = f"set_selected_competitor_price_lists:{price_format_code}"
    started_at = time.perf_counter()
    _timing(operation, "start_selection/rebuild", started_at)
    pf = _ensure_price_format(db, price_format_code)
    rows = list_global_competitor_price_lists_for_format(db=db, price_format=pf)
    _timing(operation, "load_selected_price_lists", started_at)
    counts = dict(
        db.execute(
            select(CompetitorPriceListItem.price_list_id, func.count(CompetitorPriceListItem.id))
            .where(CompetitorPriceListItem.price_list_id.in_([row.id for row in rows]))
            .group_by(CompetitorPriceListItem.price_list_id)
        ).all()
    )
    selectable_ids = {int(row.id) for row in _visible_competitor_price_list_rows(rows, counts)}
    set_competitor_assignments(db=db, price_format=pf, selected_ids=[x for x in selected_ids if int(x) in selectable_ids], coefficients=coefficients)
    sync_selected_competitor_configs(db=db, price_format_id=pf.id)
    commit_started_at = time.perf_counter()
    db.commit()
    _timing(operation, "flush/commit", commit_started_at)

    rebuild_started_at = time.perf_counter()
    rebuild_competitor_prices_for_selected(db=db, price_format_id=pf.id, commit_between_lists=True)
    _timing(operation, "rebuild_competitor_prices", rebuild_started_at)
    commit_started_at = time.perf_counter()
    db.commit()
    _timing(operation, "flush/commit", commit_started_at)

    recalculate_competitor_percentiles(db=db, price_format_id=pf.id)
    commit_started_at = time.perf_counter()
    db.commit()
    _timing(operation, "flush/commit", commit_started_at)
    _timing(operation, "finish", started_at)
    _timing(operation, "total_ms", started_at)
    return list_competitor_price_lists(db=db, price_format_code=price_format_code)


def save_selected_competitor_price_lists_only(
    *,
    db: Session,
    price_format_code: str,
    selected_ids: list[int],
    coefficients: dict[int, float] | None = None,
) -> int:
    pf = _ensure_price_format(db, price_format_code)
    rows = list_global_competitor_price_lists_for_format(db=db, price_format=pf)
    counts = dict(
        db.execute(
            select(CompetitorPriceListItem.price_list_id, func.count(CompetitorPriceListItem.id))
            .where(CompetitorPriceListItem.price_list_id.in_([row.id for row in rows]))
            .group_by(CompetitorPriceListItem.price_list_id)
        ).all()
    )
    selectable_ids = {int(row.id) for row in _visible_competitor_price_list_rows(rows, counts)}
    set_competitor_assignments(db=db, price_format=pf, selected_ids=[x for x in selected_ids if int(x) in selectable_ids], coefficients=coefficients)
    sync_selected_competitor_configs(db=db, price_format_id=pf.id)
    db.commit()
    return int(pf.id)


def import_manual_price_list_excel(
    *,
    db: Session,
    price_format_code: str,
    content: bytes,
    filename: str,
) -> CompetitorPriceList:
    pf = _ensure_price_format(db, price_format_code)
    if filename.lower().endswith(".csv"):
        text = content.decode("utf-8-sig")
        delimiter = ";" if text[:2048].count(";") >= text[:2048].count(",") else ","
        rows = [tuple(row) for row in csv.reader(io.StringIO(text), delimiter=delimiter)]
    else:
        wb = load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("empty Excel file")
    headers = [str(x or "").strip().lower() for x in rows[0]]

    aliases = {
        "sku": ["sku", "код", "артикул", "идентификатор дистрибьютора", "distributor_goods_id"],
        "name": ["наименование", "название", "name"],
        "reg_number": ["регистрационный номер", "рег номер", "reg_number"],
        "manufacturer": ["производитель", "manufacturer", "producer", "бренд"],
        "distributor_goods_name": ["наименование у дистрибьютора", "distributor_goods_name"],
        "price": ["цена", "цена дистрибьютора", "price", "goodsprice"],
        "stock": ["остаток", "stored", "stock"],
        "package_count": ["количество упаковок", "упаковок", "box", "pack"],
    }

    def idx(key: str) -> int | None:
        for alias in aliases[key]:
            if alias in headers:
                return headers.index(alias)
        return None

    sku_idx = idx("sku")
    name_idx = idx("name")
    price_idx = idx("price")
    reg_idx = idx("reg_number")
    manufacturer_idx = idx("manufacturer")
    dist_name_idx = idx("distributor_goods_name")
    stock_idx = idx("stock")
    pkg_idx = idx("package_count")
    required = {
        "внутренний код товара / SKU": sku_idx,
        "наименование товара": name_idx,
        "цена с НДС": price_idx,
        "остаток": stock_idx,
    }
    missing = [label for label, value in required.items() if value is None]
    if missing:
        raise ValueError("Файл конкурента не соответствует формату. Нет обязательных колонок: " + ", ".join(missing))

    source_key = f"manual:{datetime.utcnow().strftime('%Y%m%d%H%M%S')}"
    title = filename.rsplit(".", 1)[0] or "Manual Excel"
    row = CompetitorPriceList(
        price_format_id=pf.id,
        source_type="manual",
        source_key=source_key,
        display_name=title,
        supplier=title,
        region="branch:Без филиала; competitor:" + title,
        branch_name=DEFAULT_BRANCH,
        competitor_name=title,
        account_id="manual",
        account_login=title,
        external_price_list_id=source_key,
        price_date=date.today(),
        coefficient=1.0,
        is_selected=False,
    )
    db.add(row)
    db.flush()

    codes = []
    for raw in rows[1:]:
        if sku_idx is not None and sku_idx < len(raw):
            code = normalize_sku(raw[sku_idx]) or ""
            if code:
                codes.append(code)
    code_to_id = dict(db.execute(select(Product.code, Product.id).where(Product.code.in_(list(set(codes))))).all()) if codes else {}

    for raw in rows[1:]:
        price = _as_decimal(raw[price_idx] if price_idx < len(raw) else None)
        if price is None or price <= 0:
            continue
        sku = normalize_sku(raw[sku_idx]) if sku_idx is not None and sku_idx < len(raw) else None
        name_value = str(raw[name_idx] if name_idx is not None and name_idx < len(raw) and raw[name_idx] is not None else "")
        distributor_name_value = str(
            raw[dist_name_idx] if dist_name_idx is not None and dist_name_idx < len(raw) and raw[dist_name_idx] is not None else ""
        )
        raw_manufacturer = resolve_manufacturer(
            raw[manufacturer_idx] if manufacturer_idx is not None and manufacturer_idx < len(raw) else "",
            distributor_name_value or name_value,
            default="",
        )
        db.add(
            CompetitorPriceListItem(
                price_list_id=row.id,
                product_id=code_to_id.get(sku or ""),
                distributor_goods_id=sku or "",
                name=name_value,
                reg_number=str(raw[reg_idx] if reg_idx is not None and reg_idx < len(raw) and raw[reg_idx] is not None else ""),
                distributor_goods_name=distributor_name_value,
                distributor_price=float(price),
                stock=float(_as_decimal(raw[stock_idx]) or 0) if stock_idx is not None and stock_idx < len(raw) else None,
                package_count=float(_as_decimal(raw[pkg_idx]) or 0) if pkg_idx is not None and pkg_idx < len(raw) else None,
                raw_name=distributor_name_value or name_value,
                raw_manufacturer=raw_manufacturer,
                raw_json=json.dumps({"row": list(raw)}, ensure_ascii=False, default=str),
            )
        )

    db.flush()
    _replace_legacy_price_rows_for_list(db=db, price_list=row)
    recalculate_competitor_percentiles(db=db, price_format_id=pf.id)
    db.commit()
    return row


def export_competitor_price_list(*, db: Session, price_list_id: int, fmt: str) -> tuple[str, bytes, str]:
    data = get_competitor_price_list_items(db=db, price_list_id=price_list_id)
    rows = data["items"]
    name = str(data["meta"]["name"] or f"price_list_{price_list_id}").replace("/", "_").replace("\\", "_")

    if fmt == "csv":
        s = io.StringIO()
        writer = csv.writer(s, lineterminator="\n")
        writer.writerow([label for _, label in PRICE_LIST_EXPORT_COLUMNS])
        for row in rows:
            writer.writerow([row.get(key) if row.get(key) is not None else "" for key, _ in PRICE_LIST_EXPORT_COLUMNS])
        return f"{name}.csv", s.getvalue().encode("utf-8-sig"), "text/csv; charset=utf-8"

    wb = Workbook()
    ws = wb.active
    ws.title = "price"
    ws.append([label for _, label in PRICE_LIST_EXPORT_COLUMNS])
    for row in rows:
        ws.append([row.get(key) if row.get(key) is not None else "" for key, _ in PRICE_LIST_EXPORT_COLUMNS])
    bio = io.BytesIO()
    wb.save(bio)
    return f"{name}.xlsx", bio.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
