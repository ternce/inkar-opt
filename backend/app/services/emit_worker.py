from __future__ import annotations

import asyncio
import csv
import json
import logging
import os
import re
import shutil
import sqlite3
import socket
import threading
import time
import uuid
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from decimal import Decimal
from pathlib import Path
from typing import Any, Iterable

import httpx
from openpyxl import load_workbook
from sqlalchemy import delete, func, select, update
from sqlalchemy.orm import Session, sessionmaker

from ..config import Settings
from ..models import CompetitorPriceList, CompetitorPriceListItem, PriceFormat, PriceFormatCompetitorAssignment, RefreshJob
from .competitor_assignments import propagate_emit_assignments_to_price_formats, upsert_assignment
from .competitor_percentiles import recalculate_competitor_percentiles
from .competitor_persist import _ensure_price_format
from .manufacturers import resolve_manufacturer
from .provisor import get_access_token
from .provisor_auto_refresh import (
    REFRESH_LOCK_LEASE,
    finish_job,
    job_owner_token,
    new_owner_token,
    refresh_job_to_status,
    release_global_refresh_lock,
    release_lock,
    renew_global_refresh_lock,
    renew_lock,
    try_acquire_global_refresh_lock,
    try_acquire_lock,
)

logger = logging.getLogger(__name__)

SOURCE_TYPE = "emit"
COMPAT_SOURCE_TYPE = "provisor"
LOCK_NAME = "emit_refresh"
ACTIVE_STATUSES = ("pending", "downloading", "parsing", "normalizing", "saving", "running")
DEFAULT_STALE_TIMEOUT_SECONDS = 14_400
EMIT_NAME_MARKERS = (
    "emit",
    "amity",
    "эмити",
    "emit international",
    "amity international",
    "Р\xadРјРёС‚Рё",
    "Р\xadРјРёС‚Рё Р\x98РЅС‚РµСЂРЅРµС€РЅР»",
)
MIN_FINAL_ROWS = 1
DEFAULT_TRACE_SKU = "163571"
_emit_stale_recovery_before_mark_hook = None


@dataclass
class EmitConfig:
    enabled: bool = False
    filial_ids: list[int] = field(default_factory=lambda: [1106, 1107, 1108, 1111, 1114, 1149, 8371])
    temp_dir: str = "/tmp/emit"
    download_timeout_seconds: int = 7200
    max_file_size_gb: float = 10
    batch_insert_size: int = 5000
    delete_temp_after_success: bool = True
    cleanup_temp_hours: int = 24
    min_free_disk_gb: float = 15
    max_concurrent_filials: int = 1
    min_final_rows: int = 100
    min_row_ratio: float = 0.5
    stale_timeout_seconds: int = DEFAULT_STALE_TIMEOUT_SECONDS
    heartbeat_interval_seconds: int = 30
    max_memory_mb: int = 0
    cron: str = "0 3 * * *"
    timezone: str = "Asia/Qyzylorda"
    provisor_base_url: str = "https://api.provisor.kz"
    provisor_login: str | None = None
    provisor_password: str | None = None

    @classmethod
    def from_settings(cls, settings: Settings) -> "EmitConfig":
        return cls(
            enabled=settings.emit_worker_enabled,
            filial_ids=list(settings.emit_filial_ids),
            temp_dir=settings.emit_temp_dir,
            download_timeout_seconds=settings.emit_download_timeout_seconds,
            max_file_size_gb=settings.emit_max_file_size_gb,
            batch_insert_size=settings.emit_batch_insert_size,
            delete_temp_after_success=settings.emit_delete_temp_after_success,
            cleanup_temp_hours=settings.emit_cleanup_temp_hours,
            min_free_disk_gb=settings.emit_min_free_disk_gb,
            max_concurrent_filials=settings.emit_max_concurrent_filials,
            min_final_rows=settings.emit_min_final_rows,
            min_row_ratio=settings.emit_min_row_ratio,
            stale_timeout_seconds=settings.emit_refresh_stale_timeout_seconds,
            heartbeat_interval_seconds=settings.emit_heartbeat_interval_seconds,
            max_memory_mb=settings.emit_max_memory_mb,
            cron=settings.emit_cron,
            timezone=settings.emit_timezone,
            provisor_base_url=settings.provisor_base_url,
            provisor_login=settings.provisor_login,
            provisor_password=settings.provisor_password,
        )


@dataclass
class EmitStats:
    input_rows: int = 0
    normalized_rows: int = 0
    duplicate_rows_removed: int = 0
    zero_price_rows_skipped: int = 0
    skip_reasons: dict[str, int] = field(default_factory=dict)
    rows_without_goodsId: int = 0
    final_rows_saved: int = 0
    key_type_counts: dict[str, int] = field(default_factory=dict)
    suspicious_groups: list[dict[str, Any]] = field(default_factory=list)
    downloaded_bytes: int = 0
    file_size_bytes: int = 0
    temp_file_path: str = ""
    staging_file_path: str = ""
    stage_db_path: str = ""
    stage_db_size_mb: float = 0.0
    source_file_size_gb: float = 0.0
    max_rss_mb: float | None = None
    download_elapsed_sec: float = 0.0
    parse_dedupe_elapsed_sec: float = 0.0
    db_replace_elapsed_sec: float = 0.0
    cleanup_elapsed_sec: float = 0.0
    trace: dict[str, Any] = field(default_factory=dict)

    def to_dict(self) -> dict[str, Any]:
        return {
            "input_rows": self.input_rows,
            "normalized_rows": self.normalized_rows,
            "duplicate_rows_removed": self.duplicate_rows_removed,
            "zero_price_rows_skipped": self.zero_price_rows_skipped,
            "skip_reasons": self.skip_reasons,
            "rows_without_goodsId": self.rows_without_goodsId,
            "final_rows_saved": self.final_rows_saved,
            "key_type_counts": self.key_type_counts,
            "suspicious_groups": self.suspicious_groups,
            "downloaded_bytes": self.downloaded_bytes,
            "file_size_bytes": self.file_size_bytes,
            "temp_file_path": self.temp_file_path,
            "staging_file_path": self.staging_file_path,
            "stage_db_path": self.stage_db_path,
            "stage_db_size_mb": self.stage_db_size_mb,
            "source_file_size_gb": self.source_file_size_gb,
            "max_rss_mb": self.max_rss_mb,
            "download_elapsed_sec": self.download_elapsed_sec,
            "parse_dedupe_elapsed_sec": self.parse_dedupe_elapsed_sec,
            "db_replace_elapsed_sec": self.db_replace_elapsed_sec,
            "cleanup_elapsed_sec": self.cleanup_elapsed_sec,
            "trace": self.trace,
        }


def _json_dumps(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, default=str)


def _json_loads(value: str | None, fallback: Any) -> Any:
    try:
        data = json.loads(value or "")
        return data if data is not None else fallback
    except Exception:
        return fallback


def _parse_id_set(raw: object) -> set[str]:
    if raw is None:
        return set()
    if isinstance(raw, str):
        values = [x.strip() for x in raw.split(",")]
    elif isinstance(raw, (list, tuple, set)):
        values = [str(x).strip() for x in raw]
    else:
        values = [str(raw).strip()]
    out = set()
    for value in values:
        if value and value.lstrip("-").isdigit():
            out.add(str(int(value)))
    return out


def emit_filial_id_set(config: EmitConfig | None = None) -> set[str]:
    if config is not None:
        return {str(int(x)) for x in config.filial_ids}
    return _parse_id_set(os.getenv("EMIT_FILIAL_IDS", "1106,1107,1108,1111,1114,1149,8371"))


def is_emit_plk(*, filial_id: object = None, name: object = None, config: EmitConfig | None = None) -> bool:
    filial_text = str(filial_id or "").strip()
    if filial_text and filial_text in emit_filial_id_set(config):
        return True
    text = str(name or "").casefold()
    return any(marker.casefold() in text for marker in EMIT_NAME_MARKERS)


def _free_disk_gb(path: Path) -> float:
    usage = shutil.disk_usage(path)
    return usage.free / (1024 ** 3)


def _ensure_free_disk(path: Path, min_free_gb: float) -> None:
    path.mkdir(parents=True, exist_ok=True)
    free_gb = _free_disk_gb(path)
    if free_gb < min_free_gb:
        raise RuntimeError(f"Free disk is {free_gb:.2f}GB, below EMIT_MIN_FREE_DISK_GB={min_free_gb}")


def _stage_sidecar_paths(path: Path) -> list[Path]:
    return [path, Path(str(path) + "-wal"), Path(str(path) + "-shm")]


def _stage_total_size_mb(path: Path) -> float:
    total = 0
    for item in _stage_sidecar_paths(path):
        if item.exists():
            total += item.stat().st_size
    return round(total / (1024 * 1024), 3)


def _delete_stage_files(path: Path | None) -> None:
    if path is None:
        return
    for item in _stage_sidecar_paths(path):
        if item.exists():
            item.unlink()


def _as_decimal(value: object) -> Decimal | None:
    if value is None:
        return None
    text = str(value).strip().replace(" ", "").replace(",", ".")
    if not text:
        return None
    try:
        return Decimal(text)
    except Exception:
        return None


def _as_int(value: object) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(float(str(value).strip()))
    except Exception:
        return None


def _first_positive_decimal(*values: object) -> tuple[Decimal | None, str]:
    saw_price = False
    saw_invalid = False
    for value in values:
        price = _as_decimal(value)
        if price is None:
            continue
        saw_price = True
        if price > 0:
            return price, ""
        saw_invalid = True
    if saw_invalid or saw_price:
        return None, "invalid_price"
    return None, "missing_price"


def _increment_skip(stats: EmitStats, reason: str) -> None:
    key = reason or "normalization_error"
    stats.skip_reasons[key] = int(stats.skip_reasons.get(key) or 0) + 1
    if key in {"invalid_price", "missing_price"}:
        stats.zero_price_rows_skipped += 1


def _emit_price(row: dict[str, Any], goods: dict[str, Any]) -> tuple[Decimal | None, str]:
    if row.get("goodsPrice") not in (None, ""):
        price = _as_decimal(row.get("goodsPrice"))
        if price is not None and price > 0:
            return price, ""
        return None, "invalid_price"
    return _first_positive_decimal(
        row.get("goodsPriceWithUserDiscount"),
        goods.get("price"),
        row.get("price"),
        row.get("distributor_price"),
    )


def _nested_dict(value: object) -> dict[str, Any]:
    return value if isinstance(value, dict) else {}


def _first(row: dict[str, Any], *keys: str) -> Any:
    for key in keys:
        if key in row and row.get(key) not in (None, ""):
            return row.get(key)
    return None


def normalize_text_key(value: object) -> str:
    text = str(value or "").casefold().replace("ё", "е")
    text = re.sub(r"[^\w%/.,+-]+", " ", text, flags=re.UNICODE)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def normalize_barcode(value: object) -> str:
    digits = re.sub(r"\D+", "", str(value or ""))
    return digits if 6 <= len(digits) <= 18 else ""


def normalize_sku_key(value: object) -> str:
    text = str(value or "").strip()
    barcode = normalize_barcode(text)
    if barcode:
        return f"barcode:{barcode}"
    key = normalize_text_key(text)
    key = re.sub(r"[^a-zа-я0-9_-]+", "", key, flags=re.IGNORECASE)
    return f"sku:{key}" if key else ""


def _row_lookup(row: dict[str, Any], goods: dict[str, Any], *keys: str) -> Any:
    value = _first(row, *keys)
    if value not in (None, ""):
        return value
    return _first(goods, *keys)


def extract_variant_key(row: dict[str, Any], goods: dict[str, Any] | None = None) -> str:
    goods = goods or _nested_dict(row.get("goods"))
    for key in (
        "distributorGoodsId",
        "distributor_goods_id",
        "distributorProductId",
        "sku",
        "code",
        "barcode",
        "barCode",
        "ean",
        "gtin",
    ):
        value = _row_lookup(row, goods, key)
        normalized = normalize_sku_key(value)
        if normalized:
            return normalized
    return ""


def extract_producer_key(row: dict[str, Any], goods: dict[str, Any] | None = None, name: str = "") -> str:
    goods = goods or _nested_dict(row.get("goods"))
    producer_raw = _row_lookup(row, goods, "distributorProducer", "manufacturer", "producer", "brand")
    producer = resolve_manufacturer(producer_raw, name, default="")
    return normalize_text_key(producer)


def extract_pack_signature(row: dict[str, Any], goods: dict[str, Any] | None = None, name: str = "") -> str:
    goods = goods or _nested_dict(row.get("goods"))
    explicit_parts: list[str] = []
    for key in (
        "box",
        "pack",
        "package_count",
        "packageCount",
        "quantity",
        "qty",
        "dosage",
        "dose",
        "volume",
        "weight",
        "number",
    ):
        value = _row_lookup(row, goods, key)
        if value not in (None, ""):
            explicit_parts.append(f"{key}:{normalize_text_key(value)}")
    if explicit_parts:
        return "|".join(explicit_parts[:4])

    text = normalize_text_key(name or _row_lookup(row, goods, "fullName", "name", "distributorGoodsName"))
    patterns = [
        r"(?:\u2116|n)\s*\d+\b",
        r"\b\d+\s*(?:\u0442\u0430\u0431|\u0442\u0430\u0431\u043b|\u043a\u0430\u043f\u0441|\u043a\u0430\u043f|\u0430\u043c\u043f|\u0444\u043b|\u043f\u0430\u043a|\u0441\u0430\u0448\u0435|\u0448\u0442)\b",
        r"\b\d+(?:[,.]\d+)?\s*(?:\u043c\u043b|\u043b|\u0433|\u043c\u0433|\u043c\u043a\u0433|\u043a\u0433)\b",
        r"\b\d+(?:[,.]\d+)?\s*(?:\u043c\u0433|\u043c\u043a\u0433|\u0433)\s*/\s*(?:\u043c\u043b|\u0433)\b",
        r"\b\d+(?:[,.]\d+)?\s*%\s*\d*(?:[,.]\d+)?\s*(?:\u0433|\u043c\u043b)?\b",
    ]
    matches: list[str] = []
    for pattern in patterns:
        matches.extend(re.findall(pattern, text, flags=re.IGNORECASE))
    cleaned = [re.sub(r"\s+", "", match.casefold()) for match in matches if str(match).strip()]
    return "|".join(list(dict.fromkeys(cleaned))[:6])


def normalize_name_without_price_noise(value: object) -> str:
    text = normalize_text_key(value)
    text = re.sub(r"\b\d+(?:[,.]\d+)?\s*(?:тг|тенге|kzt|₸)\b", " ", text, flags=re.IGNORECASE)
    text = re.sub(r"\bцена\b.*$", " ", text, flags=re.IGNORECASE)
    return re.sub(r"\s+", " ", text).strip()


def normalize_emit_item(row: dict[str, Any], *, filial_id: int, filial_name: str) -> dict[str, Any] | None:
    goods = _nested_dict(row.get("goods"))
    goods_id = _as_int(_first(row, "goodsId", "goods_id", "provisor_goods_id") or goods.get("id") or goods.get("goodsId"))
    price, _reason = _emit_price(row, goods)
    if price is None:
        return None
    name = str(
        goods.get("fullName")
        or _first(row, "goods_full_name", "fullName", "name")
        or row.get("distributorGoodsName")
        or goods.get("name")
        or ""
    ).strip()
    if not name:
        return None
    producer_raw = row.get("distributorProducer") or _first(row, "manufacturer", "producer") or goods.get("producer")
    producer = resolve_manufacturer(producer_raw, name, default="")
    distributor_goods_id = str(_first(row, "distributorGoodsId", "distributor_goods_id", "sku", "code") or "").strip()
    variant_key = extract_variant_key(row, goods)
    pack_signature = extract_pack_signature(row, goods, name=name)
    producer_key = extract_producer_key(row, goods, name=name)
    normalized_name_key = normalize_name_without_price_noise(name)
    source_timestamp = str(_first(row, "insertedDate", "updatedAt", "updated_at", "source_updated_at", "timestamp") or "").strip()
    stock = _as_decimal(_first(row, "stored", "stock", "quantity"))
    pack = _as_decimal(_first(row, "box", "pack", "package_count"))
    raw_json = json.dumps(row, ensure_ascii=False, default=str)
    if len(raw_json) > 4000:
        raw_json = json.dumps(
            {
                "id": row.get("id"),
                "goodsId": goods_id,
                "distributorGoodsId": distributor_goods_id,
                "distributorGoodsName": name,
                "distributorProducer": producer,
                "price": str(price),
                "filialId": filial_id,
            },
            ensure_ascii=False,
            default=str,
        )
    return {
        "provisor_id": _as_int(row.get("id")),
        "provisor_goods_id": goods_id,
        "filial_id": _as_int(_first(row, "filialId", "filial_id")) or filial_id,
        "name": name,
        "reg_number": str(goods.get("regNumber") or row.get("regNumber") or "").strip(),
        "distributor_goods_name": name,
        "distributor_goods_id": distributor_goods_id,
        "distributor_price": float(price),
        "stock": float(stock) if stock is not None else None,
        "package_count": float(pack) if pack is not None else None,
        "expiry_date": str(_first(row, "shelfLife", "expiry_date", "expiryDate") or "").strip(),
        "raw_name": name,
        "raw_manufacturer": producer,
        "raw_json": raw_json,
        "source_item_id": _as_int(row.get("id")),
        "filial_name": filial_name,
        "variant_key": variant_key,
        "pack_signature": pack_signature,
        "producer_key": producer_key,
        "normalized_name_key": normalized_name_key,
        "source_timestamp": source_timestamp,
    }


def emit_skip_reason(row: dict[str, Any]) -> str:
    goods = _nested_dict(row.get("goods"))
    price, reason = _emit_price(row, goods)
    if price is None:
        return reason
    name = str(
        goods.get("fullName")
        or _first(row, "goods_full_name", "fullName", "name")
        or row.get("distributorGoodsName")
        or goods.get("name")
        or ""
    ).strip()
    if not name:
        return "missing_name"
    return "normalization_error"


def _dedupe_key(item: dict[str, Any]) -> tuple[str, ...]:
    goods_id = item.get("provisor_goods_id")
    variant = str(item.get("variant_key") or "").strip()
    pack = str(item.get("pack_signature") or "").strip()
    producer = str(item.get("producer_key") or "").strip()
    if goods_id:
        if variant:
            return ("goodsId+variant", str(goods_id), variant)
        return ("goodsId+pack+producer", str(goods_id), pack, producer)
    return (
        "fallback",
        str(item.get("normalized_name_key") or normalize_name_without_price_noise(item.get("name"))),
        producer,
        pack,
        variant,
    )


def _trace_sku() -> str:
    return str(os.getenv("EMIT_TRACE_SKU", DEFAULT_TRACE_SKU) or "").strip()


def _trace_goods_id() -> int | None:
    value = _trace_sku()
    try:
        return int(value)
    except Exception:
        return None


def _stage_storage_key(item: dict[str, Any]) -> tuple[str, ...]:
    """Uniquely identify a staged Emit row while preserving all positive prices."""

    return (
        *_dedupe_key(item),
        "row",
        str(item.get("provisor_id") or ""),
        str(item.get("distributor_price") or ""),
        str(item.get("stock") or ""),
        str(item.get("source_timestamp") or ""),
        str(item.get("_stage_sequence") or ""),
        str(item.get("raw_json") or ""),
    )


def _dedupe_score(item: dict[str, Any]) -> tuple[int, int, int, int, int, float, str]:
    has_goods = 1 if item.get("provisor_goods_id") else 0
    has_price = 1 if float(item.get("distributor_price") or 0) > 0 else 0
    has_variant = 1 if str(item.get("variant_key") or "").strip() else 0
    name_len = min(500, len(str(item.get("name") or "").strip()))
    has_producer = 1 if str(item.get("producer_key") or item.get("raw_manufacturer") or "").strip() else 0
    price = float(item.get("distributor_price") or 0)
    timestamp = str(item.get("source_timestamp") or "")
    return (has_price, has_goods, has_variant, name_len, has_producer, -price, timestamp)


def _candidate_quality(item: dict[str, Any]) -> tuple[int, int, int, int, int, float, str]:
    return _dedupe_score(item)


def _quality_json(item: dict[str, Any]) -> str:
    return json.dumps(list(_candidate_quality(item)), ensure_ascii=False)


def _quality_tuple(value: str | None) -> tuple[Any, ...]:
    raw = _json_loads(value, [0, 0, 0, 0])
    try:
        return tuple(raw)  # type: ignore[return-value]
    except Exception:
        return (0, 0, 0, 0, 0, 0.0, "")


def _json_sample_add(value: str | None, item: object, *, limit: int = 5) -> str:
    sample = _json_loads(value, [])
    if not isinstance(sample, list):
        sample = []
    text = str(item or "").strip()
    if text and text not in sample:
        sample.append(text)
    return json.dumps(sample[:limit], ensure_ascii=False)


def deduplicate_emit_items(items: Iterable[dict[str, Any]], stats: EmitStats) -> list[dict[str, Any]]:
    best: dict[tuple[str, ...], dict[str, Any]] = {}
    for item in items:
        stats.normalized_rows += 1
        if not item.get("provisor_goods_id"):
            stats.rows_without_goodsId += 1
        key = _dedupe_key(item)
        existing = best.get(key)
        if existing is None:
            best[key] = item
            continue
        stats.duplicate_rows_removed += 1
        if _dedupe_score(item) > _dedupe_score(existing):
            best[key] = item
    return list(best.values())


def current_rss_mb() -> float | None:
    try:
        import psutil  # type: ignore

        return round(psutil.Process(os.getpid()).memory_info().rss / (1024 * 1024), 2)
    except Exception:
        return None


def _emit_worker_identity() -> dict[str, Any]:
    return {
        "worker_owner_id": os.getenv("HOSTNAME") or os.getenv("COMPUTERNAME") or socket.gethostname(),
        "worker_pid": os.getpid(),
        "instance_id": os.getenv("RAILWAY_REPLICA_ID") or os.getenv("RENDER_INSTANCE_ID") or os.getenv("HOSTNAME") or "",
    }


def _temp_disk_usage(config: EmitConfig) -> dict[str, Any]:
    temp_dir = Path(config.temp_dir)
    total = 0
    files = 0
    if temp_dir.exists():
        for path in temp_dir.rglob("*"):
            if path.is_file():
                try:
                    total += path.stat().st_size
                    files += 1
                except OSError:
                    continue
    return {"temp_disk_bytes": total, "temp_disk_files": files}


def _raise_if_memory_exceeded(config: EmitConfig | None, *, stage: str, stats: EmitStats | None = None) -> None:
    limit = int(getattr(config, "max_memory_mb", 0) or 0)
    rss = current_rss_mb()
    if stats is not None and rss is not None:
        stats.max_rss_mb = max(float(stats.max_rss_mb or 0.0), rss)
    if limit > 0 and rss is not None and rss > limit:
        raise MemoryError(f"Emit worker RSS {rss}MB exceeded EMIT_MAX_MEMORY_MB={limit} during {stage}")


def _update_max_rss(stats: EmitStats) -> None:
    rss = current_rss_mb()
    if rss is not None:
        stats.max_rss_mb = max(float(stats.max_rss_mb or 0.0), rss)


def open_stage_db(path: Path) -> sqlite3.Connection:
    path.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(str(path))
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA synchronous=NORMAL")
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS stage_items (
            dedupe_key TEXT PRIMARY KEY,
            key_type TEXT NOT NULL DEFAULT '',
            quality_json TEXT NOT NULL,
            provisor_id INTEGER,
            provisor_goods_id INTEGER,
            goods_id_text TEXT NOT NULL DEFAULT '',
            variant_key TEXT NOT NULL DEFAULT '',
            pack_signature TEXT NOT NULL DEFAULT '',
            producer_key TEXT NOT NULL DEFAULT '',
            rows_seen INTEGER NOT NULL DEFAULT 1,
            names_sample_json TEXT NOT NULL DEFAULT '[]',
            producers_sample_json TEXT NOT NULL DEFAULT '[]',
            price_min REAL,
            price_max REAL,
            filial_id INTEGER,
            name TEXT NOT NULL DEFAULT '',
            reg_number TEXT NOT NULL DEFAULT '',
            distributor_goods_name TEXT NOT NULL DEFAULT '',
            distributor_goods_id TEXT NOT NULL DEFAULT '',
            distributor_price REAL,
            stock REAL,
            package_count REAL,
            expiry_date TEXT NOT NULL DEFAULT '',
            raw_name TEXT NOT NULL DEFAULT '',
            raw_manufacturer TEXT NOT NULL DEFAULT '',
            raw_json TEXT NOT NULL DEFAULT '{}'
        )
        """
    )
    return conn


def _stage_values(item: dict[str, Any]) -> tuple[Any, ...]:
    key = _stage_storage_key(item)
    group_key = _dedupe_key(item)
    price = item.get("distributor_price")
    return (
        json.dumps(key, ensure_ascii=False),
        group_key[0],
        _quality_json(item),
        item.get("provisor_id"),
        item.get("provisor_goods_id"),
        str(item.get("provisor_goods_id") or ""),
        item.get("variant_key") or "",
        item.get("pack_signature") or "",
        item.get("producer_key") or "",
        1,
        json.dumps([item.get("name") or ""] if item.get("name") else [], ensure_ascii=False),
        json.dumps([item.get("raw_manufacturer") or ""] if item.get("raw_manufacturer") else [], ensure_ascii=False),
        price,
        price,
        item.get("filial_id"),
        item.get("name") or "",
        item.get("reg_number") or "",
        item.get("distributor_goods_name") or item.get("name") or "",
        item.get("distributor_goods_id") or "",
        item.get("distributor_price"),
        item.get("stock"),
        item.get("package_count"),
        item.get("expiry_date") or "",
        item.get("raw_name") or item.get("name") or "",
        item.get("raw_manufacturer") or "",
        item.get("raw_json") or "{}",
    )


def _stage_upsert(conn: sqlite3.Connection, item: dict[str, Any], stats: EmitStats) -> None:
    if "_stage_sequence" not in item:
        sequence = int(getattr(stats, "_stage_sequence", 0) or 0) + 1
        setattr(stats, "_stage_sequence", sequence)
        item["_stage_sequence"] = sequence
    dedupe_key = json.dumps(_stage_storage_key(item), ensure_ascii=False)
    existing = conn.execute("SELECT quality_json FROM stage_items WHERE dedupe_key = ?", (dedupe_key,)).fetchone()
    if existing is None:
        conn.execute(
            """
            INSERT INTO stage_items (
                dedupe_key, key_type, quality_json, provisor_id, provisor_goods_id, goods_id_text,
                variant_key, pack_signature, producer_key, rows_seen, names_sample_json,
                producers_sample_json, price_min, price_max, filial_id, name, reg_number,
                distributor_goods_name, distributor_goods_id, distributor_price, stock, package_count,
                expiry_date, raw_name, raw_manufacturer, raw_json
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            _stage_values(item),
        )
        return
    stats.duplicate_rows_removed += 1
    existing_audit = conn.execute(
        "SELECT names_sample_json, producers_sample_json, price_min, price_max, rows_seen FROM stage_items WHERE dedupe_key = ?",
        (dedupe_key,),
    ).fetchone()
    price = float(item.get("distributor_price") or 0)
    conn.execute(
        """
        UPDATE stage_items
        SET rows_seen = rows_seen + 1,
            names_sample_json = ?,
            producers_sample_json = ?,
            price_min = CASE WHEN price_min IS NULL OR ? < price_min THEN ? ELSE price_min END,
            price_max = CASE WHEN price_max IS NULL OR ? > price_max THEN ? ELSE price_max END
        WHERE dedupe_key = ?
        """,
        (
            _json_sample_add(existing_audit[0] if existing_audit else "[]", item.get("name")),
            _json_sample_add(existing_audit[1] if existing_audit else "[]", item.get("raw_manufacturer")),
            price,
            price,
            price,
            price,
            dedupe_key,
        ),
    )
    if _candidate_quality(item) > _quality_tuple(existing[0]):
        conn.execute(
            """
            UPDATE stage_items
            SET key_type = ?, quality_json = ?, provisor_id = ?, provisor_goods_id = ?, goods_id_text = ?,
                variant_key = ?, pack_signature = ?, producer_key = ?, rows_seen = rows_seen,
                names_sample_json = names_sample_json, producers_sample_json = producers_sample_json,
                price_min = price_min, price_max = price_max, filial_id = ?, name = ?,
                reg_number = ?, distributor_goods_name = ?, distributor_goods_id = ?, distributor_price = ?,
                stock = ?, package_count = ?, expiry_date = ?, raw_name = ?, raw_manufacturer = ?, raw_json = ?
            WHERE dedupe_key = ?
            """,
            (
                *_stage_values(item)[1:9],
                item.get("filial_id"),
                item.get("name") or "",
                item.get("reg_number") or "",
                item.get("distributor_goods_name") or item.get("name") or "",
                item.get("distributor_goods_id") or "",
                item.get("distributor_price"),
                item.get("stock"),
                item.get("package_count"),
                item.get("expiry_date") or "",
                item.get("raw_name") or item.get("name") or "",
                item.get("raw_manufacturer") or "",
                item.get("raw_json") or "{}",
                dedupe_key,
            ),
        )


def stage_row_count(stage_db_path: Path) -> int:
    with sqlite3.connect(str(stage_db_path)) as conn:
        return int(conn.execute("SELECT COUNT(*) FROM stage_items").fetchone()[0] or 0)


def iter_stage_rows(stage_db_path: Path, *, batch_size: int) -> Iterable[list[dict[str, Any]]]:
    with sqlite3.connect(str(stage_db_path)) as conn:
        conn.row_factory = sqlite3.Row
        offset = 0
        while True:
            rows = conn.execute(
                """
                SELECT provisor_id, provisor_goods_id, filial_id, name, reg_number,
                       distributor_goods_name, distributor_goods_id, distributor_price,
                       stock, package_count, expiry_date, raw_name, raw_manufacturer, raw_json
                FROM stage_items
                ORDER BY rowid
                LIMIT ? OFFSET ?
                """,
                (batch_size, offset),
            ).fetchall()
            if not rows:
                break
            yield [dict(row) for row in rows]
            offset += len(rows)


def collect_stage_audit(conn: sqlite3.Connection, *, limit: int = 20) -> tuple[dict[str, int], list[dict[str, Any]]]:
    key_type_counts = {
        str(key_type or ""): int(count or 0)
        for key_type, count in conn.execute("SELECT key_type, COUNT(*) FROM stage_items GROUP BY key_type").fetchall()
    }
    suspicious: list[dict[str, Any]] = []
    for row in conn.execute(
        """
        SELECT goods_id_text, COUNT(*) AS variants_count,
               GROUP_CONCAT(DISTINCT variant_key) AS variants,
               MIN(price_min) AS price_min, MAX(price_max) AS price_max
        FROM stage_items
        WHERE goods_id_text != '' AND key_type = 'goodsId+variant'
        GROUP BY goods_id_text
        HAVING COUNT(*) >= 10
        ORDER BY COUNT(*) DESC
        LIMIT ?
        """,
        (limit,),
    ).fetchall():
        samples = conn.execute(
            """
            SELECT names_sample_json, producers_sample_json
            FROM stage_items
            WHERE goods_id_text = ?
            LIMIT 5
            """,
            (row[0],),
        ).fetchall()
        suspicious.append(
            {
                "reason": "same_goodsId_many_variants",
                "goodsId": row[0],
                "variants_count": int(row[1] or 0),
                "variants_sample": str(row[2] or "").split(",")[:10],
                "names_sample": [item for sample in samples for item in _json_loads(sample[0], [])][:10],
                "producers_sample": [item for sample in samples for item in _json_loads(sample[1], [])][:10],
                "prices_min": row[3],
                "prices_max": row[4],
            }
        )
    remaining = max(0, limit - len(suspicious))
    if remaining:
        for row in conn.execute(
            """
            SELECT goods_id_text, rows_seen, names_sample_json, producers_sample_json, price_min, price_max, pack_signature, producer_key
            FROM stage_items
            WHERE key_type = 'goodsId+pack+producer' AND variant_key = '' AND rows_seen >= 2
            ORDER BY rows_seen DESC
            LIMIT ?
            """,
            (remaining,),
        ).fetchall():
            suspicious.append(
                {
                    "reason": "same_goodsId_merged_without_variant",
                    "goodsId": row[0],
                    "variants_count": 1,
                    "rows_seen": int(row[1] or 0),
                    "pack_signature": row[6] or "",
                    "producer_key": row[7] or "",
                    "names_sample": _json_loads(row[2], []),
                    "producers_sample": _json_loads(row[3], []),
                    "prices_min": row[4],
                    "prices_max": row[5],
                }
            )
    remaining = max(0, limit - len(suspicious))
    if remaining:
        for row in conn.execute(
            """
            SELECT rows_seen, names_sample_json, producers_sample_json, price_min, price_max, dedupe_key
            FROM stage_items
            WHERE key_type = 'fallback' AND variant_key = '' AND pack_signature = ''
            ORDER BY rows_seen DESC
            LIMIT ?
            """,
            (remaining,),
        ).fetchall():
            suspicious.append(
                {
                    "reason": "fallback_without_barcode_or_pack",
                    "goodsId": "",
                    "variants_count": 1,
                    "rows_seen": int(row[0] or 0),
                    "dedupe_key": _json_loads(row[5], []),
                    "names_sample": _json_loads(row[1], []),
                    "producers_sample": _json_loads(row[2], []),
                    "prices_min": row[3],
                    "prices_max": row[4],
                }
            )
    return key_type_counts, suspicious[:limit]


def _iter_json_array_std(path: Path) -> Iterable[dict[str, Any]]:
    decoder = json.JSONDecoder()
    with path.open("r", encoding="utf-8-sig") as fh:
        buffer = ""
        in_array = False
        eof = False
        while not eof:
            chunk = fh.read(1024 * 1024)
            eof = not chunk
            buffer += chunk
            while True:
                buffer = buffer.lstrip()
                if not buffer:
                    break
                if not in_array:
                    if buffer[0] == "[":
                        buffer = buffer[1:]
                        in_array = True
                    elif buffer[0] == "{":
                        raise RuntimeError("Top-level JSON object Emit payload requires ijson streaming parser")
                    else:
                        raise ValueError("JSON file must start with array or object")
                if in_array:
                    buffer = buffer.lstrip()
                    if buffer.startswith("]"):
                        return
                    if buffer.startswith(","):
                        buffer = buffer[1:].lstrip()
                    try:
                        obj, idx = decoder.raw_decode(buffer)
                    except json.JSONDecodeError:
                        if eof:
                            raise
                        break
                    if isinstance(obj, dict):
                        yield obj
                    buffer = buffer[idx:]


def iter_json_rows(path: Path) -> Iterable[dict[str, Any]]:
    first_non_ws = b""
    with path.open("rb") as probe:
        while True:
            char = probe.read(1)
            if not char:
                break
            if char.strip():
                first_non_ws = char
                break
    try:
        import ijson  # type: ignore

        with path.open("rb") as fh:
            if first_non_ws == b"[":
                for row in ijson.items(fh, "item"):
                    if isinstance(row, dict):
                        yield row
                return
            for prefix in ("items.item", "data.item", "result.item", "rows.item"):
                fh.seek(0)
                found = False
                for row in ijson.items(fh, prefix):
                    found = True
                    if isinstance(row, dict):
                        yield row
                if found:
                    return
    except ModuleNotFoundError:
        if first_non_ws == b"{":
            raise RuntimeError("Top-level JSON object Emit payload requires ijson; install backend requirements")
    yield from _iter_json_array_std(path)


def iter_ndjson_rows(path: Path) -> Iterable[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig") as fh:
        for line in fh:
            line = line.strip()
            if not line:
                continue
            row = json.loads(line)
            if isinstance(row, dict):
                yield row


def iter_csv_rows(path: Path) -> Iterable[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as fh:
        sample = fh.read(4096)
        fh.seek(0)
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
        yield from csv.DictReader(fh, dialect=dialect)


def iter_xlsx_rows(path: Path) -> Iterable[dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        headers = [str(x or "").strip() for x in next(rows, ())]
        for values in rows:
            yield {headers[index]: value for index, value in enumerate(values) if index < len(headers)}
    finally:
        wb.close()


def detect_format(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return "xlsx"
    if suffix in {".csv", ".tsv"}:
        return "csv"
    if suffix in {".ndjson", ".jsonl"}:
        return "ndjson"
    with path.open("rb") as fh:
        start = fh.read(2048).lstrip()
    if start.startswith(b"["):
        return "json"
    if start.startswith(b"{"):
        lines = [line.strip() for line in start.splitlines() if line.strip()]
        if len(lines) > 1 and lines[1].startswith(b"{"):
            return "ndjson"
        return "json"
    if start[:1] in {b"\xef"}:
        return "json"
    first_line = start.splitlines()[0] if start.splitlines() else b""
    if first_line.startswith(b"{"):
        return "ndjson"
    return "csv"


def iter_source_rows(path: Path) -> Iterable[dict[str, Any]]:
    fmt = detect_format(path)
    if fmt == "xlsx":
        yield from iter_xlsx_rows(path)
    elif fmt == "csv":
        yield from iter_csv_rows(path)
    elif fmt == "ndjson":
        yield from iter_ndjson_rows(path)
    else:
        yield from iter_json_rows(path)


async def download_emit_filial(*, config: EmitConfig, filial_id: int, filial_name: str, job_callback=None) -> Path:
    temp_dir = Path(config.temp_dir)
    _ensure_free_disk(temp_dir, config.min_free_disk_gb)
    temp_path = temp_dir / f"emit_{filial_id}_{datetime.utcnow().strftime('%Y%m%d%H%M%S')}_{uuid.uuid4().hex}.json"
    max_bytes = int(config.max_file_size_gb * 1024 ** 3)
    timeout = httpx.Timeout(connect=30.0, read=float(config.download_timeout_seconds), write=30.0, pool=30.0)
    token = await get_access_token(
        base_url=config.provisor_base_url,
        login=config.provisor_login,
        password=config.provisor_password,
        timeout_seconds=config.download_timeout_seconds,
    )
    started = time.perf_counter()
    downloaded = 0
    async with httpx.AsyncClient(base_url=config.provisor_base_url, timeout=timeout) as client:
        async with client.stream(
            "GET",
            "/Price/GetByFilialId",
            params={"filialId": filial_id, "_ts": int(time.time() * 1000)},
            headers={"Authorization": f"Bearer {token}", "Cache-Control": "no-cache", "Pragma": "no-cache"},
        ) as response:
            if response.status_code >= 400:
                chunks: list[bytes] = []
                remaining = 64 * 1024
                async for chunk in response.aiter_bytes(8192):
                    if remaining <= 0:
                        break
                    chunks.append(chunk[:remaining])
                    remaining -= len(chunks[-1])
                text = b"".join(chunks)
                raise RuntimeError(f"Emit filial {filial_id} download failed: HTTP {response.status_code}: {text[:500]!r}")
            content_length = int(response.headers.get("content-length") or 0)
            if content_length > max_bytes:
                raise RuntimeError(f"Emit source exceeds EMIT_MAX_FILE_SIZE_GB before download: {content_length} bytes")
            with temp_path.open("wb") as fh:
                async for chunk in response.aiter_bytes(1024 * 1024):
                    if not chunk:
                        continue
                    downloaded += len(chunk)
                    if downloaded > max_bytes:
                        raise RuntimeError(f"Emit source exceeds EMIT_MAX_FILE_SIZE_GB: {downloaded} bytes")
                    if _free_disk_gb(temp_dir) < config.min_free_disk_gb:
                        raise RuntimeError("Free disk fell below EMIT_MIN_FREE_DISK_GB while downloading")
                    fh.write(chunk)
                    _raise_if_memory_exceeded(config, stage="download")
                    elapsed = max(0.001, time.perf_counter() - started)
                    progress = {
                        "filial_id": filial_id,
                        "filial_name": filial_name,
                        "downloaded_bytes": downloaded,
                        "downloaded_gb": round(downloaded / (1024 ** 3), 4),
                        "elapsed_sec": round(elapsed, 2),
                        "speed_mb_sec": round((downloaded / (1024 ** 2)) / elapsed, 3),
                        "temp_file_path": str(temp_path),
                    }
                    logger.info("[EMIT_DOWNLOAD_PROGRESS] %s", progress)
                    if job_callback is not None:
                        job_callback(progress)
    return temp_path


def parse_normalize_stage(
    *,
    source_path: Path,
    staging_path: Path | None = None,
    stage_db_path: Path | None = None,
    filial_id: int,
    filial_name: str,
    config: EmitConfig | None = None,
) -> EmitStats:
    started = time.perf_counter()
    cfg = config or EmitConfig(min_final_rows=MIN_FINAL_ROWS)
    stage_path = stage_db_path or staging_path
    if stage_path is None:
        raise ValueError("stage_db_path is required")
    if stage_path.exists():
        stage_path.unlink()
    stats = EmitStats(
        temp_file_path=str(source_path),
        staging_file_path=str(stage_path),
        stage_db_path=str(stage_path),
        file_size_bytes=source_path.stat().st_size,
        source_file_size_gb=round(source_path.stat().st_size / (1024 ** 3), 4),
    )
    _ensure_free_disk(Path(cfg.temp_dir), cfg.min_free_disk_gb)
    conn = open_stage_db(stage_path)
    batch_count = 0
    trace_goods_id = _trace_goods_id()
    trace_raw_prices: list[float] = []
    try:
        for row in iter_source_rows(source_path):
            stats.input_rows += 1
            try:
                item = normalize_emit_item(row, filial_id=filial_id, filial_name=filial_name)
            except Exception:
                _increment_skip(stats, "normalization_error")
                logger.exception("[EMIT_NORMALIZE_ERROR] filial_id=%s row_index=%s", filial_id, stats.input_rows)
                continue
            if item is None:
                _increment_skip(stats, emit_skip_reason(row))
                continue
            if trace_goods_id is not None and item.get("provisor_goods_id") == trace_goods_id:
                trace_raw_prices.append(float(item.get("distributor_price") or 0))
            stats.normalized_rows += 1
            if not item.get("provisor_goods_id"):
                stats.rows_without_goodsId += 1
            _stage_upsert(conn, item, stats)
            batch_count += 1
            if batch_count >= cfg.batch_insert_size:
                conn.commit()
                batch_count = 0
                _ensure_free_disk(Path(cfg.temp_dir), cfg.min_free_disk_gb)
                _raise_if_memory_exceeded(cfg, stage="parse_normalize_stage", stats=stats)
                _update_max_rss(stats)
        conn.commit()
        stats.final_rows_saved = int(conn.execute("SELECT COUNT(*) FROM stage_items").fetchone()[0] or 0)
        stats.key_type_counts, stats.suspicious_groups = collect_stage_audit(conn)
        if trace_goods_id is not None:
            stage_prices = [
                float(row[0])
                for row in conn.execute(
                    "SELECT distributor_price FROM stage_items WHERE provisor_goods_id = ? ORDER BY rowid",
                    (trace_goods_id,),
                ).fetchall()
                if row[0] is not None
            ]
            stats.trace = {
                "sku": _trace_sku(),
                "raw_rows_found": len(trace_raw_prices),
                "raw_prices": trace_raw_prices,
                "items_saved_count": len(stage_prices),
                "item_prices_in_stage": stage_prices,
            }
    finally:
        conn.close()
    stats.stage_db_size_mb = _stage_total_size_mb(stage_path)
    stats.parse_dedupe_elapsed_sec = round(time.perf_counter() - started, 3)
    _update_max_rss(stats)
    if stats.final_rows_saved < cfg.min_final_rows:
        raise RuntimeError(
            f"Emit parse produced suspiciously low row count: {stats.final_rows_saved} < EMIT_MIN_FINAL_ROWS={cfg.min_final_rows}"
        )
    logger.info(
        "[EMIT_PARSE_DEDUPE_SUMMARY] filial_id=%s input_rows=%s unique_rows=%s duplicates_removed=%s zero_price_rows_skipped=%s skip_reasons=%s rows_without_goodsId=%s key_type_counts=%s stage_db_size_mb=%s source_file_size_gb=%s max_rss_mb=%s elapsed_sec=%s",
        filial_id,
        stats.input_rows,
        stats.final_rows_saved,
        stats.duplicate_rows_removed,
        stats.zero_price_rows_skipped,
        stats.skip_reasons,
        stats.rows_without_goodsId,
        stats.key_type_counts,
        stats.stage_db_size_mb,
        stats.source_file_size_gb,
        stats.max_rss_mb,
        stats.parse_dedupe_elapsed_sec,
    )
    if stats.suspicious_groups:
        logger.warning("[EMIT_DEDUPE_SUSPICIOUS_GROUPS] filial_id=%s top=%s", filial_id, stats.suspicious_groups[:20])
    if stats.trace:
        logger.info("[EMIT_TRACE] stage=parse filial_id=%s trace=%s", filial_id, stats.trace)
    return stats


def _first_price_format(db: Session) -> PriceFormat:
    row = db.execute(select(PriceFormat).order_by(PriceFormat.id.asc())).scalars().first()
    if row is not None:
        return row
    return _ensure_price_format(db, "DEFAULT")


def _price_format_for_code(db: Session, code: str | None) -> PriceFormat:
    normalized = str(code or "").strip()
    if normalized:
        row = db.execute(select(PriceFormat).where(PriceFormat.code == normalized)).scalars().first()
        if row is not None:
            return row
    return _first_price_format(db)


def _row_count_in_staging(staging_path: Path) -> int:
    return stage_row_count(staging_path)


def _empty_match_structure_fields() -> dict[str, Any]:
    return {
        "match_key": "",
        "match_type": "unmatched",
        "match_score": None,
        "matched_sku": "",
        "normalized_name": "",
        "normalized_manufacturer": "",
        "parsed_base_name": "",
        "parsed_form": "",
        "parsed_forms_json": "",
        "parsed_dosage": None,
        "parsed_dosage_volume": None,
        "parsed_quantity": None,
        "parsed_volume": None,
        "parsed_weight": None,
        "parsed_percent_strength": None,
        "parsed_concentration": None,
        "parsed_iu_dosage": None,
        "parsed_strength_signature": "",
        "parsed_dimensions_json": "",
        "parsed_critical_tokens_json": "",
    }


def replace_emit_price_list_from_staging(
    *,
    db: Session,
    config: EmitConfig,
    filial_id: int,
    filial_name: str,
    staging_path: Path,
    stats: EmitStats,
    price_format_code: str | None = None,
) -> CompetitorPriceList:
    started = time.perf_counter()
    final_count = _row_count_in_staging(staging_path)
    if final_count < config.min_final_rows:
        raise RuntimeError(f"Emit staging produced suspiciously low row count: {final_count} < {config.min_final_rows}")
    pf = _price_format_for_code(db, price_format_code)
    source_key = f"emit:{filial_id}"
    now = datetime.utcnow()
    row = (
        db.execute(
            select(CompetitorPriceList)
            .where(CompetitorPriceList.source_type == COMPAT_SOURCE_TYPE)
            .where(CompetitorPriceList.source_key == source_key)
            .order_by(CompetitorPriceList.updated_at.desc(), CompetitorPriceList.id.desc())
        )
        .scalars()
        .first()
    )
    if row is None:
        row = CompetitorPriceList(
            price_format_id=pf.id,
            source_type=COMPAT_SOURCE_TYPE,
            source_key=source_key,
            coefficient=1.0,
            is_selected=False,
        )
        db.add(row)
        db.flush()
    previous_count = int(
        db.scalar(select(func.count(CompetitorPriceListItem.id)).where(CompetitorPriceListItem.price_list_id == row.id))
        or 0
    )
    if previous_count > 0 and final_count < int(previous_count * config.min_row_ratio):
        raise RuntimeError(
            f"Emit staging row count dropped below ratio: new={final_count}, previous={previous_count}, ratio={config.min_row_ratio}"
        )
    display_name = filial_name or f"Emit International {filial_id}"
    row.price_format_id = pf.id
    row.display_name = display_name
    row.supplier = display_name
    row.region = f"branch:{display_name}; competitor:{display_name}; account:emit; accountLogin:emit; status:success"
    row.branch_id = str(filial_id)
    row.branch_code = str(filial_id)
    row.branch_name = display_name
    row.competitor_name = display_name
    row.account_id = ""
    row.account_login = "emit"
    row.external_price_list_id = str(filial_id)
    row.sync_batch_id = now.strftime("%Y%m%d%H%M%S%f")
    row.source_updated_at = now.isoformat()
    row.last_checked_at = now
    row.last_success_at = now
    row.last_refresh_status = "success"
    row.last_refresh_message = _json_dumps(stats.to_dict())[:512]
    row.price_date = date.today()
    row.updated_at = now

    db.execute(delete(CompetitorPriceListItem).where(CompetitorPriceListItem.price_list_id == row.id))
    empty_fields = _empty_match_structure_fields()
    for rows in iter_stage_rows(staging_path, batch_size=config.batch_insert_size):
        batch: list[dict[str, Any]] = []
        for item in rows:
            mapping = {
                **empty_fields,
                "price_list_id": row.id,
                "product_id": None,
                "provisor_id": item.get("provisor_id"),
                "provisor_goods_id": item.get("provisor_goods_id"),
                "filial_id": item.get("filial_id") or filial_id,
                "name": item.get("name") or "",
                "reg_number": item.get("reg_number") or "",
                "distributor_goods_name": item.get("distributor_goods_name") or item.get("name") or "",
                "distributor_goods_id": item.get("distributor_goods_id") or "",
                "distributor_price": item.get("distributor_price"),
                "stock": item.get("stock"),
                "package_count": item.get("package_count"),
                "expiry_date": item.get("expiry_date") or "",
                "raw_name": item.get("raw_name") or item.get("name") or "",
                "raw_manufacturer": item.get("raw_manufacturer") or "",
                "raw_json": item.get("raw_json") or "{}",
            }
            batch.append(mapping)
        if batch:
            db.bulk_insert_mappings(CompetitorPriceListItem, batch)
        _ensure_free_disk(Path(config.temp_dir), config.min_free_disk_gb)
        _raise_if_memory_exceeded(config, stage="db_replace", stats=stats)
    stats.final_rows_saved = final_count
    stats.db_replace_elapsed_sec = round(time.perf_counter() - started, 3)
    db.flush()
    trace_goods_id = _trace_goods_id()
    if trace_goods_id is not None:
        db_prices = [
            float(price)
            for price in db.execute(
                select(CompetitorPriceListItem.distributor_price)
                .where(CompetitorPriceListItem.price_list_id == row.id)
                .where(CompetitorPriceListItem.provisor_goods_id == trace_goods_id)
                .where(CompetitorPriceListItem.distributor_price.is_not(None))
                .order_by(CompetitorPriceListItem.id.asc())
            ).scalars().all()
            if price is not None
        ]
        trace = dict(stats.trace or {})
        trace.update(
            {
                "sku": _trace_sku(),
                "items_saved_count": len(db_prices),
                "item_prices_in_db": db_prices,
            }
        )
        stats.trace = trace
        logger.info("[EMIT_TRACE] stage=db_replace filial_id=%s price_list_id=%s trace=%s", filial_id, row.id, trace)
    db.commit()
    logger.info(
        "[EMIT_DB_REPLACE_SUMMARY] filial_id=%s previous_rows=%s final_rows_saved=%s elapsed_sec=%s",
        filial_id,
        previous_count,
        stats.final_rows_saved,
        stats.db_replace_elapsed_sec,
    )
    return row


def _create_emit_job(
    db: Session,
    *,
    mode: str,
    filial_ids: list[int],
    requested_by: str,
    owner_token: str,
    price_format_code: str | None = None,
) -> RefreshJob:
    now = datetime.utcnow()
    job = RefreshJob(
        source_type=SOURCE_TYPE,
        mode=mode,
        status="pending",
        started_at=now,
        heartbeat_at=now,
        requested_by=requested_by,
        total_plk=len(filial_ids),
        message="Emit refresh queued.",
        metadata_json=_json_dumps(
            {
                "owner_token": owner_token,
                "filial_ids": filial_ids,
                "price_format_code": price_format_code or "",
                **_emit_worker_identity(),
                "worker_started_at": now.isoformat(),
            }
        ),
    )
    db.add(job)
    db.commit()
    db.refresh(job)
    return job


def _emit_stale_timeout_seconds(config: EmitConfig | None = None) -> int:
    if config is not None:
        return max(1, int(config.stale_timeout_seconds or DEFAULT_STALE_TIMEOUT_SECONDS))
    raw = os.getenv("EMIT_STALE_JOB_TIMEOUT_SECONDS", os.getenv("EMIT_REFRESH_STALE_TIMEOUT_SECONDS", ""))
    if raw is not None:
        try:
            return max(1, int(raw))
        except Exception:
            return DEFAULT_STALE_TIMEOUT_SECONDS
    return DEFAULT_STALE_TIMEOUT_SECONDS


def _collect_emit_temp_paths(metadata: dict[str, Any]) -> list[str]:
    paths: list[str] = []
    for key in ("temp_file_path", "staging_file_path", "stage_db_path"):
        value = metadata.get(key)
        if isinstance(value, str) and value.strip():
            paths.append(value.strip())
    for item in metadata.get("filials") or []:
        if not isinstance(item, dict):
            continue
        for key in ("temp_file_path", "staging_file_path", "stage_db_path"):
            value = item.get(key)
            if isinstance(value, str) and value.strip():
                paths.append(value.strip())
    return list(dict.fromkeys(paths))


def cleanup_stale_emit_temp_files(config: EmitConfig, metadata: dict[str, Any]) -> dict[str, Any]:
    temp_dir = Path(config.temp_dir).resolve()
    deleted = 0
    failed: list[str] = []
    for raw_path in _collect_emit_temp_paths(metadata):
        try:
            path = Path(raw_path).resolve()
            if path != temp_dir and temp_dir not in path.parents:
                continue
            candidates = [path]
            if path.suffix == ".sqlite":
                candidates.extend([Path(str(path) + "-wal"), Path(str(path) + "-shm")])
            for candidate in candidates:
                if not candidate.exists():
                    continue
                if not candidate.name.startswith("emit_"):
                    continue
                candidate.unlink()
                deleted += 1
        except Exception:
            failed.append(raw_path)
            logger.exception("[EMIT_STALE_RECOVERY] action=temp_cleanup_failed path=%s", raw_path)
    logger.info("[EMIT_STALE_RECOVERY] action=temp_cleanup files_deleted=%s files_failed=%s", deleted, len(failed))
    return {"files_deleted": deleted, "files_failed": failed}


def mark_stale_emit_jobs(db: Session, *, config: EmitConfig | None = None, now: datetime | None = None) -> list[RefreshJob]:
    now = now or datetime.utcnow()
    timeout_seconds = _emit_stale_timeout_seconds(config)
    stale_before = now - timedelta(seconds=timeout_seconds)
    rows = (
        db.execute(
            select(RefreshJob)
            .where(RefreshJob.source_type == SOURCE_TYPE)
            .where(RefreshJob.status.in_(ACTIVE_STATUSES))
            .where(RefreshJob.heartbeat_at < stale_before)
            .order_by(RefreshJob.started_at.asc().nullsfirst(), RefreshJob.id.asc())
        )
        .scalars()
        .all()
    )
    recovered: list[RefreshJob] = []
    for row in rows:
        metadata = _json_loads(row.metadata_json, {})
        if not isinstance(metadata, dict):
            metadata = {}
        token = str(metadata.get("owner_token") or "")
        old_metadata_json = row.metadata_json
        old_heartbeat = row.heartbeat_at
        age_seconds = int((now - old_heartbeat).total_seconds()) if old_heartbeat else timeout_seconds
        logger.warning(
            "[EMIT_STALE_RECOVERY] job_id=%s action=detected last_heartbeat=%s age_seconds=%s",
            row.id,
            old_heartbeat.isoformat() if old_heartbeat else "",
            age_seconds,
        )
        metadata.update(
            {
                "stale_reason": "stale_timeout",
                "recovered_at": now.isoformat(),
                "stale_age_seconds": age_seconds,
                "locks_released": [],
            }
        )
        hook = _emit_stale_recovery_before_mark_hook
        if callable(hook):
            hook(row.id)
        result = db.execute(
            update(RefreshJob)
            .where(RefreshJob.id == row.id)
            .where(RefreshJob.source_type == SOURCE_TYPE)
            .where(RefreshJob.status.in_(ACTIVE_STATUSES))
            .where(RefreshJob.heartbeat_at == old_heartbeat)
            .where(RefreshJob.metadata_json == old_metadata_json)
            .values(
                finished_at=now,
                heartbeat_at=now,
                status="interrupted",
                message=f"Emit refresh interrupted: no heartbeat for {age_seconds} seconds.",
                error_message=f"Emit refresh interrupted by stale recovery: no heartbeat for {age_seconds} seconds.",
                metadata_json=_json_dumps(metadata),
            )
        )
        if int(result.rowcount or 0) != 1:
            db.rollback()
            logger.info("[EMIT_STALE_RECOVERY] job_id=%s action=skip reason=race_recheck_failed", row.id)
            continue
        db.commit()
        cleanup_summary = cleanup_stale_emit_temp_files(config or EmitConfig(), metadata)
        metadata["temp_cleanup"] = cleanup_summary
        locks_released: list[str] = []
        if token:
            if release_lock(db, name=LOCK_NAME, owner_token=token):
                locks_released.append(LOCK_NAME)
            if release_global_refresh_lock(db, owner_token=token):
                locks_released.append("competitor_refresh_global")
        metadata["locks_released"] = locks_released
        db.execute(
            update(RefreshJob)
            .where(RefreshJob.id == row.id)
            .where(RefreshJob.status == "interrupted")
            .values(metadata_json=_json_dumps(metadata))
        )
        db.commit()
        db.refresh(row)
        logger.warning("[EMIT_STALE_RECOVERY] job_id=%s action=marked_stale locks_released=%s", row.id, locks_released)
        recovered.append(row)
    return recovered


def active_emit_job(db: Session, *, config: EmitConfig | None = None) -> RefreshJob | None:
    mark_stale_emit_jobs(db, config=config)
    return (
        db.execute(
            select(RefreshJob)
            .where(RefreshJob.source_type == SOURCE_TYPE)
            .where(RefreshJob.status.in_(ACTIVE_STATUSES))
            .order_by(RefreshJob.started_at.desc().nullslast(), RefreshJob.id.desc())
        )
        .scalars()
        .first()
    )


def latest_emit_job(db: Session, *, config: EmitConfig | None = None) -> RefreshJob | None:
    active = active_emit_job(db, config=config)
    if active is not None:
        return active
    return (
        db.execute(select(RefreshJob).where(RefreshJob.source_type == SOURCE_TYPE).order_by(RefreshJob.id.desc()))
        .scalars()
        .first()
    )


def emit_job_to_dict(job: RefreshJob | None, *, config: EmitConfig | None = None) -> dict[str, Any]:
    base = refresh_job_to_status(job)
    if job is None:
        base["source_type"] = SOURCE_TYPE
        base["is_stale"] = False
        base["stale_timeout_seconds"] = _emit_stale_timeout_seconds(config)
        return base
    now = datetime.utcnow()
    timeout_seconds = _emit_stale_timeout_seconds(config)
    heartbeat = job.heartbeat_at
    stale_age = int((now - heartbeat).total_seconds()) if heartbeat else 0
    is_stale = bool(job.status in ACTIVE_STATUSES and heartbeat is not None and stale_age > timeout_seconds)
    base.update(
        {
            "is_stale": is_stale,
            "stale_age_seconds": stale_age if is_stale or job.status == "interrupted" else 0,
            "stale_timeout_seconds": timeout_seconds,
            "last_heartbeat": heartbeat.isoformat() if heartbeat else None,
        }
    )
    meta = _json_loads(job.metadata_json, {})
    if isinstance(meta, dict):
        base.update(
            {
                "filial_id": meta.get("filial_id"),
                "filial_name": meta.get("filial_name"),
                "current_stage": meta.get("current_stage") or job.status,
                "current_filial_progress": meta.get("current_filial_progress", {}),
                "completed_filial_ids": meta.get("completed_filial_ids", meta.get("refreshed_filials", [])),
                "failed_filial_ids": meta.get("failed_filial_ids", []),
                "downloaded_bytes": meta.get("downloaded_bytes", 0),
                "file_size_bytes": meta.get("file_size_bytes", 0),
                "input_rows": meta.get("input_rows", 0),
                "final_rows_saved": meta.get("final_rows_saved", 0),
                "duplicates_removed": meta.get("duplicate_rows_removed", 0),
                "temp_file_path": meta.get("temp_file_path", ""),
                "stale_reason": meta.get("stale_reason", ""),
                "recovered_at": meta.get("recovered_at", ""),
                "stale_age_seconds": meta.get("stale_age_seconds", base.get("stale_age_seconds", 0)),
                "locks_released": meta.get("locks_released", []),
                "temp_cleanup": meta.get("temp_cleanup", {}),
                "worker_owner_id": meta.get("worker_owner_id", ""),
                "worker_pid": meta.get("worker_pid"),
                "container_instance_id": meta.get("instance_id", ""),
                "rss_memory_mb": meta.get("rss_memory_mb"),
                "temp_disk_bytes": meta.get("temp_disk_bytes", 0),
                "temp_disk_files": meta.get("temp_disk_files", 0),
                "restart_interruption_reason": meta.get("stale_reason", "") if job.status == "interrupted" else "",
                "metadata": {k: v for k, v in meta.items() if k != "owner_token"},
            }
        )
    return base


def list_emit_jobs(db: Session, *, limit: int = 50, config: EmitConfig | None = None) -> list[dict[str, Any]]:
    mark_stale_emit_jobs(db, config=config)
    rows = (
        db.execute(select(RefreshJob).where(RefreshJob.source_type == SOURCE_TYPE).order_by(RefreshJob.id.desc()).limit(limit))
        .scalars()
        .all()
    )
    return [emit_job_to_dict(row, config=config) | {"id": row.id} for row in rows]


def _recalculate_percentiles_for_emit_rows(
    db: Session,
    *,
    price_list_ids: list[int],
    price_format_code: str | None = None,
    scope_to_price_list_ids: bool = False,
) -> dict[str, Any]:
    ids = [int(item) for item in price_list_ids if int(item) > 0]
    requested_format = str(price_format_code or "").strip()
    requested_pf: PriceFormat | None = None
    warnings: list[dict[str, Any]] = []

    if requested_format:
        requested_pf = db.execute(select(PriceFormat).where(PriceFormat.code == requested_format)).scalars().first()
        if requested_pf is None:
            warning = {
                "code": "emit_format_not_found",
                "message": f"Emit refresh completed but requested PriceFormat was not found: {requested_format}",
                "requested_format_code": requested_format,
                "price_list_ids": ids,
            }
            logger.warning(
                "[EMIT_FORMAT_CONTEXT] requested_format_code=%s price_list_ids=%s warning=%s",
                requested_format,
                ids,
                warning["message"],
            )
            return {"summaries": {}, "warnings": [warning], "assigned_price_format_ids": []}
        for price_list_id in ids:
            upsert_assignment(
                db=db,
                price_format_id=int(requested_pf.id),
                competitor_price_list_id=price_list_id,
                coefficient=1.0,
                is_active=True,
            )
        db.flush()

    price_lists = {
        int(row.id): row
        for row in db.execute(select(CompetitorPriceList).where(CompetitorPriceList.id.in_(ids))).scalars().all()
    } if ids else {}

    propagation = propagate_emit_assignments_to_price_formats(db=db, emit_price_list_ids=ids) if ids else None
    if propagation is not None:
        logger.info(
            "[EMIT_FORMAT_CONTEXT] action=global_assignment_propagation price_list_ids=%s created=%s reused=%s "
            "reactivated=%s skipped_incompatible=%s affected_price_format_ids=%s",
            ids,
            propagation.created_count,
            propagation.reused_count,
            propagation.reactivated_count,
            propagation.skipped_incompatible_count,
            propagation.affected_price_format_ids,
        )
        db.flush()

    assignment_rows = list(
        db.execute(
            select(
                PriceFormatCompetitorAssignment.competitor_price_list_id,
                PriceFormatCompetitorAssignment.price_format_id,
            )
            .where(PriceFormatCompetitorAssignment.competitor_price_list_id.in_(ids))
            .where(PriceFormatCompetitorAssignment.is_active.is_(True))
        )
    ) if ids else []
    assigned_by_price_list: dict[int, list[int]] = {}
    for row in assignment_rows:
        if row.price_format_id is None:
            continue
        assigned_by_price_list.setdefault(int(row.competitor_price_list_id), []).append(int(row.price_format_id))

    for price_list_id in ids:
        assigned_ids = sorted(set(assigned_by_price_list.get(price_list_id, [])))
        price_list = price_lists.get(price_list_id)
        filial_id = getattr(price_list, "branch_id", "") or getattr(price_list, "external_price_list_id", "") or ""
        logger.info(
            "[EMIT_FORMAT_CONTEXT] filial_id=%s requested_format_code=%s price_list_id=%s assigned_price_format_ids=%s",
            filial_id or "unknown",
            requested_format or "null",
            price_list_id,
            assigned_ids,
        )
        if not assigned_ids:
            warning = {
                "code": "emit_no_active_format_assignment",
                "message": "Emit refresh completed but no active PriceFormatCompetitorAssignment found; percentiles were not recalculated",
                "filial_id": filial_id,
                "price_list_id": price_list_id,
                "requested_format_code": requested_format,
            }
            warnings.append(warning)
            logger.warning(
                "[EMIT_FORMAT_CONTEXT] filial_id=%s requested_format_code=%s price_list_id=%s assigned_price_format_ids=[] warning=%s",
                filial_id or "unknown",
                requested_format or "null",
                price_list_id,
                warning["message"],
            )

    touched_format_ids = {
        format_id
        for format_ids in assigned_by_price_list.values()
        for format_id in format_ids
    }
    summaries: dict[str, Any] = {}
    for price_format_id in sorted(touched_format_ids):
        pf = db.get(PriceFormat, price_format_id)
        if pf is None:
            continue
        summary = recalculate_competitor_percentiles(
            db=db,
            price_format_id=price_format_id,
            source_price_list_ids=ids if scope_to_price_list_ids else None,
        )
        summaries[str(pf.code or price_format_id)] = {"price_format_id": price_format_id, **summary}
        logger.info(
            "[EMIT_PERCENTILE_REBUILD] requested_format_code=%s format_code=%s price_format_id=%s rows_created=%s products_with_competitors=%s",
            requested_format,
            pf.code,
            price_format_id,
            summary.get("rows_created"),
            summary.get("products_with_competitors"),
        )
    db.commit()
    return {
        "summaries": summaries,
        "warnings": warnings,
        "assigned_price_format_ids": sorted(touched_format_ids),
        "assignment_propagation": propagation.to_dict() if propagation is not None else {},
    }


def update_emit_job(db: Session, job: RefreshJob, *, status: str, message: str, metadata: dict[str, Any] | None = None) -> None:
    token = job_owner_token(job)
    db.expire(job)
    if job.status not in ACTIVE_STATUSES:
        return
    if token:
        if not renew_global_refresh_lock(db, owner_token=token):
            return
        if not renew_lock(db, name=LOCK_NAME, owner_token=token, lease=REFRESH_LOCK_LEASE):
            return
    existing = _json_loads(job.metadata_json, {})
    if not isinstance(existing, dict):
        existing = {}
    existing.update(metadata or {})
    if token:
        existing["owner_token"] = token
    job.status = status
    job.message = message[:512]
    job.heartbeat_at = datetime.utcnow()
    job.metadata_json = _json_dumps(existing)
    db.commit()


def _emit_heartbeat_once(
    *,
    session_factory: sessionmaker,
    job_id: int,
    owner_token: str,
    config: EmitConfig,
    message: str,
) -> bool:
    try:
        with session_factory() as db:
            job = db.get(RefreshJob, job_id)
            if job is None or job.source_type != SOURCE_TYPE or job.status not in ACTIVE_STATUSES:
                return False
            if job_owner_token(job) != owner_token:
                return False
            if not renew_global_refresh_lock(db, owner_token=owner_token):
                logger.error("[EMIT_HEARTBEAT] job_id=%s action=failed lock=competitor_refresh_global", job_id)
                return False
            if not renew_lock(db, name=LOCK_NAME, owner_token=owner_token, lease=REFRESH_LOCK_LEASE):
                logger.error("[EMIT_HEARTBEAT] job_id=%s action=failed lock=%s", job_id, LOCK_NAME)
                return False
            metadata = _json_loads(job.metadata_json, {})
            if not isinstance(metadata, dict):
                metadata = {}
            metadata.update(
                {
                    **_emit_worker_identity(),
                    "rss_memory_mb": current_rss_mb(),
                    "heartbeat_interval_seconds": int(config.heartbeat_interval_seconds),
                    **_temp_disk_usage(config),
                }
            )
            metadata.setdefault("owner_token", owner_token)
            job.heartbeat_at = datetime.utcnow()
            if message:
                job.message = message[:512]
            job.metadata_json = _json_dumps(metadata)
            db.commit()
            logger.debug("[EMIT_HEARTBEAT] job_id=%s action=beat status=%s", job_id, job.status)
            return True
    except Exception:
        logger.exception("[EMIT_HEARTBEAT] job_id=%s action=exception", job_id)
        return False


def _start_emit_heartbeat_thread(
    *,
    session_factory: sessionmaker,
    job_id: int,
    owner_token: str,
    config: EmitConfig,
) -> tuple[threading.Event, threading.Thread]:
    stop_event = threading.Event()

    def beat() -> None:
        while not stop_event.wait(max(1, int(config.heartbeat_interval_seconds))):
            if not _emit_heartbeat_once(
                session_factory=session_factory,
                job_id=job_id,
                owner_token=owner_token,
                config=config,
                message="Emit refresh is running.",
            ):
                logger.warning("[EMIT_HEARTBEAT] job_id=%s action=stopped reason=beat_failed", job_id)
                return

    thread = threading.Thread(target=beat, name=f"emit-heartbeat-{job_id}", daemon=True)
    thread.start()
    return stop_event, thread


class EmitWorker:
    def __init__(self, *, session_factory: sessionmaker, config: EmitConfig):
        self.session_factory = session_factory
        self.config = config

    def create_job(
        self,
        *,
        mode: str,
        filial_ids: list[int],
        requested_by: str = "manual",
        price_format_code: str | None = None,
    ) -> tuple[RefreshJob | None, RefreshJob | None, str | None]:
        filial_ids = list(dict.fromkeys(int(x) for x in filial_ids if int(x) > 0))
        with self.session_factory() as db:
            blocker = active_emit_job(db, config=self.config)
            if blocker is not None:
                return None, blocker, None
            owner_token = new_owner_token()
            if not try_acquire_global_refresh_lock(
                db,
                owner_token=owner_token,
                source=SOURCE_TYPE,
                requested_by=requested_by,
            ):
                return None, latest_emit_job(db), None
            if not try_acquire_lock(
                db,
                name=LOCK_NAME,
                lock_type="refresh",
                owner_token=owner_token,
                lease=REFRESH_LOCK_LEASE,
                metadata={"requested_by": requested_by},
            ):
                release_global_refresh_lock(db, owner_token=owner_token)
                return None, latest_emit_job(db), None
            try:
                job = _create_emit_job(
                    db,
                    mode=mode,
                    filial_ids=filial_ids,
                    requested_by=requested_by,
                    owner_token=owner_token,
                    price_format_code=price_format_code,
                )
                return job, None, owner_token
            except Exception:
                release_lock(db, name=LOCK_NAME, owner_token=owner_token)
                release_global_refresh_lock(db, owner_token=owner_token)
                raise

    async def run_job(self, job_id: int, *, owner_token: str | None = None) -> None:
        total_started = time.perf_counter()
        heartbeat_stop: threading.Event | None = None
        heartbeat_thread: threading.Thread | None = None
        with self.session_factory() as db:
            job = db.get(RefreshJob, job_id)
            if job is None:
                return
            token = owner_token or job_owner_token(job)
            metadata = _json_loads(job.metadata_json, {})
            filial_ids = [int(x) for x in (metadata.get("filial_ids") or self.config.filial_ids)]
            job_mode = str(job.mode or "")
            job.status = "running"
            job.started_at = job.started_at or datetime.utcnow()
            job.heartbeat_at = datetime.utcnow()
            job.total_plk = len(filial_ids)
            job.message = "Emit refresh started."
            if token:
                metadata.update({**_emit_worker_identity(), "rss_memory_mb": current_rss_mb()})
                metadata["owner_token"] = token
                job.metadata_json = _json_dumps(metadata)
            db.commit()
        if token:
            heartbeat_stop, heartbeat_thread = _start_emit_heartbeat_thread(
                session_factory=self.session_factory,
                job_id=job_id,
                owner_token=token,
                config=self.config,
            )
        logger.info("[EMIT_REFRESH] job_id=%s requested_filials=%s", job_id, filial_ids)
        logger.info(
            "[EMIT_PERCENTILE_INVENTORY] stage=refresh_start job_id=%s inventory=%s",
            job_id,
            _json_dumps(emit_refresh_inventory(config=self.config, mode=job_mode, filial_ids=filial_ids)),
        )
        status = "success"
        error = ""
        aggregate = {"success": 0, "failed": 0, "filials": []}
        try:
            cleanup_temp(self.config)
            sem = asyncio.Semaphore(max(1, self.config.max_concurrent_filials))
            for filial_id in filial_ids:
                async with sem:
                    result = await self.refresh_filial(job_id=job_id, filial_id=filial_id, owner_token=token)
                    aggregate["filials"].append(result)
                    if result.get("ok"):
                        aggregate["success"] += 1
                    else:
                        aggregate["failed"] += 1
                    with self.session_factory() as progress_db:
                        progress_job = progress_db.get(RefreshJob, job_id)
                        if progress_job is not None and progress_job.status in ACTIVE_STATUSES:
                            progress_job.processed_plk = len(aggregate["filials"])
                            progress_job.success_count = int(aggregate["success"])
                            progress_job.failed_count = int(aggregate["failed"])
                            progress_job.skipped_count = len(
                                [row for row in aggregate["filials"] if isinstance(row, dict) and row.get("skipped")]
                            )
                            progress_meta = _json_loads(progress_job.metadata_json, {})
                            if not isinstance(progress_meta, dict):
                                progress_meta = {}
                            completed = [
                                int(row.get("filial_id") or 0)
                                for row in aggregate["filials"]
                                if isinstance(row, dict) and row.get("ok") and int(row.get("filial_id") or 0) > 0
                            ]
                            failed_filials = [
                                int(row.get("filial_id") or 0)
                                for row in aggregate["filials"]
                                if isinstance(row, dict) and not row.get("ok") and int(row.get("filial_id") or 0) > 0
                            ]
                            progress_meta.update(
                                {
                                    "filials": aggregate["filials"],
                                    "completed_filial_ids": completed,
                                    "failed_filial_ids": failed_filials,
                                    "filial_id": result.get("filial_id"),
                                    "current_stage": "filial_saved" if result.get("ok") else "filial_failed",
                                    "rss_memory_mb": current_rss_mb(),
                                    **_temp_disk_usage(self.config),
                                }
                            )
                            progress_job.metadata_json = _json_dumps(progress_meta)
                            progress_job.heartbeat_at = datetime.utcnow()
                            progress_job.message = (
                                f"Saved Emit filial {result.get('filial_id')}"
                                if result.get("ok")
                                else f"Emit filial {result.get('filial_id')} failed"
                            )
                            progress_db.commit()
                    logger.info(
                        "[EMIT_REFRESH] job_id=%s filial_id=%s ok=%s duration_sec=%s price_list_id=%s",
                        job_id,
                        result.get("filial_id"),
                        bool(result.get("ok")),
                        result.get("duration_sec"),
                        result.get("price_list_id"),
                    )
            if aggregate["failed"]:
                status = "error" if not aggregate["success"] else "partial_success"
        except Exception as exc:
            logger.exception("Emit refresh job failed: job_id=%s", job_id)
            status = "error"
            error = str(exc)
        finally:
            with self.session_factory() as db:
                job = db.get(RefreshJob, job_id)
                if job is not None:
                    metadata = _json_loads(job.metadata_json, {})
                    price_format_code = str((metadata or {}).get("price_format_code") or "").strip()
                    refreshed_price_list_ids = [
                        int(row.get("price_list_id") or 0)
                        for row in aggregate.get("filials", [])
                        if isinstance(row, dict) and int(row.get("price_list_id") or 0) > 0
                    ]
                    percentile_rebuild = {}
                    if status in {"success", "partial_success"} and refreshed_price_list_ids:
                        percentile_rebuild = _recalculate_percentiles_for_emit_rows(
                            db,
                            price_list_ids=refreshed_price_list_ids,
                            price_format_code=price_format_code,
                            scope_to_price_list_ids=not bool(price_format_code),
                        )
                        aggregate["percentile_rebuild"] = percentile_rebuild
                    aggregate["requested_filials"] = filial_ids
                    aggregate["refreshed_filials"] = [
                        int(row.get("filial_id") or 0)
                        for row in aggregate.get("filials", [])
                        if isinstance(row, dict) and row.get("ok") and int(row.get("filial_id") or 0) > 0
                    ]
                    aggregate["duration_sec"] = round(time.perf_counter() - total_started, 3)
                    aggregate["percentile_rebuild_formats"] = sorted((percentile_rebuild.get("summaries") or {}).keys())
                    aggregate["refresh_inventory"] = emit_refresh_inventory(
                        config=self.config,
                        mode=str(job.mode or ""),
                        filial_ids=filial_ids,
                        aggregate=aggregate,
                    )
                    aggregate["refresh_inventory"]["rebuilt_price_format_ids"] = percentile_rebuild.get("assigned_price_format_ids", [])
                    aggregate["refresh_inventory"]["rebuilt_formats"] = aggregate["percentile_rebuild_formats"]
                    logger.info(
                        "[EMIT_REFRESH] job_id=%s requested_filials=%s refreshed_filials=%s success_count=%s "
                        "failed_count=%s percentile_rebuild_formats=%s filial_durations=%s total_duration_sec=%s",
                        job_id,
                        aggregate["requested_filials"],
                        aggregate["refreshed_filials"],
                        aggregate["success"],
                        aggregate["failed"],
                        aggregate["percentile_rebuild_formats"],
                        {
                            int(row.get("filial_id") or 0): row.get("duration_sec")
                            for row in aggregate.get("filials", [])
                            if isinstance(row, dict) and int(row.get("filial_id") or 0) > 0
                        },
                        aggregate["duration_sec"],
                    )
                    logger.info(
                        "[EMIT_PERCENTILE_INVENTORY] stage=refresh_complete job_id=%s inventory=%s",
                        job_id,
                        _json_dumps(aggregate["refresh_inventory"]),
                    )
                    finish_job(
                        db,
                        job,
                        status=status,
                        message=f"Emit refresh completed: success={aggregate['success']}, failed={aggregate['failed']}."
                        if status in {"success", "partial_success"}
                        else "Emit refresh failed.",
                        error=error,
                        metadata=aggregate,
                        owner_token=token,
                        allowed_statuses=set(ACTIVE_STATUSES),
                        release_refresh=False,
                    )
                    if token:
                        release_lock(db, name=LOCK_NAME, owner_token=token)
                        release_global_refresh_lock(db, owner_token=token)
            if heartbeat_stop is not None:
                heartbeat_stop.set()
            if heartbeat_thread is not None:
                heartbeat_thread.join(timeout=5)

    async def refresh_filial(self, *, job_id: int, filial_id: int, owner_token: str | None = None) -> dict[str, Any]:
        filial_started = time.perf_counter()
        filial_name = f"Emit International {filial_id}"
        temp_path: Path | None = None
        staging_path: Path | None = None
        stats = EmitStats()
        try:
            with self.session_factory() as db:
                job = db.get(RefreshJob, job_id)
                if job is not None:
                    update_emit_job(db, job, status="downloading", message=f"Downloading Emit filial {filial_id}", metadata={"filial_id": filial_id, "filial_name": filial_name})

            def _download_progress(progress: dict[str, Any]) -> None:
                with self.session_factory() as progress_db:
                    progress_job = progress_db.get(RefreshJob, job_id)
                    if progress_job is not None:
                        update_emit_job(progress_db, progress_job, status="downloading", message=f"Downloading Emit filial {filial_id}", metadata=progress)

            download_started = time.perf_counter()
            temp_path = await download_emit_filial(config=self.config, filial_id=filial_id, filial_name=filial_name, job_callback=_download_progress)
            download_elapsed = round(time.perf_counter() - download_started, 3)
            staging_dir = Path(self.config.temp_dir)
            staging_path = staging_dir / f"emit_stage_{job_id}_{filial_id}_{uuid.uuid4().hex}.sqlite"
            with self.session_factory() as db:
                job = db.get(RefreshJob, job_id)
                if job is not None:
                    update_emit_job(db, job, status="parsing", message=f"Parsing Emit filial {filial_id}", metadata={"temp_file_path": str(temp_path), "file_size_bytes": temp_path.stat().st_size})
            stats = parse_normalize_stage(source_path=temp_path, stage_db_path=staging_path, filial_id=filial_id, filial_name=filial_name, config=self.config)
            stats.downloaded_bytes = temp_path.stat().st_size
            stats.download_elapsed_sec = download_elapsed
            with self.session_factory() as db:
                job = db.get(RefreshJob, job_id)
                if job is not None:
                    update_emit_job(db, job, status="saving", message=f"Saving Emit filial {filial_id}", metadata=stats.to_dict())
                price_list = replace_emit_price_list_from_staging(
                    db=db,
                    config=self.config,
                    filial_id=filial_id,
                    filial_name=filial_name,
                    staging_path=staging_path,
                    stats=stats,
                    price_format_code=str((job.metadata_json and _json_loads(job.metadata_json, {}).get("price_format_code")) or ""),
                )
                price_list_id = int(price_list.id)
                job = db.get(RefreshJob, job_id)
                if job is not None:
                    job.processed_plk = int(job.processed_plk or 0) + 1
                    job.success_count = int(job.success_count or 0) + 1
                    update_emit_job(db, job, status="running", message=f"Saved Emit filial {filial_id}", metadata=stats.to_dict())
            if self.config.delete_temp_after_success and temp_path.exists():
                temp_path.unlink()
            _delete_stage_files(staging_path)
            stats.cleanup_elapsed_sec = 0.0
            return {
                "ok": True,
                "filial_id": filial_id,
                "price_list_id": price_list_id,
                "duration_sec": round(time.perf_counter() - filial_started, 3),
                **stats.to_dict(),
            }
        except Exception as exc:
            logger.exception("Emit filial refresh failed: filial_id=%s", filial_id)
            with self.session_factory() as db:
                job = db.get(RefreshJob, job_id)
                if job is not None:
                    job.processed_plk = int(job.processed_plk or 0) + 1
                    job.failed_count = int(job.failed_count or 0) + 1
                    update_emit_job(db, job, status="running", message=f"Emit filial {filial_id} failed", metadata={"filial_id": filial_id, "error": str(exc), **stats.to_dict()})
            max_bytes = int(self.config.max_file_size_gb * 1024 ** 3)
            if temp_path is not None and temp_path.exists() and temp_path.stat().st_size > max_bytes:
                temp_path.unlink()
            _delete_stage_files(staging_path)
            return {
                "ok": False,
                "filial_id": filial_id,
                "error": str(exc),
                "duration_sec": round(time.perf_counter() - filial_started, 3),
                **stats.to_dict(),
            }


def cleanup_temp(config: EmitConfig) -> int:
    temp_dir = Path(config.temp_dir)
    if not temp_dir.exists():
        return 0
    cutoff = datetime.utcnow() - timedelta(hours=config.cleanup_temp_hours)
    deleted = 0
    for path in temp_dir.iterdir():
        if not path.name.startswith("emit_") and path.suffix not in {".sqlite", ".json", ".jsonl"}:
            continue
        try:
            if datetime.utcfromtimestamp(path.stat().st_mtime) < cutoff:
                path.unlink()
                deleted += 1
        except Exception:
            logger.exception("Failed to cleanup Emit temp file: %s", path)
    return deleted


def configured_filial_ids_for_mode(config: EmitConfig, *, mode: str, filial_ids: list[int] | None) -> list[int]:
    normalized_mode = str(mode or "selected").strip().lower()
    if normalized_mode == "all":
        return list(config.filial_ids)
    ids = [int(x) for x in (filial_ids or []) if int(x) > 0]
    return list(dict.fromkeys(ids))


def emit_refresh_inventory(
    *,
    config: EmitConfig,
    mode: str,
    filial_ids: list[int],
    aggregate: dict[str, Any] | None = None,
) -> dict[str, Any]:
    configured = list(dict.fromkeys(int(x) for x in config.filial_ids if int(x) > 0))
    raw_candidates = [int(x) for x in filial_ids if int(x) > 0]
    unique_filials = list(dict.fromkeys(raw_candidates))
    duplicate_filials = len(raw_candidates) - len(unique_filials)
    source_keys = [f"emit:{filial_id}" for filial_id in unique_filials]
    selected_mode_excluded = [filial_id for filial_id in configured if filial_id not in set(unique_filials)]
    filials = list((aggregate or {}).get("filials") or [])
    started = len(filials)
    succeeded = [int(row.get("filial_id") or 0) for row in filials if isinstance(row, dict) and row.get("ok")]
    failed = [int(row.get("filial_id") or 0) for row in filials if isinstance(row, dict) and not row.get("ok")]
    timed_out = [
        int(row.get("filial_id") or 0)
        for row in filials
        if isinstance(row, dict) and "timeout" in str(row.get("error") or "").casefold()
    ]
    skipped = [filial_id for filial_id in unique_filials if filial_id not in {int(row.get("filial_id") or 0) for row in filials if isinstance(row, dict)}]
    return {
        "mode": str(mode or ""),
        "business_rule": "all_configured_emit_filials" if str(mode or "").strip().lower() == "all" else "explicit_requested_emit_filials_only",
        "raw_candidates": len(raw_candidates),
        "raw_candidate_filial_ids": raw_candidates,
        "unique_plks": len(source_keys),
        "unique_source_keys": source_keys,
        "duplicates": duplicate_filials,
        "queued": len(unique_filials),
        "started": started,
        "succeeded": len([item for item in succeeded if item > 0]),
        "succeeded_filial_ids": [item for item in succeeded if item > 0],
        "failed": len([item for item in failed if item > 0]),
        "failed_filial_ids": [item for item in failed if item > 0],
        "timed_out": len([item for item in timed_out if item > 0]),
        "timed_out_filial_ids": [item for item in timed_out if item > 0],
        "skipped": len(skipped),
        "skipped_filial_ids": skipped,
        "selected_mode_excluded": len(selected_mode_excluded) if str(mode or "").strip().lower() != "all" else 0,
        "selected_mode_excluded_filial_ids": selected_mode_excluded if str(mode or "").strip().lower() != "all" else [],
    }
