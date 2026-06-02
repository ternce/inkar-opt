from __future__ import annotations

import io
import logging
import re
from decimal import Decimal

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from ..models import Product
from ..models import ProductExtra
from .ph_center_top import normalize_farmcenter_sku
from .sku import normalize_composite_sku, normalize_sku

logger = logging.getLogger(__name__)


def _norm_header(value: object) -> str:
    if value is None:
        return ""
    s = str(value).strip().lower()
    s = re.sub(r"\s+", " ", s)
    return s


def _first_nonempty_row(ws, max_scan: int = 30) -> int | None:
    for r in range(1, min(ws.max_row, max_scan) + 1):
        if any(ws.cell(row=r, column=c).value not in (None, "") for c in range(1, ws.max_column + 1)):
            return r
    return None


def _headers_map(ws) -> tuple[int, dict[str, int]]:
    header_row = _first_nonempty_row(ws)
    if header_row is None:
        return 1, {}

    headers: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        key = _norm_header(ws.cell(row=header_row, column=c).value)
        if key:
            headers[key] = c
    return header_row, headers


def _get(ws, row: int, headers: dict[str, int], *keys: str):
    for k in keys:
        col = headers.get(_norm_header(k))
        if col:
            return ws.cell(row=row, column=col).value
    return None


def _as_decimal(value: object, default: Decimal | None = None) -> Decimal | None:
    if value is None or value == "":
        return default
    try:
        if isinstance(value, (int, float, Decimal)):
            return Decimal(str(value))
        s = str(value).strip().replace(" ", "").replace(",", ".")
        return Decimal(s)
    except Exception:
        return default


def import_products_excel(
    *,
    db: Session,
    content: bytes,
    top_by_sku: dict[str, int] | None = None,
) -> tuple[list[dict], dict[str, int]]:
    """Imports warehouse/products Excel.

    Expected columns (RU):
    - Материал (SKU without leading zeros) -> normalized to 18 digits
    - Название / Наименование
    - Остатки
    - Производитель
    - Учетная себестоимость -> Product.cost

    Returns: (items, stats)
    """

    wb = load_workbook(io.BytesIO(content), data_only=True)

    ws = wb.worksheets[0] if wb.worksheets else None
    if ws is None:
        raise ValueError("Excel file has no worksheets")

    header_row, headers = _headers_map(ws)
    if not headers:
        raise ValueError("Не найдена строка заголовков")

    # Resolve column names
    key_material = ("материал", "sku", "код", "код товара")
    key_name = (
        "краткий текст материала",
        "краткий tекст материала",
        "название",
        "наименование",
        "товар",
        "name",
    )
    key_stock = ("остатки", "остаток", "количество", "stock")
    key_mfr = ("производитель", "manufacturer", "бренд")
    key_cost = (
        "учетная себ по материалу",
        "учётная себ по материалу",
        "учетная себестоимость по материалу",
        "учётная себестоимость по материалу",
        "учетная себестоимость",
        "учётная себестоимость",
        "учетная себ по материалу ",
        "учётная себ по материалу ",
        "себестоимость",
        "учетная себ",
        "учётная себ",
        "costprice",
        "cost price",
        "cost",
    )

    raw_skus: list[str] = []
    parsed_rows: list[dict] = []

    for r in range(header_row + 1, ws.max_row + 1):
        material = _get(ws, r, headers, *key_material)
        sku = normalize_composite_sku(material) or normalize_sku(material)
        if not sku:
            continue

        name = _get(ws, r, headers, *key_name)
        stock = _get(ws, r, headers, *key_stock)
        manufacturer = _get(ws, r, headers, *key_mfr)
        cost = _get(ws, r, headers, *key_cost)

        name_s = str(name).strip() if name not in (None, "") else ""
        manufacturer_s = str(manufacturer).strip() if manufacturer not in (None, "") else ""

        stock_d = _as_decimal(stock)
        cost_d = _as_decimal(cost, Decimal("0")) or Decimal("0")

        row_obj = {
            "sku": sku,
            "name": name_s,
            "stock": float(stock_d) if stock_d is not None else None,
            "manufacturer": manufacturer_s,
            "costPrice": float(cost_d),
        }
        parsed_rows.append(row_obj)
        raw_skus.append(sku)

    # Deduplicate by SKU (keep last occurrence in the file)
    rows_by_sku: dict[str, dict] = {}
    for row in parsed_rows:
        rows_by_sku[row["sku"]] = row

    uniq_rows = list(rows_by_sku.values())
    uniq_skus = list(rows_by_sku.keys())

    existing: dict[str, Product] = {}
    if uniq_skus:
        rows = db.execute(select(Product).where(Product.code.in_(uniq_skus))).scalars().all()
        existing = {p.code: p for p in rows}

    top_by_sku = top_by_sku or {}
    top_by_sku_normalized = {
        normalize_farmcenter_sku(sku): rank
        for sku, rank in top_by_sku.items()
        if normalize_farmcenter_sku(sku)
    }
    created = 0
    updated = 0
    top_matched = 0

    # upsert products first (unique by sku)
    for row in uniq_rows:
        sku = row["sku"]
        top_rank = top_by_sku_normalized.get(normalize_farmcenter_sku(sku))
        row["topRank"] = top_rank
        p = existing.get(sku)
        if p is None:
            p = Product(code=sku, name=row["name"] or sku, cost=row["costPrice"], top_rank=top_rank)
            db.add(p)
            existing[sku] = p
            created += 1
        else:
            p.name = row["name"] or p.name
            p.cost = row["costPrice"]
            p.top_rank = top_rank
            updated += 1
        if top_rank is not None:
            top_matched += 1

    db.flush()

    # upsert extras (unique by sku)
    product_ids = [p.id for p in existing.values()]
    existing_extras: dict[int, ProductExtra] = {}
    if product_ids:
        extras = db.execute(select(ProductExtra).where(ProductExtra.product_id.in_(product_ids))).scalars().all()
        existing_extras = {int(e.product_id): e for e in extras}

    for row in uniq_rows:
        sku = row["sku"]
        p = existing.get(sku)
        if p is None:
            continue

        extra = existing_extras.get(int(p.id))
        if extra is None:
            extra = ProductExtra(product_id=p.id)
            db.add(extra)
            existing_extras[int(p.id)] = extra

        extra.stock = row["stock"]
        extra.manufacturer = row["manufacturer"]

    db.commit()

    stats = {
        "rows": len(parsed_rows),
        "unique_skus": len(uniq_skus),
        "created": created,
        "updated": updated,
        "top_loaded": len(top_by_sku_normalized),
        "top_matched_by_sku": top_matched,
        "top_unmatched": max(0, len(uniq_skus) - top_matched),
    }
    logger.info("[FARMCENTER_TOP] matched_by_sku=%s", top_matched)
    logger.info("[FARMCENTER_TOP] unmatched=%s", stats["top_unmatched"])

    return uniq_rows, stats
