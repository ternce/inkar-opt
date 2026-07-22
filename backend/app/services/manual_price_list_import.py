from __future__ import annotations

import csv
import hashlib
import io
import json
import math
import re
import threading
import uuid
from dataclasses import dataclass
from datetime import date
from decimal import Decimal, InvalidOperation
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import delete, func, select
from sqlalchemy.orm import Session

from ..models import (
    CompetitorPriceList,
    CompetitorPriceListItem,
    ManualPriceListImport,
    ManualPriceListImportError,
    PriceFormatCompetitorAssignment,
    Product,
)
from .competitor_persist import _ensure_price_format
from .competitor_matching import rebuild_competitor_prices_for_selected, rematch_price_list_items_by_product
from .competitor_percentiles import DEFAULT_BRANCH, recalculate_competitor_percentiles_if_needed
from .sku import normalize_sku, normalize_sku_variants
from ..timezone import now_kz_naive


MAX_MANUAL_PLK_FILE_SIZE_BYTES = 15 * 1024 * 1024
MAX_MANUAL_PLK_ROWS = 100_000
HEADER_SCAN_ROWS = 50

_LOCKS: dict[str, threading.Lock] = {}
_LOCKS_GUARD = threading.Lock()


@dataclass(frozen=True)
class ParsedManualRow:
    row_number: int
    sku: str
    name: str
    price: Decimal
    raw: dict[str, Any]


@dataclass(frozen=True)
class RowError:
    row_number: int | None
    field: str
    raw_value: str
    error_code: str
    message: str


@dataclass(frozen=True)
class ParsedManualFile:
    file_type: str
    checksum: str
    sheet: str
    delimiter: str
    encoding: str
    total_rows: int
    empty_rows: int
    duplicate_rows: int
    conflicting_duplicate_skus: int
    valid_rows: list[ParsedManualRow]
    errors: list[RowError]
    sample_rows: list[dict[str, Any]]
    headers: dict[str, str]


def _lock_for(key: str) -> threading.Lock:
    with _LOCKS_GUARD:
        lock = _LOCKS.get(key)
        if lock is None:
            lock = threading.Lock()
            _LOCKS[key] = lock
        return lock


def _checksum(content: bytes) -> str:
    return hashlib.sha256(content).hexdigest()


def _string_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "1" if value else "0"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return ""
        if value.is_integer():
            return str(int(value))
        return format(value, "f").rstrip("0").rstrip(".")
    return str(value).replace("\ufeff", "").replace("\u00a0", " ").strip()


def _normalize_header(value: object) -> str:
    text = _string_cell(value).casefold()
    text = re.sub(r"[\s_]+", " ", text)
    text = re.sub(r"[^\wа-яё ]+", "", text, flags=re.IGNORECASE)
    return text.strip()


HEADER_ALIASES = {
    "sku": {
        "sku",
        "код",
        "артикул",
        "код товара",
        "внутренний код товара",
        "идентификатор дистрибьютора",
        "distributor goods id",
        "distributor goodsid",
        "distributor_goods_id",
    },
    "name": {
        "name",
        "название",
        "наименование",
        "наименование товара",
        "товар",
        "distributor goods name",
        "distributor_goods_name",
    },
    "price": {
        "price",
        "цена",
        "цена с ндс",
        "цена р",
        "цена руб",
        "цена поставщика",
        "distributor price",
        "goodsprice",
    },
}


def _header_map(row: tuple[object, ...]) -> dict[str, int] | None:
    normalized = [_normalize_header(cell) for cell in row]
    found: dict[str, int] = {}
    for field, aliases in HEADER_ALIASES.items():
        for index, header in enumerate(normalized):
            if header in aliases:
                found[field] = index
                break
    return found if {"sku", "name", "price"}.issubset(found) else None


def _parse_price(value: object) -> Decimal | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
            return None
        candidate = Decimal(str(value))
    else:
        text = _string_cell(value)
        if not text:
            return None
        text = text.replace(" ", "").replace("\u00a0", "")
        if "," in text and "." in text:
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", ".")
        text = re.sub(r"[^0-9.\-]", "", text)
        if not text:
            return None
        try:
            candidate = Decimal(text)
        except InvalidOperation:
            return None
    if candidate <= 0:
        return None
    return candidate.quantize(Decimal("0.0001"))


def _detect_csv(content: bytes) -> tuple[list[tuple[str, ...]], str, str]:
    encodings = ["utf-8-sig", "utf-8", "cp1251"]
    last_error: Exception | None = None
    for encoding in encodings:
        try:
            text = content.decode(encoding)
            break
        except UnicodeDecodeError as exc:
            last_error = exc
    else:
        raise ValueError(f"cannot decode CSV file: {last_error}")

    sample = text[:4096]
    delimiters = [",", ";", "\t"]
    delimiter = max(delimiters, key=lambda item: sample.count(item))
    if not sample.count(delimiter):
        try:
            delimiter = csv.Sniffer().sniff(sample).delimiter
        except csv.Error:
            delimiter = ";"
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    return [tuple(row) for row in reader], delimiter, encoding


def _xlsx_rows(content: bytes) -> tuple[list[tuple[object, ...]], str]:
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    best_rows: list[tuple[object, ...]] = []
    best_sheet = ""
    for ws in wb.worksheets:
        probe: list[tuple[object, ...]] = []
        for index, row in enumerate(ws.iter_rows(values_only=True), start=1):
            probe.append(tuple(row or ()))
            if index >= HEADER_SCAN_ROWS:
                break
        if any(_header_map(row) for row in probe):
            rows = [tuple(row or ()) for row in ws.iter_rows(values_only=True)]
            return rows, ws.title
        if not best_rows and any(any(_string_cell(cell) for cell in row) for row in probe):
            best_rows = [tuple(row or ()) for row in ws.iter_rows(values_only=True)]
            best_sheet = ws.title
    return best_rows, best_sheet


def parse_manual_price_list(content: bytes, filename: str) -> ParsedManualFile:
    if not content:
        raise ValueError("empty file")
    if len(content) > MAX_MANUAL_PLK_FILE_SIZE_BYTES:
        raise ValueError("file is too large")

    lower = filename.casefold()
    checksum = _checksum(content)
    if lower.endswith(".csv"):
        rows, delimiter, encoding = _detect_csv(content)
        file_type = "csv"
        sheet = ""
    elif lower.endswith(".xlsx") or lower.endswith(".xlsm"):
        rows, sheet = _xlsx_rows(content)
        delimiter = ""
        encoding = ""
        file_type = "xlsx"
    else:
        raise ValueError("only .xlsx, .xlsm and .csv files are supported")

    if not rows:
        raise ValueError("file contains no rows")
    if len(rows) > MAX_MANUAL_PLK_ROWS + HEADER_SCAN_ROWS:
        raise ValueError("file contains too many rows")

    header_index = None
    mapping: dict[str, int] | None = None
    for index, row in enumerate(rows[:HEADER_SCAN_ROWS]):
        mapping = _header_map(row)
        if mapping is not None:
            header_index = index
            break
    if mapping is None or header_index is None:
        raise ValueError("required columns were not found: SKU, name, price")

    header_row = rows[header_index]
    headers = {field: _string_cell(header_row[column]) if column < len(header_row) else field for field, column in mapping.items()}
    parsed: list[ParsedManualRow] = []
    errors: list[RowError] = []
    empty_rows = 0
    duplicates = 0
    by_sku: dict[str, ParsedManualRow] = {}
    conflict_skus: set[str] = set()
    total_data_rows = 0

    for offset, raw_row in enumerate(rows[header_index + 1 :], start=header_index + 2):
        if not any(_string_cell(cell) for cell in raw_row):
            empty_rows += 1
            continue
        total_data_rows += 1
        raw_values = {
            field: _string_cell(raw_row[column]) if column < len(raw_row) else ""
            for field, column in mapping.items()
        }
        sku = _string_cell(raw_values.get("sku")).strip()
        name = raw_values.get("name", "").strip()
        price = _parse_price(raw_values.get("price"))

        if not sku:
            errors.append(RowError(offset, "sku", raw_values.get("sku", ""), "missing_sku", "SKU is required"))
            continue
        if not name:
            errors.append(RowError(offset, "name", "", "missing_name", "name is required"))
            continue
        if price is None:
            errors.append(RowError(offset, "price", raw_values.get("price", ""), "invalid_price", "price must be positive"))
            continue

        parsed_row = ParsedManualRow(offset, sku, name, price, {"row": list(raw_row), "values": raw_values})
        previous = by_sku.get(sku)
        if previous is None:
            by_sku[sku] = parsed_row
            parsed.append(parsed_row)
            continue
        if previous.name == parsed_row.name and previous.price == parsed_row.price:
            duplicates += 1
            continue
        conflict_skus.add(sku)
        errors.append(RowError(offset, "sku", sku, "duplicate_sku_conflict", "SKU has conflicting duplicate rows"))

    valid_rows = [row for row in parsed if row.sku not in conflict_skus]
    for row in parsed:
        if row.sku in conflict_skus:
            errors.append(RowError(row.row_number, "sku", row.sku, "duplicate_sku_conflict", "SKU has conflicting duplicate rows"))

    return ParsedManualFile(
        file_type=file_type,
        checksum=checksum,
        sheet=sheet,
        delimiter=delimiter,
        encoding=encoding,
        total_rows=total_data_rows,
        empty_rows=empty_rows,
        duplicate_rows=duplicates,
        conflicting_duplicate_skus=len(conflict_skus),
        valid_rows=valid_rows,
        errors=errors,
        sample_rows=[{"rowNumber": row.row_number, "sku": row.sku, "name": row.name, "price": float(row.price)} for row in valid_rows[:20]],
        headers=headers,
    )


def _summary(parsed: ParsedManualFile, *, matched: int = 0, persisted: int = 0) -> dict[str, Any]:
    valid = len(parsed.valid_rows)
    invalid = len(parsed.errors)
    return {
        "fileType": parsed.file_type,
        "checksum": parsed.checksum,
        "sheet": parsed.sheet,
        "delimiter": parsed.delimiter,
        "encoding": parsed.encoding,
        "totalRows": parsed.total_rows,
        "emptyRows": parsed.empty_rows,
        "validRows": valid,
        "invalidRows": invalid,
        "duplicateRows": parsed.duplicate_rows,
        "conflictingDuplicateSkus": parsed.conflicting_duplicate_skus,
        "matchedRows": matched,
        "unmatchedRows": max(valid - matched, 0),
        "persistedRows": persisted,
        "sampleRows": parsed.sample_rows,
        "headers": parsed.headers,
        "errors": [_error_dict(error) for error in parsed.errors[:100]],
    }


def _error_dict(error: RowError) -> dict[str, Any]:
    return {
        "rowNumber": error.row_number,
        "field": error.field,
        "rawValue": error.raw_value,
        "errorCode": error.error_code,
        "message": error.message,
    }


def preview_manual_price_list(*, content: bytes, filename: str) -> dict[str, Any]:
    parsed = parse_manual_price_list(content, filename)
    return {"ok": True, "mode": "dry_run", **_summary(parsed)}


def _create_history(
    *,
    db: Session,
    parsed: ParsedManualFile | None,
    price_list_id: int | None,
    source_key: str,
    filename: str,
    status: str,
    requested_by: str,
    preserved_previous_snapshot: bool,
    matched_rows: int = 0,
    persisted_rows: int = 0,
    error_summary: str = "",
    metadata: dict[str, Any] | None = None,
) -> ManualPriceListImport:
    now = now_kz_naive()
    row = ManualPriceListImport(
        competitor_price_list_id=price_list_id,
        source_key=source_key,
        original_filename=filename,
        file_type=parsed.file_type if parsed else "",
        file_checksum=parsed.checksum if parsed else "",
        detected_sheet=parsed.sheet if parsed else "",
        detected_delimiter=parsed.delimiter if parsed else "",
        detected_encoding=parsed.encoding if parsed else "",
        status=status,
        total_rows=parsed.total_rows if parsed else 0,
        empty_rows=parsed.empty_rows if parsed else 0,
        valid_rows=len(parsed.valid_rows) if parsed else 0,
        invalid_rows=len(parsed.errors) if parsed else 0,
        duplicate_rows=parsed.duplicate_rows if parsed else 0,
        conflicting_duplicate_skus=parsed.conflicting_duplicate_skus if parsed else 0,
        matched_rows=matched_rows,
        unmatched_rows=max((len(parsed.valid_rows) if parsed else 0) - matched_rows, 0),
        persisted_rows=persisted_rows,
        preserved_previous_snapshot=preserved_previous_snapshot,
        requested_by=requested_by,
        started_at=now,
        finished_at=now,
        error_summary=error_summary,
        metadata_json=json.dumps(metadata or {}, ensure_ascii=False),
    )
    db.add(row)
    db.flush()
    if parsed:
        for error in parsed.errors[:1000]:
            db.add(
                ManualPriceListImportError(
                    import_id=row.id,
                    row_number=error.row_number,
                    field=error.field,
                    raw_value=error.raw_value[:1000],
                    error_code=error.error_code,
                    message=error.message,
                )
            )
    return row


def _manual_region(branch_name: str, competitor_name: str, address: str = "") -> str:
    parts = [f"branch:{branch_name or DEFAULT_BRANCH}", f"competitor:{competitor_name}"]
    if address:
        parts.append(f"address:{address}")
    return "; ".join(parts)


def _find_products_by_sku(db: Session, skus: list[str]) -> dict[str, int]:
    variants = sorted({variant for sku in skus for variant in [sku, *normalize_sku_variants(sku)] if variant})
    if not variants:
        return {}
    rows = db.execute(select(Product.code, Product.id).where(Product.code.in_(variants))).all()
    out: dict[str, int] = {}
    for code, product_id in rows:
        for variant in [normalize_sku(code) or str(code), *normalize_sku_variants(code)]:
            if variant:
                out[variant] = int(product_id)
    return out


def import_manual_price_list(
    *,
    db: Session,
    price_format_code: str,
    content: bytes,
    filename: str,
    display_name: str | None = None,
    competitor_name: str | None = None,
    branch_name: str | None = None,
    address: str = "",
    price_list_id: int | None = None,
    requested_by: str = "",
) -> dict[str, Any]:
    parsed = parse_manual_price_list(content, filename)
    source_key = ""
    existing: CompetitorPriceList | None = None
    if price_list_id is not None:
        existing = db.get(CompetitorPriceList, price_list_id)
        if existing is None or existing.source_type != "manual":
            raise ValueError("manual price list not found")
        source_key = existing.source_key
    lock_key = source_key or f"new:{parsed.checksum}"

    lock = _lock_for(lock_key)
    if not lock.acquire(blocking=False):
        raise ValueError("manual price list import is already running")
    try:
        pf = _ensure_price_format(db, price_format_code)
        if existing is None and price_list_id is not None:
            existing = db.get(CompetitorPriceList, price_list_id)
        if existing is not None and int(existing.price_format_id) != int(pf.id):
            raise ValueError("manual price list belongs to another price format")

        if not parsed.valid_rows:
            history = _create_history(
                db=db,
                parsed=parsed,
                price_list_id=price_list_id,
                source_key=source_key,
                filename=filename,
                status="error",
                requested_by=requested_by,
                preserved_previous_snapshot=bool(existing),
                error_summary="no valid rows to import",
            )
            db.commit()
            return {"ok": False, "status": "error", "importId": history.id, **_summary(parsed)}

        now = now_kz_naive()
        if existing is None:
            source_key = f"manual:{uuid.uuid4()}"
            title = (display_name or filename.rsplit(".", 1)[0] or "Manual price list").strip()
            competitor = (competitor_name or title).strip()
            branch = (branch_name or DEFAULT_BRANCH).strip() or DEFAULT_BRANCH
            row = CompetitorPriceList(
                price_format_id=pf.id,
                source_type="manual",
                source_key=source_key,
                display_name=title,
                supplier=competitor,
                region=_manual_region(branch, competitor, address),
                branch_name=branch,
                competitor_name=competitor,
                account_id="manual",
                account_login=requested_by or "manual",
                external_price_list_id=source_key,
                price_date=date.today(),
                coefficient=1.0,
                price_coefficient=1.0,
                is_selected=False,
                created_at=now,
                updated_at=now,
            )
            db.add(row)
            db.flush()
        else:
            row = existing
            source_key = row.source_key
            row.display_name = (display_name or row.display_name or filename.rsplit(".", 1)[0]).strip()
            row.competitor_name = (competitor_name or row.competitor_name or row.display_name).strip()
            row.supplier = row.competitor_name
            row.branch_name = (branch_name or row.branch_name or DEFAULT_BRANCH).strip() or DEFAULT_BRANCH
            row.region = _manual_region(row.branch_name, row.competitor_name, address)
            row.account_login = requested_by or row.account_login
            row.price_date = date.today()
            row.updated_at = now

        db.execute(delete(CompetitorPriceListItem).where(CompetitorPriceListItem.price_list_id == row.id))
        product_by_sku = _find_products_by_sku(db, [item.sku for item in parsed.valid_rows])
        for item in parsed.valid_rows:
            product_id = next((product_by_sku.get(variant) for variant in [item.sku, *normalize_sku_variants(item.sku)] if variant in product_by_sku), None)
            db.add(
                CompetitorPriceListItem(
                    price_list_id=row.id,
                    product_id=product_id,
                    name=item.name,
                    distributor_goods_name=item.name,
                    distributor_goods_id=item.sku,
                    distributor_price=float(item.price),
                    stock=None,
                    package_count=None,
                    match_type="sku" if product_id is not None else "unmatched",
                    match_score=100 if product_id is not None else None,
                    matched_sku=item.sku if product_id is not None else "",
                    raw_name=item.name,
                    raw_json=json.dumps(item.raw, ensure_ascii=False, default=str),
                )
            )
        db.flush()

        match_stats = rematch_price_list_items_by_product(db=db, price_list=row)
        matched = int(match_stats.get("matched") or match_stats.get("supplier_items_matched") or 0)
        persisted = len(parsed.valid_rows)
        status = "partial_success" if parsed.errors else "success"
        row.last_checked_at = now
        row.last_success_at = now
        row.source_updated_at = now.isoformat()
        row.sync_batch_id = parsed.checksum[:16]
        row.last_refresh_status = status
        row.last_refresh_message = (
            f"rows={persisted}; matched={matched}; invalid={len(parsed.errors)}; duplicates={parsed.duplicate_rows}; checksum={parsed.checksum[:12]}"
        )
        history = _create_history(
            db=db,
            parsed=parsed,
            price_list_id=int(row.id),
            source_key=source_key,
            filename=filename,
            status=status,
            requested_by=requested_by,
            preserved_previous_snapshot=False,
            matched_rows=matched,
            persisted_rows=persisted,
            error_summary=row.last_refresh_message,
            metadata={"priceListId": int(row.id), "sourceKey": source_key},
        )
        rebuild_competitor_prices_for_selected(db=db, price_format_id=int(pf.id))
        recalculate_competitor_percentiles_if_needed(db=db, price_format_id=int(pf.id))
        db.commit()
        inventory = {
            "mode": "apply",
            "price_list_id": int(row.id),
            "source_key": source_key,
            "raw_candidates": parsed.total_rows,
            "unique_plks": 1,
            "duplicates": parsed.duplicate_rows,
            "queued": 1,
            "started": 1,
            "succeeded": 1 if status in {"success", "partial_success"} else 0,
            "failed": 0,
            "timed_out": 0,
            "skipped": len(parsed.errors),
            "valid_rows": persisted,
            "matched_rows": matched,
            "unmatched_rows": max(persisted - matched, 0),
        }
        print(f"[MANUAL_PLK_IMPORT_INVENTORY] {json.dumps(inventory, ensure_ascii=False)}")
        return {
            "ok": True,
            "status": status,
            "id": int(row.id),
            "sourceKey": source_key,
            "importId": int(history.id),
            **_summary(parsed, matched=matched, persisted=persisted),
        }
    except Exception:
        db.rollback()
        raise
    finally:
        lock.release()


def fail_manual_import_history(
    *,
    db: Session,
    price_list_id: int | None,
    source_key: str,
    filename: str,
    error: str,
    requested_by: str = "",
) -> None:
    _create_history(
        db=db,
        parsed=None,
        price_list_id=price_list_id,
        source_key=source_key,
        filename=filename,
        status="error",
        requested_by=requested_by,
        preserved_previous_snapshot=bool(price_list_id),
        error_summary=error,
    )
    db.commit()


def list_manual_import_history(*, db: Session, price_list_id: int) -> list[dict[str, Any]]:
    rows = (
        db.execute(
            select(ManualPriceListImport)
            .where(ManualPriceListImport.competitor_price_list_id == price_list_id)
            .order_by(ManualPriceListImport.started_at.desc(), ManualPriceListImport.id.desc())
        )
        .scalars()
        .all()
    )
    return [
        {
            "id": row.id,
            "status": row.status,
            "filename": row.original_filename,
            "fileType": row.file_type,
            "checksum": row.file_checksum,
            "startedAt": row.started_at.isoformat() if row.started_at else "",
            "finishedAt": row.finished_at.isoformat() if row.finished_at else "",
            "totalRows": row.total_rows,
            "validRows": row.valid_rows,
            "invalidRows": row.invalid_rows,
            "duplicateRows": row.duplicate_rows,
            "matchedRows": row.matched_rows,
            "unmatchedRows": row.unmatched_rows,
            "persistedRows": row.persisted_rows,
            "preservedPreviousSnapshot": bool(row.preserved_previous_snapshot),
            "errorSummary": row.error_summary,
        }
        for row in rows
    ]


def list_manual_import_errors(*, db: Session, import_id: int) -> list[dict[str, Any]]:
    rows = (
        db.execute(
            select(ManualPriceListImportError)
            .where(ManualPriceListImportError.import_id == import_id)
            .order_by(ManualPriceListImportError.row_number.asc(), ManualPriceListImportError.id.asc())
        )
        .scalars()
        .all()
    )
    return [_error_dict(RowError(row.row_number, row.field, row.raw_value, row.error_code, row.message)) for row in rows]


def deactivate_manual_price_list(*, db: Session, price_list_id: int) -> dict[str, Any]:
    row = db.get(CompetitorPriceList, price_list_id)
    if row is None or row.source_type != "manual":
        raise ValueError("manual price list not found")
    db.execute(
        select(PriceFormatCompetitorAssignment).where(PriceFormatCompetitorAssignment.competitor_price_list_id == row.id)
    )
    for assignment in db.execute(
        select(PriceFormatCompetitorAssignment).where(PriceFormatCompetitorAssignment.competitor_price_list_id == row.id)
    ).scalars():
        assignment.is_active = False
        assignment.updated_at = now_kz_naive()
    row.last_refresh_status = "inactive"
    row.last_refresh_message = "manual price list deactivated"
    row.updated_at = now_kz_naive()
    rebuild_competitor_prices_for_selected(db=db, price_format_id=int(row.price_format_id))
    db.commit()
    return {"ok": True, "id": row.id, "status": "inactive"}


def delete_manual_price_list(*, db: Session, price_list_id: int) -> dict[str, Any]:
    row = db.get(CompetitorPriceList, price_list_id)
    if row is None or row.source_type != "manual":
        raise ValueError("manual price list not found")
    active_assignments = int(
        db.execute(
            select(func.count(PriceFormatCompetitorAssignment.id))
            .where(PriceFormatCompetitorAssignment.competitor_price_list_id == row.id)
            .where(PriceFormatCompetitorAssignment.is_active.is_(True))
        ).scalar_one()
    )
    if active_assignments:
        raise ValueError("manual price list has active assignments; deactivate it first")
    import_ids = [
        int(x)
        for x in db.execute(select(ManualPriceListImport.id).where(ManualPriceListImport.competitor_price_list_id == row.id)).scalars()
    ]
    if import_ids:
        db.execute(delete(ManualPriceListImportError).where(ManualPriceListImportError.import_id.in_(import_ids)))
    db.execute(delete(ManualPriceListImport).where(ManualPriceListImport.competitor_price_list_id == row.id))
    db.execute(delete(CompetitorPriceListItem).where(CompetitorPriceListItem.price_list_id == row.id))
    db.execute(delete(PriceFormatCompetitorAssignment).where(PriceFormatCompetitorAssignment.competitor_price_list_id == row.id))
    db.delete(row)
    db.commit()
    return {"ok": True, "id": price_list_id, "deleted": True}
