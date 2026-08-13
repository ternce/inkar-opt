from __future__ import annotations

import io
import json
import os
import re
import unicodedata
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import func, select
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.orm import Session

from ..models import BusinessList, BusinessListItem, ListItem, Product, UniversalList
from ..timezone import now_kz_naive, local_iso
from .sku import normalize_external_sku, normalize_sku_variants


SUPPORTED_LIST_TYPES = {"critical", "markup", "exclusion"}
UNIVERSAL_MARKUP_TYPES = {"critical_markup", "min_markup", "max_markup", "fixed_markup", "markup", "percentile_override"}
UNIVERSAL_FIXED_PRICE_TYPES = {"fixed_price", "min_price", "max_price", "memorandum"}
UNIVERSAL_EXCLUSION_TYPES = {"exclusion", "exclude_from_pricing", "no_bend"}
NON_NEGATIVE_PRICE_TYPES = {"fixed_price", "min_price", "max_price"}
PRICE_TYPE_ERROR_LABELS = {
    "fixed_price": "Фиксированная цена",
    "min_price": "Минимальная цена",
    "max_price": "Максимальная цена",
}
CRITICAL_MARKUP_NO_OVERRIDE = "-"
UNIVERSAL_TYPE_LABELS = {
    "фиксированная цена": "fixed_price",
    "минимальная цена": "min_price",
    "максимальная цена": "max_price",
    "минимальная наценка": "min_markup",
    "критическая наценка": "critical_markup",
    "максимальная наценка": "max_markup",
    "исключить из переоценки": "exclude_from_pricing",
    "без прогиба": "no_bend",
    "fixed markup": "fixed_markup",
    "fixed margin": "fixed_markup",
    "фикс наценка": "fixed_markup",
    "фиксированная наценка": "fixed_markup",
    "фиксированная маржа": "fixed_markup",
}
DEFAULT_MAX_UPLOAD_SIZE_MB = 10
DEFAULT_MAX_ROWS = 50_000


IDENTIFIER_ALIASES: dict[str, tuple[str, ...]] = {
    "material": ("material", "sku", "материал"),
    "article": ("article", "артикул"),
    "product_code": ("product code", "product_code", "productcode", "код товара", "код"),
    "registry_id": ("id", "ид", "id товара", "product id", "productid"),
}
IDENTIFIER_PRIORITY = ("material", "article", "product_code")
IDENTIFIER_ALIASES["material"] = (*IDENTIFIER_ALIASES["material"], "Материал", "Код", "GoodsID")
IDENTIFIER_PRIORITY_MEMORANDUM = ("registry_id", "material", "product_code")
PRODUCT_NAME_ALIASES = (
    "Артикул",
    "Наименование",
    "Название",
    "Name",
    "Product Name",
    "Наименование товара",
)

VALUE_ALIASES = (
    "value",
    "price",
    "fixed price",
    "markup",
    "percentage",
    "critical",
    "critical flag",
    "наценка",
    "критичка",
    "значение",
    "исключение",
    "exclude",
    "excluded",
)

PRICE_TYPE_VALUE_ALIASES: dict[str, tuple[str, ...]] = {
    "fixed_markup": (
        "фикс наценка",
        "фиксированная наценка",
        "наценка",
        "критичка",
        "критическая наценка",
        "значение",
        "%",
        "процент",
    ),
    "max_price": ("макс цена", "максимальная цена", "цена"),
    "min_price": ("мин цена", "минимальная цена", "цена"),
    "fixed_price": ("фикс цена", "фиксированная цена", "цена"),
}

PRICE_TYPE_VALUE_COLUMN_ERRORS: dict[str, str] = {
    "fixed_markup": "Для списка типа Фиксированная наценка нужна колонка значения",
    "max_price": "Для списка типа Максимальная цена нужна колонка значения:\nМакс цена / Максимальная цена / Цена",
    "min_price": "Для списка типа Минимальная цена нужна колонка значения:\nМин цена / Минимальная цена / Цена",
    "fixed_price": "Для списка типа Фиксированная цена нужна колонка значения:\nФикс цена / Фиксированная цена / Цена",
}
UNIVERSAL_TYPE_LABELS.update({"меморандум": "memorandum", "Меморандум": "memorandum", "РњРµРјРѕСЂР°РЅРґСѓРј": "memorandum"})
PRICE_TYPE_VALUE_ALIASES["memorandum"] = ("максимальная цена по меморандуму", "меморандум", "меморандум цена", "макс цена", "цена")
PRICE_TYPE_VALUE_ALIASES["memorandum"] = (
    *PRICE_TYPE_VALUE_ALIASES["memorandum"],
    "Меморандум",
    "Максимальная цена",
    "Цена меморандума",
    "Максимальная цена по меморандуму",
    "Предельная цена",
    "Предельная цена для оптовой реализации",
    "Предельная цена для розничной реализации",
    "Предельная оптовая цена",
    "Предельная розничная цена",
    "Максимальная оптовая цена",
    "Максимальная розничная цена",
    "Цена для оптовой реализации",
    "Цена для розничной реализации",
    "Регулируемая цена",
    "Государственная предельная цена",
    "Цена по госреестру",
    "Предельная стоимость",
)
PRICE_TYPE_VALUE_COLUMN_ERRORS["memorandum"] = "Для списка типа Меморандум нужна колонка значения:\nМаксимальная цена по меморандуму / Меморандум / Цена"
MEMORANDUM_PRICE_COLUMN_PRIORITY = tuple(
    normalize.casefold()
    for normalize in (
        "Предельная цена для оптовой реализации",
        "Предельная оптовая цена",
        "Максимальная оптовая цена",
        "Цена для оптовой реализации",
        "Меморандум",
        "Максимальная цена по меморандуму",
        "Цена меморандума",
        "Предельная цена",
        "Максимальная цена",
        "Предельная цена для розничной реализации",
        "Предельная розничная цена",
        "Максимальная розничная цена",
        "Цена для розничной реализации",
        "Регулируемая цена",
        "Государственная предельная цена",
        "Цена по госреестру",
        "Предельная стоимость",
    )
)


class ImportHeaderError(ValueError):
    def __init__(
        self,
        message: str,
        *,
        detected_headers: list[str],
        missing_fields: list[str],
        supported_aliases: dict[str, list[str]],
        status_code: int = 422,
        extra: dict[str, Any] | None = None,
    ):
        super().__init__(message)
        self.status_code = status_code
        self.detail = {
            "detail": message,
            "detectedHeaders": detected_headers,
            "missingFields": missing_fields,
            "supportedAliases": supported_aliases,
        }
        if extra:
            self.detail.update(extra)


class ImportRowLimitError(ValueError):
    def __init__(self, *, limit: int, meaningful_rows_detected: int):
        super().__init__(f"Количество заполненных строк превышает допустимый лимит {limit}")
        self.status_code = 400
        self.detail = {
            "detail": f"Количество заполненных строк превышает допустимый лимит {limit}",
            "limit": limit,
            "meaningfulRowsDetected": meaningful_rows_detected,
        }


def _missing_value_column_error(list_type: str, headers: list[str]) -> str:
    expected = PRICE_TYPE_VALUE_ALIASES.get(list_type, VALUE_ALIASES)
    prefix = PRICE_TYPE_VALUE_COLUMN_ERRORS.get(list_type, "Не найдена колонка значения")
    detected_text = ", ".join(header or "<пусто>" for header in headers) or "<нет заголовков>"
    expected_text = ", ".join(expected)
    return f"{prefix}. Обнаружены заголовки: {detected_text}. Ожидался один из заголовков: {expected_text}."

MANUFACTURER_ALIASES = ("manufacturer", "producer", "производитель")

TRUE_VALUES = {"1", "true", "yes", "critical", "y", "да", "истина"}
FALSE_VALUES = {"0", "false", "no", "n", "нет", "ложь"}


def max_upload_size_bytes() -> int:
    try:
        mb = int(os.getenv("LIST_IMPORT_MAX_UPLOAD_SIZE_MB", str(DEFAULT_MAX_UPLOAD_SIZE_MB)))
    except ValueError:
        mb = DEFAULT_MAX_UPLOAD_SIZE_MB
    return max(1, mb) * 1024 * 1024


def max_rows() -> int:
    try:
        value = int(os.getenv("LIST_IMPORT_MAX_ROWS", str(DEFAULT_MAX_ROWS)))
    except ValueError:
        value = DEFAULT_MAX_ROWS
    return max(1, value)


@dataclass
class ImportIssue:
    row: int
    code: str
    message: str
    identifier: str = ""
    name: str = ""
    field: str = ""
    column: str = ""

    def to_dict(self) -> dict[str, Any]:
        return {
            "row": self.row,
            "code": self.code,
            "message": self.message,
            "identifier": self.identifier,
            "name": self.name,
            "field": self.field,
            "column": self.column or self.field,
        }


@dataclass
class ParsedListItem:
    product: Product
    sku: str
    manufacturer: str
    value_json: dict[str, Any]
    value_decimal: Decimal | None
    value_bool: bool | None
    source_row: int
    source_identifier: str


@dataclass
class ParsedUniversalListItem:
    product: Product
    value: Decimal
    special_value: str
    source_row: int
    source_identifier: str


def normalize_header(value: object) -> str:
    text = unicodedata.normalize("NFKC", str(value or ""))
    text = text.replace("\u00a0", " ").replace("ё", "е").replace("Ё", "Е")
    text = text.strip().casefold()
    text = re.sub(r"[_\-]+", " ", text)
    text = re.sub(r"\s+", " ", text)
    return text


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, Decimal):
        if value == value.to_integral_value():
            return str(int(value))
        return format(value, "f")
    return str(value).strip()


def _identifier_text(value: object) -> str:
    text = _cell_text(value).strip()
    if not text:
        return ""
    normalized = text.replace("\u00a0", "").replace(" ", "").replace(",", ".")
    try:
        number = Decimal(normalized)
    except (InvalidOperation, ValueError):
        return text
    if number == number.to_integral_value():
        return str(int(number))
    return text


def _is_empty_row(row: tuple[Any, ...]) -> bool:
    return all(_cell_text(value) == "" for value in row)


def _cell_at(row: tuple[Any, ...], idx: int | None) -> object:
    if idx is None or idx < 0 or idx >= len(row):
        return None
    return row[idx]


def _is_meaningful_import_row(row: tuple[Any, ...], relevant_indexes: set[int]) -> bool:
    if not relevant_indexes:
        return not _is_empty_row(row)
    for idx in relevant_indexes:
        if _cell_text(_cell_at(row, idx)):
            return True
    return False


def _supported_identifier_aliases(list_type: str) -> dict[str, list[str]]:
    normalized = str(list_type or "").strip().casefold()
    normalized = UNIVERSAL_TYPE_LABELS.get(normalized, normalized)
    aliases: dict[str, list[str]] = {
        "product_code": ["Материал", "SKU", "Код", "GoodsID"],
    }
    if normalized == "memorandum":
        aliases["registry_id"] = ["ID", "ИД", "ID товара", "Product ID", "ProductID"]
        aliases["product_code"].append("Код товара")
    return aliases


def _supported_price_aliases(list_type: str) -> list[str]:
    normalized = str(list_type or "").strip().casefold()
    normalized = UNIVERSAL_TYPE_LABELS.get(normalized, normalized)
    if normalized == "memorandum":
        return list(PRICE_TYPE_VALUE_ALIASES["memorandum"])
    return list(VALUE_ALIASES)


def _find_alias_indexes(headers: list[str], aliases: tuple[str, ...]) -> list[int]:
    alias_set = {normalize_header(alias) for alias in aliases}
    return [idx for idx, header in enumerate(headers) if header in alias_set]


def _find_memorandum_price_column(headers: list[str]) -> tuple[int | None, list[int]]:
    recognized = _find_alias_indexes(headers, PRICE_TYPE_VALUE_ALIASES["memorandum"])
    if not recognized:
        return None, []
    if len(recognized) == 1:
        return recognized[0], recognized
    return None, recognized


def _find_header_indexes(
    headers: list[str],
    *,
    additional_value_aliases: tuple[str, ...] = (),
) -> tuple[dict[str, int], int | None, int | None]:
    identifier_indexes: dict[str, int] = {}
    for identifier_type, aliases in IDENTIFIER_ALIASES.items():
        alias_set = {normalize_header(alias) for alias in aliases}
        for idx, header in enumerate(headers):
            if header in alias_set:
                identifier_indexes[identifier_type] = idx
                break

    value_idx = None
    value_aliases = {normalize_header(alias) for alias in (*VALUE_ALIASES, *additional_value_aliases)}
    for idx, header in enumerate(headers):
        if header in value_aliases:
            value_idx = idx
            break

    manufacturer_idx = None
    manufacturer_aliases = {normalize_header(alias) for alias in MANUFACTURER_ALIASES}
    for idx, header in enumerate(headers):
        if header in manufacturer_aliases:
            manufacturer_idx = idx
            break
    return identifier_indexes, value_idx, manufacturer_idx


def _identifier_for_row(row: tuple[Any, ...], identifier_indexes: dict[str, int]) -> tuple[str, str]:
    for identifier_type in IDENTIFIER_PRIORITY:
        idx = identifier_indexes.get(identifier_type)
        if idx is None or idx >= len(row):
            continue
        value = _identifier_text(row[idx])
        if value:
            return identifier_type, value
    return "", ""


def _identifier_for_universal_row(
    row: tuple[Any, ...],
    identifier_indexes: dict[str, int],
    *,
    list_type: str,
) -> tuple[str, str]:
    normalized = str(list_type or "").strip().casefold()
    normalized = UNIVERSAL_TYPE_LABELS.get(normalized, normalized)
    priority = IDENTIFIER_PRIORITY_MEMORANDUM if normalized == "memorandum" else IDENTIFIER_PRIORITY
    for identifier_type in priority:
        idx = identifier_indexes.get(identifier_type)
        if idx is None or idx >= len(row):
            continue
        value = _identifier_text(row[idx])
        if value:
            return identifier_type, value
    return "", ""


def _product_lookup_keys(value: object) -> list[str]:
    raw = normalize_external_sku(value)
    keys: list[str] = []
    for key in [raw, *normalize_sku_variants(value)]:
        if key and key not in keys:
            keys.append(key)
    return keys


def find_product_by_identifier(db: Session, identifier: object) -> Product | None:
    keys = _product_lookup_keys(identifier)
    if not keys:
        return None
    return db.execute(select(Product).where(Product.code.in_(keys)).limit(1)).scalars().first()


def find_product_by_identifier_type(db: Session, identifier: object, identifier_type: str) -> tuple[Product | None, str]:
    product = find_product_by_identifier(db, identifier)
    if product is not None:
        return product, "Product.code"
    return None, "Product.code"


def _parse_bool(value: object) -> bool | None:
    text = _cell_text(value).casefold()
    if text in TRUE_VALUES:
        return True
    if text in FALSE_VALUES:
        return False
    return None


def _parse_markup_percent(value: object) -> Decimal | None:
    text = _cell_text(value).replace(" ", "").replace(",", ".")
    if not text:
        return None
    is_percent = text.endswith("%")
    if is_percent:
        text = text[:-1]
    try:
        number = Decimal(text)
    except (InvalidOperation, ValueError):
        return None
    if not is_percent and Decimal("0") < number < Decimal("1"):
        number *= Decimal("100")
    return number.quantize(Decimal("0.000001"))


def _parse_decimal_value(value: object) -> Decimal | None:
    text = _cell_text(value).replace("\u00a0", "").replace(" ", "").replace(",", ".")
    if not text:
        return None
    try:
        return Decimal(text).quantize(Decimal("0.000001"))
    except (InvalidOperation, ValueError):
        return None


def parse_list_decimal(value: object) -> Decimal | None:
    return _parse_decimal_value(value)


def _format_percent(value: Decimal) -> str:
    text = format(value.normalize(), "f")
    if "." in text:
        text = text.rstrip("0").rstrip(".")
    return f"{text}%"


def _normalize_value(list_type: str, raw_value: object) -> tuple[dict[str, Any], Decimal | None, bool | None, str | None]:
    if list_type == "critical":
        parsed = _parse_bool(raw_value)
        if parsed is None:
            return {}, None, None, "invalid critical value"
        return {"is_critical": parsed}, None, parsed, None
    if list_type == "markup":
        percent = _parse_markup_percent(raw_value)
        if percent is None:
            return {}, None, None, "invalid markup value"
        return {"markup_percent": float(percent), "display": _format_percent(percent)}, percent, None, None
    if list_type == "exclusion":
        parsed = _parse_bool(raw_value)
        if parsed is None and _cell_text(raw_value) == "":
            parsed = True
        if parsed is None:
            return {}, None, None, "invalid exclusion value"
        return {"is_excluded": parsed}, None, parsed, None
    return {}, None, None, "unsupported list type"


def _universal_import_behavior(list_type: str) -> str:
    normalized = str(list_type or "").strip().casefold()
    normalized = UNIVERSAL_TYPE_LABELS.get(normalized, normalized)
    if normalized in UNIVERSAL_MARKUP_TYPES:
        return "markup"
    if normalized in UNIVERSAL_FIXED_PRICE_TYPES:
        return "decimal"
    if normalized in UNIVERSAL_EXCLUSION_TYPES:
        return "exclusion"
    return "decimal"


def is_exclude_from_pricing_type(list_type: str) -> bool:
    normalized = str(list_type or "").strip().casefold()
    normalized = UNIVERSAL_TYPE_LABELS.get(normalized, normalized)
    return normalized in {"exclude_from_pricing", "exclusion"}


def _normalize_universal_value(list_type: str, raw_value: object) -> tuple[Decimal | None, str | None]:
    normalized = str(list_type or "").strip().casefold()
    normalized = UNIVERSAL_TYPE_LABELS.get(normalized, normalized)
    behavior = _universal_import_behavior(list_type)
    if behavior == "markup":
        value = _parse_markup_percent(raw_value)
        if value is None:
            return None, "invalid markup value"
        return value, None
    if behavior == "exclusion":
        parsed = _parse_bool(raw_value)
        if parsed is None and _cell_text(raw_value) == "":
            parsed = True
        if parsed is None:
            return None, "invalid exclusion value"
        return Decimal("1") if parsed else Decimal("0"), None
    value = _parse_decimal_value(raw_value)
    if value is None:
        return None, "Цена обязательна" if normalized == "memorandum" else "invalid numeric value"
    if normalized == "memorandum" and value <= 0:
        return None, "Цена должна быть больше нуля"
    error = validate_universal_list_price_value(normalized, value)
    if error:
        return None, error
    return value, None


def normalize_universal_list_value(list_type: str, raw_value: object) -> tuple[Decimal | None, str | None]:
    return _normalize_universal_value(list_type, raw_value)


def validate_universal_list_price_value(list_type: str, value: Decimal | None, *, sku: object = None) -> str | None:
    normalized = str(list_type or "").strip().casefold()
    normalized = UNIVERSAL_TYPE_LABELS.get(normalized, normalized)
    if normalized not in NON_NEGATIVE_PRICE_TYPES or value is None or value >= 0:
        return None
    label = PRICE_TYPE_ERROR_LABELS.get(normalized, "Цена")
    sku_text = _cell_text(sku)
    sku_part = f" SKU: {sku_text}," if sku_text else ""
    return f"{label} не может быть отрицательной.{sku_part} значение: {value:g}"


def normalize_universal_list_item_value(
    list_type: str, raw_value: object, *, sku: object = None
) -> tuple[Decimal | None, str, str | None]:
    normalized_type = str(list_type or "").strip().casefold()
    normalized_type = UNIVERSAL_TYPE_LABELS.get(normalized_type, normalized_type)
    if normalized_type == "critical_markup" and _cell_text(raw_value) == CRITICAL_MARKUP_NO_OVERRIDE:
        # The numeric column remains populated only for storage compatibility;
        # special_value is the authoritative domain value.
        return Decimal("0"), CRITICAL_MARKUP_NO_OVERRIDE, None
    value, error = _normalize_universal_value(list_type, raw_value)
    if error is None:
        error = validate_universal_list_price_value(list_type, value, sku=sku)
    return value, "", error


def _summary(total_rows: int = 0) -> dict[str, int]:
    return {
        "total_rows": total_rows,
        "processed": 0,
        "not_found": 0,
        "duplicates": 0,
        "empty_rows": 0,
        "invalid_rows": 0,
        "errors": 0,
    }


def import_business_list_excel(
    *,
    db: Session,
    content: bytes,
    filename: str,
    list_type: str,
) -> dict[str, Any]:
    list_type = str(list_type or "").strip().casefold()
    if list_type not in SUPPORTED_LIST_TYPES:
        raise ValueError("list_type must be one of: critical, markup, exclusion")
    if not filename.lower().endswith(".xlsx"):
        raise ValueError(".xlsx files are supported; .xls support requires an installed reader library")
    if len(content) > max_upload_size_bytes():
        limit_mb = max_upload_size_bytes() // (1024 * 1024)
        raise ValueError(f"file exceeds LIST_IMPORT_MAX_UPLOAD_SIZE_MB={limit_mb}")

    try:
        workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    except Exception as exc:
        raise ValueError(f"invalid Excel file: {exc}") from exc

    worksheet = workbook.active
    rows_iter = worksheet.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise ValueError("empty file")
    if _is_empty_row(header_row):
        raise ValueError("file without headers")

    headers = [normalize_header(value) for value in header_row]
    identifier_indexes, value_idx, manufacturer_idx = _find_header_indexes(headers)
    if not identifier_indexes:
        raise ValueError("missing product identifier column")
    if value_idx is None:
        raise ValueError("missing value column")

    summary = _summary()
    issues: list[ImportIssue] = []
    items_by_product: dict[int, ParsedListItem] = {}

    row_limit = max_rows()
    for excel_row_number, row in enumerate(rows_iter, start=2):
        summary["total_rows"] += 1
        if summary["total_rows"] > row_limit:
            raise ValueError(f"row count exceeds LIST_IMPORT_MAX_ROWS={row_limit}")
        if _is_empty_row(row):
            summary["empty_rows"] += 1
            continue

        identifier_type, identifier = _identifier_for_row(row, identifier_indexes)
        if not identifier:
            summary["invalid_rows"] += 1
            summary["errors"] += 1
            issues.append(ImportIssue(excel_row_number, "missing_required_field", "missing product identifier", field="identifier"))
            continue

        raw_value = row[value_idx] if value_idx is not None and value_idx < len(row) else None
        value_json, value_decimal, value_bool, error = _normalize_value(list_type, raw_value)
        if error:
            summary["invalid_rows"] += 1
            summary["errors"] += 1
            issues.append(ImportIssue(excel_row_number, "invalid_value", error, identifier=identifier, field="value"))
            continue

        product = find_product_by_identifier(db, identifier)
        if product is None:
            summary["not_found"] += 1
            issues.append(
                ImportIssue(
                    excel_row_number,
                    "product_not_found",
                    f"product not found by {identifier_type}",
                    identifier=identifier,
                    field=identifier_type,
                )
            )
            continue

        if product.id in items_by_product:
            summary["duplicates"] += 1
            issues.append(ImportIssue(excel_row_number, "duplicate_row", "duplicate product row; last row wins", identifier=identifier))

        manufacturer = _cell_text(row[manufacturer_idx]) if manufacturer_idx is not None and manufacturer_idx < len(row) else ""
        items_by_product[product.id] = ParsedListItem(
            product=product,
            sku=product.code,
            manufacturer=manufacturer,
            value_json=value_json,
            value_decimal=value_decimal,
            value_bool=value_bool,
            source_row=excel_row_number,
            source_identifier=identifier,
        )

    summary["processed"] = len(items_by_product)

    try:
        business_list = BusinessList(
            list_type=list_type,
            name=f"{list_type} import {now_kz_naive().strftime('%Y-%m-%d %H:%M:%S')}",
            original_filename=filename,
            status="imported",
            summary_json=json.dumps(summary, ensure_ascii=False),
            errors_json=json.dumps([issue.to_dict() for issue in issues], ensure_ascii=False),
            item_count=len(items_by_product),
        )
        db.add(business_list)
        db.flush()

        for item in items_by_product.values():
            db.add(
                BusinessListItem(
                    business_list_id=business_list.id,
                    product_id=item.product.id,
                    sku=item.sku,
                    product_name=item.product.name,
                    manufacturer=item.manufacturer,
                    value_json=json.dumps(item.value_json, ensure_ascii=False),
                    value_decimal=item.value_decimal,
                    value_bool=item.value_bool,
                    source_row=item.source_row,
                    source_identifier=item.source_identifier,
                )
            )
        db.commit()
    except SQLAlchemyError:
        db.rollback()
        raise
    except Exception:
        db.rollback()
        raise
    db.refresh(business_list)
    return business_list_to_dict(business_list, include_errors=True)


def import_universal_list_excel(
    *,
    db: Session,
    universal_list: UniversalList,
    content: bytes,
    filename: str,
) -> dict[str, Any]:
    if not filename.lower().endswith(".xlsx"):
        raise ValueError(".xlsx files are supported; .xls support requires an installed reader library")
    if len(content) > max_upload_size_bytes():
        limit_mb = max_upload_size_bytes() // (1024 * 1024)
        raise ValueError(f"file exceeds LIST_IMPORT_MAX_UPLOAD_SIZE_MB={limit_mb}")

    try:
        workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    except Exception as exc:
        raise ValueError(f"invalid Excel file: {exc}") from exc

    worksheet = workbook.active
    rows_iter = worksheet.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise ValueError("empty file")
    if _is_empty_row(header_row):
        raise ValueError("file without headers")

    headers = [normalize_header(value) for value in header_row]
    normalized_list_type = str(universal_list.type or "").strip().casefold()
    normalized_list_type = UNIVERSAL_TYPE_LABELS.get(normalized_list_type, normalized_list_type)
    identifier_indexes, value_idx, _manufacturer_idx = _find_header_indexes(
        headers,
        additional_value_aliases=PRICE_TYPE_VALUE_ALIASES.get(normalized_list_type, ()),
    )
    detected_headers = [_cell_text(value) for value in header_row if _cell_text(value)]
    detected_price_indexes: list[int] = []
    if normalized_list_type == "memorandum":
        value_idx, detected_price_indexes = _find_memorandum_price_column(headers)
    detected_price_columns = [
        _cell_text(header_row[idx])
        for idx in detected_price_indexes
        if idx < len(header_row)
    ]
    if normalized_list_type == "memorandum" and len(detected_price_columns) > 1:
        raise ImportHeaderError(
            "Найдено несколько колонок предельной цены",
            detected_headers=detected_headers,
            missing_fields=[],
            supported_aliases={
                "price": _supported_price_aliases(normalized_list_type),
            },
            extra={
                "detectedPriceColumns": detected_price_columns,
            },
        )
    if not identifier_indexes:
        if normalized_list_type != "memorandum":
            raise ValueError("missing product identifier column")
        raise ImportHeaderError(
            "Не найдена колонка идентификатора товара",
            detected_headers=detected_headers,
            missing_fields=["product_identifier"],
            supported_aliases=_supported_identifier_aliases(normalized_list_type),
        )
    exclusion_by_presence = is_exclude_from_pricing_type(str(universal_list.type or ""))
    if value_idx is None and not exclusion_by_presence:
        if normalized_list_type != "memorandum":
            raise ValueError(_missing_value_column_error(normalized_list_type, headers))
        raise ImportHeaderError(
            "Не найдена колонка предельной цены",
            detected_headers=detected_headers,
            missing_fields=["memorandum_price"],
            supported_aliases={
                "price": _supported_price_aliases(normalized_list_type),
            },
        )
    value_column = header_row[value_idx] if value_idx is not None and value_idx < len(header_row) else "value"
    product_name_indexes = _find_alias_indexes(headers, PRODUCT_NAME_ALIASES) if normalized_list_type == "memorandum" else []
    product_name_column = _cell_text(header_row[product_name_indexes[0]]) if product_name_indexes else ""
    relevant_row_indexes = {idx for idx in identifier_indexes.values() if idx is not None}
    if value_idx is not None:
        relevant_row_indexes.add(value_idx)
    relevant_row_indexes.update(product_name_indexes)
    identifier_priority = IDENTIFIER_PRIORITY_MEMORANDUM if normalized_list_type == "memorandum" else IDENTIFIER_PRIORITY
    selected_identifier_type = next((key for key in identifier_priority if key in identifier_indexes), "")
    selected_identifier_idx = identifier_indexes.get(selected_identifier_type)
    selected_identifier_column = (
        _cell_text(header_row[selected_identifier_idx])
        if selected_identifier_idx is not None and selected_identifier_idx < len(header_row)
        else ""
    )

    summary = _summary()
    summary.update(
        {
            "imported": 0,
            "updated": 0,
            "skipped": 0,
            "failed": 0,
            "selectedPriceColumn": _cell_text(value_column),
            "detectedPriceColumns": detected_price_columns,
            "productNameColumn": product_name_column,
            "selectedIdentifierColumn": selected_identifier_column,
            "resolvedIdentifierFields": [],
            "physicalRowsVisited": 0,
            "meaningfulDataRows": 0,
            "emptyRowsIgnored": 0,
        }
    )
    issues: list[ImportIssue] = []
    warnings: list[dict[str, Any]] = []
    resolved_identifier_fields: set[str] = set()
    items_by_product: dict[int, ParsedUniversalListItem] = {}
    row_limit = max_rows()
    for excel_row_number, row in enumerate(rows_iter, start=2):
        summary["physicalRowsVisited"] += 1
        if not _is_meaningful_import_row(row, relevant_row_indexes):
            summary["empty_rows"] += 1
            summary["emptyRowsIgnored"] += 1
            continue
        summary["total_rows"] += 1
        summary["meaningfulDataRows"] = summary["total_rows"]
        if summary["total_rows"] > row_limit:
            raise ImportRowLimitError(limit=row_limit, meaningful_rows_detected=summary["total_rows"])

        identifier_type, identifier = _identifier_for_universal_row(row, identifier_indexes, list_type=normalized_list_type)
        row_product_name = (
            _cell_text(row[product_name_indexes[0]])
            if product_name_indexes and product_name_indexes[0] < len(row)
            else ""
        )
        if not identifier:
            summary["invalid_rows"] += 1
            summary["errors"] += 1
            summary["failed"] += 1
            issues.append(ImportIssue(excel_row_number, "missing_required_field", "Идентификатор товара обязателен", name=row_product_name, field="identifier", column="identifier"))
            continue

        raw_value = row[value_idx] if value_idx is not None and value_idx < len(row) else None
        value, special_value, error = normalize_universal_list_item_value(
            str(universal_list.type or ""), raw_value, sku=identifier
        )
        if error or value is None:
            summary["invalid_rows"] += 1
            summary["errors"] += 1
            summary["failed"] += 1
            issues.append(
                ImportIssue(
                    excel_row_number,
                    "invalid_value",
                    error or "invalid value",
                    identifier=identifier,
                    name=row_product_name,
                    field="value",
                    column=str(value_column or "value"),
                )
            )
            continue

        product, resolved_identifier_field = find_product_by_identifier_type(db, identifier, identifier_type)
        if product is None:
            summary["not_found"] += 1
            summary["failed"] += 1
            missing_label = "ID" if identifier_type == "registry_id" else "SKU"
            missing_message = (
                "Товар с указанным ID отсутствует в локальном справочнике"
                if normalized_list_type == "memorandum" and identifier_type == "registry_id"
                else f"Товар с указанным {missing_label} не найден"
                if normalized_list_type == "memorandum"
                else f"product not found by {identifier_type}"
            )
            issues.append(
                ImportIssue(
                    excel_row_number,
                    "product_not_found",
                    missing_message,
                    identifier=identifier,
                    name=row_product_name,
                    field=identifier_type,
                    column=identifier_type,
                )
            )
            continue
        resolved_identifier_fields.add(resolved_identifier_field)

        if product.id in items_by_product:
            summary["duplicates"] += 1
            warning = {
                "row": excel_row_number,
                "code": "duplicate_row",
                "message": "Найден дубль ID; использована последняя строка" if identifier_type == "registry_id" else "Найден дубль SKU; использована последняя строка",
                "identifier": identifier,
                "name": row_product_name,
                "resolvedIdentifierField": resolved_identifier_field,
            }
            warnings.append(warning)
            issues.append(ImportIssue(excel_row_number, "duplicate_row", "duplicate product row; last row wins", identifier=identifier))

        items_by_product[product.id] = ParsedUniversalListItem(
            product=product,
            value=value,
            special_value=special_value,
            source_row=excel_row_number,
            source_identifier=identifier,
        )

    invalid_value_issues = [issue for issue in issues if issue.code == "invalid_value"]
    if invalid_value_issues and normalized_list_type != "memorandum":
        first = invalid_value_issues[0]
        raise ValueError(
            f"Некорректное значение в строке {first.row}: {first.message}. "
            "Допустимы числа вида 10, 10,5 или 10.5."
        )

    summary["processed"] = len(items_by_product)

    try:
        for item in items_by_product.values():
            existing = db.execute(
                select(ListItem)
                .where(ListItem.universal_list_id == universal_list.id)
                .where(ListItem.product_id == item.product.id)
            ).scalars().first()
            if existing:
                existing.value = item.value
                existing.special_value = item.special_value
                summary["updated"] += 1
            else:
                db.add(
                    ListItem(
                        universal_list_id=universal_list.id,
                        product_id=item.product.id,
                        value=item.value,
                        special_value=item.special_value,
                    )
                )
                summary["imported"] += 1
        db.commit()
    except SQLAlchemyError:
        db.rollback()
        raise
    except Exception:
        db.rollback()
        raise

    item_count = db.scalar(select(func.count(ListItem.id)).where(ListItem.universal_list_id == universal_list.id))
    if item_count is None:
        item_count = 0
    summary["skipped"] = summary["not_found"] + summary["invalid_rows"] + summary["empty_rows"]
    summary["matched"] = len(items_by_product)
    summary["resolvedIdentifierFields"] = sorted(resolved_identifier_fields)
    success = summary["failed"] == 0
    row_errors = [issue.to_dict() for issue in issues]
    return {
        "success": success,
        "status": "ok",
        "list_id": universal_list.id,
        "list_type": str(universal_list.type or ""),
        "filename": filename,
        "item_count": int(item_count),
        "totalRows": summary["total_rows"],
        "matchedRows": summary["matched"],
        "importedRows": summary["imported"],
        "updatedRows": summary["updated"],
        "unmatchedRows": summary["not_found"],
        "invalidRows": summary["invalid_rows"],
        "duplicateRows": summary["duplicates"],
        "skippedRows": summary["skipped"],
        "failedRows": summary["failed"],
        "selectedIdentifierColumn": selected_identifier_column,
        "selectedPriceColumn": summary["selectedPriceColumn"],
        "detectedPriceColumns": detected_price_columns,
        "productNameColumn": product_name_column,
        "warnings": warnings,
        "rowErrors": row_errors,
        "summary": summary,
        "errors": row_errors,
    }


def business_list_to_dict(row: BusinessList, *, include_errors: bool = False, item_preview: list[dict[str, Any]] | None = None) -> dict[str, Any]:
    payload = {
        "id": row.id,
        "list_type": row.list_type,
        "name": row.name,
        "filename": row.original_filename,
        "status": row.status,
        "item_count": row.item_count,
        "summary": json.loads(row.summary_json or "{}"),
        "created_at": local_iso(row.created_at) if row.created_at else None,
        "updated_at": local_iso(row.updated_at) if row.updated_at else None,
    }
    if include_errors:
        payload["errors"] = json.loads(row.errors_json or "[]")
    if item_preview is not None:
        payload["items"] = item_preview
    return payload


def business_list_item_to_dict(row: BusinessListItem) -> dict[str, Any]:
    return {
        "id": row.id,
        "product_id": row.product_id,
        "sku": row.sku,
        "product_name": row.product_name,
        "manufacturer": row.manufacturer,
        "value": json.loads(row.value_json or "{}"),
        "value_decimal": float(row.value_decimal) if row.value_decimal is not None else None,
        "value_bool": row.value_bool,
        "source_row": row.source_row,
        "source_identifier": row.source_identifier,
    }
