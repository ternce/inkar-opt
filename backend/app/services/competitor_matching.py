from __future__ import annotations

import json
import logging
import os
import re
import time
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal
from difflib import SequenceMatcher

from openpyxl import Workbook
from sqlalchemy import delete, select
from sqlalchemy.orm import Session

from ..models import (
    BranchCost,
    BranchStock,
    CompetitorPrice,
    CompetitorCodeMapping,
    CompetitorPriceList,
    CompetitorPriceListItem,
    Product,
    ProductExtra,
    ProductRating,
    ProductSubstituteMatch,
    SourceGoodsMatch,
)
from .competitors.code_mappings import apply_manual_mappings_to_items
from .competitor_assignments import get_assigned_competitor_price_lists
from .competitor_source_config import MULTI_PRICE_PERCENTILE_MODE, default_percentile_mode_for_source
from .manufacturers import normalize_manufacturer
from .sku import normalize_external_sku, normalize_sku, normalize_sku_variants

logger = logging.getLogger(__name__)
PROVISOR_REFERENCE_FILIAL_ID = 128
PROVISOR_REFERENCE_FILIAL_IDS = (128, 133)
MATCH_DEBUG_VERBOSE = str(os.getenv("MATCH_DEBUG_VERBOSE", "false")).strip().lower() in {"1", "true", "yes", "on"}
CRITICAL_NAME_TOKENS = frozenset(
    {
        "СТРОНГ",
        "ПЛЮС",
        "СОФТ",
        "МИНИ",
        "ФОРТЕ",
        "ЭКСТРА",
        "ГЕНТА",
        "ГК",
        "УЛЬТРА",
        "ФИАЛКА",
        "ЛИПА",
        "ХВОЩ",
        "СОСНА",
    }
)
WEIGHT_FORMS = frozenset(
    {
        "БАДЯГА",
        "БАЛЬЗАМ",
        "ВАТА",
        "ГЕЛЬ",
        "КОСМЕТИКА",
        "КРЕМ",
        "МАСЛО",
        "МАЗЬ",
        "МОЛОЧКО",
        "ПАСТА",
        "ПОРОШОК",
        "РАСТИТЕЛЬНОЕ",
        "РАСТИТЕЛЬНЫЙ",
        "РАСТИТЕЛЬНОГО",
        "САШЕ",
        "СБОР",
        "СЫРЬЕ",
        "ТРАВА",
        "ФИТО",
        "ФИТОЧАЙ",
        "ЧАЙ",
        "ШАМПУНЬ",
    }
)
MEDICINAL_GRAM_DOSAGE_TOKENS = frozenset({"ГРАН", "ГРАНУЛЫ", "МОНУРАЛ", "СМЕКТА"})
LIQUID_VOLUME_FORMS = frozenset({"СУСП", "СИРОП", "Р-Р", "КАПЛИ", "СПРЕЙ"})


def _verbose_debug(message: str, *args: object) -> None:
    if MATCH_DEBUG_VERBOSE:
        logger.debug(message, *args)


def _keeps_multi_prices_per_sku(price_list: CompetitorPriceList) -> bool:
    return default_percentile_mode_for_source(price_list) == MULTI_PRICE_PERCENTILE_MODE


def _timing(operation: str, step: str, started_at: float) -> None:
    logger.info("[TIMING] operation=%s step=%s elapsed_ms=%s", operation, step, round((time.perf_counter() - started_at) * 1000, 2))


try:
    from rapidfuzz import fuzz as rapidfuzz_fuzz
except Exception:
    rapidfuzz_fuzz = None

FORM_SYNONYMS = {
    "КАПСУЛЫ": "КАПС",
    "КАПСУЛА": "КАПС",
    "КАПС.": "КАПС",
    "КАПС": "КАПС",
    "ТАБЛЕТКИ": "ТАБ",
    "ТАБЛЕТКА": "ТАБ",
    "ТАБЛ": "ТАБ",
    "ТАБ.": "ТАБ",
    "ТАБ": "ТАБ",
    "СИРОП": "СИРОП",
    "СУСПЕНЗИЯ": "СУСП",
    "СУСП": "СУСП",
    "РАСТВОР": "Р-Р",
    "Р-Р": "Р-Р",
    "АМПУЛЫ": "АМП",
    "АМП": "АМП",
    "МАЗЬ": "МАЗЬ",
    "КРЕМ": "КРЕМ",
    "ГЕЛЬ": "ГЕЛЬ",
    "КАПЛИ": "КАПЛИ",
    "СПРЕЙ": "СПРЕЙ",
    "НАЗ": "СПРЕЙ",
    "НАЗАЛЬНЫЙ": "СПРЕЙ",
    "НАЗАЛЬНЫЕ": "СПРЕЙ",
    "ПОР": "ПОР",
    "ПОРОШОК": "ПОР",
    "ГРАН": "ГРАН",
    "ГРАНУЛЫ": "ГРАН",
    "СУПП": "СУПП",
    "СВЕЧИ": "СУПП",
    "СУППОЗИТОРИИ": "СУПП",
    "СУППОЗИТОРИЙ": "СУПП",
    "РЕК": "СУПП",
    "РЕКТ": "СУПП",
    "РЕКТАЛЬНЫЕ": "СУПП",
    "РЕКТАЛЬНЫЙ": "СУПП",
}

STRUCTURED_FIELDS = (
    "quantity",
    "volume",
    "weight",
    "dosage",
    "dosage_volume",
    "strength_signature",
    "concentration",
    "percent_strength",
    "iu_dosage",
    "form",
    "dimensions",
    "critical_tokens",
)
STRICT_MISSING_CANDIDATE_FIELDS = STRUCTURED_FIELDS
STRICT_STRUCTURE_REASONS = {
    "quantity": ("quantity_conflict", "missing_candidate_quantity"),
    "dosage": ("dosage_conflict", "missing_candidate_dosage"),
    "volume": ("volume_conflict", "missing_candidate_volume"),
    "weight": ("weight_conflict", "missing_candidate_weight"),
    "dosage_volume": ("dosage_volume_conflict", "missing_candidate_dosage_volume"),
    "strength_signature": ("strength_signature_conflict", "missing_candidate_strength_signature"),
    "concentration": ("concentration_conflict", "missing_candidate_concentration"),
    "percent_strength": ("percent_strength_conflict", "missing_candidate_percent_strength"),
    "iu_dosage": ("iu_dosage_conflict", "missing_candidate_iu_dosage"),
    "form": ("form_conflict", "missing_candidate_form"),
    "dimensions": ("dimensions_conflict", "missing_candidate_dimensions"),
    "critical_tokens": ("critical_tokens_conflict", "missing_candidate_critical_tokens"),
}
FORM_SYNONYMS.update(
    {
        "CAPS": "КАПС",
        "CAPSULE": "КАПС",
        "CAPSULES": "КАПС",

        "TAB": "ТАБ",
        "TABS": "ТАБ",
        "TABL": "ТАБ",
        "TABLET": "ТАБ",
        "TABLETS": "ТАБ",

        "SUPP": "СУПП",
        "SUPPOSITORY": "СУПП",
        "SUPPOSITORIES": "СУПП",
        "RECT": "СУПП",
        "RECTAL": "СУПП",

        "SPRAY": "СПРЕЙ",
        "NASAL": "СПРЕЙ",
        "КАПС": "КАПС",
        "КАПСУЛА": "КАПС",
        "КАПСУЛЫ": "КАПС",
        "ТАБ": "ТАБ",
        "ТАБЛ": "ТАБ",
        "ТАБЛЕТКА": "ТАБ",
        "ТАБЛЕТКИ": "ТАБ",
        "АМП": "АМП",
        "АМПУЛА": "АМП",
        "АМПУЛЫ": "АМП",
        "СПРЕЙ": "СПРЕЙ",
        "КРЕМ": "КРЕМ",
        "МАЗЬ": "МАЗЬ",
        "ГЕЛЬ": "ГЕЛЬ",
    }
)
MAX_CANDIDATE_POOL = 100
NAME_PREFILTER_THRESHOLD = 75
NAME_PREFILTER_TOP_LIMIT = 50
PARTIAL_RATIO_POOL_LIMIT = 100
STOP_WORDS = {
    "ПРОИЗВОДИТЕЛЬ",
    "СТРАНА",
    "ШТ",
    "УП",
    "УПАК",
    "ЛЕКАРСТВО",
    "Д",
    "ДЛЯ",
    "ПРИГ",
    "ПРИГОТ",
    "ПРИГОТОВЛЕНИЯ",
    "РА",
    "РАСТВОРА",
    "П",
    "О",
    "НАЗ",
    "НАЗАЛЬНЫЙ",
    "НАЗАЛЬНЫЕ",
    "ИН",
    "ИНЪЕКЦИЙ",
    "ИНЪЕКЦИИ",
    "ИНЪЕКЦИЯ",
    "ШИП",
    "ШИПУЧ",
    "ШИПУЧИЕ",
}
TOKEN_ALIASES = {
    "НЕО": {"NEO", "НEO"},
    "NEO": {"НЕО", "НEO"},
    "ЭКСТРА": {"EXTRA"},
    "EXTRA": {"ЭКСТРА"},
    "ЛОНГ": {"LONG", "ПРОЛОНГ"},
    "LONG": {"ЛОНГ", "ПРОЛОНГ"},
    "ПРОЛОНГ": {"ЛОНГ", "LONG"},
    "ФОРТЕ": {"FORTE"},
    "FORTE": {"ФОРТЕ"},
}
MANUFACTURER_STOP_WORDS = {
    "АО",
    "ООО",
    "ТОО",
    "ИП",
    "ЗАО",
    "ОАО",
    "LTD",
    "LLC",
    "GMBH",
    "D",
    "DD",
    "CO",
    "KG",
    "INC",
    "PLC",
    "PVT",
    "PRIVATE",
    "CORP",
    "COMPANY",
    "PHARMACEUTICALS",
    "PHARMA",
    "PHARMACEUTICAL",
    "KAZAKHSTAN",
    "КАЗАХСТАН",
    "RUSSIA",
    "РОССИЯ",
    "INDIA",
    "ИНДИЯ",
    "POLAND",
    "ПОЛЬША",
    "GERMANY",
    "ГЕРМАНИЯ",
    "TURKEY",
    "ТУРЦИЯ",
    "CHINA",
    "КИТАЙ",
}
MANUFACTURER_MARKER_SYNONYMS = {
    "ХФЗ": "ХИМФАРМ",
    "ХФ": "ХИМФАРМ",
    "CHIMPHARM": "ХИМФАРМ",
    "SANTO": "ХИМФАРМ",
    "BOEHRINGER": "БЕРИНГЕР",
    "INGELHEIM": "ИНГЕЛЬХАЙМ",
    "JADRAN": "ЯДРАН",
    "GALENSKI": "ГАЛЕНСКИ",
    "LABORATOIRES": "ЛАБОРАТОРИЯ",
    "LABORATORIES": "ЛАБОРАТОРИЯ",
    "ROSA": "РОЗА",
    "PHYTOPHARMA": "ФИТОФАРМА",
    "LEK": "ЛЕК",
}
LATIN_TO_CYRILLIC = str.maketrans(
    {
        "A": "А",
        "B": "В",
        "C": "С",
        "E": "Е",
        "H": "Н",
        "K": "К",
        "M": "М",
        "O": "О",
        "P": "Р",
        "T": "Т",
        "X": "Х",
        "Y": "У",
    }
)


def normalize_expiry(value: object) -> str:
    if value in (None, "", "0001-01-01T00:00:00"):
        return ""
    s = str(value).strip()
    if not s:
        return ""
    try:
        return datetime.fromisoformat(s.replace("Z", "+00:00")).date().isoformat()
    except Exception:
        return s[:10] if re.match(r"^\d{4}-\d{2}-\d{2}", s) else ""


def provisor_match_key(goods_id: object, distributor_goods_id: object, expiry_date: object) -> str:
    goods = normalize_external_sku(goods_id)
    distributor = normalize_external_sku(distributor_goods_id)
    exp = normalize_expiry(expiry_date) or "no_exp"
    return f"{goods}_{distributor}_{exp}" if goods and distributor else ""


def provisor_distributor_expiry_key(distributor_goods_id: object, expiry_date: object) -> str:
    distributor = normalize_external_sku(distributor_goods_id)
    exp = normalize_expiry(expiry_date) or "no_exp"
    return f"{distributor}_{exp}" if distributor else ""


def _to_int(value: object) -> int | None:
    if value in (None, ""):
        return None
    try:
        number = int(float(str(value).strip()))
    except Exception:
        return None
    return number if number > 0 else None


def _provisor_key(kind: str, value: object) -> str:
    return f"provisor:{kind}:{normalize_external_sku(value)}"


def _provisor_reference_sku_keys(value: object) -> list[tuple[str, str]]:
    out: list[tuple[str, str]] = []
    for key in _product_sku_keys(value):
        out.append(("reference_sku", _provisor_key("reference_sku", key)))
    return out


def _provisor_product_goods_keys(goods_id: object) -> list[tuple[str, str]]:
    goods = normalize_external_sku(goods_id)
    return [("goods_id", _provisor_key("goods_id", goods))] if goods else []


def _provisor_supplier_index_keys(
    *,
    goods_id: object,
    distributor_goods_id: object,
    expiry_date: object,
    filial_id: object,
) -> list[tuple[str, str]]:
    out: list[tuple[str, str]] = []
    goods = normalize_external_sku(goods_id)
    distributor = normalize_external_sku(distributor_goods_id)
    if goods:
        out.append(("goods_id", _provisor_key("goods_id", goods)))
    if _to_int(filial_id) in PROVISOR_REFERENCE_FILIAL_IDS:
        for raw_kind, raw_key in provisor_item_variants(goods_id, distributor_goods_id, expiry_date):
            if raw_kind in {"distributor_goods_id", "distributor_goods_id_normalized"} and raw_key:
                out.append(("reference_sku", _provisor_key("reference_sku", raw_key)))
    if goods and distributor:
        out.append(("goods_distributor_composite", _provisor_key("goods_distributor_composite", f"{goods}_{distributor}")))
    if goods and expiry_date:
        out.append(("goods_id_shelf_life", _provisor_key("goods_id_shelf_life", f"{goods}_{normalize_expiry(expiry_date) or 'no_exp'}")))
    return out


def _provisor_product_code_keys(
    *,
    goods_id: object,
    distributor_goods_id: object,
    expiry_date: object,
    filial_id: object,
) -> list[tuple[str, str]]:
    out: list[tuple[str, str]] = []
    goods = normalize_external_sku(goods_id)
    if goods:
        out.append(("goods_id", _provisor_key("goods_id", goods)))
    if _to_int(filial_id) in PROVISOR_REFERENCE_FILIAL_IDS:
        distributor = normalize_external_sku(distributor_goods_id)
        add_seen: set[str] = set()
        for raw_kind, raw_key in provisor_item_variants(goods_id, distributor, expiry_date):
            if raw_kind in {"distributor_goods_id", "distributor_goods_id_normalized"} and raw_key and raw_key not in add_seen:
                out.append(("reference_sku", raw_key))
                add_seen.add(raw_key)
    return out


def provisor_item_variants(goods_id: object, distributor_goods_id: object, expiry_date: object) -> list[tuple[str, str]]:
    variants: list[tuple[str, str]] = []

    def add(kind: str, value: str | None) -> None:
        if value and all(existing != value for _, existing in variants):
            variants.append((kind, value))

    distributor = normalize_external_sku(distributor_goods_id)
    goods = normalize_external_sku(goods_id)
    add("distributor_goods_id", distributor)
    for variant in normalize_sku_variants(distributor_goods_id):
        add("distributor_goods_id_normalized", variant)
    add("goods_id", goods)
    for variant in normalize_sku_variants(goods_id):
        add("goods_id_normalized", variant)
    add("goods_distributor_composite", f"{goods}_{distributor}" if goods and distributor else "")
    add("goods_distributor_shelf_life", provisor_match_key(goods_id, distributor_goods_id, expiry_date))
    add("distributor_goods_id_shelf_life", provisor_distributor_expiry_key(distributor_goods_id, expiry_date))
    return variants


def normalize_text(value: object) -> str:
    s = str(value or "").lower().replace("ё", "е")
    s = re.sub(r"[^0-9a-zа-я]+", " ", s, flags=re.I)
    return re.sub(r"\s+", " ", s).strip()


def normalize_text(value: object) -> str:
    s = str(value or "").lower().replace("\u0451", "\u0435")
    s = re.sub(r"[^0-9a-z\u0430-\u044f]+", " ", s, flags=re.I)
    return re.sub(r"\s+", " ", s).strip()


def normalize_manufacturer_text(value: object) -> str:
    canonical = normalize_manufacturer(value)
    if canonical:
        return canonical
    s = str(value or "").upper().replace("\u0401", "\u0415")
    s = re.sub(r"[\"'`«»„“”./,&()\\[\\]{}]+", " ", s)
    s = re.sub(r"[‐‑‒–—−-]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    words: list[str] = []
    for raw_word in s.split():
        word = MANUFACTURER_MARKER_SYNONYMS.get(raw_word, raw_word)
        if word in MANUFACTURER_STOP_WORDS:
            continue
        if len(word) <= 1 and not word.isdigit():
            continue
        words.append(word)
    s = " ".join(words)
    return normalize_text(s)


def normalized_manufacturer_tokens(value: object) -> frozenset[str]:
    return frozenset(token for token in normalize_manufacturer_text(value).split() if len(token) >= 3)


def _manufacturer_match(left: str, right: str) -> bool:
    left = normalize_manufacturer_text(left)
    right = normalize_manufacturer_text(right)
    if not left or not right:
        return False
    if left == right:
        return True
    if left.replace(" ", "") == right.replace(" ", ""):
        return True
    left_tokens = set(left.split())
    right_tokens = set(right.split())
    if not left_tokens or not right_tokens:
        return False
    if min(len(left_tokens), len(right_tokens)) >= 2:
        return left_tokens.issubset(right_tokens) or right_tokens.issubset(left_tokens)
    common = left_tokens & right_tokens
    if common and len(common) / max(len(left_tokens), len(right_tokens)) >= 0.67:
        return True
    return False


MANUFACTURER_HARD_CONFLICTS = {
    frozenset(("LEK", "SANDOZ")),
    frozenset(("BAYER", "SANDOZ")),
    frozenset(("NOBEL", "HEMOFARM")),
    frozenset(("NOBEL", "HIMFARM")),
    frozenset(("TAKEDA", "SANDOZ")),
}


def _manufacturer_similarity(left: str, right: str) -> float:
    left = normalize_manufacturer_text(left)
    right = normalize_manufacturer_text(right)
    if not left or not right:
        return 0.0
    return _score(left, right)


def _manufacturer_decision(left: str, right: str) -> tuple[str, float, bool]:
    left_norm = normalize_manufacturer_text(left)
    right_norm = normalize_manufacturer_text(right)
    if not left_norm or not right_norm:
        return "missing", 0.0, False
    if _manufacturer_match(left_norm, right_norm):
        return "alias_match", 100.0, False
    similarity = _score(left_norm, right_norm)
    if frozenset((left_norm, right_norm)) in MANUFACTURER_HARD_CONFLICTS:
        return "hard_reject", similarity, True
    if similarity >= 85:
        return "similar_match", similarity, False
    return "penalty", similarity, False


def _manufacturer_required_match(left: str, right: str) -> tuple[str, float, bool]:
    left_norm = normalize_manufacturer_text(left)
    right_norm = normalize_manufacturer_text(right)
    if not left_norm or not right_norm:
        return "missing", 0.0, False
    if frozenset((left_norm, right_norm)) in MANUFACTURER_HARD_CONFLICTS:
        return "hard_reject", _score(left_norm, right_norm), False
    if _manufacturer_match(left_norm, right_norm):
        return "alias_match", 100.0, True
    similarity = _score(left_norm, right_norm)
    if similarity >= 85:
        return "similar_match", similarity, True
    return "manufacturer_reject", similarity, False


def _structures_match_strongly(left: DrugStructure, right: DrugStructure) -> bool:
    if _structure_mismatch(left, right) is not None:
        return False
    matched, missing = _param_counts(left, right)
    return matched >= 2 and missing == 0


def normalize_drug_text(value: object) -> str:
    s = str(value or "").upper().replace("\u0401", "\u0415")
    s = re.sub(r"\bNEO\b", "НЕО", s)
    s = re.sub(r"\bEXTRA\b", "ЭКСТРА", s)
    s = re.sub(r"\bFORTE\b", "ФОРТЕ", s)
    s = re.sub(r"\bLONG\b", "ЛОНГ", s)
    s = re.sub(r"\bCAPS(?:ULES?)?\b", "\u041a\u0410\u041f\u0421", s)
    s = re.sub(r"\bTAB(?:L|LETS?)?\b", "\u0422\u0410\u0411", s)
    s = re.sub(r"\bSUPP(?:OSITOR(?:Y|IES))?\b", "\u0421\u0423\u041f\u041f", s)
    s = re.sub(r"\bRECT(?:AL)?\b", "\u0421\u0423\u041f\u041f", s)
    s = re.sub(r"\bSPRAY\s+NAS(?:AL)?\b", "\u0421\u041f\u0420\u0415\u0419", s)
    s = re.sub(r"\bSPRAY\b", "\u0421\u041f\u0420\u0415\u0419", s)
    s = re.sub(r"(\d+)\s*ML\b", r"\1МЛ", s)
    s = re.sub(r"(\d+)\s*MG\b", r"\1МГ", s)
    s = re.sub(r"(\d+)\s*G\b", r"\1Г", s)
    s = re.sub(r"\bML\b", "МЛ", s)
    s = re.sub(r"\bMG\b", "МГ", s)
    s = re.sub(r"\bG\b", "Г", s)
    s = re.sub(r"\bME\b", "МЕ", s)
    s = re.sub(r"\bIU\b", "МЕ", s)
    s = re.sub(r"\bП\s*/\s*О\b", " ", s)
    s = re.sub(r"\bД\s*/\s*ПРИГ\b", " ", s)
    s = re.sub(r"\bД\s*/\s*ИН\b", " ", s)
    s = re.sub(r"\bД\s*/\s*ИНЪЕКЦ(?:ИЙ|ИИ|ИЯ)?\b", " ", s)
    s = re.sub(r"\bР\s*-\s*РА\b", " ", s)
    s = re.sub(r"\bР\s*/\s*РА\b", " ", s)
    s = re.sub(r"\bСПРЕЙ\s+НАЗ\b", "СПРЕЙ", s)
    s = re.sub(r"\bСПРЕЙ\s+НАЗАЛЬН(?:ЫЙ|ЫЕ)\b", "СПРЕЙ", s)
    s = re.sub(r"[\"'`«»„“”]+", " ", s)
    s = re.sub(r"[‐‑‒–—−-]+", " ", s)
    s = re.sub(r"[/\\,.;:()\\[\\]{}]+", " ", s)
    s = re.sub(r"№\s*(\d+)", r"N\1", s)
    s = re.sub(r"\bN\s*(\d+)\b", r"N\1", s)
    s = re.sub(r"(\d+(?:[,.]\d+)?)\s*(МЛ|МГ|МКГ|Г|МЕ)\b", r"\1\2", s)
    s = s.translate(LATIN_TO_CYRILLIC)
    s = re.sub(r"[^0-9A-ZА-ЯЁ\s]+", " ", s)
    words = []
    for word in re.sub(r"\s+", " ", s).strip().split():
        if word in {"N"}:
            words.append(word)
        elif re.fullmatch(r"N\d+", word) or re.fullmatch(r"\d+(МЛ|МГ|МКГ|Г|МЕ)", word):
            words.append(word)
        elif re.search(r"[А-ЯЁ]", word):
            words.append(word)
        else:
            words.append(word.translate(LATIN_TO_CYRILLIC))
    s = " ".join(words)
    return re.sub(r"\s+", " ", s).strip()


def _num(value: str) -> float:
    return float(value.replace(",", "."))


def _clean_num(value: float | None) -> float | None:
    if value is None:
        return None
    rounded = round(float(value), 4)
    return int(rounded) if rounded.is_integer() else rounded


def _unitless_decimal_to_mg(value: float) -> float:
    # In supplier names decimals like 0,1 and 0,004 are normally grams.
    return value * 1000 if 0 < value < 1 else value


def _normalize_dosage_mg(value: float | None) -> float | None:
    if value is None:
        return None
    return value * 1000 if 0 < float(value) < 1 else float(value)


def _clean_tuple(values: list[float]) -> tuple[float, ...] | None:
    cleaned = tuple(float(_clean_num(value)) for value in values)
    return cleaned or None


def _dimension_value(value: str, unit: str) -> str:
    number = _clean_num(_num(value))
    return f"{number}{unit.upper()}"


@dataclass(frozen=True)
class DrugStructure:
    base_name: str
    quantity: int | None = None
    volume: float | None = None
    weight: float | None = None
    dosage: float | None = None
    raw_dosage: float | None = None
    dosage_volume: float | None = None
    strength_signature: tuple[float, ...] | None = None
    concentration: float | None = None
    percent_strength: float | None = None
    iu_dosage: float | None = None
    form: str | None = None
    forms: tuple[str, ...] | None = None
    dimensions: tuple[str, ...] | None = None
    critical_tokens: tuple[str, ...] | None = None


def parse_drug_structure(value: object) -> DrugStructure:
    raw = str(value or "").upper().replace("\u0401", "\u0415")
    raw = re.sub(r"\bCAPS(?:ULES?)?\b", "\u041a\u0410\u041f\u0421", raw)
    raw = re.sub(r"\bTAB(?:L|LETS?)?\b", "\u0422\u0410\u0411", raw)
    raw = re.sub(r"\bSUPP(?:OSITOR(?:Y|IES))?\b", "\u0421\u0423\u041f\u041f", raw)
    raw = re.sub(r"\bRECT(?:AL)?\b", "\u0421\u0423\u041f\u041f", raw)
    raw = re.sub(r"\bSPRAY\s+NAS(?:AL)?\b", "\u0421\u041f\u0420\u0415\u0419", raw)
    raw = re.sub(r"\bSPRAY\b", "\u0421\u041f\u0420\u0415\u0419", raw)
    raw = re.sub(r"(\d+(?:[,.]\d+)?)\s*ML\b", r"\1МЛ", raw)
    raw = re.sub(r"(\d+(?:[,.]\d+)?)\s*MG\b", r"\1МГ", raw)
    raw = re.sub(r"(\d+(?:[,.]\d+)?)\s*G\b", r"\1Г", raw)
    raw = re.sub(r"\bML\b", "МЛ", raw)
    raw = re.sub(r"\bMG\b", "МГ", raw)
    raw = re.sub(r"\bG\b", "Г", raw)
    raw = re.sub(r"\bME\b", "МЕ", raw)
    raw = re.sub(r"\bIU\b", "МЕ", raw)
    raw = raw.replace("№", "N")
    normalized = normalize_drug_text(value)

    quantity_m = re.search(r"\bN\s*(\d+)\b", normalized)
    dimensions: tuple[str, ...] | None = None
    dimension_pair_m = re.search(r"\b(\d+(?:[\.,]\d+)?)\s*(?:СМ|CM)\s*[XХ]\s*(\d+(?:[\.,]\d+)?)\s*(?:М|M)\b", raw)
    if dimension_pair_m:
        dimensions = (_dimension_value(dimension_pair_m.group(1), "СМ"), _dimension_value(dimension_pair_m.group(2), "М"))
    else:
        mm_m = re.search(r"\b(\d+(?:[\.,]\d+)?)\s*MM\b", raw)
        if mm_m:
            dimensions = (_dimension_value(mm_m.group(1), "MM"),)
    percent_strength: float | None = None
    percent_m = re.search(r"\b(\d+(?:[\.,]\d+)?)\s*%", raw)
    if percent_m:
        percent_strength = _num(percent_m.group(1))

    dosage: float | None = None
    raw_dosage: float | None = None
    concentration: float | None = None
    dosage_volume: float | None = None
    dose_per_volume_m = re.search(r"\b(\d+(?:[\.,]\d+)?)\s*(?:МГ)?\s*/\s*(\d+(?:[\.,]\d+)?)\s*МЛ\b", raw)
    if dose_per_volume_m:
        raw_dose_per_volume = _num(dose_per_volume_m.group(1))
        dosage = _normalize_dosage_mg(raw_dose_per_volume)
        raw_dosage = raw_dose_per_volume
        dosage_volume = _num(dose_per_volume_m.group(2))
    conc_m = re.search(r"\b(\d+(?:[\.,]\d+)?)\s*МГ\s*/\s*МЛ\b", raw)
    if conc_m:
        concentration = _num(conc_m.group(1))
    else:
        conc_m = re.search(r"\b(\d+(?:[\.,]\d+)?)\s*/\s*МЛ\b", raw)
        if conc_m:
            concentration = _unitless_decimal_to_mg(_num(conc_m.group(1)))

    dosage_m = re.search(r"\b(\d+(?:[\.,]\d+)?)\s*МГ\b", raw)
    gram_m = re.search(r"\b(\d+(?:[\.,]\d+)?)\s*Г\b", raw)
    decimal_m = re.search(r"(?<!N)\b0[\.,]\d+\b(?!\s*(?:%|/|МЛ|Г))", raw)
    words = normalized.split()
    has_weight_context = bool(set(words) & WEIGHT_FORMS)
    has_medicinal_gram_context = bool(set(words) & MEDICINAL_GRAM_DOSAGE_TOKENS)
    gram_is_weight = bool(gram_m and has_weight_context and not has_medicinal_gram_context)
    if dosage_m:
        raw_dosage = _num(dosage_m.group(1))
        dosage = _normalize_dosage_mg(raw_dosage)
    elif gram_m and not gram_is_weight:
        raw_dosage = _num(gram_m.group(1))
        dosage = raw_dosage * 1000
    elif decimal_m:
        raw_dosage = _num(decimal_m.group(0))
        dosage = _normalize_dosage_mg(raw_dosage)
    strength_signature: tuple[float, ...] | None = None
    strength_m = re.search(r"\b\d+(?:[\.,]\d+)?(?:\s*МГ)?(?:\s*/\s*\d+(?:[\.,]\d+)?(?:\s*МГ)?){2,}", raw)
    if strength_m and "МЛ" not in strength_m.group(0):
        raw_strengths = re.findall(r"\d+(?:[\.,]\d+)?", strength_m.group(0))
        strengths = [_normalize_dosage_mg(_num(value)) for value in raw_strengths]
        strength_signature = _clean_tuple([float(value) for value in strengths if value is not None])
        if strength_signature:
            dosage = strength_signature[0]
            raw_dosage = _num(raw_strengths[0])
    iu_dosage: float | None = None
    iu_m = re.search(r"\b(\d+(?:[\.,]\d+)?)\s*МЕ\b", raw)
    if iu_m:
        iu_dosage = _num(iu_m.group(1))

    forms_seen: list[str] = []
    critical_tokens = tuple(sorted(set(words) & CRITICAL_NAME_TOKENS)) or None
    for idx, word in enumerate(words):
        clean = word.strip("-")
        if clean in FORM_SYNONYMS:
            canonical_form = FORM_SYNONYMS[clean]
            if canonical_form not in forms_seen:
                forms_seen.append(canonical_form)
            continue
        if clean == "Р" and idx + 1 < len(words) and words[idx + 1].strip("-") == "Р":
            if "Р-Р" not in forms_seen:
                forms_seen.append("Р-Р")
            continue
    form: str | None = forms_seen[0] if forms_seen else None
    forms = tuple(forms_seen) or None

    volume: float | None = None
    dose_span = dose_per_volume_m.span() if dose_per_volume_m else None
    for candidate_volume_m in re.finditer(r"\b(\d+(?:[\.,]\d+)?)\s*МЛ\b", raw):
        if dose_span and candidate_volume_m.start() >= dose_span[0] and candidate_volume_m.end() <= dose_span[1]:
            continue
        volume = _num(candidate_volume_m.group(1))
    if volume is None and forms and any(x in LIQUID_VOLUME_FORMS for x in forms):
        raw_form_pattern = r"СУСП|СИРОП|РАСТВОР|Р\s*-\s*Р|Р-Р|КАПЛИ|СПРЕЙ"
        implicit_volume_m = re.search(rf"\b(\d+(?:[\.,]\d+)?)\s*(?:{raw_form_pattern})\b", raw)
        if implicit_volume_m:
            volume = _num(implicit_volume_m.group(1))

    weight: float | None = None
    weight_m = re.search(r"\b(\d+(?:[\.,]\d+)?)\s*Г\b", raw)
    has_weight_context = ((form in WEIGHT_FORMS) or bool(set(words) & WEIGHT_FORMS)) and not has_medicinal_gram_context
    if weight_m and has_weight_context:
        weight = _num(weight_m.group(1))
    elif volume is None and has_weight_context:
        before_form = re.split(r"\b(?:КРЕМ|МАЗЬ|ГЕЛЬ|ФИТО|ЧАЙ|ТРАВА|СБОР)\b", raw, maxsplit=1)[0]
        weight_numbers = re.findall(r"\b(\d+(?:[\.,]\d+)?)\b(?!\s*(?:%|/|МЛ|МГ|МЕ|Г))", before_form)
        if weight_numbers:
            weight = _num(weight_numbers[-1])
    if weight is not None and has_weight_context and gram_m:
        dosage = None
        raw_dosage = None

    base = normalized
    base = re.sub(r"\bN\s*\d+\b", " ", base)
    base = re.sub(r"\b\d+(?:[\.,]\d+)?\s*МЛ\b", " ", base)
    base = re.sub(r"\b\d+(?:[\.,]\d+)?\s*МГ\b", " ", base)
    base = re.sub(r"\b\d+(?:[\.,]\d+)?\s*МЕ\b", " ", base)
    base = re.sub(r"\b\d+(?:[\.,]\d+)?\s*Г\b", " ", base)
    base = re.sub(r"\b\d+(?:[\.,]\d+)?\s*%", " ", base)
    if percent_m:
        base = re.sub(rf"\b{re.escape(percent_m.group(1).replace(',', ' '))}\b", " ", base)
        base = re.sub(rf"\b{re.escape(percent_m.group(1).replace('.', ' '))}\b", " ", base)
    base = re.sub(r"\b0[\.,]\d+\b", " ", base)
    base = re.sub(r"\b0\s+\d+\b", " ", base)
    parsed_numeric_values = [
        volume,
        weight,
        dosage_volume,
        percent_strength,
    ]
    for parsed_value in parsed_numeric_values:
        if parsed_value is None:
            continue
        cleaned_value = _clean_num(parsed_value)
        if cleaned_value is None:
            continue
        value_text = str(cleaned_value)
        if "." in value_text:
            left, right = value_text.split(".", 1)
            base = re.sub(rf"\b{re.escape(left)}\s+{re.escape(right.rstrip('0') or '0')}\b", " ", base)
        if float(parsed_value).is_integer():
            integer_text = str(int(float(parsed_value)))
            base = re.sub(rf"\b{re.escape(integer_text)}\s+0+\b", " ", base)
            base = re.sub(rf"\b{re.escape(integer_text)}\b", " ", base)
    for raw_form in sorted(FORM_SYNONYMS, key=len, reverse=True):
        base = re.sub(rf"\b{re.escape(raw_form)}\b", " ", base)
    if form == "Р-Р":
        base = re.sub(r"\bР\s+Р\b", " ", base)
    if weight is not None:
        weight_text = str(_clean_num(weight)).replace(".", r"[\s\.,]")
        base = re.sub(rf"\b{weight_text}(?:\s+0)?\b", " ", base)
    base = " ".join(
        word
        for word in base.split()
        if word not in STOP_WORDS and word not in {"МЛ", "МГ", "МКГ", "Г", "МЕ"} and not re.fullmatch(r"0+", word)
    )
    base = re.sub(r"\s+", " ", base).strip()

    structure = DrugStructure(
        base_name=base,
        quantity=int(quantity_m.group(1)) if quantity_m else None,
        volume=_clean_num(volume),
        weight=_clean_num(weight),
        dosage=_clean_num(dosage),
        raw_dosage=_clean_num(raw_dosage),
        dosage_volume=_clean_num(dosage_volume),
        strength_signature=strength_signature,
        concentration=_clean_num(concentration),
        percent_strength=_clean_num(percent_strength),
        iu_dosage=_clean_num(iu_dosage),
        form=form,
        forms=forms,
        dimensions=dimensions,
        critical_tokens=critical_tokens,
    )
    _verbose_debug(
        "[STRUCT_PARSE] raw=%s base_name=%s dosage=%s dosage_volume=%s concentration=%s percent_strength=%s volume=%s weight=%s quantity=%s form=%s",
        value,
        structure.base_name,
        structure.dosage,
        structure.dosage_volume,
        structure.concentration,
        structure.percent_strength,
        structure.volume,
        structure.weight,
        structure.quantity,
        structure.form,
    )
    return structure


def _structure_mismatch(left: DrugStructure, right: DrugStructure) -> tuple[str, object, object] | None:
    for field in STRUCTURED_FIELDS:
        left_value = getattr(left, field)
        right_value = getattr(right, field)
        if field == "form":
            left_forms = set(left.forms or ((left.form,) if left.form else ()))
            right_forms = set(right.forms or ((right.form,) if right.form else ()))
            if left_forms and right_forms and left_forms.isdisjoint(right_forms):
                return field, tuple(sorted(left_forms)), tuple(sorted(right_forms))
            continue
        if left_value is None or right_value is None:
            continue
        if isinstance(left_value, (int, float)) and isinstance(right_value, (int, float)):
            if abs(float(left_value) - float(right_value)) > 0.001:
                return field, left_value, right_value
            continue
        if left_value != right_value:
            return field, left_value, right_value
    return None


def _strict_structure_decision(product_struct: DrugStructure, candidate_struct: DrugStructure) -> tuple[str, str | None]:
    first_missing_reason: str | None = None
    for field in STRUCTURED_FIELDS:
        product_value = getattr(product_struct, field)
        candidate_value = getattr(candidate_struct, field)
        conflict_reason, missing_reason = STRICT_STRUCTURE_REASONS.get(field, (f"{field}_conflict", f"missing_candidate_{field}"))
        if field == "form":
            product_forms = set(product_struct.forms or ((product_struct.form,) if product_struct.form else ()))
            candidate_forms = set(candidate_struct.forms or ((candidate_struct.form,) if candidate_struct.form else ()))
            if product_forms and candidate_forms:
                if product_forms.isdisjoint(candidate_forms):
                    return "reject", conflict_reason
                continue
            if product_forms and not candidate_forms and field in STRICT_MISSING_CANDIDATE_FIELDS:
                first_missing_reason = first_missing_reason or missing_reason
            continue
        if product_value is not None and candidate_value is not None:
            if isinstance(product_value, (int, float)) and isinstance(candidate_value, (int, float)):
                if abs(float(product_value) - float(candidate_value)) >= 0.001:
                    return "reject", conflict_reason
                continue
            if product_value != candidate_value:
                return "reject", conflict_reason
        elif product_value is not None and candidate_value is None and field in STRICT_MISSING_CANDIDATE_FIELDS:
            first_missing_reason = first_missing_reason or missing_reason
    if first_missing_reason:
        return "suspicious", first_missing_reason
    return "ok", None


def _structure_reason_field(reason: str | None) -> str | None:
    if not reason:
        return None
    if reason.startswith("missing_candidate_"):
        return reason.replace("missing_candidate_", "", 1)
    if reason.endswith("_conflict"):
        field = reason[: -len("_conflict")]
        if field in {
            "dosage",
            "dosage_volume",
            "strength_signature",
            "quantity",
            "volume",
            "weight",
            "concentration",
            "percent_strength",
            "iu_dosage",
            "form",
            "dimensions",
            "critical_tokens",
        }:
            return field
    return None


def _structure_values_for_reason(
    product_struct: DrugStructure,
    candidate_struct: DrugStructure,
    reason: str | None,
) -> tuple[object, object]:
    field = _structure_reason_field(reason)
    if not field:
        return None, None
    if field == "form":
        product_forms = product_struct.forms or ((product_struct.form,) if product_struct.form else None)
        candidate_forms = candidate_struct.forms or ((candidate_struct.form,) if candidate_struct.form else None)
        return product_forms, candidate_forms
    return getattr(product_struct, field, None), getattr(candidate_struct, field, None)


def _base_name_similarity(left: str, right: str) -> float:
    left = normalize_text(left)
    right = normalize_text(right)
    if not left or not right:
        return 0.0
    if left == right:
        return 100.0
    return max(_score(left, right), _token_sort_ratio(left, right), _token_set_ratio(left, right))


def _base_name_nearly_equal(left: str, right: str) -> bool:
    return _base_name_similarity(left, right) >= 97


def _first_token(value: str) -> str:
    return (normalize_text(value).split(" ", 1)[0] if normalize_text(value) else "")


def _conflict_reason(field: str) -> str:
    return {
        "quantity": "quantity_conflict",
        "volume": "volume_conflict",
        "weight": "weight_conflict",
        "dosage": "dosage_conflict",
        "dosage_volume": "dosage_volume_conflict",
        "strength_signature": "strength_signature_conflict",
        "concentration": "concentration_conflict",
        "percent_strength": "percent_strength_conflict",
        "iu_dosage": "iu_dosage_conflict",
        "form": "form_conflict",
        "dimensions": "dimensions_conflict",
        "critical_tokens": "critical_tokens_conflict",
    }.get(field, f"{field}_conflict")


def _score(a: str, b: str) -> float:
    if not a or not b:
        return 0.0
    return SequenceMatcher(None, a, b).ratio() * 100


def _token_sort_ratio(a: str, b: str) -> float:
    return _score(" ".join(sorted(a.split())), " ".join(sorted(b.split())))


def _token_set_ratio(a: str, b: str) -> float:
    a_tokens = set(a.split())
    b_tokens = set(b.split())
    if not a_tokens or not b_tokens:
        return 0.0
    common = " ".join(sorted(a_tokens & b_tokens))
    left = " ".join(sorted(a_tokens))
    right = " ".join(sorted(b_tokens))
    return max(_score(common, left), _score(common, right), _score(left, right))


def _partial_ratio(a: str, b: str) -> float:
    if not a or not b:
        return 0.0
    if rapidfuzz_fuzz is not None:
        return float(rapidfuzz_fuzz.partial_ratio(a, b))
    shorter, longer = (a, b) if len(a) <= len(b) else (b, a)
    if shorter in longer:
        return 100.0
    window = len(shorter)
    best = 0.0
    for i in range(0, max(1, len(longer) - window + 1)):
        best = max(best, _score(shorter, longer[i : i + window]))
    return best


def _similarity(a: str, b: str, *, include_partial: bool = True) -> dict[str, float]:
    scores = {
        "ratio": _score(a, b),
        "tokenSortRatio": _token_sort_ratio(a, b),
        "tokenSetRatio": _token_set_ratio(a, b),
    }
    if include_partial:
        scores["partialRatio"] = _partial_ratio(a, b)
    else:
        scores["partialRatio"] = 0.0
    scores["score"] = max(scores.values())
    return {k: round(v, 2) for k, v in scores.items()}


def _param_counts(left: DrugStructure, right: DrugStructure) -> tuple[int, int]:
    matched = 0
    missing = 0
    for field in STRUCTURED_FIELDS:
        left_value = getattr(left, field)
        right_value = getattr(right, field)
        if field == "form":
            left_forms = set(left.forms or ((left.form,) if left.form else ()))
            right_forms = set(right.forms or ((right.form,) if right.form else ()))
            if left_forms and right_forms:
                if not left_forms.isdisjoint(right_forms):
                    matched += 1
            elif bool(left_forms) != bool(right_forms):
                missing += 1
            continue
        if left_value is not None and right_value is not None:
            if isinstance(left_value, (int, float)) and isinstance(right_value, (int, float)):
                if abs(float(left_value) - float(right_value)) <= 0.001:
                    matched += 1
            elif left_value == right_value:
                matched += 1
        elif (left_value is None) != (right_value is None):
            missing += 1
    return matched, missing


def _matching_threshold(left: DrugStructure, right: DrugStructure, matched_params: int, missing_params: int) -> int:
    if left.quantity is not None and right.quantity is not None and left.quantity == right.quantity and left.form and right.form and left.form == right.form:
        return 70
    if missing_params:
        return 75 if matched_params == 0 else 72
    return 72


def _bump_reason(stats: dict, reason: object) -> None:
    key = str(reason or "unknown")
    reasons = stats.setdefault("rejectReasons", {})
    reasons[key] = int(reasons.get(key) or 0) + 1


MATCH_REJECT_REASON_KEYS = (
    "manufacturer_mismatch",
    "manufacturer_hard_reject",
    "manufacturer_penalty",
    "dosage_mismatch",
    "quantity_mismatch",
    "volume_mismatch",
    "form_mismatch",
    "empty_candidate_pool",
)


def _empty_match_reject_counts() -> dict[str, int]:
    return {key: 0 for key in MATCH_REJECT_REASON_KEYS}


def _reject_reason_for_structure_field(field: str) -> str | None:
    if field == "quantity":
        return "quantity_mismatch"
    if field == "volume":
        return "volume_mismatch"
    if field in {"dosage", "concentration", "iu_dosage"}:
        return "dosage_mismatch"
    if field == "form":
        return "form_mismatch"
    return None


def _bump_match_reject(stats: dict, reason: str, amount: int = 1) -> None:
    if reason not in MATCH_REJECT_REASON_KEYS:
        return
    reasons = stats.setdefault("detailedRejectReasons", _empty_match_reject_counts())
    reasons[reason] = int(reasons.get(reason) or 0) + int(amount)


def _merge_match_reject_counts(stats: dict, counts: dict[str, int]) -> None:
    for reason, amount in counts.items():
        _bump_match_reject(stats, reason, int(amount or 0))


def _merge_match_safety_stats(stats: dict, safety: dict[str, int]) -> None:
    mapping = {
        "structure_rejects": "structureRejects",
        "structure_suspicious": "structureSuspicious",
        "selected_suspicious": "selectedSuspicious",
        "manufacturer_hard_rejects": "manufacturerHardRejectsSafety",
        "fuzzy_rejected_low_confidence": "fuzzyRejectedLowConfidence",
    }
    for source_key, stat_key in mapping.items():
        stats[stat_key] = int(stats.get(stat_key) or 0) + int(safety.get(source_key) or 0)


PROVISOR_MATCH_COUNTER_BY_KIND = {
    "reference_sku": "matchedByReferenceDistributorSku",
    "distributor_goods_id": "matchedByDistributorGoodsId",
    "distributor_goods_id_normalized": "matchedByDistributorGoodsIdNormalized",
    "goods_id": "matchedByGoodsId",
    "substitute_goods_id": "matchedByManualSubstitute",
    "goods_id_normalized": "matchedByGoodsIdNormalized",
    "goods_distributor_composite": "matchedByGoodsDistributorComposite",
    "goods_distributor_shelf_life": "matchedByGoodsDistributorShelfLife",
    "goods_id_shelf_life": "matchedByGoodsIdShelfLife",
    "distributor_goods_id_shelf_life": "matchedByDistributorGoodsIdShelfLife",
}


def _empty_provisor_exact_counts() -> dict[str, int]:
    return {counter: 0 for counter in PROVISOR_MATCH_COUNTER_BY_KIND.values()}


def _bump_provisor_exact(stats: dict, kind: str) -> None:
    counter = PROVISOR_MATCH_COUNTER_BY_KIND.get(kind)
    if not counter:
        return
    stats[counter] = int(stats.get(counter) or 0) + 1


def _vidman_candidate_result(
    *,
    name_norm: str,
    mfr_norm: str,
    expiry_norm: str = "",
    candidates: list["ProductCandidate"],
    exact: dict[tuple[str, str], "ProductCandidate"],
    token_index: dict[str, list["ProductCandidate"]],
) -> dict:
    mfr_norm = normalize_manufacturer_text(mfr_norm)
    source_full_norm = normalize_text(normalize_drug_text(name_norm))
    source_structure = parse_drug_structure(name_norm)
    source_base_norm = normalize_text(source_structure.base_name) or source_full_norm
    exact_candidate = exact.get((source_full_norm, mfr_norm))
    safety_stats = {
        "structure_rejects": 0,
        "structure_suspicious": 0,
        "selected_suspicious": 0,
        "manufacturer_hard_rejects": 0,
        "fuzzy_rejected_low_confidence": 0,
    }
    if exact_candidate is not None:
        structure_decision, structure_reason = _strict_structure_decision(source_structure, exact_candidate.structure)
        if structure_decision == "reject":
            product_value, candidate_value = _structure_values_for_reason(source_structure, exact_candidate.structure, structure_reason)
            safety_stats["structure_rejects"] += 1
            _verbose_debug(
                "[MATCH_STRUCTURE_REJECT] product_sku=%s product_name=%s candidate_id=%s candidate_name=%s reason=%s product_value=%s candidate_value=%s",
                "",
                name_norm,
                exact_candidate.id,
                exact_candidate.name_norm,
                structure_reason,
                product_value,
                candidate_value,
            )
        else:
            matched_params, missing_params = _param_counts(source_structure, exact_candidate.structure)
            if structure_decision == "suspicious":
                safety_stats["structure_suspicious"] += 1
                safety_stats["selected_suspicious"] += 1
                _verbose_debug(
                    "[MATCH_STRUCTURE_SUSPICIOUS] product_sku=%s candidate_id=%s reason=%s name_score=%s manufacturer_decision=%s",
                    "",
                    exact_candidate.id,
                    structure_reason,
                    100,
                    "alias_match",
                )
            return {
                "candidateSku": exact_candidate.sku,
                "candidateId": exact_candidate.id,
                "candidateNameNorm": exact_candidate.name_norm,
                "candidateBaseNameNorm": exact_candidate.base_name_norm,
                "candidateManufacturerNorm": exact_candidate.manufacturer_norm,
                "sourceStructure": source_structure.__dict__,
                "candidateStructure": exact_candidate.structure.__dict__,
                "matchedParams": matched_params,
                "missingParams": missing_params,
                "nameScore": 100,
                "manufacturerScore": 100,
                "manufacturerDecision": "alias_match",
                "structureDecision": structure_decision,
                "expiryScore": 100 if expiry_norm else None,
                "expiryPenalty": 0,
                "score": 100,
                "passesThreshold": True,
                "reason": "exact_normalized_name",
                "safetyStats": safety_stats,
            }

    candidate_pool: dict[int, ProductCandidate] = {}
    first_word = source_base_norm.split(" ", 1)[0]
    for key in (
        f"base:{source_base_norm}" if source_base_norm else None,
        f"first:{first_word}" if first_word else None,
        f"prefix:{source_base_norm[:5]}" if source_base_norm else None,
        f"prefix:{source_base_norm[:3]}" if source_base_norm else None,
        *_composite_index_keys(source_base_norm, source_structure),
    ):
        if key:
            for candidate in token_index.get(key, []):
                candidate_pool[candidate.id] = candidate
    for token in _tokens(source_base_norm):
        for candidate in token_index.get(token, []):
            candidate_pool[candidate.id] = candidate
    pool = list(candidate_pool.values())
    used_broad_fallback = False
    pool_size_before_limit = len(pool)
    include_partial = len(pool) <= PARTIAL_RATIO_POOL_LIMIT

    best: tuple[tuple[int, float, float], ProductCandidate, dict, float, int, int, int, str, str] | None = None
    best_rejected: tuple[float, float, float, ProductCandidate, str, object, object] | None = None
    filtered_counts = {field: 0 for field in STRUCTURED_FIELDS}
    filtered_examples: list[dict] = []
    after_name_filter = 0
    after_strict_filter = 0
    after_manufacturer_priority = 0
    for candidate in pool[:MAX_CANDIDATE_POOL]:
        structure_decision, structure_reason = _strict_structure_decision(source_structure, candidate.structure)
        if structure_decision == "reject":
            field = _structure_reason_field(structure_reason) or "structure"
            source_value, candidate_value = _structure_values_for_reason(source_structure, candidate.structure, structure_reason)
            filtered_counts[field] += 1
            safety_stats["structure_rejects"] += 1
            if len(filtered_examples) < 5:
                filtered_examples.append(
                    {
                        "field": field,
                        "sourceValue": source_value,
                        "candidateValue": candidate_value,
                        "candidateSku": candidate.sku,
                        "candidateNameNorm": candidate.name_norm,
                        "scoreBeforeFilter": None,
                    }
                )
            if best_rejected is None:
                best_rejected = (0.0, 0.0, 0.0, candidate, structure_reason or "structure_conflict", source_value, candidate_value)
            _verbose_debug(
                "[MATCH_STRUCTURE_REJECT] product_sku=%s product_name=%s candidate_id=%s candidate_name=%s reason=%s product_value=%s candidate_value=%s",
                "",
                name_norm,
                candidate.id,
                candidate.name_norm,
                structure_reason,
                source_value,
                candidate_value,
            )
            continue
        after_strict_filter += 1
        manufacturer_decision, mfr_score, manufacturer_allowed = _manufacturer_required_match(mfr_norm, candidate.manufacturer_norm)
        if structure_decision == "suspicious":
            safety_stats["structure_suspicious"] += 1
            _verbose_debug(
                "[MATCH_STRUCTURE_SUSPICIOUS] product_sku=%s candidate_id=%s reason=%s name_score=%s manufacturer_decision=%s",
                "",
                candidate.id,
                structure_reason,
                None,
                manufacturer_decision,
            )
        if not manufacturer_allowed:
            safety_stats["manufacturer_hard_rejects"] += 1
            if best_rejected is None:
                best_rejected = (0.0, 0.0, mfr_score, candidate, "manufacturer_conflict", mfr_norm, candidate.manufacturer_norm)
            continue
        after_manufacturer_priority += 1
        similarity = _similarity(source_base_norm, candidate.base_name_norm or candidate.name_norm, include_partial=include_partial)
        name_score = similarity["score"]
        if name_score < NAME_PREFILTER_THRESHOLD:
            continue
        after_name_filter += 1
        base_score = _base_name_similarity(source_base_norm, candidate.base_name_norm or candidate.name_norm)
        candidate_base = candidate.base_name_norm or candidate.name_norm
        if base_score < 100 and (name_score < 97 or _first_token(source_base_norm) != _first_token(candidate_base)):
            safety_stats["fuzzy_rejected_low_confidence"] += 1
            if best_rejected is None or name_score > best_rejected[0]:
                best_rejected = (name_score, name_score, mfr_score, candidate, "unsafe_name_fuzzy", source_base_norm, candidate.base_name_norm)
            continue
        if manufacturer_decision == "missing" and (structure_decision != "ok" or name_score < 97):
            safety_stats["fuzzy_rejected_low_confidence"] += 1
            if best_rejected is None or name_score > best_rejected[0]:
                best_rejected = (name_score, name_score, mfr_score, candidate, "manufacturer_missing", mfr_norm, candidate.manufacturer_norm)
            continue
        if structure_decision == "ok" and name_score < 88:
            safety_stats["fuzzy_rejected_low_confidence"] += 1
            if best_rejected is None or name_score > best_rejected[0]:
                best_rejected = (name_score, name_score, mfr_score, candidate, "low_score", source_base_norm, candidate.base_name_norm)
            continue
        if structure_decision == "suspicious" and not (
            name_score >= 95
            and base_score >= 95
            and manufacturer_decision in {"alias_match", "similar_match", "missing"}
        ):
            safety_stats["fuzzy_rejected_low_confidence"] += 1
            if best_rejected is None or name_score > best_rejected[0]:
                best_rejected = (name_score, name_score, mfr_score, candidate, structure_reason or "structure_suspicious", source_base_norm, candidate.base_name_norm)
            continue
        if not mfr_norm or not candidate.manufacturer_norm:
            combined = name_score
        elif manufacturer_decision == "alias_match":
            combined = min(100.0, name_score + 15)
        elif manufacturer_decision == "similar_match":
            combined = min(100.0, name_score + 10)
        elif manufacturer_decision == "penalty":
            combined = name_score - 20
        else:
            combined = min(100.0, name_score * 0.90 + mfr_score * 0.10 + 10)
        matched_params, missing_params = _param_counts(source_structure, candidate.structure)
        threshold = 95 if structure_decision == "suspicious" else 88
        missing_penalty = min(6, missing_params * 2)
        combined -= missing_penalty
        rank = (matched_params, combined, name_score)
        if best is None or rank > best[0]:
            best = (rank, candidate, similarity, mfr_score, matched_params, missing_params, threshold, structure_decision, manufacturer_decision)

    if best is None:
        rejected_payload = None
        reason = "no_candidate"
        if best_rejected is not None:
            combined, name_score, mfr_score, candidate, field, source_value, candidate_value = best_rejected
            reason = str(field)
            rejected_payload = {
                "candidateSku": candidate.sku,
                "candidateId": candidate.id,
                "candidateNameNorm": candidate.name_norm,
                "candidateBaseNameNorm": candidate.base_name_norm,
                "candidateManufacturerNorm": candidate.manufacturer_norm,
                "candidateStructure": candidate.structure.__dict__,
                "field": field,
                "sourceValue": source_value,
                "candidateValue": candidate_value,
                "nameScore": round(name_score, 2),
                "manufacturerScore": round(mfr_score, 2),
                "score": round(combined, 2),
            }
        return {
            "candidateSku": "",
            "candidateId": None,
            "candidateNameNorm": "",
            "candidateBaseNameNorm": "",
            "candidateManufacturerNorm": "",
            "sourceBaseNameNorm": source_base_norm,
            "sourceStructure": source_structure.__dict__,
            "nameScore": 0,
            "manufacturerScore": 0,
            "expiryScore": None,
            "expiryPenalty": 0,
            "score": 0,
            "passesThreshold": False,
            "reason": reason,
            "rejectedBestCandidate": rejected_payload,
            "filteredCounts": filtered_counts,
            "filteredExamples": filtered_examples,
            "poolSize": min(after_name_filter, MAX_CANDIDATE_POOL),
            "poolSizeBeforeLimit": pool_size_before_limit,
            "afterNameFilter": after_name_filter,
            "afterStrictFilter": after_strict_filter,
            "afterManufacturerPriority": after_manufacturer_priority,
            "broadFallback": used_broad_fallback,
            "safetyStats": safety_stats,
        }

    _rank, candidate, similarity, mfr_score, matched_params, missing_params, threshold, structure_decision, manufacturer_decision = best
    name_score = similarity["score"]
    combined = name_score if not mfr_norm or not candidate.manufacturer_norm else name_score * 0.92 + mfr_score * 0.08
    combined -= min(6, missing_params * 2)
    expiry_penalty = 0
    # Product catalog rows do not currently carry expiry dates, so expiry is only
    # diagnostic here. If a candidate expiry field is added later, this branch
    # will lower the score without turning expiry into a hard filter.
    candidate_expiry = ""
    if expiry_norm and candidate_expiry and expiry_norm != candidate_expiry:
        expiry_penalty = 5
        combined -= expiry_penalty

    passes = name_score >= threshold and combined >= threshold
    reason = "ok"
    if name_score < threshold:
        reason = "low_score"
    elif combined < threshold:
        reason = "low_score"

    if passes:
        if structure_decision == "suspicious":
            safety_stats["selected_suspicious"] += 1
        if source_structure.raw_dosage is not None or candidate.structure.raw_dosage is not None:
            _verbose_debug(
                "[MATCH_DOSAGE_NORMALIZED] raw_product_dosage=%s normalized_product_dosage=%s raw_candidate_dosage=%s normalized_candidate_dosage=%s",
                source_structure.raw_dosage,
                source_structure.dosage,
                candidate.structure.raw_dosage,
                candidate.structure.dosage,
            )
        _verbose_debug(
            "[MATCH_SELECTED] product_sku=%s candidate_id=%s score=%s name_score=%s manufacturer_score=%s manufacturer_decision=%s structure_decision=%s match_type=%s",
            "",
            candidate.id,
            round(combined, 2),
            round(name_score, 2),
            round(mfr_score, 2),
            manufacturer_decision,
            structure_decision,
            "fuzzy_name_manufacturer",
        )
        _verbose_debug(
            "[MATCH_BALANCED] product_sku=%s product_name=%s candidate_name=%s structure_decision=%s name_score=%s manufacturer_decision=%s final_decision=%s",
            "",
            name_norm,
            candidate.name_norm,
            structure_decision,
            round(name_score, 2),
            manufacturer_decision,
            "select",
        )
    else:
        safety_stats["fuzzy_rejected_low_confidence"] += 1
        _verbose_debug(
            "[MATCH_BALANCED] product_sku=%s product_name=%s candidate_name=%s structure_decision=%s name_score=%s manufacturer_decision=%s final_decision=%s",
            "",
            name_norm,
            candidate.name_norm,
            structure_decision,
            round(name_score, 2),
            manufacturer_decision,
            "reject",
        )
    return {
        "candidateSku": candidate.sku,
        "candidateId": candidate.id,
        "candidateNameNorm": candidate.name_norm,
        "candidateBaseNameNorm": candidate.base_name_norm,
        "candidateManufacturerNorm": candidate.manufacturer_norm,
        "sourceBaseNameNorm": source_base_norm,
        "sourceStructure": source_structure.__dict__,
        "candidateStructure": candidate.structure.__dict__,
        "nameScore": round(name_score, 2),
        "similarity": similarity,
        "manufacturerScore": round(mfr_score, 2),
        "manufacturerDecision": manufacturer_decision,
        "structureDecision": structure_decision,
        "matchedParams": matched_params,
        "missingParams": missing_params,
        "threshold": threshold,
        "expiryScore": 100 if expiry_norm and not candidate_expiry else (100 if expiry_norm == candidate_expiry else 0 if expiry_norm else None),
        "expiryPenalty": expiry_penalty,
        "score": round(combined, 2),
        "passesThreshold": passes,
        "reason": reason,
        "filteredCounts": filtered_counts,
        "filteredExamples": filtered_examples,
        "poolSize": min(after_name_filter, MAX_CANDIDATE_POOL),
        "poolSizeBeforeLimit": pool_size_before_limit,
        "afterNameFilter": after_name_filter,
        "afterStrictFilter": after_strict_filter,
        "afterManufacturerPriority": after_manufacturer_priority,
        "broadFallback": used_broad_fallback,
        "safetyStats": safety_stats,
    }


@dataclass(frozen=True)
class ProductCandidate:
    id: int
    sku: str
    provisor_goods_id: int | None
    raw_name: str
    name_norm: str
    base_name_norm: str
    manufacturer_norm: str
    tokens: frozenset[str]
    structure: DrugStructure


@dataclass(frozen=True)
class SupplierItemCandidate:
    item: CompetitorPriceListItem
    name_norm: str
    base_name_norm: str
    manufacturer_norm: str
    tokens: frozenset[str]
    structure: DrugStructure


@dataclass(frozen=True)
class SupplierSkuCandidate:
    candidate: SupplierItemCandidate
    kind: str


@dataclass(frozen=True)
class SupplierIndexes:
    candidates: list[SupplierItemCandidate]
    token_index: dict[str, list[SupplierItemCandidate]]
    sku_index: dict[str, list[SupplierSkuCandidate]]
    composite_index: dict[str, list[SupplierItemCandidate]]


@dataclass(frozen=True)
class ProductIndexes:
    code_to_id: dict[str, int]
    candidates: list[ProductCandidate]
    exact: dict[tuple[str, str], ProductCandidate]
    token_index: dict[str, list[ProductCandidate]]
    sku_to_candidate: dict[str, ProductCandidate]
    id_to_candidate: dict[int, ProductCandidate]


def _tokens(value: str) -> frozenset[str]:
    out: set[str] = set()
    for token in value.split():
        if len(token) < 3:
            continue
        out.add(token)
        out.update(TOKEN_ALIASES.get(token, set()))
    return frozenset(x for x in out if len(x) >= 3)


def _index_key(field: str, value: object) -> str | None:
    if value in (None, ""):
        return None
    if isinstance(value, float):
        value = _clean_num(value)
    return f"{field}:{value}"


def _structured_index_key(base_name_norm: str, structure: DrugStructure) -> str | None:
    if not base_name_norm:
        return None
    return f"structured:{base_name_norm}:{structure.quantity or ''}:{structure.form or ''}"


def _composite_index_keys(base_name_norm: str, structure: DrugStructure) -> list[str]:
    if not base_name_norm:
        return []
    first_word = base_name_norm.split(" ", 1)[0]
    prefixes = [
        ("base", base_name_norm),
        ("first", first_word),
    ]
    values = [
        ("quantity", structure.quantity),
        ("dosage", structure.dosage),
    ]
    out: list[str] = []
    for prefix_kind, prefix_value in prefixes:
        if not prefix_value:
            continue
        for field, value in values:
            key_value = _index_key(field, value)
            if key_value:
                out.append(f"composite:{prefix_kind}:{prefix_value}:{key_value}")
    return out


def _key_part(value: object) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, tuple):
        return ",".join(_key_part(part) for part in value)
    if isinstance(value, float):
        value = _clean_num(value)
    return str(value)


def _dosage_signature(structure: DrugStructure) -> str:
    if structure.strength_signature:
        return "strength:" + _key_part(structure.strength_signature)
    values = [
        ("dosage", structure.dosage),
        ("dosage_volume", structure.dosage_volume),
        ("concentration", structure.concentration),
        ("iu", structure.iu_dosage),
    ]
    return "|".join(f"{name}:{_key_part(value)}" for name, value in values if value is not None)


def _supplier_composite_keys(base_name_norm: str, manufacturer_norm: str, structure: DrugStructure) -> list[str]:
    if not base_name_norm:
        return []
    form_values = list(structure.forms or ((structure.form,) if structure.form else ("",)))
    key_specs = [
        ("exact", ("base", "manufacturer", "quantity", "dosage", "volume", "weight", "form", "percent")),
        ("relaxed1", ("base", "manufacturer", "quantity", "dosage")),
        ("relaxed2", ("base", "manufacturer", "volume", "form")),
        ("relaxed3", ("base", "manufacturer", "weight", "form")),
    ]
    out: list[str] = []
    for form_value in form_values:
        parts = {
            "base": base_name_norm,
            "manufacturer": manufacturer_norm,
            "quantity": _key_part(structure.quantity),
            "dosage": _dosage_signature(structure),
            "volume": _key_part(structure.volume),
            "weight": _key_part(structure.weight),
            "form": _key_part(form_value),
            "percent": _key_part(structure.percent_strength),
        }
        for kind, fields in key_specs:
            if kind != "exact" and not any(parts[field] for field in fields[2:]):
                continue
            key = f"{kind}:" + "|".join(parts[field] for field in fields)
            if key not in out:
                out.append(key)
    return out


def _add_index(index: dict[str, list["ProductCandidate"]], key: str | None, candidate: "ProductCandidate") -> None:
    if key:
        index[key].append(candidate)


def _add_candidate_sku(sku_to_candidate: dict[str, ProductCandidate], value: object, candidate: ProductCandidate) -> None:
    raw = normalize_external_sku(value)
    if raw:
        sku_to_candidate.setdefault(raw, candidate)
    normalized = normalize_sku(value)
    if normalized:
        sku_to_candidate.setdefault(normalized, candidate)
    for variant in normalize_sku_variants(value):
        sku_to_candidate.setdefault(variant, candidate)


def _candidate_from_result(indexes: ProductIndexes, candidate_result: dict) -> ProductCandidate | None:
    candidate_id = candidate_result.get("candidateId")
    if candidate_id is not None:
        return indexes.id_to_candidate.get(int(candidate_id))
    sku = candidate_result.get("candidateSku")
    if not sku:
        return None
    for key in [sku, normalize_external_sku(sku), normalize_sku(sku), *normalize_sku_variants(sku)]:
        if key and key in indexes.sku_to_candidate:
            return indexes.sku_to_candidate[key]
    return None


def _product_indexes(db: Session) -> ProductIndexes:
    rows = (
        db.execute(
            select(Product, ProductExtra)
            .join(ProductExtra, ProductExtra.product_id == Product.id, isouter=True)
            .order_by(Product.id.asc())
        )
        .all()
    )
    code_to_id: dict[str, int] = {}
    candidates: list[ProductCandidate] = []
    exact: dict[tuple[str, str], ProductCandidate] = {}
    token_index: dict[str, list[ProductCandidate]] = defaultdict(list)
    sku_to_candidate: dict[str, ProductCandidate] = {}
    id_to_candidate: dict[int, ProductCandidate] = {}
    product_goods_counts: dict[int, int] = defaultdict(int)
    for product, _extra in rows:
        goods_id = _to_int(getattr(product, "provisor_goods_id", None))
        if goods_id:
            product_goods_counts[goods_id] += 1
    for product, extra in rows:
        raw_code = normalize_external_sku(product.code)
        if raw_code:
            code_to_id[raw_code] = int(product.id)
        sku = normalize_sku(product.code) or raw_code
        if sku:
            code_to_id[sku] = int(product.id)
        for variant in normalize_sku_variants(product.code):
            code_to_id[variant] = int(product.id)
        product_goods_id = _to_int(getattr(product, "provisor_goods_id", None))
        if product_goods_id and product_goods_counts.get(product_goods_id) == 1:
            code_to_id[_provisor_key("goods_id", product_goods_id)] = int(product.id)
        name_norm = normalize_text(normalize_drug_text(product.name))
        structure = parse_drug_structure(product.name)
        base_name_norm = normalize_text(structure.base_name)
        cand = ProductCandidate(
            id=int(product.id),
            sku=str(product.code or ""),
            provisor_goods_id=product_goods_id,
            raw_name=str(product.name or ""),
            name_norm=name_norm,
            base_name_norm=base_name_norm,
            manufacturer_norm=normalize_manufacturer_text(extra.manufacturer if extra is not None else ""),
            tokens=_tokens(base_name_norm or name_norm),
            structure=structure,
        )
        candidates.append(cand)
        id_to_candidate[cand.id] = cand
        _add_candidate_sku(sku_to_candidate, cand.sku, cand)
        exact.setdefault((cand.name_norm, cand.manufacturer_norm), cand)
        for token in cand.tokens:
            token_index[token].append(cand)
        first_word = (base_name_norm or name_norm).split(" ", 1)[0]
        _add_index(token_index, f"base:{base_name_norm}" if base_name_norm else None, cand)
        _add_index(token_index, f"first:{first_word}" if first_word else None, cand)
        _add_index(token_index, f"prefix:{(base_name_norm or name_norm)[:5]}" if (base_name_norm or name_norm) else None, cand)
        _add_index(token_index, f"prefix:{(base_name_norm or name_norm)[:3]}" if (base_name_norm or name_norm) else None, cand)
        for key in _composite_index_keys(base_name_norm, structure):
            _add_index(token_index, key, cand)
    return ProductIndexes(
        code_to_id=code_to_id,
        candidates=candidates,
        exact=exact,
        token_index=token_index,
        sku_to_candidate=sku_to_candidate,
        id_to_candidate=id_to_candidate,
    )


def _reference_sku_lookup_keys(value: object) -> list[str]:
    keys: list[str] = []
    for key in [
        normalize_external_sku(value),
        normalize_sku(value),
        *normalize_sku_variants(value),
    ]:
        if key and key not in keys:
            keys.append(key)
    return keys


def _add_reference_lookup(lookup: dict[str, set[int]], key: object, product_id: object) -> None:
    product_int = _to_int(product_id)
    if not product_int:
        return
    for lookup_key in _reference_sku_lookup_keys(key):
        lookup.setdefault(lookup_key, set()).add(product_int)


def _sync_provisor_reference_mapping_from_items(db: Session, *, account_id: str = "existing") -> dict[str, object]:
    started_at = time.perf_counter()
    rows = (
        db.execute(
            select(
                CompetitorPriceList.account_id,
                CompetitorPriceListItem.filial_id,
                CompetitorPriceListItem.distributor_goods_id,
                CompetitorPriceListItem.provisor_goods_id,
            )
            .join(CompetitorPriceList, CompetitorPriceList.id == CompetitorPriceListItem.price_list_id)
            .where(CompetitorPriceList.source_type == "provisor")
            .where(CompetitorPriceListItem.filial_id.in_(PROVISOR_REFERENCE_FILIAL_IDS))
            .where(CompetitorPriceListItem.provisor_goods_id.is_not(None))
        )
        .all()
    )
    products = db.execute(select(Product)).scalars().all()
    products_total = len(products)
    products_with_goods_id_before = sum(1 for product in products if _to_int(product.provisor_goods_id))
    products_without_goods_id_before = products_total - products_with_goods_id_before

    def empty_stats() -> dict[str, object]:
        return {
            "products_total": products_total,
            "products_with_goods_id_before": products_with_goods_id_before,
            "products_without_goods_id_before": products_without_goods_id_before,
            "matched_via_128": 0,
            "matched_via_133": 0,
            "matched_via_both_same": 0,
            "conflict_128_133": 0,
            "existing_goods_id_conflict": 0,
            "updated_total": 0,
            "products_with_goods_id_after": products_with_goods_id_before,
            "coverage_delta": 0,
            "products_with_goods_id_after_percent": round((products_with_goods_id_before / products_total) * 100, 2) if products_total else 0,
            "reference_code_values_128": 0,
            "reference_code_values_133": 0,
            "unique_to_128": 0,
            "unique_to_133": 0,
            "overlap_same_goods_id": 0,
            "overlap_conflicting_goods_id": 0,
            "conflict_examples": [],
            "existing_conflict_examples": [],
            # Backward-compatible aliases used by existing diagnostics/tests.
            "mapped_products": 0,
            "updated_products": 0,
            "created": 0,
            "updated_matches": 0,
            "skipped_no_sku": 0,
            "skipped_no_goods_id": 0,
            "skipped_product_not_found": 0,
            "updatedViaProductCode": 0,
            "updatedViaSourceGoodsMatch": 0,
            "updatedViaCompetitorCodeMapping": 0,
            "updatedViaBranchStock": 0,
            "updatedViaBranchCost": 0,
            "updatedViaProductRating": 0,
            "ambiguousMatch": 0,
            "conflict": 0,
        }

    if not rows:
        stats = empty_stats()
        logger.info(
            "[PROVISOR_REFERENCE_SYNC_DONE] account_id=%s reference_filial_ids=%s items_count=0 products_total=%s products_with_goods_id_before=%s updated_total=0 products_with_goods_id_after=%s elapsed_ms=%s",
            account_id,
            ",".join(str(x) for x in PROVISOR_REFERENCE_FILIAL_IDS),
            products_total,
            products_with_goods_id_before,
            products_with_goods_id_before,
            round((time.perf_counter() - started_at) * 1000, 2),
        )
        return stats

    logger.info(
        "[PROVISOR_REFERENCE_SYNC_START] account_id=%s reference_filial_ids=%s",
        account_id,
        ",".join(str(x) for x in PROVISOR_REFERENCE_FILIAL_IDS),
    )

    reference_maps: dict[int, dict[str, set[int]]] = {filial_id: defaultdict(set) for filial_id in PROVISOR_REFERENCE_FILIAL_IDS}
    skipped_no_sku = 0
    skipped_no_goods_id = 0
    for row_account_id, filial_id_raw, distributor_goods_id, goods_id in rows:
        filial_id = _to_int(filial_id_raw)
        if filial_id not in PROVISOR_REFERENCE_FILIAL_IDS:
            continue
        sku = normalize_external_sku(distributor_goods_id)
        goods = _to_int(goods_id)
        if not sku:
            skipped_no_sku += 1
            continue
        if not goods:
            skipped_no_goods_id += 1
            continue
        reference_maps[filial_id][sku].add(goods)

    stats = empty_stats()
    stats["skipped_no_sku"] = skipped_no_sku
    stats["skipped_no_goods_id"] = skipped_no_goods_id
    codes_128 = set(reference_maps[128])
    codes_133 = set(reference_maps[133])
    overlap = codes_128 & codes_133
    stats["reference_code_values_128"] = len(codes_128)
    stats["reference_code_values_133"] = len(codes_133)
    stats["unique_to_128"] = len(codes_128 - codes_133)
    stats["unique_to_133"] = len(codes_133 - codes_128)
    stats["overlap_same_goods_id"] = sum(1 for code in overlap if reference_maps[128][code] == reference_maps[133][code] and len(reference_maps[128][code]) == 1)
    stats["overlap_conflicting_goods_id"] = sum(1 for code in overlap if reference_maps[128][code] != reference_maps[133][code])

    updated = 0
    mapped = 0
    conflict_128_133 = 0
    existing_goods_id_conflict = 0
    skipped_product_not_found = 0
    conflict_examples: list[dict[str, object]] = []
    existing_conflict_examples: list[dict[str, object]] = []

    for product in products:
        code = normalize_external_sku(product.code)
        if not code:
            skipped_no_sku += 1
            stats["skipped_no_sku"] = skipped_no_sku
            continue
        goods_128 = reference_maps[128].get(code, set())
        goods_133 = reference_maps[133].get(code, set())
        candidates = goods_128 | goods_133
        if not candidates:
            skipped_product_not_found += 1
            continue
        if len(candidates) != 1:
            conflict_128_133 += 1
            if len(conflict_examples) < 20:
                conflict_examples.append(
                    {
                        "product_id": int(product.id),
                        "product_code": product.code,
                        "goods_128": sorted(goods_128),
                        "goods_133": sorted(goods_133),
                    }
                )
            continue
        goods = next(iter(candidates))
        mapped += 1
        existing_goods_id = _to_int(product.provisor_goods_id)
        if existing_goods_id:
            if existing_goods_id != goods:
                existing_goods_id_conflict += 1
                if len(existing_conflict_examples) < 20:
                    existing_conflict_examples.append(
                        {
                            "product_id": int(product.id),
                            "product_code": product.code,
                            "existing_goods_id": existing_goods_id,
                            "implied_goods_id": goods,
                            "goods_128": sorted(goods_128),
                            "goods_133": sorted(goods_133),
                        }
                    )
            continue
        product.provisor_goods_id = goods
        updated += 1
        if goods_128 and goods_133:
            stats["matched_via_both_same"] = int(stats["matched_via_both_same"]) + 1
        elif goods_128:
            stats["matched_via_128"] = int(stats["matched_via_128"]) + 1
        else:
            stats["matched_via_133"] = int(stats["matched_via_133"]) + 1

    products_with_goods_id_after = products_with_goods_id_before + updated
    stats["mapped_products"] = mapped
    stats["updated_products"] = updated
    stats["updatedViaProductCode"] = updated
    stats["skipped_product_not_found"] = skipped_product_not_found
    stats["conflict_128_133"] = conflict_128_133
    stats["existing_goods_id_conflict"] = existing_goods_id_conflict
    stats["updated_total"] = updated
    stats["products_with_goods_id_after"] = products_with_goods_id_after
    stats["coverage_delta"] = updated
    stats["products_with_goods_id_after_percent"] = round((products_with_goods_id_after / products_total) * 100, 2) if products_total else 0
    stats["conflict"] = conflict_128_133 + existing_goods_id_conflict
    stats["ambiguousMatch"] = conflict_128_133
    stats["conflict_examples"] = conflict_examples
    stats["existing_conflict_examples"] = existing_conflict_examples

    logger.info(
        "[PROVISOR_REFERENCE_SYNC_DONE] account_id=%s reference_filial_ids=%s items_count=%s products_total=%s products_with_goods_id_before=%s matched_via_128=%s matched_via_133=%s matched_via_both_same=%s conflict_128_133=%s existing_goods_id_conflict=%s updated_total=%s products_with_goods_id_after=%s coverage_delta=%s elapsed_ms=%s",
        account_id,
        ",".join(str(x) for x in PROVISOR_REFERENCE_FILIAL_IDS),
        len(rows),
        products_total,
        products_with_goods_id_before,
        stats["matched_via_128"],
        stats["matched_via_133"],
        stats["matched_via_both_same"],
        conflict_128_133,
        existing_goods_id_conflict,
        updated,
        products_with_goods_id_after,
        updated,
        round((time.perf_counter() - started_at) * 1000, 2),
    )
    if mapped or updated:
        _verbose_debug(
            "[PROVISOR_REFERENCE] account_id=%s reference_filial_ids=%s available=true mapped_products=%s updated_products=%s source=existing_items",
            account_id,
            ",".join(str(x) for x in PROVISOR_REFERENCE_FILIAL_IDS),
            mapped,
            updated,
        )
    return stats


def _raw_payload(item: CompetitorPriceListItem) -> dict:
    try:
        data = json.loads(item.raw_json or "{}")
        return data if isinstance(data, dict) else {}
    except Exception:
        return {}


def _price_decimal(value: object) -> Decimal | None:
    if value is None:
        return None
    try:
        dec = Decimal(str(value))
    except Exception:
        return None
    return dec if dec > 0 else None


def _json_tuple(value: object) -> tuple[str, ...] | None:
    if not value:
        return None
    try:
        parsed = json.loads(str(value))
    except Exception:
        return None
    if not isinstance(parsed, list):
        return None
    out = tuple(str(x) for x in parsed if str(x or "").strip())
    return out or None


def _float_or_none(value: object) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except Exception:
        return None


def _int_or_none(value: object) -> int | None:
    if value in (None, ""):
        return None
    try:
        return int(float(str(value)))
    except Exception:
        return None


def _signature_from_text(value: object) -> tuple[float, ...] | None:
    if not value:
        return None
    try:
        parsed = json.loads(str(value))
        if isinstance(parsed, list):
            out = tuple(float(x) for x in parsed)
            return out or None
    except Exception:
        pass
    return None


def _materialize_match_fields(item: CompetitorPriceListItem, *, raw_name: str | None = None, raw_manufacturer: str | None = None) -> DrugStructure:
    if raw_name is not None:
        item.raw_name = raw_name
    if raw_manufacturer is not None:
        item.raw_manufacturer = raw_manufacturer
    structure = parse_drug_structure(item.raw_name)
    item.normalized_name = normalize_text(normalize_drug_text(item.raw_name))
    item.normalized_manufacturer = normalize_manufacturer_text(item.raw_manufacturer)
    item.parsed_base_name = normalize_text(structure.base_name)
    item.parsed_form = structure.form or ""
    item.parsed_forms_json = json.dumps(list(structure.forms or ()), ensure_ascii=False)
    item.parsed_dosage = structure.dosage
    item.parsed_dosage_volume = structure.dosage_volume
    item.parsed_quantity = structure.quantity
    item.parsed_volume = structure.volume
    item.parsed_weight = structure.weight
    item.parsed_percent_strength = structure.percent_strength
    item.parsed_concentration = structure.concentration
    item.parsed_iu_dosage = structure.iu_dosage
    item.parsed_strength_signature = json.dumps(list(structure.strength_signature or ()), ensure_ascii=False)
    item.parsed_dimensions_json = json.dumps(list(structure.dimensions or ()), ensure_ascii=False)
    item.parsed_critical_tokens_json = json.dumps(list(structure.critical_tokens or ()), ensure_ascii=False)
    return structure


def _cached_structure_from_item(item: CompetitorPriceListItem) -> DrugStructure:
    if not (item.normalized_name and item.parsed_base_name):
        return _materialize_match_fields(item)
    return DrugStructure(
        base_name=str(item.parsed_base_name or ""),
        quantity=_int_or_none(item.parsed_quantity),
        volume=_clean_num(_float_or_none(item.parsed_volume)),
        weight=_clean_num(_float_or_none(item.parsed_weight)),
        dosage=_clean_num(_float_or_none(item.parsed_dosage)),
        raw_dosage=None,
        dosage_volume=_clean_num(_float_or_none(item.parsed_dosage_volume)),
        strength_signature=_signature_from_text(item.parsed_strength_signature),
        concentration=_clean_num(_float_or_none(item.parsed_concentration)),
        percent_strength=_clean_num(_float_or_none(item.parsed_percent_strength)),
        iu_dosage=_clean_num(_float_or_none(item.parsed_iu_dosage)),
        form=str(item.parsed_form or "") or None,
        forms=_json_tuple(item.parsed_forms_json),
        dimensions=_json_tuple(item.parsed_dimensions_json),
        critical_tokens=_json_tuple(item.parsed_critical_tokens_json),
    )


def _source_goods_matches(*, db: Session, price_format_id: int, source_type: str) -> dict[str, SourceGoodsMatch]:
    rows = (
        db.execute(
            select(SourceGoodsMatch)
            .where(SourceGoodsMatch.price_format_id == price_format_id)
            .where(SourceGoodsMatch.source_type == source_type)
        )
        .scalars()
        .all()
    )
    return {normalize_external_sku(row.distributor_goods_id): row for row in rows if row.distributor_goods_id}


def _source_goods_matches_by_goods_id(*, db: Session, price_format_id: int, source_type: str) -> dict[int, SourceGoodsMatch]:
    rows = (
        db.execute(
            select(SourceGoodsMatch)
            .where(SourceGoodsMatch.price_format_id == price_format_id)
            .where(SourceGoodsMatch.source_type == source_type)
            .where(SourceGoodsMatch.goods_id.is_not(None))
        )
        .scalars()
        .all()
    )
    out: dict[int, SourceGoodsMatch] = {}
    for row in rows:
        goods = _to_int(row.goods_id)
        if goods and (goods not in out or str(row.match_method or "").startswith("manual")):
            out[goods] = row
    return out


def _approved_provisor_substitutes_by_product(db: Session) -> dict[int, list[ProductSubstituteMatch]]:
    rows = (
        db.execute(
            select(ProductSubstituteMatch)
            .where(ProductSubstituteMatch.source_type == "provisor")
            .where(ProductSubstituteMatch.status == "approved")
            .where(ProductSubstituteMatch.source_goods_id.is_not(None))
            .order_by(ProductSubstituteMatch.product_id.asc(), ProductSubstituteMatch.priority.asc(), ProductSubstituteMatch.id.asc())
        )
        .scalars()
        .all()
    )
    out: dict[int, list[ProductSubstituteMatch]] = defaultdict(list)
    for row in rows:
        out[int(row.product_id)].append(row)
    return out


def _approved_provisor_substitutes_by_goods(db: Session) -> dict[int, list[ProductSubstituteMatch]]:
    by_goods: dict[int, list[ProductSubstituteMatch]] = defaultdict(list)
    for rows in _approved_provisor_substitutes_by_product(db).values():
        for row in rows:
            goods = _to_int(row.source_goods_id)
            if goods:
                by_goods[goods].append(row)
    return by_goods


def _matched_price_row(item: CompetitorPriceListItem) -> dict:
    return {
        "product_id": int(item.product_id),
        "price": float(item.distributor_price),
        "source_item_id": int(item.id),
        "match_type": item.match_type or "matched",
        "source_goods_id": _to_int(item.provisor_goods_id),
        "source_distributor_goods_id": item.distributor_goods_id or "",
        "source_manufacturer": item.raw_manufacturer or "",
    }


def _upsert_source_goods_match(
    *,
    db: Session,
    match_cache: dict[str, SourceGoodsMatch] | None = None,
    price_format_id: int,
    source_type: str,
    distributor_goods_id: str,
    goods_id: object,
    distributor_goods_name: str,
    distributor_producer: str,
    product_id: int,
    similarity_score: object,
    match_method: str,
) -> None:
    sku = normalize_external_sku(distributor_goods_id)
    if not sku:
        return
    row = match_cache.get(sku) if match_cache is not None else None
    if row is None and match_cache is None:
        row = (
            db.execute(
                select(SourceGoodsMatch)
                .where(SourceGoodsMatch.price_format_id == price_format_id)
                .where(SourceGoodsMatch.source_type == source_type)
                .where(SourceGoodsMatch.distributor_goods_id == sku)
            )
            .scalars()
            .first()
        )
    if row is None:
        row = SourceGoodsMatch(
            price_format_id=price_format_id,
            source_type=source_type,
            distributor_goods_id=sku,
            product_id=product_id,
        )
        db.add(row)
        if match_cache is not None:
            match_cache[sku] = row
    row.goods_id = int(goods_id) if str(goods_id or "").isdigit() else None
    row.distributor_goods_name = distributor_goods_name[:512]
    row.distributor_producer = distributor_producer[:256]
    row.product_id = product_id
    row.similarity_score = float(similarity_score) if similarity_score not in (None, "") else None
    row.match_method = match_method[:64]
    row.updated_at = datetime.utcnow()


def _unmatched_debug(
    *,
    price_list: CompetitorPriceList,
    items: list[CompetitorPriceListItem],
    code_to_id: dict[str, int],
    candidates: list["ProductCandidate"],
    exact: dict[tuple[str, str], "ProductCandidate"],
    token_index: dict[str, list["ProductCandidate"]],
    limit: int = 10,
) -> list[dict]:
    out: list[dict] = []
    for item in items:
        if item.product_id is not None and item.matched_sku:
            continue
        raw = _raw_payload(item)
        raw_inner = raw.get("raw") if isinstance(raw.get("raw"), dict) else raw
        if price_list.source_type == "provisor":
            goods_id = item.provisor_goods_id or raw_inner.get("goodsId")
            distributor_goods_id = item.distributor_goods_id or raw_inner.get("distributorGoodsId")
            expiry_date = item.expiry_date or raw_inner.get("shelfLife")
            variants = provisor_item_variants(goods_id, distributor_goods_id, expiry_date)
            raw_name = item.raw_name or item.name or item.distributor_goods_name
            raw_manufacturer = item.raw_manufacturer or raw_inner.get("distributorProducer") or ""
            candidate = _vidman_candidate_result(
                name_norm=raw_name,
                mfr_norm=normalize_manufacturer_text(raw_manufacturer),
                expiry_norm=normalize_expiry(expiry_date),
                candidates=candidates,
                exact=exact,
                token_index=token_index,
            )
            out.append(
                {
                    "itemId": item.id,
                    "name": raw_name,
                    "manufacturer": raw_manufacturer,
                    "distributorGoodsId": distributor_goods_id,
                    "expiryDate": normalize_expiry(expiry_date),
                    "bestSkuCandidate": next(({"keyType": kind, "key": value, "score": 100} for kind, value in variants if value in code_to_id), None),
                    "bestCandidate": candidate,
                    "score": candidate.get("score", 0),
                    "reason": candidate.get("reason") or "no_provisor_key_in_product_catalog",
                    "triedKeys": [{"keyType": kind, "key": value} for kind, value in variants[:5]],
                }
            )
        elif price_list.source_type == "vidman":
            raw_name = item.raw_name or item.name or item.distributor_goods_name
            raw_manufacturer = item.raw_manufacturer or ""
            candidate = _vidman_candidate_result(
                name_norm=raw_name,
                mfr_norm=normalize_manufacturer_text(raw_manufacturer),
                expiry_norm=normalize_expiry(item.expiry_date or raw_inner.get("expiryDate") or raw_inner.get("expiry_date")),
                candidates=candidates,
                exact=exact,
                token_index=token_index,
            )
            out.append(
                {
                    "itemId": item.id,
                    "name": raw_name,
                    "manufacturer": raw_manufacturer,
                    "expiryDate": item.expiry_date or "",
                    "bestCandidate": candidate,
                    "score": candidate.get("score", 0),
                    "reason": candidate.get("reason", "unknown"),
                }
            )
        else:
            out.append(
                {
                    "itemId": item.id,
                    "name": item.raw_name or item.name or item.distributor_goods_name,
                    "distributorGoodsId": item.distributor_goods_id,
                    "score": 0,
                    "reason": "no_sku_match",
                }
            )
        if len(out) >= limit:
            break
    return out


def rematch_price_list_items(
    *,
    db: Session,
    price_list: CompetitorPriceList,
    product_indexes: ProductIndexes | None = None,
    debug: bool = False,
    flush: bool = True,
) -> dict:
    started_at = time.perf_counter()
    if price_list.source_type == "provisor":
        _sync_provisor_reference_mapping_from_items(db, account_id=price_list.account_id or "existing")
    indexes = product_indexes or _product_indexes(db)
    code_to_id = indexes.code_to_id
    candidates = indexes.candidates
    exact = indexes.exact
    token_index = indexes.token_index
    indexed_at = time.perf_counter()
    product_ids = {candidate.id for candidate in candidates}
    source_matches = (
        _source_goods_matches(db=db, price_format_id=price_list.price_format_id, source_type=price_list.source_type)
        if price_list.source_type == "provisor"
        else {}
    )
    source_matches_by_goods_id = (
        _source_goods_matches_by_goods_id(db=db, price_format_id=price_list.price_format_id, source_type=price_list.source_type)
        if price_list.source_type == "provisor"
        else {}
    )
    substitute_matches_by_goods_id = _approved_provisor_substitutes_by_goods(db) if price_list.source_type == "provisor" else {}
    items = (
        db.execute(select(CompetitorPriceListItem).where(CompetitorPriceListItem.price_list_id == price_list.id))
        .scalars()
        .all()
    )
    present_provisor_goods_ids = {
        goods
        for item in items
        for goods in [_to_int(item.provisor_goods_id or _raw_payload(item).get("goodsId"))]
        if goods
    }
    product_primary_goods_ids = {
        int(candidate.id): _to_int(candidate.provisor_goods_id)
        for candidate in candidates
        if _to_int(candidate.provisor_goods_id)
    }
    loaded_at = time.perf_counter()

    stats = {
        "incoming": len(items),
        "matched": 0,
        "fuzzy": 0,
        "unmatched": 0,
        "filteredByQuantity": 0,
        "filteredByVolume": 0,
        "filteredByDosage": 0,
        "filteredByConcentration": 0,
        "filteredByForm": 0,
        "filterExamples": [],
        "matchedBySavedSku": 0,
        "matchedByManualMapping": 0,
        "manualRejected": 0,
        "matchedBySku": 0,
        "matchedByFuzzy": 0,
        **_empty_provisor_exact_counts(),
        "rejectReasons": {},
        "detailedRejectReasons": _empty_match_reject_counts(),
        "candidatePoolTotal": 0,
        "candidatePoolCount": 0,
        "candidatePoolMax": 0,
        "candidatePoolBroadFallbacks": 0,
        "structureRejects": 0,
        "structureSuspicious": 0,
        "selectedSuspicious": 0,
        "manufacturerHardRejectsSafety": 0,
        "fuzzyRejectedLowConfidence": 0,
        "_auditRows": [],
        "_unmatchedAuditRows": [],
    }
    matched_rows_for_prices: list[dict] = []
    duplicate_bucket: dict[tuple[str, str], list[CompetitorPriceListItem]] = defaultdict(list)

    def record_candidate_stats(candidate_result: dict) -> None:
        pool_size = int(candidate_result.get("poolSize") or 0)
        stats["candidatePoolTotal"] += pool_size
        stats["candidatePoolCount"] += 1
        stats["candidatePoolMax"] = max(int(stats["candidatePoolMax"]), pool_size)
        if candidate_result.get("broadFallback"):
            stats["candidatePoolBroadFallbacks"] += 1

    for item in items:
        raw = _raw_payload(item)
        raw_inner = raw.get("raw") if isinstance(raw.get("raw"), dict) else raw
        source = price_list.source_type

        item.match_type = "unmatched"
        item.match_score = None
        item.matched_sku = ""
        item.product_id = None

        item.raw_name = item.name or item.distributor_goods_name or str(raw_inner.get("distributorGoodsName") or raw_inner.get("name") or "")
        item.raw_manufacturer = str(
            raw.get("manufacturer")
            or raw_inner.get("distributorProducer")
            or raw_inner.get("manufacturer")
            or item.raw_manufacturer
            or ""
        )

        if source == "provisor":
            raw_sku = normalize_external_sku(item.distributor_goods_id)
            goods_id = item.provisor_goods_id or raw_inner.get("goodsId")
            goods_id_int = _to_int(goods_id)
            item.expiry_date = normalize_expiry(item.expiry_date or raw_inner.get("shelfLife"))
            item.match_key = provisor_match_key(goods_id, raw_sku, item.expiry_date)
            goods_match = source_matches_by_goods_id.get(goods_id_int) if goods_id_int else None
            if goods_match is not None and int(goods_match.product_id) in product_ids:
                product = indexes.id_to_candidate.get(int(goods_match.product_id))
                product_code = product.sku if product is not None else str(goods_match.product_id)
                item.product_id = int(goods_match.product_id)
                item.match_type = "provisor_goods_id"
                item.match_score = 100
                item.matched_sku = normalize_sku(product_code) or product_code
                item.match_key = _provisor_key("goods_id", goods_id_int)
                stats["matched"] += 1
                stats["matchedBySavedSku"] += 1
                stats["matchedByGoodsId"] = int(stats.get("matchedByGoodsId") or 0) + 1
                duplicate_bucket[(item.matched_sku, item.expiry_date or "no_exp")].append(item)
                logger.info(
                    "[PROVISOR_GOODS_ID_MATCH] account_id=%s filial_id=%s goods_id=%s product_id=%s product_code=%s match_type=provisor_goods_id",
                    price_list.account_id or "",
                    item.filial_id or raw_inner.get("filialId") or "",
                    goods_id_int,
                    item.product_id,
                    product_code,
                )
                continue
            substitute_rows = substitute_matches_by_goods_id.get(goods_id_int or 0) if goods_id_int else None
            if substitute_rows:
                substitute = next(
                    (
                        row
                        for row in substitute_rows
                        if int(row.product_id) in product_ids
                        and product_primary_goods_ids.get(int(row.product_id)) not in present_provisor_goods_ids
                    ),
                    None,
                )
                if substitute is not None:
                    product = indexes.id_to_candidate.get(int(substitute.product_id))
                    product_code = product.sku if product is not None else str(substitute.product_id)
                    item.product_id = int(substitute.product_id)
                    item.match_type = "provisor_manual_substitute"
                    item.match_score = 100
                    item.matched_sku = normalize_sku(product_code) or product_code
                    item.match_key = _provisor_key("goods_id", goods_id_int)
                    stats["matched"] += 1
                    stats["matchedBySavedSku"] += 1
                    duplicate_bucket[(item.matched_sku, item.expiry_date or "no_exp")].append(item)
                    logger.info(
                        "[PROVISOR_SUBSTITUTE_MATCH] product_id=%s product_sku=%s source_goods_id=%s source_name=%s source_manufacturer=%s",
                        item.product_id,
                        product_code,
                        goods_id_int,
                        substitute.source_name or item.raw_name or item.name or "",
                        substitute.source_manufacturer or item.raw_manufacturer or "",
                    )
                    continue
            deferred_saved_match = source_matches.get(raw_sku)
            saved_match = None
            if saved_match is not None and int(saved_match.product_id) in product_ids:
                item.product_id = int(saved_match.product_id)
                item.match_type = "saved_distributor_goods_id"
                item.match_score = float(saved_match.similarity_score) if saved_match.similarity_score is not None else 100
                item.matched_sku = raw_sku
                stats["matched"] += 1
                stats["matchedBySavedSku"] += 1
                duplicate_bucket[(raw_sku, item.expiry_date or "no_exp")].append(item)
                _upsert_source_goods_match(
                    db=db,
                    match_cache=source_matches,
                    price_format_id=price_list.price_format_id,
                    source_type="provisor",
                    distributor_goods_id=raw_sku,
                    goods_id=goods_id,
                    distributor_goods_name=item.raw_name,
                    distributor_producer=item.raw_manufacturer,
                    product_id=int(saved_match.product_id),
                    similarity_score=item.match_score,
                    match_method="saved_distributor_goods_id",
                )
                continue

            matched_kind = ""
            matched_sku = ""
            for kind, variant in _provisor_product_code_keys(
                goods_id=goods_id,
                distributor_goods_id=raw_sku,
                expiry_date=item.expiry_date,
                filial_id=item.filial_id or raw_inner.get("filialId"),
            ):
                if variant in code_to_id:
                    matched_kind = kind
                    matched_sku = variant
                    break
            if matched_sku:
                exact_product_candidate = indexes.id_to_candidate.get(int(code_to_id[matched_sku]))
                exact_source_structure = parse_drug_structure(item.raw_name)
                if exact_product_candidate is not None and matched_kind != "reference_sku":
                    structure_decision, structure_reason = _strict_structure_decision(exact_source_structure, exact_product_candidate.structure)
                    if structure_decision == "reject":
                        product_value, candidate_value = _structure_values_for_reason(exact_source_structure, exact_product_candidate.structure, structure_reason)
                        stats["structureRejects"] += 1
                        _verbose_debug(
                            "[MATCH_STRUCTURE_REJECT] product_sku=%s product_name=%s candidate_id=%s candidate_name=%s reason=%s product_value=%s candidate_value=%s",
                            matched_sku,
                            item.raw_name,
                            exact_product_candidate.id,
                            exact_product_candidate.name_norm,
                            structure_reason,
                            product_value,
                            candidate_value,
                        )
                        _bump_reason(stats, structure_reason)
                        stats["unmatched"] += 1
                        continue
                    if structure_decision == "suspicious":
                        stats["structureSuspicious"] += 1
                        stats["selectedSuspicious"] += 1
                item.product_id = code_to_id[matched_sku]
                item.match_type = "reference_filial_128_distributor_goods_id" if matched_kind == "reference_sku" else "provisor_goods_id"
                item.match_score = 100
                item.matched_sku = matched_sku
                stats["matched"] += 1
                stats["matchedBySku"] += 1
                stats[matched_kind] = int(stats.get(matched_kind, 0)) + 1
                _bump_provisor_exact(stats, matched_kind)
                duplicate_bucket[(matched_sku, item.expiry_date or "no_exp")].append(item)
                if matched_kind == "reference_sku":
                    logger.info(
                        "[PROVISOR_REFERENCE_SKU_MATCH] account_id=%s filial_id=%s distributor_goods_id=%s goods_id=%s product_id=%s match_type=reference_filial_128_distributor_goods_id",
                        price_list.account_id or "",
                        item.filial_id or raw_inner.get("filialId") or "",
                        raw_sku,
                        goods_id or "",
                        item.product_id,
                    )
                else:
                    logger.info(
                        "[PROVISOR_GOODS_ID_MATCH] account_id=%s filial_id=%s goods_id=%s product_id=%s product_code=%s match_type=provisor_goods_id",
                        price_list.account_id or "",
                        item.filial_id or raw_inner.get("filialId") or "",
                        goods_id or "",
                        item.product_id,
                        matched_sku,
                    )
                _upsert_source_goods_match(
                    db=db,
                    match_cache=source_matches,
                    price_format_id=price_list.price_format_id,
                    source_type="provisor",
                    distributor_goods_id=raw_sku,
                    goods_id=goods_id,
                    distributor_goods_name=item.raw_name,
                    distributor_producer=item.raw_manufacturer,
                    product_id=int(item.product_id),
                    similarity_score=100,
                    match_method=f"sku:{matched_kind}",
                )
            else:
                saved_match = deferred_saved_match
                if saved_match is not None and int(saved_match.product_id) in product_ids:
                    saved_product_candidate = indexes.id_to_candidate.get(int(saved_match.product_id))
                    saved_source_structure = parse_drug_structure(item.raw_name)
                    if saved_product_candidate is not None:
                        structure_decision, structure_reason = _strict_structure_decision(saved_source_structure, saved_product_candidate.structure)
                        if structure_decision == "reject":
                            product_value, candidate_value = _structure_values_for_reason(saved_source_structure, saved_product_candidate.structure, structure_reason)
                            stats["structureRejects"] += 1
                            _verbose_debug(
                                "[MATCH_STRUCTURE_REJECT] product_sku=%s product_name=%s candidate_id=%s candidate_name=%s reason=%s product_value=%s candidate_value=%s",
                                raw_sku,
                                item.raw_name,
                                saved_product_candidate.id,
                                saved_product_candidate.name_norm,
                                structure_reason,
                                product_value,
                                candidate_value,
                            )
                            _bump_reason(stats, structure_reason)
                            stats["unmatched"] += 1
                            continue
                        if structure_decision == "suspicious":
                            stats["structureSuspicious"] += 1
                            stats["selectedSuspicious"] += 1
                    item.product_id = int(saved_match.product_id)
                    item.match_type = "saved_distributor_goods_id"
                    item.match_score = float(saved_match.similarity_score) if saved_match.similarity_score is not None else 100
                    item.matched_sku = raw_sku
                    stats["matched"] += 1
                    stats["matchedBySavedSku"] += 1
                    duplicate_bucket[(raw_sku, item.expiry_date or "no_exp")].append(item)
                    _upsert_source_goods_match(
                        db=db,
                        match_cache=source_matches,
                        price_format_id=price_list.price_format_id,
                        source_type="provisor",
                        distributor_goods_id=raw_sku,
                        goods_id=goods_id,
                        distributor_goods_name=item.raw_name,
                        distributor_producer=item.raw_manufacturer,
                        product_id=int(saved_match.product_id),
                        similarity_score=item.match_score,
                        match_method="saved_distributor_goods_id",
                    )
                    continue

                raw_full_name = str(raw_inner.get("goods", {}).get("fullName") if isinstance(raw_inner.get("goods"), dict) else "")
                logger.info(
                    "[PROVISOR_FALLBACK_MATCH] account_id=%s filial_id=%s goods_id=%s reason=no_goods_id_mapping",
                    price_list.account_id or "",
                    item.filial_id or raw_inner.get("filialId") or "",
                    goods_id or "",
                )
                primary_candidate = _vidman_candidate_result(
                    name_norm=item.raw_name,
                    mfr_norm=normalize_manufacturer_text(item.raw_manufacturer),
                    expiry_norm=item.expiry_date,
                    candidates=candidates,
                    exact=exact,
                    token_index=token_index,
                )
                record_candidate_stats(primary_candidate)
                fallback_candidate = None
                candidate_result = primary_candidate
                if not candidate_result["passesThreshold"] and raw_full_name and raw_full_name != item.raw_name:
                    fallback_candidate = _vidman_candidate_result(
                        name_norm=raw_full_name,
                        mfr_norm=normalize_manufacturer_text(item.raw_manufacturer),
                        expiry_norm=item.expiry_date,
                        candidates=candidates,
                        exact=exact,
                        token_index=token_index,
                    )
                    record_candidate_stats(fallback_candidate)
                    if fallback_candidate.get("passesThreshold") or fallback_candidate.get("score", 0) > candidate_result.get("score", 0):
                        candidate_result = fallback_candidate

                _merge_match_safety_stats(stats, candidate_result.get("safetyStats") or {})
                filtered_counts = candidate_result.get("filteredCounts") or {}
                stats["filteredByQuantity"] += int(filtered_counts.get("quantity") or 0)
                stats["filteredByVolume"] += int(filtered_counts.get("volume") or 0)
                stats["filteredByDosage"] += int(filtered_counts.get("dosage") or 0)
                stats["filteredByConcentration"] += int(filtered_counts.get("concentration") or 0)
                stats["filteredByForm"] += int(filtered_counts.get("form") or 0)
                examples = candidate_result.get("filteredExamples") or []
                if examples and len(stats["filterExamples"]) < 10:
                    for example in examples:
                        if len(stats["filterExamples"]) >= 10:
                            break
                        stats["filterExamples"].append({"sourceName": item.raw_name, **example})

                if candidate_result["passesThreshold"]:
                    candidate = _candidate_from_result(indexes, candidate_result)
                    if candidate is not None:
                        item.product_id = candidate.id
                        item.match_type = "fuzzy_name_manufacturer"
                        item.match_score = candidate_result["score"]
                        item.matched_sku = normalize_sku(candidate.sku) or candidate.sku
                        stats["matched"] += 1
                        stats["fuzzy"] += 1
                        stats["matchedByFuzzy"] += 1
                        duplicate_bucket[(item.matched_sku, item.expiry_date or "no_exp")].append(item)
                        _upsert_source_goods_match(
                            db=db,
                            match_cache=source_matches,
                            price_format_id=price_list.price_format_id,
                            source_type="provisor",
                            distributor_goods_id=raw_sku,
                            goods_id=goods_id,
                            distributor_goods_name=item.raw_name,
                            distributor_producer=item.raw_manufacturer,
                            product_id=int(candidate.id),
                            similarity_score=candidate_result["score"],
                            match_method="fuzzy_goods_full_name" if fallback_candidate is candidate_result else "fuzzy_distributor_name",
                        )
                    else:
                        _bump_reason(stats, "candidate_sku_not_found")
                        stats["unmatched"] += 1
                else:
                    _bump_reason(stats, candidate_result.get("reason"))
                    stats["unmatched"] += 1
            continue

        if source == "vidman":
            item.expiry_date = normalize_expiry(item.expiry_date or raw_inner.get("expiryDate") or raw_inner.get("expiry_date"))
            mfr_norm = normalize_manufacturer_text(item.raw_manufacturer)
            candidate_result = _vidman_candidate_result(
                name_norm=item.raw_name,
                mfr_norm=mfr_norm,
                expiry_norm=item.expiry_date,
                candidates=candidates,
                exact=exact,
                token_index=token_index,
            )
            record_candidate_stats(candidate_result)
            _merge_match_safety_stats(stats, candidate_result.get("safetyStats") or {})
            filtered_counts = candidate_result.get("filteredCounts") or {}
            stats["filteredByQuantity"] += int(filtered_counts.get("quantity") or 0)
            stats["filteredByVolume"] += int(filtered_counts.get("volume") or 0)
            stats["filteredByDosage"] += int(filtered_counts.get("dosage") or 0)
            stats["filteredByConcentration"] += int(filtered_counts.get("concentration") or 0)
            stats["filteredByForm"] += int(filtered_counts.get("form") or 0)
            examples = candidate_result.get("filteredExamples") or []
            if examples and len(stats["filterExamples"]) < 10:
                for example in examples:
                    if len(stats["filterExamples"]) >= 10:
                        break
                    stats["filterExamples"].append(
                        {
                            "sourceName": item.raw_name,
                            **example,
                        }
                    )
            if candidate_result["passesThreshold"]:
                candidate = _candidate_from_result(indexes, candidate_result)
                if candidate is None:
                    _bump_reason(stats, "candidate_sku_not_found")
                    stats["unmatched"] += 1
                    continue
                item.product_id = candidate.id
                item.match_type = "fuzzy_name_manufacturer"
                item.match_score = candidate_result["score"]
                item.matched_sku = normalize_sku(candidate.sku) or candidate.sku
                stats["matched"] += 1
                stats["fuzzy"] += 1
            else:
                _bump_reason(stats, candidate_result.get("reason"))
                stats["unmatched"] += 1
            continue

        raw_sku = str(item.distributor_goods_id or "").strip()
        normalized = next((v for v in normalize_sku_variants(raw_sku) if v in code_to_id), None) if raw_sku else None
        item.match_key = normalized or ""
        if normalized and normalized in code_to_id:
            exact_product_candidate = indexes.id_to_candidate.get(int(code_to_id[normalized]))
            exact_source_structure = parse_drug_structure(item.raw_name)
            if exact_product_candidate is not None:
                structure_decision, structure_reason = _strict_structure_decision(exact_source_structure, exact_product_candidate.structure)
                if structure_decision == "reject":
                    product_value, candidate_value = _structure_values_for_reason(exact_source_structure, exact_product_candidate.structure, structure_reason)
                    stats["structureRejects"] += 1
                    _verbose_debug(
                        "[MATCH_STRUCTURE_REJECT] product_sku=%s product_name=%s candidate_id=%s candidate_name=%s reason=%s product_value=%s candidate_value=%s",
                        normalized,
                        item.raw_name,
                        exact_product_candidate.id,
                        exact_product_candidate.name_norm,
                        structure_reason,
                        product_value,
                        candidate_value,
                    )
                    _bump_reason(stats, structure_reason)
                    stats["unmatched"] += 1
                    continue
                if structure_decision == "suspicious":
                    stats["structureSuspicious"] += 1
                    _verbose_debug(
                        "[MATCH_STRUCTURE_SUSPICIOUS] product_sku=%s candidate_id=%s reason=%s name_score=%s manufacturer_decision=%s",
                        normalized,
                        exact_product_candidate.id,
                        structure_reason,
                        100,
                        "exact",
                    )
                    stats["selectedSuspicious"] += 1
            item.product_id = code_to_id[normalized]
            item.match_type = "sku"
            item.match_score = 100
            item.matched_sku = normalized
            stats["matched"] += 1
        else:
            _bump_reason(stats, "no_sku_match")
            stats["unmatched"] += 1

    if not _keeps_multi_prices_per_sku(price_list):
        for rows in duplicate_bucket.values():
            if len(rows) <= 1:
                continue
            prices = [_price_decimal(x.distributor_price) for x in rows]
            stocks = [_price_decimal(x.stock) or Decimal("0") for x in rows]
            min_price = min([x for x in prices if x is not None], default=None)
            stock_sum = sum(stocks, Decimal("0"))
            keeper = rows[0]
            keeper.distributor_price = float(min_price) if min_price is not None else keeper.distributor_price
            keeper.stock = float(stock_sum)
            for duplicate in rows[1:]:
                duplicate.product_id = None
                duplicate.match_type = "unmatched"
                duplicate.match_score = None
                duplicate.matched_sku = ""
    matched_at = time.perf_counter()

    actual_matched = sum(1 for item in items if item.product_id is not None and item.matched_sku)
    actual_unmatched = len(items) - actual_matched
    stats["matched"] = actual_matched
    stats["unmatched"] = actual_unmatched
    stats["matchRate"] = round((actual_matched / len(items) * 100), 2) if items else 0
    stats["candidatePoolAvg"] = (
        round(int(stats["candidatePoolTotal"]) / int(stats["candidatePoolCount"]), 2)
        if int(stats["candidatePoolCount"])
        else 0
    )
    matched_rows_for_prices = [
        _matched_price_row(item)
        for item in items
        if item.product_id is not None and item.matched_sku and item.distributor_price is not None
    ]
    stats["_matchedRows"] = matched_rows_for_prices
    if debug:
        stats["topUnmatched"] = _unmatched_debug(
            price_list=price_list,
            items=items,
            code_to_id=code_to_id,
            candidates=candidates,
            exact=exact,
            token_index=token_index,
            limit=50,
        )
    else:
        stats["topUnmatched"] = []
    debugged_at = time.perf_counter()
    stats["timingMs"] = {
        "loadProductsAndIndexes": round((indexed_at - started_at) * 1000, 2),
        "loadItems": round((loaded_at - indexed_at) * 1000, 2),
        "matching": round((matched_at - loaded_at) * 1000, 2),
        "diagnostics": round((debugged_at - matched_at) * 1000, 2),
    }

    logger.info(
        (
            "[MATCH] source=%s price_list_id=%s incoming=%s matched=%s fuzzy=%s unmatched=%s "
            "match_rate=%s%% pool_avg=%s pool_max=%s broad_fallbacks=%s "
            "filtered quantity=%s volume=%s dosage=%s concentration=%s form=%s examples=%s top_unmatched=%s"
        ),
        price_list.source_type.capitalize(),
        price_list.id,
        stats["incoming"],
        stats["matched"],
        stats["fuzzy"],
        stats["unmatched"],
        stats["matchRate"],
        stats["candidatePoolAvg"],
        stats["candidatePoolMax"],
        stats["candidatePoolBroadFallbacks"],
        stats["filteredByQuantity"],
        stats["filteredByVolume"],
        stats["filteredByDosage"],
        stats["filteredByConcentration"],
        stats["filteredByForm"],
        stats["filterExamples"][:3],
        stats["topUnmatched"][:3],
    )
    logger.info(
        "[MATCH_SAFETY] structure_rejects=%s structure_suspicious=%s selected_suspicious=%s manufacturer_hard_rejects=%s fuzzy_rejected_low_confidence=%s",
        stats.get("structureRejects", 0),
        stats.get("structureSuspicious", 0),
        stats.get("selectedSuspicious", 0),
        stats.get("manufacturerHardRejectsSafety", 0),
        stats.get("fuzzyRejectedLowConfidence", 0),
    )
    if flush:
        db.flush()
    flushed_at = time.perf_counter()
    stats["timingMs"]["flush"] = round((flushed_at - debugged_at) * 1000, 2)
    _verbose_debug("%s matching timing price_list_id=%s timing_ms=%s", price_list.source_type.capitalize(), price_list.id, stats["timingMs"])
    return stats


def _product_sku_keys(value: object) -> list[str]:
    keys: list[str] = []

    def add(key: str | None) -> None:
        if key and key not in keys:
            keys.append(key)

    add(normalize_external_sku(value))
    add(normalize_sku(value))
    for variant in normalize_sku_variants(value):
        add(variant)
    return keys


def _supplier_name_for_matching(item: CompetitorPriceListItem, raw_inner: dict) -> str:
    goods = raw_inner.get("goods") if isinstance(raw_inner.get("goods"), dict) else {}
    return str(item.raw_name or item.name or item.distributor_goods_name or goods.get("fullName") or "").strip()


def _build_supplier_indexes(
    *,
    price_list: CompetitorPriceList,
    items: list[CompetitorPriceListItem],
) -> SupplierIndexes:
    supplier_candidates: list[SupplierItemCandidate] = []
    token_index: dict[str, list[SupplierItemCandidate]] = defaultdict(list)
    sku_index: dict[str, list[SupplierSkuCandidate]] = defaultdict(list)
    composite_index: dict[str, list[SupplierItemCandidate]] = defaultdict(list)

    for item in items:
        raw = _raw_payload(item)
        raw_inner = raw.get("raw") if isinstance(raw.get("raw"), dict) else raw
        item.raw_name = _supplier_name_for_matching(item, raw_inner)
        item.raw_manufacturer = str(
            raw.get("manufacturer")
            or raw_inner.get("distributorProducer")
            or raw_inner.get("manufacturer")
            or item.raw_manufacturer
            or ""
        )
        item.expiry_date = normalize_expiry(item.expiry_date or raw_inner.get("shelfLife") or raw_inner.get("expiryDate") or raw_inner.get("expiry_date"))
        item.match_type = "unmatched"
        item.match_score = None
        item.matched_sku = ""
        item.match_key = ""
        item.product_id = None

        structure = _cached_structure_from_item(item)
        name_norm = item.normalized_name or normalize_text(normalize_drug_text(item.raw_name))
        manufacturer_norm = item.normalized_manufacturer or normalize_manufacturer_text(item.raw_manufacturer)
        base_name_norm = item.parsed_base_name or normalize_text(structure.base_name) or name_norm
        cand = SupplierItemCandidate(
            item=item,
            name_norm=name_norm,
            base_name_norm=base_name_norm,
            manufacturer_norm=manufacturer_norm,
            tokens=_tokens(base_name_norm),
            structure=structure,
        )
        supplier_candidates.append(cand)

        first_word = base_name_norm.split(" ", 1)[0]
        _add_index(token_index, f"first:{first_word}" if first_word else None, cand)
        _add_index(token_index, f"prefix:{base_name_norm[:5]}" if base_name_norm else None, cand)
        _add_index(token_index, f"prefix:{base_name_norm[:3]}" if base_name_norm else None, cand)
        _add_index(token_index, f"base:{base_name_norm}" if base_name_norm else None, cand)
        _add_index(token_index, f"manufacturer:{cand.manufacturer_norm}" if cand.manufacturer_norm else None, cand)
        for key in _composite_index_keys(base_name_norm, structure):
            _add_index(token_index, key, cand)
        for key in _supplier_composite_keys(base_name_norm, cand.manufacturer_norm, structure):
            composite_index[key].append(cand)
        if cand.manufacturer_norm:
            for key in _supplier_composite_keys(base_name_norm, "", structure):
                composite_index[key].append(cand)
        for token in cand.tokens:
            token_index[token].append(cand)

        if price_list.source_type == "provisor":
            raw_sku = normalize_external_sku(item.distributor_goods_id)
            goods_id = item.provisor_goods_id or raw_inner.get("goodsId")
            filial_id = item.filial_id or raw_inner.get("filialId")
            for kind, key in _provisor_supplier_index_keys(
                goods_id=goods_id,
                distributor_goods_id=raw_sku,
                expiry_date=item.expiry_date,
                filial_id=filial_id,
            ):
                if key:
                    sku_index[key].append(SupplierSkuCandidate(candidate=cand, kind=kind))
        else:
            for key in _product_sku_keys(item.distributor_goods_id):
                sku_index[key].append(SupplierSkuCandidate(candidate=cand, kind="sku"))

    return SupplierIndexes(
        candidates=supplier_candidates,
        token_index=token_index,
        sku_index=sku_index,
        composite_index=composite_index,
    )


def _supplier_pool_for_product(
    *,
    product: ProductCandidate,
    token_index: dict[str, list[SupplierItemCandidate]],
    max_pool: int = 20,
) -> tuple[list[SupplierItemCandidate], int]:
    pool: dict[int, SupplierItemCandidate] = {}
    base = product.base_name_norm or product.name_norm
    first_word = base.split(" ", 1)[0]
    name_keys = [
        f"base:{base}" if base else None,
        f"first:{first_word}" if first_word else None,
        f"prefix:{base[:5]}" if base else None,
        f"manufacturer:{product.manufacturer_norm}" if product.manufacturer_norm else None,
    ]
    for key in name_keys:
        if not key:
            continue
        for cand in token_index.get(key, []):
            pool[int(cand.item.id)] = cand
    if not pool:
        for token in product.tokens:
            for cand in token_index.get(token, []):
                pool[int(cand.item.id)] = cand
                if len(pool) >= max_pool * 3:
                    break
            if len(pool) >= max_pool * 3:
                break
    if not pool and product.structure.critical_tokens:
        for token in product.structure.critical_tokens:
            for alias in (token, *TOKEN_ALIASES.get(token, set())):
                for cand in token_index.get(alias, []):
                    pool[int(cand.item.id)] = cand
                    if len(pool) >= max_pool * 3:
                        break
                if len(pool) >= max_pool * 3:
                    break
            if len(pool) >= max_pool * 3:
                break
    if not pool:
        return [], 0
    filtered: list[SupplierItemCandidate] = []
    for cand in pool.values():
        if product.manufacturer_norm and cand.manufacturer_norm and product.manufacturer_norm != cand.manufacturer_norm:
            name_score = _base_name_similarity(base, cand.base_name_norm or cand.name_norm)
            if name_score < 90:
                continue
        if product.structure.quantity is not None and cand.structure.quantity is not None and product.structure.quantity != cand.structure.quantity:
            continue
        if product.structure.dosage is not None and cand.structure.dosage is not None and product.structure.dosage != cand.structure.dosage:
            continue
        filtered.append(cand)
    pool_size_before_limit = len(filtered)
    ranked = sorted(
        filtered,
        key=lambda cand: (
            product.manufacturer_norm == cand.manufacturer_norm,
            _base_name_similarity(base, cand.base_name_norm or cand.name_norm),
            -float(_price_decimal(cand.item.distributor_price) or Decimal("999999999")),
        ),
        reverse=True,
    )
    return ranked[:max_pool], pool_size_before_limit


def _structure_audit_payload(structure: DrugStructure) -> dict[str, object]:
    return {
        "base_name": structure.base_name,
        "dosage": structure.dosage,
        "dosage_volume": structure.dosage_volume,
        "strength_signature": structure.strength_signature,
        "quantity": structure.quantity,
        "volume": structure.volume,
        "weight": structure.weight,
        "form": structure.form,
        "forms": structure.forms,
        "dimensions": structure.dimensions,
        "critical_tokens": structure.critical_tokens,
    }


def _structure_audit_text(structure: DrugStructure) -> str:
    payload = _structure_audit_payload(structure)
    return ", ".join(f"{key}={value}" for key, value in payload.items())


def _missing_candidate_structure_fields(product_struct: DrugStructure, candidate_struct: DrugStructure) -> list[str]:
    fields = ("dosage", "quantity", "volume", "weight", "form")
    return [field for field in fields if getattr(product_struct, field) is not None and getattr(candidate_struct, field) is None]


def _audit_risk_reasons(
    *,
    match_type: str,
    name_score: float,
    base_score: float,
    final_score: float,
    manufacturer_decision: str,
    structure_decision: str,
    product_struct: DrugStructure,
    candidate_struct: DrugStructure,
    top_candidates: list[dict],
) -> list[str]:
    reasons: list[str] = []
    top1_top2_close = False
    if len(top_candidates) >= 2:
        top1 = float(top_candidates[0].get("score") or top_candidates[0].get("final_score") or 0)
        top2 = float(top_candidates[1].get("score") or top_candidates[1].get("final_score") or 0)
        top1_top2_close = abs(top1 - top2) < 5
    if (
        name_score >= 98
        and manufacturer_decision in {"alias_match", "similar_match"}
        and structure_decision == "ok"
        and not top1_top2_close
    ):
        return reasons
    if manufacturer_decision not in {"alias_match", "similar_match"}:
        reasons.append(f"manufacturer_{manufacturer_decision or 'unknown'}")
    if structure_decision != "ok":
        reasons.append(f"structure_{structure_decision}")
    if name_score < 95:
        reasons.append("name_score_lt_95")
    missing_fields = _missing_candidate_structure_fields(product_struct, candidate_struct)
    if missing_fields:
        reasons.append("missing_candidate_" + ",".join(missing_fields))
    if final_score < 95:
        reasons.append("final_score_lt_95")
    if base_score < 100:
        reasons.append("base_name_differs")
    if manufacturer_decision in {"missing", "penalty", "manufacturer_reject"}:
        reason = f"manufacturer_{manufacturer_decision}"
        if reason not in reasons:
            reasons.append(reason)
    if top1_top2_close:
        reasons.append("top1_top2_close")
    return reasons


def _audit_sort_key(row: dict) -> tuple[int, int, int, int, float]:
    reason = str(row.get("risk_reason") or "")
    return (
        0 if "manufacturer_penalty" in reason or "manufacturer_missing" in reason else 1,
        0 if str(row.get("structure_decision") or "") == "suspicious" else 1,
        0 if float(row.get("name_score") or 0) < 95 else 1,
        0 if "top1_top2_close" in reason else 1,
        float(row.get("final_score") or 0),
    )


def _write_matching_audit_xlsx(
    *,
    path: str,
    selected_rows: list[dict],
    unmatched_rows: list[dict],
    manufacturer_rows: list[dict],
    summary: dict[str, object],
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "selected_risky"
    columns = [
        "product_sku",
        "product_name",
        "product_manufacturer",
        "product_structure",
        "candidate_name",
        "candidate_manufacturer",
        "candidate_structure",
        "product_parsed_base_name",
        "product_parsed_forms",
        "product_parsed_dosage",
        "product_parsed_volume",
        "candidate_parsed_base_name",
        "candidate_parsed_forms",
        "candidate_parsed_dosage",
        "candidate_parsed_volume",
        "price",
        "source_price_list",
        "match_type",
        "name_score",
        "base_score",
        "manufacturer_decision",
        "manufacturer_score",
        "structure_decision",
        "structure_reason",
        "final_score",
        "risk_reason",
        "candidate_pool_size",
        "top_3_candidates",
    ]
    ws.append(columns)
    for row in sorted(selected_rows, key=_audit_sort_key)[:300]:
        ws.append([row.get(column, "") for column in columns])
    ws.freeze_panes = "A2"
    for column_cells in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_len + 2, 10), 80)

    unmatched_ws = wb.create_sheet("unmatched_candidates")
    unmatched_columns = [
        "product_sku",
        "product_name",
        "product_manufacturer",
        "best_candidate_name",
        "best_candidate_manufacturer",
        "name_score",
        "manufacturer_decision",
        "structure_decision",
        "reject_reason",
        "candidate_pool_size",
        "top_3_candidates",
    ]
    unmatched_ws.append(unmatched_columns)
    for row in unmatched_rows[:300]:
        unmatched_ws.append([row.get(column, "") for column in unmatched_columns])
    unmatched_ws.freeze_panes = "A2"
    for column_cells in unmatched_ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in column_cells)
        unmatched_ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_len + 2, 10), 80)

    manufacturer_ws = wb.create_sheet("manufacturer_alias_candidates")
    manufacturer_columns = [
        "product_manufacturer",
        "candidate_manufacturer",
        "normalized_product_manufacturer",
        "normalized_candidate_manufacturer",
        "product_name",
        "candidate_name",
        "score",
        "action",
    ]
    manufacturer_ws.append(manufacturer_columns)
    for row in manufacturer_rows[:300]:
        manufacturer_ws.append([row.get(column, "") for column in manufacturer_columns])
    manufacturer_ws.freeze_panes = "A2"
    for column_cells in manufacturer_ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in column_cells)
        manufacturer_ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_len + 2, 10), 80)

    summary_ws = wb.create_sheet("summary")
    summary_ws.append(["metric", "value"])
    for key, value in summary.items():
        summary_ws.append([key, value])
    summary_ws.column_dimensions["A"].width = 36
    summary_ws.column_dimensions["B"].width = 18
    wb.save(path)


def _manufacturer_alias_candidate_row(
    *,
    product: ProductCandidate,
    supplier: SupplierItemCandidate,
    score: float,
) -> dict:
    return {
        "product_manufacturer": product.manufacturer_norm,
        "candidate_manufacturer": supplier.manufacturer_norm,
        "normalized_product_manufacturer": normalize_manufacturer_text(product.manufacturer_norm),
        "normalized_candidate_manufacturer": normalize_manufacturer_text(supplier.manufacturer_norm),
        "product_name": product.raw_name or product.name_norm,
        "candidate_name": supplier.item.raw_name or supplier.item.name or supplier.item.distributor_goods_name or "",
        "score": round(score, 2),
        "action": "add_alias_or_confirm_reject",
    }


def _find_exact_composite_supplier(
    *,
    product: ProductCandidate,
    composite_index: dict[str, list[SupplierItemCandidate]],
    used_item_ids: set[int],
) -> tuple[SupplierItemCandidate | None, dict]:
    candidates: dict[int, SupplierItemCandidate] = {}
    manufacturer_keys = [product.manufacturer_norm]
    if product.manufacturer_norm:
        manufacturer_keys.append("")
    for manufacturer_norm in manufacturer_keys:
        for key in _supplier_composite_keys(product.base_name_norm or product.name_norm, manufacturer_norm, product.structure):
            for cand in composite_index.get(key, []):
                if int(cand.item.id) not in used_item_ids:
                    candidates[int(cand.item.id)] = cand
    diagnostics = {
        "manufacturerRejects": 0,
        "aliasCandidates": [],
        "topCandidates": [],
    }
    accepted: list[SupplierItemCandidate] = []
    for cand in candidates.values():
        structure_decision, _structure_reason = _strict_structure_decision(product.structure, cand.structure)
        name_score = _base_name_similarity(product.base_name_norm or product.name_norm, cand.base_name_norm or cand.name_norm)
        manufacturer_decision, mfr_score, manufacturer_ok = _manufacturer_required_match(product.manufacturer_norm, cand.manufacturer_norm)
        diagnostics["topCandidates"].append(
            {
                "candidate_id": int(cand.item.id),
                "name": cand.item.raw_name or cand.item.name or cand.item.distributor_goods_name or "",
                "manufacturer": cand.manufacturer_norm,
                "name_score": round(name_score, 2),
                "base_score": round(name_score, 2),
                "manufacturer_decision": manufacturer_decision,
                "manufacturer_score": round(mfr_score, 2),
                "structure_decision": structure_decision,
                "reject_reason": "" if structure_decision == "ok" and (manufacturer_ok or manufacturer_decision == "missing") else manufacturer_decision,
                "final_score": round(name_score, 2),
            }
        )
        if structure_decision != "ok":
            continue
        if manufacturer_decision in {"manufacturer_reject", "penalty", "hard_reject"}:
            diagnostics["manufacturerRejects"] += 1
            if name_score >= 97:
                diagnostics["aliasCandidates"].append(_manufacturer_alias_candidate_row(product=product, supplier=cand, score=name_score))
            continue
        if manufacturer_decision == "missing" or manufacturer_ok:
            accepted.append(cand)
    if not accepted:
        diagnostics["topCandidates"] = sorted(
            diagnostics["topCandidates"],
            key=lambda row: float(row.get("final_score") or 0),
            reverse=True,
        )[:3]
        return None, diagnostics
    selected = min(accepted, key=lambda cand: float(_price_decimal(cand.item.distributor_price) or Decimal("999999999")))
    diagnostics["topCandidates"] = sorted(
        diagnostics["topCandidates"],
        key=lambda row: (row.get("reject_reason") == "", float(row.get("final_score") or 0)),
        reverse=True,
    )[:3]
    return selected, diagnostics


def _best_supplier_for_product(
    *,
    product: ProductCandidate,
    pool: list[SupplierItemCandidate],
    used_item_ids: set[int],
    source_type: str = "",
) -> tuple[SupplierItemCandidate | None, float, str, dict]:
    best: tuple[tuple[int, float, float, float], SupplierItemCandidate, float, float, int, int, int, str, str] | None = None
    filtered_counts = {field: 0 for field in STRUCTURED_FIELDS}
    reject_counts = _empty_match_reject_counts()
    manufacturer_counts = {
        "alias_match": 0,
        "similar_match": 0,
        "penalty": 0,
        "hard_reject": 0,
        "manufacturer_reject": 0,
    }
    safety_stats = {
        "structure_rejects": 0,
        "structure_suspicious": 0,
        "selected_suspicious": 0,
        "manufacturer_hard_rejects": 0,
        "fuzzy_rejected_low_confidence": 0,
    }
    best_rejected: tuple[float, str] | None = None
    manufacturer_alias_candidates: list[dict] = []
    product_base = product.base_name_norm or product.name_norm
    initial_pool = len(pool)
    include_partial = initial_pool <= PARTIAL_RATIO_POOL_LIMIT
    candidate_pool_for_scoring: list[SupplierItemCandidate] = []
    for supplier in pool:
        if int(supplier.item.id) in used_item_ids:
            continue
        candidate_pool_for_scoring.append(supplier)

    after_name_filter = 0

    after_strict_filter = 0
    after_manufacturer_priority = 0
    final_candidates: list[dict] = []
    rejected_candidates: list[dict] = []

    def record_rejected_candidate(
        supplier: SupplierItemCandidate,
        *,
        name_score: float,
        base_score: float,
        manufacturer_decision: str = "",
        structure_decision: str = "",
        reject_reason: str,
        final_score: float | None = None,
    ) -> None:
        rejected_candidates.append(
            {
                "candidate_id": int(supplier.item.id),
                "name": supplier.item.raw_name or supplier.item.name or supplier.item.distributor_goods_name or "",
                "manufacturer": supplier.manufacturer_norm,
                "name_score": round(float(name_score or 0), 2),
                "base_score": round(float(base_score or 0), 2),
                "manufacturer_decision": manufacturer_decision,
                "structure_decision": structure_decision,
                "reject_reason": reject_reason,
                "final_score": round(float(final_score if final_score is not None else name_score or 0), 2),
            }
        )

    for supplier in candidate_pool_for_scoring[:MAX_CANDIDATE_POOL]:
        supplier_base = supplier.base_name_norm or supplier.name_norm
        _similarity_payload = _similarity(product_base, supplier_base, include_partial=include_partial)
        name_score = _similarity_payload["score"]
        base_score = _base_name_similarity(product_base, supplier_base)
        structure_decision, structure_reason = _strict_structure_decision(product.structure, supplier.structure)
        if structure_decision == "reject":
            field = _structure_reason_field(structure_reason) or "structure"
            product_value, supplier_value = _structure_values_for_reason(product.structure, supplier.structure, structure_reason)
            filtered_counts[field] += 1
            detailed_reason = _reject_reason_for_structure_field(field)
            if detailed_reason:
                reject_counts[detailed_reason] += 1
            safety_stats["structure_rejects"] += 1
            reason = structure_reason or _conflict_reason(field)
            if best_rejected is None:
                best_rejected = (0.0, reason)
            record_rejected_candidate(
                supplier,
                name_score=name_score,
                base_score=base_score,
                structure_decision=structure_decision,
                reject_reason=reason,
            )
            _verbose_debug(
                "[MATCH_STRUCTURE_REJECT] product_sku=%s product_name=%s candidate_id=%s candidate_name=%s reason=%s product_value=%s candidate_value=%s",
                product.sku,
                product.name_norm,
                supplier.item.id,
                supplier.item.raw_name or supplier.item.name,
                reason,
                product_value,
                supplier_value,
            )
            _verbose_debug(
                "[MATCH_BALANCED] product_sku=%s product_name=%s candidate_name=%s structure_decision=%s name_score=%s manufacturer_decision=%s final_decision=%s",
                product.sku,
                product.name_norm,
                supplier.item.raw_name or supplier.item.name,
                structure_decision,
                None,
                "",
                "reject",
            )
            continue

        after_strict_filter += 1
        manufacturer_decision, mfr_score, manufacturer_allowed = _manufacturer_required_match(
            product.manufacturer_norm,
            supplier.manufacturer_norm,
        )
        if structure_decision == "suspicious":
            safety_stats["structure_suspicious"] += 1
            _verbose_debug(
                "[MATCH_STRUCTURE_SUSPICIOUS] product_sku=%s candidate_id=%s reason=%s name_score=%s manufacturer_decision=%s",
                product.sku,
                supplier.item.id,
                structure_reason,
                None,
                manufacturer_decision,
            )
        if manufacturer_decision in manufacturer_counts:
            manufacturer_counts[manufacturer_decision] += 1
        if not manufacturer_allowed:
            reject_counts["manufacturer_hard_reject"] += 1
            safety_stats["manufacturer_hard_rejects"] += 1
            if best_rejected is None:
                best_rejected = (0.0, "manufacturer_conflict")
            if structure_decision == "ok" and name_score >= 97 and manufacturer_decision in {"manufacturer_reject", "hard_reject"}:
                manufacturer_alias_candidates.append(_manufacturer_alias_candidate_row(product=product, supplier=supplier, score=name_score))
            record_rejected_candidate(
                supplier,
                name_score=name_score,
                base_score=base_score,
                manufacturer_decision=manufacturer_decision,
                structure_decision=structure_decision,
                reject_reason="manufacturer_conflict",
            )
            _verbose_debug(
                "[MATCH_MANUFACTURER] source=%s product_sku=%s candidate_id=%s name_score=%s manufacturer_score=%s decision=%s product_manufacturer=%s candidate_manufacturer=%s",
                source_type,
                product.sku,
                supplier.item.id,
                None,
                round(mfr_score, 2),
                manufacturer_decision,
                product.manufacturer_norm,
                supplier.manufacturer_norm,
            )
            continue
        after_manufacturer_priority += 1
        if name_score < NAME_PREFILTER_THRESHOLD:
            record_rejected_candidate(
                supplier,
                name_score=name_score,
                base_score=base_score,
                manufacturer_decision=manufacturer_decision,
                structure_decision=structure_decision,
                reject_reason="name_prefilter",
            )
            continue
        after_name_filter += 1
        if base_score < 100 and (name_score < 97 or _first_token(product_base) != _first_token(supplier_base)):
            safety_stats["fuzzy_rejected_low_confidence"] += 1
            if best_rejected is None or name_score > best_rejected[0]:
                best_rejected = (name_score, "unsafe_name_fuzzy")
            record_rejected_candidate(
                supplier,
                name_score=name_score,
                base_score=base_score,
                manufacturer_decision=manufacturer_decision,
                structure_decision=structure_decision,
                reject_reason="unsafe_name_fuzzy",
            )
            continue
        if manufacturer_decision == "missing" and (structure_decision != "ok" or name_score < 97):
            safety_stats["fuzzy_rejected_low_confidence"] += 1
            if best_rejected is None or name_score > best_rejected[0]:
                best_rejected = (name_score, "manufacturer_missing")
            record_rejected_candidate(
                supplier,
                name_score=name_score,
                base_score=base_score,
                manufacturer_decision=manufacturer_decision,
                structure_decision=structure_decision,
                reject_reason="manufacturer_missing",
            )
            continue
        if structure_decision == "ok" and name_score < 88:
            safety_stats["fuzzy_rejected_low_confidence"] += 1
            if best_rejected is None or name_score > best_rejected[0]:
                best_rejected = (name_score, "low_score")
            record_rejected_candidate(
                supplier,
                name_score=name_score,
                base_score=base_score,
                manufacturer_decision=manufacturer_decision,
                structure_decision=structure_decision,
                reject_reason="low_score",
            )
            continue
        if structure_decision == "suspicious" and not (
            name_score >= 95
            and base_score >= 95
            and manufacturer_decision in {"alias_match", "similar_match", "missing"}
        ):
            safety_stats["fuzzy_rejected_low_confidence"] += 1
            if best_rejected is None or name_score > best_rejected[0]:
                best_rejected = (name_score, structure_reason or "structure_suspicious")
            record_rejected_candidate(
                supplier,
                name_score=name_score,
                base_score=base_score,
                manufacturer_decision=manufacturer_decision,
                structure_decision=structure_decision,
                reject_reason=structure_reason or "structure_suspicious",
            )
            continue
        combined = (
            name_score
            if not product.manufacturer_norm or not supplier.manufacturer_norm
            else min(100.0, name_score + 15)
            if manufacturer_decision == "alias_match"
            else min(100.0, name_score + 10)
            if manufacturer_decision == "similar_match"
            else name_score - (15 if _structures_match_strongly(product.structure, supplier.structure) else 25)
            if manufacturer_decision == "penalty"
            else name_score
        )
        if manufacturer_decision == "penalty":
            reject_counts["manufacturer_penalty"] += 1
        matched_params, missing_params = _param_counts(product.structure, supplier.structure)
        threshold = 95 if structure_decision == "suspicious" else 88
        combined -= min(6, missing_params * 2)
        supplier_price = float(_price_decimal(supplier.item.distributor_price) or Decimal("999999999"))
        manufacturer_rank = 2 if manufacturer_decision == "alias_match" else 1 if manufacturer_decision == "similar_match" else 0
        rank = (matched_params, combined, name_score, manufacturer_rank, -supplier_price)
        final_candidates.append(
            {
                "candidate_id": int(supplier.item.id),
                "name": supplier.item.raw_name or supplier.item.name or supplier.item.distributor_goods_name or "",
                "manufacturer": supplier.manufacturer_norm,
                "name_score": round(name_score, 2),
                "base_score": round(base_score, 2),
                "manufacturer_score": round(mfr_score, 2),
                "manufacturer_decision": manufacturer_decision,
                "structure_decision": structure_decision,
                "score": round(combined, 2),
                "final_score": round(combined, 2),
                "reject_reason": "",
                "matched_params": matched_params,
            }
        )
        if best is None or rank > best[0]:
            best = (rank, supplier, combined, name_score, matched_params, missing_params, threshold, structure_decision, manufacturer_decision)
    top_final_candidates = sorted(
        final_candidates,
        key=lambda row: (int(row.get("matched_params") or 0), float(row.get("score") or 0), float(row.get("name_score") or 0)),
        reverse=True,
    )[:3]
    rejected_top_candidates = sorted(
        rejected_candidates,
        key=lambda row: (float(row.get("final_score") or 0), float(row.get("name_score") or 0), float(row.get("base_score") or 0)),
        reverse=True,
    )[:3]
    selected_id = int(best[1].item.id) if best is not None else None
    _verbose_debug(
        "[MATCH_POOL] source=%s product_sku=%s initial_pool=%s after_name_filter=%s after_strict_filter=%s after_manufacturer_priority=%s selected=%s top3=%s",
        source_type,
        product.sku,
        initial_pool,
        after_name_filter,
        after_strict_filter,
        after_manufacturer_priority,
        selected_id,
        json.dumps(top_final_candidates, ensure_ascii=False),
    )
    if best is None:
        reason = best_rejected[1] if best_rejected else "no_candidate"
        return None, 0, reason, {
            "filteredCounts": filtered_counts,
            "rejectCounts": reject_counts,
            "manufacturerCounts": manufacturer_counts,
            "safetyStats": safety_stats,
            "topCandidates": top_final_candidates or rejected_top_candidates,
            "rejectedTopCandidates": rejected_top_candidates,
            "manufacturerAliasCandidates": manufacturer_alias_candidates,
            "poolStats": {
                "initialPool": initial_pool,
                "afterNameFilter": after_name_filter,
                "afterStrictFilter": after_strict_filter,
                "afterManufacturerPriority": after_manufacturer_priority,
            },
        }
    _rank, supplier, combined, name_score, _matched_params, _missing_params, threshold, structure_decision, selected_manufacturer_decision = best
    _selected_decision_check, selected_mfr_score, _selected_hard_conflict = _manufacturer_decision(product.manufacturer_norm, supplier.manufacturer_norm)
    if name_score < threshold or combined < threshold:
        safety_stats["fuzzy_rejected_low_confidence"] += 1
        low_score_top_candidates = [
            {
                **row,
                "reject_reason": row.get("reject_reason") or "low_score",
                "final_score": row.get("final_score", row.get("score", 0)),
            }
            for row in top_final_candidates
        ]
        return supplier, round(combined, 2), "low_score", {
            "filteredCounts": filtered_counts,
            "rejectCounts": reject_counts,
            "manufacturerCounts": manufacturer_counts,
            "safetyStats": safety_stats,
            "threshold": threshold,
            "topCandidates": low_score_top_candidates or rejected_top_candidates,
            "rejectedTopCandidates": low_score_top_candidates or rejected_top_candidates,
            "manufacturerAliasCandidates": manufacturer_alias_candidates,
            "poolStats": {
                "initialPool": initial_pool,
                "afterNameFilter": after_name_filter,
                "afterStrictFilter": after_strict_filter,
                "afterManufacturerPriority": after_manufacturer_priority,
            },
        }
    if product.structure.raw_dosage is not None or supplier.structure.raw_dosage is not None:
        _verbose_debug(
            "[MATCH_DOSAGE_NORMALIZED] raw_product_dosage=%s normalized_product_dosage=%s raw_candidate_dosage=%s normalized_candidate_dosage=%s",
            product.structure.raw_dosage,
            product.structure.dosage,
            supplier.structure.raw_dosage,
            supplier.structure.dosage,
        )
    _verbose_debug(
        "[MATCH_SELECTED] source=%s product_sku=%s candidate_id=%s score=%s name_score=%s manufacturer_score=%s manufacturer_decision=%s structure_decision=%s match_type=%s candidate_name=%s",
        source_type,
        product.sku,
        supplier.item.id,
        round(combined, 2),
        round(name_score, 2),
        round(selected_mfr_score, 2),
        selected_manufacturer_decision,
        structure_decision,
        "fuzzy_name_manufacturer",
        supplier.item.raw_name or supplier.item.name,
    )
    _verbose_debug(
        "[MATCH_BALANCED] product_sku=%s product_name=%s candidate_name=%s structure_decision=%s name_score=%s manufacturer_decision=%s final_decision=%s",
        product.sku,
        product.name_norm,
        supplier.item.raw_name or supplier.item.name,
        structure_decision,
        round(name_score, 2),
        selected_manufacturer_decision,
        "select",
    )
    if selected_manufacturer_decision == "penalty":
        _verbose_debug(
            "[MATCH_MANUFACTURER] source=%s product_sku=%s candidate_id=%s decision=suspicious_penalty score=%s name_score=%s product_manufacturer=%s candidate_manufacturer=%s",
            source_type,
            product.sku,
            supplier.item.id,
            round(combined, 2),
            round(name_score, 2),
            product.manufacturer_norm,
            supplier.manufacturer_norm,
        )
    if structure_decision == "suspicious":
        safety_stats["selected_suspicious"] += 1
    return supplier, round(combined, 2), "ok", {
        "filteredCounts": filtered_counts,
        "rejectCounts": reject_counts,
        "manufacturerCounts": manufacturer_counts,
        "safetyStats": safety_stats,
        "threshold": threshold,
        "selected": {
            "nameScore": round(name_score, 2),
            "baseScore": round(_base_name_similarity(product_base, supplier.base_name_norm or supplier.name_norm), 2),
            "manufacturerDecision": selected_manufacturer_decision,
            "manufacturerScore": round(selected_mfr_score, 2),
            "structureDecision": structure_decision,
            "structureReason": None,
        },
        "topCandidates": top_final_candidates,
        "rejectedTopCandidates": rejected_top_candidates,
        "manufacturerAliasCandidates": manufacturer_alias_candidates,
        "poolStats": {
            "initialPool": initial_pool,
            "afterNameFilter": after_name_filter,
            "afterStrictFilter": after_strict_filter,
            "afterManufacturerPriority": after_manufacturer_priority,
        },
    }


def rematch_price_list_items_by_product(
    *,
    db: Session,
    price_list: CompetitorPriceList,
    product_indexes: ProductIndexes | None = None,
    flush: bool = True,
) -> dict:
    started_at = time.perf_counter()
    if price_list.source_type == "provisor" and product_indexes is None:
        _sync_provisor_reference_mapping_from_items(db, account_id=price_list.account_id or "existing")
    indexes = product_indexes or _product_indexes(db)
    products = indexes.candidates
    indexed_at = time.perf_counter()
    items = (
        db.execute(select(CompetitorPriceListItem).where(CompetitorPriceListItem.price_list_id == price_list.id))
        .scalars()
        .all()
    )
    loaded_at = time.perf_counter()
    supplier_indexes = _build_supplier_indexes(price_list=price_list, items=items)
    supplier_candidates = supplier_indexes.candidates
    supplier_token_index = supplier_indexes.token_index
    supplier_sku_index = supplier_indexes.sku_index
    supplier_composite_index = supplier_indexes.composite_index
    supplier_indexed_at = time.perf_counter()
    is_provisor_reference_price_list = price_list.source_type == "provisor" and any(
        _to_int(item.filial_id) in PROVISOR_REFERENCE_FILIAL_IDS for item in items
    )
    product_rows_by_id: dict[int, Product] = {}
    if price_list.source_type == "provisor" and is_provisor_reference_price_list and products:
        product_rows_by_id = {
            int(row.id): row
            for row in db.execute(select(Product).where(Product.id.in_([product.id for product in products]))).scalars().all()
        }

    stats = {
        "incoming": len(items),
        "products_total": len(products),
        "products_with_price": 0,
        "coverage_percent": 0,
        "supplier_items_total": len(items),
        "supplier_items_matched": 0,
        "productCount": len(products),
        "matched": 0,
        "matchedProducts": 0,
        "fuzzy": 0,
        "unmatched": 0,
        "filteredByQuantity": 0,
        "filteredByVolume": 0,
        "filteredByDosage": 0,
        "filteredByConcentration": 0,
        "filteredByForm": 0,
        "filterExamples": [],
        "matchedBySavedSku": 0,
        "matchedBySku": 0,
        "matchedByExactComposite": 0,
        "matchedByFuzzy": 0,
        "matchedByExactName": 0,
        "matchedWithManufacturerAlias": 0,
        "matchedWithManufacturerPenalty": 0,
        "hardManufacturerRejects": 0,
        **_empty_provisor_exact_counts(),
        "referenceMappedProducts": 0,
        "referenceUpdatedProducts": 0,
        "referenceUnmappedProducts": 0,
        "missingProductGoodsId": 0,
        "rejectReasons": {},
        "detailedRejectReasons": _empty_match_reject_counts(),
        "candidatePoolTotal": 0,
        "candidatePoolCount": 0,
        "candidatePoolMax": 0,
        "candidatePoolBroadFallbacks": 0,
        "structureRejects": 0,
        "structureSuspicious": 0,
        "selectedSuspicious": 0,
        "manufacturerHardRejectsSafety": 0,
        "fuzzyRejectedLowConfidence": 0,
        "stage1ExactComposite": 0,
        "stage2Fuzzy": 0,
        "fuzzyProductsCount": 0,
        "manufacturerRejects": 0,
        "_auditRows": [],
        "_unmatchedAuditRows": [],
        "_manufacturerAliasRows": [],
    }
    source_matches = (
        _source_goods_matches(db=db, price_format_id=price_list.price_format_id, source_type=price_list.source_type)
        if price_list.source_type == "provisor"
        else {}
    )
    source_matches_by_goods_id = (
        _source_goods_matches_by_goods_id(db=db, price_format_id=price_list.price_format_id, source_type=price_list.source_type)
        if price_list.source_type == "provisor"
        else {}
    )
    substitute_matches_by_product = _approved_provisor_substitutes_by_product(db) if price_list.source_type == "provisor" else {}
    present_provisor_goods_ids = {
        goods
        for item in items
        for goods in [_to_int(item.provisor_goods_id or (_raw_payload(item).get("goodsId")))]
        if goods
    }
    source_matches_by_product: dict[int, list[SourceGoodsMatch]] = defaultdict(list)
    for saved in source_matches.values():
        source_matches_by_product[int(saved.product_id)].append(saved)
    manual_mapping_result = apply_manual_mappings_to_items(db=db, price_list=price_list, items=items)
    used_item_ids: set[int] = set(manual_mapping_result.get("itemIds") or set())
    manual_mapped_product_ids: set[int] = set(manual_mapping_result.get("productIds") or set())
    stats["matchedByManualMapping"] = int(manual_mapping_result.get("applied") or 0)
    stats["manualRejected"] = int(manual_mapping_result.get("rejected") or 0)

    for product in products:
        if int(product.id) in manual_mapped_product_ids:
            continue
        matched_supplier: SupplierItemCandidate | None = None
        score = 0.0
        match_type = ""
        matched_key = normalize_sku(product.sku) or product.sku
        matched_kind = ""
        selected_audit_meta: dict = {}

        exact_composite_supplier = None
        exact_composite_meta: dict = {}
        if price_list.source_type != "provisor":
            exact_composite_supplier, exact_composite_meta = _find_exact_composite_supplier(
                product=product,
                composite_index=supplier_composite_index,
                used_item_ids=used_item_ids,
            )
            stats["manufacturerRejects"] += int(exact_composite_meta.get("manufacturerRejects") or 0)
            stats["_manufacturerAliasRows"].extend(exact_composite_meta.get("aliasCandidates") or [])
        if exact_composite_supplier is not None:
            matched_supplier = exact_composite_supplier
            score = 100.0
            match_type = "exact_composite"
            matched_key = f"composite:{product.base_name_norm or product.name_norm}"
            matched_kind = "exact_composite"
            stats["matchedByExactComposite"] += 1
            stats["stage1ExactComposite"] += 1
            selected_structure_decision, selected_structure_reason = _strict_structure_decision(product.structure, matched_supplier.structure)
            selected_manufacturer_decision, selected_mfr_score, _selected_hard = _manufacturer_decision(product.manufacturer_norm, matched_supplier.manufacturer_norm)
            selected_audit_meta = {
                "selected": {
                    "nameScore": 100,
                    "baseScore": round(_base_name_similarity(product.base_name_norm or product.name_norm, matched_supplier.base_name_norm or matched_supplier.name_norm), 2),
                    "manufacturerDecision": selected_manufacturer_decision,
                    "manufacturerScore": round(selected_mfr_score, 2),
                    "structureDecision": selected_structure_decision,
                    "structureReason": selected_structure_reason,
                },
                "topCandidates": exact_composite_meta.get("topCandidates") or [],
            }

        if matched_supplier is None and price_list.source_type == "provisor":
            exact_keys: list[tuple[str, str]] = []
            product_provisor_goods_id = product.provisor_goods_id
            substitute_rows_for_product = substitute_matches_by_product.get(int(product.id), [])
            substitute_keys_considered = False
            if is_provisor_reference_price_list:
                exact_keys.extend(_provisor_reference_sku_keys(product.sku))
            if product_provisor_goods_id:
                manual_or_saved_goods_match = source_matches_by_goods_id.get(_to_int(product_provisor_goods_id) or 0)
                if manual_or_saved_goods_match is None or int(manual_or_saved_goods_match.product_id) == int(product.id):
                    exact_keys.extend(_provisor_product_goods_keys(product_provisor_goods_id))
            elif not is_provisor_reference_price_list:
                stats["missingProductGoodsId"] += 1
            primary_goods_present = bool(_to_int(product_provisor_goods_id) and _to_int(product_provisor_goods_id) in present_provisor_goods_ids)
            if primary_goods_present and substitute_rows_for_product:
                logger.info(
                    "[PROVISOR_SUBSTITUTE_NOT_USED] reason=primary_found product_id=%s product_sku=%s primary_goods_id=%s price_list_id=%s",
                    product.id,
                    product.sku,
                    product_provisor_goods_id or "",
                    price_list.id,
                )
            if not primary_goods_present:
                for substitute in substitute_rows_for_product:
                    substitute_goods_id = _to_int(substitute.source_goods_id)
                    if substitute_goods_id:
                        exact_keys.append(("substitute_goods_id", _provisor_key("goods_id", substitute_goods_id)))
                        substitute_keys_considered = True
        elif matched_supplier is None:
            exact_keys = [("sku", key) for key in _product_sku_keys(product.sku)]
            substitute_keys_considered = False
            substitute_rows_for_product = []
        else:
            exact_keys = []
            substitute_keys_considered = False
            substitute_rows_for_product = []

        for key_kind, key in exact_keys:
            exact_candidates = []
            for sku_cand in supplier_sku_index.get(key, []):
                cand = sku_cand.candidate
                if int(cand.item.id) in used_item_ids:
                    continue
                if price_list.source_type == "provisor" and key_kind in {"goods_id", "reference_sku", "substitute_goods_id"}:
                    exact_candidates.append(sku_cand)
                    continue
                structure_decision, structure_reason = _strict_structure_decision(product.structure, cand.structure)
                if structure_decision == "reject":
                    product_value, candidate_value = _structure_values_for_reason(product.structure, cand.structure, structure_reason)
                    stats["structureRejects"] += 1
                    field = _structure_reason_field(structure_reason) or ""
                    detailed_reason = _reject_reason_for_structure_field(field)
                    if detailed_reason:
                        _bump_match_reject(stats, detailed_reason)
                    _verbose_debug(
                        "[MATCH_STRUCTURE_REJECT] product_sku=%s product_name=%s candidate_id=%s candidate_name=%s reason=%s product_value=%s candidate_value=%s",
                        product.sku,
                        product.name_norm,
                        cand.item.id,
                        cand.item.raw_name or cand.item.name,
                        structure_reason,
                        product_value,
                        candidate_value,
                    )
                    continue
                if structure_decision == "suspicious":
                    stats["structureSuspicious"] += 1
                manufacturer_decision, manufacturer_score, manufacturer_ok = _manufacturer_required_match(product.manufacturer_norm, cand.manufacturer_norm)
                _verbose_debug(
                    "[MATCH] product=%s candidate=%s manufacturer_match=%s manufacturer_decision=%s",
                    product.sku,
                    cand.item.id,
                    manufacturer_ok,
                    manufacturer_decision,
                )
                if not manufacturer_ok:
                    name_score = _base_name_similarity(product.base_name_norm or product.name_norm, cand.base_name_norm or cand.name_norm)
                    if structure_decision == "ok" and name_score >= 97 and manufacturer_decision in {"manufacturer_reject", "hard_reject"}:
                        stats["_manufacturerAliasRows"].append(_manufacturer_alias_candidate_row(product=product, supplier=cand, score=name_score))
                    _verbose_debug(
                        "[MATCH] product=%s candidate=%s exact_key=%s manufacturer_mismatch_rejected=true manufacturer_score=%s",
                        product.sku,
                        cand.item.id,
                        key_kind,
                        round(manufacturer_score, 2),
                    )
                    stats["manufacturerHardRejectsSafety"] += 1
                    stats["manufacturerRejects"] += 1
                    continue
                if manufacturer_decision == "missing" and structure_decision != "ok":
                    stats["fuzzyRejectedLowConfidence"] += 1
                    continue
                exact_candidates.append(sku_cand)
            if exact_candidates:
                selected_sku_candidate = min(
                    exact_candidates,
                    key=lambda sku_cand: float(_price_decimal(sku_cand.candidate.item.distributor_price) or Decimal("999999999")),
                )
                matched_supplier = selected_sku_candidate.candidate
                score = 100.0
                if price_list.source_type == "provisor" and key_kind == "goods_id":
                    match_type = "provisor_goods_id"
                elif price_list.source_type == "provisor" and key_kind == "substitute_goods_id":
                    match_type = "provisor_manual_substitute"
                elif price_list.source_type == "provisor" and key_kind == "reference_sku":
                    match_type = "reference_filial_128_distributor_goods_id"
                else:
                    match_type = "sku"
                matched_key = key
                matched_kind = "substitute_goods_id" if key_kind == "substitute_goods_id" else selected_sku_candidate.kind
                stats["matchedBySku"] += 1
                if price_list.source_type == "provisor":
                    _bump_provisor_exact(stats, matched_kind)
                if price_list.source_type == "provisor" and key_kind in {"goods_id", "reference_sku", "substitute_goods_id"}:
                    selected_structure_decision = "skipped_exact_provisor_id"
                    _selected_structure_reason = ""
                    selected_manufacturer_decision = "skipped_exact_provisor_id"
                    selected_mfr_score = 100.0
                    if key_kind == "substitute_goods_id":
                        substitute_meta = next(
                            (
                                row
                                for row in substitute_matches_by_product.get(int(product.id), [])
                                if _provisor_key("goods_id", row.source_goods_id) == key
                            ),
                            None,
                        )
                        logger.info(
                            "[PROVISOR_SUBSTITUTE_MATCH] product_id=%s product_sku=%s source_goods_id=%s source_name=%s source_manufacturer=%s",
                            product.id,
                            product.sku,
                            matched_supplier.item.provisor_goods_id or "",
                            (substitute_meta.source_name if substitute_meta else "") or matched_supplier.item.raw_name or matched_supplier.item.name or "",
                            (substitute_meta.source_manufacturer if substitute_meta else "") or matched_supplier.item.raw_manufacturer or "",
                        )
                        logger.info(
                            "[PROVISOR_SUBSTITUTE_APPLIED] product_id=%s product_sku=%s primary_goods_id=%s substitute_goods_id=%s price=%s price_list_id=%s",
                            product.id,
                            product.sku,
                            product_provisor_goods_id or "",
                            matched_supplier.item.provisor_goods_id or "",
                            matched_supplier.item.distributor_price or "",
                            price_list.id,
                        )
                    elif key_kind == "goods_id":
                        logger.info(
                            "[PROVISOR_GOODS_ID_MATCH] account_id=%s filial_id=%s goods_id=%s product_id=%s product_code=%s match_type=provisor_goods_id",
                            price_list.account_id or "",
                            matched_supplier.item.filial_id or "",
                            product_provisor_goods_id or matched_supplier.item.provisor_goods_id or "",
                            product.id,
                            product.sku,
                        )
                    else:
                        logger.info(
                            "[PROVISOR_REFERENCE_SKU_MATCH] account_id=%s filial_id=%s distributor_goods_id=%s goods_id=%s product_id=%s match_type=reference_filial_128_distributor_goods_id",
                            price_list.account_id or "",
                            matched_supplier.item.filial_id or "",
                            matched_supplier.item.distributor_goods_id or "",
                            matched_supplier.item.provisor_goods_id or "",
                            product.id,
                        )
                else:
                    selected_structure_decision, _selected_structure_reason = _strict_structure_decision(product.structure, matched_supplier.structure)
                    if selected_structure_decision == "suspicious":
                        stats["selectedSuspicious"] += 1
                    selected_manufacturer_decision, selected_mfr_score, _selected_hard = _manufacturer_decision(product.manufacturer_norm, matched_supplier.manufacturer_norm)
                selected_audit_meta = {
                    "selected": {
                        "nameScore": 100,
                        "baseScore": round(_base_name_similarity(product.base_name_norm or product.name_norm, matched_supplier.base_name_norm or matched_supplier.name_norm), 2),
                        "manufacturerDecision": selected_manufacturer_decision,
                        "manufacturerScore": round(selected_mfr_score, 2),
                        "structureDecision": selected_structure_decision,
                        "structureReason": _selected_structure_reason,
                    },
                    "topCandidates": [],
                }
                _verbose_debug(
                    "[MATCH_SELECTED] product_sku=%s candidate_id=%s score=%s name_score=%s manufacturer_score=%s manufacturer_decision=%s structure_decision=%s match_type=%s",
                    product.sku,
                    matched_supplier.item.id,
                    round(score, 2),
                    100,
                    round(selected_mfr_score, 2),
                    selected_manufacturer_decision,
                    selected_structure_decision,
                    match_type,
                )
                break

        if (
            matched_supplier is None
            and price_list.source_type == "provisor"
            and substitute_keys_considered
            and substitute_rows_for_product
        ):
            logger.info(
                "[PROVISOR_SUBSTITUTE_NOT_USED] reason=no_substitute_price product_id=%s product_sku=%s primary_goods_id=%s price_list_id=%s",
                product.id,
                product.sku,
                product_provisor_goods_id or "",
                price_list.id,
            )

        if matched_supplier is None and price_list.source_type == "provisor":
            for saved in source_matches_by_product.get(int(product.id), []):
                saved_goods_owner = source_matches_by_goods_id.get(_to_int(saved.goods_id) or 0) if saved.goods_id else None
                if saved_goods_owner is not None and int(saved_goods_owner.product_id) != int(product.id):
                    continue
                raw_key = normalize_external_sku(saved.distributor_goods_id)
                saved_keys = []
                if saved.goods_id:
                    saved_keys.append(_provisor_key("goods_id", saved.goods_id))
                if raw_key and str(saved.match_method or "") == "reference_filial_128_distributor_goods_id":
                    saved_keys.append(_provisor_key("reference_sku", raw_key))
                exact_candidates = []
                for saved_key in saved_keys:
                    for sku_cand in supplier_sku_index.get(saved_key, []):
                        if int(sku_cand.candidate.item.id) in used_item_ids:
                            continue
                        exact_candidates.append(sku_cand)
                if exact_candidates:
                    matched_supplier = exact_candidates[0].candidate
                    score = float(saved.similarity_score) if saved.similarity_score is not None else 100.0
                    if saved.goods_id:
                        match_type = "provisor_goods_id"
                        matched_key = _provisor_key("goods_id", saved.goods_id)
                        matched_kind = "goods_id"
                        _bump_provisor_exact(stats, "goods_id")
                    else:
                        match_type = "reference_filial_128_distributor_goods_id"
                        matched_key = _provisor_key("reference_sku", raw_key)
                        matched_kind = "reference_sku"
                        _bump_provisor_exact(stats, "reference_sku")
                    stats["matchedBySavedSku"] += 1
                    selected_structure_decision = "skipped_exact_provisor_id"
                    _selected_structure_reason = ""
                    selected_manufacturer_decision = "skipped_exact_provisor_id"
                    selected_mfr_score = 100.0
                    selected_audit_meta = {
                        "selected": {
                            "nameScore": 100,
                            "baseScore": round(_base_name_similarity(product.base_name_norm or product.name_norm, matched_supplier.base_name_norm or matched_supplier.name_norm), 2),
                            "manufacturerDecision": selected_manufacturer_decision,
                            "manufacturerScore": round(selected_mfr_score, 2),
                            "structureDecision": selected_structure_decision,
                            "structureReason": _selected_structure_reason,
                        },
                        "topCandidates": [],
                    }
                    _verbose_debug(
                        "[MATCH_SELECTED] product_sku=%s candidate_id=%s score=%s name_score=%s manufacturer_score=%s manufacturer_decision=%s structure_decision=%s match_type=%s",
                        product.sku,
                        matched_supplier.item.id,
                        round(score, 2),
                        100,
                        round(selected_mfr_score, 2),
                        selected_manufacturer_decision,
                        selected_structure_decision,
                        match_type,
                    )
                    if match_type == "provisor_goods_id":
                        logger.info(
                            "[PROVISOR_GOODS_ID_MATCH] account_id=%s filial_id=%s goods_id=%s product_id=%s product_code=%s match_type=provisor_goods_id",
                            price_list.account_id or "",
                            matched_supplier.item.filial_id or "",
                            saved.goods_id or matched_supplier.item.provisor_goods_id or "",
                            product.id,
                            product.sku,
                        )
                    else:
                        logger.info(
                            "[PROVISOR_REFERENCE_SKU_MATCH] account_id=%s filial_id=%s distributor_goods_id=%s goods_id=%s product_id=%s match_type=reference_filial_128_distributor_goods_id",
                            price_list.account_id or "",
                            matched_supplier.item.filial_id or "",
                            matched_supplier.item.distributor_goods_id or "",
                            matched_supplier.item.provisor_goods_id or "",
                            product.id,
                        )
                    break

        if matched_supplier is None:
            if price_list.source_type == "provisor":
                product_goods = product_provisor_goods_id if "product_provisor_goods_id" in locals() else product.provisor_goods_id
                logger.info(
                    "[PROVISOR_FALLBACK_MATCH] account_id=%s filial_id=%s goods_id=%s reason=no_goods_id_mapping",
                    price_list.account_id or "",
                    price_list.branch_id or price_list.source_key or "",
                    product_goods or "",
                )
            stats["fuzzyProductsCount"] += 1
            pool, pool_before_limit = _supplier_pool_for_product(product=product, token_index=supplier_token_index, max_pool=20)
            if not pool:
                _bump_match_reject(stats, "empty_candidate_pool")
            stats["candidatePoolTotal"] += len(pool)
            stats["candidatePoolCount"] += 1
            stats["candidatePoolMax"] = max(int(stats["candidatePoolMax"]), len(pool))
            if pool_before_limit > 20:
                stats["candidatePoolLimited"] = int(stats.get("candidatePoolLimited") or 0) + 1
            candidate, candidate_score, reason, meta = _best_supplier_for_product(
                product=product,
                pool=pool,
                used_item_ids=used_item_ids,
                source_type=price_list.source_type,
            )
            filtered_counts = meta.get("filteredCounts") or {}
            _merge_match_reject_counts(stats, meta.get("rejectCounts") or {})
            _merge_match_safety_stats(stats, meta.get("safetyStats") or {})
            manufacturer_counts = meta.get("manufacturerCounts") or {}
            stats["hardManufacturerRejects"] += int(manufacturer_counts.get("hard_reject") or 0)
            stats["manufacturerRejects"] += int(manufacturer_counts.get("manufacturer_reject") or 0)
            stats["_manufacturerAliasRows"].extend(meta.get("manufacturerAliasCandidates") or [])
            stats["filteredByQuantity"] += int(filtered_counts.get("quantity") or 0)
            stats["filteredByVolume"] += int(filtered_counts.get("volume") or 0)
            stats["filteredByDosage"] += int(filtered_counts.get("dosage") or 0)
            stats["filteredByConcentration"] += int(filtered_counts.get("concentration") or 0)
            stats["filteredByForm"] += int(filtered_counts.get("form") or 0)
            if candidate is not None and reason == "ok":
                matched_supplier = candidate
                score = candidate_score
                match_type = "fuzzy_name_manufacturer"
                matched_kind = "fuzzy"
                stats["fuzzy"] += 1
                stats["matchedByFuzzy"] += 1
                stats["stage2Fuzzy"] += 1
                if candidate_score >= 95:
                    stats["matchedByExactName"] += 1
                selected_decision, _selected_similarity, _selected_hard = _manufacturer_decision(
                    product.manufacturer_norm,
                    candidate.manufacturer_norm,
                )
                if selected_decision in {"alias_match", "similar_match"}:
                    stats["matchedWithManufacturerAlias"] += 1
                elif selected_decision == "penalty":
                    stats["matchedWithManufacturerPenalty"] += 1
                selected_audit_meta = meta
            else:
                top_candidates = meta.get("rejectedTopCandidates") or meta.get("topCandidates") or exact_composite_meta.get("topCandidates") or []
                if pool or top_candidates:
                    best_top = top_candidates[0] if top_candidates else {}
                    pool_stats = meta.get("poolStats") or {}
                    stats["_unmatchedAuditRows"].append(
                        {
                            "product_sku": product.sku,
                            "product_name": product.raw_name or product.name_norm,
                            "product_manufacturer": product.manufacturer_norm,
                            "best_candidate_name": best_top.get("name", ""),
                            "best_candidate_manufacturer": best_top.get("manufacturer", ""),
                            "name_score": best_top.get("name_score", ""),
                            "manufacturer_decision": best_top.get("manufacturer_decision", ""),
                            "structure_decision": best_top.get("structure_decision", ""),
                            "reject_reason": reason,
                            "candidate_pool_size": pool_stats.get("initialPool", len(pool)),
                            "top_3_candidates": json.dumps(top_candidates[:3], ensure_ascii=False),
                        }
                    )
                _bump_reason(stats, reason)
                continue

        if matched_supplier is None:
            _bump_reason(stats, "no_candidate")
            continue

        item = matched_supplier.item
        item.product_id = product.id
        item.match_type = match_type
        item.match_score = score
        item.matched_sku = matched_key
        item.match_key = matched_key
        used_item_ids.add(int(item.id))
        stats["matchedProducts"] += 1
        selected_meta = selected_audit_meta.get("selected") or {}
        top_candidates = selected_audit_meta.get("topCandidates") or []
        name_score_audit = float(selected_meta.get("nameScore") if selected_meta.get("nameScore") is not None else score)
        base_score_audit = float(selected_meta.get("baseScore") if selected_meta.get("baseScore") is not None else _base_name_similarity(product.base_name_norm or product.name_norm, matched_supplier.base_name_norm or matched_supplier.name_norm))
        manufacturer_decision_audit = str(selected_meta.get("manufacturerDecision") or _manufacturer_required_match(product.manufacturer_norm, matched_supplier.manufacturer_norm)[0])
        manufacturer_score_audit = float(selected_meta.get("manufacturerScore") if selected_meta.get("manufacturerScore") is not None else _manufacturer_required_match(product.manufacturer_norm, matched_supplier.manufacturer_norm)[1])
        structure_decision_audit = str(selected_meta.get("structureDecision") or _strict_structure_decision(product.structure, matched_supplier.structure)[0])
        structure_reason_audit = selected_meta.get("structureReason")
        if structure_reason_audit is None:
            _structure_decision_tmp, structure_reason_audit = _strict_structure_decision(product.structure, matched_supplier.structure)
        risk_reasons = _audit_risk_reasons(
            match_type=match_type,
            name_score=name_score_audit,
            base_score=base_score_audit,
            final_score=float(score or 0),
            manufacturer_decision=manufacturer_decision_audit,
            structure_decision=structure_decision_audit,
            product_struct=product.structure,
            candidate_struct=matched_supplier.structure,
            top_candidates=top_candidates,
        )
        if risk_reasons:
            pool_stats = selected_audit_meta.get("poolStats") or {}
            stats["_auditRows"].append(
                {
                    "product_sku": product.sku,
                    "product_name": product.raw_name or product.name_norm,
                    "product_manufacturer": product.manufacturer_norm,
                    "product_structure": _structure_audit_text(product.structure),
                    "candidate_name": item.raw_name or item.name or item.distributor_goods_name,
                    "candidate_manufacturer": matched_supplier.manufacturer_norm,
                    "candidate_structure": _structure_audit_text(matched_supplier.structure),
                    "product_parsed_base_name": product.structure.base_name,
                    "product_parsed_forms": json.dumps(list(product.structure.forms or ()), ensure_ascii=False),
                    "product_parsed_dosage": product.structure.dosage,
                    "product_parsed_volume": product.structure.volume,
                    "candidate_parsed_base_name": matched_supplier.structure.base_name,
                    "candidate_parsed_forms": json.dumps(list(matched_supplier.structure.forms or ()), ensure_ascii=False),
                    "candidate_parsed_dosage": matched_supplier.structure.dosage,
                    "candidate_parsed_volume": matched_supplier.structure.volume,
                    "price": float(item.distributor_price) if item.distributor_price is not None else None,
                    "source_price_list": f"{price_list.source_type}:{price_list.source_key}",
                    "match_type": match_type,
                    "name_score": round(name_score_audit, 2),
                    "base_score": round(base_score_audit, 2),
                    "manufacturer_decision": manufacturer_decision_audit,
                    "manufacturer_score": round(manufacturer_score_audit, 2),
                    "structure_decision": structure_decision_audit,
                    "structure_reason": structure_reason_audit or "",
                    "final_score": round(float(score or 0), 2),
                    "risk_reason": "; ".join(risk_reasons),
                    "candidate_pool_size": pool_stats.get("initialPool", ""),
                    "top_3_candidates": json.dumps(top_candidates[:3], ensure_ascii=False),
                }
            )
        if (
            price_list.source_type == "provisor"
            and is_provisor_reference_price_list
            and matched_kind == "reference_sku"
            and item.provisor_goods_id
        ):
            product_row = product_rows_by_id.get(int(product.id))
            if product_row is not None and _to_int(product_row.provisor_goods_id) != _to_int(item.provisor_goods_id):
                product_row.provisor_goods_id = _to_int(item.provisor_goods_id)
                stats["referenceUpdatedProducts"] += 1
            debug_product_goods_id = (
                _to_int(getattr(product_row, "provisor_goods_id", None))
                if product_row is not None
                else _to_int(product.provisor_goods_id)
            )
            stats["referenceMappedProducts"] += 1
            _verbose_debug(
                "[MATCH_PROVISOR_DEBUG] product_sku=%s product_provisor_goods_id=%s supplier_goods_id=%s supplier_distributor_goods_id=%s match_type=reference_sku",
                product.sku,
                debug_product_goods_id or "",
                item.provisor_goods_id or "",
                item.distributor_goods_id or "",
            )
        elif price_list.source_type == "provisor" and matched_kind in {"goods_id", "saved", "fuzzy"}:
            debug_product_goods_id = product_provisor_goods_id if price_list.source_type == "provisor" else product.provisor_goods_id
            _verbose_debug(
                "[MATCH_PROVISOR_DEBUG] product_sku=%s product_provisor_goods_id=%s supplier_goods_id=%s supplier_distributor_goods_id=%s match_type=%s",
                product.sku,
                debug_product_goods_id or "",
                item.provisor_goods_id or "",
                item.distributor_goods_id or "",
                matched_kind,
            )
        if price_list.source_type == "provisor" and item.distributor_goods_id and match_type != "provisor_manual_substitute":
            raw = _raw_payload(item)
            raw_inner = raw.get("raw") if isinstance(raw.get("raw"), dict) else raw
            _upsert_source_goods_match(
                db=db,
                match_cache=source_matches,
                price_format_id=price_list.price_format_id,
                source_type="provisor",
                distributor_goods_id=normalize_external_sku(item.distributor_goods_id),
                goods_id=item.provisor_goods_id or raw_inner.get("goodsId"),
                distributor_goods_name=item.raw_name,
                distributor_producer=item.raw_manufacturer,
                product_id=int(product.id),
                similarity_score=score,
                match_method=match_type[:64],
            )

    if _keeps_multi_prices_per_sku(price_list):
        products_by_goods_id = {
            _to_int(product.provisor_goods_id): product
            for product in products
            if _to_int(product.provisor_goods_id)
        }
        for item in items:
            goods_id = _to_int(item.provisor_goods_id or (_raw_payload(item).get("goodsId")))
            product = products_by_goods_id.get(goods_id or 0)
            if product is None:
                continue
            item.product_id = int(product.id)
            item.match_type = "provisor_goods_id"
            item.match_score = 100
            item.matched_sku = normalize_sku(product.sku) or product.sku
            item.match_key = _provisor_key("goods_id", goods_id)

    matched_at = time.perf_counter()
    matched_items = [item for item in items if item.product_id is not None and item.matched_sku]
    matched_product_ids = {int(item.product_id) for item in matched_items if item.product_id is not None}
    stats["matched"] = len(matched_items)
    stats["products_with_price"] = len(matched_product_ids)
    stats["supplier_items_matched"] = len(matched_items)
    stats["unmatched"] = max(0, len(products) - int(stats["matchedProducts"]))
    stats["coverage_percent"] = round((len(matched_product_ids) / len(products) * 100), 2) if products else 0
    stats["matchRate"] = stats["coverage_percent"]
    stats["coverage"] = stats["coverage_percent"]
    stats["debug_supplier_match_rate"] = round((len(matched_items) / len(items) * 100), 2) if items else 0
    if price_list.source_type == "provisor" and is_provisor_reference_price_list:
        stats["referenceUnmappedProducts"] = max(0, len(products) - int(stats["referenceMappedProducts"]))
    stats["candidatePoolAvg"] = (
        round(int(stats["candidatePoolTotal"]) / int(stats["candidatePoolCount"]), 2)
        if int(stats["candidatePoolCount"])
        else 0
    )
    stats["_matchedRows"] = [
        _matched_price_row(item)
        for item in matched_items
        if item.distributor_price is not None
    ]
    stats["topUnmatched"] = []
    debugged_at = time.perf_counter()
    stats["timingMs"] = {
        "loadProductsAndIndexes": round((indexed_at - started_at) * 1000, 2),
        "loadItems": round((loaded_at - indexed_at) * 1000, 2),
        "supplierIndexes": round((supplier_indexed_at - loaded_at) * 1000, 2),
        "matching": round((matched_at - supplier_indexed_at) * 1000, 2),
        "diagnostics": round((debugged_at - matched_at) * 1000, 2),
    }
    if price_list.source_type == "vidman":
        vidman_timing_payload = [
            ("parse_structure", stats["timingMs"]["supplierIndexes"]),
            ("build_supplier_indexes", stats["timingMs"]["supplierIndexes"]),
            ("candidate_pool", stats["timingMs"]["matching"]),
            ("exact_match", stats["timingMs"]["matching"]),
            ("fuzzy_match", stats["timingMs"]["matching"]),
            ("save_results", stats["timingMs"]["diagnostics"]),
        ]
        for stage, elapsed_ms in vidman_timing_payload:
            logger.info(
                "[VIDMAN_MATCH_TIMING] stage=%s elapsed_ms=%s items=%s pool_size=%s",
                stage,
                elapsed_ms,
                len(items),
                stats.get("candidatePoolAvg", 0),
            )
    logger.info(
        "[MATCH] source=%s price_list_id=%s product_count=%s supplier_items=%s coverage=%s%% matched=%s pool_avg=%s pool_max=%s broad_fallbacks=0",
        price_list.source_type,
        price_list.id,
        len(products),
        len(items),
        stats["coverage"],
        stats["matchedProducts"],
        stats["candidatePoolAvg"],
        stats["candidatePoolMax"],
    )
    logger.info(
        "[MATCH_REJECTS] source=%s price_list_id=%s manufacturer_mismatch=%s manufacturer_hard_reject=%s manufacturer_penalty=%s dosage_mismatch=%s quantity_mismatch=%s volume_mismatch=%s form_mismatch=%s empty_candidate_pool=%s",
        price_list.source_type,
        price_list.id,
        stats["detailedRejectReasons"].get("manufacturer_mismatch", 0),
        stats["detailedRejectReasons"].get("manufacturer_hard_reject", 0),
        stats["detailedRejectReasons"].get("manufacturer_penalty", 0),
        stats["detailedRejectReasons"].get("dosage_mismatch", 0),
        stats["detailedRejectReasons"].get("quantity_mismatch", 0),
        stats["detailedRejectReasons"].get("volume_mismatch", 0),
        stats["detailedRejectReasons"].get("form_mismatch", 0),
        stats["detailedRejectReasons"].get("empty_candidate_pool", 0),
    )
    logger.info(
        "[MATCH_SAFETY] structure_rejects=%s structure_suspicious=%s selected_suspicious=%s manufacturer_hard_rejects=%s fuzzy_rejected_low_confidence=%s",
        stats.get("structureRejects", 0),
        stats.get("structureSuspicious", 0),
        stats.get("selectedSuspicious", 0),
        stats.get("manufacturerHardRejectsSafety", 0),
        stats.get("fuzzyRejectedLowConfidence", 0),
    )
    logger.info(
        "[MATCH_STAGE_STATS] stage1_exact_composite=%s stage2_fuzzy=%s fuzzy_products_count=%s avg_candidate_pool=%s max_candidate_pool=%s manufacturer_rejects=%s elapsed_ms=%s",
        stats.get("stage1ExactComposite", 0),
        stats.get("stage2Fuzzy", 0),
        stats.get("fuzzyProductsCount", 0),
        stats.get("candidatePoolAvg", 0),
        stats.get("candidatePoolMax", 0),
        stats.get("manufacturerRejects", 0),
        round((time.perf_counter() - started_at) * 1000, 2),
    )
    if price_list.source_type == "vidman":
        _verbose_debug(
            "[MATCH_VIDMAN] products_total=%s supplier_items=%s matched=%s coverage=%s%% matchedByExactName=%s matchedByFuzzy=%s matchedWithManufacturerAlias=%s matchedWithManufacturerPenalty=%s hardManufacturerRejects=%s",
            len(products),
            len(items),
            stats["matchedProducts"],
            stats["coverage"],
            stats.get("matchedByExactName", 0),
            stats.get("matchedByFuzzy", 0),
            stats.get("matchedWithManufacturerAlias", 0),
            stats.get("matchedWithManufacturerPenalty", 0),
            stats.get("hardManufacturerRejects", 0),
        )
    if price_list.source_type == "provisor":
        if is_provisor_reference_price_list:
            _verbose_debug(
                "[PROVISOR_REFERENCE] account_id=%s reference_filial_ids=%s available=true mapped_products=%s updated_products=%s unmapped_products=%s",
                price_list.account_id or "",
                ",".join(str(x) for x in PROVISOR_REFERENCE_FILIAL_IDS),
                stats.get("referenceMappedProducts", 0),
                stats.get("referenceUpdatedProducts", 0),
                stats.get("referenceUnmappedProducts", 0),
            )
        else:
            _verbose_debug(
                "[PROVISOR_REFERENCE] account_id=%s reference_filial_ids=%s available=false mapped_products=0 updated_products=0",
                price_list.account_id or "",
                ",".join(str(x) for x in PROVISOR_REFERENCE_FILIAL_IDS),
            )
            _verbose_debug(
                "[PROVISOR_REFERENCE_SKIP] account_id=%s reason=reference_filial_not_available using_existing_product_provisor_goods_id=true missingProductGoodsId=%s",
                price_list.account_id or "",
                stats.get("missingProductGoodsId", 0),
            )
        _verbose_debug(
            "[MATCH_PROVISOR] total_products=%s matched_total=%s matchedByReferenceDistributorSku=%s matchedByGoodsId=%s matchedByComposite=%s matchedBySavedSku=%s matchedByFuzzy=%s missingProductGoodsId=%s coverage=%s%%",
            len(products),
            stats["matchedProducts"],
            stats.get("matchedByReferenceDistributorSku", 0),
            stats.get("matchedByGoodsId", 0),
            stats.get("matchedByGoodsDistributorComposite", 0),
            stats.get("matchedBySavedSku", 0),
            stats.get("matchedByFuzzy", 0),
            stats.get("missingProductGoodsId", 0),
            stats["coverage"],
        )
    if flush:
        db.flush()
    return stats


def explain_unmatched_for_price_list(*, db: Session, price_list: CompetitorPriceList, limit: int = 10) -> list[dict]:
    indexes = _product_indexes(db)
    items = (
        db.execute(
            select(CompetitorPriceListItem)
            .where(CompetitorPriceListItem.price_list_id == price_list.id)
            .where((CompetitorPriceListItem.matched_sku == "") | (CompetitorPriceListItem.product_id.is_(None)))
            .order_by(CompetitorPriceListItem.id.asc())
            .limit(limit)
        )
        .scalars()
        .all()
    )
    return _unmatched_debug(
        price_list=price_list,
        items=items,
        code_to_id=indexes.code_to_id,
        candidates=indexes.candidates,
        exact=indexes.exact,
        token_index=indexes.token_index,
        limit=limit,
    )


def explain_vidman_unmatched(*, db: Session, price_format_id: int, limit: int = 10) -> list[dict]:
    indexes = _product_indexes(db)
    selected_ids = [
        int(item.price_list.id)
        for item in get_assigned_competitor_price_lists(db=db, price_format_id=price_format_id)
        if item.price_list.source_type == "vidman"
    ]
    rows = (
        db.execute(
            select(CompetitorPriceList, CompetitorPriceListItem)
            .join(CompetitorPriceListItem, CompetitorPriceListItem.price_list_id == CompetitorPriceList.id)
            .where(CompetitorPriceList.id.in_(selected_ids))
            .where(CompetitorPriceList.source_type == "vidman")
            .where((CompetitorPriceListItem.matched_sku == "") | (CompetitorPriceListItem.product_id.is_(None)))
            .limit(limit)
        )
        .all()
        if selected_ids
        else []
    )
    out: list[dict] = []
    for price_list, item in rows:
        raw_name = item.raw_name or item.name or item.distributor_goods_name
        raw_manufacturer = item.raw_manufacturer or ""
        mfr_norm = normalize_manufacturer_text(raw_manufacturer)
        candidate = _vidman_candidate_result(
            name_norm=raw_name,
            mfr_norm=mfr_norm,
            expiry_norm=normalize_expiry(item.expiry_date or ""),
            candidates=indexes.candidates,
            exact=indexes.exact,
            token_index=indexes.token_index,
        )
        out.append(
            {
                "priceListId": price_list.id,
                "source": f"{price_list.source_type}:{price_list.source_key}",
                "priceListName": price_list.display_name or price_list.supplier,
                "rawName": raw_name,
                "rawManufacturer": raw_manufacturer,
                "normalizedName": normalize_text(normalize_drug_text(raw_name)),
                "normalizedManufacturer": mfr_norm,
                "bestCandidate": candidate,
            }
        )
    return out


def rebuild_competitor_prices_for_selected(
    *,
    db: Session,
    price_format_id: int,
    commit_between_lists: bool = False,
) -> dict[str, dict]:
    total_started_at = time.perf_counter()
    operation = f"rebuild_competitor_prices:{price_format_id}"
    _timing(operation, "start_selection/rebuild", total_started_at)
    selected = [item.price_list for item in get_assigned_competitor_price_lists(db=db, price_format_id=price_format_id)]
    _timing(operation, "load_selected_price_lists", total_started_at)

    summary: dict[str, dict] = {}
    audit_rows: list[dict] = []
    unmatched_audit_rows: list[dict] = []
    manufacturer_alias_rows: list[dict] = []
    matched_by_source: list[tuple[str, str, object, list[dict]]] = []
    _sync_provisor_reference_mapping_from_items(db)
    index_started_at = time.perf_counter()
    product_indexes = _product_indexes(db)
    index_finished_at = time.perf_counter()
    _timing(operation, "product_indexes_build", index_started_at)
    logger.info("[MATCH] product_count=%s selected_price_lists=%s", len(product_indexes.candidates), len(selected))
    rematch_ms: dict[str, float] = {}
    for price_list in selected:
        rematch_started_at = time.perf_counter()
        stats = rematch_price_list_items_by_product(db=db, price_list=price_list, product_indexes=product_indexes, flush=False)
        matched_rows = stats.pop("_matchedRows", [])
        audit_rows.extend(stats.pop("_auditRows", []))
        unmatched_audit_rows.extend(stats.pop("_unmatchedAuditRows", []))
        manufacturer_alias_rows.extend(stats.pop("_manufacturerAliasRows", []))
        rematch_finished_at = time.perf_counter()
        src = f"{price_list.source_type}:{price_list.source_key}"
        _timing(f"{operation}:{src}", "rematch_price_list", rematch_started_at)
        summary[src] = stats
        rematch_ms[src] = round((rematch_finished_at - rematch_started_at) * 1000, 2)
        if commit_between_lists:
            commit_started_at = time.perf_counter()
            db.commit()
            _timing(f"{operation}:{src}", "flush/commit", commit_started_at)
        matched_by_source.append(
            (
                src,
                price_list.supplier or price_list.display_name,
                price_list.price_date,
                matched_rows,
            )
        )
    _timing(operation, "collect_matched_items", total_started_at)

    delete_started_at = time.perf_counter()
    db.execute(
        delete(CompetitorPrice)
        .where(CompetitorPrice.price_format_id == price_format_id)
        .where(CompetitorPrice.product_id.is_not(None))
    )
    delete_finished_at = time.perf_counter()
    _timing(operation, "delete_old_competitor_prices", delete_started_at)

    insert_started_at = time.perf_counter()
    new_rows: list[CompetitorPrice] = []
    for src, supplier, price_date, matched_items in matched_by_source:
        for matched in matched_items:
            product_id = int(matched.get("product_id") or 0)
            distributor_price = float(matched.get("price") or 0)
            source_item_id = matched.get("source_item_id")
            reason = str(matched.get("match_type") or "matched")
            _verbose_debug(
                "[PRICE] product=%s supplier=%s selected_price=%s source_price_list=%s source_item=%s reason=%s",
                product_id,
                supplier,
                distributor_price,
                src,
                source_item_id,
                reason,
            )
            new_rows.append(
                CompetitorPrice(
                    price_format_id=price_format_id,
                    product_id=product_id,
                    source_name=src,
                    supplier=supplier,
                    price_date=price_date,
                    coefficient=1.0,
                    source_price=distributor_price,
                    match_type=reason,
                    source_item_id=int(source_item_id) if source_item_id is not None else None,
                    source_goods_id=matched.get("source_goods_id"),
                    source_distributor_goods_id=str(matched.get("source_distributor_goods_id") or ""),
                    source_manufacturer=str(matched.get("source_manufacturer") or ""),
                )
            )
    if new_rows:
        db.bulk_save_objects(new_rows)
        db.flush()
    insert_finished_at = time.perf_counter()
    selected_total = sum(int(stats.get("matchedProducts") or stats.get("matched") or 0) for stats in summary.values())
    unmatched_with_top3 = 0
    unmatched_without_top3 = 0
    for row in unmatched_audit_rows:
        raw_top3 = row.get("top_3_candidates")
        has_top3 = False
        if isinstance(raw_top3, str) and raw_top3.strip():
            try:
                has_top3 = bool(json.loads(raw_top3))
            except Exception:
                has_top3 = raw_top3.strip() not in {"[]", "null", "None"}
        elif raw_top3:
            has_top3 = True
        if has_top3:
            unmatched_with_top3 += 1
        else:
            unmatched_without_top3 += 1
    audit_summary = {
        "total_products": len(product_indexes.candidates),
        "selected_total": selected_total,
        "selected_safe_count": max(0, selected_total - len(audit_rows)),
        "selected_risky_count": len(audit_rows),
        "unmatched_with_candidates_count": len(unmatched_audit_rows),
        "unmatched_candidates_with_top3": unmatched_with_top3,
        "unmatched_candidates_without_top3": unmatched_without_top3,
        "manufacturer_alias_candidates_count": len(manufacturer_alias_rows),
        "structure_suspicious_count": sum(int(stats.get("structureSuspicious") or 0) for stats in summary.values()),
        "structure_reject_count": sum(int(stats.get("structureRejects") or 0) for stats in summary.values()),
        "manufacturer_reject_count": sum(int(stats.get("manufacturerHardRejectsSafety") or stats.get("hardManufacturerRejects") or 0) for stats in summary.values()),
        "fuzzy_count": sum(int(stats.get("matchedByFuzzy") or stats.get("fuzzy") or 0) for stats in summary.values()),
    }
    audit_path = os.path.abspath(os.path.join(os.getcwd(), "suspicious_matches.xlsx"))
    _write_matching_audit_xlsx(
        path=audit_path,
        selected_rows=audit_rows,
        unmatched_rows=unmatched_audit_rows,
        manufacturer_rows=manufacturer_alias_rows,
        summary=audit_summary,
    )
    summary["_audit"] = {"path": audit_path, **audit_summary}
    _verbose_debug("[MATCH_AUDIT] path=%s rows=%s summary=%s", audit_path, len(audit_rows), audit_summary)
    _timing(operation, "insert_new_competitor_prices", insert_started_at)
    _timing(operation, "flush/commit", insert_started_at)
    _timing(operation, "finish", total_started_at)
    _timing(operation, "total_ms", total_started_at)
    _verbose_debug(
        "Rebuild competitor prices timing price_format_id=%s selected=%s timing_ms=%s",
        price_format_id,
        len(selected),
        {
            "productIndex": round((index_finished_at - index_started_at) * 1000, 2),
            "rematchBySource": rematch_ms,
            "collectMatchedItems": 0,
            "deleteOldCompetitorPrices": round((delete_finished_at - delete_started_at) * 1000, 2),
            "insertNewCompetitorPricesAndFlush": round((insert_finished_at - insert_started_at) * 1000, 2),
            "total": round((insert_finished_at - total_started_at) * 1000, 2),
        },
    )
    return summary
