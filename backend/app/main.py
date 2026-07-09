from __future__ import annotations
import sys
import asyncio

if sys.platform.startswith("win"):
    asyncio.set_event_loop_policy(asyncio.WindowsProactorEventLoopPolicy())
import logging
import json
import time
import difflib
import re
import unicodedata
from dataclasses import replace
from datetime import date, datetime, timedelta
from decimal import Decimal
from urllib.parse import quote
from fastapi import Body, Depends, FastAPI, File, Form, HTTPException, Query, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
import httpx
from sqlalchemy import delete, func, select, update
from sqlalchemy.exc import OperationalError, SQLAlchemyError
from sqlalchemy.orm import Session
import io
import csv
from pathlib import Path
import os
from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter

from .config import get_settings
from . import data
from .db import SessionLocal, init_db
from .deps import (
    assigned_branch_ids,
    can_see_all_branches,
    current_user_to_dict,
    get_current_user,
    get_db,
    require_write_access,
    user_can_access_branch,
)
from .models import (
    CalculatedPrice,
    BusinessList,
    BusinessListItem,
    CompetitorCodeMapping,
    CompetitorPrice,
    CompetitorPriceList,
    CompetitorPriceListItem,
    CompetitorPriceListItem,
    ListItem,
    MarkupRange,
    NoCompetitorMarkupRange,
    BendRange,
    BranchCost,
    BranchStock,
    Job,
    PriceSourceAccount,
    PriceFormat,
    PriceFormatCompetitorAssignment,
    PriceList,
    PricingContext,
    PricingWorkflowRun,
    Product,
    ProductExtra,
    ProductRating,
    ProductSubstituteMatch,
    ReferenceUpdateStatus,
    RefreshJob,
    SourceGoodsMatch,
    UniversalList,
    UniversalListPriceFormat,
    Holding,
    Counterparty,
    DeliveryPoint,
    CounterpartyPriceFormat,
    AppUser,
)
from .schemas import (
    CalculatePricesRequest,
    CalculatePricesResponse,
    CreateUniversalListRequest,
    CreateUniversalListResponse,
    UploadExcelResponse,
)
from .services.excel_import import import_excel
from .services.pricing import (
    AMBIGUOUS_LIST_TYPES,
    calculate_price_zone,
    resolve_competitor_price,
    calculate_prices,
    get_markup_percent_by_range,
    lowest_available_competitor_price,
    margin_percent_from_price,
    normalize_list_type,
)
from .services.provisor import ProvisorAuthError, get_prices_by_filial_id
from .services.widman_client import WidmanInvalidCredentialsError
from .services.competitor_persist import persist_phcenter_report, persist_provisor_prices
from .services.competitor_price_lists import (
    export_competitor_price_list,
    get_competitor_price_list_items,
    import_manual_price_list_excel,
    list_competitor_price_lists,
    mark_unified_price_list_checked,
    save_selected_competitor_price_lists_only,
    set_selected_competitor_price_lists,
    sync_selected_competitor_configs,
    upsert_provisor_price_list,
    upsert_unified_price_list,
)
from .services.competitor_assignments import (
    get_all_assigned_competitor_price_lists,
    get_assigned_competitor_price_lists,
    get_assignment,
    price_format_branch_matches,
    upsert_assignment,
)
from .services.price_source_accounts import (
    account_to_dict,
    adapter_for_source,
    credentials_from_row,
    test_account_connection,
    upsert_account,
)
from .services.price_sources import VidmanPriceService
from .services.pricing_rules.rules import (
    apply_pricing_rule_to_format,
    copy_pricing_rule,
    delete_pricing_rule,
    get_pricing_rule,
    list_pricing_rules,
    pricing_rule_application_status,
    pricing_rule_to_dict,
    upsert_pricing_rule,
)
from .services.pricing_rules.templates import (
    copy_template,
    list_rounding_rules,
    list_templates,
    rounding_to_dict,
    template_to_dict,
    upsert_rounding_rule,
    upsert_template,
)
from .services.products_excel_import import import_products_excel
from .services.products_view import get_products_with_competitor_top5
from .services.universal_list_import import (
    business_list_item_to_dict,
    business_list_to_dict,
    import_business_list_excel,
    import_universal_list_excel,
    find_product_by_identifier,
    is_exclude_from_pricing_type,
    max_upload_size_bytes,
    normalize_universal_list_value,
    normalize_universal_list_item_value,
    parse_list_decimal,
)
from .timezone import local_display, local_iso, now_kz_naive
from .services.ph_center_top import FarmcenterTopService
from .services.competitor_matching import (
    PROVISOR_REFERENCE_FILIAL_ID,
    PROVISOR_REFERENCE_FILIAL_IDS,
    explain_unmatched_for_price_list,
    explain_vidman_unmatched,
    parse_drug_structure,
    provisor_item_variants,
    rebuild_competitor_prices_for_selected,
)
from .services.competitor_percentiles import recalculate_competitor_percentiles_if_needed
from .services.sku import normalize_external_sku, normalize_sku, normalize_sku_variants
from .services.competitors.management import list_competitor_sources
from .services.competitors.mappings.read_models import (
    delete_substitute_mapping,
    list_competitor_mappings,
)
from .services.competitors.code_mappings import (
    apply_mapping_to_matching_items,
    find_products_for_mapping,
    list_catalog_code_mappings,
    list_code_mappings,
    mapping_source_payload,
    mapping_to_dict,
    platform_from_value,
    upsert_code_mapping,
)
from .services.competitors.percentiles.read_models import (
    export_percentile_product_rows,
    list_percentile_product_rows,
    list_percentile_sources,
    percentile_coverage_audit,
    percentile_trace,
)
from .services.jobs import create_job, get_active_job, job_to_dict, schedule_job, update_job
from .services.references.batch import import_reference_batch
from .services.references.imports import import_reference_excel
from .services.references.ratings import RATING_DATA_TYPES, import_top_rating_excel
from .services.references.sources import ReferenceFilePayload, make_reference_source
from .services.references.statuses import import_job_to_dict, list_reference_imports, list_reference_statuses, reference_readiness_matrix
from .services.references.templates import build_reference_template, reference_template_filename
from .services.references.types import BRANCHES, REFERENCE_TYPES
from .services.pricing_workflow.analytics import analytics_for_run, build_workflow_analytics
from .services.pricing_workflow.contexts import list_contexts
from .services.pricing_workflow.exports import export_workflow_run
from .services.pricing_workflow.snapshot import loads_snapshot
from .services.pricing_workflow.validation import build_workflow_status
from .services.pricing_workflow.workflow import (
    create_workflow_run,
    list_workflow_competitors,
    list_workflow_price_formats,
    price_format_to_workflow_dict,
    run_to_dict,
)
from .services.provisor_auto_refresh import (
    active_or_stale_refresh_job,
    all_refresh_targets,
    create_skipped_job,
    finish_job as finish_refresh_job,
    heartbeat as refresh_job_heartbeat,
    job_owner_token as refresh_job_owner_token,
    latest_refresh_job,
    normalize_mode as normalize_refresh_mode,
    new_owner_token as new_refresh_owner_token,
    release_scheduler_lock,
    renew_scheduler_lock,
    refresh_job_to_status,
    selected_refresh_targets,
    start_job as start_refresh_job,
    target_counts as refresh_target_counts,
    try_acquire_scheduler_lock,
    try_create_refresh_job,
    update_progress_from_result as update_refresh_job_progress,
)
from .services.emit_worker import (
    EmitConfig,
    EmitWorker,
    configured_filial_ids_for_mode,
    emit_job_to_dict,
    is_emit_plk,
    latest_emit_job,
    list_emit_jobs,
)


logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    force=True,
)
logging.getLogger().setLevel(logging.INFO)
for _logger_name in ("uvicorn", "uvicorn.error", "uvicorn.access", "fastapi", "backend", "backend.app"):
    logging.getLogger(_logger_name).setLevel(logging.INFO)

app = FastAPI(title="aptekaopt-backend", version="0.1.0")
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)
_provisor_auto_refresh_scheduler = None
_provisor_auto_refresh_scheduler_token: str | None = None
_provisor_auto_refresh_scheduler_renew_task = None
_emit_refresh_scheduler = None
_emit_worker: EmitWorker | None = None

settings = get_settings()
PRICE_LISTS_FETCH_TIMEOUT_SECONDS = 30
PRICE_LIST_FETCH_TIMEOUT_SECONDS = 30
PROVISOR_PRICE_TOTAL_TIMEOUT_SECONDS = int(os.getenv("PROVISOR_PRICE_TOTAL_TIMEOUT_SECONDS", "120"))
PRICE_LIST_REFRESH_TTL = timedelta(hours=2)
PROVISOR_PRICE_UNHEALTHY_TIMEOUTS = 3
PROVISOR_PRICE_UNHEALTHY_SKIP_FOR = timedelta(hours=6)
DEFAULT_PROVISOR_EXCLUDED_FILIAL_IDS = "1052,1076,1106,1107,1108,1111,1114,1149,1049"
_provisor_price_health: dict[tuple[str, str], dict[str, object]] = {}


def _timing(operation: str, step: str, started_at: float) -> None:
    logger.info("[TIMING] operation=%s step=%s elapsed_ms=%s", operation, step, round((time.perf_counter() - started_at) * 1000, 2))


def _provisor_health_key(account_id: object, filial_id: object) -> tuple[str, str]:
    return str(account_id or ""), str(filial_id or "")


def _is_provisor_price_unhealthy(*, account_id: object, filial_id: object) -> datetime | None:
    key = _provisor_health_key(account_id, filial_id)
    state = _provisor_price_health.get(key)
    if not state:
        return None
    skip_until = state.get("skip_until")
    if isinstance(skip_until, datetime) and skip_until > now_kz_naive():
        return skip_until
    if isinstance(skip_until, datetime):
        _provisor_price_health.pop(key, None)
    return None


def _record_provisor_price_success(*, account_id: object, filial_id: object) -> None:
    _provisor_price_health.pop(_provisor_health_key(account_id, filial_id), None)


def _record_provisor_price_timeout(*, account_id: object, filial_id: object) -> datetime | None:
    key = _provisor_health_key(account_id, filial_id)
    state = _provisor_price_health.setdefault(key, {"timeouts": 0})
    timeouts = int(state.get("timeouts") or 0) + 1
    state["timeouts"] = timeouts
    if timeouts >= PROVISOR_PRICE_UNHEALTHY_TIMEOUTS:
        skip_until = now_kz_naive() + PROVISOR_PRICE_UNHEALTHY_SKIP_FOR
        state["skip_until"] = skip_until
        return skip_until
    return None


def _parse_id_set(raw: object) -> set[str]:
    if raw is None:
        return set()
    if isinstance(raw, str):
        values = [x.strip() for x in raw.split(",")]
    elif isinstance(raw, (list, tuple, set)):
        values = [str(x).strip() for x in raw]
    else:
        values = [str(raw).strip()]
    return {str(int(x)) for x in values if x and x.lstrip("-").isdigit()}


def _payload_id_set(payload: dict, *keys: str) -> set[str]:
    for key in keys:
        if key in payload:
            return _parse_id_set(payload.get(key))
    return set()


def _refresh_source_from_payload(payload: dict) -> str:
    source = str(payload.get("source") or payload.get("sourceType") or payload.get("source_type") or "all").strip().lower()
    if source not in {"provisor", "vidman", "all"}:
        raise HTTPException(status_code=400, detail="source must be one of: provisor, vidman, all")
    return source


def _refresh_job_type(source: str) -> str:
    return f"refresh_price_lists:{source}"


def _refresh_job_key(format_code: str, source: str) -> str:
    return f"refresh:{format_code}:{source}"


def _get_any_active_refresh_job(db: Session, format_code: str) -> Job | None:
    for source in ("all", "provisor", "vidman"):
        job = get_active_job(db=db, job_type=_refresh_job_type(source), format_code=format_code)
        if job is not None:
            return job
    return get_active_job(db=db, job_type="refresh_price_lists", format_code=format_code)


def _provisor_excluded_filial_ids(account_config: dict[str, object] | None = None) -> set[str]:
    ids = _parse_id_set(os.getenv("PROVISOR_EXCLUDED_FILIAL_IDS", DEFAULT_PROVISOR_EXCLUDED_FILIAL_IDS))
    ids.update(_parse_id_set(os.getenv("EMIT_FILIAL_IDS", ",".join(str(x) for x in settings.emit_filial_ids))))
    config = account_config or {}
    for key in ("excludedFilialIds", "excluded_filial_ids", "heavyFilialIds", "heavy_filial_ids"):
        ids.update(_parse_id_set(config.get(key)))
    return ids

app.add_middleware(
    CORSMiddleware,
    allow_origins=settings.cors_allow_origins if settings.environment != "dev" else ["*"],
    allow_credentials=True,
    allow_methods=["*"] ,
    allow_headers=["*"] ,
)


def _env_flag(name: str) -> bool:
    v = os.getenv(name)
    if v is None:
        return False
    return v.strip().lower() in {"1", "true", "yes", "y", "on"}


def _env_int(name: str, default: int, *, minimum: int = 1) -> int:
    try:
        return max(minimum, int(os.getenv(name, str(default))))
    except Exception:
        return max(minimum, default)


def _seed_price_formats_if_missing() -> None:
    # Dev convenience: ensure at least a couple of price formats exist,
    # otherwise the UI cannot start.
    # Can be disabled via DISABLE_AUTO_SEED=1.
    if settings.environment == "prod":
        return
    if _env_flag("DISABLE_AUTO_SEED"):
        return

    from .db import SessionLocal

    try:
        with SessionLocal() as db:
            existing_codes = set(db.execute(select(PriceFormat.code)).scalars().all())
            to_add = [x for x in data.PRICE_FORMATS if x.get("code") and x["code"] not in existing_codes]
            if not to_add:
                return

            for x in to_add:
                db.add(
                    PriceFormat(
                        code=str(x.get("code", "")).strip(),
                        name=str(x.get("name", "")).strip(),
                        branch=str(x.get("branch", "")).strip(),
                    )
                )
            db.commit()
    except Exception:
        # Do not prevent service startup in dev; just log.
        import traceback

        traceback.print_exc()


@app.on_event("startup")
async def _startup() -> None:
    try:
        init_db()
        _seed_price_formats_if_missing()
        _start_provisor_auto_refresh_scheduler()
        _start_emit_refresh_scheduler()
    except Exception:
        # In production deployments (e.g., Railway) the database might be configured
        # after the first deploy or might be temporarily unavailable.
        # We prefer the service to come up (serve UI + mock data) and log the error.
        import traceback

        traceback.print_exc()
        if settings.environment != "prod":
            raise


def _start_provisor_auto_refresh_scheduler() -> None:
    global _provisor_auto_refresh_scheduler, _provisor_auto_refresh_scheduler_token, _provisor_auto_refresh_scheduler_renew_task
    if settings.environment != "prod":
        logger.info("[PROVISOR_AUTO_REFRESH] scheduler disabled: environment=%s", settings.environment)
        return
    if not settings.provisor_auto_refresh_enabled:
        logger.info("[PROVISOR_AUTO_REFRESH] scheduler disabled by PROVISOR_AUTO_REFRESH_ENABLED=false")
        return
    if _provisor_auto_refresh_scheduler is not None:
        return
    owner_token = new_refresh_owner_token()
    with SessionLocal() as db:
        if not try_acquire_scheduler_lock(db, owner_token=owner_token):
            logger.info("[PROVISOR_AUTO_REFRESH] scheduler not started: another process owns scheduler lease")
            return
    try:
        from apscheduler.schedulers.asyncio import AsyncIOScheduler
        from apscheduler.triggers.cron import CronTrigger
    except Exception:
        with SessionLocal() as db:
            release_scheduler_lock(db, owner_token=owner_token)
        logger.exception("[PROVISOR_AUTO_REFRESH] APScheduler is not installed; scheduler not started")
        return
    mode = normalize_refresh_mode(settings.provisor_auto_refresh_mode)
    trigger = CronTrigger.from_crontab(settings.provisor_auto_refresh_cron)
    scheduler_timezone = settings.provisor_auto_refresh_timezone
    scheduler = AsyncIOScheduler(timezone=scheduler_timezone)
    scheduler.add_job(
        lambda: asyncio.create_task(_start_provisor_refresh_background(mode=mode, requested_by="scheduler")),
        trigger=trigger,
        id="provisor_auto_refresh",
        name="Provisor PLK auto refresh",
        max_instances=1,
        coalesce=True,
        misfire_grace_time=300,
    )
    scheduler.start()
    _provisor_auto_refresh_scheduler = scheduler
    _provisor_auto_refresh_scheduler_token = owner_token
    try:
        _provisor_auto_refresh_scheduler_renew_task = asyncio.create_task(_renew_provisor_scheduler_ownership(owner_token))
    except RuntimeError:
        logger.warning("[PROVISOR_AUTO_REFRESH] scheduler ownership renew task not started: no running event loop")
    logger.info(
        "[PROVISOR_AUTO_REFRESH] scheduler started cron=%s timezone=%s mode=%s max_parallel_accounts=%s max_parallel_plk=%s",
        settings.provisor_auto_refresh_cron,
        scheduler_timezone,
        mode,
        settings.provisor_auto_refresh_max_parallel_accounts,
        settings.provisor_auto_refresh_max_parallel_plk,
    )


async def _renew_provisor_scheduler_ownership(owner_token: str) -> None:
    global _provisor_auto_refresh_scheduler
    while True:
        await asyncio.sleep(30)
        with SessionLocal() as db:
            renewed = renew_scheduler_lock(db, owner_token=owner_token)
        if renewed:
            continue
        logger.error("[PROVISOR_AUTO_REFRESH] scheduler lease lost; shutting down local scheduler")
        scheduler = _provisor_auto_refresh_scheduler
        if scheduler is not None:
            scheduler.shutdown(wait=False)
            _provisor_auto_refresh_scheduler = None
        return


@app.on_event("shutdown")
def _shutdown() -> None:
    _shutdown_provisor_auto_refresh_scheduler()
    _shutdown_emit_refresh_scheduler()


def _shutdown_provisor_auto_refresh_scheduler() -> None:
    global _provisor_auto_refresh_scheduler, _provisor_auto_refresh_scheduler_token, _provisor_auto_refresh_scheduler_renew_task
    if _provisor_auto_refresh_scheduler_renew_task is not None:
        _provisor_auto_refresh_scheduler_renew_task.cancel()
        _provisor_auto_refresh_scheduler_renew_task = None
    if _provisor_auto_refresh_scheduler is not None:
        _provisor_auto_refresh_scheduler.shutdown(wait=False)
        _provisor_auto_refresh_scheduler = None
        logger.info("[PROVISOR_AUTO_REFRESH] scheduler shutdown")
    if _provisor_auto_refresh_scheduler_token:
        with SessionLocal() as db:
            release_scheduler_lock(db, owner_token=_provisor_auto_refresh_scheduler_token)
        _provisor_auto_refresh_scheduler_token = None


def _emit_worker_instance() -> EmitWorker:
    global _emit_worker
    if _emit_worker is None:
        _emit_worker = EmitWorker(session_factory=SessionLocal, config=EmitConfig.from_settings(settings))
    return _emit_worker


def _start_emit_refresh_scheduler() -> None:
    global _emit_refresh_scheduler
    config = EmitConfig.from_settings(settings)
    if not config.enabled:
        logger.info("[EMIT_REFRESH] scheduler disabled")
        return
    if _emit_refresh_scheduler is not None:
        return
    try:
        from apscheduler.schedulers.asyncio import AsyncIOScheduler
        from apscheduler.triggers.cron import CronTrigger
    except Exception:
        logger.exception("[EMIT_REFRESH] APScheduler is not installed; scheduler not started")
        return
    scheduler = AsyncIOScheduler(timezone=config.timezone)
    scheduler.add_job(
        lambda: asyncio.create_task(_start_emit_refresh_background(mode="all", requested_by="scheduler")),
        trigger=CronTrigger.from_crontab(config.cron),
        id="emit_refresh",
        name="Emit/Amity International refresh",
        max_instances=1,
        coalesce=True,
        misfire_grace_time=300,
    )
    scheduler.start()
    _emit_refresh_scheduler = scheduler
    logger.info("[EMIT_REFRESH] scheduler started cron=%s timezone=%s", config.cron, config.timezone)


def _shutdown_emit_refresh_scheduler() -> None:
    global _emit_refresh_scheduler
    if _emit_refresh_scheduler is not None:
        _emit_refresh_scheduler.shutdown(wait=False)
        _emit_refresh_scheduler = None
        logger.info("[EMIT_REFRESH] scheduler shutdown")


async def _start_emit_refresh_background(*, mode: str, requested_by: str, filial_ids: list[int] | None = None) -> None:
    worker = _emit_worker_instance()
    target_ids = configured_filial_ids_for_mode(worker.config, mode=mode, filial_ids=filial_ids)
    if not target_ids:
        logger.warning("[EMIT_REFRESH] skipped: no filial IDs requested")
        return
    with SessionLocal() as db:
        if active_or_stale_refresh_job(db) is not None:
            logger.warning("[EMIT_REFRESH] skipped: normal Provisor refresh is active or stale")
            return
    job, blocker, owner_token = worker.create_job(mode=mode, filial_ids=target_ids, requested_by=requested_by)
    if job is None:
        logger.warning("[EMIT_REFRESH] skipped: another Emit job is active blocker_id=%s", getattr(blocker, "id", None))
        return
    asyncio.create_task(worker.run_job(int(job.id), owner_token=owner_token))


def _fmt_dt(dt: datetime | None) -> str:
    return local_display(dt)


def _fmt_d(d: date | None) -> str:
    if not d:
        return ""
    return d.strftime("%d.%m.%Y")


def _is_sqlite_locked(exc: BaseException) -> bool:
    return "database is locked" in str(exc).lower()


def _is_sqlite_full(exc: BaseException) -> bool:
    text = str(exc).lower()
    return "database or disk is full" in text or "disk is full" in text


def _top_rank_value(row: dict) -> int | None:
    value = row.get("topRank")
    if value is None:
        value = row.get("top_position")
    if value is None:
        value = row.get("is_top")
    if value in (None, ""):
        return None
    try:
        return int(float(str(value).strip()))
    except Exception:
        return None


def _is_top_filter_value(value: object) -> bool:
    if value in (None, ""):
        return False
    try:
        return int(float(str(value).strip())) != 0
    except Exception:
        return str(value).strip() != "0"


def _is_missing_price_value(value: object) -> bool:
    if value in (None, ""):
        return True
    try:
        return float(value) <= 0
    except Exception:
        return False


def _apply_generated_price_export_state(
    rows: list[dict],
    *,
    search: str | None = None,
    top_filter: str = "all",
    top_sort: str = "default",
) -> list[dict]:
    out = list(rows)
    top_filter = (top_filter or "all").strip().lower()
    top_sort = (top_sort or "default").strip().lower()
    q = (search or "").strip().lower()

    if top_filter in {"top", "top_only"}:
        out = [row for row in out if _is_top_filter_value(_top_rank_value(row))]
    elif top_filter in {"no_top", "non_top", "non_top_only"}:
        out = [row for row in out if not _is_top_filter_value(_top_rank_value(row))]

    if q:
        def matches(row: dict) -> bool:
            values = [
                row.get("sku"),
                row.get("name"),
                row.get("manufacturer"),
            ]
            return any(q in str(value or "").lower() for value in values)

        out = [row for row in out if matches(row)]

    if top_sort in {"asc", "desc"}:
        reverse = top_sort == "desc"
        out = sorted(
            out,
            key=lambda row: (
                _top_rank_value(row) is None,
                _top_rank_value(row) if _top_rank_value(row) is not None else 0,
            ),
            reverse=reverse,
        )
        if reverse:
            with_top = [row for row in out if _top_rank_value(row) is not None]
            without_top = [row for row in out if _top_rank_value(row) is None]
            out = with_top + without_top

    return out


def _generated_price_export_rows(
    db: Session,
    price_list_id: str,
    *,
    search: str | None = None,
    top_filter: str = "all",
    top_sort: str = "default",
) -> tuple[PriceList, PriceFormat, list[dict], list[dict]]:
    row = (
        db.execute(
            select(PriceList, PriceFormat)
            .join(PriceFormat, PriceFormat.id == PriceList.price_format_id)
            .where(PriceList.number == price_list_id)
        )
        .first()
    )
    if not row:
        raise HTTPException(status_code=404, detail="price list not found")
    pl, pf = row
    products = get_products_with_competitor_top5(
        db=db,
        price_format_code=pf.code,
        price_list_number=pl.number,
    )
    calculated_skus = set(
        db.execute(
            select(Product.code)
            .join(CalculatedPrice, CalculatedPrice.product_id == Product.id)
            .where(CalculatedPrice.price_list_id == pl.id)
        ).scalars().all()
    )
    products = [row for row in products if str(row.get("sku") or "") in calculated_skus]
    calculated_by_sku = {
        str(product.code): cp
        for cp, product in db.execute(
            select(CalculatedPrice, Product)
            .join(Product, Product.id == CalculatedPrice.product_id)
            .where(CalculatedPrice.price_list_id == pl.id)
        ).all()
    }
    for product_row in products:
        cp = calculated_by_sku.get(str(product_row.get("sku") or ""))
        if cp is not None:
            product_row["modelLog"] = _combined_export_log(db, cp, pf)
    competitor_columns = []
    if products:
        competitor_columns = [
            x
            for x in products[0].get("competitorColumns", [])
            if isinstance(x, dict) and x.get("source")
        ]
    products = _apply_generated_price_export_state(
        products,
        search=search,
        top_filter=top_filter,
        top_sort=top_sort,
    )
    return pl, pf, products, competitor_columns


def _frontend_dist_dir() -> Path | None:
    # Optionally override in Railway via FRONTEND_DIST.
    # Default is repoRoot/front/dist (works in our Docker build).
    import os

    override = os.getenv("FRONTEND_DIST")
    if override:
        p = Path(override).resolve()
        return p if p.exists() else None

    repo_root = Path(__file__).resolve().parents[2]
    dist = repo_root / "front" / "dist"
    return dist if dist.exists() else None


@app.get("/health")
def health():
    return {"status": "ok"}


@app.get("/api/jobs/{job_id}")
def get_job(job_id: str, db: Session = Depends(get_db)):
    job = db.get(Job, job_id)
    if job is None:
        raise HTTPException(status_code=404, detail="job not found")
    return job_to_dict(job)


@app.get("/api/ph-center/prices-analysis")
async def ph_center_prices_analysis(
    region: int = Query(..., ge=1),
    price_mode: int = Query(0, ge=0),
    distributors: str = Query(..., min_length=1),
    persist: bool = Query(False),
    persist_only: bool = Query(False),
    format_code: str | None = Query(None),
    db: Session = Depends(get_db),
):
    """Proxy to ph.center to avoid browser CORS/auth issues.

    Calls: /api/Report/PricesAnalysis?region=..&price_mode=..&distributors=..
    """

    base = settings.phcenter_base_url.rstrip("/")
    url = f"{base}/api/Report/PricesAnalysis"

    headers: dict[str, str] = {}
    token_raw = (settings.phcenter_token or "").strip()
    if not token_raw:
        if persist_only:
            return {"_persist": {"ok": False, "error": "PHCENTER_TOKEN is not configured"}}
        raise HTTPException(status_code=500, detail="PHCENTER_TOKEN is not configured")

    # Accept both forms in env:
    # - token only:        PHCENTER_TOKEN=abc123
    # - full header value: PHCENTER_TOKEN=Bearer abc123
    token = token_raw
    if token.lower().startswith("bearer "):
        token = token[7:].strip()
    if not token:
        if persist_only:
            return {"_persist": {"ok": False, "error": "PHCENTER_TOKEN is not configured"}}
        raise HTTPException(status_code=500, detail="PHCENTER_TOKEN is not configured")

    headers["Authorization"] = f"Bearer {token}"

    distributors_s = distributors.strip()
    if not distributors_s:
        raise HTTPException(status_code=400, detail="distributors is required")

    # Accept comma-separated list of distributor codes.
    parts = [p.strip() for p in distributors_s.split(",") if p.strip()]
    if not parts:
        raise HTTPException(status_code=400, detail="distributors is required")

    parsed: list[int] = []
    for p in parts:
        try:
            n = int(p)
        except Exception:
            raise HTTPException(status_code=400, detail=f"invalid distributor code: {p}")
        if n <= 0:
            raise HTTPException(status_code=400, detail=f"invalid distributor code: {p}")
        parsed.append(n)

    # Normalize (dedupe, preserve order)
    uniq: list[int] = []
    seen: set[int] = set()
    for n in parsed:
        if n in seen:
            continue
        seen.add(n)
        uniq.append(n)

    distributors_param = ",".join(str(n) for n in uniq)

    try:
        async with httpx.AsyncClient(timeout=60.0) as client:
            res = await client.get(
                url,
                params={
                    "region": region,
                    "price_mode": price_mode,
                    "distributors": distributors_param,
                },
                headers=headers,
            )
    except httpx.RequestError as e:
        if persist_only:
            return {"_persist": {"ok": False, "error": f"ph.center request failed: {e}"}}
        raise HTTPException(status_code=502, detail=f"ph.center request failed: {e}")

    if res.status_code in (401, 403):
        if persist_only:
            return {"_persist": {"ok": False, "error": f"ph.center auth failed: HTTP {res.status_code}"}}
        # Auth errors should be visible to UI as-is.
        raise HTTPException(status_code=401, detail={"status": res.status_code, "body": res.text})

    if res.status_code >= 400:
        detail = None
        try:
            detail = res.json()
        except Exception:
            detail = res.text
        if persist_only:
            return {"_persist": {"ok": False, "error": f"ph.center HTTP {res.status_code}"}}
        raise HTTPException(status_code=502, detail={"status": res.status_code, "body": detail})

    try:
        payload = res.json()
    except Exception:
        if persist_only:
            return {"_persist": {"ok": False, "error": "ph.center returned non-JSON response"}}
        raise HTTPException(status_code=502, detail="ph.center returned non-JSON response")

    if persist:
        if not isinstance(format_code, str) or not format_code.strip():
            raise HTTPException(status_code=400, detail="format_code is required when persist=1")
        try:
            stats = persist_phcenter_report(
                db=db,
                price_format_code=format_code.strip(),
                report=payload if isinstance(payload, dict) else {},
                distributor_codes=uniq,
            )
            if isinstance(payload, dict):
                payload["_persist"] = {"ok": True, **stats.to_dict()}
        except Exception as e:
            # Do not break UI rendering if persisting fails.
            if isinstance(payload, dict):
                payload["_persist"] = {"ok": False, "error": str(e)}

    if persist and persist_only:
        # minimal response for background sync
        if isinstance(payload, dict) and isinstance(payload.get("_persist"), dict):
            return {"_persist": payload["_persist"]}
        return {"_persist": {"ok": False, "error": "persist_only: missing persist result"}}

    return payload


@app.get("/api/provisor/prices")
async def provisor_prices(
    filialId: int = Query(..., ge=1),
    persist: bool = Query(False),
    persist_only: bool = Query(False),
    format_code: str | None = Query(None),
    db: Session = Depends(get_db),
):
    """Proxy to api.provisor.kz with server-side auth.

    Calls: /Price/GetByFilialId?filialId=..

    Auth is performed using PROVISOR_LOGIN/PROVISOR_PASSWORD env vars.
    Tokens are cached in-memory and refreshed when needed.
    """

    try:
        base = settings.provisor_base_url.rstrip("/")
        items = await get_prices_by_filial_id(
            base_url=base,
            login=settings.provisor_login,
            password=settings.provisor_password,
            filial_id=filialId,
        )

        if persist:
            if not isinstance(format_code, str) or not format_code.strip():
                raise HTTPException(status_code=400, detail="format_code is required when persist=1")
            try:
                stats = persist_provisor_prices(
                    db=db,
                    price_format_code=format_code.strip(),
                    filial_id=filialId,
                    items=[x for x in items if isinstance(x, dict)],
                )
                upsert_provisor_price_list(
                    db=db,
                    price_format_code=format_code.strip(),
                    filial_id=filialId,
                    items=[x for x in items if isinstance(x, dict)],
                )
                if persist_only:
                    return {"_persist": {"ok": True, **stats.to_dict()}}
                return {"items": items, "_persist": {"ok": True, **stats.to_dict()}}
            except Exception as e:
                if persist_only:
                    return {"_persist": {"ok": False, "error": str(e)}}
                return {"items": items, "_persist": {"ok": False, "error": str(e)}}

        return items
    except ProvisorAuthError as e:
        msg = str(e)
        if "not configured" in msg:
            raise HTTPException(status_code=500, detail=msg)
        raise HTTPException(status_code=502, detail=msg)
    except httpx.ReadTimeout:
        raise HTTPException(status_code=504, detail=f"Provisor timeout for filialId={filialId}")
    except httpx.RequestError as e:
        raise HTTPException(status_code=502, detail=f"Provisor request failed: {e}")
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.get("/api/dashboard")
def get_dashboard():
    # No mock dashboard in MVP: return empty scaffold.
    return {
        "priceFormats": [],
        "recentPriceLists": [],
        "assignments": [],
        "activeLists": [],
        "contractors": [],
    }


def _branch_id_for_name(branch_name: str) -> str:
    normalized = _branch_norm(branch_name)
    return next((str(row["id"]) for row in BRANCHES if _branch_norm(row["name"]) == normalized), "")


def _filter_price_formats_for_user(rows: list[PriceFormat], user: AppUser) -> list[PriceFormat]:
    if can_see_all_branches(user):
        return rows
    return [row for row in rows if user_can_access_branch(user, _branch_id_for_name(row.branch), row.branch)]


def _ensure_price_format_access(pf: PriceFormat, user: AppUser) -> None:
    if not user_can_access_branch(user, _branch_id_for_name(pf.branch), pf.branch):
        raise HTTPException(status_code=403, detail="branch is not assigned to current user")


@app.get("/api/current-user")
def get_current_user_endpoint(current_user: AppUser = Depends(get_current_user)):
    return current_user_to_dict(current_user)


@app.get("/api/price-formats")
def get_price_formats(db: Session = Depends(get_db), current_user: AppUser = Depends(get_current_user)):
    rows = db.execute(select(PriceFormat).order_by(PriceFormat.code.asc())).scalars().all()
    rows = _filter_price_formats_for_user(rows, current_user)
    if not rows:
        return []
    return [{"id": x.id, "name": x.name, "code": x.code, "branch": x.branch} for x in rows]


@app.post("/api/price-formats")
def create_price_format(
    payload: dict = Body(...),
    db: Session = Depends(get_db),
    current_user: AppUser = Depends(require_write_access),
):
    code = str(payload.get("code") or "").strip()
    name = str(payload.get("name") or code).strip()
    branch = str(payload.get("branch") or "").strip()
    if not code:
        raise HTTPException(status_code=400, detail="code is required")
    if not name:
        raise HTTPException(status_code=400, detail="name is required")
    if branch and not user_can_access_branch(current_user, _branch_id_for_name(branch), branch):
        raise HTTPException(status_code=403, detail="branch is not assigned to current user")
    existing = db.execute(select(PriceFormat).where(PriceFormat.code == code)).scalars().first()
    if existing is not None:
        raise HTTPException(status_code=400, detail="price format code already exists")

    pricing_rule_id = payload.get("pricingRuleId") or payload.get("pricing_rule_id")
    pricing_rule_name = str(payload.get("pricingRule") or payload.get("pricing_rule") or "").strip()
    row = PriceFormat(code=code, name=name, branch=branch, pricing_rule=pricing_rule_name)
    if pricing_rule_id not in (None, "", "none"):
        row.pricing_rule_id = int(pricing_rule_id)
        if not pricing_rule_name:
            try:
                rule = get_pricing_rule(db=db, rule_id=int(pricing_rule_id))
                row.pricing_rule = rule.name
            except ValueError:
                raise HTTPException(status_code=400, detail="pricing rule not found")
    db.add(row)
    db.commit()
    db.refresh(row)
    return {"id": row.id, "name": row.name, "code": row.code, "branch": row.branch}


@app.get("/api/pricing-workflow/contexts")
def get_pricing_workflow_contexts(db: Session = Depends(get_db), current_user: AppUser = Depends(get_current_user)):
    contexts = list_contexts(db=db)
    if not can_see_all_branches(current_user):
        contexts = [
            row for row in contexts
            if user_can_access_branch(current_user, str(row.get("branchId") or ""), str(row.get("region") or ""))
        ]
    price_formats = _filter_price_formats_for_user(
        db.execute(select(PriceFormat).order_by(PriceFormat.code.asc())).scalars().all(),
        current_user,
    )
    return {
        "contexts": contexts,
        "priceFormats": [price_format_to_workflow_dict(row) for row in price_formats],
    }


@app.get("/api/pricing-workflow/competitors")
def get_pricing_workflow_competitors(
    price_format_id: int = Query(..., ge=1),
    db: Session = Depends(get_db),
    current_user: AppUser = Depends(get_current_user),
):
    try:
        pf = db.get(PriceFormat, price_format_id)
        if pf is None:
            raise ValueError("price format not found")
        _ensure_price_format_access(pf, current_user)
        payload = list_workflow_competitors(db=db, price_format_id=price_format_id)
        percentile_sources = list_percentile_sources(db=db, price_format_code=pf.code if pf else None)
        payload["percentileSources"] = percentile_sources
        return payload
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))


@app.get("/api/pricing-workflow/status")
def get_pricing_workflow_status(
    pricing_context_id: int = Query(..., ge=1),
    price_format_id: int = Query(..., ge=1),
    competitor_source_ids: str | None = Query(None),
    db: Session = Depends(get_db),
    current_user: AppUser = Depends(get_current_user),
):
    ids: list[int] = []
    if competitor_source_ids:
        for part in competitor_source_ids.split(","):
            part = part.strip()
            if part.isdigit():
                ids.append(int(part))
    pf = db.get(PriceFormat, price_format_id)
    if pf is None:
        raise HTTPException(status_code=404, detail="price format not found")
    _ensure_price_format_access(pf, current_user)
    return build_workflow_status(
        db=db,
        pricing_context_id=pricing_context_id,
        price_format_id=price_format_id,
        competitor_source_ids=ids,
    )


def _branch_norm(value: object) -> str:
    return str(value or "").strip().casefold()


def _date_days_old(value: date | datetime | None) -> int | None:
    if value is None:
        return None
    d = value.date() if isinstance(value, datetime) else value
    return max(0, (date.today() - d).days)


def _workflow_context_for_branch(db: Session, branch_id: str) -> PricingContext:
    branch_id = str(branch_id or "").strip()
    row = (
        db.execute(
            select(PricingContext)
            .where(PricingContext.branch_id == branch_id)
            .where(PricingContext.region == branch_id)
            .where(PricingContext.sales_channel == "default")
        )
        .scalars()
        .first()
    )
    if row is not None:
        return row
    row = PricingContext(
        branch_id=branch_id,
        region=branch_id,
        sales_channel="default",
        name=branch_id or "Без филиала",
        is_active=True,
    )
    db.add(row)
    db.commit()
    db.refresh(row)
    return row


def _format_assignments(db: Session, pf: PriceFormat) -> list[CompetitorPriceList]:
    return [item.price_list for item in get_assigned_competitor_price_lists(db=db, price_format_id=int(pf.id))]


def _latest_price_list_for_format(db: Session, pf: PriceFormat) -> tuple[PriceList | None, int]:
    row = (
        db.execute(
            select(PriceList)
            .where(PriceList.price_format_id == pf.id)
            .order_by(PriceList.created_at.desc(), PriceList.id.desc())
            .limit(1)
        )
        .scalars()
        .first()
    )
    if row is None:
        return None, 0
    sku_count = int(
        db.execute(select(func.count(CalculatedPrice.id)).where(CalculatedPrice.price_list_id == row.id)).scalar() or 0
    )
    return row, sku_count


def _branch_format_row(db: Session, pf: PriceFormat) -> dict:
    assignments = _format_assignments(db, pf)
    latest, sku_count = _latest_price_list_for_format(db, pf)
    return {
        "id": pf.id,
        "code": pf.code,
        "name": pf.name,
        "branch": pf.branch,
        "pricingRule": pf.pricing_rule or "",
        "pricingRuleId": int(pf.pricing_rule_id) if pf.pricing_rule_id is not None else None,
        "roundingRuleId": int(pf.rounding_rule_id) if pf.rounding_rule_id is not None else None,
        "appliedRule": pricing_rule_application_status(db=db, pf=pf),
        "assignedPlkCount": len(assignments),
        "lastGeneratedAt": local_iso(latest.created_at) if latest else "",
        "lastActivationDate": latest.activation_date.isoformat() if latest and latest.activation_date else "",
        "lastRunStatus": latest.status if latest else "нет расчёта",
        "lastPriceListNumber": latest.number if latest else "",
        "lastSkuCount": sku_count,
    }


@app.get("/api/pricing-workflow/branch-formats")
def get_pricing_workflow_branch_formats(
    branch_id: str = Query("", alias="branch_id"),
    db: Session = Depends(get_db),
    current_user: AppUser = Depends(get_current_user),
):
    if branch_id and not user_can_access_branch(current_user, _branch_id_for_name(branch_id), branch_id):
        raise HTTPException(status_code=403, detail="branch is not assigned to current user")
    rows = db.execute(select(PriceFormat).order_by(PriceFormat.code.asc())).scalars().all()
    rows = _filter_price_formats_for_user(rows, current_user)
    if branch_id:
        rows = [row for row in rows if _branch_norm(row.branch) == _branch_norm(branch_id)]
    return [_branch_format_row(db, row) for row in rows]


def _readiness_item(kind: str, label: str, status: str, message: str) -> dict:
    return {"kind": kind, "label": label, "status": status, "message": message}


def _format_readiness(db: Session, pf: PriceFormat, branch_id: str) -> dict:
    product_count = int(db.execute(select(func.count(Product.id))).scalar() or 0)
    stock_count = int(db.execute(select(func.count(BranchStock.id)).where(BranchStock.branch_id == branch_id)).scalar() or 0)
    cost_count = int(db.execute(select(func.count(BranchCost.id)).where(BranchCost.branch_id == branch_id)).scalar() or 0)
    global_rating_count = int(
        db.execute(select(func.count(ProductRating.id)).where(ProductRating.rating_type == "global")).scalar() or 0
    )
    local_rating_count = int(
        db.execute(
            select(func.count(ProductRating.id))
            .where(ProductRating.branch_id == branch_id)
            .where(ProductRating.rating_type == "local")
        ).scalar()
        or 0
    )
    assignments = _format_assignments(db, pf)
    active_assignments = assignments
    stale_sources = [row for row in active_assignments if (_date_days_old(row.price_date) or 999) > 2]
    missing_dates = [row for row in active_assignments if row.price_date is None]
    markup_count = int(db.execute(select(func.count(MarkupRange.id)).where(MarkupRange.price_format_id == pf.id)).scalar() or 0)
    bend_count = int(db.execute(select(func.count(BendRange.id)).where(BendRange.price_format_id == pf.id)).scalar() or 0)
    no_comp_count = int(
        db.execute(select(func.count(NoCompetitorMarkupRange.id)).where(NoCompetitorMarkupRange.price_format_id == pf.id)).scalar() or 0
    )

    items = [
        _readiness_item("stock", "Остатки", "ok" if stock_count > 0 else "warning", f"Строк остатков: {stock_count}"),
        _readiness_item("cost", "Себестоимость", "ok" if cost_count > 0 else "warning", f"Строк себестоимости: {cost_count}"),
        _readiness_item("rating_global", "Рейтинг общий", "ok" if global_rating_count > 0 else "warning", f"Строк общего рейтинга: {global_rating_count}"),
        _readiness_item("rating_local", "Рейтинг локальный", "ok" if local_rating_count > 0 else "warning", f"Строк локального рейтинга: {local_rating_count}"),
        _readiness_item(
            "competitors",
            "Назначенные ПЛК",
            "ok" if active_assignments else "error",
            f"Активных источников: {len(active_assignments)}",
        ),
        _readiness_item(
            "competitor_freshness",
            "Актуальность конкурентов",
            "warning" if stale_sources or missing_dates else "ok",
            f"Устаревших источников: {len(stale_sources)}, без даты: {len(missing_dates)}",
        ),
        _readiness_item(
            "pricing_rule",
            "Правило ЦО",
            "ok" if pf.pricing_rule_id or pf.pricing_rule else "error",
            pf.pricing_rule or ("Правило выбрано" if pf.pricing_rule_id else "Не выбрано правило ЦО"),
        ),
        _readiness_item("markup", "Рекомендованные наценки", "ok" if markup_count > 0 else "warning", f"Диапазонов: {markup_count}"),
        _readiness_item("bend", "Прогибы", "ok" if bend_count > 0 else "warning", f"Диапазонов: {bend_count}"),
        _readiness_item("no_competitor", "Наценки без конкурентов", "ok" if no_comp_count > 0 else "warning", f"Диапазонов: {no_comp_count}"),
        _readiness_item("products", "Товары", "ok" if product_count > 0 else "error", f"Товаров в базе: {product_count}"),
    ]
    errors = [item for item in items if item["status"] == "error"]
    warnings = [item for item in items if item["status"] == "warning"]
    return {
        "formatCode": pf.code,
        "formatName": pf.name,
        "status": "error" if errors else "warning" if warnings else "ok",
        "canGenerate": not errors,
        "items": items,
        "errors": errors,
        "warnings": warnings,
    }


@app.get("/api/pricing-workflow/readiness")
def get_pricing_workflow_readiness(
    branch_id: str = Query("", alias="branch_id"),
    format_codes: str = Query("", alias="format_codes"),
    db: Session = Depends(get_db),
    current_user: AppUser = Depends(get_current_user),
):
    if branch_id and not user_can_access_branch(current_user, _branch_id_for_name(branch_id), branch_id):
        raise HTTPException(status_code=403, detail="branch is not assigned to current user")
    codes = [part.strip() for part in format_codes.split(",") if part.strip()]
    stmt = select(PriceFormat).order_by(PriceFormat.code.asc())
    if codes:
        stmt = stmt.where(PriceFormat.code.in_(codes))
    rows = db.execute(stmt).scalars().all()
    rows = _filter_price_formats_for_user(rows, current_user)
    if branch_id:
        rows = [row for row in rows if _branch_norm(row.branch) == _branch_norm(branch_id)]
    payload = [_format_readiness(db, row, branch_id) for row in rows]
    return {"branchId": branch_id, "items": payload, "canGenerate": all(item["canGenerate"] for item in payload)}


@app.post("/api/pricing-workflow/generate-batch")
def post_pricing_workflow_generate_batch(
    payload: dict = Body(...),
    db: Session = Depends(get_db),
    current_user: AppUser = Depends(require_write_access),
):
    branch_id = str(payload.get("branch_id") or payload.get("branchId") or "").strip()
    if branch_id and not user_can_access_branch(current_user, _branch_id_for_name(branch_id), branch_id):
        raise HTTPException(status_code=403, detail="branch is not assigned to current user")
    codes = payload.get("format_codes") or payload.get("formatCodes") or []
    if not isinstance(codes, list) or not codes:
        raise HTTPException(status_code=400, detail="format_codes must be a non-empty list")
    activation_date = payload.get("activation_date") or payload.get("activationDate") or payload.get("date_start") or payload.get("dateStart")
    comment = str(payload.get("comment") or "")
    context = _workflow_context_for_branch(db, branch_id)
    batch_id = f"batch-{now_kz_naive().strftime('%Y%m%d%H%M%S%f')}"
    items: list[dict] = []
    completed = 0
    failed = 0

    for idx, code in enumerate(codes, start=1):
        pf = db.execute(select(PriceFormat).where(PriceFormat.code == str(code))).scalars().first()
        if pf is None:
            failed += 1
            items.append({"formatCode": code, "status": "failed", "progress": 100, "message": "ЦФ не найден", "error": "price format not found"})
            continue
        _ensure_price_format_access(pf, current_user)
        assignments = _format_assignments(db, pf)
        competitor_sources = [
            {
                "id": row.id,
                "sourceType": row.source_type,
                "name": row.display_name or row.supplier,
                "coefficient": float(row.coefficient or 1),
                "enabled": True,
            }
            for row in assignments
        ]
        try:
            run = create_workflow_run(
                db=db,
                payload={
                    "pricing_context_id": context.id,
                    "price_format_id": pf.id,
                    "competitor_sources": competitor_sources,
                    "percentile_sources": [],
                    "activation_date": activation_date,
                    "user": current_user.username,
                    "comment": comment,
                },
            )
            if run.status == "success":
                completed += 1
                status = "success"
                message = "Прайс-лист сформирован"
                error = ""
            else:
                failed += 1
                status = "failed"
                message = "Расчёт завершился с ошибкой"
                error = run.error or ""
            items.append(
                {
                    "formatCode": pf.code,
                    "formatName": pf.name,
                    "status": status,
                    "progress": 100,
                    "message": message,
                    "error": error,
                    "workflowRunId": run.id,
                    "priceListNumber": run.price_list_number,
                }
            )
        except Exception as exc:
            db.rollback()
            failed += 1
            items.append(
                {
                    "formatCode": pf.code,
                    "formatName": pf.name,
                    "status": "failed",
                    "progress": 100,
                    "message": "Ошибка запуска расчёта",
                    "error": str(exc),
                }
            )

    result = {
        "id": batch_id,
        "branchId": branch_id,
        "status": "failed" if failed and not completed else "warning" if failed else "success",
        "totalFormats": len(codes),
        "completedFormats": completed,
        "failedFormats": failed,
        "startedAt": local_iso(now_kz_naive()),
        "finishedAt": local_iso(now_kz_naive()),
        "items": items,
    }
    if failed and not completed:
        raise HTTPException(status_code=400, detail={"message": "Не удалось сформировать выбранные ценовые форматы", "items": items})
    return result


@app.get("/api/pricing-workflow/runs/{run_id}")
def get_pricing_workflow_run_status(
    run_id: int,
    db: Session = Depends(get_db),
    current_user: AppUser = Depends(get_current_user),
):
    run = db.get(PricingWorkflowRun, run_id)
    if run is None:
        raise HTTPException(status_code=404, detail="workflow run not found")
    pf = db.get(PriceFormat, run.price_format_id)
    if pf is not None:
        _ensure_price_format_access(pf, current_user)
    return run_to_dict(db=db, run=run, include_items=False)


@app.get("/api/pricing-workflow/recent-results")
def get_pricing_workflow_recent_results(
    branch_id: str = Query("", alias="branch_id"),
    db: Session = Depends(get_db),
    current_user: AppUser = Depends(get_current_user),
):
    if branch_id and not user_can_access_branch(current_user, _branch_id_for_name(branch_id), branch_id):
        raise HTTPException(status_code=403, detail="branch is not assigned to current user")
    stmt = (
        select(PriceList, PriceFormat)
        .join(PriceFormat, PriceFormat.id == PriceList.price_format_id)
        .order_by(PriceList.created_at.desc(), PriceList.id.desc())
        .limit(30)
    )
    rows = db.execute(stmt).all()
    rows = [(pl, pf) for pl, pf in rows if user_can_access_branch(current_user, _branch_id_for_name(pf.branch), pf.branch)]
    if branch_id:
        rows = [(pl, pf) for pl, pf in rows if _branch_norm(pf.branch) == _branch_norm(branch_id)]
    out = []
    for pl, pf in rows[:15]:
        sku_count = int(db.execute(select(func.count(CalculatedPrice.id)).where(CalculatedPrice.price_list_id == pl.id)).scalar() or 0)
        out.append(
            {
                "number": pl.number,
                "format": pf.code,
                "branch": pf.branch,
                "date": _fmt_dt(pl.created_at),
                "activationDate": _fmt_d(pl.activation_date),
                "user": pl.user,
                "status": pl.status,
                "skuCount": sku_count,
            }
        )
    return out


def _price_list_by_identifier(db: Session, price_list_id: str) -> tuple[PriceList, PriceFormat]:
    condition = PriceList.number == price_list_id
    if str(price_list_id).isdigit():
        condition = condition | (PriceList.id == int(price_list_id))
    row = (
        db.execute(
            select(PriceList, PriceFormat)
            .join(PriceFormat, PriceFormat.id == PriceList.price_format_id)
            .where(condition)
        )
        .first()
    )
    if row is None:
        raise HTTPException(status_code=404, detail="price list not found")
    return row[0], row[1]


def _generated_price_list_summary(db: Session, pl: PriceList, pf: PriceFormat) -> dict:
    sku_count = int(db.execute(select(func.count(CalculatedPrice.id)).where(CalculatedPrice.price_list_id == pl.id)).scalar() or 0)
    with_competitors = int(
        db.execute(
            select(func.count(CalculatedPrice.id))
            .where(CalculatedPrice.price_list_id == pl.id)
            .where(CalculatedPrice.competitor_price.is_not(None))
        ).scalar()
        or 0
    )
    return {
        "id": pl.id,
        "number": pl.number,
        "format": pf.code,
        "formatName": pf.name,
        "branch": pf.branch,
        "pricingRule": pf.pricing_rule or "",
        "pricingRuleId": int(pf.pricing_rule_id) if pf.pricing_rule_id is not None else None,
        "date": _fmt_dt(pl.created_at),
        "createdAt": local_iso(pl.created_at) if pl.created_at else "",
        "activationDate": _fmt_d(pl.activation_date),
        "user": pl.user,
        "generatedBy": getattr(pl, "generated_by", "") or pl.user,
        "skuCount": sku_count,
        "withCompetitors": with_competitors,
        "withoutCompetitors": max(0, sku_count - with_competitors),
        "status": pl.status or "generated",
        "hasSnapshot": bool(getattr(pl, "run_snapshot_json", "") and getattr(pl, "run_snapshot_json", "") != "{}"),
        "revision": "",
        "comment": "",
    }


@app.get("/api/generated-price-lists")
def get_generated_price_lists(
    branch: str | None = Query(None),
    format_code: str | None = Query(None),
    date_from: str | None = Query(None),
    date_to: str | None = Query(None),
    search: str | None = Query(None),
    db: Session = Depends(get_db),
    current_user: AppUser = Depends(get_current_user),
):
    stmt = (
        select(PriceList, PriceFormat)
        .join(PriceFormat, PriceFormat.id == PriceList.price_format_id)
        .order_by(PriceList.created_at.desc(), PriceList.id.desc())
    )
    if branch:
        stmt = stmt.where(PriceFormat.branch == branch)
    if format_code:
        stmt = stmt.where(PriceFormat.code == format_code)
    if date_from:
        try:
            stmt = stmt.where(PriceList.created_at >= datetime.fromisoformat(date_from))
        except Exception:
            pass
    if date_to:
        try:
            stmt = stmt.where(PriceList.created_at < datetime.fromisoformat(date_to) + timedelta(days=1))
        except Exception:
            pass
    if search:
        like = f"%{search.strip()}%"
        stmt = stmt.where((PriceList.number.ilike(like)) | (PriceFormat.code.ilike(like)) | (PriceFormat.name.ilike(like)))
    rows = db.execute(stmt.limit(200)).all()
    rows = [(pl, pf) for pl, pf in rows if user_can_access_branch(current_user, _branch_id_for_name(pf.branch), pf.branch)]
    return [_generated_price_list_summary(db, pl, pf) for pl, pf in rows]


@app.get("/api/generated-price-lists/{price_list_id}")
def get_generated_price_list_card(
    price_list_id: str,
    db: Session = Depends(get_db),
    current_user: AppUser = Depends(get_current_user),
):
    pl, pf = _price_list_by_identifier(db, price_list_id)
    _ensure_price_format_access(pf, current_user)
    analytics = build_workflow_analytics(db=db, price_list_id=pl.id)
    payload = _generated_price_list_summary(db, pl, pf)
    payload["analytics"] = analytics
    payload["snapshot"] = analytics.get("snapshot") or {}
    payload["runSources"] = analytics.get("runSources") or {}
    payload["runRule"] = analytics.get("runRule") or {}
    payload["runLists"] = analytics.get("runLists") or []
    payload["runReferenceVersions"] = analytics.get("runReferenceVersions") or {}
    payload["runPercentileConfig"] = analytics.get("runPercentileConfig") or {}
    return payload


def _source_label(source_name: str) -> str:
    if not source_name:
        return ""
    if source_name.startswith("percentile:"):
        return source_name.replace("percentile:", "Персентиль: ")
    if ":" in source_name:
        source, key = source_name.split(":", 1)
        return f"{source} {key}"
    return source_name


def _zone_reference_for_saved_row(
    db: Session | None,
    cp: CalculatedPrice,
    price_format_id: int | None = None,
    product_id: int | None = None,
) -> object:
    stored = cp.lowest_competitor_price if cp.lowest_competitor_price is not None else cp.competitor_price
    if stored is not None or db is None or price_format_id is None:
        return stored
    return lowest_available_competitor_price(db, int(price_format_id), int(product_id or cp.product_id))


def _calculated_zone(
    cp: CalculatedPrice,
    db: Session | None = None,
    price_format_id: int | None = None,
    product_id: int | None = None,
) -> str | None:
    zone, _reference, _deviation = calculate_price_zone(
        cp.final_price,
        chosen_competitor_price=cp.chosen_competitor_price,
        lowest_competitor_price=_zone_reference_for_saved_row(db, cp, price_format_id, product_id),
    )
    return zone or "no-data"


def _calculated_zone_reference_and_deviation(
    cp: CalculatedPrice,
    db: Session | None = None,
    price_format_id: int | None = None,
    product_id: int | None = None,
) -> tuple[float | None, float | None]:
    _zone, reference, deviation = calculate_price_zone(
        cp.final_price,
        chosen_competitor_price=cp.chosen_competitor_price,
        lowest_competitor_price=_zone_reference_for_saved_row(db, cp, price_format_id, product_id),
    )
    return (
        float(reference) if reference is not None else None,
        float(deviation * 100) if deviation is not None else None,
    )


def _applied_list_summary(db: Session, cp: CalculatedPrice) -> dict | None:
    list_id = getattr(cp, "applied_list_id", None)
    if list_id is None:
        return None
    row = db.get(UniversalList, int(list_id))
    return {
        "applied_rule_type": cp.applied_rule_type or "",
        "applied_rule_value": float(cp.applied_rule_value) if cp.applied_rule_value is not None else None,
        "list_id": int(list_id),
        "list_name": row.name if row else str(list_id),
        "list_code": row.code if row else "",
        "ambiguous": (cp.applied_rule_type or "") in AMBIGUOUS_LIST_TYPES,
    }


def _list_override_log(
    applied: dict | None,
    *,
    global_margin_percent: float | None = None,
    pricing_margin_percent: float | None = None,
    final_price: float | None = None,
) -> dict | None:
    if not applied:
        return None
    list_type = str(applied.get("applied_rule_type") or "")
    value = applied.get("applied_rule_value")
    display_value = f"{value:g}%" if value is not None and list_type in {"fixed_markup", "critical_markup"} else f"{value:g}" if value is not None else "—"
    actions = {
        "fixed_markup": "МДЦ рассчитана по марже из списка и применена как финальная цена; конкуренты и прогиб не применялись.",
        "critical_markup": "Маржа для расчёта заменена значением списка.",
        "fixed_price": "Применена фиксированная цена из списка.",
        "min_price": "Применено ограничение минимальной цены.",
        "max_price": "Применено ограничение максимальной цены.",
        "no_bend": "Прогиб отключён значением списка.",
        "exclude_from_pricing": "Позиция исключена из расчёта значением списка.",
    }
    affected_fields = {
        "fixed_markup": "Маржа для расчёта МДЦ",
        "critical_markup": "Маржа для расчёта МДЦ",
        "fixed_price": "Финальная цена",
        "min_price": "Минимальная финальная цена",
        "max_price": "Максимальная финальная цена",
        "no_bend": "Прогиб",
        "exclude_from_pricing": "Участие в расчёте и экспорте",
    }
    action = actions.get(list_type, "Применено значение активного списка.")
    if list_type == "fixed_markup" and pricing_margin_percent is not None:
        action = f"МДЦ рассчитана по марже из списка {pricing_margin_percent:g}% и применена как финальная цена; конкуренты и прогиб не применялись."
    elif list_type == "critical_markup" and pricing_margin_percent is not None:
        if global_margin_percent is not None:
            action = (
                f"Маржа для расчёта переопределена с {global_margin_percent:g}% "
                f"на {pricing_margin_percent:g}%."
            )
        else:
            action = f"Маржа для расчёта установлена в {pricing_margin_percent:g}%."
    list_changed_final_price = None
    list_effect_message = action
    if list_type == "min_price" and value is not None and final_price is not None:
        list_changed_final_price = final_price <= float(value)
        if final_price > float(value):
            list_effect_message = "Список проверен, но не изменил цену, потому что рассчитанная цена уже выше минимальной цены."
            action = list_effect_message
    elif list_type == "max_price" and value is not None and final_price is not None:
        list_changed_final_price = final_price >= float(value)
        if final_price < float(value):
            list_effect_message = "Список проверен, но не изменил цену, потому что рассчитанная цена уже ниже максимальной цены."
            action = list_effect_message
    elif list_type == "fixed_price":
        list_changed_final_price = True
    return {
        "listName": applied.get("list_name") or "",
        "listCode": applied.get("list_code") or "",
        "listType": list_type,
        "value": value,
        "displayValue": display_value,
        "affectedField": affected_fields.get(list_type, "Параметр расчёта"),
        "action": action,
        "ambiguous": bool(applied.get("ambiguous")),
        "listMatched": True,
        "listApplied": True,
        "listChangedFinalPrice": list_changed_final_price,
        "listEffectMessage": list_effect_message,
    }


def _list_override_log_text(log: dict | None) -> str:
    if not log:
        return ""
    lines = [
        "Применен список:",
        f"Название: {log.get('listName') or ''}",
        f"Код: {log.get('listCode') or ''}",
        f"Тип: {log.get('listType') or ''}",
        f"Значение: {log.get('displayValue') or ''}",
    ]
    message = str(log.get("listEffectMessage") or log.get("action") or "").strip()
    if message:
        lines.extend(["", message])
    return "\n".join(lines)


def _combined_export_log(db: Session, cp: CalculatedPrice, pf: PriceFormat) -> str:
    pricing_log = str(cp.applied_reason or "")
    applied_list = _applied_list_summary(db, cp)
    global_margin_percent = None
    try:
        global_margin_percent = float(get_markup_percent_by_range(db, pf.id, Decimal(str(float(cp.cost or 0)))) * Decimal("100"))
    except ValueError:
        pass
    list_log = _list_override_log(
        applied_list,
        global_margin_percent=global_margin_percent,
        pricing_margin_percent=float(cp.mdc_markup_percent)
        if getattr(cp, "mdc_markup_percent", None) is not None
        else float(cp.markup_percent_used)
        if cp.markup_percent_used is not None
        else None,
        final_price=float(cp.final_price) if cp.final_price is not None else None,
    )
    list_text = _list_override_log_text(list_log)
    return "\n\n".join(part for part in [pricing_log, list_text] if part)


def _pricing_log(db: Session, cp: CalculatedPrice, product: Product, source: CompetitorPrice | None) -> list[dict]:
    cost = float(cp.cost or 0)
    final = float(cp.final_price or 0)
    margin = margin_percent_from_price(cost, final)
    pricing_margin = (
        float(cp.mdc_markup_percent)
        if getattr(cp, "mdc_markup_percent", None) is not None
        else float(cp.markup_percent_used)
        if cp.markup_percent_used is not None
        else (float(cp.base_price or 0) - cost) / cost * 100
        if cost
        else None
    )
    log = [
        {"label": "Себестоимость", "value": round(cost, 4), "description": "Загруженная себестоимость товара на момент расчёта."},
        {"label": "Маржа (наценка)", "value": f"{pricing_margin:.2f}%" if pricing_margin is not None else "нет данных", "description": "Эффективная маржа, использованная pricing engine для расчёта МДЦ."},
        {"label": "Фактическая маржа", "value": f"{margin:.2f}%" if margin is not None else "нет данных", "description": "Маржа по финальной цене и себестоимости."},
        {"label": "МДЦ", "value": float(cp.base_price or 0), "description": "Минимально допустимая цена, сохранённая в результате расчёта."},
        {
            "label": "Конкуренты",
            "value": float(cp.competitor_price) if cp.competitor_price is not None else "нет цены конкурента",
            "description": f"Источник: {_source_label(source.source_name) if source else 'не определён'}",
        },
        {"label": "Прогиб", "value": float(cp.price_from_competitor) if cp.price_from_competitor is not None else "не применён", "description": "Цена после применения прогиба к конкурентной цене, если она была рассчитана."},
        {"label": "Выбор конкурента", "value": _source_label(source.source_name) if source else "не применялся", "description": "Для отображения выбран ближайший сохранённый конкурентный источник по товару."},
        {"label": "Отклонения", "value": _calculated_zone(cp), "description": "Зона показывает положение финальной цены относительно конкурента."},
        {"label": "Percentile", "value": "сработал" if source and str(source.source_name or "").startswith("percentile:") else "не сработал", "description": _source_label(source.source_name) if source else ""},
        {"label": "Финальная причина", "value": cp.applied_reason or "", "description": "Человекочитаемый лог, сохранённый pricing engine."},
    ]
    return log


def _generated_item_dict(
    db: Session,
    cp: CalculatedPrice,
    product: Product,
    pf: PriceFormat,
    extra: ProductExtra | None,
    ratings: dict[str, int | None] | None = None,
) -> dict:
    cost = float(cp.cost or 0)
    final = float(cp.final_price or 0)
    actual_margin = margin_percent_from_price(cost, final)
    actual_margin_percent = round(float(actual_margin), 2) if actual_margin is not None else None
    source_label = _source_label(cp.applied_source_name or "")
    applied_list = _applied_list_summary(db, cp)
    global_margin_percent = None
    try:
        global_margin_percent = float(get_markup_percent_by_range(db, pf.id, Decimal(str(cost))) * Decimal("100"))
    except ValueError:
        pass
    list_override_log = _list_override_log(
        applied_list,
        global_margin_percent=global_margin_percent,
        pricing_margin_percent=float(cp.mdc_markup_percent)
        if getattr(cp, "mdc_markup_percent", None) is not None
        else float(cp.markup_percent_used)
        if cp.markup_percent_used is not None
        else None,
        final_price=final,
    )
    ratings = ratings if ratings is not None else _product_ratings_by_id(db, [int(product.id)], str(pf.branch or "")).get(int(product.id), {})
    global_rating = ratings.get("global")
    local_rating = ratings.get("local")
    zone_reference = _zone_reference_for_saved_row(db, cp, pf.id, product.id)
    return {
        "sku": product.code,
        "name": product.name,
        "topRank": int(product.top_rank) if product.top_rank is not None else None,
        "isTop": product.top_rank is not None,
        "is_top": int(product.top_rank) if product.top_rank is not None else 0,
        "globalRating": global_rating,
        "localRating": local_rating,
        "global_rating": global_rating,
        "local_rating": local_rating,
        "manufacturer": extra.manufacturer if extra else "",
        "stock": float(extra.stock) if extra and extra.stock is not None else None,
        "cost": cost,
        "basePrice": float(cp.base_price or 0),
        "mdc": float(cp.base_price or 0),
        "bestCompetitorPrice": float(cp.competitor_price) if cp.competitor_price is not None else None,
        "lowestCompetitorPrice": float(zone_reference) if zone_reference is not None else None,
        "chosenCompetitorPrice": float(cp.chosen_competitor_price) if cp.chosen_competitor_price is not None else None,
        "selectedCompetitorPrice": float(cp.chosen_competitor_price) if cp.chosen_competitor_price is not None else None,
        "priceAfterBend": float(cp.price_from_competitor) if cp.price_from_competitor is not None else None,
        "bendPercentUsed": float(cp.bend_percent_used) if cp.bend_percent_used is not None else None,
        "effectiveMarkupPercent": float(cp.markup_percent_used) if cp.markup_percent_used is not None else None,
        "markupPercentUsed": float(cp.markup_percent_used) if cp.markup_percent_used is not None else None,
        "mdcMarkupPercent": float(cp.mdc_markup_percent) if getattr(cp, "mdc_markup_percent", None) is not None else None,
        "mdcPrice": float(cp.mdc_price) if getattr(cp, "mdc_price", None) is not None else None,
        "competitorCandidatePrice": float(cp.competitor_candidate_price) if getattr(cp, "competitor_candidate_price", None) is not None else None,
        "finalPrice": final,
        "markupPercent": float(cp.markup_percent_used) if cp.markup_percent_used is not None else actual_margin_percent,
        "actualMarginPercent": actual_margin_percent,
        "zone": _calculated_zone(cp, db, pf.id, product.id),
        "priceSource": source_label,
        "appliedSourceName": cp.applied_source_name or "",
        "appliedSourceType": cp.applied_source_type or "",
        "percentileSource": source_label if cp.used_percentile else "",
        "usedPercentile": bool(cp.used_percentile),
        "usedSubstitute": bool(cp.used_substitute),
        "appliedListIds": cp.applied_list_ids or "[]",
        "appliedRuleType": cp.applied_rule_type or "",
        "appliedRuleValue": float(cp.applied_rule_value) if cp.applied_rule_value is not None else None,
        "appliedListId": int(cp.applied_list_id) if cp.applied_list_id is not None else None,
        "appliedListName": applied_list["list_name"] if applied_list else "",
        "appliedRuleAmbiguous": bool(applied_list and applied_list["ambiguous"]),
        "appliedRule": applied_list,
        "pricingCalculationLog": cp.applied_reason or "",
        "listOverrideLog": list_override_log,
        "ratingGlobal": global_rating,
        "ratingLocal": local_rating,
        "pricingReason": cp.applied_reason or "",
        "pricingRule": cp.applied_rule_name or pf.pricing_rule or "",
        "pricingRuleVersion": cp.applied_rule_version or "",
        "log": _pricing_log(db, cp, product, None),
    }


def _product_ratings_by_id(db: Session, product_ids: list[int], branch_id: str) -> dict[int, dict[str, int | None]]:
    """Return the newest global and branch-local ProductRating values per product."""
    if not product_ids:
        return {}
    rows = (
        db.execute(
            select(ProductRating)
            .where(ProductRating.product_id.in_(product_ids))
            .where(
                (ProductRating.rating_type == "global")
                | ((ProductRating.rating_type == "local") & (ProductRating.branch_id == branch_id))
            )
            .order_by(ProductRating.updated_at.desc(), ProductRating.id.desc())
        )
        .scalars()
        .all()
    )
    result: dict[int, dict[str, int | None]] = {}
    for row in rows:
        ratings = result.setdefault(int(row.product_id), {})
        if row.rating_type not in ratings:
            ratings[row.rating_type] = int(row.rating) if row.rating is not None else None
    return result


def _competitor_column_key(source_type: object, source_key: object) -> str:
    return f"{str(source_type or '').strip()}:{str(source_key or '').strip()}"


def _competitor_column_title(value: object, fallback: object = "") -> str:
    """Remove only adjacent duplicate labels from a rendered competitor title."""
    raw = str(value or fallback or "").strip()
    parts = [part.strip() for part in raw.split("—") if part.strip()]
    unique_parts: list[str] = []
    for part in parts:
        if unique_parts and unique_parts[-1].casefold() == part.casefold():
            continue
        unique_parts.append(part)
    return " — ".join(unique_parts) if unique_parts else raw


def _competitor_columns_for_price_list(db: Session, pl: PriceList, pf: PriceFormat) -> list[dict]:
    run_sources = loads_snapshot(pl.run_sources_json, {})
    selected = (run_sources.get("selectedCompetitorSources") or []) if isinstance(run_sources, dict) else []
    columns: list[dict] = []
    if isinstance(selected, list) and selected:
        for index, source in enumerate(selected):
            if not isinstance(source, dict):
                continue
            key = _competitor_column_key(source.get("sourceType"), source.get("sourceKey"))
            if key == ":":
                continue
            title = _competitor_column_title(
                source.get("displayName") or source.get("competitorName") or source.get("supplier"),
                key,
            )
            columns.append(
                {
                    "id": source.get("id") or index + 1,
                    "key": key,
                    "title": title,
                    "sourceType": source.get("sourceType") or "",
                    "priceListId": source.get("id"),
                    "competitorName": source.get("competitorName") or source.get("supplier") or title,
                    "sourceKey": source.get("sourceKey") or "",
                    "coefficient": float(source.get("coefficient") or 1),
                    "priceDate": source.get("priceDate") or "",
                }
            )
    if columns:
        seen: set[str] = set()
        out: list[dict] = []
        for column in columns:
            if column["key"] in seen:
                continue
            seen.add(column["key"])
            out.append(column)
        return out

    return [
        {
            "id": item.price_list.id,
            "key": _competitor_column_key(item.price_list.source_type, item.price_list.source_key),
            "title": _competitor_column_title(
                item.price_list.display_name or item.price_list.competitor_name or item.price_list.supplier,
                item.price_list.source_key,
            ),
            "sourceType": item.price_list.source_type,
            "priceListId": item.price_list.id,
            "competitorName": item.price_list.competitor_name or item.price_list.supplier or item.price_list.display_name,
            "sourceKey": item.price_list.source_key,
            "coefficient": float(item.assignment.coefficient or 1),
            "priceDate": item.price_list.price_date.isoformat() if item.price_list.price_date else "",
        }
        for item in get_assigned_competitor_price_lists(db=db, price_format_id=int(pf.id))
    ]


def _competitor_prices_by_product(
    db: Session,
    *,
    pf: PriceFormat,
    product_ids: list[int],
    columns: list[dict],
) -> dict[int, dict[str, dict]]:
    if not product_ids or not columns:
        return {}
    source_names = [str(column.get("key") or "") for column in columns if column.get("key")]
    coefficient_by_source = {str(column.get("key") or ""): float(column.get("coefficient") or 1) for column in columns}
    rows = (
        db.execute(
            select(CompetitorPrice)
            .where(CompetitorPrice.price_format_id == pf.id)
            .where(CompetitorPrice.product_id.in_(product_ids))
            .where(CompetitorPrice.source_name.in_(source_names))
            .where(CompetitorPrice.source_price.is_not(None))
            .order_by(CompetitorPrice.product_id.asc(), CompetitorPrice.source_name.asc(), CompetitorPrice.source_price.asc())
        )
        .scalars()
        .all()
    )
    out: dict[int, dict[str, dict]] = {}
    for row in rows:
        product_id = int(row.product_id or 0)
        source_name = row.source_name or ""
        raw_price = float(row.source_price) if row.source_price is not None else None
        if not product_id or not source_name or raw_price is None:
            continue
        coefficient = coefficient_by_source.get(source_name, 1.0)
        price = raw_price * coefficient
        current = out.setdefault(product_id, {}).get(source_name)
        if current is not None and current.get("price") is not None and float(current["price"]) <= price:
            continue
        match_type = row.match_type or ""
        out[product_id][source_name] = {
            "price": price,
            "sourcePrice": raw_price,
            "coefficient": coefficient,
            "sourceName": source_name,
            "matchedBy": match_type,
            "isManualMapping": match_type == "manual_code_mapping",
            "isSubstitute": match_type == "provisor_manual_substitute",
        }
    return out


@app.get("/api/generated-price-lists/{price_list_id}/items")
def get_generated_price_list_items(
    price_list_id: str,
    q: str | None = Query(None),
    zone: str | None = Query(None),
    top_filter: str = Query("all"),
    sort: str | None = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(100, ge=20, le=500),
    db: Session = Depends(get_db),
    current_user: AppUser = Depends(get_current_user),
):
    pl, pf = _price_list_by_identifier(db, price_list_id)
    _ensure_price_format_access(pf, current_user)
    stmt = select(CalculatedPrice, Product).join(Product, Product.id == CalculatedPrice.product_id).where(CalculatedPrice.price_list_id == pl.id)
    if q:
        like = f"%{q.strip()}%"
        stmt = stmt.where((Product.code.ilike(like)) | (Product.name.ilike(like)))
    if sort == "final_desc":
        stmt = stmt.order_by(CalculatedPrice.final_price.desc())
    elif sort == "final_asc":
        stmt = stmt.order_by(CalculatedPrice.final_price.asc())
    elif sort == "markup_desc":
        stmt = stmt.order_by((CalculatedPrice.final_price - CalculatedPrice.cost).desc())
    else:
        stmt = stmt.order_by(Product.name.asc(), Product.code.asc())
    all_rows = db.execute(stmt).all()
    if zone and zone != "__all__":
        all_rows = [
            (cp, product)
            for cp, product in all_rows
            if _calculated_zone(cp, db, pf.id, product.id) == zone
        ]
    ratings_by_product = _product_ratings_by_id(
        db, [int(product.id) for _, product in all_rows], str(pf.branch or "")
    )
    top_filter_normalized = (top_filter or "all").strip().lower()
    if top_filter_normalized in {"top", "top_only"}:
        all_rows = [(cp, product) for cp, product in all_rows if ratings_by_product.get(int(product.id), {}).get("global") is not None]
    elif top_filter_normalized in {"no_top", "non_top", "non_top_only"}:
        all_rows = [(cp, product) for cp, product in all_rows if ratings_by_product.get(int(product.id), {}).get("global") is None]
    total = len(all_rows)
    rows = all_rows[(page - 1) * page_size : page * page_size]
    product_ids = [int(product.id) for _, product in rows]
    extras = {
        row.product_id: row
        for row in db.execute(select(ProductExtra).where(ProductExtra.product_id.in_(product_ids))).scalars().all()
    } if rows else {}
    competitor_columns = _competitor_columns_for_price_list(db, pl, pf)
    competitor_prices = _competitor_prices_by_product(db, pf=pf, product_ids=product_ids, columns=competitor_columns)
    items = []
    for cp, product in rows:
        item = _generated_item_dict(
            db, cp, product, pf, extras.get(product.id), ratings_by_product.get(int(product.id), {})
        )
        item["competitorPrices"] = competitor_prices.get(int(product.id), {})
        items.append(item)
    return {
        "competitorColumns": competitor_columns,
        "items": items,
        "page": page,
        "pageSize": page_size,
        "total": total,
    }


@app.get("/api/generated-price-lists/{price_list_id}/analytics")
def get_generated_price_list_analytics(
    price_list_id: str,
    db: Session = Depends(get_db),
    current_user: AppUser = Depends(get_current_user),
):
    pl, pf = _price_list_by_identifier(db, price_list_id)
    _ensure_price_format_access(pf, current_user)
    return build_workflow_analytics(db=db, price_list_id=pl.id)


@app.get("/api/generated-price-lists/{price_list_id}/export.{fmt}")
def export_generated_price_list(
    price_list_id: str,
    fmt: str,
    top_filter: str = Query("all"),
    db: Session = Depends(get_db),
    current_user: AppUser = Depends(get_current_user),
):
    if fmt not in {"csv", "xlsx"}:
        raise HTTPException(status_code=400, detail="fmt must be csv or xlsx")
    pl, pf = _price_list_by_identifier(db, price_list_id)
    _ensure_price_format_access(pf, current_user)
    items_payload = get_generated_price_list_items(
        price_list_id=price_list_id,
        q=None,
        zone=None,
        top_filter=top_filter,
        sort=None,
        page=1,
        page_size=100000,
        db=db,
        current_user=current_user,
    )
    rows_payload = items_payload["items"]
    competitor_columns = items_payload.get("competitorColumns") or []
    headers = [
        ("sku", "SKU"),
        ("name", "Наименование"),
        ("globalRating", "Рейтинг глобальный"),
        ("localRating", "Рейтинг локальный"),
        ("manufacturer", "Производитель"),
        ("stock", "Остаток"),
        ("cost", "Себестоимость"),
        ("basePrice", "Базовая цена"),
        ("mdc", "МДЦ"),
        ("bestCompetitorPrice", "Лучшая цена конкурента"),
        ("lowestCompetitorPrice", "Мин. цена конкурента"),
        ("chosenCompetitorPrice", "Выбранная цена конкурента"),
        ("priceAfterBend", "Цена после прогиба"),
        ("bendPercentUsed", "Прогиб %"),
        ("markupPercentUsed", "Наценка правила %"),
        ("finalPrice", "Финальная цена"),
        ("markupPercent", "Наценка %"),
        ("actualMarginPercent", "Фактическая маржа %"),
        ("zone", "Зона"),
        ("priceSource", "Источник цены"),
        ("percentileSource", "Источник персентиля"),
        ("pricingReason", "Причина применения цены"),
        ("pricingCalculationLog", "Лог расчета цены"),
        ("listOverrideLogText", "Лог применения списка"),
        ("listOverrideType", "Тип списка"),
        ("listOverrideName", "Название списка"),
        ("listOverrideCode", "Код списка"),
        ("listOverrideValue", "Значение списка"),
        ("listOverrideAction", "Диагностика списка"),
        ("listOverrideChangedFinalPrice", "Список изменил финальную цену"),
        ("pricingRule", "Правило ЦО"),
    ]
    dynamic_headers = [(f"competitor:{column.get('key')}", column.get("title") or column.get("key")) for column in competitor_columns]
    headers = headers[:10] + dynamic_headers + headers[10:]

    def _export_cell(row: dict, key: str) -> object:
        if key.startswith("competitor:"):
            column_key = key.replace("competitor:", "", 1)
            cell = (row.get("competitorPrices") or {}).get(column_key) or {}
            price = cell.get("price")
            return "" if _is_missing_price_value(price) else price
        if key.startswith("listOverride"):
            log = row.get("listOverrideLog") if isinstance(row.get("listOverrideLog"), dict) else {}
            if key == "listOverrideLogText":
                return _list_override_log_text(log)
            if key == "listOverrideType":
                return log.get("listType") or ""
            if key == "listOverrideName":
                return log.get("listName") or ""
            if key == "listOverrideCode":
                return log.get("listCode") or ""
            if key == "listOverrideValue":
                return log.get("displayValue") or log.get("value") or ""
            if key == "listOverrideAction":
                return log.get("listEffectMessage") or log.get("action") or ""
            if key == "listOverrideChangedFinalPrice":
                changed = log.get("listChangedFinalPrice")
                if changed is True:
                    return "да"
                if changed is False:
                    return "нет"
                return ""
        value = row.get(key, "")
        mojibake_dash_values = {"\u0432\u0402\u201d", "\u0420\u0406\u0420\u201a\u0432\u0402\u045c"}
        competitor_price_keys = {"bestCompetitorPrice", "lowestCompetitorPrice", "chosenCompetitorPrice"}
        if (
            value in (None, "")
            or (isinstance(value, str) and value in mojibake_dash_values)
            or (key in competitor_price_keys and _is_missing_price_value(value))
        ):
            return ""
        return value

    filename = f"{pl.number}.{fmt}"
    if fmt == "csv":
        s = io.StringIO()
        writer = csv.writer(s, lineterminator="\n")
        writer.writerow([label for _, label in headers])
        for row in rows_payload:
            writer.writerow([_export_cell(row, key) for key, _ in headers])
        return StreamingResponse(
            iter([s.getvalue().encode("utf-8-sig")]),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
        )
    wb = Workbook()
    ws = wb.active
    ws.title = "Прайс"
    ws.append([label for _, label in headers])
    for row in rows_payload:
        ws.append([_export_cell(row, key) for key, _ in headers])
    long_text_headers = {"Лог расчета цены", "Лог применения списка", "Диагностика списка"}
    for column_cells in ws.columns:
        header = str(column_cells[0].value or "")
        max_len = max(
            len(line)
            for cell in column_cells
            for line in str(cell.value if cell.value is not None else "").splitlines()
        )
        width = min(max(max_len + 2, 10), 80 if header in long_text_headers else 32)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = width
        if header in long_text_headers:
            for cell in column_cells:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    bio = io.BytesIO()
    wb.save(bio)
    return StreamingResponse(
        iter([bio.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
    )


@app.get("/api/generated-price-lists/{price_list_id}/compare/{other_id}")
def compare_generated_price_lists(
    price_list_id: str,
    other_id: str,
    db: Session = Depends(get_db),
    current_user: AppUser = Depends(get_current_user),
):
    pl, pf = _price_list_by_identifier(db, price_list_id)
    other, other_pf = _price_list_by_identifier(db, other_id)
    _ensure_price_format_access(pf, current_user)
    _ensure_price_format_access(other_pf, current_user)
    current = {
        cp.product_id: cp
        for cp in db.execute(select(CalculatedPrice).where(CalculatedPrice.price_list_id == pl.id)).scalars().all()
    }
    previous = {
        cp.product_id: cp
        for cp in db.execute(select(CalculatedPrice).where(CalculatedPrice.price_list_id == other.id)).scalars().all()
    }
    product_ids = list(current.keys() & previous.keys())[:500]
    products = {row.id: row for row in db.execute(select(Product).where(Product.id.in_(product_ids))).scalars().all()} if product_ids else {}
    rows = []
    for product_id in product_ids:
        cur = current[product_id]
        old = previous[product_id]
        old_price = float(old.final_price or 0)
        new_price = float(cur.final_price or 0)
        rows.append(
            {
                "sku": products.get(product_id).code if product_id in products else "",
                "name": products.get(product_id).name if product_id in products else "",
                "oldPrice": old_price,
                "newPrice": new_price,
                "changePercent": round((new_price - old_price) / old_price * 100, 2) if old_price else None,
                "oldZone": _calculated_zone(old),
                "newZone": _calculated_zone(cur),
            }
        )
    return {"base": pl.number, "other": other.number, "items": rows}


LIST_TYPE_LABELS = {
    "fixed_price": "Фиксированная цена",
    "min_price": "Минимальная цена",
    "max_price": "Максимальная цена",
    "min_markup": "Минимальная наценка",
    "critical_markup": "Критическая наценка",
    "max_markup": "Максимальная наценка",
    "no_bend": "Без прогиба",
    "percentile_override": "Percentile override",
    "exclude_from_pricing": "Исключить из переоценки",
}

LIST_TYPE_LABELS.setdefault("fixed_markup", "Фиксированная наценка")
LIST_TYPE_CODES = {v: k for k, v in LIST_TYPE_LABELS.items()}
PERCENT_LIST_TYPES = {"fixed_markup", "critical_markup", "min_markup", "max_markup", "percentile_override", "markup"}
PRICE_LIST_TYPES = {"fixed_price", "min_price", "max_price"}
BOOLEAN_LIST_TYPES = {"exclude_from_pricing", "no_bend", "exclusion"}


def _list_type_code(value: str | None) -> str:
    text = (value or "").strip()
    return normalize_list_type(LIST_TYPE_CODES.get(text, text or "fixed_price")) or "fixed_price"


def _list_type_label(value: str | None) -> str:
    code = _list_type_code(value)
    return LIST_TYPE_LABELS.get(code, value or "")


def _format_list_item_value(list_type: str | None, value: object, special_value: str = "") -> str:
    if _list_type_code(list_type) == "critical_markup" and special_value == "-":
        return "-"
    if value is None:
        return ""
    try:
        numeric = float(value)
    except Exception:
        return ""
    text = f"{numeric:g}"
    type_code = _list_type_code(list_type)
    if type_code in PERCENT_LIST_TYPES:
        return f"{text}%"
    if type_code in PRICE_LIST_TYPES:
        return text
    if type_code in BOOLEAN_LIST_TYPES:
        return "Да" if numeric != 0 else "Нет"
    return text


def _parse_list_date(value: object) -> date | None:
    text = str(value or "").strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d.%m.%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    raise HTTPException(status_code=400, detail="date must be YYYY-MM-DD or DD.MM.YYYY")


def _list_status_label(value: str | None) -> str:
    text = (value or "").strip().lower()
    if text in {"active", "активен", "активный"} or text.startswith("актив"):
        return "Активный"
    if text in {"archived", "архив", "архивный"} or text.startswith("архив"):
        return "Архивный"
    return "Неактивный"


def _effective_list_status(ul: UniversalList, *, today: date | None = None) -> dict:
    today = today or date.today()
    raw_label = _list_status_label(ul.status)
    if raw_label != "Активный":
        return {
            "code": "inactive",
            "label": "Неактивный",
            "active": False,
            "reason": "Список отключён вручную.",
        }
    if ul.start_date is not None and ul.start_date > today:
        return {
            "code": "not_started",
            "label": "Не начался",
            "active": False,
            "reason": f"Список начнёт применяться {ul.start_date.isoformat()}.",
        }
    if ul.end_date is not None and ul.end_date < today:
        return {
            "code": "expired",
            "label": "Истёк",
            "active": False,
            "reason": f"Срок действия списка закончился {ul.end_date.isoformat()}.",
        }
    return {
        "code": "active",
        "label": "Активный",
        "active": True,
        "reason": "Список сейчас применим по статусу и периоду действия.",
    }


def _list_bindings(db: Session, list_ids: list[int]) -> dict[int, list[PriceFormat]]:
    if not list_ids:
        return {}
    rows = db.execute(
        select(UniversalListPriceFormat.universal_list_id, PriceFormat)
        .join(PriceFormat, PriceFormat.id == UniversalListPriceFormat.price_format_id)
        .where(UniversalListPriceFormat.universal_list_id.in_(list_ids))
        .order_by(PriceFormat.code.asc())
    ).all()
    out: dict[int, list[PriceFormat]] = {}
    for list_id, pf in rows:
        out.setdefault(int(list_id), []).append(pf)
    return out


def _sync_universal_list_bindings(db: Session, ul: UniversalList, format_codes: list[str] | None) -> None:
    if format_codes is None:
        return
    codes = [str(code).strip() for code in format_codes if str(code).strip()]
    db.execute(delete(UniversalListPriceFormat).where(UniversalListPriceFormat.universal_list_id == ul.id))
    ul.price_format_id = None
    if not codes:
        return
    formats = db.execute(select(PriceFormat).where(PriceFormat.code.in_(codes))).scalars().all()
    for pf in formats:
        db.add(UniversalListPriceFormat(universal_list_id=ul.id, price_format_id=pf.id))
    if len(formats) == 1:
        ul.price_format_id = formats[0].id


def _universal_list_row(db: Session, ul: UniversalList, item_count: int, bindings: list[PriceFormat] | None) -> dict:
    bindings = bindings or []
    is_global = ul.price_format_id is None and not bindings
    raw_status = _list_status_label(ul.status)
    today = date.today()
    effective_status = _effective_list_status(ul, today=today)
    return {
        "id": ul.id,
        "name": ul.name,
        "code": ul.code or f"UL_{ul.id:06d}",
        "type": _list_type_code(ul.type),
        "typeLabel": _list_type_label(ul.type),
        "active": bool(effective_status["active"]),
        "status": effective_status["label"],
        "rawStatus": raw_status,
        "effectiveStatus": effective_status["code"],
        "effectiveStatusLabel": effective_status["label"],
        "effectiveStatusReason": effective_status["reason"],
        "dateValidity": {
            "today": today.isoformat(),
            "startsInFuture": bool(ul.start_date is not None and ul.start_date > today),
            "expired": bool(ul.end_date is not None and ul.end_date < today),
        },
        "itemsCount": int(item_count or 0),
        "priceFormats": [{"code": pf.code, "name": pf.name, "branch": pf.branch} for pf in bindings],
        "scope": "global" if is_global else "formats",
        "startDate": _fmt_d(ul.start_date),
        "endDate": _fmt_d(ul.end_date),
        "updatedAt": _fmt_dt(ul.created_at),
        "comment": "",
    }


async def _read_upload_limited(file: UploadFile, *, max_bytes: int) -> bytes:
    chunks: list[bytes] = []
    total = 0
    while True:
        chunk = await file.read(1024 * 1024)
        if not chunk:
            break
        total += len(chunk)
        if total > max_bytes:
            limit_mb = max_bytes // (1024 * 1024)
            raise ValueError(f"file exceeds LIST_IMPORT_MAX_UPLOAD_SIZE_MB={limit_mb}")
        chunks.append(chunk)
    return b"".join(chunks)


@app.post("/lists/import")
async def import_business_list(
    list_type: str = Form(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    try:
        content = await _read_upload_limited(file, max_bytes=max_upload_size_bytes())
        return import_business_list_excel(
            db=db,
            content=content,
            filename=file.filename or "upload.xlsx",
            list_type=list_type,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except SQLAlchemyError as exc:
        db.rollback()
        raise HTTPException(status_code=500, detail="database error during list import") from exc


@app.get("/lists")
def get_business_lists(
    list_type: str | None = Query(None),
    limit: int = Query(100, ge=1, le=500),
    offset: int = Query(0, ge=0),
    db: Session = Depends(get_db),
):
    stmt = select(BusinessList).order_by(BusinessList.created_at.desc(), BusinessList.id.desc())
    if list_type:
        stmt = stmt.where(BusinessList.list_type == list_type.strip().casefold())
    rows = db.execute(stmt.offset(offset).limit(limit)).scalars().all()
    return [business_list_to_dict(row) for row in rows]


@app.get("/lists/{list_id}")
def get_business_list(list_id: int, limit: int = Query(100, ge=1, le=500), db: Session = Depends(get_db)):
    row = db.execute(select(BusinessList).where(BusinessList.id == list_id)).scalars().first()
    if not row:
        raise HTTPException(status_code=404, detail="list not found")
    items = db.execute(
        select(BusinessListItem)
        .where(BusinessListItem.business_list_id == row.id)
        .order_by(BusinessListItem.source_row.asc(), BusinessListItem.id.asc())
        .limit(limit)
    ).scalars().all()
    return business_list_to_dict(
        row,
        include_errors=True,
        item_preview=[business_list_item_to_dict(item) for item in items],
    )


@app.delete("/lists/{list_id}")
def delete_business_list(list_id: int, db: Session = Depends(get_db)):
    row = db.execute(select(BusinessList).where(BusinessList.id == list_id)).scalars().first()
    if not row:
        raise HTTPException(status_code=404, detail="list not found")
    db.delete(row)
    db.commit()
    return {"status": "deleted", "id": list_id}


@app.get("/api/lists-management")
def get_lists_management(
    search: str | None = Query(None),
    type: str | None = Query(None),
    status: str | None = Query(None),
    db: Session = Depends(get_db),
):
    stmt = select(UniversalList).order_by(UniversalList.created_at.desc(), UniversalList.id.desc())
    if search:
        like = f"%{search.strip()}%"
        stmt = stmt.where((UniversalList.name.ilike(like)) | (UniversalList.code.ilike(like)))
    if type and type != "__all__":
        type_label = LIST_TYPE_LABELS.get(type, type)
        stmt = stmt.where((UniversalList.type == type) | (UniversalList.type == type_label))
    rows = db.execute(stmt.limit(300)).scalars().all()
    if status and status != "__all__":
        needle = str(status).strip().lower()
        filtered = []
        for row in rows:
            effective_status = _effective_list_status(row)
            if effective_status["code"].lower().startswith(needle) or effective_status["label"].lower().startswith(needle):
                filtered.append(row)
        rows = filtered
    counts = dict(
        db.execute(
            select(ListItem.universal_list_id, func.count(ListItem.id))
            .where(ListItem.universal_list_id.in_([row.id for row in rows]) if rows else False)
            .group_by(ListItem.universal_list_id)
        ).all()
    ) if rows else {}
    bindings = _list_bindings(db, [row.id for row in rows])
    return [_universal_list_row(db, row, counts.get(row.id, 0), bindings.get(row.id)) for row in rows]


@app.get("/api/lists-management/{list_id}")
def get_lists_management_card(list_id: int, db: Session = Depends(get_db)):
    ul = db.execute(select(UniversalList).where(UniversalList.id == list_id)).scalars().first()
    if not ul:
        raise HTTPException(status_code=404, detail="list not found")
    bindings = _list_bindings(db, [ul.id]).get(ul.id, [])
    items = db.execute(
        select(ListItem, Product, ProductExtra)
        .join(Product, Product.id == ListItem.product_id)
        .outerjoin(ProductExtra, ProductExtra.product_id == Product.id)
        .where(ListItem.universal_list_id == ul.id)
        .order_by(Product.name.asc(), Product.code.asc())
    ).all()
    payload = _universal_list_row(db, ul, len(items), bindings)
    payload["items"] = [
        {
            "sku": p.code,
            "name": p.name,
            "manufacturer": extra.manufacturer if extra else "",
            "value": "-" if item.special_value == "-" else (float(item.value) if item.value is not None else None),
            "valueDisplay": _format_list_item_value(ul.type, item.value, item.special_value),
            "comment": "",
        }
        for item, p, extra in items
    ]
    return payload


@app.post("/api/lists-management")
def create_lists_management(payload: dict = Body(...), db: Session = Depends(get_db)):
    ul = UniversalList(
        code=(payload.get("code") or "").strip() or None,
        name=(payload.get("name") or "Новый список").strip(),
        type=LIST_TYPE_LABELS.get(payload.get("type"), payload.get("type") or "Фиксированная цена"),
        status="Активный" if payload.get("active", True) else "Неактивный",
        start_date=_parse_list_date(payload.get("startDate")),
        end_date=_parse_list_date(payload.get("endDate")),
    )
    ul.type = _list_type_code(ul.type)
    db.add(ul)
    db.flush()
    if not ul.code:
        ul.code = f"UL_{ul.id:06d}"
    _sync_universal_list_bindings(db, ul, payload.get("formatCodes"))
    db.commit()
    return {"id": ul.id}


@app.patch("/api/lists-management/{list_id}")
def update_lists_management(list_id: int, payload: dict = Body(...), db: Session = Depends(get_db)):
    ul = db.execute(select(UniversalList).where(UniversalList.id == list_id)).scalars().first()
    if not ul:
        raise HTTPException(status_code=404, detail="list not found")
    if "name" in payload:
        ul.name = str(payload.get("name") or ul.name).strip()
    if "code" in payload:
        ul.code = str(payload.get("code") or ul.code or "").strip() or None
    if "type" in payload:
        ul.type = LIST_TYPE_LABELS.get(payload.get("type"), payload.get("type") or ul.type)
        ul.type = _list_type_code(ul.type)
    if "active" in payload:
        ul.status = "Активный" if payload.get("active") else "Неактивный"
    if "status" in payload:
        ul.status = str(payload.get("status") or ul.status)
    if "startDate" in payload:
        ul.start_date = _parse_list_date(payload.get("startDate"))
    if "endDate" in payload:
        ul.end_date = _parse_list_date(payload.get("endDate"))
    _sync_universal_list_bindings(db, ul, payload.get("formatCodes") if "formatCodes" in payload else None)
    db.commit()
    return {"status": "ok"}


@app.post("/api/lists-management/{list_id}/copy")
def copy_lists_management(list_id: int, db: Session = Depends(get_db)):
    ul = db.execute(select(UniversalList).where(UniversalList.id == list_id)).scalars().first()
    if not ul:
        raise HTTPException(status_code=404, detail="list not found")
    copied = UniversalList(
        code=None,
        name=f"{ul.name} — копия",
        type=ul.type,
        status="Неактивный",
        start_date=ul.start_date,
        end_date=ul.end_date,
        price_format_id=ul.price_format_id,
    )
    db.add(copied)
    db.flush()
    copied.code = f"UL_{copied.id:06d}"
    for link in db.execute(select(UniversalListPriceFormat).where(UniversalListPriceFormat.universal_list_id == ul.id)).scalars().all():
        db.add(UniversalListPriceFormat(universal_list_id=copied.id, price_format_id=link.price_format_id))
    for item in db.execute(select(ListItem).where(ListItem.universal_list_id == ul.id)).scalars().all():
        db.add(
            ListItem(
                universal_list_id=copied.id,
                product_id=item.product_id,
                value=item.value,
                special_value=item.special_value,
            )
        )
    db.commit()
    return {"id": copied.id}


@app.post("/api/lists-management/{list_id}/items")
def upsert_lists_management_item(list_id: int, payload: dict = Body(...), db: Session = Depends(get_db)):
    ul = db.execute(select(UniversalList).where(UniversalList.id == list_id)).scalars().first()
    if not ul:
        raise HTTPException(status_code=404, detail="list not found")
    sku = str(payload.get("sku") or "").strip()
    product = find_product_by_identifier(db, sku)
    if not product:
        raise HTTPException(status_code=404, detail="product not found")
    raw_value = payload.get("value")
    if _list_type_code(ul.type) == "critical_markup" and str(raw_value or "").strip() == "-":
        value, special_value, value_error = Decimal("0"), "-", None
    elif is_exclude_from_pricing_type(str(ul.type or "")):
        value, value_error = normalize_universal_list_value(str(ul.type or ""), raw_value)
        special_value = ""
    else:
        value, special_value, value_error = parse_list_decimal(raw_value), "", None
    if value is None:
        raise HTTPException(status_code=400, detail=value_error or "invalid numeric value")
    item = db.execute(
        select(ListItem).where(ListItem.universal_list_id == ul.id).where(ListItem.product_id == product.id)
    ).scalars().first()
    if item:
        item.value = value
        item.special_value = special_value
    else:
        db.add(
            ListItem(
                universal_list_id=ul.id,
                product_id=product.id,
                value=value,
                special_value=special_value,
            )
        )
    db.commit()
    return {"status": "ok"}


@app.post("/api/lists-management/{list_id}/import")
async def import_lists_management_items(list_id: int, file: UploadFile = File(...), db: Session = Depends(get_db)):
    ul = db.execute(select(UniversalList).where(UniversalList.id == list_id)).scalars().first()
    if not ul:
        raise HTTPException(status_code=404, detail="list not found")
    content = await file.read()
    wb = load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    headers = [str(cell.value or "").strip().lower() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    sku_idx = headers.index("sku") if "sku" in headers else 0
    exclusion_by_presence = is_exclude_from_pricing_type(str(ul.type or ""))
    value_idx = headers.index("value") if "value" in headers else (None if exclusion_by_presence else 1)
    imported = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        sku = str(row[sku_idx] or "").strip()
        if not sku:
            continue
        product = find_product_by_identifier(db, sku)
        if not product:
            continue
        raw_value = row[value_idx] if value_idx is not None and value_idx < len(row) else None
        if _list_type_code(ul.type) == "critical_markup" and str(raw_value or "").strip() == "-":
            value, special_value = Decimal("0"), "-"
        elif exclusion_by_presence:
            value, _value_error = normalize_universal_list_value(str(ul.type or ""), raw_value)
            special_value = ""
        else:
            value, special_value = parse_list_decimal(raw_value), ""
        if value is None:
            continue
        item = db.execute(
            select(ListItem).where(ListItem.universal_list_id == ul.id).where(ListItem.product_id == product.id)
        ).scalars().first()
        if item:
            item.value = value
            item.special_value = special_value
        else:
            db.add(
                ListItem(
                    universal_list_id=ul.id,
                    product_id=product.id,
                    value=value,
                    special_value=special_value,
                )
            )
        imported += 1
    db.commit()
    return {"status": "ok", "imported": imported}


@app.post("/api/lists-management/{list_id}/import-excel")
async def import_lists_management_excel(list_id: int, file: UploadFile = File(...), db: Session = Depends(get_db)):
    ul = db.execute(select(UniversalList).where(UniversalList.id == list_id)).scalars().first()
    if not ul:
        raise HTTPException(status_code=404, detail="list not found")
    try:
        content = await _read_upload_limited(file, max_bytes=max_upload_size_bytes())
        return import_universal_list_excel(
            db=db,
            universal_list=ul,
            content=content,
            filename=file.filename or "upload.xlsx",
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except SQLAlchemyError as exc:
        db.rollback()
        raise HTTPException(status_code=500, detail="database error during list import") from exc


@app.get("/api/lists-management/{list_id}/export.{fmt}")
def export_lists_management(list_id: int, fmt: str, db: Session = Depends(get_db)):
    if fmt not in {"csv", "xlsx"}:
        raise HTTPException(status_code=400, detail="fmt must be csv or xlsx")
    payload = get_lists_management_card(list_id=list_id, db=db)
    headers = ["SKU", "Наименование", "Производитель", "Значение правила", "Комментарий"]
    filename = f"{payload['code']}.{fmt}"
    if fmt == "csv":
        s = io.StringIO()
        writer = csv.writer(s, lineterminator="\n")
        writer.writerow(headers)
        for row in payload["items"]:
            writer.writerow([row["sku"], row["name"], row["manufacturer"], row["value"], row["comment"]])
        return StreamingResponse(iter([s.getvalue().encode("utf-8-sig")]), media_type="text/csv; charset=utf-8", headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"})
    wb = Workbook()
    ws = wb.active
    ws.title = "list"
    ws.append(headers)
    for row in payload["items"]:
        ws.append([row["sku"], row["name"], row["manufacturer"], row["value"], row["comment"]])
    bio = io.BytesIO()
    wb.save(bio)
    return StreamingResponse(iter([bio.getvalue()]), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"})


def _contractor_row(db: Session, link: CounterpartyPriceFormat, holding: Holding | None, cp: Counterparty | None, point: DeliveryPoint | None, pf: PriceFormat) -> dict:
    selected_sources = len(get_assigned_competitor_price_lists(db=db, price_format_id=int(pf.id)))
    latest_pl = db.execute(select(PriceList).where(PriceList.price_format_id == pf.id).order_by(PriceList.created_at.desc()).limit(1)).scalars().first()
    ref_rows = db.execute(select(ReferenceUpdateStatus).where(ReferenceUpdateStatus.branch_name == pf.branch).limit(20)).scalars().all()
    return {
        "id": link.id,
        "holding": holding.name if holding else "",
        "counterparty": cp.name if cp else "",
        "pharmacy": point.name if point else "",
        "region": pf.branch,
        "branch": pf.branch,
        "formatCode": pf.code,
        "formatName": pf.name,
        "status": link.status or "active",
        "updatedAt": _fmt_dt(link.updated_at),
        "selectedSources": selected_sources,
        "latestPriceList": latest_pl.number if latest_pl else "",
        "latestPriceListDate": _fmt_dt(latest_pl.created_at) if latest_pl else "",
        "references": [{"type": row.data_type, "status": row.status, "lastUpdatedAt": _fmt_dt(row.last_updated_at), "rows": row.rows_count} for row in ref_rows],
    }


@app.get("/api/contractors")
def get_contractors(
    search: str | None = Query(None),
    branch: str | None = Query(None),
    format_code: str | None = Query(None),
    db: Session = Depends(get_db),
):
    stmt = (
        select(CounterpartyPriceFormat, Holding, Counterparty, DeliveryPoint, PriceFormat)
        .join(PriceFormat, PriceFormat.id == CounterpartyPriceFormat.price_format_id)
        .outerjoin(Holding, Holding.id == CounterpartyPriceFormat.holding_id)
        .outerjoin(Counterparty, Counterparty.id == CounterpartyPriceFormat.counterparty_id)
        .outerjoin(DeliveryPoint, DeliveryPoint.id == CounterpartyPriceFormat.delivery_point_id)
        .order_by(PriceFormat.branch.asc(), Counterparty.name.asc(), DeliveryPoint.name.asc())
    )
    if branch and branch != "__all__":
        stmt = stmt.where(PriceFormat.branch == branch)
    if format_code and format_code != "__all__":
        stmt = stmt.where(PriceFormat.code == format_code)
    rows = db.execute(stmt.limit(300)).all()
    payload = [_contractor_row(db, link, holding, cp, point, pf) for link, holding, cp, point, pf in rows]
    if not payload:
        formats = db.execute(select(PriceFormat).order_by(PriceFormat.branch.asc(), PriceFormat.code.asc()).limit(100)).scalars().all()
        payload = [
            {
                "id": pf.id,
                "holding": "Без холдинга",
                "counterparty": "Не назначен",
                "pharmacy": "",
                "region": pf.branch,
                "branch": pf.branch,
                "formatCode": pf.code,
                "formatName": pf.name,
                "status": "draft",
                "updatedAt": "",
                "selectedSources": 0,
                "latestPriceList": "",
                "latestPriceListDate": "",
                "references": [],
            }
            for pf in formats
        ]
    if search:
        needle = search.strip().lower()
        payload = [row for row in payload if any(needle in str(value).lower() for value in row.values() if not isinstance(value, list))]
    return payload


@app.get("/api/contractors/{row_id}")
def get_contractor_card(row_id: int, db: Session = Depends(get_db)):
    link = db.execute(select(CounterpartyPriceFormat).where(CounterpartyPriceFormat.id == row_id)).scalars().first()
    if not link:
        pf = db.execute(select(PriceFormat).where(PriceFormat.id == row_id)).scalars().first()
        if not pf:
            raise HTTPException(status_code=404, detail="contractor row not found")
        return {
            "id": row_id,
            "holding": "Без холдинга",
            "counterparty": "Не назначен",
            "pharmacy": "",
            "branch": pf.branch,
            "formatCode": pf.code,
            "formatName": pf.name,
            "status": "draft",
            "priceFormats": [{"code": pf.code, "name": pf.name, "branch": pf.branch}],
            "assignments": [],
            "recentPriceLists": get_generated_price_lists(branch=pf.branch, format_code=pf.code, date_from=None, date_to=None, search=None, db=db)[:5],
            "references": [],
        }
    holding = db.get(Holding, link.holding_id) if link.holding_id else None
    cp = db.get(Counterparty, link.counterparty_id) if link.counterparty_id else None
    point = db.get(DeliveryPoint, link.delivery_point_id) if link.delivery_point_id else None
    pf = db.get(PriceFormat, link.price_format_id)
    if not pf:
        raise HTTPException(status_code=404, detail="price format not found")
    base = _contractor_row(db, link, holding, cp, point, pf)
    base["priceFormats"] = [{"code": pf.code, "name": pf.name, "branch": pf.branch}]
    base["assignments"] = [
        {
            "source": item.price_list.display_name or item.price_list.source_key,
            "competitor": item.price_list.competitor_name or item.price_list.supplier,
            "region": item.price_list.region or item.price_list.branch_name,
            "login": item.price_list.account_login,
            "coefficient": float(item.assignment.coefficient or 1),
            "active": bool(item.assignment.is_active),
        }
        for item in get_assigned_competitor_price_lists(db=db, price_format_id=int(pf.id), active_only=False)
    ]
    base["recentPriceLists"] = get_generated_price_lists(branch=pf.branch, format_code=pf.code, date_from=None, date_to=None, search=None, db=db)[:5]
    return base


@app.post("/api/contractors/import")
async def import_contractors(file: UploadFile = File(...), db: Session = Depends(get_db)):
    content = await file.read()
    wb = load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    headers = [str(cell.value or "").strip().lower() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    def idx(name: str, fallback: int) -> int:
        return headers.index(name) if name in headers else fallback
    holding_i, cp_i, pharmacy_i, branch_i, format_i = idx("holding", 0), idx("counterparty", 1), idx("pharmacy", 2), idx("branch", 3), idx("format_code", 4)
    imported = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        format_code = str(row[format_i] or "").strip()
        if not format_code:
            continue
        pf = db.execute(select(PriceFormat).where(PriceFormat.code == format_code)).scalars().first()
        if not pf:
            pf = PriceFormat(code=format_code, name=format_code, branch=str(row[branch_i] or "").strip())
            db.add(pf)
            db.flush()
        holding_name = str(row[holding_i] or "").strip() or "Без холдинга"
        cp_name = str(row[cp_i] or "").strip() or "Не указан"
        point_name = str(row[pharmacy_i] or "").strip()
        holding = db.execute(select(Holding).where(Holding.name == holding_name)).scalars().first()
        if not holding:
            holding = Holding(name=holding_name, branch_id=pf.branch)
            db.add(holding)
            db.flush()
        cp = db.execute(select(Counterparty).where(Counterparty.name == cp_name).where(Counterparty.branch_id == pf.branch)).scalars().first()
        if not cp:
            cp = Counterparty(name=cp_name, holding_id=str(holding.id), branch_id=pf.branch)
            db.add(cp)
            db.flush()
        point = None
        if point_name:
            point = db.execute(select(DeliveryPoint).where(DeliveryPoint.name == point_name).where(DeliveryPoint.counterparty_id == str(cp.id))).scalars().first()
            if not point:
                point = DeliveryPoint(name=point_name, counterparty_id=str(cp.id), branch_id=pf.branch)
                db.add(point)
                db.flush()
        link = db.execute(
            select(CounterpartyPriceFormat)
            .where(CounterpartyPriceFormat.counterparty_id == cp.id)
            .where(CounterpartyPriceFormat.delivery_point_id == (point.id if point else None))
            .where(CounterpartyPriceFormat.price_format_id == pf.id)
        ).scalars().first()
        if not link:
            db.add(CounterpartyPriceFormat(holding_id=holding.id, counterparty_id=cp.id, delivery_point_id=point.id if point else None, price_format_id=pf.id))
        imported += 1
    db.commit()
    return {"status": "ok", "imported": imported}


def _latest_analytics_price_list(db: Session, branch: str | None, format_code: str | None) -> tuple[PriceList, PriceFormat]:
    stmt = select(PriceList, PriceFormat).join(PriceFormat, PriceFormat.id == PriceList.price_format_id).order_by(PriceList.created_at.desc())
    if branch and branch != "__all__":
        stmt = stmt.where(PriceFormat.branch == branch)
    if format_code and format_code != "__all__":
        stmt = stmt.where(PriceFormat.code == format_code)
    row = db.execute(stmt.limit(1)).first()
    if not row:
        raise HTTPException(status_code=404, detail="price list not found")
    return row[0], row[1]


@app.get("/api/price-list-analytics")
def get_price_list_analytics_dashboard(
    price_list_id: str | None = Query(None),
    branch: str | None = Query(None),
    format_code: str | None = Query(None),
    db: Session = Depends(get_db),
    current_user: AppUser = Depends(get_current_user),
):
    pl, pf = _price_list_by_identifier(db, price_list_id) if price_list_id else _latest_analytics_price_list(db, branch, format_code)
    _ensure_price_format_access(pf, current_user)
    analytics = build_workflow_analytics(db=db, price_list_id=pl.id)
    summary = analytics.get("summary", {})
    rows = db.execute(select(CalculatedPrice, Product).join(Product, Product.id == CalculatedPrice.product_id).where(CalculatedPrice.price_list_id == pl.id)).all()
    total = max(1, len(rows))
    zones = [
        {"code": "left", "label": "ЛП", "count": int(summary.get("leftZone") or 0), "percent": round((int(summary.get("leftZone") or 0) / total) * 100, 2), "change": 0},
        {"code": "optimal", "label": "Зона логичности", "count": int(summary.get("optimalZone") or 0), "percent": round((int(summary.get("optimalZone") or 0) / total) * 100, 2), "change": 0},
        {"code": "right", "label": "ПП", "count": int(summary.get("rightZone") or 0), "percent": round((int(summary.get("rightZone") or 0) / total) * 100, 2), "change": 0},
        {"code": "no-data", "label": "Зона без цен", "count": int(summary.get("withoutCompetitors") or 0), "percent": round((int(summary.get("withoutCompetitors") or 0) / total) * 100, 2), "change": 0},
    ]
    markups: list[float] = []
    bends: list[float] = []
    changed: list[dict] = []
    increased_count = 0
    decreased_count = 0
    unchanged_count = 0
    universal_list_count = 0
    for cp, product in rows:
        cost = float(cp.cost or 0)
        final = float(cp.final_price or 0)
        base = float(cp.base_price or 0)
        margin = margin_percent_from_price(cost, final)
        if margin is not None:
            markups.append(round(float(margin), 2))
        if cp.bend_percent_used is not None and cp.price_from_competitor is not None:
            bends.append(round(float(cp.bend_percent_used), 2))
        elif cp.price_from_competitor is not None and cp.chosen_competitor_price:
            bends.append(round((float(cp.chosen_competitor_price) - float(cp.price_from_competitor)) / float(cp.chosen_competitor_price) * 100, 2))
        if base:
            change_percent = round((final - base) / base * 100, 2)
            changed.append({"sku": product.code, "name": product.name, "oldPrice": base, "newPrice": final, "changePercent": change_percent, "zone": _calculated_zone(cp)})
            if change_percent > 0.01:
                increased_count += 1
            elif change_percent < -0.01:
                decreased_count += 1
            else:
                unchanged_count += 1
        raw_list_ids = str(cp.applied_list_ids or "").strip()
        if raw_list_ids and raw_list_ids not in {"[]", "null", "None"}:
            universal_list_count += 1
    buckets = [
        {"label": "<0", "count": len([x for x in markups if x < 0])},
        {"label": "0-5", "count": len([x for x in markups if 0 <= x < 5])},
        {"label": "5-10", "count": len([x for x in markups if 5 <= x < 10])},
        {"label": "10-20", "count": len([x for x in markups if 10 <= x < 20])},
        {"label": "20+", "count": len([x for x in markups if x >= 20])},
    ]
    percentile_count = int(summary.get("percentileUsage") or 0)
    changed_sorted = sorted(changed, key=lambda row: abs(row["changePercent"] or 0), reverse=True)[:25]
    repricing = {
        "changedCount": int(summary.get("changedPriceCount") or len([row for row in changed if abs(row["changePercent"] or 0) > 0.01])),
        "increasedCount": increased_count,
        "decreasedCount": decreased_count,
        "unchangedCount": unchanged_count,
        "averageChangePercent": round(sum([row["changePercent"] for row in changed]) / len(changed), 2) if changed else 0,
        "averageBendPercent": round(sum(bends) / len(bends), 2) if bends else 0,
        "maxIncrease": max([row["changePercent"] for row in changed], default=0),
        "maxDecrease": min([row["changePercent"] for row in changed], default=0),
        "withoutCompetitors": int(summary.get("withoutCompetitors") or 0),
        "noCompetitorRuleApplied": int(summary.get("noCompetitorRuleApplied") or 0),
        "withSubstitute": int(analytics.get("productsWithSubstituteMatches") or 0),
        "withPercentile": percentile_count,
        "withUniversalLists": universal_list_count,
        "belowMdcPrevented": len([cp for cp, _ in rows if float(cp.final_price or 0) <= float(cp.base_price or 0)]),
    }
    return {
        "priceList": _generated_price_list_summary(db, pl, pf),
        "summary": summary,
        "rightZoneReasons": analytics.get("rightZoneReasons") or {},
        "snapshot": analytics.get("snapshot") or {},
        "runSources": analytics.get("runSources") or {},
        "runRule": analytics.get("runRule") or {},
        "runLists": analytics.get("runLists") or [],
        "runReferenceVersions": analytics.get("runReferenceVersions") or {},
        "runPercentileConfig": analytics.get("runPercentileConfig") or {},
        "zones": zones,
        "charts": {
            "zoneDistribution": zones,
            "markupHistogram": buckets,
            "competitorUsage": [
                {"source": item.get("label") or _source_label(item.get("source") or ""), "count": int(item.get("skuCount") or 0)}
                for item in (analytics.get("competitorUsage") or [])
            ],
            "percentileUsage": [
                {"label": item.get("label") or _source_label(item.get("source") or ""), "count": int(item.get("skuCount") or 0)}
                for item in (analytics.get("percentileUsage") or [])
            ] or [{"label": "percentile", "count": percentile_count}, {"label": "regular", "count": max(0, len(rows) - percentile_count)}],
            "noCompetitors": [{"label": "Без конкурентной цены", "count": int(summary.get("withoutCompetitors") or 0)}, {"label": "С конкурентной ценой", "count": int(summary.get("withCompetitors") or 0)}],
            "topChangedProducts": analytics.get("topChangedProducts") or changed_sorted,
        },
        "repricing": repricing,
    }


@app.get("/api/price-list-analytics/export.{fmt}")
def export_price_list_analytics(
    fmt: str,
    price_list_id: str | None = Query(None),
    branch: str | None = Query(None),
    format_code: str | None = Query(None),
    db: Session = Depends(get_db),
    current_user: AppUser = Depends(get_current_user),
):
    if fmt not in {"csv", "xlsx"}:
        raise HTTPException(status_code=400, detail="fmt must be csv or xlsx")
    payload = get_price_list_analytics_dashboard(
        price_list_id=price_list_id,
        branch=branch,
        format_code=format_code,
        db=db,
        current_user=current_user,
    )
    rows = [
        ["Прайс-лист", payload["priceList"]["number"]],
        ["Снимок расчета", "да" if payload.get("snapshot") else "устаревший формат / без снимка"],
        ["Версия схемы снимка", (payload.get("snapshot") or {}).get("schemaVersion", "")],
        ["Сформировал", payload["priceList"].get("generatedBy", "")],
        ["Всего SKU", payload["summary"].get("skuTotal", 0)],
        ["С конкурентной ценой", payload["summary"].get("withCompetitors", 0)],
        ["Без конкурентной цены", payload["summary"].get("withoutCompetitors", 0)],
        ["ЛП", payload["summary"].get("leftZone", 0)],
        ["Зона логичности", payload["summary"].get("optimalZone", 0)],
        ["ПП", payload["summary"].get("rightZone", 0)],
        ["Зона без цен", payload["summary"].get("withoutCompetitors", 0)],
        ["Применена логика без конкурентов", payload["summary"].get("noCompetitorRuleApplied", 0)],
        ["Средняя наценка", payload["summary"].get("averageMarkup", 0)],
        ["Средняя цена", payload["summary"].get("averageFinalPrice", 0)],
        ["Использовано персентилей", payload["summary"].get("percentileUsage", 0)],
        ["Использовано замен", payload["summary"].get("substituteUsage", 0)],
        ["Изменено цен", payload["summary"].get("changedPriceCount", 0)],
        ["Цены повышены", payload.get("repricing", {}).get("increasedCount", 0)],
        ["Цены снижены", payload.get("repricing", {}).get("decreasedCount", 0)],
        ["Без изменений", payload.get("repricing", {}).get("unchangedCount", 0)],
        ["Универсальные списки", payload.get("repricing", {}).get("withUniversalLists", 0)],
    ]
    filename = f"itogi_co_{payload['priceList']['number']}.{fmt}"
    if fmt == "csv":
        s = io.StringIO()
        writer = csv.writer(s, lineterminator="\n")
        writer.writerows(rows)
        return StreamingResponse(iter([s.getvalue().encode("utf-8-sig")]), media_type="text/csv; charset=utf-8", headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"})
    wb = Workbook()
    ws = wb.active
    ws.title = "Итоги ЦО"
    for row in rows:
        ws.append(row)
    bio = io.BytesIO()
    wb.save(bio)
    return StreamingResponse(iter([bio.getvalue()]), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"})


@app.post("/api/pricing-workflow/generate")
def post_pricing_workflow_generate(
    payload: dict = Body(...),
    db: Session = Depends(get_db),
    current_user: AppUser = Depends(require_write_access),
):
    try:
        payload = payload or {}
        pf = db.get(PriceFormat, int(payload.get("price_format_id") or payload.get("priceFormatId") or 0))
        if pf is not None:
            _ensure_price_format_access(pf, current_user)
        payload["user"] = current_user.username
        run = create_workflow_run(db=db, payload=payload)
        status_code = 200 if run.status == "success" else 400
        result = run_to_dict(db=db, run=run, include_items=True)
        if status_code != 200:
            raise HTTPException(status_code=status_code, detail=result.get("error") or "workflow failed")
        return result
    except HTTPException:
        raise
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        db.rollback()
        logger.exception("pricing workflow generate failed")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/pricing-workflow/results/{run_id}")
def get_pricing_workflow_result(run_id: int, db: Session = Depends(get_db), current_user: AppUser = Depends(get_current_user)):
    run = db.get(PricingWorkflowRun, run_id)
    if run is None:
        raise HTTPException(status_code=404, detail="workflow run not found")
    pf = db.get(PriceFormat, run.price_format_id)
    if pf is not None:
        _ensure_price_format_access(pf, current_user)
    return run_to_dict(db=db, run=run, include_items=True)


@app.get("/api/pricing-workflow/analytics/{run_id}")
def get_pricing_workflow_analytics(run_id: int, db: Session = Depends(get_db), current_user: AppUser = Depends(get_current_user)):
    run = db.get(PricingWorkflowRun, run_id)
    if run is None:
        raise HTTPException(status_code=404, detail="workflow run not found")
    pf = db.get(PriceFormat, run.price_format_id)
    if pf is not None:
        _ensure_price_format_access(pf, current_user)
    return analytics_for_run(db=db, run=run)


@app.get("/api/pricing-workflow/export/{run_id}")
def get_pricing_workflow_export(
    run_id: int,
    format: str = Query("csv"),
    include: str = Query("price"),
    db: Session = Depends(get_db),
    current_user: AppUser = Depends(get_current_user),
):
    try:
        run = db.get(PricingWorkflowRun, run_id)
        if run is not None:
            pf = db.get(PriceFormat, run.price_format_id)
            if pf is not None:
                _ensure_price_format_access(pf, current_user)
        filename, content, media_type = export_workflow_run(db=db, run_id=run_id, fmt=format, include=include)
    except ValueError as e:
        raise HTTPException(status_code=404 if "not found" in str(e) else 400, detail=str(e))
    safe_filename = quote(filename)
    return StreamingResponse(
        iter([content]),
        media_type=media_type,
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{safe_filename}"},
    )


@app.get("/api/price-lists")
def get_price_lists(
    format_code: str | None = None,
    status: str | None = None,
    branch: str | None = None,
    db: Session = Depends(get_db),
    current_user: AppUser = Depends(get_current_user),
):
    stmt = (
        select(PriceList, PriceFormat)
        .join(PriceFormat, PriceFormat.id == PriceList.price_format_id)
        .order_by(PriceList.created_at.desc())
    )

    if format_code:
        stmt = stmt.where(PriceFormat.code == format_code)
    if status and status != "Все":
        stmt = stmt.where(PriceList.status == status)
    if branch and branch != "Все":
        stmt = stmt.where(PriceFormat.branch == branch)

    rows = db.execute(stmt).all()
    rows = [(pl, pf) for pl, pf in rows if user_can_access_branch(current_user, _branch_id_for_name(pf.branch), pf.branch)]
    if not rows:
        return []

    return [
        {
            "date": _fmt_dt(pl.created_at),
            "number": pl.number,
            "format": pf.code,
            "activationDate": _fmt_d(pl.activation_date),
            "user": pl.user,
            "status": pl.status,
            "branch": pf.branch,
        }
        for (pl, pf) in rows
    ]


@app.get("/api/price-lists/latest")
def get_latest_price_list(format_code: str = Query(..., min_length=1), db: Session = Depends(get_db)):
    row = (
        db.execute(
            select(PriceList, PriceFormat)
            .join(PriceFormat, PriceFormat.id == PriceList.price_format_id)
            .where(PriceFormat.code == format_code)
            .order_by(PriceList.created_at.desc(), PriceList.id.desc())
            .limit(1)
        )
        .first()
    )
    if not row:
        return None
    pl, pf = row
    return {
        "date": _fmt_dt(pl.created_at),
        "number": pl.number,
        "format": pf.code,
        "activationDate": _fmt_d(pl.activation_date),
        "user": pl.user,
        "status": pl.status,
        "branch": pf.branch,
    }


@app.get("/api/price-lists/{price_list_id}/analysis")
def get_price_list_analysis(price_list_id: str, db: Session = Depends(get_db)):
    pl_pf = db.execute(
        select(PriceList, PriceFormat)
        .join(PriceFormat, PriceFormat.id == PriceList.price_format_id)
        .where(PriceList.number == price_list_id)
    ).first()

    if not pl_pf:
        raise HTTPException(status_code=404, detail="price list not found")

    pl, pf = pl_pf

    calc_rows = db.execute(
        select(CalculatedPrice, Product)
        .join(Product, Product.id == CalculatedPrice.product_id)
        .where(CalculatedPrice.price_list_id == pl.id)
        .order_by(Product.name.asc())
    ).all()

    if not calc_rows:
        return {
            "id": price_list_id,
            "meta": {
                "date": _fmt_dt(pl.created_at),
                "number": pl.number,
                "format": pf.code,
                "activationDate": _fmt_d(pl.activation_date),
                "user": pl.user,
                "status": pl.status,
                "branch": pf.branch,
            },
            "distribution": [
                {"name": "ЛП", "value": 0, "fill": "#EF4444"},
                {"name": "Зона логичности", "value": 0, "fill": "#10B981"},
                {"name": "ПП", "value": 0, "fill": "#F59E0B"},
                {"name": "Зона без цен", "value": 0, "fill": "#64748B"},
            ],
            "products": [],
        }

    def _zone_name(z: str) -> str:
        if z == "left":
            return "ЛП"
        if z == "optimal":
            return "Зона логичности"
        if z == "right":
            return "ПП"
        if z == "no-data":
            return "Зона без цен"
        return "—"

    zone_counts = {"left": 0, "optimal": 0, "right": 0, "no-data": 0}
    products: list[dict] = []
    for cp, product in calc_rows:
        competitor_price = float(cp.competitor_price) if cp.competitor_price is not None else None
        resolved = type("AppliedSource", (), {"applied_source": cp.applied_source_name or ""})()
        reference_price, deviation = _calculated_zone_reference_and_deviation(cp)

        zone = _calculated_zone(cp)
        if zone in zone_counts:
            zone_counts[zone] += 1

        products.append(
            {
                "product": product.name,
                "price": float(cp.final_price),
                "cost": float(cp.cost),
                "competitorPrice": competitor_price,
                "zoneReferencePrice": reference_price,
                "deviation": deviation,
                "source": resolved.applied_source if competitor_price is not None else "Наценка по ЦФ",
                "zone": zone,
            }
        )

    distribution = [
        {"name": _zone_name("left"), "value": zone_counts["left"], "fill": "#EF4444"},
        {"name": _zone_name("optimal"), "value": zone_counts["optimal"], "fill": "#10B981"},
        {"name": _zone_name("right"), "value": zone_counts["right"], "fill": "#F59E0B"},
        {"name": _zone_name("no-data"), "value": zone_counts["no-data"], "fill": "#64748B"},
    ]

    return {
        "id": price_list_id,
        "meta": {
            "date": _fmt_dt(pl.created_at),
            "number": pl.number,
            "format": pf.code,
            "activationDate": _fmt_d(pl.activation_date),
            "user": pl.user,
            "status": pl.status,
            "branch": pf.branch,
        },
        "distribution": distribution,
        "products": products,
    }


@app.get("/api/competitors")
def get_competitors_available():
    return data.COMPETITORS_AVAILABLE


@app.get("/api/debug/matching")
def debug_matching(
    format_code: str | None = Query(None),
    rebuild: bool = Query(False),
    row_limit: int = Query(50, ge=0, le=500),
    vidman_limit: int = Query(10, ge=0, le=50),
    db: Session = Depends(get_db),
):
    pf = None
    if format_code:
        pf = db.execute(select(PriceFormat).where(PriceFormat.code == format_code)).scalars().first()

    rebuild_summary = None
    if rebuild and pf is not None:
        rebuild_summary = rebuild_competitor_prices_for_selected(db=db, price_format_id=pf.id)
        recalculate_competitor_percentiles_if_needed(db=db, price_format_id=pf.id)
        db.commit()

    list_stmt = select(CompetitorPriceList)
    lists = db.execute(list_stmt).scalars().all()
    assigned_items = get_assigned_competitor_price_lists(db=db, price_format_id=int(pf.id)) if pf is not None else []
    assigned_ids = {int(item.price_list.id) for item in assigned_items}
    list_ids = [x.id for x in lists]

    items_total = 0
    items_matched = 0
    items_unmatched = 0
    prices_total = 0
    by_source: list[dict] = []
    selected_diagnostics: list[dict] = []
    sample_provisor_match_keys: list[str] = []
    sample_product_skus: list[str] = []
    exact_intersection_count = 0
    provisor_matched_by_composite = 0
    provisor_matched_by_distributor_only = 0

    if list_ids:
        items_total = db.scalar(
            select(func.count(CompetitorPriceListItem.id)).where(CompetitorPriceListItem.price_list_id.in_(list_ids))
        ) or 0
        items_matched = db.scalar(
            select(func.count(CompetitorPriceListItem.id))
            .where(CompetitorPriceListItem.price_list_id.in_(list_ids))
            .where(CompetitorPriceListItem.matched_sku != "")
            .where(CompetitorPriceListItem.product_id.is_not(None))
        ) or 0
        items_unmatched = int(items_total) - int(items_matched)

        for row in lists:
            total = db.scalar(
                select(func.count(CompetitorPriceListItem.id)).where(CompetitorPriceListItem.price_list_id == row.id)
            ) or 0
            matched = db.scalar(
                select(func.count(CompetitorPriceListItem.id))
                .where(CompetitorPriceListItem.price_list_id == row.id)
                .where(CompetitorPriceListItem.matched_sku != "")
                .where(CompetitorPriceListItem.product_id.is_not(None))
            ) or 0
            by_source.append(
                {
                    "id": row.id,
                    "priceListId": row.external_price_list_id or row.source_key,
                    "selected": int(row.id) in assigned_ids,
                    "source": row.source_type,
                    "sourceKey": row.source_key,
                    "account": row.account_login or row.account_id,
                    "name": row.display_name or row.supplier,
                    "totalItems": int(total),
                    "matchedItems": int(matched),
                    "unmatchedItems": int(total) - int(matched),
                    "matchRate": round((int(matched) / int(total) * 100), 2) if int(total) else 0,
                }
            )

    if pf is not None:
        product_codes = [normalize_external_sku(x) for x in db.execute(select(Product.code).limit(50)).scalars().all()]
        sample_product_skus = [x for x in product_codes if x][:20]
        product_variant_set: set[str] = set()
        for code in db.execute(select(Product.code)).scalars().all():
            for variant in normalize_sku_variants(code):
                product_variant_set.add(variant)
        selected_provisor_ids = [int(item.price_list.id) for item in assigned_items if item.price_list.source_type == "provisor"]
        provisor_rows = (
            db.execute(
                select(CompetitorPriceListItem)
                .join(CompetitorPriceList, CompetitorPriceList.id == CompetitorPriceListItem.price_list_id)
                .where(CompetitorPriceList.id.in_(selected_provisor_ids))
                .where(CompetitorPriceList.source_type == "provisor")
                .limit(5000)
            )
            .scalars()
            .all()
            if selected_provisor_ids
            else []
        )
        provisor_keys: list[str] = []
        for item in provisor_rows:
            try:
                raw = json.loads(item.raw_json or "{}")
                raw_inner = raw.get("raw") if isinstance(raw.get("raw"), dict) else raw
            except Exception:
                raw_inner = {}
            goods_id = item.provisor_goods_id or raw_inner.get("goodsId")
            expiry_date = item.expiry_date or raw_inner.get("shelfLife")
            distributor_goods_id = item.distributor_goods_id or raw_inner.get("distributorGoodsId")
            variants = provisor_item_variants(goods_id, distributor_goods_id, expiry_date)
            composite = next((value for kind, value in variants if kind == "composite"), "")
            if composite:
                provisor_keys.append(composite)
                if len(sample_provisor_match_keys) < 20:
                    sample_provisor_match_keys.append(composite)
            if composite and composite in product_variant_set:
                exact_intersection_count += 1
        provisor_matched_by_composite = db.scalar(
            select(func.count(CompetitorPriceListItem.id))
            .join(CompetitorPriceList, CompetitorPriceList.id == CompetitorPriceListItem.price_list_id)
            .where(CompetitorPriceList.id.in_(selected_provisor_ids))
            .where(CompetitorPriceList.source_type == "provisor")
            .where(CompetitorPriceListItem.matched_sku != "")
            .where(CompetitorPriceListItem.match_key == CompetitorPriceListItem.matched_sku)
        ) if selected_provisor_ids else 0
        provisor_matched_by_distributor_only = db.scalar(
            select(func.count(CompetitorPriceListItem.id))
            .join(CompetitorPriceList, CompetitorPriceList.id == CompetitorPriceListItem.price_list_id)
            .where(CompetitorPriceList.id.in_(selected_provisor_ids))
            .where(CompetitorPriceList.source_type == "provisor")
            .where(CompetitorPriceListItem.matched_sku != "")
            .where(CompetitorPriceListItem.match_key != CompetitorPriceListItem.matched_sku)
        ) if selected_provisor_ids else 0

    price_stmt = select(func.count(CompetitorPrice.id)).where(CompetitorPrice.product_id.is_not(None))
    if pf is not None:
        price_stmt = price_stmt.where(CompetitorPrice.price_format_id == pf.id)
    prices_total = db.scalar(price_stmt) or 0

    selected_lists = [x for x in lists if int(x.id) in assigned_ids]
    selected_sources = [f"{x.source_type}:{x.source_key}" for x in selected_lists]
    source_by_key = {f"{x.source_type}:{x.source_key}": x for x in selected_lists}

    selected_item_counts: dict[int, tuple[int, int]] = {}
    for row in selected_lists:
        total = db.scalar(
            select(func.count(CompetitorPriceListItem.id)).where(CompetitorPriceListItem.price_list_id == row.id)
        ) or 0
        matched = db.scalar(
            select(func.count(CompetitorPriceListItem.id))
            .where(CompetitorPriceListItem.price_list_id == row.id)
            .where(CompetitorPriceListItem.matched_sku != "")
            .where(CompetitorPriceListItem.product_id.is_not(None))
        ) or 0
        selected_item_counts[row.id] = (int(total), int(matched))
        extra: dict = {}
        if row.source_type == "provisor":
            with_distributor_sku = db.scalar(
                select(func.count(CompetitorPriceListItem.id))
                .where(CompetitorPriceListItem.price_list_id == row.id)
                .where(CompetitorPriceListItem.distributor_goods_id != "")
            ) or 0
            composite_matched = db.scalar(
                select(func.count(CompetitorPriceListItem.id))
                .where(CompetitorPriceListItem.price_list_id == row.id)
                .where(CompetitorPriceListItem.matched_sku != "")
                .where(CompetitorPriceListItem.match_key == CompetitorPriceListItem.matched_sku)
            ) or 0
            distributor_only_matched = db.scalar(
                select(func.count(CompetitorPriceListItem.id))
                .where(CompetitorPriceListItem.price_list_id == row.id)
                .where(CompetitorPriceListItem.matched_sku != "")
                .where(CompetitorPriceListItem.match_key != CompetitorPriceListItem.matched_sku)
            ) or 0
            extra = {
                "skuSource": "goodsId_distributorGoodsId_shelfLife",
                "goodsIdUsedAsSku": False,
                "goodsIdUsedInCompositeKey": True,
                "itemsWithDistributorGoodsId": int(with_distributor_sku),
                "itemsWithoutDistributorGoodsId": int(total) - int(with_distributor_sku),
                "matchedByCompositeKey": int(composite_matched),
                "matchedByDistributorGoodsIdOnly": int(distributor_only_matched),
            }
        selected_diagnostics.append(
            {
                "priceListId": row.external_price_list_id or row.source_key,
                "internalId": row.id,
                "source": row.source_type,
                "sourceKey": row.source_key,
                "account": row.account_login or row.account_id,
                "name": row.display_name or row.supplier,
                "totalItems": int(total),
                "matchedItems": int(matched),
                "unmatchedItems": int(total) - int(matched),
                "matchRate": round((int(matched) / int(total) * 100), 2) if int(total) else 0,
                "topUnmatched": explain_unmatched_for_price_list(db=db, price_list=row, limit=5),
                **extra,
            }
        )

    price_by_product_source: dict[tuple[int, str], float] = {}
    if selected_sources and pf is not None:
        price_rows = db.execute(
            select(CompetitorPrice.product_id, CompetitorPrice.source_name, CompetitorPrice.source_price)
            .where(CompetitorPrice.price_format_id == pf.id)
            .where(CompetitorPrice.product_id.is_not(None))
            .where(CompetitorPrice.source_name.in_(selected_sources))
            .where(CompetitorPrice.source_price.is_not(None))
        ).all()
        for product_id, source_name, source_price in price_rows:
            if product_id is None or source_price is None:
                continue
            key = (int(product_id), str(source_name))
            value = float(source_price)
            prev = price_by_product_source.get(key)
            if prev is None or value < prev:
                price_by_product_source[key] = value

    matched_items_by_product_list: dict[tuple[int, int], list[CompetitorPriceListItem]] = {}
    if selected_lists:
        matched_items = db.execute(
            select(CompetitorPriceListItem)
            .where(CompetitorPriceListItem.price_list_id.in_([x.id for x in selected_lists]))
            .where(CompetitorPriceListItem.product_id.is_not(None))
            .where(CompetitorPriceListItem.matched_sku != "")
        ).scalars().all()
        for item in matched_items:
            if item.product_id is None:
                continue
            matched_items_by_product_list.setdefault((int(item.product_id), int(item.price_list_id)), []).append(item)

    row_diagnostics: list[dict] = []
    if pf is not None and row_limit > 0:
        products = db.execute(select(Product).order_by(Product.code.asc()).limit(row_limit)).scalars().all()
        for product in products:
            found: dict[str, float] = {}
            missing: dict[str, dict] = {}
            for source in selected_sources:
                price_list = source_by_key[source]
                price = price_by_product_source.get((int(product.id), source))
                label = price_list.display_name or price_list.supplier or source
                if price is not None:
                    found[label] = price
                    continue

                matched_rows = matched_items_by_product_list.get((int(product.id), int(price_list.id)), [])
                if matched_rows:
                    positive_price = any(x.distributor_price is not None and float(x.distributor_price) > 0 for x in matched_rows)
                    reason = "old_data_not_saved" if positive_price else "matched_sku_exists_but_no_price"
                else:
                    total, matched = selected_item_counts.get(price_list.id, (0, 0))
                    reason = "source_item_unmatched" if total > 0 and matched == 0 else "no_matched_sku"
                missing[label] = {
                    "source": source,
                    "reason": reason,
                    "selected": True,
                }
            row_diagnostics.append(
                {
                    "sku": product.code,
                    "name": product.name,
                    "selectedCompetitorPriceLists": [
                        {
                            "source": source,
                            "name": source_by_key[source].display_name or source_by_key[source].supplier or source,
                        }
                        for source in selected_sources
                    ],
                    "pricesFound": found,
                    "pricesMissing": missing,
                }
            )

    return {
        "formatCode": format_code or "",
        "products": int(db.scalar(select(func.count(Product.id))) or 0),
        "priceLists": len(lists),
        "selectedPriceLists": len(selected_lists),
        "competitorItems": int(items_total),
        "withMatchedSku": int(items_matched),
        "withoutMatchedSku": int(items_unmatched),
        "competitorPricesForCalculation": int(prices_total),
        "sampleProvisorMatchKeys": sample_provisor_match_keys,
        "sampleProductSkus": sample_product_skus,
        "exactIntersectionCount": int(exact_intersection_count),
        "provisorMatchedByCompositeKey": int(provisor_matched_by_composite),
        "provisorMatchedByDistributorGoodsIdOnly": int(provisor_matched_by_distributor_only),
        "rebuildSummary": rebuild_summary,
        "selectedPriceListDiagnostics": selected_diagnostics,
        "rowDiagnostics": row_diagnostics,
        "vidmanUnmatchedTop": explain_vidman_unmatched(db=db, price_format_id=pf.id, limit=vidman_limit)
        if pf is not None and vidman_limit > 0
        else [],
        "bySource": by_source,
    }


@app.get("/api/price-formats/{format_code}/competitors")
def get_competitors_assigned(format_code: str):
    assigned_ids = data.COMPETITORS_ASSIGNED_BY_FORMAT.get(format_code, [])
    assigned = [x for x in data.COMPETITORS_AVAILABLE if x["id"] in assigned_ids]
    return {"format": format_code, "assigned": assigned, "assignedIds": assigned_ids}


@app.post("/api/price-formats/{format_code}/competitors")
def set_competitors_assigned(format_code: str, payload: dict, db: Session = Depends(get_db)):
    ids = payload.get("assignedIds")
    if not isinstance(ids, list) or not all(isinstance(x, int) for x in ids):
        raise HTTPException(status_code=400, detail="assignedIds must be list[int]")
    data.COMPETITORS_ASSIGNED_BY_FORMAT[format_code] = ids

    # Persist selection to DB so pricing uses the selected sources.
    # We map competitor.id -> competitor.name as source_name.
    pf = db.execute(select(PriceFormat).where(PriceFormat.code == format_code)).scalars().first()
    if pf is None:
        pf = PriceFormat(code=format_code, name=format_code)
        db.add(pf)
        db.flush()

    selected = [x for x in data.COMPETITORS_AVAILABLE if x.get("id") in ids]
    selected_source_names = {str(x.get("name") or "").strip() for x in selected if str(x.get("name") or "").strip()}

    # Delete configs that are no longer selected
    existing_cfg = db.execute(
        select(CompetitorPrice)
        .where(CompetitorPrice.price_format_id == pf.id)
        .where(CompetitorPrice.product_id.is_(None))
    ).scalars().all()

    for row in existing_cfg:
        if (row.source_name or "") not in selected_source_names:
            db.delete(row)

    # Upsert selected configs
    for comp in selected:
        source_name = str(comp.get("name") or "").strip()
        if not source_name:
            continue

        coeff = comp.get("coefficient")
        coefficient = float(coeff) if isinstance(coeff, (int, float)) else 1.0
        supplier = str(comp.get("supplier") or "").strip() or None

        cfg = db.execute(
            select(CompetitorPrice)
            .where(CompetitorPrice.price_format_id == pf.id)
            .where(CompetitorPrice.product_id.is_(None))
            .where(CompetitorPrice.source_name == source_name)
        ).scalars().first()

        if cfg is None:
            cfg = CompetitorPrice(
                price_format_id=pf.id,
                product_id=None,
                source_name=source_name,
                coefficient=coefficient,
            )
            db.add(cfg)

        cfg.coefficient = coefficient
        cfg.supplier = supplier

    db.commit()
    return {"format": format_code, "assignedIds": ids}


@app.get("/api/price-formats/{format_code}/competitor-price-lists")
def get_competitor_price_lists(
    format_code: str,
    account_id: str | None = Query(None),
    region: str | None = Query(None),
    db: Session = Depends(get_db),
):
    return list_competitor_price_lists(
        db=db,
        price_format_code=format_code,
        account_id=account_id,
        region=region,
    )


@app.get("/api/competitors/price-lists")
def get_competitor_sources(
    format_code: str = Query(..., min_length=1),
    account_id: str | None = Query(None),
    region: str | None = Query(None),
    db: Session = Depends(get_db),
):
    return list_competitor_sources(
        db=db,
        price_format_code=format_code,
        account_id=account_id,
        region=region,
    )


def _provisor_dedupe_key(row: CompetitorPriceList) -> tuple[str, ...]:
    def key(value: object) -> str:
        return str(value or "").strip().casefold()

    return (
        str(row.source_type or ""),
        str(row.source_key or ""),
        str(row.account_id or ""),
        str(row.external_price_list_id or ""),
        key(row.branch_id or row.branch_code or row.branch_name or row.region),
        key(row.competitor_name or row.supplier or row.display_name),
    )


def _product_sku_lookup(db: Session) -> set[str]:
    lookup: set[str] = set()
    for code in db.execute(select(Product.code)).scalars().all():
        for key in [normalize_external_sku(code), normalize_sku(code), *normalize_sku_variants(code)]:
            if key:
                lookup.add(key)
    return lookup


_INVISIBLE_SPACE_CHARS = {"\u00a0", "\u1680", "\u180e", "\u200b", "\u200c", "\u200d", "\u202f", "\u205f", "\u2060", "\u3000", "\ufeff"}


def _is_unicode_space_or_hidden(ch: str) -> bool:
    return ch.isspace() or ch in _INVISIBLE_SPACE_CHARS or unicodedata.category(ch) in {"Zs", "Zl", "Zp"}


def _strip_unicode_spaces(value: object) -> str:
    if value is None:
        return ""
    raw = str(value)
    left = 0
    right = len(raw)
    while left < right and _is_unicode_space_or_hidden(raw[left]):
        left += 1
    while right > left and _is_unicode_space_or_hidden(raw[right - 1]):
        right -= 1
    return raw[left:right]


def _remove_leading_zeroes(value: object) -> str:
    raw = _strip_unicode_spaces(value)
    stripped = raw.lstrip("0")
    return stripped or ("0" if raw else "")


def _numeric_only_sku(value: object) -> str:
    return re.sub(r"\D+", "", _strip_unicode_spaces(value))


def _numeric_without_leading_zeroes(value: object) -> str:
    digits = _numeric_only_sku(value)
    return digits.lstrip("0") or ("0" if digits else "")


def _sku_text_diagnostics(value: object) -> dict:
    raw = "" if value is None else str(value)
    hidden = []
    for ch in raw:
        if _is_unicode_space_or_hidden(ch) and ch not in {" ", "\t", "\r", "\n"}:
            label = unicodedata.name(ch, "UNKNOWN")
            code = f"U+{ord(ch):04X}"
            if code not in [x["codepoint"] for x in hidden]:
                hidden.append({"codepoint": code, "name": label})
    return {
        "hasLeadingOrTrailingSpace": raw != raw.strip(),
        "hasUnicodeEdgeSpace": raw != _strip_unicode_spaces(raw),
        "hasAnyWhitespace": any(ch.isspace() or ch in _INVISIBLE_SPACE_CHARS for ch in raw),
        "hiddenUnicodeSpaces": hidden,
        "casefold": raw.casefold(),
        "trim": raw.strip(),
        "unicodeTrim": _strip_unicode_spaces(raw),
        "removeLeadingZeros": _remove_leading_zeroes(raw),
        "numericOnly": _numeric_only_sku(raw),
        "numericWithoutLeadingZeros": _numeric_without_leading_zeroes(raw),
        "punctuationRemovedUpper": re.sub(r"[-/.\s]+", "", _strip_unicode_spaces(raw)).upper(),
    }


def _sample_values(values: set[str], *, limit: int = 3) -> list[str]:
    return sorted(v for v in values if v)[:limit]


def _provisor_reference_filial_audit(db: Session, *, limit: int = 10) -> dict:
    product_skus = _product_sku_lookup(db)
    product_rows = (
        db.execute(select(Product, ProductExtra).join(ProductExtra, ProductExtra.product_id == Product.id, isouter=True))
        .all()
    )
    products_by_code: dict[str, tuple[Product, ProductExtra | None]] = {}
    code_trim: set[str] = set()
    code_no_zeroes: set[str] = set()
    code_numeric: set[str] = set()
    code_numeric_no_zeroes: set[str] = set()
    code_casefold: set[str] = set()
    code_punctuationless_upper: set[str] = set()
    candidate_codes: list[str] = []
    product_name_text: set[str] = set()
    product_extra_text: set[str] = set()
    product_codes_by_variant: dict[str, set[str]] = {}

    def add_variant(bucket: dict[str, set[str]], key: str, code: str) -> None:
        if key:
            bucket.setdefault(key, set()).add(code)

    for product, extra in product_rows:
        code = str(product.code or "")
        products_by_code[code] = (product, extra)
        candidate_codes.append(code)
        code_trim.add(code.strip())
        code_no_zeroes.add(_remove_leading_zeroes(code))
        code_numeric.add(_numeric_only_sku(code))
        code_numeric_no_zeroes.add(_numeric_without_leading_zeroes(code))
        code_casefold.add(code.casefold())
        code_punctuationless_upper.add(re.sub(r"[-/.\s]+", "", _strip_unicode_spaces(code)).upper())
        product_name_text.add(str(product.name or "").casefold())
        if extra is not None:
            product_extra_text.add(str(extra.manufacturer or "").casefold())
        for variant in {
            normalize_external_sku(code),
            normalize_sku(code) or "",
            *normalize_sku_variants(code),
            code.strip(),
            _remove_leading_zeroes(code),
            _numeric_only_sku(code),
            _numeric_without_leading_zeroes(code),
            code.casefold(),
            re.sub(r"[-/.\s]+", "", _strip_unicode_spaces(code)).upper(),
        }:
            add_variant(product_codes_by_variant, variant, code)

    external_sku_fields: dict[str, set[str]] = {
        "BranchStock.sku": {str(v or "") for v in db.execute(select(BranchStock.sku)).scalars().all()},
        "BranchCost.sku": {str(v or "") for v in db.execute(select(BranchCost.sku)).scalars().all()},
        "ProductRating.sku": {str(v or "") for v in db.execute(select(ProductRating.sku)).scalars().all()},
        "CompetitorCodeMapping.our_sku": {str(v or "") for v in db.execute(select(CompetitorCodeMapping.our_sku)).scalars().all()},
        "SourceGoodsMatch.distributor_goods_id": {
            str(v or "")
            for v in db.execute(
                select(SourceGoodsMatch.distributor_goods_id).where(SourceGoodsMatch.source_type == "provisor")
            )
            .scalars()
            .all()
        },
        "ProductSubstituteMatch.source_distributor_goods_id": {
            str(v or "")
            for v in db.execute(
                select(ProductSubstituteMatch.source_distributor_goods_id).where(ProductSubstituteMatch.source_type == "provisor")
            )
            .scalars()
            .all()
        },
    }
    external_sku_normalized: dict[str, set[str]] = {
        field: {key for value in values for key in {value, value.strip(), normalize_external_sku(value), normalize_sku(value) or "", *normalize_sku_variants(value)} if key}
        for field, values in external_sku_fields.items()
    }

    rows = (
        db.execute(
            select(CompetitorPriceList, CompetitorPriceListItem)
            .join(CompetitorPriceListItem, CompetitorPriceListItem.price_list_id == CompetitorPriceList.id)
            .where(CompetitorPriceList.source_type == "provisor")
            .where(CompetitorPriceListItem.filial_id.in_(PROVISOR_REFERENCE_FILIAL_IDS))
            .order_by(CompetitorPriceListItem.id.asc())
        )
        .all()
    )
    rows_with_goods_id = 0
    matched_sku = 0
    normalization_summary = {
        "trim": 0,
        "removeLeadingZeros": 0,
        "numericOnly": 0,
        "numericWithoutLeadingZeros": 0,
        "caseInsensitive": 0,
        "hyphenSlashDotSpaceInsensitive": 0,
        "existingNormalizeSkuVariants": 0,
    }
    other_field_summary = {
        "Product.nameContainsDistributorGoodsId": 0,
        "ProductExtra.manufacturerContainsDistributorGoodsId": 0,
        **{field: 0 for field in external_sku_fields},
    }
    missing_examples: list[dict] = []

    def strategy_matches(value: object) -> dict[str, bool]:
        raw = "" if value is None else str(value)
        existing_variants = {key for key in [normalize_external_sku(raw), normalize_sku(raw), *normalize_sku_variants(raw)] if key}
        return {
            "trim": raw.strip() in code_trim,
            "removeLeadingZeros": _remove_leading_zeroes(raw) in code_no_zeroes,
            "numericOnly": _numeric_only_sku(raw) in code_numeric,
            "numericWithoutLeadingZeros": _numeric_without_leading_zeroes(raw) in code_numeric_no_zeroes,
            "caseInsensitive": raw.casefold() in code_casefold,
            "hyphenSlashDotSpaceInsensitive": re.sub(r"[-/.\s]+", "", _strip_unicode_spaces(raw)).upper() in code_punctuationless_upper,
            "existingNormalizeSkuVariants": any(key in product_skus for key in existing_variants),
        }

    def other_field_matches(value: object) -> dict[str, bool]:
        raw = "" if value is None else str(value)
        raw_cf = raw.casefold()
        external_variants = {key for key in [raw, raw.strip(), normalize_external_sku(raw), normalize_sku(raw) or "", *normalize_sku_variants(raw)] if key}
        matches = {
            "Product.nameContainsDistributorGoodsId": bool(raw_cf and any(raw_cf in name for name in product_name_text)),
            "ProductExtra.manufacturerContainsDistributorGoodsId": bool(raw_cf and any(raw_cf in manufacturer for manufacturer in product_extra_text)),
        }
        matches.update({field: any(key in values for key in external_variants) for field, values in external_sku_normalized.items()})
        return matches

    def closest_candidates(value: object, *, candidate_limit: int = 5) -> list[dict]:
        raw = "" if value is None else str(value)
        candidate_keys = [
            raw,
            raw.strip(),
            normalize_external_sku(raw),
            normalize_sku(raw) or "",
            _remove_leading_zeroes(raw),
            _numeric_only_sku(raw),
            _numeric_without_leading_zeroes(raw),
            raw.casefold(),
            re.sub(r"[-/.\s]+", "", _strip_unicode_spaces(raw)).upper(),
        ]
        codes: list[str] = []
        reasons: dict[str, set[str]] = {}
        for key in candidate_keys:
            for code in _sample_values(product_codes_by_variant.get(key, set()), limit=candidate_limit):
                if code not in codes:
                    codes.append(code)
                reasons.setdefault(code, set()).add("normalized-equal")
        for code in difflib.get_close_matches(raw, candidate_codes, n=candidate_limit, cutoff=0.72):
            if code not in codes:
                codes.append(code)
            reasons.setdefault(code, set()).add("difflib")
        out = []
        for code in codes[:candidate_limit]:
            product, extra = products_by_code.get(code, (None, None))  # type: ignore[assignment]
            out.append(
                {
                    "productId": product.id if product is not None else None,
                    "code": code,
                    "name": product.name if product is not None else "",
                    "manufacturer": extra.manufacturer if extra is not None else "",
                    "reason": sorted(reasons.get(code, set())),
                }
            )
        return out

    for price_list, item in rows:
        if item.provisor_goods_id is not None:
            rows_with_goods_id += 1
        variants = [
            normalize_external_sku(item.distributor_goods_id),
            normalize_sku(item.distributor_goods_id),
            *normalize_sku_variants(item.distributor_goods_id),
        ]
        if any(key and key in product_skus for key in variants):
            matched_sku += 1
            continue

        strategy = strategy_matches(item.distributor_goods_id)
        for key, value in strategy.items():
            if value:
                normalization_summary[key] += 1
        field_matches = other_field_matches(item.distributor_goods_id)
        for key, value in field_matches.items():
            if value:
                other_field_summary[key] += 1
        if len(missing_examples) < limit:
            name = item.raw_name or item.name or item.distributor_goods_name
            missing_examples.append(
                {
                    "priceListId": price_list.id,
                    "sourceKey": price_list.source_key,
                    "accountId": price_list.account_id,
                    "accountLogin": price_list.account_login,
                    "externalPriceListId": price_list.external_price_list_id,
                    "displayName": price_list.display_name,
                    "branchId": price_list.branch_id,
                    "branchName": price_list.branch_name,
                    "competitorName": price_list.competitor_name,
                    "sourceUpdatedAt": price_list.source_updated_at,
                    "priceDate": price_list.price_date.isoformat() if price_list.price_date else None,
                    "itemId": item.id,
                    "goodsId": item.provisor_goods_id,
                    "distributorGoodsId": item.distributor_goods_id,
                    "normalizedDistributorGoodsId": normalize_sku(item.distributor_goods_id),
                    "externalNormalizedDistributorGoodsId": normalize_external_sku(item.distributor_goods_id),
                    "name": name,
                    "manufacturer": item.raw_manufacturer or item.normalized_manufacturer,
                    "normalization": _sku_text_diagnostics(item.distributor_goods_id),
                    "normalizationWouldMatchProductCode": strategy,
                    "matchedOtherFields": [key for key, value in field_matches.items() if value],
                    "closestProductCodeCandidates": closest_candidates(item.distributor_goods_id),
                }
            )
    return {
        "referenceFilialId": PROVISOR_REFERENCE_FILIAL_ID,
        "referenceFilialIds": list(PROVISOR_REFERENCE_FILIAL_IDS),
        "rowsTotal": len(rows),
        "rowsWithGoodsId": rows_with_goods_id,
        "distributorGoodsIdMatchedProducts": matched_sku,
        "skuNotFound": max(0, len(rows) - matched_sku),
        "skuNotFoundNormalizationSummary": normalization_summary,
        "skuNotFoundOtherFieldSummary": other_field_summary,
        "topSkuNotFoundExamples": missing_examples,
    }


@app.get("/api/competitors/provisor-diagnostics")
def get_provisor_diagnostics(
    format_code: str = Query(..., min_length=1),
    region: str | None = Query(None),
    db: Session = Depends(get_db),
):
    pf = db.execute(select(PriceFormat).where(PriceFormat.code == format_code)).scalars().first()
    if pf is None:
        raise HTTPException(status_code=404, detail="price format not found")

    rows = (
        db.execute(
            select(CompetitorPriceList)
            .where(CompetitorPriceList.source_type == "provisor")
            .order_by(CompetitorPriceList.updated_at.desc(), CompetitorPriceList.id.desc())
        )
        .scalars()
        .all()
    )
    row_ids = [int(row.id) for row in rows]
    item_counts = (
        dict(
            db.execute(
                select(CompetitorPriceListItem.price_list_id, func.count(CompetitorPriceListItem.id))
                .where(CompetitorPriceListItem.price_list_id.in_(row_ids))
                .group_by(CompetitorPriceListItem.price_list_id)
            ).all()
        )
        if row_ids
        else {}
    )
    branch_matched = [row for row in rows if price_format_branch_matches(row, pf, region)]
    usable_rows_with_items = [row for row in rows if int(item_counts.get(row.id, 0)) > 0]
    rows_with_items = [row for row in branch_matched if int(item_counts.get(row.id, 0)) > 0]
    deduped: dict[tuple[str, ...], CompetitorPriceList] = {}
    for row in rows_with_items:
        key = _provisor_dedupe_key(row)
        current = deduped.get(key)
        if current is None or (row.updated_at or datetime.min, int(row.id or 0)) > (current.updated_at or datetime.min, int(current.id or 0)):
            deduped[key] = row

    assignments = (
        db.execute(
            select(PriceFormatCompetitorAssignment, CompetitorPriceList)
            .join(CompetitorPriceList, CompetitorPriceList.id == PriceFormatCompetitorAssignment.competitor_price_list_id)
            .where(PriceFormatCompetitorAssignment.price_format_id == pf.id)
            .where(CompetitorPriceList.source_type == "provisor")
        )
        .all()
    )
    active_price_list_ids = [int(row.id) for assignment, row in assignments if assignment.is_active]
    match_type_distribution = [
        {"matchType": str(match_type or "unmatched"), "rows": int(count)}
        for match_type, count in (
            db.execute(
                select(CompetitorPriceListItem.match_type, func.count(CompetitorPriceListItem.id))
                .where(CompetitorPriceListItem.price_list_id.in_(active_price_list_ids))
                .group_by(CompetitorPriceListItem.match_type)
                .order_by(func.count(CompetitorPriceListItem.id).desc())
            ).all()
            if active_price_list_ids
            else []
        )
    ]
    product_goods_ids = {
        int(goods_id)
        for goods_id in db.execute(select(Product.provisor_goods_id).where(Product.provisor_goods_id.is_not(None))).scalars().all()
        if goods_id is not None
    }
    unmatched_rows = (
        db.execute(
            select(CompetitorPriceListItem.provisor_goods_id)
            .where(CompetitorPriceListItem.price_list_id.in_(active_price_list_ids))
            .where(CompetitorPriceListItem.provisor_goods_id.is_not(None))
            .where((CompetitorPriceListItem.product_id.is_(None)) | (CompetitorPriceListItem.matched_sku == ""))
        )
        .scalars()
        .all()
        if active_price_list_ids
        else []
    )
    unmatched_with_goods_id = len(unmatched_rows)
    unmatched_goods_id_exists = sum(1 for goods_id in unmatched_rows if int(goods_id) in product_goods_ids)
    total_products = int(db.scalar(select(func.count(Product.id))) or 0)
    products_with_goods = int(db.scalar(select(func.count(Product.id)).where(Product.provisor_goods_id.is_not(None))) or 0)

    active_lists = [
        {
            "id": row.id,
            "sourceKey": row.source_key,
            "accountId": row.account_id,
            "accountLogin": row.account_login,
            "branchName": row.branch_name,
            "competitorName": row.competitor_name,
            "itemsCount": int(item_counts.get(row.id, 0)),
        }
        for assignment, row in assignments
        if assignment.is_active
    ]
    visible_ids = {int(row.id) for row in deduped.values()}
    selected_ids = {int(row.id) for _, row in assignments}

    return {
        "formatCode": format_code,
        "branch": region or pf.branch or "",
        "visibility": {
            "totalProvisorGlobalPool": len(rows),
            "totalProvisor": len(rows),
            "usableWithItems": len(usable_rows_with_items),
            "visibleForFormatBranch": len(visible_ids),
            "hiddenByBranch": max(0, len(rows) - len(branch_matched)),
            "hiddenZeroItems": max(0, len(branch_matched) - len(rows_with_items)),
            "zeroItemsHidden": max(0, len(rows) - len(usable_rows_with_items)),
            "hiddenByDedupe": max(0, len(rows_with_items) - len(visible_ids)),
            "returnedGlobalPool": len(usable_rows_with_items),
            "selectedTotal": len(selected_ids),
            "activeTotal": len(active_price_list_ids),
        },
        "coverage": {
            "totalProducts": total_products,
            "productsWithProvisorGoodsId": products_with_goods,
            "productsWithoutProvisorGoodsId": max(0, total_products - products_with_goods),
            "activeProvisorPriceLists": active_lists,
            "matchTypeDistribution": match_type_distribution,
            "unmatchedRowsWithGoodsId": unmatched_with_goods_id,
            "unmatchedRowsWhoseGoodsIdExistsInProduct": unmatched_goods_id_exists,
        },
        "referenceFilialCoverage": _provisor_reference_filial_audit(db),
    }


@app.get("/api/competitors/percentiles")
def get_competitor_percentile_sources(
    format_code: str | None = Query(None),
    db: Session = Depends(get_db),
):
    return list_percentile_sources(db=db, price_format_code=format_code)


@app.get("/api/competitors/percentile-rows")
def get_competitor_percentile_rows(
    format_code: str = Query(...),
    region: str = Query(""),
    competitor: str = Query(""),
    q: str = Query(""),
    percentile_filter: str = Query("all"),
    competitor_filter: str = Query("all"),
    sort: str = Query("sku"),
    direction: str = Query("asc"),
    page: int = Query(1, ge=1),
    page_size: int = Query(100, ge=20, le=500),
    db: Session = Depends(get_db),
):
    return list_percentile_product_rows(
        db=db,
        price_format_code=format_code,
        region=region,
        competitor=competitor,
        q=q,
        percentile_filter=percentile_filter,
        competitor_filter=competitor_filter,
        sort=sort,
        direction=direction,
        page=page,
        page_size=page_size,
    )


@app.get("/api/competitors/percentile-trace")
def get_competitor_percentile_trace(
    format_code: str = Query(...),
    region: str = Query(...),
    competitor: str = Query(...),
    sku: str = Query(...),
    db: Session = Depends(get_db),
):
    return percentile_trace(
        db=db,
        price_format_code=format_code,
        region=region,
        competitor=competitor,
        sku=sku,
    )


@app.get("/api/competitors/percentile-coverage-audit")
def get_competitor_percentile_coverage_audit(
    format_code: str = Query(...),
    region: str = Query(...),
    competitor: str = Query(...),
    db: Session = Depends(get_db),
):
    return percentile_coverage_audit(
        db=db,
        price_format_code=format_code,
        region=region,
        competitor=competitor,
    )


@app.get("/api/competitors/percentile-rows/export.{fmt}")
def export_competitor_percentile_rows_endpoint(
    fmt: str,
    format_code: str = Query(...),
    region: str = Query(""),
    competitor: str = Query(""),
    q: str = Query(""),
    percentile_filter: str = Query("all"),
    competitor_filter: str = Query("all"),
    sort: str = Query("sku"),
    direction: str = Query("asc"),
    db: Session = Depends(get_db),
):
    if fmt not in {"csv", "xlsx"}:
        raise HTTPException(status_code=400, detail="fmt must be csv or xlsx")
    filename, content, media_type = export_percentile_product_rows(
        db=db,
        price_format_code=format_code,
        fmt=fmt,
        region=region,
        competitor=competitor,
        q=q,
        percentile_filter=percentile_filter,
        competitor_filter=competitor_filter,
        sort=sort,
        direction=direction,
    )
    return StreamingResponse(
        io.BytesIO(content),
        media_type=media_type,
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
    )


def _assignment_percentile_source_name(source_id: object) -> str:
    return f"percentile:{str(source_id or '').strip()}"


def _selected_competitor_rows(db: Session, pf: PriceFormat) -> list[CompetitorPriceList]:
    return [item.price_list for item in get_assigned_competitor_price_lists(db=db, price_format_id=int(pf.id))]


def _assignment_row_from_price_list(row: CompetitorPriceList, items_count: int = 0, assignment: object | None = None) -> dict:
    return {
        "id": str(row.id),
        "sourceId": row.id,
        "sourceType": row.source_type,
        "sourceKey": row.source_key,
        "sourceName": row.display_name or row.supplier or f"{row.source_type}:{row.source_key}",
        "region": row.branch_name or row.region or "",
        "branchName": row.branch_name or "",
        "competitorName": row.competitor_name or row.supplier or row.display_name or "",
        "accountId": row.account_id or "",
        "accountLogin": row.account_login or "",
        "coefficient": float(getattr(assignment, "coefficient", row.coefficient) or 1.0),
        "priceDate": row.price_date.isoformat() if row.price_date else "",
        "itemsCount": int(items_count or 0),
        "active": bool(getattr(assignment, "is_active", row.is_selected)),
        "isPercentile": False,
    }


def _assignment_row_from_percentile(source: dict, cfg: CompetitorPrice | None = None) -> dict:
    source_id = str(source.get("id") or "")
    return {
        "id": _assignment_percentile_source_name(source_id),
        "sourceId": source_id,
        "sourceType": "percentile",
        "sourceKey": source_id,
        "sourceName": source.get("name") or f"{source.get('region') or ''} - P{source.get('percentile') or ''}",
        "region": source.get("region") or "",
        "branchName": source.get("region") or "",
        "competitorName": source.get("competitor") or "Эмити",
        "accountId": "percentile",
        "accountLogin": f"Персентиль {source.get('percentile') or ''}".strip(),
        "coefficient": float(cfg.coefficient if cfg is not None and cfg.coefficient is not None else 1.0),
        "priceDate": str(source.get("generatedAt") or "")[:10],
        "itemsCount": int(source.get("skuCount") or 0),
        "active": True,
        "isPercentile": True,
    }


@app.get("/api/price-formats/{format_code}/competitor-assignments")
def get_competitor_assignments(format_code: str, db: Session = Depends(get_db), current_user: AppUser = Depends(get_current_user)):
    pf = db.execute(select(PriceFormat).where(PriceFormat.code == format_code)).scalars().first()
    if pf is None:
        raise HTTPException(status_code=404, detail="price format not found")
    _ensure_price_format_access(pf, current_user)

    selected = get_assigned_competitor_price_lists(db=db, price_format_id=int(pf.id))
    counts = (
        dict(
            db.execute(
                select(CompetitorPriceListItem.price_list_id, func.count(CompetitorPriceListItem.id))
                .where(CompetitorPriceListItem.price_list_id.in_([item.price_list.id for item in selected]))
                .group_by(CompetitorPriceListItem.price_list_id)
            ).all()
        )
        if selected
        else {}
    )
    rows = [
        _assignment_row_from_price_list(item.price_list, int(counts.get(item.price_list.id, 0)), item.assignment)
        for item in selected
    ]

    percentile_cfgs = (
        db.execute(
            select(CompetitorPrice)
            .where(CompetitorPrice.price_format_id == pf.id)
            .where(CompetitorPrice.product_id.is_(None))
            .where(CompetitorPrice.source_name.like("percentile:%"))
        )
        .scalars()
        .all()
    )
    cfg_by_source_name = {cfg.source_name: cfg for cfg in percentile_cfgs}
    for source in list_percentile_sources(db=db, price_format_code=format_code):
        source_id = str(source.get("id") or "")
        cfg = cfg_by_source_name.get(_assignment_percentile_source_name(source_id))
        if cfg is not None:
            rows.append(_assignment_row_from_percentile(source, cfg))
    return rows


@app.post("/api/price-formats/{format_code}/competitor-assignments")
def post_competitor_assignment(format_code: str, payload: dict = Body(...), db: Session = Depends(get_db), current_user: AppUser = Depends(require_write_access)):
    pf = db.execute(select(PriceFormat).where(PriceFormat.code == format_code)).scalars().first()
    if pf is None:
        raise HTTPException(status_code=404, detail="price format not found")
    _ensure_price_format_access(pf, current_user)

    source_type = str(payload.get("sourceType") or "").strip()
    source_id = payload.get("sourceId")
    try:
        coefficient = float(payload.get("coefficient", 1.0) or 1.0)
    except Exception:
        coefficient = 1.0

    if source_type == "percentile":
        source_name = _assignment_percentile_source_name(source_id)
        existing = (
            db.execute(
                select(CompetitorPrice)
                .where(CompetitorPrice.price_format_id == pf.id)
                .where(CompetitorPrice.product_id.is_(None))
                .where(CompetitorPrice.source_name == source_name)
            )
            .scalars()
            .first()
        )
        if existing is not None:
            raise HTTPException(status_code=409, detail="Источник уже назначен")
        cfg = CompetitorPrice(
            price_format_id=pf.id,
            product_id=None,
            source_name=source_name,
            supplier=str(payload.get("sourceName") or "percentile"),
            coefficient=coefficient,
        )
        db.add(cfg)
        db.commit()
        return {"status": "ok"}

    try:
        source_id_int = int(source_id)
    except Exception:
        raise HTTPException(status_code=400, detail="sourceId must be int for competitor price list")

    row = db.get(CompetitorPriceList, source_id_int)
    if row is None:
        raise HTTPException(status_code=404, detail="source not found")
    existing_assignment = get_assignment(db=db, price_format_id=int(pf.id), competitor_price_list_id=source_id_int)
    if existing_assignment is not None and existing_assignment.is_active:
        raise HTTPException(status_code=409, detail="Источник уже назначен")
    upsert_assignment(db=db, price_format_id=int(pf.id), competitor_price_list_id=source_id_int, coefficient=coefficient, is_active=True)
    sync_selected_competitor_configs(db=db, price_format_id=pf.id)
    db.commit()
    return {"status": "ok"}


@app.patch("/api/price-formats/{format_code}/competitor-assignments/{assignment_id}")
def patch_competitor_assignment(format_code: str, assignment_id: str, payload: dict = Body(...), db: Session = Depends(get_db), current_user: AppUser = Depends(require_write_access)):
    pf = db.execute(select(PriceFormat).where(PriceFormat.code == format_code)).scalars().first()
    if pf is None:
        raise HTTPException(status_code=404, detail="price format not found")
    _ensure_price_format_access(pf, current_user)
    try:
        coefficient = float(payload.get("coefficient", 1.0) or 1.0)
    except Exception:
        coefficient = 1.0
    active = payload.get("active")

    if assignment_id.startswith("percentile:"):
        source_name = _assignment_percentile_source_name(assignment_id.removeprefix("percentile:"))
        cfg = (
            db.execute(
                select(CompetitorPrice)
                .where(CompetitorPrice.price_format_id == pf.id)
                .where(CompetitorPrice.product_id.is_(None))
                .where(CompetitorPrice.source_name == source_name)
            )
            .scalars()
            .first()
        )
        if cfg is None:
            raise HTTPException(status_code=404, detail="assignment not found")
        if active is False:
            db.delete(cfg)
        else:
            cfg.coefficient = coefficient
        db.commit()
        return {"status": "ok"}

    try:
        source_id_int = int(assignment_id)
    except Exception:
        raise HTTPException(status_code=400, detail="assignment id must be int or percentile source id")
    row = db.get(CompetitorPriceList, source_id_int)
    assignment = get_assignment(db=db, price_format_id=int(pf.id), competitor_price_list_id=source_id_int)
    if row is None or assignment is None:
        raise HTTPException(status_code=404, detail="assignment not found")
    assignment.coefficient = coefficient
    if active is not None:
        assignment.is_active = bool(active)
    assignment.updated_at = now_kz_naive()
    sync_selected_competitor_configs(db=db, price_format_id=pf.id)
    db.commit()
    return {"status": "ok"}


@app.delete("/api/price-formats/{format_code}/competitor-assignments/{assignment_id}")
def delete_competitor_assignment(format_code: str, assignment_id: str, db: Session = Depends(get_db), current_user: AppUser = Depends(require_write_access)):
    pf = db.execute(select(PriceFormat).where(PriceFormat.code == format_code)).scalars().first()
    if pf is None:
        raise HTTPException(status_code=404, detail="price format not found")
    _ensure_price_format_access(pf, current_user)
    if assignment_id.startswith("percentile:"):
        source_name = _assignment_percentile_source_name(assignment_id.removeprefix("percentile:"))
        cfg = (
            db.execute(
                select(CompetitorPrice)
                .where(CompetitorPrice.price_format_id == pf.id)
                .where(CompetitorPrice.product_id.is_(None))
                .where(CompetitorPrice.source_name == source_name)
            )
            .scalars()
            .first()
        )
        if cfg is not None:
            db.delete(cfg)
            db.commit()
        return {"status": "ok"}
    try:
        source_id_int = int(assignment_id)
    except Exception:
        raise HTTPException(status_code=400, detail="assignment id must be int or percentile source id")
    row = db.get(CompetitorPriceList, source_id_int)
    assignment = get_assignment(db=db, price_format_id=int(pf.id), competitor_price_list_id=source_id_int)
    if row is None or assignment is None:
        raise HTTPException(status_code=404, detail="assignment not found")
    assignment.is_active = False
    assignment.updated_at = now_kz_naive()
    sync_selected_competitor_configs(db=db, price_format_id=pf.id)
    db.commit()
    return {"status": "ok"}


@app.get("/api/competitors/mappings")
def get_competitor_mappings(
    search: str | None = Query(None),
    limit: int = Query(200, ge=1, le=1000),
    db: Session = Depends(get_db),
):
    return list_competitor_mappings(db=db, search=search, limit=limit)


@app.delete("/api/competitors/mappings/substitutes/{mapping_id}")
def delete_competitor_substitute_mapping(mapping_id: int, db: Session = Depends(get_db)):
    try:
        delete_substitute_mapping(db=db, mapping_id=mapping_id)
        return {"status": "ok"}
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))


def _price_format_id_or_none(db: Session, format_code: str | None) -> int | None:
    if not format_code:
        return None
    pf = db.execute(select(PriceFormat).where(PriceFormat.code == format_code.strip())).scalars().first()
    return int(pf.id) if pf is not None else None


def _code_mapping_source_payload(payload: dict, platform: str, db: Session) -> dict:
    item_id = payload.get("itemId") or payload.get("item_id")
    if item_id not in (None, ""):
        item = db.get(CompetitorPriceListItem, int(item_id))
        if item is None:
            raise HTTPException(status_code=404, detail="competitor item not found")
        price_list = db.get(CompetitorPriceList, item.price_list_id)
        if price_list is None or price_list.source_type != platform:
            raise HTTPException(status_code=400, detail="competitor item platform mismatch")
        return mapping_source_payload(platform, item)
    return {
        "source_external_key": payload.get("sourceExternalKey") or payload.get("source_external_key"),
        "source_match_key": payload.get("sourceMatchKey") or payload.get("source_match_key"),
        "source_name": payload.get("sourceName") or payload.get("source_name") or "",
        "source_manufacturer": payload.get("sourceManufacturer") or payload.get("source_manufacturer") or "",
        "source_dosage_form": payload.get("sourceDosageForm") or payload.get("source_dosage_form") or "",
        "source_normalized_name": payload.get("sourceNormalizedName") or payload.get("source_normalized_name") or "",
    }


@app.get("/api/competitors/code-mappings")
def get_competitor_code_mappings(
    platform: str = Query("provisor"),
    status: str = Query("all"),
    format_code: str | None = Query(None),
    source_q: str = Query(""),
    product_q: str = Query(""),
    limit: int = Query(200, ge=1, le=1000),
    db: Session = Depends(get_db),
):
    try:
        normalized_platform = platform_from_value(platform)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    return list_code_mappings(
        db=db,
        platform=normalized_platform,
        price_format_id=_price_format_id_or_none(db, format_code),
        status=status,
        source_q=source_q,
        product_q=product_q,
        limit=limit,
    )


@app.get("/api/competitors/code-mappings/catalog-view")
def get_competitor_code_mappings_catalog_view(
    platform: str = Query("provisor"),
    status: str = Query("all"),
    format_code: str | None = Query(None),
    source_q: str = Query(""),
    product_q: str = Query(""),
    page: int = Query(1, ge=1),
    limit: int = Query(300, ge=1, le=1000),
    include_candidates: bool = Query(True),
    db: Session = Depends(get_db),
):
    try:
        normalized_platform = platform_from_value(platform)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    return list_catalog_code_mappings(
        db=db,
        platform=normalized_platform,
        price_format_id=_price_format_id_or_none(db, format_code),
        status=status,
        source_q=source_q,
        product_q=product_q,
        page=page,
        limit=limit,
        include_candidates=include_candidates,
    )


@app.get("/api/products/search")
def search_products_for_mapping(
    q: str = Query(..., min_length=1),
    limit: int = Query(30, ge=1, le=100),
    db: Session = Depends(get_db),
):
    return find_products_for_mapping(db=db, q=q, limit=limit)


@app.get("/api/competitor-items/search")
def search_competitor_items_for_mapping(
    platform: str = Query("provisor"),
    q: str = Query(..., min_length=1),
    format_code: str | None = Query(None),
    limit: int = Query(30, ge=1, le=100),
    db: Session = Depends(get_db),
):
    try:
        normalized_platform = platform_from_value(platform)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    result = list_code_mappings(
        db=db,
        platform=normalized_platform,
        price_format_id=_price_format_id_or_none(db, format_code),
        status="all",
        source_q=q,
        product_q="",
        limit=limit,
    )
    return result.get("items", [])


@app.post("/api/competitors/code-mappings")
def create_competitor_code_mapping(payload: dict = Body(...), db: Session = Depends(get_db)):
    try:
        platform = platform_from_value(payload.get("platform"))
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    status = str(payload.get("status") or "mapped").strip().lower()
    if status not in {"mapped", "unmapped", "rejected"}:
        raise HTTPException(status_code=400, detail="status must be mapped, unmapped or rejected")
    product = None
    product_id = payload.get("ourProductId") or payload.get("productId") or payload.get("our_product_id")
    if status == "mapped":
        if product_id in (None, ""):
            raise HTTPException(status_code=400, detail="ourProductId is required for mapped status")
        product = db.get(Product, int(product_id))
        if product is None:
            raise HTTPException(status_code=404, detail="product not found")
    source_payload = _code_mapping_source_payload(payload, platform, db)
    row = upsert_code_mapping(
        db=db,
        platform=platform,
        product=product,
        source_payload=source_payload,
        status=status,
        confidence=payload.get("confidence", 100),
        created_by=str(payload.get("createdBy") or payload.get("created_by") or ""),
    )
    if platform == "provisor" and status == "mapped" and product is not None:
        item_id = payload.get("itemId") or payload.get("item_id")
        source_item = db.get(CompetitorPriceListItem, int(item_id)) if item_id not in (None, "") else None
        if source_item is not None and source_item.provisor_goods_id is not None:
            product.provisor_goods_id = int(source_item.provisor_goods_id)
    db.flush()
    touched = apply_mapping_to_matching_items(db=db, mapping=row, product=product, clear=status == "unmapped")
    db.commit()
    db.refresh(row)
    extra = db.get(ProductExtra, product.id) if product is not None else None
    data = mapping_to_dict(row, product, extra)
    data["touchedItems"] = touched
    return data


@app.post("/api/competitors/code-mappings/{mapping_id}/unmap")
def unmap_competitor_code_mapping(mapping_id: int, db: Session = Depends(get_db)):
    row = db.get(CompetitorCodeMapping, mapping_id)
    if row is None:
        raise HTTPException(status_code=404, detail="mapping not found")
    try:
        platform_from_value(row.platform)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    mapped_product = db.get(Product, row.our_product_id) if row.our_product_id else None
    if (
        row.platform == "provisor"
        and mapped_product is not None
        and str(mapped_product.provisor_goods_id or "") == str(row.source_external_key or "")
    ):
        mapped_product.provisor_goods_id = None
    row.status = "unmapped"
    row.our_product_id = None
    row.our_sku = ""
    row.confidence = None
    row.approved_at = None
    row.updated_at = now_kz_naive()
    touched = apply_mapping_to_matching_items(db=db, mapping=row, product=None, clear=True)
    db.commit()
    db.refresh(row)
    data = mapping_to_dict(row)
    data["touchedItems"] = touched
    return data


@app.post("/api/competitors/code-mappings/{mapping_id}/reject")
def reject_competitor_code_mapping(mapping_id: int, db: Session = Depends(get_db)):
    row = db.get(CompetitorCodeMapping, mapping_id)
    if row is None:
        raise HTTPException(status_code=404, detail="mapping not found")
    try:
        platform_from_value(row.platform)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    row.status = "rejected"
    row.our_product_id = None
    row.our_sku = ""
    row.confidence = None
    row.approved_at = None
    row.updated_at = now_kz_naive()
    touched = apply_mapping_to_matching_items(db=db, mapping=row, product=None)
    db.commit()
    db.refresh(row)
    data = mapping_to_dict(row)
    data["touchedItems"] = touched
    return data


@app.get("/api/pricing-rules/markup-templates")
def get_markup_templates(db: Session = Depends(get_db)):
    return list_templates(db=db, kind="markup")


@app.post("/api/pricing-rules/markup-templates")
def create_markup_template(payload: dict = Body(...), db: Session = Depends(get_db), current_user: AppUser = Depends(require_write_access)):
    try:
        return template_to_dict(upsert_template(db=db, kind="markup", payload=payload), "markup")
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.patch("/api/pricing-rules/markup-templates/{template_id}")
def patch_markup_template(template_id: int, payload: dict = Body(...), db: Session = Depends(get_db), current_user: AppUser = Depends(require_write_access)):
    try:
        return template_to_dict(upsert_template(db=db, kind="markup", template_id=template_id, payload=payload), "markup")
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.post("/api/pricing-rules/markup-templates/{template_id}/copy")
def copy_markup_template_endpoint(template_id: int, db: Session = Depends(get_db), current_user: AppUser = Depends(require_write_access)):
    try:
        return template_to_dict(copy_template(db=db, kind="markup", template_id=template_id), "markup")
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))


@app.get("/api/pricing-rules/bend-templates")
def get_bend_templates(db: Session = Depends(get_db)):
    return list_templates(db=db, kind="bend")


@app.post("/api/pricing-rules/bend-templates")
def create_bend_template(payload: dict = Body(...), db: Session = Depends(get_db), current_user: AppUser = Depends(require_write_access)):
    try:
        return template_to_dict(upsert_template(db=db, kind="bend", payload=payload), "bend")
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.patch("/api/pricing-rules/bend-templates/{template_id}")
def patch_bend_template(template_id: int, payload: dict = Body(...), db: Session = Depends(get_db), current_user: AppUser = Depends(require_write_access)):
    try:
        return template_to_dict(upsert_template(db=db, kind="bend", template_id=template_id, payload=payload), "bend")
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.post("/api/pricing-rules/bend-templates/{template_id}/copy")
def copy_bend_template_endpoint(template_id: int, db: Session = Depends(get_db), current_user: AppUser = Depends(require_write_access)):
    try:
        return template_to_dict(copy_template(db=db, kind="bend", template_id=template_id), "bend")
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))


@app.get("/api/pricing-rules/no-competitor-templates")
def get_no_competitor_templates(db: Session = Depends(get_db)):
    return list_templates(db=db, kind="no_competitor")


@app.post("/api/pricing-rules/no-competitor-templates")
def create_no_competitor_template(payload: dict = Body(...), db: Session = Depends(get_db), current_user: AppUser = Depends(require_write_access)):
    try:
        return template_to_dict(upsert_template(db=db, kind="no_competitor", payload=payload), "no_competitor")
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.patch("/api/pricing-rules/no-competitor-templates/{template_id}")
def patch_no_competitor_template(template_id: int, payload: dict = Body(...), db: Session = Depends(get_db), current_user: AppUser = Depends(require_write_access)):
    try:
        return template_to_dict(upsert_template(db=db, kind="no_competitor", template_id=template_id, payload=payload), "no_competitor")
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.post("/api/pricing-rules/no-competitor-templates/{template_id}/copy")
def copy_no_competitor_template_endpoint(template_id: int, db: Session = Depends(get_db), current_user: AppUser = Depends(require_write_access)):
    try:
        return template_to_dict(copy_template(db=db, kind="no_competitor", template_id=template_id), "no_competitor")
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))


@app.get("/api/pricing-rules/rounding-rules")
def get_rounding_rules(db: Session = Depends(get_db)):
    return list_rounding_rules(db=db)


@app.post("/api/pricing-rules/rounding-rules")
def create_rounding_rule(payload: dict = Body(...), db: Session = Depends(get_db), current_user: AppUser = Depends(require_write_access)):
    try:
        return rounding_to_dict(upsert_rounding_rule(db=db, payload=payload))
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.patch("/api/pricing-rules/rounding-rules/{rule_id}")
def patch_rounding_rule(rule_id: int, payload: dict = Body(...), db: Session = Depends(get_db), current_user: AppUser = Depends(require_write_access)):
    try:
        return rounding_to_dict(upsert_rounding_rule(db=db, rule_id=rule_id, payload=payload))
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.get("/api/pricing-rules")
def get_pricing_rules(db: Session = Depends(get_db)):
    return list_pricing_rules(db=db)


@app.post("/api/pricing-rules")
def create_pricing_rule(payload: dict = Body(...), db: Session = Depends(get_db), current_user: AppUser = Depends(require_write_access)):
    try:
        return pricing_rule_to_dict(upsert_pricing_rule(db=db, payload=payload))
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.get("/api/pricing-rules/{rule_id}")
def get_pricing_rule_endpoint(rule_id: int, db: Session = Depends(get_db)):
    try:
        return pricing_rule_to_dict(get_pricing_rule(db=db, rule_id=rule_id))
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))


@app.patch("/api/pricing-rules/{rule_id}")
def patch_pricing_rule(rule_id: int, payload: dict = Body(...), db: Session = Depends(get_db), current_user: AppUser = Depends(require_write_access)):
    try:
        return pricing_rule_to_dict(upsert_pricing_rule(db=db, rule_id=rule_id, payload=payload))
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.delete("/api/pricing-rules/{rule_id}")
def delete_pricing_rule_endpoint(rule_id: int, db: Session = Depends(get_db), current_user: AppUser = Depends(require_write_access)):
    try:
        delete_pricing_rule(db=db, rule_id=rule_id)
        return {"status": "ok"}
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))


@app.post("/api/pricing-rules/{rule_id}/copy")
def copy_pricing_rule_endpoint(rule_id: int, db: Session = Depends(get_db), current_user: AppUser = Depends(require_write_access)):
    try:
        return pricing_rule_to_dict(copy_pricing_rule(db=db, rule_id=rule_id))
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))


@app.post("/api/price-formats/{format_code}/pricing-rule")
def apply_pricing_rule_endpoint(
    format_code: str,
    payload: dict = Body(...),
    db: Session = Depends(get_db),
    current_user: AppUser = Depends(require_write_access),
):
    try:
        pf = db.execute(select(PriceFormat).where(PriceFormat.code == format_code)).scalars().first()
        if pf is not None:
            _ensure_price_format_access(pf, current_user)
        return apply_pricing_rule_to_format(db=db, format_code=format_code, rule_id=int(payload.get("pricingRuleId")))
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.get("/api/references/types")
def get_reference_types():
    return REFERENCE_TYPES


@app.get("/api/references/branches")
def get_reference_branches(current_user: AppUser = Depends(get_current_user)):
    if can_see_all_branches(current_user):
        return BRANCHES
    ids = assigned_branch_ids(current_user)
    return [row for row in BRANCHES if str(row["id"]) in ids or user_can_access_branch(current_user, str(row["id"]), str(row["name"]))]


@app.get("/api/references/status")
def get_reference_status(db: Session = Depends(get_db), current_user: AppUser = Depends(get_current_user)):
    rows = list_reference_statuses(db=db)
    return [
        row for row in rows
        if user_can_access_branch(current_user, str(row.get("branchId") or ""), str(row.get("branchName") or ""))
    ]


@app.get("/api/references/readiness-matrix")
def get_reference_readiness_matrix(db: Session = Depends(get_db), current_user: AppUser = Depends(get_current_user)):
    matrix = reference_readiness_matrix(db=db)
    matrix["rows"] = [
        row for row in matrix["rows"]
        if user_can_access_branch(current_user, str(row.get("branchId") or ""), str(row.get("branchName") or ""))
    ]
    return matrix


@app.get("/api/references/imports")
def get_reference_imports(
    limit: int = Query(100, ge=1, le=500),
    db: Session = Depends(get_db),
    current_user: AppUser = Depends(get_current_user),
):
    rows = list_reference_imports(db=db, limit=limit)
    if can_see_all_branches(current_user):
        return rows
    allowed = assigned_branch_ids(current_user)
    out = []
    for row in rows:
        try:
            branch_ids = set(json.loads(row.get("branchIds") or "[]"))
        except Exception:
            branch_ids = set()
        if branch_ids & allowed:
            out.append(row)
    return out


@app.get("/api/references/template")
def download_reference_template(
    data_type: str = Query(..., min_length=1),
    current_user: AppUser = Depends(get_current_user),
):
    try:
        content = build_reference_template(data_type)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    filename = quote(reference_template_filename(data_type))
    return StreamingResponse(
        io.BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"},
    )


@app.post("/api/references/import")
async def post_reference_import(
    data_type: str = Query(..., min_length=1),
    branch_ids: str = Query(..., min_length=1),
    user_name: str = Query("UI"),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: AppUser = Depends(require_write_access),
):
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="empty file")
    branches = [x.strip() for x in branch_ids.split(",") if x.strip()]
    for branch_id in branches:
        if not user_can_access_branch(current_user, branch_id, ""):
            raise HTTPException(status_code=403, detail="branch is not assigned to current user")
    try:
        if data_type in RATING_DATA_TYPES:
            return import_top_rating_excel(
                db=db,
                data_type=data_type,
                branch_ids=branches,
                content=content,
                filename=file.filename or "reference.xlsx",
                user_name=current_user.username or user_name,
            )
        row = import_reference_excel(
            db=db,
            data_type=data_type,
            branch_ids=branches,
            content=content,
            filename=file.filename or "reference.xlsx",
            user_name=current_user.username or user_name,
        )
        return import_job_to_dict(row)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.post("/api/references/import/batch")
async def post_reference_batch_import(
    data_types: str = Form(..., min_length=1),
    branch_ids: str = Form(..., min_length=1),
    user_name: str = Form("UI"),
    source_type: str = Form("excel"),
    files: list[UploadFile] = File(...),
    db: Session = Depends(get_db),
    current_user: AppUser = Depends(require_write_access),
):
    selected_types = [x.strip() for x in data_types.split(",") if x.strip()]
    branches = [x.strip() for x in branch_ids.split(",") if x.strip()]
    for branch_id in branches:
        if not user_can_access_branch(current_user, branch_id, ""):
            raise HTTPException(status_code=403, detail="branch is not assigned to current user")
    if not selected_types:
        raise HTTPException(status_code=400, detail="data_types is required")
    if len(selected_types) != len(files):
        raise HTTPException(status_code=400, detail="files count must match data_types count")

    payloads: list[ReferenceFilePayload] = []
    for data_type, file in zip(selected_types, files):
        content = await file.read()
        if not content:
            raise HTTPException(status_code=400, detail=f"empty file for {data_type}")
        payloads.append(
            ReferenceFilePayload(
                data_type=data_type,
                filename=file.filename or f"{data_type}.xlsx",
                content=content,
            )
        )

    try:
        source = make_reference_source(source_type, payloads)
        return import_reference_batch(db=db, source=source, selected_branch_ids=branches, user_name=current_user.username or user_name)
    except NotImplementedError as e:
        raise HTTPException(status_code=501, detail=str(e))
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.get("/api/price-source-accounts")
def get_price_source_accounts(db: Session = Depends(get_db)):
    rows = (
        db.execute(select(PriceSourceAccount).order_by(PriceSourceAccount.source_type.asc(), PriceSourceAccount.login.asc()))
        .scalars()
        .all()
    )
    return [account_to_dict(row) for row in rows]


@app.get("/api/provisor/accounts")
def get_provisor_accounts(db: Session = Depends(get_db)):
    price_list_counts = dict(
        db.execute(
            select(CompetitorPriceList.account_id, func.count(func.distinct(CompetitorPriceList.source_key)))
            .where(CompetitorPriceList.source_type == "provisor")
            .where(CompetitorPriceList.account_id != "")
            .group_by(CompetitorPriceList.account_id)
        ).all()
    )
    rows = (
        db.execute(
            select(PriceSourceAccount)
            .where(PriceSourceAccount.source_type == "provisor")
            .where(PriceSourceAccount.is_active.is_(True))
            .order_by(PriceSourceAccount.login.asc())
        )
        .scalars()
        .all()
    )
    return [
        {
            "id": row.id,
            "login": row.login,
            "status": row.status or "not_checked",
            "price_list_count": int(price_list_counts.get(str(row.id)) if price_list_counts.get(str(row.id)) is not None else row.price_lists_count or 0),
            "last_success_at": row.last_success_at.isoformat() if row.last_success_at else None,
        }
        for row in rows
    ]


@app.post("/api/price-source-accounts")
def create_price_source_account(payload: dict = Body(...), db: Session = Depends(get_db)):
    try:
        row = upsert_account(
            db=db,
            source_type=str(payload.get("sourceType") or payload.get("source_type") or "").strip(),
            login=str(payload.get("login") or "").strip(),
            password=str(payload.get("password") or ""),
            config=payload.get("config") if isinstance(payload.get("config"), dict) else {},
        )
        return account_to_dict(row)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.post("/api/price-source-accounts/{account_id}/test")
async def test_price_source_account(
    account_id: int,
    format_code: str | None = Query(None),
    db: Session = Depends(get_db),
):
    operation = "price_source_account_test"
    started_at = time.perf_counter()
    _timing(operation, "start", started_at)
    try:
        logger.info("POST /api/price-source-accounts/%s/test format_code=%s", account_id, format_code)
        row = await test_account_connection(db=db, account_id=account_id, price_format_code=format_code)
        _timing(operation, "finish", started_at)
        _timing(operation, "total_ms", started_at)
        payload = account_to_dict(row)
        logger.info(
            "Price source account test response: account_id=%s sourceType=%s status=%s priceListsCount=%s statusMessage=%s",
            payload.get("id"),
            payload.get("sourceType"),
            payload.get("status"),
            payload.get("priceListsCount"),
            payload.get("statusMessage"),
        )
        return payload
    except ValueError as e:
        logger.exception("[TIMING] operation=%s step=exception elapsed_ms=%s", operation, round((time.perf_counter() - started_at) * 1000, 2))
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        logger.exception("[TIMING] operation=%s step=exception elapsed_ms=%s", operation, round((time.perf_counter() - started_at) * 1000, 2))
        logger.exception("Price source account test endpoint failed: account_id=%s format_code=%s", account_id, format_code)
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/api/price-source-accounts/{account_id}")
def delete_price_source_account(account_id: int, db: Session = Depends(get_db)):
    row = db.get(PriceSourceAccount, account_id)
    if row is None:
        raise HTTPException(status_code=404, detail="account not found")
    row.is_active = False
    row.updated_at = now_kz_naive()
    db.commit()
    return {"status": "ok"}


@app.post("/api/price-formats/{format_code}/competitor-price-lists/selection")
async def post_competitor_price_list_selection(format_code: str, payload: dict = Body(...), db: Session = Depends(get_db)):
    operation = f"competitor_price_list_selection:{format_code}"
    selection_started_at = time.perf_counter()
    _timing(operation, "start_selection/rebuild", selection_started_at)
    ids_raw = payload.get("selectedIds") or []
    if not isinstance(ids_raw, list):
        raise HTTPException(status_code=400, detail="selectedIds must be list[int]")
    selected_ids: list[int] = []
    for x in ids_raw:
        try:
            selected_ids.append(int(x))
        except Exception:
            raise HTTPException(status_code=400, detail="selectedIds must be list[int]")

    coefficients: dict[int, float] = {}
    coeff_raw = payload.get("coefficients")
    if isinstance(coeff_raw, dict):
        for key, value in coeff_raw.items():
            try:
                coefficients[int(key)] = float(value)
            except Exception:
                continue

    existing = get_active_job(db=db, job_type="generate_price", format_code=format_code)
    if existing is not None:
        return {"job_id": existing.id, "status": existing.status, "message": "Такая задача уже выполняется"}
    refresh_job = _get_any_active_refresh_job(db=db, format_code=format_code)
    if refresh_job is not None:
        raise HTTPException(status_code=409, detail="данные ещё не готовы: дождитесь завершения обновления прайсов")

    try:
        price_format_id = save_selected_competitor_price_lists_only(
            db=db,
            price_format_code=format_code,
            selected_ids=selected_ids,
            coefficients=coefficients,
        )
        job = create_job(
            db=db,
            job_type="generate_price",
            format_code=format_code,
            price_format_id=price_format_id,
            message="Создана задача формирования прайса",
        )
        schedule_job(
            job.id,
            lambda job_db, job_row: _run_generate_price_job(job_db, job_row, price_format_id=price_format_id, payload=payload or {}),
        )
        _timing(operation, "finish", selection_started_at)
        _timing(operation, "total_ms", selection_started_at)
        return {"job_id": job.id, "status": job.status}
    except OperationalError as e:
        db.rollback()
        logger.exception("[TIMING] operation=%s step=exception elapsed_ms=%s", operation, round((time.perf_counter() - selection_started_at) * 1000, 2))
        raise HTTPException(
            status_code=503 if _is_sqlite_locked(e) else 500,
            detail="База занята, повторите операцию через несколько секунд" if _is_sqlite_locked(e) else str(e),
        )
    except Exception as e:
        db.rollback()
        logger.exception("[TIMING] operation=%s step=exception elapsed_ms=%s", operation, round((time.perf_counter() - selection_started_at) * 1000, 2))
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/competitor-price-lists/{price_list_id}")
def get_competitor_price_list(price_list_id: int, db: Session = Depends(get_db)):
    try:
        return get_competitor_price_list_items(db=db, price_list_id=price_list_id)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))


@app.get("/api/competitor-price-lists/{price_list_id}/export.{fmt}")
def export_competitor_price_list_endpoint(price_list_id: int, fmt: str, db: Session = Depends(get_db)):
    if fmt not in {"csv", "xlsx"}:
        raise HTTPException(status_code=400, detail="fmt must be csv or xlsx")

    try:
        filename, content, media_type = export_competitor_price_list(
            db=db,
            price_list_id=price_list_id,
            fmt=fmt
        )
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))

    safe_filename = quote(filename)

    return StreamingResponse(
        iter([content]),
        media_type=media_type,
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{safe_filename}"
        },
    )


@app.post("/api/price-formats/{format_code}/competitor-price-lists/upload-excel")
async def upload_competitor_price_list_excel(
    format_code: str,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="empty file")
    try:
        row = import_manual_price_list_excel(
            db=db,
            price_format_code=format_code,
            content=content,
            filename=file.filename or "manual.xlsx",
        )
        return {"id": row.id, "name": row.display_name}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


def _provisor_auto_refresh_payload(*, account_ids: list[str], filial_ids: set[str] | None, mode: str) -> dict:
    payload = {
        "source": "provisor",
        "accountIds": [int(x) for x in account_ids if str(x).isdigit()],
        "forceRefresh": True,
        "runRebuildAfterRefresh": False,
        "maxParallelAccounts": settings.provisor_auto_refresh_max_parallel_accounts,
        "maxParallelPlk": settings.provisor_auto_refresh_max_parallel_plk,
        "keepLastSuccess": settings.provisor_auto_refresh_keep_last_success,
    }
    if mode == "selected":
        payload["provisorFilialIds"] = [int(x) for x in sorted(filial_ids or set()) if str(x).isdigit()]
    return payload


def _heartbeat_provisor_refresh_job(job_id: int, owner_token: str, message: str | None = None) -> bool:
    with SessionLocal() as heartbeat_db:
        current = heartbeat_db.get(RefreshJob, job_id)
        if current is None:
            return False
        return refresh_job_heartbeat(heartbeat_db, current, message=message, owner_token=owner_token)


async def _run_provisor_refresh_job(job_id: int, *, mode: str, requested_by: str, owner_token: str | None = None) -> None:
    db = SessionLocal()
    heartbeat_task = None
    try:
        job = db.get(RefreshJob, job_id)
        if job is None:
            logger.error("[PROVISOR_AUTO_REFRESH] job not found: %s", job_id)
            return
        owner_token = owner_token or refresh_job_owner_token(job)
        if not owner_token:
            logger.error("[PROVISOR_AUTO_REFRESH] job has no owner token: %s", job_id)
            return
        mode = normalize_refresh_mode(mode)
        targets = selected_refresh_targets(db) if mode == "selected" else all_refresh_targets(db)
        if not targets:
            message = (
                "No selected Provisor PLK configured for auto-refresh."
                if mode == "selected"
                else "No active Provisor accounts configured for auto-refresh."
            )
            finish_refresh_job(db, job, status="failed", message=message, error=message, owner_token=owner_token, release_refresh=True)
            logger.warning("[PROVISOR_AUTO_REFRESH] job failed before start: %s", message)
            return
        total_accounts, total_plk = refresh_target_counts(targets, db, mode=mode)
        if mode == "selected" and total_plk <= 0:
            message = "No selected Provisor PLK configured for auto-refresh."
            finish_refresh_job(db, job, status="failed", message=message, error=message, owner_token=owner_token, release_refresh=True)
            logger.warning("[PROVISOR_AUTO_REFRESH] job failed before start: %s", message)
            return
        started = start_refresh_job(
            db,
            job,
            total_accounts=total_accounts,
            total_plk=total_plk,
            metadata={"requested_by": requested_by, "targets": {fmt: {acc: sorted(ids) for acc, ids in by_acc.items()} for fmt, by_acc in targets.items()}},
            owner_token=owner_token,
        )
        if not started:
            logger.warning("[PROVISOR_AUTO_REFRESH] job start skipped because token/status no longer match: job_id=%s", job_id)
            return

        async def beat() -> None:
            while True:
                await asyncio.sleep(30)
                if not _heartbeat_provisor_refresh_job(job_id, owner_token, "Обновляем ПЛК Provisor..."):
                    return

        heartbeat_task = asyncio.create_task(beat())
        logger.info("[PROVISOR_AUTO_REFRESH] job started id=%s mode=%s requested_by=%s", job_id, mode, requested_by)
        aggregate = {"results": [], "errors": [], "accounts": set(), "processed_plk": 0, "success": 0, "failed": 0, "skipped": 0}
        for format_code, by_account in targets.items():
            for account_id, filial_ids in by_account.items():
                db.expire_all()
                job = db.get(RefreshJob, job_id)
                if job is None:
                    return
                if not _heartbeat_provisor_refresh_job(job_id, owner_token, f"Обновляем ПЛК Provisor для аккаунта {account_id}..."):
                    logger.warning("[PROVISOR_AUTO_REFRESH] heartbeat failed before account refresh; stopping job_id=%s", job_id)
                    return
                payload = _provisor_auto_refresh_payload(account_ids=[account_id], filial_ids=filial_ids, mode=mode)
                logger.info("[PROVISOR_AUTO_REFRESH] account started job_id=%s format=%s account_id=%s plk=%s", job_id, format_code, account_id, sorted(filial_ids) if mode == "selected" else "all")
                result = await _run_refresh_price_lists_logic(format_code=format_code, payload=payload, db=db, job=None)
                progress = result.get("progress") or {}
                aggregate["results"].append({"format_code": format_code, "account_id": account_id})
                aggregate["errors"].extend(result.get("errors") or [])
                aggregate["accounts"].add(account_id)
                aggregate["processed_plk"] += int(progress.get("processed") or 0)
                aggregate["success"] += int(progress.get("success") or 0)
                aggregate["failed"] += int(progress.get("errors") or 0)
                aggregate["skipped"] += int(progress.get("skipped") or 0)
                db.expire_all()
                job = db.get(RefreshJob, job_id)
                if job is not None:
                    if not update_refresh_job_progress(db, job, result, owner_token=owner_token):
                        logger.warning("[PROVISOR_AUTO_REFRESH] progress update skipped because token/status no longer match: job_id=%s", job_id)
                        return
                    job.processed_accounts = min(int(job.total_accounts or 0), len(aggregate["accounts"]))
                    job.processed_plk = int(aggregate["processed_plk"])
                    job.success_count = int(aggregate["success"])
                    job.failed_count = int(aggregate["failed"])
                    job.skipped_count = int(aggregate["skipped"])
                    job.heartbeat_at = now_kz_naive()
                    db.commit()
                logger.info("[PROVISOR_AUTO_REFRESH] account finished job_id=%s account_id=%s success=%s failed=%s skipped=%s", job_id, account_id, int(progress.get("success") or 0), int(progress.get("errors") or 0), int(progress.get("skipped") or 0))
        db.expire_all()
        job = db.get(RefreshJob, job_id)
        if job is not None:
            failed = int(aggregate["failed"] or 0)
            status = "failed" if failed and not int(aggregate["success"] or 0) else "success"
            message = f"Обновление Provisor завершено: успешно={aggregate['success']}, ошибок={aggregate['failed']}, пропущено={aggregate['skipped']}." if status == "success" else "Обновление Provisor завершилось с ошибкой."
            finished = finish_refresh_job(
                db,
                job,
                status=status,
                message=message,
                error="; ".join(str(x) for x in aggregate["errors"][:5]),
                metadata={"result_count": len(aggregate["results"])},
                owner_token=owner_token,
                allowed_statuses={"running"},
                release_refresh=True,
            )
            if not finished:
                logger.warning("[PROVISOR_AUTO_REFRESH] final status skipped because token/status no longer match: job_id=%s", job_id)
        logger.info("[PROVISOR_AUTO_REFRESH] job completed id=%s", job_id)
    except Exception as e:
        db.rollback()
        logger.exception("[PROVISOR_AUTO_REFRESH] job failed id=%s", job_id)
        job = db.get(RefreshJob, job_id)
        if job is not None:
            finish_refresh_job(
                db,
                job,
                status="failed",
                message="Обновление Provisor завершилось с ошибкой.",
                error=str(e),
                owner_token=owner_token,
                allowed_statuses={"pending", "running"},
                release_refresh=True,
            )
    finally:
        if heartbeat_task is not None:
            heartbeat_task.cancel()
        db.close()


@app.get("/api/price-sources/refresh/status")
def get_price_sources_refresh_status(db: Session = Depends(get_db)):
    return refresh_job_to_status(latest_refresh_job(db))


@app.post("/api/price-sources/refresh/provisor/auto/run-now", status_code=202)
async def run_provisor_auto_refresh_now(payload: dict = Body(default={}), db: Session = Depends(get_db)):
    mode = normalize_refresh_mode(str(payload.get("mode") or settings.provisor_auto_refresh_mode))
    blocker = active_or_stale_refresh_job(db)
    if blocker is not None:
        message = "Обновление Provisor уже выполняется." if blocker.status != "stale" else "Обновление Provisor зависло; завершите его перед запуском новой задачи."
        create_skipped_job(db, mode=mode, requested_by="manual", message=message)
        raise HTTPException(status_code=409, detail=message)
    job, blocker, owner_token = try_create_refresh_job(db, mode=mode, requested_by="manual")
    if job is None:
        message = "Обновление Provisor уже выполняется."
        if blocker is not None and blocker.status == "stale":
            message = "Обновление Provisor зависло; завершите его перед запуском новой задачи."
        create_skipped_job(db, mode=mode, requested_by="manual", message=message)
        raise HTTPException(status_code=409, detail=message)
    assert owner_token is not None
    asyncio.create_task(_run_provisor_refresh_job(int(job.id), mode=job.mode, requested_by="manual", owner_token=owner_token))
    return {"job_id": job.id, **refresh_job_to_status(job)}


@app.post("/api/price-sources/refresh/provisor/auto/resolve-stale")
def resolve_stale_provisor_auto_refresh(db: Session = Depends(get_db)):
    blocker = active_or_stale_refresh_job(db)
    if blocker is None or blocker.status != "stale":
        raise HTTPException(status_code=404, detail="Нет зависшей задачи обновления Provisor для завершения.")
    finish_refresh_job(
        db,
        blocker,
        status="failed",
        message="Зависшая задача обновления Provisor помечена как ошибочная оператором.",
        error=blocker.error_message or "stale refresh resolved",
        owner_token=refresh_job_owner_token(blocker),
        allowed_statuses={"stale"},
        release_refresh=True,
    )
    logger.warning("[PROVISOR_AUTO_REFRESH] stale job marked failed id=%s", blocker.id)
    return refresh_job_to_status(blocker)


@app.get("/api/emit/refresh/status")
def get_emit_refresh_status(db: Session = Depends(get_db)):
    return emit_job_to_dict(latest_emit_job(db))


@app.post("/api/emit/refresh/run-now", status_code=202)
async def run_emit_refresh_now(payload: dict = Body(default={}), db: Session = Depends(get_db)):
    worker = _emit_worker_instance()
    mode = str(payload.get("mode") or "selected").strip().lower()
    if mode not in {"selected", "all"}:
        raise HTTPException(status_code=400, detail="mode must be selected or all")
    filial_ids = [int(x) for x in (payload.get("filial_ids") or payload.get("filialIds") or []) if str(x).strip().lstrip("-").isdigit()]
    target_ids = configured_filial_ids_for_mode(worker.config, mode=mode, filial_ids=filial_ids)
    if not target_ids:
        raise HTTPException(status_code=400, detail="No Emit filial IDs requested")
    if active_or_stale_refresh_job(db) is not None:
        raise HTTPException(status_code=409, detail="Обычное обновление Provisor выполняется или зависло; дождитесь завершения перед запуском обновления Emit.")
    price_format_code = str(payload.get("format_code") or payload.get("formatCode") or "").strip()
    if price_format_code:
        pf = db.execute(select(PriceFormat).where(PriceFormat.code == price_format_code)).scalars().first()
        if pf is None:
            raise HTTPException(status_code=400, detail=f"Ценовой формат не найден: {price_format_code}")
    job, blocker, owner_token = worker.create_job(
        mode=mode,
        filial_ids=target_ids,
        requested_by="manual",
        price_format_code=price_format_code,
    )
    if job is None:
        raise HTTPException(status_code=409, detail={"message": "Emit refresh is already running.", "job": emit_job_to_dict(blocker)})
    asyncio.create_task(worker.run_job(int(job.id), owner_token=owner_token))
    return {"job_id": job.id, **emit_job_to_dict(job)}


@app.get("/api/emit/refresh/jobs")
def get_emit_refresh_jobs(limit: int = Query(default=50, ge=1, le=200), db: Session = Depends(get_db)):
    return {"items": list_emit_jobs(db, limit=limit)}


@app.get("/api/emit/refresh/jobs/{job_id}")
def get_emit_refresh_job(job_id: int, db: Session = Depends(get_db)):
    job = db.get(RefreshJob, job_id)
    if job is None or job.source_type != "emit":
        raise HTTPException(status_code=404, detail="Emit refresh job not found")
    return {"id": job.id, **emit_job_to_dict(job)}


async def _start_provisor_refresh_background(*, mode: str, requested_by: str) -> None:
    db = SessionLocal()
    try:
        job, blocker, owner_token = try_create_refresh_job(db, mode=mode, requested_by=requested_by)
        if blocker is not None:
            message = "Обновление Provisor уже выполняется." if blocker.status != "stale" else "Обновление Provisor зависло; завершите его перед запуском новой задачи."
            create_skipped_job(db, mode=mode, requested_by=requested_by, message=message)
            logger.warning("[PROVISOR_AUTO_REFRESH] skipped new job: blocker_id=%s blocker_status=%s", blocker.id, blocker.status)
            return
        if job is None:
            create_skipped_job(db, mode=mode, requested_by=requested_by, message="Обновление Provisor уже выполняется.")
            logger.warning("[PROVISOR_AUTO_REFRESH] skipped new job: refresh lock is owned by another process")
            return
        assert owner_token is not None
        asyncio.create_task(_run_provisor_refresh_job(int(job.id), mode=job.mode, requested_by=requested_by, owner_token=owner_token))
    finally:
        db.close()


@app.post("/api/price-formats/{format_code}/competitor-price-lists/refresh")
async def refresh_competitor_price_lists(format_code: str, payload: dict = Body(default={}), db: Session = Depends(get_db)):
    refresh_source = _refresh_source_from_payload(payload or {})
    refresh_job_type = _refresh_job_type(refresh_source)
    job_key = _refresh_job_key(format_code, refresh_source)
    logger.info("[SOURCE_REFRESH_JOB] job_key=%s source=%s format_code=%s", job_key, refresh_source, format_code)
    existing = get_active_job(db=db, job_type=refresh_job_type, format_code=format_code)
    if existing is not None:
        return {"job_id": existing.id, "status": existing.status, "source": refresh_source, "message": "Такая задача уже выполняется"}
    pf = db.execute(select(PriceFormat).where(PriceFormat.code == format_code)).scalars().first()
    job = create_job(
        db=db,
        job_type=refresh_job_type,
        format_code=format_code,
        price_format_id=int(pf.id) if pf else None,
        message="Создана задача обновления прайс-листов",
    )
    schedule_job(job.id, lambda job_db, job_row: _run_refresh_price_lists_job(job_db, job_row, format_code=format_code, payload=payload or {}))
    return {"job_id": job.id, "status": job.status, "source": refresh_source}


async def _run_refresh_price_lists_logic(format_code: str, payload: dict, db: Session, job: Job | None = None):
    refresh_started_at = time.perf_counter()
    force_refresh = bool(payload.get("forceRefresh") or payload.get("force_refresh"))
    refresh_source = _refresh_source_from_payload(payload)
    target_account_ids = _payload_id_set(payload, "accountIds", "account_ids", "accountId", "account_id")
    target_provisor_filial_ids = _payload_id_set(
        payload,
        "provisorFilialIds",
        "provisor_filial_ids",
        "filialIds",
        "filial_ids",
    )
    operation = f"refresh_competitor_price_lists:{format_code}:{refresh_source}"
    run_rebuild_raw = payload.get("runRebuildAfterRefresh", payload.get("run_rebuild_after_refresh"))
    run_rebuild_after_refresh = run_rebuild_raw is True or (
        isinstance(run_rebuild_raw, str) and run_rebuild_raw.strip().lower() in {"1", "true", "yes", "on"}
    )
    _timing(operation, "start_refresh", refresh_started_at)
    logger.info("[REFRESH] JOB START format_code=%s source=%s", format_code, refresh_source)
    print(f"[REFRESH] JOB START format_code={format_code} source={refresh_source}", flush=True)
    refreshed: list[dict] = []
    errors: list[str] = []
    account_statuses: list[dict] = []
    progress = {
        "total": 0,
        "processed": 0,
        "success": 0,
        "skipped": 0,
        "errors": 0,
        "skipped_heavy": 0,
        "skipped_heavy_filials": [],
        "skipped_auth_error_count": 0,
        "success_with_items": 0,
        "success_zero_items": 0,
        "timeout": 0,
        "heavy_excluded": 0,
        "fetch_duration_seconds": [],
        "bySource": {},
        "accounts_requested": [],
        "accounts_processed": [],
        "accounts_skipped": [],
    }

    accounts = (
        db.execute(select(PriceSourceAccount).where(PriceSourceAccount.is_active.is_(True)))
        .scalars()
        .all()
    )
    if refresh_source != "all":
        accounts = [account for account in accounts if account.source_type == refresh_source]
    requested_account_ids = sorted(int(x) for x in target_account_ids) if target_account_ids else []
    eligible_account_ids = {int(account.id) for account in accounts}
    skipped_requested_account_ids = [account_id for account_id in requested_account_ids if account_id not in eligible_account_ids]
    if target_account_ids:
        accounts = [account for account in accounts if str(int(account.id)) in target_account_ids]
    processed_account_ids = [int(account.id) for account in accounts]
    progress["accounts_requested"] = requested_account_ids
    progress["accounts_processed"] = processed_account_ids
    progress["accounts_skipped"] = skipped_requested_account_ids
    pf_for_refresh = db.execute(select(PriceFormat).where(PriceFormat.code == format_code)).scalars().first()
    _timing(operation, "load_accounts", refresh_started_at)
    source_started_at = time.perf_counter()
    source_counts = {
        source: sum(1 for account in accounts if account.source_type == source)
        for source in sorted({str(account.source_type or "") for account in accounts})
        if source
    }
    if refresh_source != "all":
        source_counts.setdefault(refresh_source, len(accounts))
    for source, accounts_count in source_counts.items():
        logger.info(
            "[REFRESH_SOURCE_START] source=%s format_code=%s accounts_count=%s",
            source,
            format_code,
            accounts_count,
        )
        print(
            f"[REFRESH_SOURCE_START] source={source} format_code={format_code} accounts_count={accounts_count}",
            flush=True,
        )
    logger.info("[REFRESH] ACCOUNTS total=%s source=%s", len(accounts), refresh_source)
    print(f"[REFRESH] ACCOUNTS total={len(accounts)} source={refresh_source}", flush=True)
    if job is not None:
        update_job(db, job, status="running", progress=10, message="Загружаем аккаунты", log_level="info")

    # Backward-compatible fallback: old env-based Provisor integration still works
    # even before the admin creates source accounts in settings.
    if not accounts and settings.provisor_login and settings.provisor_password:
        try:
            fallback = upsert_account(
                db=db,
                source_type="provisor",
                login=settings.provisor_login,
                password=settings.provisor_password,
                config={"filialIds": payload.get("provisorFilialIds") or []},
            )
            accounts = [fallback]
            if refresh_source not in {"all", "provisor"}:
                accounts = []
        except Exception as e:
            errors.append(f"provisor/env: {e}")

    # External price sources are sensitive to parallel long-running item loads.
    # Keep timeout per active price list, but avoid overloading shared sessions/APIs.
    provisor_account_parallel = _env_int("PROVISOR_AUTO_REFRESH_MAX_PARALLEL_ACCOUNTS", 2)
    provisor_plk_parallel = _env_int("PROVISOR_AUTO_REFRESH_MAX_PARALLEL_PLK", 1)
    if payload.get("maxParallelAccounts") is not None or payload.get("max_parallel_accounts") is not None:
        provisor_account_parallel = max(1, int(payload.get("maxParallelAccounts") or payload.get("max_parallel_accounts") or provisor_account_parallel))
    if payload.get("maxParallelPlk") is not None or payload.get("max_parallel_plk") is not None:
        provisor_plk_parallel = max(1, int(payload.get("maxParallelPlk") or payload.get("max_parallel_plk") or provisor_plk_parallel))
    source_limits = {"provisor": provisor_plk_parallel, "vidman": 1}
    account_sem = asyncio.Semaphore(provisor_account_parallel if refresh_source == "provisor" else 4)
    source_sems = {source: asyncio.Semaphore(limit) for source, limit in source_limits.items()}
    logger.info(
        "[REFRESH] accounts_parallel=%s fetch_lists_timeout=%ss",
        provisor_account_parallel if refresh_source == "provisor" else 4,
        PRICE_LISTS_FETCH_TIMEOUT_SECONDS,
    )
    print(
        f"[REFRESH] accounts_parallel={provisor_account_parallel if refresh_source == 'provisor' else 4} fetch_lists_timeout={PRICE_LISTS_FETCH_TIMEOUT_SECONDS}s",
        flush=True,
    )

    async def fetch_account(account: PriceSourceAccount) -> dict:
        status = {
            "source": refresh_source,
            "sourceType": account.source_type,
            "accountId": account.id,
            "login": account.login,
            "ok": False,
            "priceListsCount": 0,
            "success": 0,
            "skipped": 0,
            "errors": 0,
            "processed": 0,
            "total": 0,
            "message": "",
            "results": [],
            "skipped_timeout": 0,
            "skipped_unchanged": 0,
            "skipped_heavy_count": 0,
            "skipped_heavy_filials": [],
            "skipped_auth_error_count": 0,
            "success_with_items": 0,
            "success_zero_items": 0,
            "timeout": 0,
            "heavy_excluded": 0,
            "fetch_duration_seconds": [],
            "skipped_price_lists": [],
        }
        async with account_sem:
            account_operation = f"{operation}:account:{account.id}"
            account_started_at = time.perf_counter()
            shared_vidman_client = None
            try:
                if account.source_type == "vidman" and str(account.status or "").strip().lower() == "auth_error":
                    logger.info("[VIDMAN_AUTH_ERROR_SKIPPED] account_id=%s reason=auth_error_status", account.id)
                    status.update(
                        {
                            "source": "vidman",
                            "status": "auth_error",
                            "ok": False,
                            "priceListsCount": 0,
                            "priceLists": 0,
                            "items": 0,
                            "skipped": 1,
                            "skipped_auth_error_count": 1,
                            "errors": 0,
                            "message": "Skipped Vidman account because status=auth_error",
                            "statusMessage": "Skipped Vidman account because status=auth_error",
                            "results": [],
                        }
                    )
                    return status
                adapter = adapter_for_source(account.source_type)
                credentials = credentials_from_row(account)
                excluded_provisor_filial_ids = _provisor_excluded_filial_ids(credentials.config) if account.source_type == "provisor" else set()
                shared_vidman_client = (
                    adapter._client(credentials, timeout=PRICE_LIST_FETCH_TIMEOUT_SECONDS)
                    if isinstance(adapter, VidmanPriceService) and account.source_type == "vidman"
                    else None
                )
                shared_vidman_login_lock = asyncio.Lock()
                if job is not None:
                    update_job(db, job, status="running", progress=25, message=f"Авторизация источника {account.source_type}", log_level="info")
                fetch_lists_started_at = time.perf_counter()
                logger.info(
                    "[REFRESH] FETCH_LISTS START source=%s account_id=%s",
                    account.source_type,
                    account.id,
                )
                if account.source_type == "vidman":
                    logger.info("[VW_REFRESH_START] account_id=%s", account.id)
                print(
                    f"[REFRESH] FETCH_LISTS START source={account.source_type} account_id={account.id}",
                    flush=True,
                )
                fetch_lists_started_wall = local_iso(now_kz_naive())
                try:
                    async with source_sems.get(account.source_type, asyncio.Semaphore(2)):
                        if shared_vidman_client is not None and isinstance(adapter, VidmanPriceService):
                            price_lists = await asyncio.wait_for(
                                adapter.fetch_price_lists_with_client(credentials, shared_vidman_client),
                                timeout=PRICE_LISTS_FETCH_TIMEOUT_SECONDS,
                            )
                        else:
                            price_lists = await asyncio.wait_for(
                                adapter.fetch_price_lists(credentials),
                                timeout=PRICE_LISTS_FETCH_TIMEOUT_SECONDS,
                            )
                except (asyncio.TimeoutError, httpx.TimeoutException) as e:
                    elapsed_ms = round((time.perf_counter() - fetch_lists_started_at) * 1000, 2)
                    logger.warning(
                        "[TIMEOUT_DEBUG] price_id=%s price_name=%s source=%s started_at=%s elapsed=%s stage=%s exception=%r",
                        "",
                        "",
                        account.source_type,
                        fetch_lists_started_wall,
                        elapsed_ms,
                        "get_price_lists",
                        e,
                    )
                    raise
                _timing(account_operation, "fetch_price_lists", fetch_lists_started_at)
                fetch_lists_elapsed_ms = round((time.perf_counter() - fetch_lists_started_at) * 1000, 2)
                logger.info(
                    "[REFRESH] FETCH_LISTS DONE source=%s account_id=%s elapsed_ms=%s total=%s",
                    account.source_type,
                    account.id,
                    fetch_lists_elapsed_ms,
                    len(price_lists),
                )
                print(
                    f"[REFRESH] FETCH_LISTS DONE source={account.source_type} account_id={account.id} "
                    f"elapsed_ms={fetch_lists_elapsed_ms} total={len(price_lists)}",
                    flush=True,
                )
                if job is not None:
                    update_job(db, job, status="running", progress=40, message=f"Получаем список прайсов {account.source_type}", log_level="info")
                status["total"] = len(price_lists)
                logger.info(
                    "%s fetch_price_lists returned %s items for account_id=%s during refresh",
                    account.source_type.capitalize(),
                    len(price_lists),
                    account.id,
                )
                if account.source_type == "provisor":
                    if target_provisor_filial_ids:
                        price_lists = [
                            row
                            for row in price_lists
                            if str(getattr(row, "price_list_id", "") or "").strip() in target_provisor_filial_ids
                        ]
                        status["total"] = len(price_lists)
                    has_reference_filial = any(
                        str(getattr(row, "price_list_id", "") or "").strip() in {str(x) for x in PROVISOR_REFERENCE_FILIAL_IDS}
                        for row in price_lists
                    )
                    if not has_reference_filial:
                        logger.info(
                            "[PROVISOR_REFERENCE] account_id=%s reference_filial_ids=%s available=false mapped_products=0 updated_products=0",
                            account.id,
                            ",".join(str(x) for x in PROVISOR_REFERENCE_FILIAL_IDS),
                        )
                        logger.info(
                            "[PROVISOR_REFERENCE_SKIP] account_id=%s reason=reference_filial_not_available using_existing_product_provisor_goods_id=true",
                            account.id,
                        )
                local_price_list_state: dict[str, tuple[str, int, datetime | None]] = {}
                if pf_for_refresh is not None and price_lists:
                    source_keys = [f"{account.id}:{getattr(row, 'price_list_id', '')}" for row in price_lists]
                    existing_rows = (
                        db.execute(
                            select(CompetitorPriceList)
                            .where(CompetitorPriceList.price_format_id == pf_for_refresh.id)
                            .where(CompetitorPriceList.source_type == account.source_type)
                            .where(CompetitorPriceList.account_id == str(account.id))
                            .where(CompetitorPriceList.source_key.in_(source_keys))
                        )
                        .scalars()
                        .all()
                    )
                    existing_ids = [int(row.id) for row in existing_rows]
                    item_counts = (
                        dict(
                            db.execute(
                                select(CompetitorPriceListItem.price_list_id, func.count(CompetitorPriceListItem.id))
                                .where(CompetitorPriceListItem.price_list_id.in_(existing_ids))
                                .group_by(CompetitorPriceListItem.price_list_id)
                            ).all()
                        )
                        if existing_ids
                        else {}
                    )
                    for row in existing_rows:
                        local_price_list_state[str(row.source_key or "")] = (
                            str(row.source_updated_at or ""),
                            int(item_counts.get(row.id, 0)),
                            row.updated_at,
                        )
                if account.source_type == "vidman":
                    for row in price_lists:
                        price_id = str(getattr(row, "price_list_id", "") or "")
                        price_name = str(getattr(row, "price_list_name", "") or getattr(row, "distributor_name", "") or price_id)
                        new_date = str(getattr(row, "source_updated_at", "") or "")
                        old_date = local_price_list_state.get(f"{account.id}:{price_id}", ("", 0, None))[0]
                        logger.info(
                            "[VW_PRICE_LIST_FOUND] account_id=%s price_id=%s price_name=%s old_date=%s new_date=%s",
                            account.id,
                            price_id,
                            price_name,
                            old_date,
                            new_date,
                        )

                active_fetch_count = 0
                active_fetch_lock = asyncio.Lock()
                logger.info(
                    "[REFRESH] total_price_lists=%s concurrency=%s timeout=%ss",
                    len(price_lists),
                    source_limits.get(account.source_type, 3),
                    PRICE_LIST_FETCH_TIMEOUT_SECONDS,
                )
                print(
                    f"[REFRESH] total_price_lists={len(price_lists)} "
                    f"concurrency={source_limits.get(account.source_type, 3)} "
                    f"timeout={PRICE_LIST_FETCH_TIMEOUT_SECONDS}s",
                    flush=True,
                )
                async def fetch_one(price_list):
                    async with source_sems.get(account.source_type, asyncio.Semaphore(2)):
                        nonlocal active_fetch_count
                        fetch_items_started_at = time.perf_counter()
                        fetch_items_started_wall = local_iso(now_kz_naive())
                        timeout_stage = "precheck"
                        price_list_id = str(getattr(price_list, "price_list_id", ""))
                        price_list_timeout = (
                            PROVISOR_PRICE_TOTAL_TIMEOUT_SECONDS
                            if account.source_type == "provisor"
                            else PRICE_LIST_FETCH_TIMEOUT_SECONDS
                        )
                        async with active_fetch_lock:
                            active_fetch_count += 1
                            active_now = active_fetch_count
                        logger.info(
                            "[REFRESH] START price_list=%s active=%s/%s",
                            price_list_id,
                            active_now,
                            source_limits.get(account.source_type, 3),
                        )
                        print(
                            f"[REFRESH] START price_list={price_list_id} "
                            f"active={active_now}/{source_limits.get(account.source_type, 3)}",
                            flush=True,
                        )
                        try:
                            source_updated_at = str(getattr(price_list, "source_updated_at", "") or "").strip()
                            local_key = f"{account.id}:{price_list_id}"
                            local_updated_at, local_items_count, local_row_updated_at = local_price_list_state.get(local_key, ("", 0, None))
                            price_list_name = str(getattr(price_list, "price_list_name", "") or getattr(price_list, "distributor_name", "") or price_list_id)
                            logger.info(
                                "[PRICE_REFRESH_START] source=%s account_id=%s price_id=%s price_name=%s timeout_seconds=%s",
                                account.source_type,
                                account.id,
                                price_list_id,
                                price_list_name,
                                price_list_timeout,
                            )
                            print(
                                f"[PRICE_REFRESH_START] source={account.source_type} account_id={account.id} "
                                f"price_id={price_list_id} price_name={price_list_name} timeout_seconds={price_list_timeout}",
                                flush=True,
                            )
                            if account.source_type == "provisor" and (
                                price_list_id in excluded_provisor_filial_ids
                                or is_emit_plk(filial_id=price_list_id, name=price_list_name)
                            ):
                                elapsed_ms = round((time.perf_counter() - fetch_items_started_at) * 1000, 2)
                                logger.info(
                                    "[PROVISOR_HEAVY_SKIPPED] account_id=%s filial_id=%s filial_name=%s reason=%s",
                                    account.id,
                                    price_list_id,
                                    price_list_name,
                                    "excluded_emit_or_heavy_filial",
                                )
                                print(
                                    f"[PROVISOR_HEAVY_SKIPPED] account_id={account.id} filial_id={price_list_id} "
                                    f"filial_name={price_list_name} reason=excluded_emit_or_heavy_filial",
                                    flush=True,
                                )
                                return {
                                    "ok": False,
                                    "skipped": True,
                                    "skipped_heavy": True,
                                    "priceList": price_list,
                                    "items": [],
                                    "elapsed_ms": elapsed_ms,
                                    "timeout_limit_seconds": price_list_timeout,
                                    "skippedInfo": {
                                        "id": price_list_id,
                                        "name": price_list_name,
                                        "source": account.source_type,
                                        "reason": "excluded_emit_or_heavy_filial",
                                        "fetch_duration_seconds": round(elapsed_ms / 1000, 3),
                                        "timeout_limit_seconds": price_list_timeout,
                                    },
                                }
                            if account.source_type == "provisor" and not force_refresh:
                                skip_until = _is_provisor_price_unhealthy(account_id=account.id, filial_id=price_list_id)
                                if skip_until is not None:
                                    elapsed_ms = round((time.perf_counter() - fetch_items_started_at) * 1000, 2)
                                    logger.warning(
                                        "[PROVISOR_PRICE_SKIPPED_UNHEALTHY] account_id=%s filial_id=%s price_name=%s skip_until=%s",
                                        account.id,
                                        price_list_id,
                                        price_list_name,
                                        skip_until.isoformat(),
                                    )
                                    return {
                                        "ok": False,
                                        "skipped": True,
                                        "priceList": price_list,
                                    "items": [],
                                    "elapsed_ms": elapsed_ms,
                                    "timeout_limit_seconds": price_list_timeout,
                                    "skippedInfo": {
                                        "id": price_list_id,
                                        "name": price_list_name,
                                        "source": account.source_type,
                                        "reason": "unhealthy_timeout_cache",
                                        "fetch_duration_seconds": round(elapsed_ms / 1000, 3),
                                        "timeout_limit_seconds": price_list_timeout,
                                    },
                                }
                            if account.source_type == "vidman":
                                logger.info(
                                    "[VW_PRICE_LIST_REFRESH_START] account_id=%s price_id=%s price_name=%s old_date=%s new_date=%s items_count=%s",
                                    account.id,
                                    price_list_id,
                                    price_list_name,
                                    local_updated_at,
                                    source_updated_at,
                                    local_items_count,
                                )
                            if (
                                account.source_type == "vidman"
                                and not force_refresh
                                and source_updated_at
                                and local_updated_at == source_updated_at
                                and local_items_count > 0
                            ):
                                elapsed_ms = round((time.perf_counter() - fetch_items_started_at) * 1000, 2)
                                logger.info("[REFRESH] source=%s price_list=%s action=skip_unchanged", account.source_type, price_list_id)
                                logger.info(
                                    "[VIDMAN] price_list=%s action=skip_unchanged updated_at=%s local_updated_at=%s",
                                    price_list_id,
                                    source_updated_at,
                                    local_updated_at,
                                )
                                logger.info(
                                    "[VW_PRICE_LIST_SKIPPED] account_id=%s price_id=%s price_name=%s old_date=%s new_date=%s items_count=%s reason=%s",
                                    account.id,
                                    price_list_id,
                                    price_list_name,
                                    local_updated_at,
                                    source_updated_at,
                                    local_items_count,
                                    "unchanged",
                                )
                                message = f"Прайс {price_list_id} не изменился, пропускаем загрузку товаров"
                                if job is not None:
                                    update_job(
                                        db,
                                        job,
                                        status="running",
                                        progress=60,
                                        message=message,
                                        log_level="info",
                                        log_meta={
                                            "id": price_list_id,
                                            "source": account.source_type,
                                            "updatedAt": source_updated_at,
                                            "action": "skip_unchanged",
                                        },
                                    )
                                return {
                                    "ok": False,
                                    "skipped_unchanged": True,
                                    "priceList": price_list,
                                    "items": [],
                                    "localItemsCount": local_items_count,
                                    "elapsed_ms": elapsed_ms,
                                    "timeout_limit_seconds": price_list_timeout,
                                }
                            if (
                                account.source_type == "provisor"
                                and not force_refresh
                                and local_items_count > 0
                                and local_row_updated_at is not None
                                and now_kz_naive() - local_row_updated_at < PRICE_LIST_REFRESH_TTL
                            ):
                                elapsed_ms = round((time.perf_counter() - fetch_items_started_at) * 1000, 2)
                                logger.info("[REFRESH] source=%s price_list=%s action=skip_unchanged", account.source_type, price_list_id)
                                logger.info(
                                    "[PROVISOR] filial=%s action=skip_unchanged reason=ttl local_updated_at=%s",
                                    price_list_id,
                                    local_updated_at,
                                )
                                return {
                                    "ok": False,
                                    "skipped_unchanged": True,
                                    "priceList": price_list,
                                    "items": [],
                                    "localItemsCount": local_items_count,
                                    "elapsed_ms": elapsed_ms,
                                    "timeout_limit_seconds": price_list_timeout,
                                }
                            if account.source_type == "vidman":
                                timeout_stage = "login"
                                logger.info(
                                    "[VIDMAN] price_list=%s updated_at=%s local_updated_at=%s force_refresh=%s action=fetch_changed",
                                    price_list_id,
                                    source_updated_at,
                                    local_updated_at,
                                    force_refresh,
                                )
                                logger.info("[VIDMAN] price_list=%s action=fetch_items start", price_list_id)
                            logger.info("[REFRESH] source=%s price_list=%s action=fetch", account.source_type, price_list_id)
                            if shared_vidman_client is not None and isinstance(adapter, VidmanPriceService):
                                async with shared_vidman_login_lock:
                                    if not shared_vidman_client._logged_in:
                                        logger.info("[VIDMAN] account_id=%s action=login_once_before_item_fetch", account.id)
                                        await shared_vidman_client.login()
                                timeout_stage = "get_price_items"
                                coro = adapter.fetch_price_list_items_with_client(shared_vidman_client, price_list)
                            else:
                                timeout_stage = "get_price_items"
                                coro = adapter.fetch_price_list_items(credentials, price_list)
                            if account.source_type == "provisor":
                                fetch_task = asyncio.create_task(coro)
                                done, pending = await asyncio.wait({fetch_task}, timeout=price_list_timeout)
                                if pending:
                                    fetch_task.cancel()
                                    fetch_task.add_done_callback(lambda task: task.exception() if not task.cancelled() else None)
                                    logger.warning(
                                        "[PROVISOR_PRICE_CANCELLED] account_id=%s filial_id=%s price_name=%s message=%s",
                                        account.id,
                                        price_list_id,
                                        price_list_name,
                                        "cancelled after total timeout",
                                    )
                                    raise asyncio.TimeoutError(f"Provisor filial {price_list_id} total timeout > {price_list_timeout}s")
                                items = next(iter(done)).result()
                            else:
                                items = await asyncio.wait_for(coro, timeout=price_list_timeout)
                            timeout_stage = "parsing"
                            if account.source_type == "vidman" and isinstance(adapter, VidmanPriceService):
                                items = adapter._unified_items_from_rows(price_list=price_list, rows=items)
                            elapsed_ms = round((time.perf_counter() - fetch_items_started_at) * 1000, 2)
                            if account.source_type == "vidman":
                                logger.info(
                                    "[VW_PRICE_LIST_ITEMS_FETCHED] account_id=%s price_id=%s price_name=%s old_date=%s new_date=%s items_count=%s",
                                    account.id,
                                    price_list_id,
                                    price_list_name,
                                    local_updated_at,
                                    source_updated_at,
                                    len(items or []),
                                )
                            if account.source_type == "provisor":
                                _record_provisor_price_success(account_id=account.id, filial_id=price_list_id)
                                provisor_updated_at = ""
                                for item in items or []:
                                    raw_item = getattr(item, "raw", None)
                                    if isinstance(raw_item, dict):
                                        provisor_updated_at = str(raw_item.get("insertedDate") or "").strip()
                                        if provisor_updated_at:
                                            break
                                if provisor_updated_at:
                                    price_list = replace(price_list, source_updated_at=provisor_updated_at)
                                    if local_updated_at == provisor_updated_at and local_items_count > 0:
                                        logger.info("[REFRESH] source=%s price_list=%s action=skip_unchanged", account.source_type, price_list_id)
                                        logger.info(
                                            "[PROVISOR] filial=%s action=skip_unchanged updated_at=%s",
                                            price_list_id,
                                            provisor_updated_at,
                                        )
                                        return {
                                            "ok": False,
                                            "skipped_unchanged": True,
                                            "priceList": price_list,
                                            "items": [],
                                            "localItemsCount": local_items_count,
                                            "elapsed_ms": elapsed_ms,
                                            "timeout_limit_seconds": price_list_timeout,
                                        }
                                    logger.info(
                                        "[PROVISOR] filial=%s action=fetch_changed updated_at=%s local_updated_at=%s",
                                        price_list_id,
                                        provisor_updated_at,
                                        local_updated_at,
                                    )
                            if account.source_type == "vidman":
                                logger.info(
                                    "[VIDMAN] price_list=%s action=fetch_items done items=%s elapsed_ms=%s",
                                    price_list_id,
                                    len(items or []),
                                    elapsed_ms,
                                )
                            logger.info(
                                "[REFRESH] DONE price_list=%s elapsed_ms=%s items=%s",
                                price_list_id,
                                elapsed_ms,
                                len(items or []),
                            )
                            logger.info(
                                "[PRICE_REFRESH_DONE] source=%s account_id=%s price_id=%s price_name=%s items_count=%s elapsed_ms=%s",
                                account.source_type,
                                account.id,
                                price_list_id,
                                price_list_name,
                                len(items or []),
                                elapsed_ms,
                            )
                            print(
                                f"[REFRESH] DONE price_list={price_list_id} "
                                f"elapsed_ms={elapsed_ms} items={len(items or [])}",
                                flush=True,
                            )
                            _timing(
                                f"{account_operation}:price_list:{price_list_id}",
                                "fetch_price_list_items",
                                fetch_items_started_at,
                            )
                            logger.info(
                                "%s fetched items: account_id=%s price_list_id=%s total=%s",
                                account.source_type.capitalize(),
                                account.id,
                                getattr(price_list, "price_list_id", ""),
                                len(items or []),
                            )
                            return {
                                "ok": True,
                                "priceList": price_list,
                                "items": items,
                                "oldDate": local_updated_at,
                                "newDate": source_updated_at,
                                "localItemsCount": local_items_count,
                                "elapsed_ms": elapsed_ms,
                                "timeout_limit_seconds": price_list_timeout,
                            }
                        except (asyncio.TimeoutError, httpx.TimeoutException):
                            price_list_name = str(getattr(price_list, "price_list_name", "") or getattr(price_list, "distributor_name", "") or price_list_id)
                            warning_message = f"Прайс {price_list_id} пропущен (timeout {PRICE_LIST_FETCH_TIMEOUT_SECONDS}s)"
                            elapsed_ms = round((time.perf_counter() - fetch_items_started_at) * 1000, 2)
                            if account.source_type == "provisor":
                                warning_message = f"Provisor price {price_list_id} skipped (timeout {price_list_timeout}s)"
                                skip_until = _record_provisor_price_timeout(account_id=account.id, filial_id=price_list_id)
                                logger.warning(
                                    "[PROVISOR_PRICE_TOTAL_TIMEOUT] account_id=%s filial_id=%s price_name=%s elapsed_ms=%s timeout_seconds=%s stage=%s",
                                    account.id,
                                    price_list_id,
                                    price_list_name,
                                    elapsed_ms,
                                    price_list_timeout,
                                    timeout_stage,
                                )
                                if skip_until is not None:
                                    logger.warning(
                                        "[PROVISOR_PRICE_SKIPPED_UNHEALTHY] account_id=%s filial_id=%s price_name=%s skip_until=%s",
                                        account.id,
                                        price_list_id,
                                        price_list_name,
                                        skip_until.isoformat(),
                                    )
                            logger.warning(
                                "[TIMEOUT_DEBUG] price_id=%s price_name=%s source=%s started_at=%s elapsed=%s stage=%s exception=%s",
                                price_list_id,
                                price_list_name,
                                account.source_type,
                                fetch_items_started_wall,
                                elapsed_ms,
                                timeout_stage,
                                "TimeoutError",
                            )
                            logger.warning(
                                "[REFRESH] SKIP price_list=%s reason=%s elapsed_ms=%s",
                                price_list_id,
                                f"timeout_{price_list_timeout}s",
                                elapsed_ms,
                            )
                            logger.warning(
                                "[PRICE_REFRESH_TIMEOUT] source=%s account_id=%s price_id=%s price_name=%s elapsed_ms=%s timeout_seconds=%s",
                                account.source_type,
                                account.id,
                                price_list_id,
                                price_list_name,
                                elapsed_ms,
                                price_list_timeout,
                            )
                            print(
                                f"[REFRESH] SKIP price_list={price_list_id} "
                                f"reason=timeout_{price_list_timeout}s elapsed_ms={elapsed_ms}",
                                flush=True,
                            )
                            logger.exception(
                                "[TIMING] operation=%s step=exception elapsed_ms=%s",
                                f"{account_operation}:price_list:{price_list_id}",
                                elapsed_ms,
                            )
                            logger.warning(warning_message)
                            if account.source_type == "vidman":
                                logger.warning(
                                    "[VW_PRICE_LIST_SKIPPED] account_id=%s price_id=%s price_name=%s old_date=%s new_date=%s items_count=%s reason=%s",
                                    account.id,
                                    price_list_id,
                                    price_list_name,
                                    local_updated_at,
                                    source_updated_at,
                                    0,
                                    f"timeout_{price_list_timeout}s",
                                )
                            if job is not None:
                                update_job(
                                    db,
                                    job,
                                    status="running",
                                    progress=60,
                                    message=warning_message,
                                    log_level="warning",
                                    log_meta={
                                        "id": price_list_id,
                                        "name": price_list_name,
                                        "source": account.source_type,
                                        "reason": f"timeout_{price_list_timeout}s",
                                    },
                                )
                            return {
                                "ok": False,
                                "timeout": True,
                                "priceList": price_list,
                                "items": [],
                                "elapsed_ms": elapsed_ms,
                                "timeout_limit_seconds": price_list_timeout,
                                "error": warning_message,
                                "skippedInfo": {
                                    "id": price_list_id,
                                    "name": price_list_name,
                                    "source": account.source_type,
                                    "reason": f"timeout_{price_list_timeout}s",
                                    "fetch_duration_seconds": round(elapsed_ms / 1000, 3),
                                    "timeout_limit_seconds": price_list_timeout,
                                },
                            }
                        except WidmanInvalidCredentialsError:
                            elapsed_ms = round((time.perf_counter() - fetch_items_started_at) * 1000, 2)
                            logger.warning(
                                "[REFRESH] SKIP price_list=%s reason=%s elapsed_ms=%s",
                                price_list_id,
                                "invalid_credentials",
                                elapsed_ms,
                            )
                            print(
                                f"[REFRESH] SKIP price_list={price_list_id} "
                                f"reason=invalid_credentials elapsed_ms={elapsed_ms}",
                                flush=True,
                            )
                            logger.exception(
                                "[TIMING] operation=%s step=exception elapsed_ms=%s",
                                f"{account_operation}:price_list:{price_list_id}",
                                elapsed_ms,
                            )
                            logger.exception(
                                "[VW_REFRESH_ERROR] account_id=%s price_id=%s price_name=%s old_date=%s new_date=%s items_count=%s exception=%s",
                                account.id,
                                price_list_id,
                                price_list_name,
                                local_updated_at,
                                source_updated_at,
                                0,
                                "invalid_credentials",
                            )
                            return {
                                "ok": False,
                                "auth_error": True,
                                "priceList": price_list,
                                "error": "Неверный логин или пароль",
                            }
                        except Exception as e:
                            elapsed_ms = round((time.perf_counter() - fetch_items_started_at) * 1000, 2)
                            logger.warning(
                                "[REFRESH] SKIP price_list=%s reason=%s elapsed_ms=%s",
                                price_list_id,
                                e.__class__.__name__,
                                elapsed_ms,
                            )
                            print(
                                f"[REFRESH] SKIP price_list={price_list_id} "
                                f"reason={e.__class__.__name__} elapsed_ms={elapsed_ms}",
                                flush=True,
                            )
                            logger.exception(
                                "[TIMING] operation=%s step=exception elapsed_ms=%s",
                                f"{account_operation}:price_list:{price_list_id}",
                                elapsed_ms,
                            )
                            logger.exception(
                                "Failed to fetch price list items: source=%s account_id=%s price_list_id=%s",
                                account.source_type,
                                account.id,
                                getattr(price_list, "price_list_id", ""),
                            )
                            if account.source_type == "vidman":
                                logger.exception(
                                    "[VW_REFRESH_ERROR] account_id=%s price_id=%s price_name=%s old_date=%s new_date=%s items_count=%s exception=%s",
                                    account.id,
                                    price_list_id,
                                    price_list_name,
                                    local_updated_at,
                                    source_updated_at,
                                    0,
                                    repr(e),
                                )
                            return {"ok": False, "priceList": price_list, "error": str(e)}
                        finally:
                            async with active_fetch_lock:
                                active_fetch_count = max(0, active_fetch_count - 1)

                tasks = [asyncio.create_task(fetch_one(pl)) for pl in price_lists]
                results = []
                total_price_lists = max(1, len(tasks))
                try:
                    for done in asyncio.as_completed(tasks):
                        result = await done
                        results.append(result)
                        processed_now = len(results)
                        if job is not None:
                            progress_now = 40 + int((processed_now / total_price_lists) * 20)
                            update_job(
                                db,
                                job,
                                status="running",
                                progress=progress_now,
                                message=f"Загружаем товары {account.source_type}: {processed_now}/{total_price_lists}",
                                log_level="info",
                                log_meta={"processed": processed_now, "total": total_price_lists},
                            )
                finally:
                    if shared_vidman_client is not None:
                        await shared_vidman_client.close()
                for result in results:
                    status["processed"] += 1
                    price_list_for_metrics = result.get("priceList") if isinstance(result, dict) else None
                    elapsed_ms_for_metrics = float(result.get("elapsed_ms") or 0) if isinstance(result, dict) else 0.0
                    if price_list_for_metrics is not None and elapsed_ms_for_metrics > 0:
                        status["fetch_duration_seconds"].append(
                            {
                                "priceListId": str(getattr(price_list_for_metrics, "price_list_id", "") or ""),
                                "duration": round(elapsed_ms_for_metrics / 1000, 3),
                                "timeoutLimit": result.get("timeout_limit_seconds"),
                            }
                        )
                    if isinstance(result, Exception):
                        status["errors"] += 1
                        status["results"].append({"ok": False, "error": str(result)})
                    elif result.get("ok"):
                        status["success"] += 1
                        if result.get("items"):
                            status["success_with_items"] += 1
                        else:
                            status["success_zero_items"] += 1
                        status["results"].append(result)
                    elif result.get("timeout"):
                        status["skipped"] += 1
                        status["skipped_timeout"] += 1
                        status["timeout"] += 1
                        if result.get("skippedInfo"):
                            status["skipped_price_lists"].append(result["skippedInfo"])
                        status["results"].append(result)
                    elif result.get("skipped_heavy"):
                        status["skipped"] += 1
                        status["skipped_heavy_count"] += 1
                        status["heavy_excluded"] += 1
                        if result.get("skippedInfo"):
                            status["skipped_heavy_filials"].append(result["skippedInfo"])
                        status["results"].append(result)
                    elif result.get("skipped_unchanged"):
                        status["skipped"] += 1
                        status["skipped_unchanged"] += 1
                        status["results"].append(result)
                    elif result.get("skipped"):
                        status["skipped"] += 1
                        if result.get("skippedInfo"):
                            status["skipped_price_lists"].append(result["skippedInfo"])
                        status["results"].append(result)
                    else:
                        status["errors"] += 1
                        status["results"].append(result)

                if any(isinstance(result, dict) and result.get("auth_error") for result in results):
                    status.update(
                        {
                            "source": "vidman",
                            "status": "auth_error",
                            "ok": False,
                            "priceListsCount": 0,
                            "priceLists": 0,
                            "items": 0,
                            "success": 0,
                            "skipped": 1,
                            "skipped_auth_error_count": 1,
                            "errors": 0,
                            "message": "Неверный логин или пароль",
                            "statusMessage": "Неверный логин или пароль",
                            "results": [],
                        }
                    )
                else:
                    status["ok"] = status["success"] > 0
                    status["priceListsCount"] = status["success"]
                    if status.get("skipped_timeout"):
                        status["message"] = (
                            f"Обновление завершено, но {status['skipped_timeout']} прайсов пропущено по таймауту"
                        )
                    else:
                        status["message"] = (
                            f"Обновлено: {status['success']}, без изменений: {status.get('skipped_unchanged', 0)}, "
                            f"timeout: {status.get('skipped_timeout', 0)}, heavy: {status.get('skipped_heavy_count', 0)}, ошибки: {status['errors']}"
                        )
            except WidmanInvalidCredentialsError:
                logger.exception("[TIMING] operation=%s step=exception elapsed_ms=%s", account_operation, round((time.perf_counter() - account_started_at) * 1000, 2))
                logger.warning(
                    "Vidman account skipped because of invalid credentials: account_id=%s login=%s",
                    account.id,
                    account.login,
                )
                logger.info("[VIDMAN_AUTH_ERROR_SKIPPED] account_id=%s reason=auth_error_status", account.id)
                status.update(
                    {
                        "source": "vidman",
                        "status": "auth_error",
                        "ok": False,
                        "priceListsCount": 0,
                        "priceLists": 0,
                        "items": 0,
                        "skipped": 1,
                        "skipped_auth_error_count": 1,
                        "errors": 0,
                        "message": "Неверный логин или пароль",
                        "statusMessage": "Неверный логин или пароль",
                        "results": [],
                    }
                )
            except Exception as e:
                logger.exception("[TIMING] operation=%s step=exception elapsed_ms=%s", account_operation, round((time.perf_counter() - account_started_at) * 1000, 2))
                logger.exception(
                    "price source refresh failed before item fetch: source=%s account_id=%s login=%s",
                    account.source_type,
                    account.id,
                    account.login,
                )
                status["errors"] += 1
                status["message"] = str(e)
            if shared_vidman_client is not None:
                await shared_vidman_client.close()
            if account.source_type == "vidman":
                logger.info(
                    "[VW_REFRESH_DONE] account_id=%s price_id=%s price_name=%s old_date=%s new_date=%s items_count=%s reason=%s",
                    account.id,
                    "",
                    "",
                    "",
                    "",
                    status.get("success", 0),
                    status.get("message", ""),
                )
            return status

    fetched_statuses = await asyncio.gather(*(fetch_account(account) for account in accounts), return_exceptions=True)

    source_slowest_price_lists: dict[str, list[dict[str, object]]] = {}
    for account, status in zip(accounts, fetched_statuses):
        account_id = int(account.id)
        account_source_type = str(account.source_type or "")
        account_login = str(account.login or "")
        if isinstance(status, Exception):
            logger.exception("price source refresh task failed: account_id=%s", account_id, exc_info=status)
            status = {
                "sourceType": account_source_type,
                "accountId": account_id,
                "login": account_login,
                "ok": False,
                "priceListsCount": 0,
                "success": 0,
                "skipped": 0,
                "errors": 1,
                "processed": 0,
                    "total": 0,
                    "message": str(status),
                    "results": [],
                    "skipped_timeout": 0,
                    "skipped_unchanged": 0,
                    "skipped_heavy_count": 0,
                    "skipped_heavy_filials": [],
                    "skipped_auth_error_count": 0,
                    "skipped_price_lists": [],
                }

        saved_count = 0
        if job is not None:
            update_job(db, job, status="running", progress=80, message="Сохраняем прайсы в БД", log_level="info")
        result_rows = [x for x in status.get("results", []) if isinstance(x, dict)]
        source_slowest_price_lists.setdefault(account_source_type, []).extend(
            sorted(
                [
                    {
                        "account_id": account_id,
                        "price_id": str(getattr(x.get("priceList"), "price_list_id", "") or ""),
                        "price_name": str(
                            getattr(x.get("priceList"), "price_list_name", "")
                            or getattr(x.get("priceList"), "distributor_name", "")
                            or getattr(x.get("priceList"), "price_list_id", "")
                            or ""
                        ),
                        "elapsed_ms": float(x.get("elapsed_ms") or 0),
                        "reason": str((x.get("skippedInfo") or {}).get("reason") or ("timeout" if x.get("timeout") else "")),
                    }
                    for x in result_rows
                    if float(x.get("elapsed_ms") or 0) > 0
                ],
                key=lambda row: float(row.get("elapsed_ms") or 0),
                reverse=True,
            )[:5]
        )
        total_to_save = sum(1 for x in result_rows if x.get("ok") and x.get("priceList") is not None)
        saving_index = 0
        for result in result_rows:
            price_list = result.get("priceList")
            if price_list is not None and result.get("skipped_unchanged"):
                try:
                    mark_unified_price_list_checked(
                        db=db,
                        price_format_code=format_code,
                        price_list=price_list,
                        status="checked_unchanged",
                        message="Проверено успешно; исходный прайс-лист не изменился",
                    )
                except Exception:
                    logger.exception(
                        "Failed to mark unchanged price list check: source=%s account_id=%s price_list_id=%s",
                        account_source_type,
                        account_id,
                        getattr(price_list, "price_list_id", ""),
                    )
            elif price_list is not None and result.get("timeout"):
                try:
                    mark_unified_price_list_checked(
                        db=db,
                        price_format_code=format_code,
                        price_list=price_list,
                        status="timeout",
                        message=str(result.get("error") or "Refresh timed out"),
                    )
                except Exception:
                    logger.exception(
                        "Failed to mark timeout price list check: source=%s account_id=%s price_list_id=%s",
                        account_source_type,
                        account_id,
                        getattr(price_list, "price_list_id", ""),
                    )
            if result.get("ok") and price_list is not None:
                saving_index += 1
                price_list_id = str(getattr(price_list, "price_list_id", "") or "")
                price_list_name = str(
                    getattr(price_list, "price_list_name", "")
                    or getattr(price_list, "distributor_name", "")
                    or price_list_id
                )
                if job is not None:
                    update_job(
                        db,
                        job,
                        status="running",
                        progress=80,
                        message=f"Сохраняем прайс {saving_index}/{max(1, total_to_save)}: {price_list_name}",
                        log_level="info",
                        log_meta={
                            "source": account_source_type,
                            "accountId": account_id,
                            "priceListId": price_list_id,
                            "index": saving_index,
                            "total": total_to_save,
                        },
                    )
                save_started_at = time.perf_counter()
                try:
                    items = result.get("items") or []
                    if not items and int(result.get("localItemsCount") or 0) > 0:
                        mark_unified_price_list_checked(
                            db=db,
                            price_format_code=format_code,
                            price_list=price_list,
                            status="success_zero_items",
                            message="Проверено успешно; пустой ответ сохранил существующие строки",
                        )
                        refreshed.append(
                            {
                                "sourceType": account_source_type,
                                "sourceKey": f"{account_id}:{price_list_id}",
                                "accountId": account_id,
                                "name": price_list_name,
                                "itemsCount": 0,
                                "preservedItemsCount": int(result.get("localItemsCount") or 0),
                                "status": "success_zero_items",
                                "fetchDurationSeconds": round(float(result.get("elapsed_ms") or 0) / 1000, 3),
                                "timeoutLimitSeconds": result.get("timeout_limit_seconds"),
                            }
                        )
                        continue
                    saved = upsert_unified_price_list(
                        db=db,
                        price_format_code=format_code,
                        price_list=price_list,
                        items=items,
                        status="updated",
                        run_matching=False,
                    )
                    db_replace_elapsed_ms = round((time.perf_counter() - save_started_at) * 1000, 2)
                    if account_source_type == "provisor":
                        logger.info(
                            "[PROVISOR_PLK_DB_REPLACE_TIMING] account_id=%s filial_id=%s rows=%s db_replace_elapsed_ms=%s",
                            account_id,
                            price_list_id,
                            len(items),
                            db_replace_elapsed_ms,
                        )
                    saved_source_type = str(saved.source_type or "")
                    saved_source_key = str(saved.source_key or "")
                    saved_display_name = str(saved.display_name or price_list_name)
                    if account_source_type == "vidman":
                        logger.info(
                            "[VW_PRICE_LIST_SAVED] account_id=%s price_id=%s price_name=%s old_date=%s new_date=%s items_count=%s",
                            account_id,
                            price_list_id,
                            price_list_name,
                            result.get("oldDate") or "",
                            getattr(price_list, "source_updated_at", "") or saved.source_updated_at or "",
                            len(items),
                        )
                    _timing(f"{operation}:account:{account_id}:price_list:{price_list_id}", "normalize/map_items", save_started_at)
                    _timing(f"{operation}:account:{account_id}:price_list:{price_list_id}", "save_to_db", save_started_at)
                    refreshed.append(
                        {
                            "sourceType": saved_source_type,
                            "sourceKey": saved_source_key,
                            "accountId": account_id,
                            "name": saved_display_name,
                            "itemsCount": len(items),
                            "fetchDurationSeconds": round(float(result.get("elapsed_ms") or 0) / 1000, 3),
                            "timeoutLimitSeconds": result.get("timeout_limit_seconds"),
                        }
                    )
                    saved_count += 1
                    if job is not None:
                        update_job(
                            db,
                            job,
                            status="running",
                            progress=80,
                            message=f"Сохранён прайс {saving_index}/{max(1, total_to_save)}: {saved_display_name}",
                            log_level="info",
                            log_meta={
                                "source": saved_source_type,
                                "sourceKey": saved_source_key,
                                "accountId": account_id,
                                "itemsCount": len(items),
                            },
                        )
                except Exception as e:
                    elapsed_ms = round((time.perf_counter() - save_started_at) * 1000, 2)
                    db.rollback()
                    logger.exception(
                        "[TIMING] operation=%s step=exception elapsed_ms=%s",
                        f"{operation}:account:{account_id}:price_list:{price_list_id}",
                        elapsed_ms,
                    )
                    logger.exception(
                        "Failed to save price list: source=%s account_id=%s price_list_id=%s",
                        account_source_type,
                        account_id,
                        price_list_id,
                    )
                    if account_source_type == "vidman":
                        logger.exception(
                            "[VW_REFRESH_ERROR] account_id=%s price_id=%s price_name=%s old_date=%s new_date=%s items_count=%s exception=%s",
                            account_id,
                            price_list_id,
                            price_list_name,
                            result.get("oldDate") or "",
                            getattr(price_list, "source_updated_at", "") or "",
                            len(result.get("items") or []),
                            repr(e),
                        )
                    if _is_sqlite_full(e):
                        raise RuntimeError(
                            "SQLite database or disk is full while saving competitor price lists. "
                            "Free disk space or reduce stored raw payloads, then run refresh again."
                        ) from e
                    status["errors"] = int(status.get("errors") or 0) + 1
                    status["success"] = max(0, int(status.get("success") or 0) - 1)
                    errors.append(f"{account_source_type} / {account_login} / {price_list_name}: {e}")
            elif result.get("error"):
                errors.append(f"{account_source_type} / {account_login}: {result.get('error')}")

        account_row = db.get(PriceSourceAccount, account_id)
        if account_row is not None:
            available_count = (
                saved_count
                + int(status.get("skipped_unchanged") or 0)
                + int(status.get("skipped_heavy_count") or 0)
                + int(status.get("success_zero_items") or 0)
            )
            account_row.status = "connected" if available_count > 0 else "auth_error"
            account_row.status_message = str(status.get("message") or "")[:512]
            if available_count > 0:
                account_row.last_success_at = now_kz_naive()
            account_row.price_lists_count = available_count
            account_row.updated_at = now_kz_naive()
        commit_started_at = time.perf_counter()
        db.commit()
        _timing(f"{operation}:account:{account_id}", "commit", commit_started_at)

        status["ok"] = (
            saved_count
            + int(status.get("skipped_unchanged") or 0)
            + int(status.get("skipped_heavy_count") or 0)
            + int(status.get("success_zero_items") or 0)
        ) > 0
        status["priceListsCount"] = saved_count + int(status.get("skipped_unchanged") or 0) + int(status.get("success_zero_items") or 0)
        status.pop("results", None)
        account_statuses.append(status)
        if status.get("skipped_price_lists"):
            errors.extend(
                [
                    f"{item.get('source')} / {item.get('name') or item.get('id')}: нет ответа за {PRICE_LIST_FETCH_TIMEOUT_SECONDS} секунд"
                    for item in status.get("skipped_price_lists", [])
                ]
            )

        source_summary = progress["bySource"].setdefault(
            account_source_type,
            {"total": 0, "processed": 0, "success": 0, "skipped": 0, "errors": 0},
        )
        for key in ("total", "processed", "success", "skipped", "errors"):
            source_summary[key] += int(status.get(key) or 0)
            progress[key] += int(status.get(key) or 0)
        source_summary["skipped_timeout"] = int(source_summary.get("skipped_timeout") or 0) + int(status.get("skipped_timeout") or 0)
        progress["skipped_timeout"] = int(progress.get("skipped_timeout") or 0) + int(status.get("skipped_timeout") or 0)
        source_summary["skipped_unchanged"] = int(source_summary.get("skipped_unchanged") or 0) + int(status.get("skipped_unchanged") or 0)
        progress["skipped_unchanged"] = int(progress.get("skipped_unchanged") or 0) + int(status.get("skipped_unchanged") or 0)
        source_summary["skipped_heavy"] = int(source_summary.get("skipped_heavy") or 0) + int(status.get("skipped_heavy_count") or 0)
        progress["skipped_heavy"] = int(progress.get("skipped_heavy") or 0) + int(status.get("skipped_heavy_count") or 0)
        source_summary["skipped_auth_error_count"] = int(source_summary.get("skipped_auth_error_count") or 0) + int(status.get("skipped_auth_error_count") or 0)
        progress["skipped_auth_error_count"] = int(progress.get("skipped_auth_error_count") or 0) + int(status.get("skipped_auth_error_count") or 0)
        source_summary["success_with_items"] = int(source_summary.get("success_with_items") or 0) + int(status.get("success_with_items") or 0)
        progress["success_with_items"] = int(progress.get("success_with_items") or 0) + int(status.get("success_with_items") or 0)
        source_summary["success_zero_items"] = int(source_summary.get("success_zero_items") or 0) + int(status.get("success_zero_items") or 0)
        progress["success_zero_items"] = int(progress.get("success_zero_items") or 0) + int(status.get("success_zero_items") or 0)
        source_summary["timeout"] = int(source_summary.get("timeout") or 0) + int(status.get("timeout") or 0)
        progress["timeout"] = int(progress.get("timeout") or 0) + int(status.get("timeout") or 0)
        source_summary["heavy_excluded"] = int(source_summary.get("heavy_excluded") or 0) + int(status.get("heavy_excluded") or 0)
        progress["heavy_excluded"] = int(progress.get("heavy_excluded") or 0) + int(status.get("heavy_excluded") or 0)
        source_summary.setdefault("fetch_duration_seconds", [])
        progress.setdefault("fetch_duration_seconds", [])
        source_summary["fetch_duration_seconds"].extend(status.get("fetch_duration_seconds") or [])
        progress["fetch_duration_seconds"].extend(status.get("fetch_duration_seconds") or [])
        source_summary.setdefault("skipped_heavy_filials", [])
        progress.setdefault("skipped_heavy_filials", [])
        source_summary["skipped_heavy_filials"].extend(status.get("skipped_heavy_filials") or [])
        progress["skipped_heavy_filials"].extend(status.get("skipped_heavy_filials") or [])

    for source, summary in progress.get("bySource", {}).items():
        slowest = sorted(
            source_slowest_price_lists.get(source, []),
            key=lambda row: float(row.get("elapsed_ms") or 0),
            reverse=True,
        )[:5]
        summary["slowest_price_lists"] = slowest
        elapsed_ms = round((time.perf_counter() - source_started_at) * 1000, 2)
        logger.info(
            "[REFRESH_SOURCE_SUMMARY] source=%s accounts_count=%s total=%s success=%s timeout=%s failed=%s skipped=%s skipped_heavy=%s skipped_auth_error=%s elapsed_ms=%s slowest_price_lists=%s",
            source,
            source_counts.get(source, 0),
            int(summary.get("total") or 0),
            int(summary.get("success") or 0),
            int(summary.get("skipped_timeout") or 0),
            int(summary.get("errors") or 0),
            int(summary.get("skipped") or 0),
            int(summary.get("skipped_heavy") or 0),
            int(summary.get("skipped_auth_error_count") or 0),
            elapsed_ms,
            slowest,
        )
        print(
            f"[REFRESH_SOURCE_SUMMARY] source={source} accounts_count={source_counts.get(source, 0)} "
            f"total={int(summary.get('total') or 0)} success={int(summary.get('success') or 0)} "
            f"timeout={int(summary.get('skipped_timeout') or 0)} failed={int(summary.get('errors') or 0)} "
            f"skipped={int(summary.get('skipped') or 0)} skipped_heavy={int(summary.get('skipped_heavy') or 0)} "
            f"skipped_auth_error={int(summary.get('skipped_auth_error_count') or 0)} "
            f"elapsed_ms={elapsed_ms} slowest_price_lists={slowest}",
            flush=True,
        )

    rebuild_summary = {}
    if pf_for_refresh is not None and run_rebuild_after_refresh:
        if job is not None:
            update_job(db, job, status="running", progress=90, message="Пересобираем цены выбранных конкурентов", log_level="info")
        rebuild_started_at = time.perf_counter()
        sync_selected_competitor_configs(db=db, price_format_id=pf_for_refresh.id)
        rebuild_summary = rebuild_competitor_prices_for_selected(db=db, price_format_id=pf_for_refresh.id, commit_between_lists=True)
        recalculate_competitor_percentiles_if_needed(db=db, price_format_id=pf_for_refresh.id)
        db.commit()
        _timing(operation, "rebuild_selected_competitor_prices", rebuild_started_at)
    else:
        logger.info(
            '[REFRESH_NO_MATCHING] format_code=%s message="Refresh completed without matching/rebuild"',
            format_code,
        )

    _timing(operation, "finish", refresh_started_at)
    _timing(operation, "total_ms", refresh_started_at)
    updated_count = int(progress.get("success") or 0)
    skipped_unchanged = int(progress.get("skipped_unchanged") or 0)
    skipped_timeout = int(progress.get("skipped_timeout") or 0)
    error_count = int(progress.get("errors") or 0)
    progress["updated_count"] = updated_count
    progress["error_count"] = error_count
    progress["skipped_heavy_count"] = int(progress.get("skipped_heavy") or 0)
    return {
        "source": refresh_source,
        "accounts_requested": requested_account_ids,
        "accounts_processed": processed_account_ids,
        "accounts_skipped": skipped_requested_account_ids,
        "updated_count": updated_count,
        "skipped_unchanged": skipped_unchanged,
        "skipped_timeout": skipped_timeout,
        "skipped_heavy_count": int(progress.get("skipped_heavy") or 0),
        "skipped_heavy_filials": progress.get("skipped_heavy_filials") or [],
        "skipped_auth_error_count": int(progress.get("skipped_auth_error_count") or 0),
        "error_count": error_count,
        "refreshed": refreshed,
        "errors": errors,
        "accounts": account_statuses,
        "progress": progress,
        "rebuild": rebuild_summary,
        "items": list_competitor_price_lists(db=db, price_format_code=format_code),
    }


async def _run_refresh_price_lists_job(db: Session, job: Job, *, format_code: str, payload: dict) -> dict:
    update_job(db, job, status="running", progress=10, message="Загружаем аккаунты", log_level="info")
    result = await _run_refresh_price_lists_logic(format_code=format_code, payload=payload, db=db, job=job)
    job = db.get(Job, job.id)
    if job is not None:
        skipped_timeout = int((result.get("progress") or {}).get("skipped_timeout") or 0) if isinstance(result, dict) else 0
        message = (
            f"Обновление завершено, но {skipped_timeout} прайсов пропущено по таймауту"
            if skipped_timeout
            else "Обновление прайс-листов готово"
        )
        update_job(db, job, status="running", progress=100, message=message, result=result, log_level="warning" if skipped_timeout else "info")
    return result


async def _run_generate_price_job(db: Session, job: Job, *, price_format_id: int, payload: dict | None = None) -> dict:
    return await asyncio.to_thread(_run_generate_price_job_thread, job.id, price_format_id, payload or {})


def _run_generate_price_job_thread(job_id: str, price_format_id: int, payload: dict) -> dict:
    db = SessionLocal()
    try:
        job = db.get(Job, job_id)
        if job is None:
            raise ValueError(f"job not found: {job_id}")
        return _run_generate_price_job_sync(db, job, price_format_id=price_format_id, payload=payload)
    finally:
        db.close()


def _run_generate_price_job_sync(db: Session, job: Job, *, price_format_id: int, payload: dict | None = None) -> dict:
    payload = payload or {}
    logger.info("[GENERATE] start price_format_id=%s", price_format_id)
    update_job(db, job, status="running", progress=10, message="Загружаем выбранные прайсы", log_level="info")
    products_loaded = int(db.execute(select(func.count(Product.id))).scalar_one() or 0)
    logger.info("[GENERATE] products_loaded=%s", products_loaded)
    update_job(db, job, status="running", progress=25, message="Строим индексы товаров", log_level="info")
    update_job(db, job, status="running", progress=45, message="Matching выбранных прайсов", log_level="info")
    selected_price_lists_count = len(get_assigned_competitor_price_lists(db=db, price_format_id=price_format_id))
    logger.info(
        "[GENERATE_MATCHING_START] format_code=%s selected_price_lists_count=%s",
        job.format_code or payload.get("format_code") or payload.get("price_format_code") or "",
        selected_price_lists_count,
    )
    pf_for_generation = db.get(PriceFormat, price_format_id)
    percentile_mode = bool(pf_for_generation and (pf_for_generation.competitor_price_mode or "regular") == "percentile")
    if percentile_mode:
        summary = {"percentile_mode": {"skippedRawCompetitorPriceRebuild": True}}
    else:
        summary = rebuild_competitor_prices_for_selected(db=db, price_format_id=price_format_id, commit_between_lists=True)
    update_job(db, job, status="running", progress=70, message="Пересборка competitor_prices завершена", result={"summary": summary}, log_level="info")
    if not percentile_mode:
        recalculate_competitor_percentiles_if_needed(db=db, price_format_id=price_format_id)
    db.commit()
    competitor_prices_loaded = int(
        db.execute(
            select(func.count(CompetitorPrice.id))
            .where(CompetitorPrice.price_format_id == price_format_id)
            .where(CompetitorPrice.product_id.is_not(None))
        ).scalar_one()
        or 0
    )
    if percentile_mode:
        competitor_prices_loaded = int(
            db.execute(
                select(func.count(CompetitorPricePercentile.id))
                .where(CompetitorPricePercentile.price_format_id == price_format_id)
                .where(CompetitorPricePercentile.value.is_not(None))
            ).scalar_one()
            or 0
        )
        logger.info("[GENERATE] percentile_rows_loaded=%s", competitor_prices_loaded)
    coverage_values = [float(row.get("coverage") or row.get("matchRate") or 0) for row in summary.values() if isinstance(row, dict)]
    coverage = round(sum(coverage_values) / len(coverage_values), 2) if coverage_values else 0
    logger.info("[GENERATE] competitor_prices_loaded=%s", competitor_prices_loaded)
    logger.info("[GENERATE] coverage=%s", coverage)
    if competitor_prices_loaded <= 0:
        raise ValueError("данные ещё не готовы: competitor_prices не сформированы")

    price_format_code = str(payload.get("price_format_code") or payload.get("format_code") or job.format_code or "").strip()
    activation_date = None
    raw_activation = payload.get("activation_date")
    if isinstance(raw_activation, str) and raw_activation.strip():
        activation_date = date.fromisoformat(raw_activation.strip())
    as_of = activation_date or date.today()
    price_list_number = payload.get("price_list_number")
    if not isinstance(price_list_number, str) or not price_list_number.strip():
        price_list_number = f"{price_format_code}_{as_of.isoformat()}"
    user = payload.get("user")
    user_s = user.strip() if isinstance(user, str) else "UI"
    region_id = int(payload["region"]) if payload.get("region") is not None else None
    update_job(db, job, status="running", progress=90, message="Записываем результат", log_level="info")
    calculated_count = calculate_prices(
        db=db,
        price_format_code=price_format_code,
        price_list_number=price_list_number,
        as_of=as_of,
        activation_date=activation_date,
        user=user_s,
        region_id=region_id,
    )
    if calculated_count == 0:
        raise ValueError("Нет товаров для расчёта. Сначала загрузите Excel с товарами.")
    return {"summary": summary, "price_list_number": price_list_number, "calculated_count": calculated_count}


@app.get("/api/price-formats/{format_code}/lists")
def get_lists_for_format(format_code: str):
    return data.LISTS_BY_FORMAT.get(format_code, [])


@app.get("/api/price-formats/{format_code}/counterparties")
def get_counterparties_for_format(format_code: str):
    return data.COUNTERPARTIES_BY_FORMAT.get(format_code, [])


@app.get("/api/price-formats/{format_code}/settings")
def get_settings_for_format(
    format_code: str,
    db: Session = Depends(get_db),
    current_user: AppUser | None = Depends(get_current_user),
):
    pf = db.execute(select(PriceFormat).where(PriceFormat.code == format_code)).scalars().first()
    if not pf:
        return data.PRICING_SETTINGS_BY_FORMAT.get(format_code) or data.PRICING_SETTINGS_BY_FORMAT.get(
            "ИПЛ_01_001"
        )

    if current_user is not None:
        _ensure_price_format_access(pf, current_user)

    ranges = db.execute(
        select(MarkupRange)
        .where(MarkupRange.price_format_id == pf.id)
        .order_by(MarkupRange.cost_from.asc())
    ).scalars().all()

    bend_rows = db.execute(
        select(BendRange)
        .where(BendRange.price_format_id == pf.id)
        .order_by(BendRange.price_from.asc())
    ).scalars().all()

    no_competitor_rows = db.execute(
        select(NoCompetitorMarkupRange)
        .where(NoCompetitorMarkupRange.price_format_id == pf.id)
        .order_by(NoCompetitorMarkupRange.cost_from.asc())
    ).scalars().all()

    return {
        "name": pf.code,
        "branch": pf.branch,
        "pricingRule": pf.pricing_rule or "",
        "pricingRuleId": int(pf.pricing_rule_id) if pf.pricing_rule_id is not None else None,
        "roundingRuleId": int(pf.rounding_rule_id) if pf.rounding_rule_id is not None else None,
        "appliedRule": pricing_rule_application_status(db=db, pf=pf),
        "competitorPriceMode": pf.competitor_price_mode or "regular",
        "percentileNumber": int(pf.percentile_number or 10),
        "deflectionPercent": float(pf.progib or 0),
        "includeVAT": True,
        "useMinCompetitor": True,
        "considerStock": False,
        "recommendedMarkups": [
            {
                "id": idx + 1,
                "lowerBound": float(r.cost_from),
                "upperBound": float(r.cost_to) if r.cost_to is not None else 99999999,
                "markupPercent": float(r.markup_percent) * 100,
            }
            for idx, r in enumerate(ranges)
        ],
        "bendRanges": [
            {
                "id": idx + 1,
                "priceFrom": float(r.price_from),
                "bendPercent": float(r.bend_percent),
            }
            for idx, r in enumerate(bend_rows)
        ],
        "noCompetitorMarkups": [
            {
                "id": idx + 1,
                "lowerBound": float(r.cost_from),
                "upperBound": float(r.cost_to) if r.cost_to is not None else 99999999,
                "markupPercent": float(r.markup_percent) * 100,
            }
            for idx, r in enumerate(no_competitor_rows)
        ],
    }


@app.put("/api/price-formats/{format_code}/settings")
def put_settings_for_format(
    format_code: str,
    payload: dict = Body(...),
    db: Session = Depends(get_db),
    current_user: AppUser = Depends(require_write_access),
):
    pf = db.execute(select(PriceFormat).where(PriceFormat.code == format_code)).scalars().first()
    if pf is not None:
        _ensure_price_format_access(pf, current_user)
    if pf is None:
        pf = PriceFormat(code=format_code, name=payload.get("name") or format_code)
        db.add(pf)
        db.flush()

    if isinstance(payload.get("branch"), str):
        pf.branch = payload["branch"]
    if isinstance(payload.get("pricingRule"), str):
        pf.pricing_rule = payload["pricingRule"]
    if "roundingRuleId" in payload and payload.get("roundingRuleId") in (None, "", "none"):
        pf.rounding_rule_id = None
    if payload.get("roundingRuleId") not in (None, "", "none"):
        try:
            pf.rounding_rule_id = int(payload.get("roundingRuleId"))
        except Exception:
            pass
    if payload.get("pricingRuleId") not in (None, ""):
        try:
            apply_pricing_rule_to_format(db=db, format_code=format_code, rule_id=int(payload.get("pricingRuleId")))
            return get_settings_for_format(format_code=format_code, db=db, current_user=current_user)
        except Exception:
            db.rollback()
            raise
    if payload.get("competitorPriceMode") in {"regular", "percentile"}:
        pf.competitor_price_mode = payload["competitorPriceMode"]
    if payload.get("percentileNumber") is not None:
        try:
            pct = int(payload.get("percentileNumber"))
            if pct in {10, 20, 30, 40, 50, 60, 70, 80, 90}:
                pf.percentile_number = pct
        except Exception:
            pass

    deflection = payload.get("deflectionPercent")
    if deflection is not None:
        try:
            pf.progib = float(deflection)
        except Exception:
            pass

    # Replace markup ranges
    rec = payload.get("recommendedMarkups")
    if isinstance(rec, list):
        db.execute(delete(MarkupRange).where(MarkupRange.price_format_id == pf.id))
        for row in rec:
            if not isinstance(row, dict):
                continue
            lb = row.get("lowerBound")
            ub = row.get("upperBound")
            mp = row.get("markupPercent")
            try:
                lb_f = float(lb)
                ub_f = float(ub) if ub is not None else None
                mp_f = float(mp)
            except Exception:
                continue

            db.add(
                MarkupRange(
                    price_format_id=pf.id,
                    cost_from=lb_f,
                    cost_to=ub_f,
                    markup_percent=mp_f / 100.0,
                )
            )

    # Replace bend ranges (step table by competitor price)
    bends = payload.get("bendRanges")
    if isinstance(bends, list):
        db.execute(delete(BendRange).where(BendRange.price_format_id == pf.id))
        for row in bends:
            if not isinstance(row, dict):
                continue
            pf_raw = row.get("priceFrom")
            bp_raw = row.get("bendPercent")
            try:
                pf_f = float(pf_raw)
                bp_f = float(bp_raw)
            except Exception:
                continue

            db.add(
                BendRange(
                    price_format_id=pf.id,
                    price_from=pf_f,
                    bend_percent=bp_f,
                )
            )

    no_comp = payload.get("noCompetitorMarkups")
    if isinstance(no_comp, list):
        db.execute(delete(NoCompetitorMarkupRange).where(NoCompetitorMarkupRange.price_format_id == pf.id))
        for row in no_comp:
            if not isinstance(row, dict):
                continue
            lb = row.get("lowerBound")
            ub = row.get("upperBound")
            mp = row.get("markupPercent")
            try:
                lb_f = float(lb)
                ub_f = float(ub) if ub is not None else None
                mp_f = float(mp)
            except Exception:
                continue

            db.add(
                NoCompetitorMarkupRange(
                    price_format_id=pf.id,
                    cost_from=lb_f,
                    cost_to=ub_f,
                    markup_percent=mp_f / 100.0,
                )
            )

    db.commit()
    return get_settings_for_format(format_code=format_code, db=db, current_user=current_user)


@app.post("/upload-excel", response_model=UploadExcelResponse)
async def upload_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="empty file")

    try:
        counts = import_excel(db=db, content=content)
        return UploadExcelResponse(**counts)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.post("/api/upload-excel", response_model=UploadExcelResponse)
async def upload_excel_api(file: UploadFile = File(...), db: Session = Depends(get_db)):
    return await upload_excel(file=file, db=db)


@app.post("/api/products/upload-excel")
async def upload_products_excel(
    file: UploadFile = File(...),
    region: int = Query(8, ge=1),
    price_mode: int = Query(0, ge=0),
    distributors: str = Query("4"),
    db: Session = Depends(get_db),
):
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="empty file")

    top_by_sku: dict[str, int] = {}
    warnings: list[str] = []
    distributor_ids: list[int] = []
    for part in str(distributors or "").split(","):
        part = part.strip()
        if not part:
            continue
        try:
            value = int(part)
        except Exception:
            continue
        if value > 0 and value not in distributor_ids:
            distributor_ids.append(value)
    if distributor_ids:
        try:
            top_by_sku = await FarmcenterTopService(
                base_url=settings.phcenter_base_url,
                token=settings.phcenter_token,
            ).load_top_by_sku(region=region, price_mode=price_mode, distributors=distributor_ids)
        except Exception as e:
            logger.warning("[FARMCENTER_TOP] failed: %s", e)
            warnings.append("Не удалось загрузить TOP-рейтинг Фармцентра")

    try:
        items, stats = import_products_excel(db=db, content=content, top_by_sku=top_by_sku)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Excel import failed: {e}")

    return {"stats": stats, "items": items, "warnings": warnings}


@app.get("/api/products/with-competitor-prices")
def products_with_competitor_prices(
    format_code: str | None = Query(None),
    price_list_number: str | None = Query(None),
    region: int | None = Query(None, ge=1),
    db: Session = Depends(get_db),
):
    return get_products_with_competitor_top5(
        db=db,
        price_format_code=format_code,
        price_list_number=price_list_number,
        region_id=region,
    )


def _substitute_to_dict(row: ProductSubstituteMatch) -> dict:
    return {
        "id": row.id,
        "productId": row.product_id,
        "sourceType": row.source_type,
        "sourceGoodsId": row.source_goods_id,
        "sourceDistributorGoodsId": row.source_distributor_goods_id,
        "sourceName": row.source_name,
        "sourceManufacturer": row.source_manufacturer,
        "status": row.status,
        "priority": row.priority,
        "comment": row.comment,
        "createdAt": _fmt_dt(row.created_at),
        "updatedAt": _fmt_dt(row.updated_at),
    }


def _structure_warnings(product_name: object, source_name: object) -> list[str]:
    product_structure = parse_drug_structure(product_name)
    source_structure = parse_drug_structure(source_name)
    warnings: list[str] = []
    checks = [
        ("dosage", "дозировка отличается"),
        ("form", "форма выпуска отличается"),
        ("quantity", "количество отличается"),
        ("volume", "объем отличается"),
        ("weight", "вес отличается"),
        ("concentration", "концентрация отличается"),
    ]
    for attr, label in checks:
        product_value = getattr(product_structure, attr, None)
        source_value = getattr(source_structure, attr, None)
        if product_value is not None and source_value is not None and product_value != source_value:
            warnings.append(label)
    return warnings


def _refresh_provisor_substitute_prices_for_product(
    *,
    db: Session,
    product: Product,
    price_format_code: str | None = None,
    region_id: int | None = None,
) -> tuple[list[int], list[dict]]:
    started_at = time.perf_counter()
    substitutes = (
        db.execute(
            select(ProductSubstituteMatch)
            .where(ProductSubstituteMatch.product_id == product.id)
            .where(ProductSubstituteMatch.source_type == "provisor")
            .where(ProductSubstituteMatch.status == "approved")
            .where(ProductSubstituteMatch.source_goods_id.is_not(None))
            .order_by(ProductSubstituteMatch.priority.asc(), ProductSubstituteMatch.id.asc())
        )
        .scalars()
        .all()
    )
    selected_lists: list[CompetitorPriceList]
    touched_price_format_ids: list[int]
    if price_format_code:
        pf = db.execute(select(PriceFormat).where(PriceFormat.code == price_format_code.strip())).scalars().first()
        if pf is not None:
            assigned = get_assigned_competitor_price_lists(db=db, price_format_id=int(pf.id))
            selected_lists = [item.price_list for item in assigned if item.price_list.source_type == "provisor"]
            touched_price_format_ids = [int(pf.id)]
        else:
            selected_lists = []
            touched_price_format_ids = []
    else:
        assigned = get_all_assigned_competitor_price_lists(db=db)
        selected_lists = [item.price_list for item in assigned if item.price_list.source_type == "provisor"]
        touched_price_format_ids = sorted({int(item.assignment.price_format_id) for item in assigned})
    selected_sources = [f"{row.source_type}:{row.source_key}" for row in selected_lists]
    if selected_sources:
        db.execute(
            delete(CompetitorPrice)
            .where(CompetitorPrice.product_id == product.id)
            .where(CompetitorPrice.source_name.in_(selected_sources))
        )

    primary_goods_id = product.provisor_goods_id
    inserted = 0
    for price_list in selected_lists:
        source_name = f"{price_list.source_type}:{price_list.source_key}"
        chosen_item = None
        match_type = ""
        if primary_goods_id:
            chosen_item = (
                db.execute(
                    select(CompetitorPriceListItem)
                    .where(CompetitorPriceListItem.price_list_id == price_list.id)
                    .where(CompetitorPriceListItem.provisor_goods_id == primary_goods_id)
                    .where(CompetitorPriceListItem.distributor_price.is_not(None))
                    .order_by(CompetitorPriceListItem.distributor_price.asc(), CompetitorPriceListItem.id.asc())
                    .limit(1)
                )
                .scalars()
                .first()
            )
            if chosen_item is not None:
                match_type = "provisor_goods_id"
                if substitutes:
                    logger.info(
                        "[PROVISOR_SUBSTITUTE_NOT_USED] reason=primary_found product_id=%s product_sku=%s primary_goods_id=%s price_list_id=%s",
                        product.id,
                        product.code,
                        primary_goods_id,
                        price_list.id,
                    )

        if chosen_item is None:
            for substitute in substitutes:
                substitute_goods_id = int(substitute.source_goods_id)
                candidate = (
                    db.execute(
                        select(CompetitorPriceListItem)
                        .where(CompetitorPriceListItem.price_list_id == price_list.id)
                        .where(CompetitorPriceListItem.provisor_goods_id == substitute_goods_id)
                        .where(CompetitorPriceListItem.distributor_price.is_not(None))
                        .order_by(CompetitorPriceListItem.distributor_price.asc(), CompetitorPriceListItem.id.asc())
                        .limit(1)
                    )
                    .scalars()
                    .first()
                )
                if candidate is None:
                    continue
                chosen_item = candidate
                match_type = "provisor_manual_substitute"
                logger.info(
                    "[PROVISOR_SUBSTITUTE_APPLIED] product_id=%s product_sku=%s primary_goods_id=%s substitute_goods_id=%s price=%s price_list_id=%s",
                    product.id,
                    product.code,
                    primary_goods_id or "",
                    substitute_goods_id,
                    chosen_item.distributor_price or "",
                    price_list.id,
                )
                break
            if chosen_item is None and substitutes:
                logger.info(
                    "[PROVISOR_SUBSTITUTE_NOT_USED] reason=no_substitute_price product_id=%s product_sku=%s primary_goods_id=%s price_list_id=%s",
                    product.id,
                    product.code,
                    primary_goods_id or "",
                    price_list.id,
                )

        if chosen_item is None or chosen_item.distributor_price is None:
            continue

        chosen_item.product_id = product.id
        chosen_item.match_type = match_type
        chosen_item.match_score = 100
        chosen_item.matched_sku = str(product.code or "")
        chosen_item.match_key = f"provisor:goods_id:{chosen_item.provisor_goods_id or ''}"
        db.add(
            CompetitorPrice(
                price_format_id=price_list.price_format_id,
                product_id=product.id,
                source_name=source_name,
                supplier=price_list.supplier or price_list.display_name,
                price_date=price_list.price_date,
                coefficient=1.0,
                source_price=float(chosen_item.distributor_price),
                match_type=match_type,
                source_item_id=chosen_item.id,
                source_goods_id=chosen_item.provisor_goods_id,
                source_distributor_goods_id=chosen_item.distributor_goods_id or "",
                source_manufacturer=chosen_item.raw_manufacturer or "",
            )
        )
        inserted += 1

    db.flush()
    elapsed_ms = round((time.perf_counter() - started_at) * 1000, 2)
    logger.info("[SUBSTITUTE_INCREMENTAL_REFRESH] product_id=%s elapsed_ms=%s", product.id, elapsed_ms)
    rows: list[dict] = []
    for price_format_id in touched_price_format_ids:
        pf = db.get(PriceFormat, price_format_id)
        if pf is None:
            continue
        rows.extend(
            get_products_with_competitor_top5(
                db=db,
                price_format_code=pf.code,
                region_id=region_id,
                product_id=product.id,
            )
        )
    return touched_price_format_ids, rows


@app.get("/api/products/{product_id}/substitutes")
def get_product_substitutes(product_id: int, db: Session = Depends(get_db)):
    product = db.get(Product, product_id)
    if product is None:
        raise HTTPException(status_code=404, detail="product not found")
    rows = (
        db.execute(
            select(ProductSubstituteMatch)
            .where(ProductSubstituteMatch.product_id == product_id)
            .order_by(ProductSubstituteMatch.priority.asc(), ProductSubstituteMatch.id.asc())
        )
        .scalars()
        .all()
    )
    return {
        "product": {
            "id": product.id,
            "sku": product.code,
            "name": product.name,
            "provisorGoodsId": product.provisor_goods_id,
        },
        "items": [_substitute_to_dict(row) for row in rows],
    }


@app.get("/api/provisor/items/search")
def search_provisor_items(
    q: str = Query(..., min_length=1),
    format_code: str | None = Query(None),
    limit: int = Query(20, ge=1, le=50),
    db: Session = Depends(get_db),
):
    query = q.strip()
    if not query:
        return []
    stmt = (
        select(CompetitorPriceListItem, CompetitorPriceList)
        .join(CompetitorPriceList, CompetitorPriceList.id == CompetitorPriceListItem.price_list_id)
        .where(CompetitorPriceList.source_type == "provisor")
    )
    if format_code:
        pf = db.execute(select(PriceFormat).where(PriceFormat.code == format_code.strip())).scalars().first()
        if pf is not None:
            assigned_ids = [int(item.price_list.id) for item in get_assigned_competitor_price_lists(db=db, price_format_id=int(pf.id))]
            if assigned_ids:
                stmt = stmt.where(CompetitorPriceList.id.in_(assigned_ids))

    if query.isdigit():
        stmt = stmt.where(
            (CompetitorPriceListItem.provisor_goods_id == int(query))
            | (CompetitorPriceListItem.distributor_goods_id.ilike(f"%{query}%"))
        )
    else:
        like = f"%{query}%"
        stmt = stmt.where(
            (CompetitorPriceListItem.name.ilike(like))
            | (CompetitorPriceListItem.distributor_goods_name.ilike(like))
            | (CompetitorPriceListItem.raw_name.ilike(like))
            | (CompetitorPriceListItem.raw_manufacturer.ilike(like))
        )
    rows = db.execute(stmt.order_by(CompetitorPriceListItem.id.desc()).limit(limit * 3)).all()
    seen: set[tuple[int | None, str]] = set()
    out: list[dict] = []
    for item, price_list in rows:
        key = (item.provisor_goods_id, item.distributor_goods_id or "")
        if key in seen:
            continue
        seen.add(key)
        out.append(
            {
                "id": item.id,
                "goodsId": item.provisor_goods_id,
                "distributorGoodsId": item.distributor_goods_id,
                "name": item.raw_name or item.name or item.distributor_goods_name,
                "manufacturer": item.raw_manufacturer,
                "price": float(item.distributor_price) if item.distributor_price is not None else None,
                "source": f"{price_list.source_type}:{price_list.source_key}",
                "sourceLabel": price_list.display_name or price_list.supplier or price_list.source_key,
            }
        )
        if len(out) >= limit:
            break
    return out


@app.post("/api/products/{product_id}/substitutes")
def create_product_substitute(product_id: int, payload: dict = Body(...), db: Session = Depends(get_db)):
    product = db.get(Product, product_id)
    if product is None:
        raise HTTPException(status_code=404, detail="product not found")
    goods_id = payload.get("sourceGoodsId") or payload.get("goodsId")
    try:
        source_goods_id = int(goods_id) if goods_id not in (None, "") else None
    except Exception:
        raise HTTPException(status_code=400, detail="sourceGoodsId must be int")
    if source_goods_id is None:
        raise HTTPException(status_code=400, detail="sourceGoodsId is required")
    source_name = str(payload.get("sourceName") or payload.get("name") or "").strip()
    source_manufacturer = str(payload.get("sourceManufacturer") or payload.get("manufacturer") or "").strip()
    source_distributor_goods_id = str(payload.get("sourceDistributorGoodsId") or payload.get("distributorGoodsId") or "").strip()

    extra = db.get(ProductExtra, product.id)
    warnings = []
    if source_goods_id != product.provisor_goods_id:
        warnings.append("другой goodsId")
    if source_distributor_goods_id and source_distributor_goods_id != str(product.code):
        warnings.append("другая SKU")
    if source_manufacturer and extra is not None and (extra.manufacturer or "").strip().lower() != source_manufacturer.lower():
        warnings.append("другой производитель")
    warnings.extend(_structure_warnings(product.name, source_name))

    row = (
        db.execute(
            select(ProductSubstituteMatch)
            .where(ProductSubstituteMatch.product_id == product.id)
            .where(ProductSubstituteMatch.source_type == "provisor")
            .where(ProductSubstituteMatch.source_goods_id == source_goods_id)
        )
        .scalars()
        .first()
    )
    if row is None:
        row = ProductSubstituteMatch(product_id=product.id, source_type="provisor", source_goods_id=source_goods_id)
        db.add(row)
    row.source_distributor_goods_id = source_distributor_goods_id
    row.source_name = source_name
    row.source_manufacturer = source_manufacturer
    row.status = str(payload.get("status") or "approved").strip() or "approved"
    row.priority = int(payload.get("priority") or 100)
    row.comment = str(payload.get("comment") or "").strip()
    row.updated_at = now_kz_naive()
    db.flush()

    region_id = None
    raw_region = payload.get("region")
    if raw_region not in (None, ""):
        try:
            region_id = int(raw_region)
        except Exception:
            region_id = None
    price_format_ids, updated_rows = _refresh_provisor_substitute_prices_for_product(
        db=db,
        product=product,
        price_format_code=str(payload.get("formatCode") or payload.get("format_code") or "").strip() or None,
        region_id=region_id,
    )
    db.commit()
    db.refresh(row)
    return {"item": _substitute_to_dict(row), "warnings": warnings, "updatedRows": updated_rows, "priceFormatIds": price_format_ids}


@app.post("/calculate-prices", response_model=CalculatePricesResponse)
def calculate_prices_endpoint(payload: CalculatePricesRequest, db: Session = Depends(get_db)):
    as_of = payload.activation_date or date.today()
    price_list_number = payload.price_list_number or f"{payload.price_format_code}_{as_of.isoformat()}"

    try:
        count = calculate_prices(
            db=db,
            price_format_code=payload.price_format_code,
            price_list_number=price_list_number,
            as_of=as_of,
            activation_date=payload.activation_date,
            user=payload.user,
            region_id=payload.region,
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    if count == 0:
        # Most common reason in a fresh environment: products not imported yet.
        raise HTTPException(
            status_code=400,
            detail="Нет товаров для расчёта. Сначала загрузите Excel с товарами (код/наименование/себестоимость) через /api/upload-excel.",
        )

    return CalculatePricesResponse(price_list_number=price_list_number, calculated_count=count)


@app.post("/api/calculate-prices", response_model=CalculatePricesResponse)
def calculate_prices_api(payload: CalculatePricesRequest, db: Session = Depends(get_db)):
    return calculate_prices_endpoint(payload=payload, db=db)


@app.post("/api/generate-price-list")
def generate_price_list(payload: dict = Body(...), db: Session = Depends(get_db)):
    price_format_code = payload.get("price_format_code") or payload.get("format_code")
    if not isinstance(price_format_code, str) or not price_format_code.strip():
        raise HTTPException(status_code=400, detail="price_format_code is required")
    price_format_code = price_format_code.strip()

    activation_date = None
    raw_activation = payload.get("activation_date")
    if isinstance(raw_activation, str) and raw_activation.strip():
        try:
            activation_date = date.fromisoformat(raw_activation.strip())
        except Exception:
            raise HTTPException(status_code=400, detail="activation_date must be ISO date (YYYY-MM-DD)")

    as_of = activation_date or date.today()
    price_list_number = payload.get("price_list_number")
    if not isinstance(price_list_number, str) or not price_list_number.strip():
        price_list_number = f"{price_format_code}_{as_of.isoformat()}"
    user = payload.get("user")
    user_s = user.strip() if isinstance(user, str) else ""

    region_id = None
    raw_region = payload.get("region")
    if raw_region is not None:
        try:
            region_id = int(raw_region)
        except Exception:
            raise HTTPException(status_code=400, detail="region must be int")
        if region_id <= 0:
            raise HTTPException(status_code=400, detail="region must be >= 1")

    try:
        pf_for_selection = (
            db.execute(select(PriceFormat).where(PriceFormat.code == price_format_code)).scalars().first()
        )
        if pf_for_selection is None:
            raise HTTPException(status_code=400, detail=f"Формат прайса не найден: {price_format_code}")

        selected_count = len(get_assigned_competitor_price_lists(db=db, price_format_id=int(pf_for_selection.id)))
        if selected_count <= 0:
            raise HTTPException(
                status_code=400,
                detail="Выберите хотя бы один прайс-лист конкурента перед формированием прайса.",
            )

        calculated_count = calculate_prices(
            db=db,
            price_format_code=price_format_code,
            price_list_number=price_list_number,
            as_of=as_of,
            activation_date=activation_date,
            user=user_s,
            region_id=region_id,
        )
    except HTTPException:
        raise
    except ValueError as e:
        logger.exception("generate-price-list validation error for format=%s", price_format_code)
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        db.rollback()
        logger.exception(
            "generate-price-list failed: format=%s price_list=%s region=%s payload=%s",
            price_format_code,
            price_list_number,
            region_id,
            {k: v for k, v in payload.items() if k not in {"password", "token"}},
        )
        raise HTTPException(status_code=400, detail=f"Не удалось сформировать прайс: {e}")

    if calculated_count == 0:
        raise HTTPException(
            status_code=400,
            detail="Нет товаров для расчёта. Сначала загрузите Excel с товарами (код/наименование/себестоимость) через /api/upload-excel.",
        )

    pl = db.execute(select(PriceList).where(PriceList.number == price_list_number)).scalars().first()
    if not pl:
        logger.error("generate-price-list finished without PriceList row: %s", price_list_number)
        raise HTTPException(status_code=400, detail="Прайс не был создан. Проверьте настройки формата и выбранные прайс-листы.")

    items_rows = db.execute(
        select(CalculatedPrice, Product)
        .join(Product, Product.id == CalculatedPrice.product_id)
        .where(CalculatedPrice.price_list_id == pl.id)
        .order_by(Product.name.asc())
    ).all()

    competitor_view_by_sku = {
        str(row.get("sku") or ""): row
        for row in get_products_with_competitor_top5(
            db=db,
            price_format_code=price_format_code,
            price_list_number=price_list_number,
            region_id=region_id,
        )
    }

    items = [
        {
            "productId": p.id,
            "sku": p.code,
            "name": p.name,
            "price": float(cp.final_price),
            "cost": float(cp.cost),
            "competitorPrice": float(cp.competitor_price) if cp.competitor_price is not None else None,
            "competitors": {
                str(col.get("label") or col.get("source") or ""): prices[str(col.get("source"))]
                for col in (competitor_view_by_sku.get(str(p.code), {}).get("competitorColumns") or [])
                for prices in [competitor_view_by_sku.get(str(p.code), {}).get("competitorPricesBySource") or {}]
                if str(col.get("source")) in prices
            },
            "log": cp.applied_reason or "",
            "pricingCalculationLog": cp.applied_reason or "",
            "listOverrideLog": _list_override_log(_applied_list_summary(db, cp)),
            "zone": _calculated_zone(cp),
        }
        for (cp, p) in items_rows
    ]

    stats = {
        "lp": sum(1 for x in items if x.get("competitorPrice") is not None and x.get("zone") == "left"),
        "zl": sum(1 for x in items if x.get("competitorPrice") is not None and x.get("zone") == "optimal"),
        "pp": sum(1 for x in items if x.get("competitorPrice") is not None and x.get("zone") == "right"),
        "no_mck": sum(1 for x in items if x.get("competitorPrice") is None),
    }

    return {
        "price_list_id": price_list_number,
        "calculated_count": calculated_count,
        "items": items,
        "stats": stats,
    }


@app.get("/competitor-prices")
def competitor_prices(
    price_format_code: str = Query(...),
    product_code: str | None = Query(None),
    db: Session = Depends(get_db),
):
    pf = db.execute(select(PriceFormat).where(PriceFormat.code == price_format_code)).scalars().first()
    if not pf:
        raise HTTPException(status_code=404, detail="price format not found")

    stmt = select(Product.code, CompetitorPrice).join(
        Product, Product.id == CompetitorPrice.product_id, isouter=True
    ).where(CompetitorPrice.price_format_id == pf.id)

    if product_code:
        stmt = stmt.where(Product.code == product_code)

    rows = db.execute(stmt).all()
    return [
        {
            "product_code": code,
            "source_name": cp.source_name,
            "supplier": cp.supplier,
            "coefficient": float(cp.coefficient or 1.0),
            "source_price": float(cp.source_price) if cp.source_price is not None else None,
            "price_date": cp.price_date,
        }
        for (code, cp) in rows
    ]


@app.get("/price-list")
def get_price_list(price_list_number: str = Query(...), db: Session = Depends(get_db)):
    pl = db.execute(select(PriceList).where(PriceList.number == price_list_number)).scalars().first()
    if not pl:
        raise HTTPException(status_code=404, detail="price list not found")

    rows = db.execute(
        select(CalculatedPrice, Product)
        .join(Product, Product.id == CalculatedPrice.product_id)
        .where(CalculatedPrice.price_list_id == pl.id)
        .order_by(Product.name.asc())
    ).all()

    return [
        {
            "product": p.name,
            "price": float(cp.final_price),
            "cost": float(cp.cost),
            "zone": _calculated_zone(cp),
        }
        for (cp, p) in rows
    ]


@app.get("/analytics")
def analytics(price_list_number: str = Query(...), db: Session = Depends(get_db)):
    # reuse analysis response shape for now
    return get_price_list_analysis(price_list_id=price_list_number, db=db)


@app.get("/api/universal-lists")
def get_universal_lists(db: Session = Depends(get_db)):
    rows = db.execute(select(UniversalList).order_by(UniversalList.created_at.desc())).scalars().all()
    if not rows:
        # Заглушка: отдаём как в текущем UI
        return [
            {
                "id": 1,
                "name": "Прямые контракты",
                "type": "Фикс цена",
                "status": "Активен",
                "period": "01.03.2026 - 31.03.2026",
                "itemsCount": 0,
            },
            {
                "id": 2,
                "name": "Ограничения сверху",
                "type": "Макс. наценка",
                "status": "Неактивен",
                "period": "01.01.2026 - 01.01.2027",
                "itemsCount": 0,
            },
        ]

    counts = dict(
        db.execute(
            select(ListItem.universal_list_id, func.count(ListItem.id))
            .group_by(ListItem.universal_list_id)
        ).all()
    )

    def _map_status(s: str) -> str:
        s_l = (s or "").strip().lower()
        if s_l.startswith("актив"):
            return "Активен"
        if s_l.startswith("черн"):
            return "Черновик"
        return "Неактивен"

    return [
        {
            "id": ul.id,
            "name": ul.name,
            "type": ul.type,
            "status": _map_status(ul.status),
            "period": f"{_fmt_d(ul.start_date)} - {_fmt_d(ul.end_date)}".strip(),
            "itemsCount": int(counts.get(ul.id, 0)),
        }
        for ul in rows
    ]


@app.post("/api/universal-lists", response_model=CreateUniversalListResponse)
def create_universal_list(payload: CreateUniversalListRequest = Body(...), db: Session = Depends(get_db)):
    price_format_id: int | None = None
    if payload.price_format_code:
        pf = (
            db.execute(select(PriceFormat).where(PriceFormat.code == payload.price_format_code))
            .scalars()
            .first()
        )
        if not pf:
            raise HTTPException(status_code=404, detail="price format not found")
        price_format_id = pf.id

    ul = UniversalList(
        code=None,
        name=payload.name.strip(),
        status=payload.status.strip(),
        type=payload.type.strip(),
        start_date=payload.start_date,
        end_date=payload.end_date,
        price_format_id=price_format_id,
    )

    db.add(ul)
    db.flush()  # получаем ul.id

    if not ul.code:
        ul.code = f"UL_{ul.id:06d}"

    db.commit()
    return CreateUniversalListResponse(id=ul.id)


@app.delete("/api/universal-lists/{list_id}")
def delete_universal_list(list_id: int, db: Session = Depends(get_db)):
    ul = db.execute(select(UniversalList).where(UniversalList.id == list_id)).scalars().first()
    if not ul:
        raise HTTPException(status_code=404, detail="list not found")

    logger.info("[DELETE_LIST] list_id=%s", ul.id)
    calculated_prices_before = int(
        db.scalar(select(func.count(CalculatedPrice.id)).where(CalculatedPrice.applied_list_id == ul.id))
        or 0
    )
    logger.info("[DELETE_LIST] list_id=%s calculated_prices_before=%s", ul.id, calculated_prices_before)

    update_result = db.execute(
        update(CalculatedPrice)
        .where(CalculatedPrice.applied_list_id == ul.id)
        .values(applied_list_id=None)
        .execution_options(synchronize_session=False)
    )
    rows_updated = int(update_result.rowcount or 0)
    db.flush()

    calculated_prices_after = int(
        db.scalar(select(func.count(CalculatedPrice.id)).where(CalculatedPrice.applied_list_id == ul.id))
        or 0
    )
    logger.info(
        "[DELETE_LIST] list_id=%s rows_updated=%s calculated_prices_after=%s",
        ul.id,
        rows_updated,
        calculated_prices_after,
    )
    if calculated_prices_after != 0:
        logger.error(
            "[DELETE_LIST] list_id=%s delete_started=false calculated_prices_after=%s",
            ul.id,
            calculated_prices_after,
        )
        db.rollback()
        raise HTTPException(
            status_code=409,
            detail=f"Cannot delete list {ul.id}: calculated_prices still references it ({calculated_prices_after} rows).",
        )

    impacted_prices = (
        db.execute(
            select(CalculatedPrice)
            .where(
                CalculatedPrice.applied_list_ids.like(f"%{ul.id}%")
            )
        )
        .scalars()
        .all()
    )
    for row in impacted_prices:
        try:
            raw_ids = json.loads(row.applied_list_ids or "[]")
            if isinstance(raw_ids, list):
                cleaned_ids = [item for item in raw_ids if str(item) != str(ul.id)]
                row.applied_list_ids = json.dumps(cleaned_ids, ensure_ascii=False)
        except Exception:
            row.applied_list_ids = "[]"

    db.execute(delete(UniversalListPriceFormat).where(UniversalListPriceFormat.universal_list_id == ul.id))
    db.execute(delete(ListItem).where(ListItem.universal_list_id == ul.id))
    logger.info("[DELETE_LIST] list_id=%s delete_started=true", ul.id)
    db.execute(delete(UniversalList).where(UniversalList.id == ul.id))
    db.commit()
    return {"status": "ok"}


@app.get("/api/universal-lists/{list_id}")
def get_universal_list_details(list_id: int, db: Session = Depends(get_db)):
    ul = db.execute(select(UniversalList).where(UniversalList.id == list_id)).scalars().first()
    if not ul:
        raise HTTPException(status_code=404, detail="list not found")

    items = db.execute(
        select(ListItem, Product)
        .join(Product, Product.id == ListItem.product_id)
        .where(ListItem.universal_list_id == ul.id)
        .order_by(Product.name.asc())
    ).all()

    status_l = (ul.status or "").strip().lower()
    status_ui = "Активен" if status_l.startswith("актив") else "Неактивен"

    return {
        "id": ul.id,
        "name": ul.name,
        "type": ul.type,
        "status": status_ui,
        "period": {"start": _fmt_d(ul.start_date), "end": _fmt_d(ul.end_date)},
        "items": [
            {
                "code": p.code,
                "name": p.name,
                "value": f"{float(li.value):.2f}",
            }
            for (li, p) in items
        ],
        "linkedPriceLists": [],
    }


@app.get("/api/price-lists/{price_list_id}/export.csv")
def export_price_list_csv(
    price_list_id: str,
    search: str | None = Query(None),
    top_filter: str = Query("all"),
    top_sort: str = Query("default"),
    db: Session = Depends(get_db),
):
    pl, _pf, rows, competitor_columns = _generated_price_export_rows(
        db,
        price_list_id,
        search=search,
        top_filter=top_filter,
        top_sort=top_sort,
    )

    buffer = io.StringIO()
    writer = csv.writer(buffer, delimiter=",", lineterminator="\n")
    writer.writerow(
        [
            "Номер прайс-листа",
            "SKU",
            "Название",
            "Остатки",
            "Производитель",
            "Себестоимость",
        ]
        + ["ТОП"]
        + [str(col.get("label") or col.get("source") or "") for col in competitor_columns]
        + ["Модельная цена", "Логи"]
    )

    for row in rows:
        by_source = row.get("competitorPricesBySource") if isinstance(row.get("competitorPricesBySource"), dict) else {}
        writer.writerow(
            [
                pl.number,
                row.get("sku") or "",
                row.get("name") or "",
                "" if row.get("stock") is None else row.get("stock"),
                row.get("manufacturer") or "",
                "" if row.get("costPrice") is None else row.get("costPrice"),
                "" if _top_rank_value(row) is None else _top_rank_value(row),
            ]
            + ["" if by_source.get(str(col.get("source"))) is None else by_source.get(str(col.get("source"))) for col in competitor_columns]
            + [
                "" if row.get("modelPrice") is None else row.get("modelPrice"),
                row.get("modelLog") or "",
            ]
        )

    buffer.seek(0)
    filename = f"{pl.number}.csv"
    return StreamingResponse(
        iter([buffer.getvalue()]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename=\"{filename}\""},
    )


@app.get("/api/price-lists/{price_list_id}/export.xlsx")
def export_price_list_xlsx(
    price_list_id: str,
    search: str | None = Query(None),
    top_filter: str = Query("all"),
    top_sort: str = Query("default"),
    db: Session = Depends(get_db),
):
    pl, _pf, rows, competitor_columns = _generated_price_export_rows(
        db,
        price_list_id,
        search=search,
        top_filter=top_filter,
        top_sort=top_sort,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Прайс"
    ws.append(
        [
            "Номер прайс-листа",
            "SKU",
            "Название",
            "Остатки",
            "Производитель",
            "Себестоимость",
        ]
        + ["ТОП"]
        + [str(col.get("label") or col.get("source") or "") for col in competitor_columns]
        + ["Модельная цена", "Логи"]
    )
    for row in rows:
        by_source = row.get("competitorPricesBySource") if isinstance(row.get("competitorPricesBySource"), dict) else {}
        ws.append(
            [
                pl.number,
                row.get("sku") or "",
                row.get("name") or "",
                row.get("stock"),
                row.get("manufacturer") or "",
                row.get("costPrice"),
                _top_rank_value(row),
            ]
            + [by_source.get(str(col.get("source"))) for col in competitor_columns]
            + [
                row.get("modelPrice"),
                row.get("modelLog") or "",
            ]
        )

    if ws.max_row >= 2:
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        ws.conditional_formatting.add(
            f"G2:G{ws.max_row}",
            CellIsRule(operator="between", formula=["1", "100"], fill=green_fill),
        )

    bio = io.BytesIO()
    wb.save(bio)
    filename = f"{pl.number}.xlsx"
    return StreamingResponse(
        iter([bio.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=\"{filename}\""},
    )


@app.get("/api/phcenter/prices-analysis")
async def phcenter_prices_analysis(
    region: int = Query(...),
    price_mode: int = Query(...),
    distributors: int = Query(...),
):
    token = settings.phcenter_token
    if not token:
        raise HTTPException(status_code=500, detail="PHCENTER_TOKEN is not configured")

    authorization = token if token.lower().startswith("bearer ") else f"Bearer {token}"

    url = f"{settings.phcenter_base_url.rstrip('/')}/api/Report/PricesAnalysis"

    async with httpx.AsyncClient(timeout=30.0) as client:
        resp = await client.get(
            url,
            params={"region": region, "price_mode": price_mode, "distributors": distributors},
            headers={"Authorization": authorization},
        )

    if resp.status_code >= 400:
        raise HTTPException(status_code=resp.status_code, detail=resp.text)

    # Возвращаем как есть (обычно JSON)
    try:
        return resp.json()
    except Exception:
        return {"raw": resp.text}


@app.get("/{full_path:path}", include_in_schema=False)
def serve_frontend(full_path: str):
    # Serve SPA build when present (Railway / production). In dev, Vite serves UI.
    dist = _frontend_dist_dir()
    if dist is None:
        raise HTTPException(status_code=404, detail="frontend is not built")

    # Don't turn missing API endpoints into index.html.
    if full_path.startswith("api/") or full_path in {"openapi.json", "docs", "redoc", "health"}:
        raise HTTPException(status_code=404, detail="not found")

    requested = (dist / full_path).resolve()
    # Prevent path traversal.
    if dist not in requested.parents and requested != dist:
        raise HTTPException(status_code=404, detail="not found")

    if requested.is_file():
        return FileResponse(requested)

    index = dist / "index.html"
    if index.exists():
        return FileResponse(index)

    raise HTTPException(status_code=404, detail="index.html not found")
