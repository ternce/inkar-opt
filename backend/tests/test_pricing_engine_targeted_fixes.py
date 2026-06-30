from datetime import date
from decimal import Decimal
import csv
import io

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from backend.app.db import Base
from backend.app.deps import get_db
from backend.app.models import (
    BendRange,
    CalculatedPrice,
    CompetitorPrice,
    CompetitorPriceList,
    CompetitorPriceListItem,
    CompetitorPricePercentile,
    ListItem,
    MarkupRange,
    NoCompetitorMarkupRange,
    PriceFormatCompetitorAssignment,
    PriceFormat,
    PriceList,
    Product,
    ProductExtra,
    ProductRating,
    RoundingRule,
    UniversalList,
    UniversalListPriceFormat,
)
from backend.app.services.competitors.code_mappings import list_catalog_code_mappings
from backend.app.services.competitors.percentiles.read_models import (
    list_percentile_product_rows,
    list_percentile_sources,
    percentile_coverage_audit,
)
from backend.app.main import _competitor_column_title, _generated_item_dict, app
from backend.app.services.competitor_percentiles import KAZAKHSTAN_REGION, KAZAKHSTAN_SCOPE, REGIONAL_SCOPE, recalculate_competitor_percentiles
from backend.app.services.competitor_source_config import MULTI_PRICE_PERCENTILE_MODE
from backend.app.services.pricing import calculate_price_for_product, calculate_price_zone, calculate_prices
from backend.app.services.pricing_workflow.analytics import build_workflow_analytics
from backend.app.services.pricing_workflow.exports import _export_zone


def _session():
    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)
    return sessionmaker(bind=engine)()


def _format(db, *, rounding=None):
    pf = PriceFormat(code="0001", name="0001", branch="Almaty")
    db.add(pf)
    db.flush()
    if rounding is not None:
        pf.rounding_rule_id = rounding.id
    db.add(MarkupRange(price_format_id=pf.id, cost_from=0, cost_to=None, markup_percent=0.15))
    db.add(NoCompetitorMarkupRange(price_format_id=pf.id, cost_from=0, cost_to=None, markup_percent=0.03))
    db.add(BendRange(price_format_id=pf.id, price_from=0, bend_percent=0.5))
    db.add(BendRange(price_format_id=pf.id, price_from=2000, bend_percent=0.2))
    db.add(BendRange(price_format_id=pf.id, price_from=5000, bend_percent=0.1))
    db.flush()
    return pf


def _product(db, *, code="SKU-1", cost=2500):
    row = Product(code=code, name=code, cost=cost)
    db.add(row)
    db.flush()
    return row


def _competitor(db, pf, product, source, price):
    db.add(CompetitorPrice(price_format_id=pf.id, product_id=None, source_name=source, supplier=source, coefficient=1))
    db.add(CompetitorPrice(price_format_id=pf.id, product_id=product.id, source_name=source, supplier=source, source_price=price, coefficient=1))
    db.flush()


def _list_item(db, product, list_type, value, *, pf=None, status="active", start_date=None, end_date=None):
    row = UniversalList(
        code=f"UL-{list_type}-{product.code}",
        name=f"{list_type} list",
        type=list_type,
        status=status,
        start_date=start_date,
        end_date=end_date,
        price_format_id=pf.id if pf is not None else None,
    )
    db.add(row)
    db.flush()
    db.add(ListItem(universal_list_id=row.id, product_id=product.id, value=value))
    db.flush()
    return row


@pytest.mark.parametrize(
    ("raw", "expected"),
    [
        ("Медсервис (Костанай) — Медсервис (Костанай) — Aksai4/83", "Медсервис (Костанай) — Aksai4/83"),
        ("Стофарм средняя цена (Астана) — Стофарм средняя цена (Астана) — magMereke", "Стофарм средняя цена (Астана) — magMereke"),
        ("Optimus pharm (Астана) — Optimus pharm (Астана) — 8mkr8a", "Optimus pharm (Астана) — 8mkr8a"),
        ("Источник — Другой филиал — account", "Источник — Другой филиал — account"),
    ],
)
def test_competitor_column_title_removes_only_adjacent_duplicates(raw, expected):
    assert _competitor_column_title(raw) == expected


def test_bend_is_selected_by_competitor_price_not_product_cost():
    db = _session()
    pf = _format(db)
    product = _product(db, cost=2500)
    _competitor(db, pf, product, "manual:expensive", 9000)

    price, debug = calculate_price_for_product(db=db, product=product, price_format=pf, as_of=date.today())

    assert float(debug["bend_percent_used"]) == 0.1
    assert float(debug["chosen_competitor_price"]) == 9000
    assert float(price) == 8991.0


def test_competitor_iteration_chooses_first_candidate_at_or_above_mdc():
    db = _session()
    rounding = RoundingRule(code="R01", name="R01", mode="math", precision=2, step=0.01)
    db.add(rounding)
    db.flush()
    pf = _format(db, rounding=rounding)
    product = _product(db, cost=2500)
    _competitor(db, pf, product, "manual:low", 2850)
    _competitor(db, pf, product, "manual:second", 2900)
    _competitor(db, pf, product, "manual:high", 3000)

    price, debug = calculate_price_for_product(db=db, product=product, price_format=pf, as_of=date.today())

    assert debug["chosen_competitor_source"] == "manual:high"
    assert debug["chosen_competitor_rank"] == 3
    assert float(price) == 2994.0
    assert debug["zone"] == "right"


def test_competitor_price_never_falls_below_mdc_after_rounding():
    db = _session()
    rounding = RoundingRule(code="DOWN", name="DOWN", mode="down", precision=2, step=0.01)
    db.add(rounding)
    db.flush()
    pf = _format(db, rounding=rounding)
    db.query(MarkupRange).filter(MarkupRange.price_format_id == pf.id).delete()
    db.add(MarkupRange(price_format_id=pf.id, cost_from=0, cost_to=None, markup_percent=0))
    product = _product(db, cost=100.009)
    _competitor(db, pf, product, "manual:low", 100)

    price, debug = calculate_price_for_product(db=db, product=product, price_format=pf, as_of=date.today())

    assert float(debug["base_price"]) == 100.009
    assert float(price) == 100.01
    assert price >= debug["base_price"]


def test_rounding_rule_from_price_format_is_used():
    db = _session()
    rounding = RoundingRule(code="UP05", name="UP05", mode="up", precision=2, step=0.05)
    db.add(rounding)
    db.flush()
    pf = _format(db, rounding=rounding)
    product = _product(db, cost=10)
    _competitor(db, pf, product, "manual:source", 12.01)

    price, _debug = calculate_price_for_product(db=db, product=product, price_format=pf, as_of=date.today())

    assert float(price) == 11.95


def test_no_competitor_candidate_below_mdc_is_bumped_to_mdc():
    db = _session()
    pf = _format(db)
    product = _product(db, cost=100)

    price, debug = calculate_price_for_product(db=db, product=product, price_format=pf, as_of=date.today())

    assert debug["reason"] == "no_competitor_markup_bumped_to_mdc"
    assert float(price) == 117.65
    assert price >= debug["base_price"]


def test_no_competitor_candidate_above_mdc_is_kept():
    db = _session()
    pf = _format(db)
    db.query(NoCompetitorMarkupRange).filter(NoCompetitorMarkupRange.price_format_id == pf.id).delete()
    db.add(NoCompetitorMarkupRange(price_format_id=pf.id, cost_from=0, cost_to=None, markup_percent=0.20))
    db.flush()
    product = _product(db, cost=100)

    price, debug = calculate_price_for_product(db=db, product=product, price_format=pf, as_of=date.today())

    assert debug["reason"] == "no_competitor_markup"
    assert float(price) == 125.0
    assert debug["zone"] is None
    assert debug["log"] == "Нет цен выбранных конкурентов. Применена глобальная шкала маржи для товаров без конкурентов."
    assert "20" not in debug["log"]


def test_no_competitor_list_override_keeps_percentage_only_in_second_log():
    db = _session()
    pf = _format(db)
    product = _product(db, cost=100)
    row = _list_item(db, product, "fixed_markup", 10.5, pf=pf)

    calculate_prices(
        db=db,
        price_format_code=pf.code,
        price_list_number="PL-NO-COMP-LIST-LOG",
        as_of=date.today(),
        activation_date=None,
        user="test",
        force_new_price_list=True,
    )

    cp = db.query(CalculatedPrice).one()
    payload = _generated_item_dict(db, cp, product, pf, None)
    assert "%" not in payload["pricingCalculationLog"]
    assert payload["listOverrideLog"]["listType"] == "fixed_markup"
    assert payload["listOverrideLog"]["displayValue"] == "10.5%"
    assert payload["appliedListId"] == row.id


@pytest.mark.parametrize(
    ("cost", "margin_fraction", "expected_mdc"),
    [
        (250, 0.25, 333.33),
        (1000, 0.25, 1333.33),
    ],
)
def test_global_margin_range_uses_margin_formula(cost, margin_fraction, expected_mdc):
    db = _session()
    pf = _format(db)
    db.query(MarkupRange).filter(MarkupRange.price_format_id == pf.id).delete()
    db.add(MarkupRange(price_format_id=pf.id, cost_from=0, cost_to=None, markup_percent=margin_fraction))
    db.flush()
    product = _product(db, cost=cost)

    _price, debug = calculate_price_for_product(db=db, product=product, price_format=pf, as_of=date.today())

    assert round(float(debug["mdc_price"]), 2) == expected_mdc


def test_fixed_markup_25_uses_margin_formula_for_mdc():
    db = _session()
    pf = _format(db)
    product = _product(db, cost=250)
    _list_item(db, product, "fixed_markup", 25, pf=pf)

    _price, debug = calculate_price_for_product(db=db, product=product, price_format=pf, as_of=date.today())

    assert float(debug["mdc_markup_percent"]) == 25.0
    assert round(float(debug["mdc_price"]), 2) == 333.33


def test_fixed_price_list_overrides_generated_price():
    db = _session()
    pf = _format(db)
    product = _product(db, cost=100)
    row = _list_item(db, product, "fixed_price", 1200, pf=pf)

    price, debug = calculate_price_for_product(db=db, product=product, price_format=pf, as_of=date.today())

    assert float(price) == 1200
    assert debug["reason"] == "fixed_price_list"
    assert debug["applied_list_ids"] == [row.id]


def test_fixed_price_2500_is_final_price_not_percent_markup():
    db = _session()
    pf = _format(db)
    product = _product(db, cost=100)
    row = _list_item(db, product, "fixed_price", 2500, pf=pf)

    price, debug = calculate_price_for_product(db=db, product=product, price_format=pf, as_of=date.today())

    assert float(price) == 2500.0
    assert float(debug["mdc_markup_percent"]) == 15.0
    assert round(float(debug["mdc_price"]), 2) == 117.65
    assert debug["applied_rule_type"] == "fixed_price"
    assert debug["applied_list_ids"] == [row.id]


def test_fixed_price_zone_uses_lowest_available_competitor_when_chosen_is_empty():
    db = _session()
    rounding = RoundingRule(code="R01", name="R01", mode="math", precision=2, step=0.01)
    db.add(rounding)
    db.flush()
    pf = _format(db, rounding=rounding)
    product = _product(db, cost=100, code="FIXED-ZONE")
    _list_item(db, product, "fixed_price", 243.39, pf=pf)
    db.add(
        CompetitorPrice(
            price_format_id=pf.id,
            product_id=product.id,
            source_name="manual:available",
            supplier="manual:available",
            source_price=244.43,
            coefficient=1,
        )
    )
    db.flush()

    price, debug = calculate_price_for_product(db=db, product=product, price_format=pf, as_of=date.today())

    assert float(price) == 243.39
    assert debug["competitor_price"] is None
    assert debug["chosen_competitor_price"] is None
    assert debug["selected_competitor_price"] is None
    assert float(debug["lowest_competitor_price"]) == 244.43
    assert debug["zone"] == "left"


def test_fixed_markup_zone_uses_lowest_available_competitor_when_list_bypasses_selection():
    db = _session()
    rounding = RoundingRule(code="R-FM-ZONE", name="R-FM-ZONE", mode="math", precision=2, step=0.01)
    db.add(rounding)
    db.flush()
    pf = _format(db, rounding=rounding)
    product = _product(db, cost=100, code="FIXED-MARKUP-ZONE")
    _list_item(db, product, "fixed_markup", 10, pf=pf)
    db.add(
        CompetitorPrice(
            price_format_id=pf.id,
            product_id=product.id,
            source_name="manual:available",
            supplier="manual:available",
            source_price=115,
            coefficient=1,
        )
    )
    db.flush()

    price, debug = calculate_price_for_product(db=db, product=product, price_format=pf, as_of=date.today())

    assert round(float(price), 2) == 111.11
    assert debug["competitor_price"] is None
    assert debug["chosen_competitor_price"] is None
    assert float(debug["lowest_competitor_price"]) == 115.0
    assert debug["zone"] == "left"


def test_fixed_markup_list_overrides_default_markup():
    db = _session()
    pf = _format(db)
    product = _product(db, cost=100)
    row = _list_item(db, product, "fixed_markup", 5, pf=pf)

    price, debug = calculate_price_for_product(db=db, product=product, price_format=pf, as_of=date.today())

    assert float(price) == 105.26
    assert debug["reason"] == "fixed_markup_list_final"
    assert debug["applied_list_ids"] == [row.id]
    assert debug["applied_rule_type"] == "fixed_markup"
    assert float(debug["mdc_markup_percent"]) == 5.0
    assert float(debug["mdc_price"]) == 105.26
    assert float(debug["final_price"]) == 105.26
    assert debug["chosen_competitor_price"] is None
    assert debug["competitor_candidate_price"] is None


def test_percentile_rebuild_includes_products_without_competitors():
    db = _session()
    pf = _format(db)
    with_competitor = _product(db, code="PCT-WITH", cost=100)
    without_competitor = _product(db, code="PCT-WITHOUT", cost=100)
    price_list = CompetitorPriceList(
        price_format_id=pf.id,
        source_type="manual",
        source_key="pct-source",
        display_name="Percentile Source",
        supplier="Percentile Source",
        branch_name="Almaty",
        competitor_name="Competitor A",
    )
    db.add(price_list)
    db.flush()
    db.add(
        PriceFormatCompetitorAssignment(
            price_format_id=pf.id,
            competitor_price_list_id=price_list.id,
            is_active=True,
            coefficient=1,
        )
    )
    db.add(
        CompetitorPriceListItem(
            price_list_id=price_list.id,
            product_id=with_competitor.id,
            distributor_price=150,
        )
    )
    db.flush()

    summary = recalculate_competitor_percentiles(db=db, price_format_id=pf.id)
    with_rows = (
        db.query(CompetitorPricePercentile)
        .filter(CompetitorPricePercentile.price_format_id == pf.id)
        .filter(CompetitorPricePercentile.product_id == with_competitor.id)
        .filter(CompetitorPricePercentile.percentile_scope == REGIONAL_SCOPE)
        .all()
    )
    without_rows = (
        db.query(CompetitorPricePercentile)
        .filter(CompetitorPricePercentile.price_format_id == pf.id)
        .filter(CompetitorPricePercentile.product_id == without_competitor.id)
        .filter(CompetitorPricePercentile.percentile_scope == REGIONAL_SCOPE)
        .all()
    )

    assert summary["rows_created"] == 20
    assert summary["products_processed"] == 2
    assert summary["products_with_competitors"] == 1
    assert summary["products_without_competitors"] == 1
    assert len(with_rows) == 5
    assert len(without_rows) == 5
    assert {float(row.value) for row in with_rows} == {150.0}
    assert all(row.value is None for row in without_rows)


def test_percentile_source_sku_count_matches_full_product_scope():
    db = _session()
    pf = _format(db)
    with_competitor = _product(db, code="PCT-SOURCE-WITH", cost=100)
    _product(db, code="PCT-SOURCE-WITHOUT", cost=100)
    price_list = CompetitorPriceList(
        price_format_id=pf.id,
        source_type="manual",
        source_key="pct-source-count",
        display_name="Percentile Count Source",
        supplier="Percentile Count Source",
        branch_name="Almaty",
        competitor_name="Competitor B",
    )
    db.add(price_list)
    db.flush()
    db.add(
        PriceFormatCompetitorAssignment(
            price_format_id=pf.id,
            competitor_price_list_id=price_list.id,
            is_active=True,
            coefficient=1,
        )
    )
    db.add(
        CompetitorPriceListItem(
            price_list_id=price_list.id,
            product_id=with_competitor.id,
            distributor_price=200,
        )
    )
    db.flush()

    recalculate_competitor_percentiles(db=db, price_format_id=pf.id)

    sources = list_percentile_sources(db=db, price_format_code=pf.code)
    regional_sources = [source for source in sources if source.get("scope") == REGIONAL_SCOPE]

    assert len(regional_sources) == 5
    assert {source["skuCount"] for source in regional_sources} == {2}
    assert {source["percentile"] for source in regional_sources} == {10, 20, 30, 40, 60}


def test_percentile_rebuild_uses_latest_duplicate_row_per_account_sku():
    db = _session()
    pf = _format(db)
    product = _product(db, code="PCT-DUP", cost=100)
    price_list = CompetitorPriceList(
        price_format_id=pf.id,
        source_type="manual",
        source_key="pct-dup",
        display_name="Account 1",
        supplier="Regular Competitor",
        branch_name="Almaty",
        competitor_name="Regular Competitor",
        account_login="Account 1",
    )
    db.add(price_list)
    db.flush()
    db.add(PriceFormatCompetitorAssignment(price_format_id=pf.id, competitor_price_list_id=price_list.id, is_active=True))
    db.add(CompetitorPriceListItem(price_list_id=price_list.id, product_id=product.id, distributor_price=100))
    db.flush()
    db.add(CompetitorPriceListItem(price_list_id=price_list.id, product_id=product.id, distributor_price=200))
    db.flush()

    recalculate_competitor_percentiles(db=db, price_format_id=pf.id)

    rows = (
        db.query(CompetitorPricePercentile)
        .filter(CompetitorPricePercentile.price_format_id == pf.id)
        .filter(CompetitorPricePercentile.product_id == product.id)
        .filter(CompetitorPricePercentile.branch_name == "Almaty")
        .filter(CompetitorPricePercentile.competitor_name == "Regular Competitor")
        .filter(CompetitorPricePercentile.percentile_scope == REGIONAL_SCOPE)
        .all()
    )
    assert {float(row.value) for row in rows} == {200.0}
    assert {row.source_count for row in rows} == {1}


def test_multi_price_percentile_source_uses_all_rows_for_one_sku():
    db = _session()
    pf = _format(db)
    product = _product(db, code="000000000001015510", cost=100)
    price_list = CompetitorPriceList(
        price_format_id=pf.id,
        source_type="manual",
        source_key="emit-almaty",
        display_name="Emit Almaty",
        supplier="Emit",
        branch_name="Almaty",
        competitor_name="Emit",
        account_login="emit",
    )
    db.add(price_list)
    db.flush()
    db.add(
        PriceFormatCompetitorAssignment(
            price_format_id=pf.id,
            competitor_price_list_id=price_list.id,
            is_active=True,
        )
    )
    for price in [7120, 7140, 7090, 7200, 7155, 7105]:
        db.add(CompetitorPriceListItem(price_list_id=price_list.id, product_id=product.id, distributor_price=price))
    db.flush()

    recalculate_competitor_percentiles(db=db, price_format_id=pf.id)

    rows = (
        db.query(CompetitorPricePercentile)
        .filter(CompetitorPricePercentile.price_format_id == pf.id)
        .filter(CompetitorPricePercentile.product_id == product.id)
        .filter(CompetitorPricePercentile.branch_name == "Almaty")
        .filter(CompetitorPricePercentile.competitor_name == "Emit")
        .filter(CompetitorPricePercentile.percentile_scope == REGIONAL_SCOPE)
        .all()
    )
    by_percentile = {row.percentile: float(row.value) for row in rows}
    assert by_percentile == {10: 7097.5, 20: 7105.0, 30: 7112.5, 40: 7120.0, 60: 7140.0}
    assert set(by_percentile.values()).isdisjoint({7090.0, 7155.0, 7200.0})
    assert {row.source_count for row in rows} == {1}
    assert {row.used_price_count for row in rows} == {6}
    assert {row.status for row in rows} == {"Calculated"}


def test_emit_multi_price_same_list_goods_id_keeps_all_prices():
    db = _session()
    pf = _format(db)
    product = _product(db, code="PCT-EMIT-GOODS", cost=100)
    product.provisor_goods_id = 163571
    price_list = CompetitorPriceList(
        price_format_id=pf.id,
        source_type="provisor",
        source_key="emit:1106",
        display_name="Emit International 1106",
        supplier="Emit International 1106",
        branch_name="Emit International 1106",
        competitor_name="Emit International 1106",
        account_login="emit",
    )
    db.add(price_list)
    db.flush()
    db.add(PriceFormatCompetitorAssignment(price_format_id=pf.id, competitor_price_list_id=price_list.id, is_active=True))
    for price in [8688.26, 8989.73, 9358.46]:
        db.add(
            CompetitorPriceListItem(
                price_list_id=price_list.id,
                product_id=product.id,
                provisor_goods_id=163571,
                filial_id=1106,
                distributor_price=price,
                matched_sku=product.code,
            )
        )
    db.flush()

    recalculate_competitor_percentiles(db=db, price_format_id=pf.id)

    rows = (
        db.query(CompetitorPricePercentile)
        .filter(CompetitorPricePercentile.price_format_id == pf.id)
        .filter(CompetitorPricePercentile.product_id == product.id)
        .filter(CompetitorPricePercentile.branch_name == "Emit International 1106")
        .filter(CompetitorPricePercentile.competitor_name == "Emit International 1106")
        .filter(CompetitorPricePercentile.percentile_scope == REGIONAL_SCOPE)
        .all()
    )
    by_percentile = {row.percentile: float(row.value) for row in rows}
    assert {row.used_price_count for row in rows} == {3}
    assert len(set(round(value, 4) for value in by_percentile.values())) > 1
    assert round(by_percentile[10], 3) == 8748.554
    assert round(by_percentile[60], 3) == 9063.476


def test_percentile_price_generation_uses_precomputed_rows_without_raw_rebuild(monkeypatch):
    import backend.app.services.pricing as pricing_service

    db = _session()
    pf = _format(db)
    pf.competitor_price_mode = "percentile"
    pf.percentile_number = 10
    product = _product(db, code="PCT-PRECOMPUTED", cost=100)
    db.add(
        CompetitorPricePercentile(
            price_format_id=pf.id,
            product_id=product.id,
            branch_name="Emit International 1106",
            competitor_name="Emit International 1106",
            percentile_scope=REGIONAL_SCOPE,
            percentile=10,
            value=120,
            source_count=1,
            price_count=3,
            used_price_count=3,
            status="Calculated",
        )
    )
    db.flush()

    def fail_raw_rebuild(**_kwargs):
        raise AssertionError("raw competitor_price_list_items rebuild should not run during percentile generation")

    monkeypatch.setattr(pricing_service, "rebuild_competitor_prices_for_selected", fail_raw_rebuild)

    count = calculate_prices(
        db=db,
        price_format_code=pf.code,
        price_list_number="PCT-PRECOMPUTED-PL",
        as_of=date.today(),
        activation_date=None,
        user="test",
        force_new_price_list=True,
    )

    assert count == 1
    calculated = db.query(CalculatedPrice).filter(CalculatedPrice.product_id == product.id).one()
    assert calculated.used_percentile is True
    assert float(calculated.lowest_competitor_price) == 120.0


def test_multi_price_percentile_ignores_invalid_prices_and_one_price_status():
    db = _session()
    pf = _format(db)
    product = _product(db, code="PCT-MULTI-INVALID", cost=100)
    price_list = CompetitorPriceList(
        price_format_id=pf.id,
        source_type="manual",
        source_key="future-source",
        display_name="Future Source",
        supplier="Future Source",
        branch_name="Almaty",
        competitor_name="Future Source",
    )
    db.add(price_list)
    db.flush()
    db.add(
        PriceFormatCompetitorAssignment(
            price_format_id=pf.id,
            competitor_price_list_id=price_list.id,
            is_active=True,
            percentile_mode=MULTI_PRICE_PERCENTILE_MODE,
        )
    )
    for price in [None, 0, -1, 777]:
        db.add(CompetitorPriceListItem(price_list_id=price_list.id, product_id=product.id, distributor_price=price))
    db.flush()

    recalculate_competitor_percentiles(db=db, price_format_id=pf.id)

    rows = (
        db.query(CompetitorPricePercentile)
        .filter(CompetitorPricePercentile.price_format_id == pf.id)
        .filter(CompetitorPricePercentile.product_id == product.id)
        .filter(CompetitorPricePercentile.branch_name == "Almaty")
        .filter(CompetitorPricePercentile.competitor_name == "Future Source")
        .filter(CompetitorPricePercentile.percentile_scope == REGIONAL_SCOPE)
        .all()
    )
    assert {float(row.value) for row in rows} == {777.0}
    assert {row.source_count for row in rows} == {1}
    assert {row.used_price_count for row in rows} == {1}
    assert {row.status for row in rows} == {"Calculated from one price"}


def test_emit_percentile_coverage_counts_all_matched_positive_products():
    db = _session()
    pf = _format(db)
    products = [_product(db, code=f"PCT-EMIT-COV-{idx}", cost=100) for idx in range(1, 4)]
    price_list = CompetitorPriceList(
        price_format_id=pf.id,
        source_type="provisor",
        source_key="emit:1108",
        display_name="Emit International 1108",
        supplier="Emit International 1108",
        branch_name="Emit International 1108",
        competitor_name="Emit International 1108",
        account_login="emit",
    )
    db.add(price_list)
    db.flush()
    db.add(PriceFormatCompetitorAssignment(price_format_id=pf.id, competitor_price_list_id=price_list.id, is_active=True))
    for idx, product in enumerate(products, start=1):
        product.provisor_goods_id = 200000 + idx
        db.add(
            CompetitorPriceListItem(
                price_list_id=price_list.id,
                product_id=product.id,
                provisor_goods_id=product.provisor_goods_id,
                filial_id=1108,
                distributor_price=1000 + idx,
                matched_sku=product.code,
            )
        )
    db.flush()

    recalculate_competitor_percentiles(db=db, price_format_id=pf.id)

    audit = percentile_coverage_audit(
        db=db,
        price_format_code=pf.code,
        region="Emit International 1108",
        competitor="Emit International 1108",
    )
    browser = list_percentile_product_rows(
        db=db,
        price_format_code=pf.code,
        region="Emit International 1108",
        competitor="Emit International 1108",
        percentile_filter="has_percentile",
        page_size=100,
    )

    assert audit["counts"]["rawEmitRowsImported"] == 3
    assert audit["counts"]["distinctMatchedProductIdsWithPositivePrice"] == 3
    assert audit["counts"]["productsPassedIntoPercentileCalculation"] == 3
    assert audit["counts"]["productsWithStoredPercentileRows"] == 3
    assert audit["dropReasons"]["positiveMatchedProductsMissingStoredPercentiles"] == 0
    assert browser["summary"]["productsWithPercentile"] == 3
    assert browser["total"] == 3


def test_kazakhstan_percentiles_are_calculated_from_regional_percentiles():
    db = _session()
    pf = _format(db)
    product = _product(db, code="PCT-KZ", cost=100)

    def add_source(region: str, account: str, prices: list[float], competitor: str = "Emiti") -> None:
        price_list = CompetitorPriceList(
            price_format_id=pf.id,
            source_type="manual",
            source_key=f"{region}-{account}",
            display_name=account,
            supplier=competitor,
            branch_name=region,
            competitor_name=competitor,
            account_login=account,
        )
        db.add(price_list)
        db.flush()
        db.add(PriceFormatCompetitorAssignment(price_format_id=pf.id, competitor_price_list_id=price_list.id, is_active=True))
        for price in prices:
            db.add(CompetitorPriceListItem(price_list_id=price_list.id, product_id=product.id, distributor_price=price))

    add_source("Almaty", "Account A", [100])
    add_source("Astana", "Account B", [300])
    add_source("Almaty", "Medservice Account", [1], competitor="Medservice")
    db.flush()

    recalculate_competitor_percentiles(db=db, price_format_id=pf.id)

    kz_rows = (
        db.query(CompetitorPricePercentile)
        .filter(CompetitorPricePercentile.price_format_id == pf.id)
        .filter(CompetitorPricePercentile.product_id == product.id)
        .filter(CompetitorPricePercentile.branch_name == KAZAKHSTAN_REGION)
        .filter(CompetitorPricePercentile.competitor_name == "Emiti")
        .filter(CompetitorPricePercentile.percentile_scope == KAZAKHSTAN_SCOPE)
        .all()
    )
    by_percentile = {row.percentile: float(row.value) for row in kz_rows}
    assert by_percentile[10] == 120.0
    assert by_percentile[20] == 140.0
    assert by_percentile[30] == 160.0
    assert by_percentile[40] == 180.0
    assert by_percentile[60] == 220.0
    assert {row.source_count for row in kz_rows} == {2}


def _percentile_client_with_rows():
    engine = create_engine("sqlite://", connect_args={"check_same_thread": False}, poolclass=StaticPool)
    Base.metadata.create_all(engine)
    Session = sessionmaker(bind=engine)

    def override_db():
        with Session() as db:
            yield db

    app.dependency_overrides[get_db] = override_db
    with Session() as db:
        pf = _format(db)
        pf.code = "PCT-API"
        pf.branch = "BRANCH-1"
        with_competitor = _product(db, code="PCT-API-WITH", cost=100)
        with_competitor.provisor_goods_id = 163571
        without_competitor = _product(db, code="PCT-API-WITHOUT", cost=100)
        db.add_all(
            [
                ProductExtra(product_id=with_competitor.id, manufacturer="Maker A"),
                ProductExtra(product_id=without_competitor.id, manufacturer="Maker B"),
                ProductRating(branch_id="", product_id=with_competitor.id, sku=with_competitor.code, rating_type="global", rating=10),
                ProductRating(branch_id="BRANCH-1", product_id=with_competitor.id, sku=with_competitor.code, rating_type="local", rating=20),
            ]
        )
        def add_source(region: str, competitor: str, account: str, price: float) -> None:
            price_list = CompetitorPriceList(
                price_format_id=pf.id,
                source_type="manual",
                source_key=f"{region}-{competitor}-{account}",
                display_name=account,
                supplier=competitor,
                branch_name=region,
                competitor_name=competitor,
                account_login=account,
            )
            db.add(price_list)
            db.flush()
            db.add(
                PriceFormatCompetitorAssignment(
                    price_format_id=pf.id,
                    competitor_price_list_id=price_list.id,
                    is_active=True,
                    coefficient=1,
                )
            )
            db.add(
                CompetitorPriceListItem(
                    price_list_id=price_list.id,
                    product_id=with_competitor.id,
                    provisor_goods_id=with_competitor.provisor_goods_id,
                    distributor_price=price,
                )
            )

        for idx, price in enumerate([705, 702.23, 710, 685, 690], start=1):
            add_source("Almaty", "Emiti", f"Apteka {idx} Emiti", price)
        add_source("Almaty", "Medservice", "Apteka 1 Medservice", 1)
        add_source("Astana", "Emiti", "Apteka 1 Emiti Astana", 999)
        recalculate_competitor_percentiles(db=db, price_format_id=pf.id)
        db.commit()
    return TestClient(app)


def test_percentile_rows_endpoint_returns_all_rows_and_counters():
    client = _percentile_client_with_rows()
    try:
        response = client.get("/api/competitors/percentile-rows?format_code=PCT-API&region=Almaty&competitor=Emiti&page_size=100")
        assert response.status_code == 200
        payload = response.json()
        by_sku = {row["sku"]: row for row in payload["items"]}

        assert payload["total"] == 2
        assert payload["selectedRegion"] == "Almaty"
        assert payload["selectedCompetitor"] == "Emiti"
        assert len(payload["priceColumns"]) == 5
        assert set(by_sku) == {"PCT-API-WITH", "PCT-API-WITHOUT"}
        pct = by_sku["PCT-API-WITH"]["percentiles"]
        assert round(float(pct["10"]), 3) == 687.0
        assert round(float(pct["20"]), 3) == 689.0
        assert round(float(pct["30"]), 3) == 692.446
        assert round(float(pct["40"]), 3) == 697.338
        assert round(float(pct["60"]), 3) == 703.338
        assert by_sku["PCT-API-WITH"]["competitorCount"] == 5
        assert sorted(v for v in by_sku["PCT-API-WITH"]["branchPrices"].values() if v is not None) == [685, 690, 702.23, 705, 710]
        assert by_sku["PCT-API-WITH"]["status"] == "Рассчитан"
        assert all(value is None for value in by_sku["PCT-API-WITHOUT"]["percentiles"].values())
        assert by_sku["PCT-API-WITHOUT"]["competitorCount"] == 0
        assert by_sku["PCT-API-WITHOUT"]["status"] == "Нет данных"
        assert payload["summary"] == {
            "totalProducts": 2,
            "productsWithPercentile": 1,
            "productsWithoutPercentile": 1,
            "productsWithCompetitors": 1,
            "productsWithoutCompetitors": 1,
            "coveragePercent": 50.0,
        }

        medservice = client.get("/api/competitors/percentile-rows?format_code=PCT-API&region=Almaty&competitor=Medservice&page_size=100")
        assert medservice.status_code == 200
        medservice_row = {row["sku"]: row for row in medservice.json()["items"]}["PCT-API-WITH"]
        assert medservice_row["percentiles"]["10"] == 1

        astana = client.get("/api/competitors/percentile-rows?format_code=PCT-API&region=Astana&competitor=Emiti&page_size=100")
        assert astana.status_code == 200
        astana_row = {row["sku"]: row for row in astana.json()["items"]}["PCT-API-WITH"]
        assert astana_row["percentiles"]["10"] == 999
    finally:
        app.dependency_overrides.pop(get_db, None)


def test_percentile_rows_endpoint_falls_back_when_stale_generated_list_filters_are_sent():
    client = _percentile_client_with_rows()
    try:
        response = client.get(
            "/api/competitors/percentile-rows"
            "?format_code=PCT-API&region=888_Almaty_2026-07-01_wf44&competitor=888_Almaty_2026-07-01_wf44&page_size=100"
        )
        assert response.status_code == 200
        payload = response.json()
        by_sku = {row["sku"]: row for row in payload["items"]}

        assert payload["selectedRegion"] == "Almaty"
        assert payload["selectedCompetitor"] == "Emiti"
        assert payload["total"] == 2
        assert len(payload["priceColumns"]) == 5
        assert round(float(by_sku["PCT-API-WITH"]["percentiles"]["10"]), 3) == 687.0
        assert by_sku["PCT-API-WITH"]["competitorCount"] == 5
    finally:
        app.dependency_overrides.pop(get_db, None)


def test_percentile_trace_endpoint_returns_raw_prices_used():
    client = _percentile_client_with_rows()
    try:
        response = client.get(
            "/api/competitors/percentile-trace"
            "?format_code=PCT-API&region=Almaty&competitor=Emiti&sku=PCT-API-WITH"
        )
        assert response.status_code == 200
        payload = response.json()

        assert payload["found"] is True
        assert payload["competitor"] == "Emiti"
        assert payload["region"] == "Almaty"
        assert payload["sku"] == "PCT-API-WITH"
        assert payload["productId"] > 0
        assert payload["rawRowsCount"] == 5
        assert len(payload["sourceAccountIds"]) == 5
        assert {row["goodsId"] for row in payload["rawPricesUsed"]} == {163571}
        assert [row["price"] for row in payload["rawPricesUsed"]] == [705.0, 702.23, 710.0, 685.0, 690.0]
        assert payload["sortedPrices"] == [685.0, 690.0, 702.23, 705.0, 710.0]
        assert payload["priceCount"] == 5
        assert payload["usedPriceCount"] == 5
        assert round(float(payload["percentiles"]["10"]), 3) == 687.0
        assert round(float(payload["percentiles"]["60"]), 3) == 703.338
    finally:
        app.dependency_overrides.pop(get_db, None)


def test_percentile_csv_export_contains_all_rows():
    client = _percentile_client_with_rows()
    try:
        response = client.get("/api/competitors/percentile-rows/export.csv?format_code=PCT-API&region=Almaty&competitor=Emiti")
        assert response.status_code == 200
        text = response.content.decode("utf-8-sig")
        rows = list(csv.DictReader(io.StringIO(text)))

        assert len(rows) == 2
        assert {row["Код"] for row in rows} == {"PCT-API-WITH", "PCT-API-WITHOUT"}
        assert "Персентиль 10_Emiti" in rows[0]
        assert "Персентиль 60_Emiti" in rows[0]
        assert {row["Status"] for row in rows} == {"Рассчитан", "Нет данных"}
    finally:
        app.dependency_overrides.pop(get_db, None)


def test_percentile_xlsx_export_contains_all_rows():
    client = _percentile_client_with_rows()
    try:
        response = client.get("/api/competitors/percentile-rows/export.xlsx?format_code=PCT-API&region=Almaty&competitor=Emiti")
        assert response.status_code == 200
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))

        assert len(rows) == 3
        assert rows[0][0:2] == ("Код", "Название")
        assert "Персентиль 10_Emiti" in rows[0]
        assert "Персентиль 60_Emiti" in rows[0]
        assert {row[0] for row in rows[1:]} == {"PCT-API-WITH", "PCT-API-WITHOUT"}
    finally:
        app.dependency_overrides.pop(get_db, None)


def test_fixed_markup_5_5_overrides_global_20_and_is_persisted_in_logs():
    db = _session()
    pf = _format(db)
    db.query(MarkupRange).filter(MarkupRange.price_format_id == pf.id).delete()
    db.add(MarkupRange(price_format_id=pf.id, cost_from=0, cost_to=None, markup_percent=0.20))
    product = _product(db, code="FIXED-5-5", cost=1000)
    row = _list_item(db, product, "fixed_markup", 5.5, pf=pf)

    calculate_prices(
        db=db,
        price_format_code=pf.code,
        price_list_number="PL-FIXED-5-5",
        as_of=date.today(),
        activation_date=None,
        user="test",
        force_new_price_list=True,
    )

    cp = db.query(CalculatedPrice).one()
    payload = _generated_item_dict(db, cp, product, pf, None)
    log_by_label = {line["label"]: line["value"] for line in payload["log"]}

    assert float(cp.markup_percent_used) == 5.5
    assert float(cp.mdc_markup_percent) == 5.5
    assert round(float(cp.mdc_price), 2) == 1058.20
    assert cp.applied_list_id == row.id
    assert cp.applied_list_ids == f"[{row.id}]"
    assert log_by_label["Маржа (наценка)"] == "5.50%"
    assert payload["listOverrideLog"]["affectedField"] == "Маржа для расчёта МДЦ"
    assert "конкуренты и прогиб не применялись" in payload["listOverrideLog"]["action"]


def test_exact_lists_management_audit_skus_apply_overrides_during_generation():
    db = _session()
    rounding = RoundingRule(code="R01", name="R01", mode="math", precision=2, step=0.01)
    db.add(rounding)
    db.flush()
    pf = _format(db, rounding=rounding)
    db.query(MarkupRange).filter(MarkupRange.price_format_id == pf.id).delete()
    db.add(MarkupRange(price_format_id=pf.id, cost_from=0, cost_to=None, markup_percent=0.20))
    db.flush()

    specs = [
        ("000000000001017768", 8589.64, "fixed_price", Decimal("12100")),
        ("000000000001017769", 5257.77, "fixed_price", Decimal("9000")),
        ("000000000001017758", 3000, "fixed_price", Decimal("4200")),
        ("000000000001003454", 617.50, "max_price", Decimal("900")),
        ("000000000001005967", 559.45, "max_price", Decimal("820")),
        ("000000000001005754", 678.87, "max_price", Decimal("1000")),
        ("000000000001015510", 5000, "min_price", Decimal("7300")),
        ("000000000001015509", 1733.14, "min_price", Decimal("2200")),
        ("000000000001013903", 383, "fixed_markup", Decimal("5.5")),
        ("000000000001015508", 6239.30, "fixed_markup", Decimal("3.3")),
        ("000000000001002221", 10656, "critical_markup", Decimal("4.6")),
        ("000000000001006722", 2312.75, "critical_markup", Decimal("15.5")),
    ]
    expected_by_sku = {sku: (list_type, value, Decimal(str(cost))) for sku, cost, list_type, value in specs}

    for sku, cost, list_type, value in specs:
        product = _product(db, code=sku, cost=cost)
        ul = UniversalList(
            code=f"UL-AUDIT-{list_type}-{sku}",
            name=f"audit {list_type}",
            type=list_type,
            status="active",
            price_format_id=None,
        )
        db.add(ul)
        db.flush()
        db.add(UniversalListPriceFormat(universal_list_id=ul.id, price_format_id=pf.id))
        db.add(ListItem(universal_list_id=ul.id, product_id=product.id, value=value))
    db.flush()

    calculate_prices(
        db=db,
        price_format_code=pf.code,
        price_list_number="PL-EXACT-LISTS-AUDIT",
        as_of=date.today(),
        activation_date=None,
        user="test",
        force_new_price_list=True,
    )

    products_by_id = {row.id: row for row in db.query(Product).all()}
    calculated = db.query(CalculatedPrice).all()
    assert len(calculated) == len(specs)

    for cp in calculated:
        product = products_by_id[cp.product_id]
        list_type, expected_value, cost = expected_by_sku[product.code]
        payload = _generated_item_dict(db, cp, product, pf, None)
        list_log = payload["listOverrideLog"]
        log_by_label = {line["label"]: line["value"] for line in payload["log"]}

        assert cp.applied_rule_type == list_type
        assert float(cp.applied_rule_value) == float(expected_value)
        assert list_log["listType"] == list_type
        assert list_log["displayValue"] in {f"{float(expected_value):g}", f"{float(expected_value):g}%"}
        assert "Маржа (наценка)" in log_by_label
        assert "Фактическая маржа" in log_by_label

        if list_type == "fixed_price":
            assert Decimal(str(cp.final_price)) == expected_value
            assert payload["finalPrice"] == float(expected_value)
        elif list_type == "min_price":
            assert Decimal(str(cp.final_price)) >= expected_value
        elif list_type == "max_price":
            assert Decimal(str(cp.final_price)) <= expected_value
        else:
            expected_mdc = cost / (Decimal("1") - expected_value / Decimal("100"))
            assert Decimal(str(cp.mdc_markup_percent)) == expected_value
            assert Decimal(str(cp.markup_percent_used)) == expected_value
            assert Decimal(str(cp.mdc_price)).quantize(Decimal("0.01")) == expected_mdc.quantize(Decimal("0.01"))
            assert log_by_label["Маржа (наценка)"] == f"{float(expected_value):.2f}%"


def test_fixed_markup_one_is_one_percent_not_one_hundred_percent():
    db = _session()
    pf = _format(db)
    product = _product(db, cost=100)
    row = _list_item(db, product, "fixed_markup", 1, pf=pf)

    price, debug = calculate_price_for_product(db=db, product=product, price_format=pf, as_of=date.today())

    assert float(price) == 101.01
    assert debug["applied_rule_type"] == "fixed_markup"
    assert debug["applied_list_ids"] == [row.id]
    assert float(debug["mdc_markup_percent"]) == 1.0
    assert round(float(debug["mdc_price"]), 2) == 101.01


def test_lists_management_active_fixed_markup_m2m_overrides_default_markup():
    db = _session()
    pf = _format(db)
    product = _product(db, cost=1000)
    row = UniversalList(code="UL-FIXED-MARKUP", name="Fixed markup", type="fixed_markup", status="Активный")
    db.add(row)
    db.flush()
    db.add(ListItem(universal_list_id=row.id, product_id=product.id, value=33))
    db.add(UniversalListPriceFormat(universal_list_id=row.id, price_format_id=pf.id))
    db.flush()

    price, debug = calculate_price_for_product(db=db, product=product, price_format=pf, as_of=date.today())

    assert float(price) == 1492.54
    assert debug["reason"] == "fixed_markup_list_final"
    assert debug["applied_list_ids"] == [row.id]
    assert debug["applied_rule_type"] == "fixed_markup"
    assert float(debug["applied_rule_value"]) == 33.0
    assert float(debug["mdc_markup_percent"]) == 33.0
    assert round(float(debug["mdc_price"]), 2) == 1492.54
    assert debug["applied_list_name"] == "Fixed markup"
    assert "Конкуренты и прогиб не применялись" in debug["log"]
    assert "Applied Rule" not in debug["log"]


@pytest.mark.parametrize(
    ("critical_margin", "expected_mdc", "expected_price"),
    [(10, 111.11, 111.12), (35, 153.85, 153.85)],
)
def test_critical_markup_overrides_global_margin_for_mdc_and_no_competitor(critical_margin, expected_mdc, expected_price):
    db = _session()
    pf = _format(db)
    db.query(MarkupRange).filter(MarkupRange.price_format_id == pf.id).delete()
    db.query(NoCompetitorMarkupRange).filter(NoCompetitorMarkupRange.price_format_id == pf.id).delete()
    db.add(MarkupRange(price_format_id=pf.id, cost_from=0, cost_to=None, markup_percent=0.25))
    db.add(NoCompetitorMarkupRange(price_format_id=pf.id, cost_from=0, cost_to=None, markup_percent=0.25))
    product = _product(db, cost=100)
    row = _list_item(db, product, "critical_markup", critical_margin, pf=pf)

    price, debug = calculate_price_for_product(db=db, product=product, price_format=pf, as_of=date.today())

    assert round(float(price), 2) == expected_price
    assert float(debug["effective_markup_percent"]) == critical_margin
    assert float(debug["mdc_markup_percent"]) == critical_margin
    assert round(float(debug["mdc_price"]), 2) == expected_mdc
    assert debug["applied_list_ids"] == [row.id]


def test_critical_markup_still_uses_competitor_workflow():
    db = _session()
    pf = _format(db)
    db.query(BendRange).filter(BendRange.price_format_id == pf.id).delete()
    db.add(BendRange(price_format_id=pf.id, price_from=0, bend_percent=0))
    db.flush()
    product = _product(db, cost=100)
    row = _list_item(db, product, "critical_markup", 10, pf=pf)
    _competitor(db, pf, product, "manual:high", 200)

    price, debug = calculate_price_for_product(db=db, product=product, price_format=pf, as_of=date.today())

    assert float(price) == 200.0
    assert debug["reason"] == "competitor_bend"
    assert debug["applied_list_ids"] == [row.id]
    assert debug["applied_rule_type"] == "critical_markup"
    assert round(float(debug["mdc_price"]), 2) == 111.11
    assert float(debug["chosen_competitor_price"]) == 200
    assert float(debug["competitor_candidate_price"]) == 200


def test_fixed_markup_bypasses_competitors_and_uses_mdc_as_final_price():
    db = _session()
    pf = _format(db)
    db.query(MarkupRange).filter(MarkupRange.price_format_id == pf.id).delete()
    db.query(NoCompetitorMarkupRange).filter(NoCompetitorMarkupRange.price_format_id == pf.id).delete()
    db.query(BendRange).filter(BendRange.price_format_id == pf.id).delete()
    db.add(MarkupRange(price_format_id=pf.id, cost_from=0, cost_to=None, markup_percent=0.25))
    db.add(NoCompetitorMarkupRange(price_format_id=pf.id, cost_from=0, cost_to=None, markup_percent=0.25))
    db.add(BendRange(price_format_id=pf.id, price_from=0, bend_percent=5))
    db.flush()
    product = _product(db, code="000000000001000218", cost=278)
    row = UniversalList(code="010101", name="010101", type="fixed_markup", status="active")
    db.add(row)
    db.flush()
    db.add(ListItem(universal_list_id=row.id, product_id=product.id, value=5.5))
    db.add(UniversalListPriceFormat(universal_list_id=row.id, price_format_id=pf.id))
    _competitor(db, pf, product, "manual:c1", 242.95)
    _competitor(db, pf, product, "manual:c2", 275)
    _competitor(db, pf, product, "manual:c3", 315)
    db.flush()

    price, debug = calculate_price_for_product(db=db, product=product, price_format=pf, as_of=date.today())

    assert float(debug["effective_markup_percent"]) == 5.5
    assert round(float(debug["mdc_price"]), 2) == 294.18
    assert debug["selected_competitor_price"] is None
    assert debug["chosen_competitor_price"] is None
    assert float(debug["bend_percent_used"]) == 0
    assert debug["competitor_candidate_price"] is None
    assert debug["price_from_competitor"] is None
    assert round(float(price), 2) == 294.18
    assert round(float(debug["final_price"]), 2) == 294.18
    assert debug["applied_rule_type"] == "fixed_markup"
    assert float(debug["applied_rule_value"]) == 5.5
    assert debug["applied_list_id"] == row.id
    assert debug["applied_list_name"] == "010101"
    assert debug["rejected_competitors"] == []


def test_fixed_markup_sets_mdc_final_without_competitor_candidate():
    db = _session()
    pf = _format(db)
    db.query(BendRange).filter(BendRange.price_format_id == pf.id).delete()
    db.add(BendRange(price_format_id=pf.id, price_from=0, bend_percent=0))
    db.flush()
    product = _product(db, cost=90.72)
    row = _list_item(db, product, "fixed_markup", 33, pf=pf)
    _competitor(db, pf, product, "manual:low", 12)

    price, debug = calculate_price_for_product(db=db, product=product, price_format=pf, as_of=date.today())

    assert float(price) == 135.4
    assert debug["reason"] == "fixed_markup_list_final"
    assert debug["applied_list_ids"] == [row.id]
    assert debug["applied_rule_type"] == "fixed_markup"
    assert float(debug["applied_rule_value"]) == 33.0
    assert float(debug["mdc_markup_percent"]) == 33.0
    assert round(float(debug["mdc_price"]), 2) == 135.4
    assert debug["competitor_candidate_price"] is None
    assert debug["chosen_competitor_price"] is None
    assert float(debug["final_price"]) == 135.4


def test_fixed_markup_does_not_apply_bend():
    db = _session()
    pf = _format(db)
    db.query(BendRange).filter(BendRange.price_format_id == pf.id).delete()
    db.add(BendRange(price_format_id=pf.id, price_from=0, bend_percent=20))
    db.flush()
    product = _product(db, cost=100)
    row = _list_item(db, product, "fixed_markup", 10, pf=pf)
    _competitor(db, pf, product, "manual:high", 1000)

    price, debug = calculate_price_for_product(db=db, product=product, price_format=pf, as_of=date.today())

    assert float(price) == 111.11
    assert debug["reason"] == "fixed_markup_list_final"
    assert debug["applied_list_ids"] == [row.id]
    assert debug["applied_rule_type"] == "fixed_markup"
    assert float(debug["applied_rule_value"]) == 10.0
    assert float(debug["mdc_markup_percent"]) == 10.0
    assert float(debug["mdc_price"]) == 111.11
    assert debug["competitor_candidate_price"] is None
    assert debug["chosen_competitor_price"] is None
    assert debug["price_from_competitor"] is None
    assert float(debug["bend_percent_used"]) == 0
    assert float(debug["final_price"]) == 111.11


def test_fixed_markup_bypasses_multiple_competitors():
    db = _session()
    pf = _format(db)
    db.query(BendRange).filter(BendRange.price_format_id == pf.id).delete()
    db.add(BendRange(price_format_id=pf.id, price_from=0, bend_percent=5))
    db.flush()
    product = _product(db, cost=100)
    row = _list_item(db, product, "fixed_markup", 5.5, pf=pf)
    _competitor(db, pf, product, "manual:low", 90)
    _competitor(db, pf, product, "manual:mid", 200)
    _competitor(db, pf, product, "manual:high", 300)

    price, debug = calculate_price_for_product(db=db, product=product, price_format=pf, as_of=date.today())

    assert float(price) == 105.82
    assert float(debug["mdc_price"]) == 105.82
    assert debug["applied_list_ids"] == [row.id]
    assert debug["list_matched"] is True
    assert debug["list_applied"] is True
    assert debug["list_changed_final_price"] is True
    assert debug["selected_competitor_price"] is None
    assert debug["competitor_candidate_price"] is None
    assert debug["price_from_competitor"] is None
    assert float(debug["bend_percent_used"]) == 0
    assert "competitors and bend were bypassed" in debug["list_effect_message"]


def test_min_and_max_markup_lists_apply_bounds():
    db = _session()
    pf = _format(db)
    product_min = _product(db, code="MIN-MARKUP", cost=100)
    product_max = _product(db, code="MAX-MARKUP", cost=100)
    _competitor(db, pf, product_min, "manual:min", 105)
    _competitor(db, pf, product_max, "manual:max", 200)
    min_list = _list_item(db, product_min, "min_markup", 25, pf=pf)
    max_list = _list_item(db, product_max, "max_markup", 10, pf=pf)

    min_price, min_debug = calculate_price_for_product(db=db, product=product_min, price_format=pf, as_of=date.today())
    max_price, max_debug = calculate_price_for_product(db=db, product=product_max, price_format=pf, as_of=date.today())

    assert float(min_price) == 133.34
    assert min_debug["applied_list_ids"] == [min_list.id]
    assert float(max_price) == 199.0
    assert max_debug["applied_list_ids"] == [max_list.id]


def test_min_and_max_price_lists_apply_bounds():
    db = _session()
    pf = _format(db)
    product_min = _product(db, code="MIN-PRICE", cost=100)
    product_max = _product(db, code="MAX-PRICE", cost=100)
    _competitor(db, pf, product_min, "manual:min-price", 105)
    _competitor(db, pf, product_max, "manual:max-price", 200)
    min_list = _list_item(db, product_min, "min_price", 150, pf=pf)
    max_list = _list_item(db, product_max, "max_price", 130, pf=pf)

    min_price, min_debug = calculate_price_for_product(db=db, product=product_min, price_format=pf, as_of=date.today())
    max_price, max_debug = calculate_price_for_product(db=db, product=product_max, price_format=pf, as_of=date.today())

    assert float(min_price) == 150.0
    assert min_debug["applied_list_ids"] == [min_list.id]
    assert float(max_price) == 130.0
    assert max_debug["applied_list_ids"] == [max_list.id]


def test_min_price_noop_reports_checked_but_unchanged():
    db = _session()
    pf = _format(db)
    product = _product(db, cost=100)
    row = _list_item(db, product, "min_price", 50, pf=pf)

    price, debug = calculate_price_for_product(db=db, product=product, price_format=pf, as_of=date.today())

    assert float(price) > 50
    assert debug["applied_list_ids"] == [row.id]
    assert debug["list_matched"] is True
    assert debug["list_applied"] is True
    assert debug["list_changed_final_price"] is False
    assert "did not change final price" in debug["list_effect_message"]


def test_max_price_noop_reports_checked_but_unchanged():
    db = _session()
    pf = _format(db)
    product = _product(db, cost=100)
    row = _list_item(db, product, "max_price", 1000, pf=pf)

    price, debug = calculate_price_for_product(db=db, product=product, price_format=pf, as_of=date.today())

    assert float(price) < 1000
    assert debug["applied_list_ids"] == [row.id]
    assert debug["list_matched"] is True
    assert debug["list_applied"] is True
    assert debug["list_changed_final_price"] is False
    assert "did not change final price" in debug["list_effect_message"]


def test_min_price_1200_clamps_final_price_upward():
    db = _session()
    pf = _format(db)
    product = _product(db, cost=100)
    row = _list_item(db, product, "min_price", 1200, pf=pf)
    _competitor(db, pf, product, "manual:below-min-price", 500)

    price, debug = calculate_price_for_product(db=db, product=product, price_format=pf, as_of=date.today())

    assert float(price) == 1200.0
    assert debug["reason"] == "min_price_floor"
    assert debug["applied_list_ids"] == [row.id]
    assert debug["applied_rule_type"] == "min_price"
    assert float(debug["applied_rule_value"]) == 1200.0


@pytest.mark.parametrize(
    ("calculated_price", "min_price", "rounding_step", "expected"),
    [
        (7226, 7300, 72.26, 7300),
        (2163, 2200, 43.26, 2200),
        (8000, 7300, 0.01, 8000),
    ],
)
def test_min_price_is_hard_final_bound_after_rounding(calculated_price, min_price, rounding_step, expected):
    db = _session()
    rounding = RoundingRule(code=f"DOWN-{rounding_step}", name="Down", mode="down", precision=2, step=rounding_step)
    db.add(rounding)
    db.flush()
    pf = _format(db, rounding=rounding)
    db.query(BendRange).filter(BendRange.price_format_id == pf.id).delete()
    db.add(BendRange(price_format_id=pf.id, price_from=0, bend_percent=0))
    product = _product(db, code=f"MIN-{calculated_price}", cost=100)
    row = _list_item(db, product, "min_price", min_price, pf=pf)
    _competitor(db, pf, product, f"manual:{calculated_price}", calculated_price)

    price, debug = calculate_price_for_product(db=db, product=product, price_format=pf, as_of=date.today())

    assert float(price) == expected
    assert float(debug["final_price"]) == expected
    assert debug["applied_list_ids"] == [row.id]


def test_api_payload_keeps_pricing_and_max_price_list_logs_independent():
    db = _session()
    pf = _format(db)
    db.query(BendRange).filter(BendRange.price_format_id == pf.id).delete()
    db.add(BendRange(price_format_id=pf.id, price_from=0, bend_percent=0))
    product = _product(db, code="DUAL-LOG-MAX", cost=100)
    row = _list_item(db, product, "max_price", 820, pf=pf)
    _competitor(db, pf, product, "manual:first", 900)

    calculate_prices(
        db=db,
        price_format_code=pf.code,
        price_list_number="PL-DUAL-LOG-MAX",
        as_of=date.today(),
        activation_date=None,
        user="test",
        force_new_price_list=True,
    )

    cp = db.query(CalculatedPrice).one()
    payload = _generated_item_dict(db, cp, product, pf, None)
    assert payload["pricingCalculationLog"].startswith("Цена рассчитана относительно 1-й цены конкурента")
    assert payload["listOverrideLog"] == {
        "listName": "max_price list",
        "listCode": "UL-max_price-DUAL-LOG-MAX",
        "listType": "max_price",
        "value": 820.0,
        "displayValue": "820",
        "affectedField": "Максимальная финальная цена",
        "action": "Применено ограничение максимальной цены.",
        "ambiguous": False,
        "listMatched": True,
        "listApplied": True,
        "listChangedFinalPrice": True,
        "listEffectMessage": "Применено ограничение максимальной цены.",
    }
    assert payload["appliedListId"] == row.id
    assert float(payload["finalPrice"]) == 820.0


def test_max_price_3000_clamps_final_price_downward():
    db = _session()
    pf = _format(db)
    product = _product(db, cost=100)
    row = _list_item(db, product, "max_price", 3000, pf=pf)
    _competitor(db, pf, product, "manual:above-max-price", 5000)

    price, debug = calculate_price_for_product(db=db, product=product, price_format=pf, as_of=date.today())

    assert float(price) == 3000.0
    assert debug["reason"] == "max_price_cap"
    assert debug["applied_list_ids"] == [row.id]
    assert debug["applied_rule_type"] == "max_price"
    assert float(debug["applied_rule_value"]) == 3000.0


def test_exclusion_list_marks_product_without_normal_pricing():
    db = _session()
    pf = _format(db)
    product = _product(db, cost=100)
    row = _list_item(db, product, "exclude_from_pricing", 1, pf=pf)

    price, debug = calculate_price_for_product(db=db, product=product, price_format=pf, as_of=date.today())

    assert float(price) == 100.0
    assert debug["reason"] == "exclude_from_pricing_list"
    assert debug["excluded_from_pricing"] is True
    assert debug["applied_list_ids"] == [row.id]


def test_excluded_product_is_absent_from_generated_price_rows():
    db = _session()
    pf = _format(db)
    excluded = _product(db, code="EXCLUDED", cost=100)
    included = _product(db, code="INCLUDED", cost=100)
    _list_item(db, excluded, "exclude_from_pricing", 1, pf=pf)

    count = calculate_prices(
        db=db,
        price_format_code=pf.code,
        price_list_number="PL-EXCLUSION",
        as_of=date.today(),
        activation_date=None,
        user="test",
        force_new_price_list=True,
    )

    assert count == 1
    rows = db.query(CalculatedPrice).all()
    assert [row.product_id for row in rows] == [included.id]


def test_exclusion_removes_existing_row_when_price_list_is_regenerated():
    db = _session()
    pf = _format(db)
    excluded = _product(db, code="LATE-EXCLUDED", cost=100)
    included = _product(db, code="STILL-INCLUDED", cost=100)

    initial_count = calculate_prices(
        db=db,
        price_format_code=pf.code,
        price_list_number="PL-REGENERATED-EXCLUSION",
        as_of=date.today(),
        activation_date=None,
        user="test",
    )
    assert initial_count == 2
    assert db.query(CalculatedPrice).filter(CalculatedPrice.product_id == excluded.id).one_or_none() is not None

    _list_item(db, excluded, "exclude_from_pricing", 1, pf=pf)
    regenerated_count = calculate_prices(
        db=db,
        price_format_code=pf.code,
        price_list_number="PL-REGENERATED-EXCLUSION",
        as_of=date.today(),
        activation_date=None,
        user="test",
    )

    assert regenerated_count == 1
    assert db.query(CalculatedPrice).filter(CalculatedPrice.product_id == excluded.id).one_or_none() is None
    assert db.query(CalculatedPrice).filter(CalculatedPrice.product_id == included.id).one_or_none() is not None


def test_no_bend_disables_bend_only_for_matching_product():
    db = _session()
    pf = _format(db)
    product = _product(db, code="NO-BEND", cost=100)
    control = _product(db, code="CONTROL", cost=100)
    row = _list_item(db, product, "no_bend", 1, pf=pf)
    _competitor(db, pf, product, "manual:no-bend", 200)
    _competitor(db, pf, control, "manual:control", 200)

    price, debug = calculate_price_for_product(db=db, product=product, price_format=pf, as_of=date.today())
    control_price, control_debug = calculate_price_for_product(db=db, product=control, price_format=pf, as_of=date.today())

    assert float(price) == 200.0
    assert float(debug["bend_percent_used"]) == 0.0
    assert debug["applied_list_ids"] == [row.id]
    assert float(control_price) == 199.0
    assert float(control_debug["bend_percent_used"]) == 0.5
    assert control_debug["applied_list_ids"] == []


def test_same_type_list_conflict_aborts_generation_with_clear_message():
    db = _session()
    pf = _format(db)
    product = _product(db, code="CONFLICT-SKU", cost=100)
    first = _list_item(db, product, "fixed_price", 111, pf=pf)
    second = _list_item(db, product, "fixed_price", 222, pf=pf)

    with pytest.raises(ValueError) as exc_info:
        calculate_prices(
            db=db,
            price_format_code=pf.code,
            price_list_number="PL-CONFLICT",
            as_of=date.today(),
            activation_date=None,
            user="test",
            force_new_price_list=True,
        )

    message = str(exc_info.value)
    assert "Lists Management conflicts detected" in message
    assert "SKU=CONFLICT-SKU" in message
    assert "type=fixed_price" in message
    assert f"#{first.id}" in message
    assert f"#{second.id}" in message


def test_unconfirmed_rule_is_explicitly_marked_ambiguous_in_diagnostics_and_api():
    db = _session()
    pf = _format(db)
    product = _product(db, code="AMBIGUOUS", cost=100)
    row = _list_item(db, product, "min_markup", 25, pf=pf)

    calculate_prices(
        db=db,
        price_format_code=pf.code,
        price_list_number="PL-AMBIGUOUS",
        as_of=date.today(),
        activation_date=None,
        user="test",
        force_new_price_list=True,
    )

    cp = db.query(CalculatedPrice).one()
    payload = _generated_item_dict(db, cp, product, pf, None)
    assert cp.applied_rule_type == "min_markup"
    assert "AMBIGUOUS" not in cp.applied_reason
    assert payload["appliedRuleAmbiguous"] is True
    assert payload["appliedRule"] == {
        "applied_rule_type": "min_markup",
        "applied_rule_value": 25.0,
        "list_id": row.id,
        "list_name": "min_markup list",
        "list_code": "UL-min_markup-AMBIGUOUS",
        "ambiguous": True,
    }
    assert payload["listOverrideLog"]["ambiguous"] is True


def test_many_to_many_list_assignment_affects_matching_format_only():
    db = _session()
    pf = _format(db)
    other_pf = PriceFormat(code="0002", name="0002", branch="Almaty")
    db.add(other_pf)
    db.flush()
    db.add(MarkupRange(price_format_id=other_pf.id, cost_from=0, cost_to=None, markup_percent=0.15))
    product = _product(db, cost=100)
    row = UniversalList(code="UL-M2M", name="M2M", type="fixed_price", status="active")
    db.add(row)
    db.flush()
    db.add(ListItem(universal_list_id=row.id, product_id=product.id, value=777))
    db.add(UniversalListPriceFormat(universal_list_id=row.id, price_format_id=pf.id))
    db.flush()

    price, debug = calculate_price_for_product(db=db, product=product, price_format=pf, as_of=date.today())
    other_price, other_debug = calculate_price_for_product(db=db, product=product, price_format=other_pf, as_of=date.today())

    assert float(price) == 777.0
    assert debug["applied_list_ids"] == [row.id]
    assert float(other_price) == 117.65
    assert other_debug["applied_list_ids"] == []


def test_inactive_and_out_of_date_lists_are_ignored():
    db = _session()
    pf = _format(db)
    product = _product(db, cost=100)
    _list_item(db, product, "fixed_price", 1, pf=pf, status="inactive")
    _list_item(db, product, "fixed_price", 2, pf=pf, start_date=date(2030, 1, 1))
    _list_item(db, product, "fixed_price", 3, pf=pf, end_date=date(2020, 1, 1))

    price, debug = calculate_price_for_product(db=db, product=product, price_format=pf, as_of=date.today())

    assert float(price) == 117.65
    assert debug["applied_list_ids"] == []


def test_calculate_prices_stores_explicit_competitor_fields():
    db = _session()
    pf = _format(db)
    product = _product(db, cost=2500)
    _competitor(db, pf, product, "manual:low", 2900)
    _competitor(db, pf, product, "manual:second", 3000)

    calculate_prices(
        db=db,
        price_format_code=pf.code,
        price_list_number="PL-1",
        as_of=date.today(),
        activation_date=None,
        user="test",
        force_new_price_list=True,
    )

    cp = db.query(CalculatedPrice).one()
    assert float(cp.lowest_competitor_price) == 2900
    assert float(cp.chosen_competitor_price) == 3000
    assert float(cp.price_from_competitor) == 2994.0
    assert float(cp.bend_percent_used) == 0.2
    assert float(cp.markup_percent_used) == 15
    assert cp.zone == "right"


def test_calculate_prices_stores_applied_list_diagnostics():
    db = _session()
    pf = _format(db)
    product = _product(db, cost=100)
    row = _list_item(db, product, "fixed_price", 333, pf=pf)

    calculate_prices(
        db=db,
        price_format_code=pf.code,
        price_list_number="PL-LIST",
        as_of=date.today(),
        activation_date=None,
        user="test",
        force_new_price_list=True,
    )

    cp = db.query(CalculatedPrice).one()
    assert float(cp.final_price) == 333
    assert cp.applied_list_ids == f"[{row.id}]"
    assert cp.applied_list_id == row.id
    assert cp.applied_rule_type == "fixed_price"
    assert float(cp.applied_rule_value) == 333


def test_calculate_prices_stores_mdc_markup_diagnostics():
    db = _session()
    pf = _format(db)
    db.query(BendRange).filter(BendRange.price_format_id == pf.id).delete()
    db.add(BendRange(price_format_id=pf.id, price_from=0, bend_percent=0))
    db.flush()
    product = _product(db, cost=90.72)
    row = _list_item(db, product, "fixed_markup", 33, pf=pf)
    _competitor(db, pf, product, "manual:high", 200)

    calculate_prices(
        db=db,
        price_format_code=pf.code,
        price_list_number="PL-MDC-DIAG",
        as_of=date.today(),
        activation_date=None,
        user="test",
        force_new_price_list=True,
    )

    cp = db.query(CalculatedPrice).one()
    assert float(cp.final_price) == 135.4
    assert cp.applied_list_ids == f"[{row.id}]"
    assert cp.applied_list_id == row.id
    assert cp.applied_rule_type == "fixed_markup"
    assert float(cp.applied_rule_value) == 33
    assert float(cp.mdc_markup_percent) == 33
    assert round(float(cp.mdc_price), 2) == 135.4
    assert cp.competitor_candidate_price is None
    assert cp.chosen_competitor_price is None
    assert "Цена рассчитана" in cp.applied_reason
    assert "Конкуренты и прогиб не применялись" in cp.applied_reason
    assert "Applied Rule" not in cp.applied_reason


def test_generated_item_payload_exposes_applied_list_rule_details():
    db = _session()
    pf = _format(db)
    product = _product(db, cost=100)
    row = _list_item(db, product, "fixed_price", 2500, pf=pf)

    calculate_prices(
        db=db,
        price_format_code=pf.code,
        price_list_number="PL-ITEM-DIAG",
        as_of=date.today(),
        activation_date=None,
        user="test",
        force_new_price_list=True,
    )

    cp = db.query(CalculatedPrice).one()
    payload = _generated_item_dict(db, cp, product, pf, None)
    assert payload["appliedRuleType"] == "fixed_price"
    assert payload["appliedRuleValue"] == 2500.0
    assert payload["appliedListId"] == row.id
    assert payload["appliedListName"] == "fixed_price list"
    assert payload["appliedRule"] == {
        "applied_rule_type": "fixed_price",
        "applied_rule_value": 2500.0,
        "list_id": row.id,
        "list_name": "fixed_price list",
        "list_code": "UL-fixed_price-SKU-1",
        "ambiguous": False,
    }
    assert payload["pricingCalculationLog"] == cp.applied_reason
    assert payload["listOverrideLog"] == {
        "listName": "fixed_price list",
        "listCode": "UL-fixed_price-SKU-1",
        "listType": "fixed_price",
        "value": 2500.0,
        "displayValue": "2500",
        "affectedField": "Финальная цена",
        "action": "Применена фиксированная цена из списка.",
        "ambiguous": False,
        "listMatched": True,
        "listApplied": True,
        "listChangedFinalPrice": True,
        "listEffectMessage": "Применена фиксированная цена из списка.",
    }
    assert "fixed_price" not in payload["pricingCalculationLog"]


def test_generated_item_payload_recomputes_zone_reference_when_saved_best_competitor_is_empty():
    db = _session()
    pf = _format(db)
    product = _product(db, cost=100, code="PAYLOAD-ZONE")
    _list_item(db, product, "fixed_price", 243.39, pf=pf)
    db.add(
        CompetitorPrice(
            price_format_id=pf.id,
            product_id=product.id,
            source_name="manual:available",
            supplier="manual:available",
            source_price=244.43,
            coefficient=1,
        )
    )
    db.flush()

    calculate_prices(
        db=db,
        price_format_code=pf.code,
        price_list_number="PL-PAYLOAD-ZONE",
        as_of=date.today(),
        activation_date=None,
        user="test",
        force_new_price_list=True,
    )

    cp = db.query(CalculatedPrice).one()
    cp.competitor_price = None
    cp.lowest_competitor_price = None
    db.flush()

    payload = _generated_item_dict(db, cp, product, pf, None)
    assert payload["bestCompetitorPrice"] is None
    assert payload["lowestCompetitorPrice"] == 244.43
    assert payload["zone"] == "left"


def test_zone_uses_lowest_competitor_even_when_chosen_competitor_is_different():
    zone, reference, deviation = calculate_price_zone(
        2894.2,
        chosen_competitor_price=2900,
        lowest_competitor_price=2850,
    )

    assert zone == "optimal"
    assert float(reference) == 2850
    assert float(deviation) > 0


@pytest.mark.parametrize(
    ("final_price", "expected_zone"),
    [(99, "left"), (100, "optimal"), (103, "optimal"), (103.01, "right")],
)
def test_zone_boundaries_are_relative_to_lowest_competitor(final_price, expected_zone):
    zone, reference, _deviation = calculate_price_zone(
        final_price,
        chosen_competitor_price=150,
        lowest_competitor_price=100,
    )

    assert zone == expected_zone
    assert float(reference) == 100


@pytest.mark.parametrize(
    ("final_price", "expected_zone"),
    [(999, "left"), (1000, "optimal"), (1030, "optimal"), (1031, "right")],
)
def test_zone_business_rule_exact_c1_1000_cases(final_price, expected_zone):
    zone, reference, _deviation = calculate_price_zone(
        final_price,
        chosen_competitor_price=1500,
        lowest_competitor_price=1000,
    )

    assert zone == expected_zone
    assert float(reference) == 1000


def test_zone_falls_back_to_lowest_competitor_when_chosen_missing():
    zone, reference, deviation = calculate_price_zone(
        1020,
        chosen_competitor_price=None,
        lowest_competitor_price=1000,
    )

    assert zone == "optimal"
    assert float(reference) == 1000
    assert round(float(deviation), 2) == 0.02


def test_zone_is_empty_without_any_competitor_reference():
    zone, reference, deviation = calculate_price_zone(
        1020,
        chosen_competitor_price=None,
        lowest_competitor_price=None,
    )

    assert zone is None
    assert reference is None
    assert deviation is None


def test_generated_item_markup_is_rule_markup_and_actual_margin_is_separate():
    db = _session()
    pf = _format(db)
    product = _product(db, cost=100)
    pl = PriceList(number="PL-XLSX", price_format_id=pf.id, user="test")
    db.add(pl)
    db.flush()
    cp = CalculatedPrice(
        price_list_id=pl.id,
        product_id=product.id,
        cost=100,
        base_price=115,
        competitor_price=None,
        final_price=120,
        markup_percent_used=15,
        mdc_markup_percent=10,
        zone="",
    )
    db.add(cp)
    db.flush()

    item = _generated_item_dict(db, cp, product, pf, None)

    assert item["markupPercent"] == 15
    assert item["actualMarginPercent"] == 16.67
    log_by_label = {line["label"]: line["value"] for line in item["log"]}
    assert log_by_label["Маржа (наценка)"] == "10.00%"
    assert log_by_label["Фактическая маржа"] == "16.67%"
    assert item["zone"] == "no-data"
    assert _export_zone(cp) == "no-data"


def test_legacy_generated_item_recovers_pricing_percent_from_stored_base_price():
    db = _session()
    pf = _format(db)
    product = _product(db, cost=100)
    pl = PriceList(number="PL-LEGACY", price_format_id=pf.id, user="test")
    db.add(pl)
    db.flush()
    cp = CalculatedPrice(
        price_list_id=pl.id,
        product_id=product.id,
        cost=100,
        base_price=115,
        final_price=120,
        zone="",
    )
    db.add(cp)
    db.flush()

    item = _generated_item_dict(db, cp, product, pf, None)
    log_by_label = {line["label"]: line["value"] for line in item["log"]}

    assert log_by_label["Маржа (наценка)"] == "15.00%"
    assert log_by_label["Фактическая маржа"] == "16.67%"


def test_mapping_response_exposes_mapping_and_generated_coverage():
    db = _session()
    pf = _format(db)
    product_a = _product(db, code="A", cost=10)
    product_b = _product(db, code="B", cost=20)
    pl = PriceList(number="PL-COV", price_format_id=pf.id, user="test")
    db.add(pl)
    db.flush()
    db.add(CalculatedPrice(price_list_id=pl.id, product_id=product_a.id, cost=10, base_price=11, competitor_price=12, final_price=12, zone="right"))
    db.add(CalculatedPrice(price_list_id=pl.id, product_id=product_b.id, cost=20, base_price=22, competitor_price=None, final_price=21, zone=""))
    db.flush()

    payload = list_catalog_code_mappings(db=db, platform="provisor", price_format_id=pf.id, limit=10)

    metric = payload["metrics"][0]
    assert "mappingCoveragePercent" in metric
    assert metric["generatedPricingCoverage"]["total"] == 2
    assert metric["generatedPricingCoverage"]["withCompetitors"] == 1
    assert metric["generatedPricingCoverage"]["withoutCompetitors"] == 1
    assert metric["generatedPricingCoverage"]["coveragePercent"] == 50


def test_analytics_uses_selected_price_list_and_explicit_bend():
    db = _session()
    pf = _format(db)
    product_a = _product(db, code="A", cost=100)
    product_b = _product(db, code="B", cost=100)
    first = PriceList(number="FIRST", price_format_id=pf.id, user="test")
    second = PriceList(number="SECOND", price_format_id=pf.id, user="test")
    db.add_all([first, second])
    db.flush()
    db.add(CalculatedPrice(price_list_id=first.id, product_id=product_a.id, cost=100, base_price=115, competitor_price=120, chosen_competitor_price=120, price_from_competitor=119.4, bend_percent_used=0.5, final_price=119.4, zone="left"))
    db.add(CalculatedPrice(price_list_id=second.id, product_id=product_b.id, cost=100, base_price=115, competitor_price=None, final_price=103, zone=""))
    db.flush()

    analytics = build_workflow_analytics(db=db, price_list_id=first.id)

    assert analytics["summary"]["skuTotal"] == 1
    assert analytics["summary"]["withCompetitors"] == 1
    assert analytics["summary"]["withoutCompetitors"] == 0
    assert analytics["summary"]["averageBendPercent"] == 0.5

    no_competitor_analytics = build_workflow_analytics(db=db, price_list_id=second.id)
    assert no_competitor_analytics["summary"]["withoutCompetitors"] == 1
    assert "noDataZone" not in no_competitor_analytics["summary"]
    assert "noIntersection" not in no_competitor_analytics["summary"]
    assert [item["name"] for item in no_competitor_analytics["zones"]] == ["left", "optimal", "right", "no-data"]
    assert {item["name"]: item["value"] for item in no_competitor_analytics["zones"]}["no-data"] == 1


def test_analytics_splits_right_zone_reasons():
    db = _session()
    pf = _format(db)
    products = [_product(db, code=f"R{i}", cost=100) for i in range(3)]
    pl = PriceList(number="RIGHTS", price_format_id=pf.id, user="test")
    db.add(pl)
    db.flush()
    db.add(CalculatedPrice(price_list_id=pl.id, product_id=products[0].id, cost=100, base_price=115, competitor_price=100, final_price=115, zone="right"))
    db.add(CalculatedPrice(price_list_id=pl.id, product_id=products[1].id, cost=100, base_price=90, competitor_price=100, chosen_competitor_price=130, price_from_competitor=129, final_price=129, zone="right"))
    db.add(CalculatedPrice(price_list_id=pl.id, product_id=products[2].id, cost=100, base_price=90, competitor_price=100, final_price=120, zone="right", applied_list_ids="[7]"))
    db.flush()

    analytics = build_workflow_analytics(db=db, price_list_id=pl.id)

    assert analytics["rightZoneReasons"]["right_due_to_mdc_floor"] == 1
    assert analytics["rightZoneReasons"]["right_due_to_chosen_higher_competitor"] == 0
    assert analytics["rightZoneReasons"]["right_due_to_universal_override"] == 1


def test_analytics_recalculates_old_rows_without_chosen_competitor():
    db = _session()
    pf = _format(db)
    product = _product(db, code="OLD", cost=100)
    pl = PriceList(number="OLD-ZONE", price_format_id=pf.id, user="test")
    db.add(pl)
    db.flush()
    db.add(
        CalculatedPrice(
            price_list_id=pl.id,
            product_id=product.id,
            cost=100,
            base_price=115,
            competitor_price=100,
            final_price=102,
            zone="right",
        )
    )
    db.flush()

    analytics = build_workflow_analytics(db=db, price_list_id=pl.id)

    assert analytics["summary"]["optimalZone"] == 1
    assert analytics["summary"]["rightZone"] == 0
