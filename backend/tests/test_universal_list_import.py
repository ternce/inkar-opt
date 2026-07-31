from __future__ import annotations

import io
import json
from datetime import date, datetime, timedelta

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from sqlalchemy import create_engine, func, select
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from backend.app.db import Base
from backend.app.deps import get_db
from backend.app.main import _generated_price_export_rows, app
from backend.app.models import (
    BusinessList,
    BusinessListItem,
    BranchCost,
    BranchStock,
    CalculatedPrice,
    ListItem,
    MarkupRange,
    PriceFormat,
    PriceList,
    Product,
    ReferenceUpdateStatus,
    UniversalList,
    UniversalListPriceFormat,
)
from backend.app.services.pricing import calculate_prices


def _client():
    engine = create_engine(
        "sqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    Session = sessionmaker(bind=engine)

    def override_db():
        db = Session()
        try:
            yield db
        finally:
            db.close()

    app.dependency_overrides[get_db] = override_db
    return TestClient(app), Session


def _xlsx(rows: list[list[object]]) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    for row in rows:
        worksheet.append(row)
    bio = io.BytesIO()
    workbook.save(bio)
    return bio.getvalue()


def _xlsx_with_styled_tail(rows: list[list[object]], *, styled_row: int) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    for row in rows:
        worksheet.append(row)
    worksheet.cell(row=styled_row, column=1).fill = PatternFill(fill_type="solid", fgColor="FFFF00")
    bio = io.BytesIO()
    workbook.save(bio)
    return bio.getvalue()


def _activate_all_products_for_generation(db, *, branch_id="1"):
    products = db.query(Product).all()
    for product in products:
        if db.execute(
            select(BranchStock).where(BranchStock.branch_id == branch_id).where(BranchStock.product_id == product.id)
        ).scalars().first() is None:
            db.add(BranchStock(branch_id=branch_id, product_id=product.id, sku=product.code, stock=1))
        if db.execute(
            select(BranchCost).where(BranchCost.branch_id == branch_id).where(BranchCost.product_id == product.id)
        ).scalars().first() is None:
            db.add(BranchCost(branch_id=branch_id, product_id=product.id, sku=product.code, cost=float(product.cost or 0)))
    status = db.execute(
        select(ReferenceUpdateStatus)
        .where(ReferenceUpdateStatus.branch_id == branch_id)
        .where(ReferenceUpdateStatus.data_type == "stock")
    ).scalars().first()
    if status is None:
        status = ReferenceUpdateStatus(branch_id=branch_id, data_type="stock")
        db.add(status)
    status.branch_name = branch_id
    status.status = "success"
    status.rows_count = len(products)
    status.last_updated_at = datetime(2026, 1, 1, 8, 0, 0)
    cost_status = db.execute(
        select(ReferenceUpdateStatus)
        .where(ReferenceUpdateStatus.branch_id == branch_id)
        .where(ReferenceUpdateStatus.data_type == "cost")
    ).scalars().first()
    if cost_status is None:
        cost_status = ReferenceUpdateStatus(branch_id=branch_id, data_type="cost")
        db.add(cost_status)
    cost_status.branch_name = branch_id
    cost_status.status = "success"
    cost_status.rows_count = len(products)
    cost_status.last_updated_at = datetime(2026, 1, 1, 8, 0, 0)
    db.flush()


def _upload(client: TestClient, rows: list[list[object]], list_type: str):
    return client.post(
        "/lists/import",
        data={"list_type": list_type},
        files={"file": ("list.xlsx", _xlsx(rows), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )


def _seed_products(Session):
    db = Session()
    try:
        db.add_all(
            [
                Product(code="12345", name="Product 12345", cost=1, provisor_goods_id=1003437),
                Product(code="A-77", name="Article A77", cost=1, provisor_goods_id=1005949),
                Product(code="000000000001005594", name="Ко-Ирбесан 300мг/12,5мг №14 таб", cost=1, provisor_goods_id=777777),
                Product(code="PC-9", name="Product Code 9", cost=1),
                Product(code="000000000000000222", name="Padded Product", cost=1),
            ]
        )
        db.commit()
    finally:
        db.close()


def test_critical_list_with_one_zero_values():
    client, Session = _client()
    _seed_products(Session)

    response = _upload(client, [["Material", "Value"], ["12345", 1], ["A-77", 0]], "critical")

    assert response.status_code == 200
    payload = response.json()
    assert payload["summary"]["processed"] == 2
    list_id = payload["id"]
    details = client.get(f"/lists/{list_id}").json()
    values = {item["sku"]: item["value"]["is_critical"] for item in details["items"]}
    assert values == {"12345": True, "A-77": False}


def test_markup_list_normalizes_20():
    client, Session = _client()
    _seed_products(Session)

    response = _upload(client, [["SKU", "Markup"], ["12345", 20]], "markup")

    assert response.status_code == 200
    details = client.get(f"/lists/{response.json()['id']}").json()
    assert details["items"][0]["value"]["markup_percent"] == 20.0
    assert details["items"][0]["value"]["display"] == "20%"


def test_business_list_import_timestamps_are_kazakhstan_offset():
    client, Session = _client()
    _seed_products(Session)

    response = _upload(client, [["SKU", "Markup"], ["12345", 20]], "markup")

    assert response.status_code == 200
    assert response.json()["created_at"].endswith("+05:00")


def test_markup_list_normalizes_20_percent():
    client, Session = _client()
    _seed_products(Session)

    response = _upload(client, [["SKU", "Markup"], ["12345", "20%"]], "markup")

    assert response.status_code == 200
    details = client.get(f"/lists/{response.json()['id']}").json()
    assert details["items"][0]["value"]["markup_percent"] == 20.0
    assert details["items"][0]["value"]["display"] == "20%"


def test_markup_list_normalizes_fraction():
    client, Session = _client()
    _seed_products(Session)

    response = _upload(client, [["SKU", "Markup"], ["12345", 0.2]], "markup")

    assert response.status_code == 200
    details = client.get(f"/lists/{response.json()['id']}").json()
    assert details["items"][0]["value"]["markup_percent"] == 20.0
    assert details["items"][0]["value"]["display"] == "20%"


def test_markup_list_normalizes_one_as_one_percent():
    client, Session = _client()
    _seed_products(Session)

    response = _upload(client, [["SKU", "Markup"], ["12345", 1]], "markup")

    assert response.status_code == 200
    details = client.get(f"/lists/{response.json()['id']}").json()
    assert details["items"][0]["value"]["markup_percent"] == 1.0
    assert details["items"][0]["value"]["display"] == "1%"


def test_markup_list_normalizes_integer_percents():
    client, Session = _client()
    _seed_products(Session)

    response = _upload(client, [["SKU", "Markup"], ["12345", 2], ["A-77", 20]], "markup")

    assert response.status_code == 200
    details = client.get(f"/lists/{response.json()['id']}").json()
    values = {item["sku"]: item["value"]["markup_percent"] for item in details["items"]}
    assert values == {"12345": 2.0, "A-77": 20.0}


def test_shuffled_column_order_and_product_code_lookup():
    client, Session = _client()
    _seed_products(Session)

    response = _upload(client, [["Value", "Manufacturer", "Product Code"], [1, "Maker", "PC-9"]], "exclusion")

    assert response.status_code == 200
    details = client.get(f"/lists/{response.json()['id']}").json()
    assert details["items"][0]["sku"] == "PC-9"
    assert details["items"][0]["manufacturer"] == "Maker"
    assert details["items"][0]["value"]["is_excluded"] is True


def test_duplicate_products_last_row_wins():
    client, Session = _client()
    _seed_products(Session)

    response = _upload(client, [["Material", "Critical"], ["12345", 0], ["12345", 1]], "critical")

    assert response.status_code == 200
    payload = response.json()
    assert payload["summary"]["processed"] == 1
    assert payload["summary"]["duplicates"] == 1
    details = client.get(f"/lists/{payload['id']}").json()
    assert details["items"][0]["value"]["is_critical"] is True
    assert details["errors"][0]["code"] == "duplicate_row"


def test_unknown_sku_reported():
    client, Session = _client()
    _seed_products(Session)

    response = _upload(client, [["SKU", "Value"], ["UNKNOWN", 1]], "critical")

    assert response.status_code == 200
    payload = response.json()
    assert payload["summary"]["not_found"] == 1
    assert payload["summary"]["processed"] == 0
    assert payload["errors"][0]["code"] == "product_not_found"


def test_empty_file_rejected():
    client, _Session = _client()

    response = _upload(client, [], "critical")

    assert response.status_code == 400
    assert "file without headers" in response.json()["detail"] or "empty file" in response.json()["detail"]


def test_file_without_headers_rejected():
    client, _Session = _client()

    response = _upload(client, [[None, None], ["12345", 1]], "critical")

    assert response.status_code == 400
    assert "file without headers" in response.json()["detail"]


def test_upload_size_limit_rejected(monkeypatch):
    client, _Session = _client()
    monkeypatch.setenv("LIST_IMPORT_MAX_UPLOAD_SIZE_MB", "1")

    response = client.post(
        "/lists/import",
        data={"list_type": "critical"},
        files={"file": ("list.xlsx", b"x" * (1024 * 1024 + 1), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert response.status_code == 400
    assert "LIST_IMPORT_MAX_UPLOAD_SIZE_MB=1" in response.json()["detail"]


def test_row_limit_rejected_and_does_not_persist(monkeypatch):
    client, Session = _client()
    _seed_products(Session)
    monkeypatch.setenv("LIST_IMPORT_MAX_ROWS", "1")

    response = _upload(client, [["SKU", "Value"], ["12345", 1], ["A-77", 1]], "critical")

    assert response.status_code == 400
    assert "LIST_IMPORT_MAX_ROWS=1" in response.json()["detail"]
    db = Session()
    try:
        assert db.scalar(select(func.count(BusinessList.id))) == 0
        assert db.scalar(select(func.count(BusinessListItem.id))) == 0
    finally:
        db.close()


def test_database_error_rolls_back_and_returns_normalized_error(monkeypatch):
    client, Session = _client()
    _seed_products(Session)

    def broken_commit(self):
        raise SQLAlchemyError("forced commit failure")

    monkeypatch.setattr(Session.class_, "commit", broken_commit)

    response = _upload(client, [["SKU", "Value"], ["12345", 1]], "critical")

    assert response.status_code == 500
    assert response.json()["detail"] == "database error during list import"
    db = Session()
    try:
        assert db.scalar(select(func.count(BusinessList.id))) == 0
        assert db.scalar(select(func.count(BusinessListItem.id))) == 0
    finally:
        db.close()


def test_delete_list_removes_items():
    client, Session = _client()
    _seed_products(Session)
    response = _upload(client, [["SKU", "Value"], ["12345", 1]], "critical")
    list_id = response.json()["id"]

    delete_response = client.delete(f"/lists/{list_id}")

    assert delete_response.status_code == 200
    db = Session()
    try:
        assert db.scalar(select(BusinessList).where(BusinessList.id == list_id)) is None
        assert db.scalar(select(BusinessListItem).where(BusinessListItem.business_list_id == list_id)) is None
    finally:
        db.close()


def test_delete_universal_list_cleans_bindings_items_and_calculated_price_refs():
    client, Session = _client()
    db = Session()
    try:
        price_format = PriceFormat(code="003", name="Format 003", branch="A")
        product = Product(code="SKU-DEL", name="Delete Product", cost=10)
        universal_list = UniversalList(code="UL-DEL", name="Delete Me", type="fixed_price", status="active")
        price_list = PriceList(number="PL-DEL", price_format_id=1, activation_date=date.today(), user="test")
        db.add_all([price_format, product, universal_list])
        db.flush()
        price_list.price_format_id = price_format.id
        db.add(price_list)
        db.flush()
        db.add_all(
            [
                UniversalListPriceFormat(universal_list_id=universal_list.id, price_format_id=price_format.id),
                ListItem(universal_list_id=universal_list.id, product_id=product.id, value=1),
                CalculatedPrice(
                    price_list_id=price_list.id,
                    product_id=product.id,
                    cost=10,
                    base_price=10,
                    final_price=11,
                    applied_list_id=universal_list.id,
                    applied_list_ids=json.dumps([universal_list.id, 999]),
                ),
            ]
        )
        db.commit()
        list_id = universal_list.id
    finally:
        db.close()

    response = client.delete(f"/api/universal-lists/{list_id}")

    assert response.status_code == 200
    db = Session()
    try:
        assert db.get(UniversalList, list_id) is None
        assert db.scalar(select(func.count(UniversalListPriceFormat.id)).where(UniversalListPriceFormat.universal_list_id == list_id)) == 0
        assert db.scalar(select(func.count(ListItem.id)).where(ListItem.universal_list_id == list_id)) == 0
        calculated = db.execute(select(CalculatedPrice)).scalars().one()
        assert calculated.applied_list_id is None
        assert json.loads(calculated.applied_list_ids) == [999]
    finally:
        db.close()


def test_lists_management_patch_accepts_dd_mm_yyyy_and_format_codes():
    client, Session = _client()
    db = Session()
    try:
        price_format_1 = PriceFormat(code="0001", name="Format 1", branch="A")
        price_format_2 = PriceFormat(code="003", name="Format 3", branch="B")
        universal_list = UniversalList(code="OLD", name="Old", type="fixed_price", status="active")
        db.add_all([price_format_1, price_format_2, universal_list])
        db.commit()
        list_id = universal_list.id
    finally:
        db.close()

    response = client.patch(
        f"/api/lists-management/{list_id}",
        json={
            "name": "Критичка 1",
            "code": "01",
            "type": "critical_markup",
            "active": False,
            "startDate": "13.06.2026",
            "endDate": "14.06.2026",
            "formatCodes": ["0001", "003"],
        },
    )

    assert response.status_code == 200
    db = Session()
    try:
        saved = db.get(UniversalList, list_id)
        assert saved is not None
        assert saved.name == "Критичка 1"
        assert saved.code == "01"
        assert saved.type != "fixed_price"
        assert saved.start_date.isoformat() == "2026-06-13"
        assert saved.end_date.isoformat() == "2026-06-14"
        assert db.scalar(select(func.count(UniversalListPriceFormat.id))) == 2
    finally:
        db.close()


def test_lists_management_patch_bad_date_returns_json_400():
    client, Session = _client()
    db = Session()
    try:
        universal_list = UniversalList(code="01", name="List", type="critical_markup", status="active")
        db.add(universal_list)
        db.commit()
        list_id = universal_list.id
    finally:
        db.close()

    response = client.patch(f"/api/lists-management/{list_id}", json={"startDate": "13/06/2026"})

    assert response.status_code == 400
    assert response.json() == {"detail": "date must be YYYY-MM-DD or DD.MM.YYYY"}


def test_lists_management_response_exposes_effective_status_from_dates():
    client, Session = _client()
    today = date.today()
    db = Session()
    try:
        db.add_all(
            [
                UniversalList(code="ACTIVE", name="Active", type="fixed_price", status="Активный", start_date=today - timedelta(days=1), end_date=today + timedelta(days=1)),
                UniversalList(code="EXPIRED", name="Expired", type="fixed_price", status="Активный", start_date=today - timedelta(days=3), end_date=today - timedelta(days=1)),
                UniversalList(code="FUTURE", name="Future", type="fixed_price", status="Активный", start_date=today + timedelta(days=1), end_date=None),
                UniversalList(code="OFF", name="Off", type="fixed_price", status="Неактивный", start_date=today - timedelta(days=1), end_date=today + timedelta(days=1)),
            ]
        )
        db.commit()
    finally:
        db.close()

    rows = {row["code"]: row for row in client.get("/api/lists-management").json()}

    assert rows["ACTIVE"]["rawStatus"] == "Активный"
    assert rows["ACTIVE"]["effectiveStatus"] == "active"
    assert rows["ACTIVE"]["status"] == "Активный"
    assert rows["ACTIVE"]["active"] is True
    assert rows["EXPIRED"]["rawStatus"] == "Активный"
    assert rows["EXPIRED"]["effectiveStatus"] == "expired"
    assert rows["EXPIRED"]["status"] == "Истёк"
    assert rows["EXPIRED"]["active"] is False
    assert rows["EXPIRED"]["dateValidity"]["expired"] is True
    assert rows["FUTURE"]["effectiveStatus"] == "not_started"
    assert rows["FUTURE"]["status"] == "Не начался"
    assert rows["FUTURE"]["active"] is False
    assert rows["FUTURE"]["dateValidity"]["startsInFuture"] is True
    assert rows["OFF"]["effectiveStatus"] == "inactive"
    assert rows["OFF"]["rawStatus"] == "Неактивный"

    active_rows = client.get("/api/lists-management?status=active").json()
    assert [row["code"] for row in active_rows] == ["ACTIVE"]
    expired_rows = client.get("/api/lists-management?status=expired").json()
    assert [row["code"] for row in expired_rows] == ["EXPIRED"]


def test_lists_management_import_excel_adds_items_to_existing_list():
    client, Session = _client()
    _seed_products(Session)
    create_response = client.post(
        "/api/lists-management",
        json={"code": "01", "name": "Критичка 1", "type": "critical_markup", "active": True},
    )
    assert create_response.status_code == 200
    list_id = create_response.json()["id"]

    response = client.post(
        f"/api/lists-management/{list_id}/import-excel",
        files={
            "file": (
                "critical.xlsx",
                _xlsx([["SKU", "Value"], ["12345", "20%"], ["UNKNOWN", "10%"], ["A-77", 0.3]]),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["list_id"] == list_id
    assert payload["summary"]["total_rows"] == 3
    assert payload["summary"]["processed"] == 2
    assert payload["summary"]["not_found"] == 1
    assert payload["item_count"] == 2

    card = client.get(f"/api/lists-management/{list_id}").json()
    assert card["itemsCount"] == 2
    values = {item["sku"]: item["value"] for item in card["items"]}
    assert values == {"12345": 20.0, "A-77": 30.0}


def test_memorandum_import_excel_accepts_business_headers_and_reports_row_errors():
    client, Session = _client()
    _seed_products(Session)
    created = client.post(
        "/api/lists-management",
        json={"code": "MEMO-IMPORT", "name": "Memorandum", "type": "memorandum", "active": True},
    )
    list_id = created.json()["id"]

    content = _xlsx(
        [
            ["Материал", "Артикул", "Производитель", "Меморандум"],
            ["12345", "Product 12345", "Maker", "6325,5"],
            ["A-77", "Article A77", "Maker", 1800],
            ["A-77", "Article A77 duplicate", "Maker", "1700.5"],
            ["", "No SKU", "Maker", 100],
            ["PC-9", "Bad price", "Maker", 0],
            ["UNKNOWN", "Unknown", "Maker", 200],
        ]
    )

    response = client.post(
        f"/api/lists-management/{list_id}/import-excel",
        files={"file": ("memo.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["success"] is False
    assert payload["totalRows"] == 6
    assert payload["importedRows"] == 2
    assert payload["updatedRows"] == 0
    assert payload["failedRows"] == 3
    assert payload["skippedRows"] == 3
    assert payload["summary"]["duplicates"] == 1
    assert payload["warnings"][0]["code"] == "duplicate_row"
    messages = [error["message"] for error in payload["errors"]]
    assert "Идентификатор товара обязателен" in messages
    assert "Цена должна быть больше нуля" in messages
    assert "Товар с указанным SKU не найден" in messages
    assert any(error["column"] == "Меморандум" for error in payload["errors"])

    card = client.get(f"/api/lists-management/{list_id}").json()
    values = {item["sku"]: item["value"] for item in card["items"]}
    assert values == {"12345": 6325.5, "A-77": 1700.5}


@pytest.mark.parametrize("price_header", ["Максимальная цена", "Цена меморандума", "Максимальная цена по меморандуму", "Предельная цена"])
@pytest.mark.parametrize("sku_header", ["Материал", "SKU", "Код", "GoodsID"])
def test_memorandum_import_excel_accepts_header_aliases(sku_header, price_header):
    client, Session = _client()
    _seed_products(Session)
    list_id = client.post(
        "/api/lists-management",
        json={"code": f"MEMO-{sku_header}-{price_header}", "name": "Memorandum", "type": "memorandum", "active": True},
    ).json()["id"]

    response = client.post(
        f"/api/lists-management/{list_id}/import-excel",
        files={
            "file": (
                "memo.xlsx",
                _xlsx([[sku_header, price_header], ["12345", "99,5"]]),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    assert response.json()["summary"]["processed"] == 1
    assert client.get(f"/api/lists-management/{list_id}").json()["items"][0]["value"] == 99.5


def test_memorandum_import_excel_invalid_file_returns_safe_400_not_500():
    client, Session = _client()
    _seed_products(Session)
    list_id = client.post(
        "/api/lists-management",
        json={"code": "MEMO-BAD-FILE", "name": "Memorandum", "type": "memorandum", "active": True},
    ).json()["id"]

    response = client.post(
        f"/api/lists-management/{list_id}/import-excel",
        files={"file": ("memo.xlsx", b"not an xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert response.status_code == 400
    assert response.status_code != 500


def test_memorandum_import_counts_only_meaningful_rows_with_styled_tail_beyond_limit(monkeypatch):
    client, Session = _client()
    _seed_products(Session)
    monkeypatch.setenv("LIST_IMPORT_MAX_ROWS", "2000")
    list_id = client.post(
        "/api/lists-management",
        json={"code": "MEMO-STYLED-TAIL", "name": "Memorandum", "type": "memorandum", "active": True},
    ).json()["id"]
    rows = [["ID", "Name", "Предельная цена для оптовой реализации"]]
    rows.extend([[f"9{idx:06d}", f"Registry item {idx}", 100] for idx in range(1091)])

    response = client.post(
        f"/api/lists-management/{list_id}/import-excel",
        files={
            "file": (
                "memo.xlsx",
                _xlsx_with_styled_tail(rows, styled_row=60000),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["totalRows"] == 1091
    assert payload["summary"]["meaningfulDataRows"] == 1091
    assert payload["summary"]["physicalRowsVisited"] > 50000
    assert payload["summary"]["emptyRowsIgnored"] > 50000


def test_memorandum_import_ignores_empty_styled_and_whitespace_rows(monkeypatch):
    client, Session = _client()
    _seed_products(Session)
    monkeypatch.setenv("LIST_IMPORT_MAX_ROWS", "1")
    list_id = client.post(
        "/api/lists-management",
        json={"code": "MEMO-EMPTY-ROWS", "name": "Memorandum", "type": "memorandum", "active": True},
    ).json()["id"]

    response = client.post(
        f"/api/lists-management/{list_id}/import-excel",
        files={
            "file": (
                "memo.xlsx",
                _xlsx_with_styled_tail(
                    [
                        ["Материал", "Артикул", "Меморандум"],
                        ["   ", "\t", "  "],
                        [None, None, None],
                        ["12345", "Product 12345", 100],
                    ],
                    styled_row=100,
                ),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["totalRows"] == 1
    assert payload["importedRows"] == 1
    assert payload["summary"]["emptyRowsIgnored"] >= 2


def test_universal_import_header_row_does_not_count_toward_row_limit(monkeypatch):
    client, Session = _client()
    _seed_products(Session)
    monkeypatch.setenv("LIST_IMPORT_MAX_ROWS", "1")
    list_id = client.post(
        "/api/lists-management",
        json={"code": "UL-HEADER-NOT-COUNTED", "name": "Fixed", "type": "fixed_price", "active": True},
    ).json()["id"]

    response = client.post(
        f"/api/lists-management/{list_id}/import-excel",
        files={"file": ("fixed.xlsx", _xlsx([["Материал", "Фикс цена"], ["12345", 100]]), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert response.status_code == 200
    assert response.json()["totalRows"] == 1


def test_universal_import_exact_row_limit_is_accepted(monkeypatch):
    client, Session = _client()
    _seed_products(Session)
    monkeypatch.setenv("LIST_IMPORT_MAX_ROWS", "3")
    list_id = client.post(
        "/api/lists-management",
        json={"code": "UL-EXACT-LIMIT", "name": "Fixed", "type": "fixed_price", "active": True},
    ).json()["id"]

    response = client.post(
        f"/api/lists-management/{list_id}/import-excel",
        files={
            "file": (
                "fixed.xlsx",
                _xlsx([["Материал", "Фикс цена"], ["12345", 100], ["A-77", 200], ["PC-9", 300]]),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    assert response.json()["totalRows"] == 3
    assert response.json()["importedRows"] == 3


def test_universal_import_meaningful_rows_over_limit_returns_structured_400_and_stops(monkeypatch):
    client, Session = _client()
    _seed_products(Session)
    monkeypatch.setenv("LIST_IMPORT_MAX_ROWS", "2")
    list_id = client.post(
        "/api/lists-management",
        json={"code": "UL-OVER-LIMIT", "name": "Fixed", "type": "fixed_price", "active": True},
    ).json()["id"]

    response = client.post(
        f"/api/lists-management/{list_id}/import-excel",
        files={
            "file": (
                "fixed.xlsx",
                _xlsx([["Материал", "Фикс цена"], ["12345", 100], [None, None], ["A-77", 200], ["PC-9", 300], ["UNKNOWN", "bad"]]),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 400
    assert response.status_code != 500
    detail = response.json()["detail"]
    assert detail["detail"] == "Количество заполненных строк превышает допустимый лимит 2"
    assert detail["limit"] == 2
    assert detail["meaningfulRowsDetected"] == 3
    assert client.get(f"/api/lists-management/{list_id}").json()["items"] == []


def _create_memorandum_list(client: TestClient, code: str = "MEMO-TEST") -> int:
    return client.post(
        "/api/lists-management",
        json={"code": code, "name": "Memorandum", "type": "memorandum", "active": True},
    ).json()["id"]


def test_memorandum_import_excel_accepts_gos_reestr_id_name_wholesale_price():
    client, Session = _client()
    _seed_products(Session)
    list_id = _create_memorandum_list(client, "MEMO-GOS-REESTR")

    response = client.post(
        f"/api/lists-management/{list_id}/import-excel",
        files={
            "file": (
                "Гос Реестр.xlsx",
                _xlsx(
                    [
                        ["ID", "Name", "Предельная цена для оптовой реализации"],
                        [1005594, "Ко-Ирбесан 300мг/12,5мг №14 таб", 6545.96],
                    ]
                ),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["importedRows"] == 1
    assert payload["matchedRows"] == 1
    assert payload["unmatchedRows"] == 0
    assert payload["selectedIdentifierColumn"] == "ID"
    assert payload["selectedPriceColumn"] == "Предельная цена для оптовой реализации"
    assert payload["summary"]["resolvedIdentifierFields"] == ["Product.code"]
    item = client.get(f"/api/lists-management/{list_id}").json()["items"][0]
    assert item["sku"] == "000000000001005594"
    assert item["value"] == 6545.96


@pytest.mark.parametrize("identifier_header", ["ID", "ИД", "ID товара", "Product ID", "ProductID"])
def test_memorandum_import_excel_resolves_new_id_aliases_by_normalized_product_code(identifier_header):
    client, Session = _client()
    _seed_products(Session)
    list_id = _create_memorandum_list(client, f"MEMO-ID-{identifier_header}")

    response = client.post(
        f"/api/lists-management/{list_id}/import-excel",
        files={
            "file": (
                "memo.xlsx",
                _xlsx([[identifier_header, "Меморандум"], [1005594, 100]]),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    assert response.json()["summary"]["resolvedIdentifierFields"] == ["Product.code"]
    assert client.get(f"/api/lists-management/{list_id}").json()["items"][0]["sku"] == "000000000001005594"


@pytest.mark.parametrize("identifier_header", ["Материал", "SKU", "Код", "GoodsID", "Код товара"])
def test_memorandum_import_excel_keeps_existing_code_identifier_aliases(identifier_header):
    client, Session = _client()
    _seed_products(Session)
    list_id = _create_memorandum_list(client, f"MEMO-CODE-{identifier_header}")

    response = client.post(
        f"/api/lists-management/{list_id}/import-excel",
        files={
            "file": (
                "memo.xlsx",
                _xlsx([[identifier_header, "Меморандум"], ["12345", "11,5"]]),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    assert response.json()["summary"]["resolvedIdentifierFields"] == ["Product.code"]
    assert client.get(f"/api/lists-management/{list_id}").json()["items"][0]["value"] == 11.5


@pytest.mark.parametrize(
    "price_header",
    [
        "Предельная цена для оптовой реализации",
        "Предельная цена для розничной реализации",
        "Предельная оптовая цена",
        "Предельная розничная цена",
        "Максимальная оптовая цена",
        "Максимальная розничная цена",
        "Цена для оптовой реализации",
        "Цена для розничной реализации",
        "Регулируемая цена",
        "Государственная предельная цена",
        "Цена по госреестру",
        "Предельная стоимость",
    ],
)
def test_memorandum_import_excel_accepts_new_price_aliases(price_header):
    client, Session = _client()
    _seed_products(Session)
    list_id = _create_memorandum_list(client, f"MEMO-PRICE-{price_header}")

    response = client.post(
        f"/api/lists-management/{list_id}/import-excel",
        files={
            "file": (
                "memo.xlsx",
                _xlsx([["ID", price_header], [1005594, "77.7"]]),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    assert response.json()["selectedPriceColumn"] == price_header
    assert client.get(f"/api/lists-management/{list_id}").json()["items"][0]["value"] == 77.7


def test_memorandum_import_excel_header_normalization_spaces_breaks_nbsp_and_yo():
    client, Session = _client()
    _seed_products(Session)
    list_id = _create_memorandum_list(client, "MEMO-NORMALIZED-HEADERS")

    response = client.post(
        f"/api/lists-management/{list_id}/import-excel",
        files={
            "file": (
                "memo.xlsx",
                _xlsx([[" id ", "Цена\u00a0по\nгосреёстру"], [1005594.0, "12,25"]]),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    assert client.get(f"/api/lists-management/{list_id}").json()["items"][0]["value"] == 12.25


def test_memorandum_import_excel_numeric_id_string_with_decimal_resolves_when_integer():
    client, Session = _client()
    _seed_products(Session)
    list_id = _create_memorandum_list(client, "MEMO-ID-DECIMAL-STRING")

    response = client.post(
        f"/api/lists-management/{list_id}/import-excel",
        files={
            "file": (
                "memo.xlsx",
                _xlsx([["ID", "Меморандум"], ["1005594.0", "20,5"]]),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    assert client.get(f"/api/lists-management/{list_id}").json()["items"][0]["sku"] == "000000000001005594"


def test_memorandum_import_excel_unknown_id_reports_row_error():
    client, Session = _client()
    _seed_products(Session)
    list_id = _create_memorandum_list(client, "MEMO-UNKNOWN-ID")

    response = client.post(
        f"/api/lists-management/{list_id}/import-excel",
        files={
            "file": (
                "memo.xlsx",
                _xlsx([["ID", "Name", "Меморандум"], [1003437, "FDP MEDLAC 5г фл+р-р", 20]]),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["failedRows"] == 1
    assert payload["unmatchedRows"] == 1
    assert payload["rowErrors"][0]["identifier"] == "1003437"
    assert payload["rowErrors"][0]["name"] == "FDP MEDLAC 5г фл+р-р"
    assert payload["rowErrors"][0]["message"] == "Товар с указанным ID отсутствует в локальном справочнике"


def test_memorandum_import_excel_missing_products_do_not_abort_valid_registry_rows():
    client, Session = _client()
    _seed_products(Session)
    list_id = _create_memorandum_list(client, "MEMO-MIXED-REGISTRY")

    response = client.post(
        f"/api/lists-management/{list_id}/import-excel",
        files={
            "file": (
                "memo.xlsx",
                _xlsx(
                    [
                        ["ID", "Name", "Предельная цена для оптовой реализации"],
                        [1003437, "FDP MEDLAC 5г фл+р-р", 10],
                        [1005594, "Ко-Ирбесан 300мг/12,5мг №14 таб", 20],
                    ]
                ),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    payload = response.json()
    assert response.status_code == 200
    assert payload["matchedRows"] == 1
    assert payload["unmatchedRows"] == 1
    assert payload["importedRows"] == 1
    assert client.get(f"/api/lists-management/{list_id}").json()["items"][0]["value"] == 20.0


def test_memorandum_import_excel_does_not_use_provisor_goods_id_for_registry_id():
    client, Session = _client()
    _seed_products(Session)
    list_id = _create_memorandum_list(client, "MEMO-NO-PROVISOR-ID")

    response = client.post(
        f"/api/lists-management/{list_id}/import-excel",
        files={
            "file": (
                "memo.xlsx",
                _xlsx([["ID", "Меморандум"], [1003437, 100]]),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    payload = response.json()
    assert response.status_code == 200
    assert payload["matchedRows"] == 0
    assert payload["unmatchedRows"] == 1
    assert client.get(f"/api/lists-management/{list_id}").json()["items"] == []


def test_memorandum_import_excel_does_not_match_by_name_alone():
    client, Session = _client()
    _seed_products(Session)
    list_id = _create_memorandum_list(client, "MEMO-NO-NAME-MATCH")

    response = client.post(
        f"/api/lists-management/{list_id}/import-excel",
        files={
            "file": (
                "memo.xlsx",
                _xlsx([["ID", "Name", "Меморандум"], [9999999, "Ко-Ирбесан 300мг/12,5мг №14 таб", 100]]),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    assert response.json()["unmatchedRows"] == 1
    assert client.get(f"/api/lists-management/{list_id}").json()["items"] == []


@pytest.mark.parametrize("price", ["4 699,50", "4\u00a0699,50", "4699.50"])
def test_memorandum_import_excel_registry_prices_parse_spaces_nbsp_comma_and_dot(price):
    client, Session = _client()
    _seed_products(Session)
    list_id = _create_memorandum_list(client, f"MEMO-PRICE-FORMAT-{price}")

    response = client.post(
        f"/api/lists-management/{list_id}/import-excel",
        files={
            "file": (
                "memo.xlsx",
                _xlsx([["ID", "Предельная цена для оптовой реализации"], [1005594, price]]),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    assert client.get(f"/api/lists-management/{list_id}").json()["items"][0]["value"] == 4699.5


def test_memorandum_import_excel_duplicate_registry_ids_use_last_row():
    client, Session = _client()
    _seed_products(Session)
    list_id = _create_memorandum_list(client, "MEMO-DUP-REGISTRY-ID")

    response = client.post(
        f"/api/lists-management/{list_id}/import-excel",
        files={
            "file": (
                "memo.xlsx",
                _xlsx(
                    [
                        ["ID", "Name", "Меморандум"],
                        [1005594, "Ко-Ирбесан 300мг/12,5мг №14 таб", 100],
                        [1005594, "Ко-Ирбесан 300мг/12,5мг №14 таб", 200],
                    ]
                ),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    payload = response.json()
    assert response.status_code == 200
    assert payload["duplicateRows"] == 1
    assert payload["warnings"][0]["code"] == "duplicate_row"
    assert client.get(f"/api/lists-management/{list_id}").json()["items"][0]["value"] == 200.0


def test_memorandum_import_excel_multiple_price_columns_returns_422():
    client, Session = _client()
    _seed_products(Session)
    list_id = _create_memorandum_list(client, "MEMO-MULTI-PRICE")

    response = client.post(
        f"/api/lists-management/{list_id}/import-excel",
        files={
            "file": (
                "memo.xlsx",
                _xlsx(
                    [
                        ["ID", "Предельная цена для розничной реализации", "Предельная цена для оптовой реализации"],
                        [1005594, 999, 111],
                    ]
                ),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 422
    detail = response.json()["detail"]
    assert detail["detail"] == "Найдено несколько колонок предельной цены"
    assert detail["detectedPriceColumns"] == [
        "Предельная цена для розничной реализации",
        "Предельная цена для оптовой реализации",
    ]


def test_memorandum_import_excel_unsupported_headers_return_structured_422_not_500():
    client, Session = _client()
    _seed_products(Session)
    list_id = _create_memorandum_list(client, "MEMO-BAD-HEADERS")

    response = client.post(
        f"/api/lists-management/{list_id}/import-excel",
        files={
            "file": (
                "memo.xlsx",
                _xlsx([["Unknown", "Some Price"], ["12345", 10]]),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 422
    assert response.status_code != 500
    detail = response.json()["detail"]
    assert detail["detectedHeaders"] == ["Unknown", "Some Price"]
    assert "product_identifier" in detail["missingFields"]
    assert "product_code" in detail["supportedAliases"]


def test_critical_markup_manual_save_accepts_dash_and_keeps_empty_invalid():
    client, Session = _client()
    _seed_products(Session)
    created = client.post(
        "/api/lists-management",
        json={"name": "Critical", "type": "critical_markup", "active": True},
    )
    list_id = created.json()["id"]

    saved = client.post(
        f"/api/lists-management/{list_id}/items",
        json={"sku": "12345", "value": "-"},
    )
    assert saved.status_code == 200
    card = client.get(f"/api/lists-management/{list_id}").json()
    assert card["items"][0]["value"] == "-"
    assert card["items"][0]["valueDisplay"] == "-"

    numeric = client.post(
        f"/api/lists-management/{list_id}/items",
        json={"sku": "12345", "value": "7.5"},
    )
    assert numeric.status_code == 200
    assert client.get(f"/api/lists-management/{list_id}").json()["items"][0]["value"] == 7.5

    empty = client.post(
        f"/api/lists-management/{list_id}/items",
        json={"sku": "A-77", "value": ""},
    )
    assert empty.status_code == 400


def test_critical_markup_excel_import_accepts_dash():
    client, Session = _client()
    _seed_products(Session)
    created = client.post(
        "/api/lists-management",
        json={"name": "Critical", "type": "critical_markup", "active": True},
    )
    list_id = created.json()["id"]

    response = client.post(
        f"/api/lists-management/{list_id}/import-excel",
        files={
            "file": (
                "critical.xlsx",
                _xlsx([["SKU", "Value"], ["12345", "-"], ["A-77", 5]]),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    assert response.json()["summary"]["processed"] == 2
    assert response.json()["summary"]["invalid_rows"] == 0
    values = {
        item["sku"]: item["value"]
        for item in client.get(f"/api/lists-management/{list_id}").json()["items"]
    }
    assert values == {"12345": "-", "A-77": 5.0}
    db = Session()
    try:
        assert db.scalar(select(func.count(BusinessList.id))) == 0
        assert db.scalar(select(func.count(ListItem.id)).where(ListItem.universal_list_id == list_id)) == 2
    finally:
        db.close()


@pytest.mark.parametrize(
    ("list_type", "header"),
    [
        ("max_price", "Макс цена"),
        ("max_price", "Максимальная цена"),
        ("max_price", "Цена"),
        ("min_price", "Мин цена"),
        ("min_price", "Минимальная цена"),
        ("min_price", "Цена"),
        ("fixed_price", "Фикс цена"),
        ("fixed_price", "Фиксированная цена"),
        ("fixed_price", "Цена"),
    ],
)
def test_lists_management_import_excel_accepts_price_type_business_headers(list_type, header):
    client, Session = _client()
    _seed_products(Session)
    create_response = client.post(
        "/api/lists-management",
        json={"code": f"UL-{list_type}", "name": list_type, "type": list_type, "active": True},
    )
    list_id = create_response.json()["id"]

    response = client.post(
        f"/api/lists-management/{list_id}/import-excel",
        files={"file": ("prices.xlsx", _xlsx([["Материал", header], ["12345", 1234.5]]), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert response.status_code == 200
    assert response.json()["summary"]["processed"] == 1
    card = client.get(f"/api/lists-management/{list_id}").json()
    assert card["items"][0]["value"] == 1234.5


@pytest.mark.parametrize(
    ("list_type", "expected_detail"),
    [
        ("max_price", "Для списка типа Максимальная цена нужна колонка значения:\nМакс цена / Максимальная цена / Цена"),
        ("min_price", "Для списка типа Минимальная цена нужна колонка значения:\nМин цена / Минимальная цена / Цена"),
        ("fixed_price", "Для списка типа Фиксированная цена нужна колонка значения:\nФикс цена / Фиксированная цена / Цена"),
    ],
)
def test_lists_management_import_excel_returns_business_error_for_missing_price_value_column(list_type, expected_detail):
    client, Session = _client()
    _seed_products(Session)
    create_response = client.post(
        "/api/lists-management",
        json={"code": f"UL-{list_type}", "name": list_type, "type": list_type, "active": True},
    )
    list_id = create_response.json()["id"]

    response = client.post(
        f"/api/lists-management/{list_id}/import-excel",
        files={"file": ("prices.xlsx", _xlsx([["Материал"], ["12345"]]), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert response.status_code == 400
    detail = response.json()["detail"]
    assert expected_detail in detail
    assert "Обнаружены заголовки: материал" in detail
    assert "Ожидался один из заголовков:" in detail


@pytest.mark.parametrize("header", ["Фикс наценка", "Критичка"])
def test_fixed_markup_import_accepts_template_value_headers(header):
    client, Session = _client()
    _seed_products(Session)
    list_id = client.post(
        "/api/lists-management",
        json={"code": f"UL-FM-{header}", "name": "Fixed markup", "type": "fixed_markup", "active": True},
    ).json()["id"]

    response = client.post(
        f"/api/lists-management/{list_id}/import-excel",
        files={"file": ("fixed-markup.xlsx", _xlsx([
            ["Материал", "Артикул", "Производитель", header],
            ["12345", "", "Maker", 10],
        ]), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert response.status_code == 200
    assert response.json()["summary"]["processed"] == 1


@pytest.mark.parametrize("raw_value", ["10,5", "10.5"])
def test_fixed_markup_import_accepts_comma_and_dot_decimals(raw_value):
    client, Session = _client()
    _seed_products(Session)
    list_id = client.post(
        "/api/lists-management",
        json={"code": f"UL-FM-{raw_value}", "name": "Fixed markup", "type": "fixed_markup", "active": True},
    ).json()["id"]

    response = client.post(
        f"/api/lists-management/{list_id}/import-excel",
        files={"file": ("fixed-markup.xlsx", _xlsx([["Материал", "Фикс наценка"], ["12345", raw_value]]), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert response.status_code == 200
    card = client.get(f"/api/lists-management/{list_id}").json()
    assert card["items"][0]["value"] == 10.5


def test_fixed_markup_unknown_header_returns_detected_and_expected_aliases():
    client, Session = _client()
    _seed_products(Session)
    list_id = client.post(
        "/api/lists-management",
        json={"code": "UL-FM-UNKNOWN", "name": "Fixed markup", "type": "fixed_markup", "active": True},
    ).json()["id"]

    response = client.post(
        f"/api/lists-management/{list_id}/import-excel",
        files={"file": ("fixed-markup.xlsx", _xlsx([["Материал", "Артикул", "Производитель", "Неизвестно"], ["12345", "", "Maker", 10]]), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert response.status_code == 400
    detail = response.json()["detail"]
    assert "Обнаружены заголовки: материал, артикул, производитель, неизвестно" in detail
    assert "фикс наценка" in detail
    assert "критичка" in detail


def test_invalid_markup_decimal_returns_json_400_with_clear_detail():
    client, Session = _client()
    _seed_products(Session)
    list_id = client.post(
        "/api/lists-management",
        json={"code": "UL-FM-BAD", "name": "Fixed markup", "type": "fixed_markup", "active": True},
    ).json()["id"]

    response = client.post(
        f"/api/lists-management/{list_id}/import-excel",
        files={"file": ("fixed-markup.xlsx", _xlsx([["Материал", "Фикс наценка"], ["12345", "ten"]]), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert response.status_code == 400
    assert response.headers["content-type"].startswith("application/json")
    assert "Некорректное значение в строке 2" in response.json()["detail"]


@pytest.mark.parametrize(
    ("list_type", "raw_value", "expected_value", "expected_display"),
    [
        ("critical_markup", 1, 1.0, "1%"),
        ("fixed_markup", 1, 1.0, "1%"),
        ("fixed_markup", 33, 33.0, "33%"),
        ("fixed_markup", 0.2, 20.0, "20%"),
        ("fixed_price", 2500, 2500.0, "2500"),
        ("fixed_price", 120.5, 120.5, "120.5"),
        ("fixed_price", 1000, 1000.0, "1000"),
        ("fixed_price", "1000.5", 1000.5, "1000.5"),
        ("fixed_price", "1000,5", 1000.5, "1000.5"),
        ("min_price", 1200, 1200.0, "1200"),
        ("max_price", 3000, 3000.0, "3000"),
        ("min_markup", 2, 2.0, "2%"),
        ("max_markup", 0.2, 20.0, "20%"),
        ("percentile_override", 0.2, 20.0, "20%"),
        ("exclude_from_pricing", 1, 1.0, "Да"),
        ("exclude_from_pricing", "нет", 0.0, "Нет"),
        ("no_bend", "да", 1.0, "Да"),
        ("no_bend", 0, 0.0, "Нет"),
    ],
)
def test_lists_management_import_excel_normalizes_by_selected_list_type(list_type, raw_value, expected_value, expected_display):
    client, Session = _client()
    _seed_products(Session)
    create_response = client.post(
        "/api/lists-management",
        json={"code": f"UL-{list_type}", "name": list_type, "type": list_type, "active": True},
    )
    assert create_response.status_code == 200
    list_id = create_response.json()["id"]

    response = client.post(
        f"/api/lists-management/{list_id}/import-excel",
        files={
            "file": (
                "list.xlsx",
                _xlsx([["Material", "Value"], ["12345", raw_value]]),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    assert response.json()["summary"]["processed"] == 1
    card = client.get(f"/api/lists-management/{list_id}").json()
    assert card["items"][0]["value"] == expected_value
    assert card["items"][0]["valueDisplay"] == expected_display
    db = Session()
    try:
        saved = db.get(UniversalList, list_id)
        assert saved is not None
        assert saved.type == list_type
    finally:
        db.close()


def test_lists_management_manual_item_accepts_comma_decimal():
    client, Session = _client()
    _seed_products(Session)
    create_response = client.post(
        "/api/lists-management",
        json={"code": "UL-COMMA", "name": "Comma Decimal", "type": "fixed_price", "active": True},
    )
    assert create_response.status_code == 200
    list_id = create_response.json()["id"]

    response = client.post(f"/api/lists-management/{list_id}/items", json={"sku": "12345", "value": "1000,5"})

    assert response.status_code == 200
    card = client.get(f"/api/lists-management/{list_id}").json()
    assert card["items"][0]["value"] == 1000.5


def test_lists_management_manual_item_normalizes_sku_before_linking_product():
    client, Session = _client()
    _seed_products(Session)
    create_response = client.post(
        "/api/lists-management",
        json={"code": "UL-SKU-NORMALIZED", "name": "Normalized SKU", "type": "fixed_markup", "active": True},
    )
    list_id = create_response.json()["id"]

    response = client.post(f"/api/lists-management/{list_id}/items", json={"sku": "222", "value": "5,5"})

    assert response.status_code == 200
    card = client.get(f"/api/lists-management/{list_id}").json()
    assert card["items"][0]["sku"] == "000000000000000222"
    assert card["items"][0]["value"] == 5.5


def test_exclude_from_pricing_manual_item_allows_sku_without_value():
    client, Session = _client()
    _seed_products(Session)
    create_response = client.post(
        "/api/lists-management",
        json={"code": "UL-EXCLUDE-MANUAL", "name": "Exclude Manual", "type": "exclude_from_pricing", "active": True},
    )
    list_id = create_response.json()["id"]

    response = client.post(f"/api/lists-management/{list_id}/items", json={"sku": "12345"})

    assert response.status_code == 200
    card = client.get(f"/api/lists-management/{list_id}").json()
    assert card["items"][0]["sku"] == "12345"
    assert card["items"][0]["value"] == 1.0
    assert card["items"][0]["valueDisplay"] == "Да"


def test_exclude_from_pricing_sku_only_excel_is_absent_from_generation_and_export():
    client, Session = _client()
    db = Session()
    try:
        price_format = PriceFormat(code="PF-EXCLUDE", name="PF Exclude", branch="Almaty")
        excluded = Product(code="EXCLUDED-SKU", name="Excluded", cost=100)
        included = Product(code="INCLUDED-SKU", name="Included", cost=100)
        db.add_all([price_format, excluded, included])
        db.flush()
        db.add(MarkupRange(price_format_id=price_format.id, cost_from=0, cost_to=None, markup_percent=0.15))
        db.commit()
    finally:
        db.close()

    create_response = client.post(
        "/api/lists-management",
        json={
            "code": "UL-EXCLUDE-SKU-ONLY",
            "name": "Exclude SKU only",
            "type": "exclude_from_pricing",
            "active": True,
            "formatCodes": ["PF-EXCLUDE"],
        },
    )
    assert create_response.status_code == 200
    list_id = create_response.json()["id"]
    import_response = client.post(
        f"/api/lists-management/{list_id}/import-excel",
        files={
            "file": (
                "exclude.xlsx",
                _xlsx([["SKU"], ["EXCLUDED-SKU"]]),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert import_response.status_code == 200
    assert import_response.json()["summary"]["processed"] == 1

    db = Session()
    try:
        imported_item = db.execute(
            select(ListItem)
            .join(Product, Product.id == ListItem.product_id)
            .where(ListItem.universal_list_id == list_id)
            .where(Product.code == "EXCLUDED-SKU")
        ).scalars().one()
        assert float(imported_item.value) == 1.0
        _activate_all_products_for_generation(db)

        count = calculate_prices(
            db=db,
            price_format_code="PF-EXCLUDE",
            price_list_number="PL-EXCLUDE-SKU-ONLY",
            as_of=date.today(),
            activation_date=None,
            user="test",
            force_new_price_list=True,
        )
        assert count == 1
        calculated_skus = db.execute(
            select(Product.code)
            .join(CalculatedPrice, CalculatedPrice.product_id == Product.id)
        ).scalars().all()
        assert calculated_skus == ["INCLUDED-SKU"]

        _pl, _pf, export_rows, _competitor_columns = _generated_price_export_rows(
            db,
            "PL-EXCLUDE-SKU-ONLY",
        )
        assert [row["sku"] for row in export_rows] == ["INCLUDED-SKU"]
    finally:
        db.close()

    csv_response = client.get("/api/price-lists/PL-EXCLUDE-SKU-ONLY/export.csv")
    assert csv_response.status_code == 200
    assert "INCLUDED-SKU" in csv_response.text
    assert "EXCLUDED-SKU" not in csv_response.text

    xlsx_response = client.get("/api/price-lists/PL-EXCLUDE-SKU-ONLY/export.xlsx")
    assert xlsx_response.status_code == 200
    workbook = load_workbook(io.BytesIO(xlsx_response.content), data_only=True)
    exported_cells = [str(cell.value or "") for row in workbook.active.iter_rows() for cell in row]
    assert "INCLUDED-SKU" in exported_cells
    assert "EXCLUDED-SKU" not in exported_cells


def test_generated_csv_and_xlsx_export_keep_only_pricing_log_without_list():
    client, Session = _client()
    db = Session()
    try:
        price_format = PriceFormat(code="PF-EXPORT-NO-LIST", name="PF Export No List", branch="Almaty")
        product = Product(code="SKU-NO-LIST", name="No List", cost=100)
        db.add_all([price_format, product])
        db.flush()
        db.add(MarkupRange(price_format_id=price_format.id, cost_from=0, cost_to=None, markup_percent=0.15))
        _activate_all_products_for_generation(db)
        calculate_prices(
            db=db,
            price_format_code="PF-EXPORT-NO-LIST",
            price_list_number="PL-EXPORT-NO-LIST",
            as_of=date.today(),
            activation_date=None,
            user="test",
            force_new_price_list=True,
        )
    finally:
        db.close()

    csv_response = client.get("/api/price-lists/PL-EXPORT-NO-LIST/export.csv")
    assert csv_response.status_code == 200
    assert "SKU-NO-LIST" in csv_response.text
    assert "Применен список" not in csv_response.text

    xlsx_response = client.get("/api/price-lists/PL-EXPORT-NO-LIST/export.xlsx")
    assert xlsx_response.status_code == 200
    workbook = load_workbook(io.BytesIO(xlsx_response.content), data_only=True)
    exported_cells = [str(cell.value or "") for row in workbook.active.iter_rows() for cell in row]
    assert any("SKU-NO-LIST" in cell for cell in exported_cells)
    assert not any("Применен список" in cell for cell in exported_cells)


def test_generated_csv_and_xlsx_export_append_list_log_when_list_applies():
    client, Session = _client()
    db = Session()
    try:
        price_format = PriceFormat(code="PF-EXPORT-LIST", name="PF Export List", branch="Almaty")
        product = Product(code="SKU-WITH-LIST", name="With List", cost=100)
        db.add_all([price_format, product])
        db.flush()
        db.add(MarkupRange(price_format_id=price_format.id, cost_from=0, cost_to=None, markup_percent=0.15))
        universal_list = UniversalList(
            code="UL_EXPORT_FIXED_MARKUP",
            name="Export Fixed Markup",
            type="fixed_markup",
            status="active",
            price_format_id=price_format.id,
        )
        db.add(universal_list)
        db.flush()
        db.add(ListItem(universal_list_id=universal_list.id, product_id=product.id, value=5.5))
        _activate_all_products_for_generation(db)
        calculate_prices(
            db=db,
            price_format_code="PF-EXPORT-LIST",
            price_list_number="PL-EXPORT-LIST",
            as_of=date.today(),
            activation_date=None,
            user="test",
            force_new_price_list=True,
        )
    finally:
        db.close()

    csv_response = client.get("/api/price-lists/PL-EXPORT-LIST/export.csv")
    assert csv_response.status_code == 200
    assert "SKU-WITH-LIST" in csv_response.text
    assert "Применен список" in csv_response.text
    assert "Export Fixed Markup" in csv_response.text
    assert "fixed_markup" in csv_response.text
    assert "5.5%" in csv_response.text
    assert "Конкуренты и прогиб не применялись" in csv_response.text

    xlsx_response = client.get("/api/price-lists/PL-EXPORT-LIST/export.xlsx")
    assert xlsx_response.status_code == 200
    workbook = load_workbook(io.BytesIO(xlsx_response.content), data_only=True)
    exported_cells = [str(cell.value or "") for row in workbook.active.iter_rows() for cell in row]
    joined = "\n".join(exported_cells)
    assert "SKU-WITH-LIST" in joined
    assert "Применен список" in joined
    assert "Export Fixed Markup" in joined
    assert "fixed_markup" in joined
    assert "5.5%" in joined
    assert "Конкуренты и прогиб не применялись" in joined
