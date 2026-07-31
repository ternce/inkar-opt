from __future__ import annotations

import csv
import io
from datetime import date, timedelta
from decimal import Decimal

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine, event
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from backend.app.db import Base
from backend.app.deps import get_db
from backend.app.main import app
from backend.app.models import (
    BranchCost,
    BranchStock,
    CalculatedPrice,
    ListItem,
    MarkupRange,
    PriceFormat,
    PriceList,
    Product,
    ReferenceUpdateStatus,
    RoundingRule,
    UniversalList,
    UniversalListPriceFormat,
)
from backend.app.services.pricing import (
    LIST_TYPE_MEMORANDUM,
    _active_lists_for_format,
    _build_pricing_preload,
    calculate_price_for_product,
    calculate_prices,
)


def _session_factory():
    engine = create_engine("sqlite://", connect_args={"check_same_thread": False}, poolclass=StaticPool)
    Base.metadata.create_all(engine)
    return engine, sessionmaker(bind=engine)


def _pf(db, *, code="PF", rounding_rule=None):
    row = PriceFormat(code=code, name=code, branch="1", rounding_rule_id=rounding_rule.id if rounding_rule else None)
    db.add(row)
    db.flush()
    db.add(MarkupRange(price_format_id=row.id, cost_from=0, cost_to=None, markup_percent=Decimal("0.20")))
    return row


def _product(db, *, code="SKU", cost=Decimal("100")):
    row = Product(code=code, name=code, cost=cost)
    db.add(row)
    db.flush()
    return row


def _list(db, pf, product, list_type, value, *, active=True, start=None, end=None, linked=True, name=None):
    row = UniversalList(
        code=f"UL-{list_type}-{product.code}-{value}",
        name=name or f"{list_type} {value}",
        type=list_type,
        status="active" if active else "inactive",
        start_date=start,
        end_date=end,
    )
    db.add(row)
    db.flush()
    if linked:
        db.add(UniversalListPriceFormat(universal_list_id=row.id, price_format_id=pf.id))
    db.add(ListItem(universal_list_id=row.id, product_id=product.id, value=Decimal(str(value))))
    db.flush()
    return row


def _price(db, pf, product, *, as_of=None):
    as_of = as_of or date.today()
    active = _active_lists_for_format(db, int(pf.id), as_of)
    preload = _build_pricing_preload(db, price_format=pf, products=[product], active_lists=active, branch_id="1")
    return calculate_price_for_product(
        db=db,
        product=product,
        price_format=pf,
        as_of=as_of,
        region_id=1,
        active_lists=active,
        cost_override=product.cost,
        pricing_preload=preload,
    )


def _activate_products_for_generation(db, products, *, branch_id="1"):
    for product in products:
        db.add(BranchStock(branch_id=branch_id, product_id=product.id, sku=product.code, stock=1))
        db.add(BranchCost(branch_id=branch_id, product_id=product.id, sku=product.code, cost=product.cost))
    db.add(ReferenceUpdateStatus(branch_id=branch_id, branch_name=branch_id, data_type="stock", status="success", rows_count=len(products)))
    db.add(ReferenceUpdateStatus(branch_id=branch_id, branch_name=branch_id, data_type="cost", status="success", rows_count=len(products)))
    db.flush()


def test_create_memorandum_list_and_reject_non_positive_item():
    engine, Session = _session_factory()

    def override_db():
        with Session() as db:
            yield db

    app.dependency_overrides[get_db] = override_db
    with Session() as db:
        pf = _pf(db)
        product = _product(db)
        db.commit()
        pf_code = pf.code
        sku = product.code

    client = TestClient(app)
    created = client.post(
        "/api/lists-management",
        json={"name": "Меморандум 2026", "type": "memorandum", "active": True, "formatCodes": [pf_code]},
    )
    assert created.status_code == 200
    list_id = created.json()["id"]

    invalid = client.post(f"/api/lists-management/{list_id}/items", json={"sku": sku, "value": 0})
    assert invalid.status_code == 400
    valid = client.post(f"/api/lists-management/{list_id}/items", json={"sku": sku, "value": 115})
    assert valid.status_code == 200

    details = client.get(f"/api/lists-management/{list_id}").json()
    assert details["type"] == "memorandum"
    assert details["typeLabel"] == "Меморандум"
    app.dependency_overrides.pop(get_db, None)
    engine.dispose()


def test_memorandum_assigned_to_format_caps_final_price_and_records_fields():
    _engine, Session = _session_factory()
    with Session() as db:
        pf = _pf(db)
        product = _product(db, cost=Decimal("100"))
        memo = _list(db, pf, product, LIST_TYPE_MEMORANDUM, 115, name="Меморандум 2026")
        db.commit()

        price, debug = _price(db, pf, product)

        assert price == Decimal("115.00")
        assert debug["memorandum_applied"] is True
        assert debug["memorandum_max_price"] == Decimal("115.000000")
        assert debug["memorandum_below_mdc"] is True
        assert debug["memorandum_list_id"] == memo.id
        assert debug["memorandum_diagnostic_code"] == "memorandum_below_mdc"
        assert "меморандум" in debug["log"].casefold()


def test_unassigned_expired_and_future_memorandum_do_not_apply():
    _engine, Session = _session_factory()
    today = date.today()
    with Session() as db:
        pf = _pf(db)
        other_pf = _pf(db, code="OTHER")
        product = _product(db)
        _list(db, other_pf, product, LIST_TYPE_MEMORANDUM, 90)
        _list(db, pf, product, LIST_TYPE_MEMORANDUM, 91, start=today - timedelta(days=10), end=today - timedelta(days=1))
        _list(db, pf, product, LIST_TYPE_MEMORANDUM, 92, start=today + timedelta(days=1))
        db.commit()

        price, debug = _price(db, pf, product, as_of=today)

        assert price == Decimal("125.00")
        assert debug["memorandum_max_price"] is None


def test_unassigned_memorandum_list_is_global_by_design():
    _engine, Session = _session_factory()
    with Session() as db:
        pf = _pf(db)
        product = _product(db)
        _list(db, pf, product, LIST_TYPE_MEMORANDUM, 115, linked=False)
        db.commit()

        price, debug = _price(db, pf, product)

        assert price == Decimal("115.000000")
        assert debug["memorandum_applied"] is True


def test_memorandum_below_candidate_caps_but_candidate_below_memorandum_stays_unchanged():
    _engine, Session = _session_factory()
    with Session() as db:
        pf = _pf(db)
        capped = _product(db, code="CAP", cost=Decimal("100"))
        unchanged = _product(db, code="KEEP", cost=Decimal("100"))
        _list(db, pf, capped, LIST_TYPE_MEMORANDUM, 115)
        _list(db, pf, unchanged, LIST_TYPE_MEMORANDUM, 140)
        db.commit()

        capped_price, capped_debug = _price(db, pf, capped)
        unchanged_price, unchanged_debug = _price(db, pf, unchanged)

        assert capped_price == Decimal("115.000000")
        assert capped_debug["memorandum_applied"] is True
        assert unchanged_price == Decimal("125.00")
        assert unchanged_debug["memorandum_applied"] is False


def test_memorandum_does_not_filter_generation_assortment_and_caps_only_matching_products():
    _engine, Session = _session_factory()
    with Session() as db:
        pf = _pf(db)
        capped = _product(db, code="MEMO-CAPPED", cost=Decimal("100"))
        no_memo = _product(db, code="NO-MEMO", cost=Decimal("100"))
        memo_above_price = _product(db, code="MEMO-ABOVE", cost=Decimal("100"))
        _list(db, pf, capped, LIST_TYPE_MEMORANDUM, 115)
        _list(db, pf, memo_above_price, LIST_TYPE_MEMORANDUM, 140)
        _activate_products_for_generation(db, [capped, no_memo, memo_above_price])
        db.commit()

        count = calculate_prices(
            db=db,
            price_format_code=pf.code,
            price_list_number="PL-MEMO-NOT-FILTER",
            as_of=date.today(),
            activation_date=None,
            user="test",
            force_new_price_list=True,
        )

        rows = {
            product_code: row
            for product_code, row in db.query(Product.code, CalculatedPrice)
            .join(CalculatedPrice, CalculatedPrice.product_id == Product.id)
            .all()
        }
        assert count == 3
        assert set(rows) == {"MEMO-CAPPED", "NO-MEMO", "MEMO-ABOVE"}
        assert float(rows["MEMO-CAPPED"].final_price) == 115.0
        assert rows["MEMO-CAPPED"].memorandum_applied is True
        assert float(rows["NO-MEMO"].final_price) == 125.0
        assert rows["NO-MEMO"].memorandum_max_price is None
        assert rows["NO-MEMO"].memorandum_applied is False
        assert float(rows["MEMO-ABOVE"].final_price) == 125.0
        assert float(rows["MEMO-ABOVE"].memorandum_max_price) == 140.0
        assert rows["MEMO-ABOVE"].memorandum_applied is False


def test_fixed_and_min_price_above_memorandum_are_capped():
    _engine, Session = _session_factory()
    with Session() as db:
        pf = _pf(db)
        fixed = _product(db, code="FIXED")
        min_price = _product(db, code="MIN")
        fixed_markup = _product(db, code="FMARKUP")
        _list(db, pf, fixed, "fixed_price", 130)
        _list(db, pf, fixed, LIST_TYPE_MEMORANDUM, 115)
        _list(db, pf, min_price, "min_price", 140)
        _list(db, pf, min_price, LIST_TYPE_MEMORANDUM, 115)
        _list(db, pf, fixed_markup, "fixed_markup", 30)
        _list(db, pf, fixed_markup, LIST_TYPE_MEMORANDUM, 115)
        db.commit()

        fixed_price, fixed_debug = _price(db, pf, fixed)
        min_result, min_debug = _price(db, pf, min_price)
        markup_result, markup_debug = _price(db, pf, fixed_markup)

        assert fixed_price == Decimal("115.000000")
        assert fixed_debug["price_before_memorandum"] == Decimal("130.000000")
        assert min_result == Decimal("115.000000")
        assert min_debug["price_before_memorandum"] == Decimal("140.000000")
        assert markup_result == Decimal("115.000000")
        assert markup_debug["price_before_memorandum"] == Decimal("142.86")


def test_lower_normal_max_price_wins_before_memorandum():
    _engine, Session = _session_factory()
    with Session() as db:
        pf = _pf(db)
        product = _product(db)
        _list(db, pf, product, "max_price", 110)
        _list(db, pf, product, LIST_TYPE_MEMORANDUM, 115)
        db.commit()

        price, debug = _price(db, pf, product)

        assert price == Decimal("110.000000")
        assert debug["memorandum_max_price"] == Decimal("115.000000")
        assert debug["memorandum_applied"] is False


def test_multiple_memorandum_rows_use_lowest_and_report_duplicate_conflict():
    _engine, Session = _session_factory()
    with Session() as db:
        pf = _pf(db)
        product = _product(db)
        _list(db, pf, product, LIST_TYPE_MEMORANDUM, 118)
        _list(db, pf, product, LIST_TYPE_MEMORANDUM, 115)
        db.commit()

        price, debug = _price(db, pf, product)

        assert price == Decimal("115.000000")
        assert debug["memorandum_duplicate_conflict"] is True
        assert debug["memorandum_max_price"] == Decimal("115.000000")


def test_rounding_cannot_push_final_price_above_memorandum():
    _engine, Session = _session_factory()
    with Session() as db:
        rounding = RoundingRule(code="STEP10", name="Step 10", mode="math", precision=0, step=10)
        db.add(rounding)
        db.flush()
        pf = _pf(db, rounding_rule=rounding)
        product = _product(db, cost=Decimal("959.20"))
        _list(db, pf, product, LIST_TYPE_MEMORANDUM, Decimal("1199"))
        db.commit()

        price, debug = _price(db, pf, product)

        assert price <= Decimal("1199")
        assert price == Decimal("1199.000000")
        assert debug["memorandum_applied"] is True


def test_excluded_product_remains_excluded_despite_memorandum():
    _engine, Session = _session_factory()
    with Session() as db:
        pf = _pf(db)
        product = _product(db)
        _list(db, pf, product, "exclude_from_pricing", 1)
        _list(db, pf, product, LIST_TYPE_MEMORANDUM, 115)
        db.commit()

        price, debug = _price(db, pf, product)

        assert price == Decimal("100")
        assert debug["excluded_from_pricing"] is True
        assert debug["memorandum_applied"] is False


def test_generated_result_export_includes_memorandum_columns():
    engine, Session = _session_factory()

    def override_db():
        with Session() as db:
            yield db

    app.dependency_overrides[get_db] = override_db
    with Session() as db:
        pf = _pf(db)
        product = _product(db)
        price_list = PriceList(number="MEMO-PL", price_format_id=pf.id, status="generated")
        memo = UniversalList(code="MEMO", name="Меморандум 2026", type=LIST_TYPE_MEMORANDUM, status="active")
        db.add_all([price_list, memo])
        db.flush()
        db.add(
            CalculatedPrice(
                price_list_id=price_list.id,
                product_id=product.id,
                cost=100,
                base_price=125,
                final_price=115,
                applied_reason="memo applied",
                applied_source_name="percentile:P30",
                memorandum_max_price=115,
                price_before_memorandum=130,
                memorandum_applied=True,
                memorandum_below_mdc=True,
                memorandum_list_id=memo.id,
                memorandum_list_name=memo.name,
                memorandum_diagnostic_code="memorandum_below_mdc",
            )
        )
        db.commit()

    client = TestClient(app)
    payload = client.get("/api/generated-price-lists/MEMO-PL/items?page_size=20")
    assert payload.status_code == 200
    row = payload.json()["items"][0]
    assert row["memorandumApplied"] is True
    assert row["memorandumBelowMdc"] is True
    assert row["appliedSourceName"] == "percentile:P30"

    csv_response = client.get("/api/generated-price-lists/MEMO-PL/export.csv")
    assert csv_response.status_code == 200
    csv_row = list(csv.DictReader(io.StringIO(csv_response.content.decode("utf-8-sig"))))[0]
    assert csv_row["Максимальная цена по меморандуму"] == "115.0"
    assert csv_row["Меморандум применён"] == "да"
    assert csv_row["Меморандум ниже МДЦ"] == "да"

    xlsx_response = client.get("/api/generated-price-lists/MEMO-PL/export.xlsx")
    assert xlsx_response.status_code == 200
    workbook = load_workbook(io.BytesIO(xlsx_response.content), read_only=True)
    sheet = workbook.active
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    values = [cell.value for cell in next(sheet.iter_rows(min_row=2, max_row=2))]
    row_by_header = dict(zip(headers, values))
    assert row_by_header["Максимальная цена по меморандуму"] == 115.0
    assert row_by_header["Меморандум применён"] == "да"
    assert row_by_header["Меморандум ниже МДЦ"] == "да"
    app.dependency_overrides.pop(get_db, None)
    engine.dispose()


def test_memorandum_preload_keeps_product_loop_query_count_constant():
    engine, Session = _session_factory()

    def run(product_count: int) -> int:
        with Session() as db:
            pf = _pf(db, code=f"PF-{product_count}")
            products = [_product(db, code=f"SKU-{product_count}-{idx}") for idx in range(product_count)]
            for product in products:
                _list(db, pf, product, LIST_TYPE_MEMORANDUM, 115)
            db.commit()
            active = _active_lists_for_format(db, int(pf.id), date.today())
            preload = _build_pricing_preload(db, price_format=pf, products=products, active_lists=active, branch_id="1")
            count = 0

            def before_cursor_execute(*_args):
                nonlocal count
                count += 1

            event.listen(engine, "before_cursor_execute", before_cursor_execute)
            try:
                for product in products:
                    calculate_price_for_product(
                        db=db,
                        product=product,
                        price_format=pf,
                        as_of=date.today(),
                        region_id=1,
                        active_lists=active,
                        cost_override=product.cost,
                        pricing_preload=preload,
                    )
            finally:
                event.remove(engine, "before_cursor_execute", before_cursor_execute)
            return count

    assert run(5) == run(100)
