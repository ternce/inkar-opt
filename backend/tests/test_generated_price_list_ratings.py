from __future__ import annotations

import csv
import io

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from backend.app.db import Base
from backend.app.deps import get_db
from backend.app.main import app
from backend.app.models import CalculatedPrice, PriceFormat, PriceList, Product, ProductRating, UniversalList


def _client_with_rating_matrix():
    engine = create_engine("sqlite://", connect_args={"check_same_thread": False}, poolclass=StaticPool)
    Base.metadata.create_all(engine)
    Session = sessionmaker(bind=engine)

    def override_db():
        with Session() as db:
            yield db

    app.dependency_overrides[get_db] = override_db
    with Session() as db:
        price_format = PriceFormat(code="RATING-PF", name="Rating PF", branch="BRANCH-1")
        db.add(price_format)
        db.flush()
        price_list = PriceList(number="RATING-PL", price_format_id=price_format.id, status="generated")
        db.add(price_list)
        db.flush()
        universal_list = UniversalList(code="UL-RATING", name="Rating fixed markup", type="fixed_markup", status="active")
        db.add(universal_list)
        db.flush()
        products = [
            Product(code="BOTH", name="Both", cost=100, top_rank=999),
            Product(code="GLOBAL", name="Global", cost=100),
            Product(code="LOCAL", name="Local", cost=100),
            Product(code="NONE", name="None", cost=100, top_rank=888),
        ]
        db.add_all(products)
        db.flush()
        for index, product in enumerate(products):
            db.add(
                CalculatedPrice(
                    price_list_id=price_list.id,
                    product_id=product.id,
                    cost=100,
                    base_price=120,
                    final_price=130 + index,
                    applied_reason="pricing log for BOTH" if product.code == "BOTH" else "",
                    applied_list_id=universal_list.id if product.code == "BOTH" else None,
                    applied_rule_type="fixed_markup" if product.code == "BOTH" else "",
                    applied_rule_value=10.5 if product.code == "BOTH" else None,
                    rating_global=900 + index,
                    rating_local=800 + index,
                )
            )
        db.add_all(
            [
                ProductRating(branch_id="", product_id=products[0].id, sku="BOTH", rating_type="global", rating=1),
                ProductRating(branch_id="BRANCH-1", product_id=products[0].id, sku="BOTH", rating_type="local", rating=11),
                ProductRating(branch_id="", product_id=products[1].id, sku="GLOBAL", rating_type="global", rating=2),
                ProductRating(branch_id="BRANCH-1", product_id=products[2].id, sku="LOCAL", rating_type="local", rating=12),
                # A rating from another branch must not leak into this generated list.
                ProductRating(branch_id="BRANCH-2", product_id=products[3].id, sku="NONE", rating_type="local", rating=99),
            ]
        )
        db.commit()
    return TestClient(app)


def test_generated_payload_and_exports_use_global_and_local_product_ratings():
    client = _client_with_rating_matrix()

    response = client.get("/api/generated-price-lists/RATING-PL/items?page_size=20")
    assert response.status_code == 200
    by_sku = {row["sku"]: row for row in response.json()["items"]}
    assert (by_sku["BOTH"]["globalRating"], by_sku["BOTH"]["localRating"]) == (1, 11)
    assert (by_sku["GLOBAL"]["globalRating"], by_sku["GLOBAL"]["localRating"]) == (2, None)
    assert (by_sku["LOCAL"]["globalRating"], by_sku["LOCAL"]["localRating"]) == (None, 12)
    assert (by_sku["NONE"]["globalRating"], by_sku["NONE"]["localRating"]) == (None, None)
    assert by_sku["BOTH"]["global_rating"] == 1
    assert by_sku["BOTH"]["local_rating"] == 11
    assert {sku: row["finalPrice"] for sku, row in by_sku.items()} == {
        "BOTH": 130.0,
        "GLOBAL": 131.0,
        "LOCAL": 132.0,
        "NONE": 133.0,
    }

    csv_response = client.get("/api/generated-price-lists/RATING-PL/export.csv")
    assert csv_response.status_code == 200
    csv_rows = list(csv.DictReader(io.StringIO(csv_response.content.decode("utf-8-sig"))))
    csv_header = csv_rows[0].keys()
    assert "Рейтинг глобальный" in csv_header
    assert "Рейтинг локальный" in csv_header
    assert not any("PharmCenter Top" in value or "Топ фарм-центра" in value for value in csv_header)

    both_csv = next(row for row in csv_rows if row["SKU"] == "BOTH")
    assert both_csv["Лог расчета цены"] == "pricing log for BOTH"
    assert "Rating fixed markup" in both_csv["Лог применения списка"]
    assert "UL-RATING" in both_csv["Лог применения списка"]
    assert both_csv["Тип списка"] == "fixed_markup"
    assert both_csv["Название списка"] == "Rating fixed markup"
    assert both_csv["Код списка"] == "UL-RATING"
    assert both_csv["Значение списка"] == "10.5%"
    assert both_csv["Диагностика списка"]

    xlsx_response = client.get("/api/generated-price-lists/RATING-PL/export.xlsx")
    assert xlsx_response.status_code == 200
    sheet = load_workbook(io.BytesIO(xlsx_response.content)).active
    xlsx_rows = list(sheet.iter_rows(values_only=True))
    xlsx_header = list(xlsx_rows[0])
    assert "Рейтинг глобальный" in xlsx_header
    assert "Рейтинг локальный" in xlsx_header
    assert not any(value and ("PharmCenter Top" in value or "Топ фарм-центра" in value) for value in xlsx_header)

    xlsx_dicts = [dict(zip(xlsx_header, row)) for row in xlsx_rows[1:]]
    both_xlsx = next(row for row in xlsx_dicts if row["SKU"] == "BOTH")
    assert both_xlsx["Лог расчета цены"] == "pricing log for BOTH"
    assert "Rating fixed markup" in both_xlsx["Лог применения списка"]
    assert both_xlsx["Тип списка"] == "fixed_markup"
    assert both_xlsx["Код списка"] == "UL-RATING"
    assert both_xlsx["Значение списка"] == "10.5%"
    log_column = xlsx_header.index("Лог применения списка") + 1
    assert sheet.column_dimensions[sheet.cell(row=1, column=log_column).column_letter].width >= 20

    app.dependency_overrides.pop(get_db, None)
