import os
import time
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed

import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry


# =========================
# НАСТРОЙКИ
# =========================

TOKEN = os.getenv("PROVISOR_TOKEN", "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJqdGkiOiJiM2E3N2U3NS02ODEwLTRiY2YtOTIyMS1jNGEwZGI5ZWRjNDUiLCJodHRwOi8vc2NoZW1hcy54bWxzb2FwLm9yZy93cy8yMDA1LzA1L2lkZW50aXR5L2NsYWltcy9uYW1laWRlbnRpZmllciI6IjVlZmIyOTZkLTk5NTItNGI5ZS1iYjcwLTgzZmFkMzE2MTM4MyIsImh0dHA6Ly9zY2hlbWFzLnhtbHNvYXAub3JnL3dzLzIwMDUvMDUvaWRlbnRpdHkvY2xhaW1zL25hbWUiOiJBa3NhaTQvODMiLCJodHRwOi8vc2NoZW1hcy54bWxzb2FwLm9yZy93cy8yMDA1LzA1L2lkZW50aXR5L2NsYWltcy9oYXNoIjoiQVFBQUFBRUFBQ2NRQUFBQUVFTTJ3azFLU1U5TEhlaW91eUZuZmIxRk1WZi85NzVZOXRzajRVNjRNZGQvTnFuQkRjeGRpZWY3MjlTNjRPRmNYdz09IiwiQ29tcGFueVR5cGUiOiLQkNC_0YLQtdC60LAiLCJodHRwOi8vc2NoZW1hcy5taWNyb3NvZnQuY29tL3dzLzIwMDgvMDYvaWRlbnRpdHkvY2xhaW1zL3JvbGUiOiJQaGFybWFjeSIsImV4cCI6MTc3Njk2Nzg5MiwiaXNzIjoiaHR0cHM6Ly9QaGFybWNlbnRlci5reiIsImF1ZCI6IlBoYXJtYWNldXRpY2FsIG1hcmtldCJ9.E4UbZGiFEsdbw7641DBrJtNXzNlCUiAznSDSAp4ZPBI").strip()

if not TOKEN:
    raise ValueError("Не найден токен. Укажи переменную окружения PROVISOR_TOKEN.")

FILIALS = {


    
    128: "Инкар Алматы",
    159: "Медсервис (Актау)",
    1106: "Эмити Интернешнл (Актау)",
}

FILIAL_IDS = list(FILIALS.keys())

BASE_URL = "https://api.provisor.kz"

HEADERS = {
    "Authorization": f"Bearer {TOKEN}",
    "Accept": "application/json",
    "User-Agent": "Mozilla/5.0",
    "Connection": "keep-alive",
}

EXPORT_DIR = "exports"
os.makedirs(EXPORT_DIR, exist_ok=True)

# Сетевые настройки
CONNECT_TIMEOUT = 20
READ_TIMEOUT = 300
MAX_RETRIES = 5
RETRY_BACKOFF = 2
MAX_WORKERS = min(4, max(1, len(FILIAL_IDS)))


# =========================
# SESSION / RETRY
# =========================

def create_session() -> requests.Session:
    session = requests.Session()

    retry = Retry(
        total=MAX_RETRIES,
        connect=MAX_RETRIES,
        read=MAX_RETRIES,
        backoff_factor=RETRY_BACKOFF,
        status_forcelist=[429, 500, 502, 503, 504],
        allowed_methods=frozenset(["GET"]),
        raise_on_status=False,
    )

    adapter = HTTPAdapter(
        max_retries=retry,
        pool_connections=20,
        pool_maxsize=20,
    )

    session.mount("http://", adapter)
    session.mount("https://", adapter)
    session.headers.update(HEADERS)

    return session


# =========================
# API
# =========================

def fetch_price_data(session: requests.Session, filial_id: int) -> list:
    url = f"{BASE_URL}/Price/GetByFilialId?filialId={filial_id}"
    filial_name = FILIALS.get(filial_id, str(filial_id))

    print(f"[LOAD] filial={filial_id} | {filial_name}")

    started = time.perf_counter()

    try:
        response = session.get(
            url,
            timeout=(CONNECT_TIMEOUT, READ_TIMEOUT),
        )
        response.raise_for_status()
    except requests.exceptions.ReadTimeout:
        raise TimeoutError(
            f"Превышено время ожидания ответа для filial_id={filial_id}"
        )
    except requests.exceptions.ConnectTimeout:
        raise TimeoutError(
            f"Не удалось установить соединение для filial_id={filial_id}"
        )
    except requests.exceptions.RequestException as e:
        raise RuntimeError(
            f"Ошибка запроса для filial_id={filial_id}: {e}"
        ) from e

    try:
        data = response.json()
    except ValueError as e:
        raise ValueError(
            f"Некорректный JSON для filial_id={filial_id}"
        ) from e

    if not isinstance(data, list):
        raise ValueError(f"API вернул не список для filial_id={filial_id}")

    elapsed = time.perf_counter() - started
    print(f"[OK] filial={filial_id} | rows={len(data)} | {elapsed:.1f} sec")

    return data


# =========================
# DATA
# =========================

def normalize_shelf_life(value) -> str | None:
    if value in (None, "", "0001-01-01T00:00:00"):
        return None

    dt = pd.to_datetime(value, errors="coerce")
    if pd.isna(dt):
        return None

    return dt.strftime("%Y-%m-%d")


def flatten_item(item: dict) -> dict:
    goods = item.get("goods") or {}
    filial = item.get("filial") or {}

    distributor_goods_id = item.get("distributorGoodsId")
    distributor_goods_id = str(distributor_goods_id).strip() if distributor_goods_id else None

    shelf_life = normalize_shelf_life(item.get("shelfLife"))

    batch = item.get("batch")
    batch = str(batch).strip() if batch not in (None, "") else None

    return {
        "row_id": item.get("id"),
        "goods_id": item.get("goodsId"),
        "filial_id": item.get("filialId"),
        "filial_name": filial.get("name"),
        "distributor_goods_id": distributor_goods_id,
        "goods_full_name": goods.get("fullName"),
        "distributor_goods_name": item.get("distributorGoodsName"),
        "distributor_producer": item.get("distributorProducer"),
        "price": item.get("goodsPrice"),
        "stored": item.get("stored"),
        "shelf_life": shelf_life,
        "batch": batch,
        "price_status": item.get("priceStatus"),
        "pack": item.get("pack"),
        "box": item.get("box"),
        "multiplicity": item.get("multiplicity"),
        "min_order": item.get("minOrder"),
    }


def build_dataframe(data: list) -> pd.DataFrame:
    df = pd.DataFrame([flatten_item(x) for x in data])

    if df.empty:
        return df

    df["price"] = pd.to_numeric(df["price"], errors="coerce")
    df["stored"] = pd.to_numeric(df["stored"], errors="coerce")
    df["goods_id"] = pd.to_numeric(df["goods_id"], errors="coerce")
    df["filial_id"] = pd.to_numeric(df["filial_id"], errors="coerce")

    return df


# =========================
# PREPARE
# =========================

def build_keys(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()

    df["goods_id_str"] = df["goods_id"].fillna(-1).astype("Int64").astype(str)
    df["distributor_goods_id_key"] = df["distributor_goods_id"].fillna("no_sku")
    df["shelf_life_key"] = df["shelf_life"].fillna("no_exp")
    df["batch_key"] = df["batch"].fillna("no_batch")

    df["SKU"] = (
        df["goods_id_str"]
        + "_"
        + df["distributor_goods_id_key"]
        + "_"
        + df["shelf_life_key"]
        # + "_" + df["batch_key"]  # включи, если партия должна участвовать в уникальности
    )

    df["Название"] = df["goods_full_name"].fillna(df["distributor_goods_name"])
    df["Срок годности"] = df["shelf_life"].fillna("")
    df["SKU дистрибьютора"] = df["distributor_goods_id"].fillna("")
    df["Производитель дистрибьютора"] = df["distributor_producer"].fillna("")

    return df


def aggregate_duplicates_inside_filial(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df

    group_cols = ["filial_id", "SKU"]

    aggregated = (
        df.groupby(group_cols, as_index=False)
        .agg({
            "goods_id": "first",
            "distributor_goods_id": "first",
            "Название": "first",
            "Срок годности": "first",
            "Производитель дистрибьютора": "first",
            "price": "min",
            "stored": "sum",
        })
    )

    return aggregated


# =========================
# PIVOT
# =========================

def build_pivot(df: pd.DataFrame, filial_ids: list[int]) -> pd.DataFrame:
    if df.empty:
        raise ValueError("Пустой DataFrame, нечего сводить.")

    df = df[df["goods_id"].notna()].copy()
    df = build_keys(df)
    df = aggregate_duplicates_inside_filial(df)

    base = df[
        [
            "SKU",
            "goods_id",
            "distributor_goods_id",
            "Название",
            "Срок годности",
            "Производитель дистрибьютора",
        ]
    ].drop_duplicates()

    base = base.rename(columns={
        "goods_id": "goodsId",
        "distributor_goods_id": "SKU дистрибьютора",
    })

    result = base.copy()

    for fid in filial_ids:
        label = FILIALS.get(fid, str(fid))

        subset = df[df["filial_id"] == fid][["SKU", "price", "stored"]].copy()
        subset = subset.rename(columns={
            "price": f"Цена {label}",
            "stored": f"Остаток {label}",
        })

        result = result.merge(subset, on="SKU", how="left")

    result = result.sort_values(
        by=["goodsId", "SKU дистрибьютора", "Срок годности", "SKU"],
        na_position="last",
    ).reset_index(drop=True)

    return result


# =========================
# EXCEL FORMAT
# =========================

def format_excel(file_path: str) -> None:
    wb = load_workbook(file_path)
    ws = wb.active

    for col in ws.columns:
        col_letter = col[0].column_letter
        header = str(col[0].value) if col[0].value else ""

        if header == "Название":
            ws.column_dimensions[col_letter].width = 70
            continue

        if header in {"SKU", "SKU дистрибьютора"}:
            ws.column_dimensions[col_letter].width = 28
            continue

        if header == "Производитель дистрибьютора":
            ws.column_dimensions[col_letter].width = 28
            continue

        max_length = 0
        for cell in col:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))

        ws.column_dimensions[col_letter].width = min(max_length + 2, 25)

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A2"

    wb.save(file_path)


# =========================
# WORKER
# =========================

def load_filial_df(fid: int) -> pd.DataFrame | None:
    session = create_session()

    try:
        data = fetch_price_data(session, fid)
        df = build_dataframe(data)

        if df.empty:
            print(f"[WARN] filial={fid} вернул пустой DataFrame")
            return None

        return df

    except Exception as e:
        print(f"[ERROR] filial={fid}: {e}")
        return None

    finally:
        session.close()


# =========================
# MAIN
# =========================

def main() -> None:
    total_started = time.perf_counter()
    frames: list[pd.DataFrame] = []

    print(f"[INFO] Старт. Филиалов: {len(FILIAL_IDS)} | workers={MAX_WORKERS}")

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        future_to_fid = {
            executor.submit(load_filial_df, fid): fid
            for fid in FILIAL_IDS
        }

        for future in as_completed(future_to_fid):
            fid = future_to_fid[future]

            try:
                df = future.result()
                if df is not None and not df.empty:
                    frames.append(df)
                    print(f"[APPEND] filial={fid} добавлен | rows={len(df)}")
                else:
                    print(f"[SKIP] filial={fid} без данных")
            except Exception as e:
                print(f"[FUTURE ERROR] filial={fid}: {e}")

    if not frames:
        raise Exception("Нет данных ни по одному филиалу.")

    print("[INFO] Объединяем все DataFrame...")
    df_all = pd.concat(frames, ignore_index=True)
    print(f"[INFO] Общий объём строк до pivot: {len(df_all)}")

    print("[INFO] Строим pivot...")
    pivot = build_pivot(df_all, FILIAL_IDS)
    print(f"[INFO] Строк после pivot: {len(pivot)}")

    now = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    file_path = os.path.join(EXPORT_DIR, f"prices_with_shelf_life_{now}.xlsx")

    print("[INFO] Сохраняем Excel...")
    pivot.to_excel(file_path, index=False)

    print("[INFO] Форматируем Excel...")
    format_excel(file_path)

    total_elapsed = time.perf_counter() - total_started

    print(f"\n[OK] Файл готов: {os.path.abspath(file_path)}")
    print(f"[INFO] Всего строк в отчёте: {len(pivot)}")
    print(f"[INFO] Общее время выполнения: {total_elapsed:.1f} sec")


if __name__ == "__main__":
    main()